import os
import sys
import io
import xlsxwriter
import xlwt
import openpyxl
import xlwt
from xlwt import *
from django.http import HttpResponse
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from balcon.models import *
from inno.funciones import *
#
#
# with transaction.atomic():
#     try:
#         libre_origen = '/homologacion_turismo.xls'
#         fuentecabecera = easyxf(
#             'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf(
#             'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#
#         output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
#         # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#         libdestino = xlwt.Workbook()
#         hojadestino = libdestino.add_sheet('Sheet1')
#         fil = 0
#         columnas = [(u"CEDULA", 6000),
#                     (u"APELLIDOS Y NOMBRES", 6000),
#                     (u"OBSERVACIÓN", 6000)
#                     ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#         fila = 1
#         titulacion = 0
#
#         lin = 0
#         # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
#         #miarchivo = openpyxl.load_workbook("turlinea.xlsx")
#         miarchivo = openpyxl.load_workbook(path_anexo)
#         # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")
#
#         ws = miarchivo.get_sheet_by_name("novenod")
#         worksheet = ws
#         c = 0
#         cont = 0
#         periodo_id = 224
#         carrera_id = 134
#         mallaantigua_id = 199
#         mallanueva_id = 487
#         for row in worksheet.iter_rows(min_row=0):
#             if lin >= 0:
#                 currentValues, cadena = [], ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 identificacion = currentValues[0]
#
#                 if not identificacion:
#                     break
#
#                 matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
#                                                      inscripcion__persona__cedula=identificacion).first()
#                 cont += 1
#                 matricula.pasoayuda = True
#                 matricula.save()
#                 print(u"%s - %s" % (matricula, cont))
#                 inscripcion = matricula.inscripcion
#                 hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
#                 hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
#                 hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)
#
#                 itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)
#
#                 for itinerario in itinerarios:
#                     itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
#                                                                       nivel_id=itinerario.nivel_id)
#                     # estadodo culmiado, en curso y pendiente
#                     practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
#                                                                                            inscripcion=inscripcion,
#                                                                                            estadosolicitud=2,
#                                                                                            culminada=False,
#                                                                                            itinerariomalla=itinerario)
#                     for pcurso in practicasencurso:
#                         pcurso.culminada = True
#                         pcurso.save()
#
#                     practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
#                                                                                               inscripcion=inscripcion,
#                                                                                               estadosolicitud=4,
#                                                                                               culminada=False,
#                                                                                               itinerariomalla=itinerario)
#                     for pendiente in practicaspendientes:
#                         pendiente.estadosolicitud = 2
#                         pendiente.culminada = True
#                         pendiente.save()
#
#                 practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
#                 horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)
#
#                 if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
#                                                    malla_id=mallaantigua_id).exists():
#                     imantigua = \
#                         InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
#                             0]
#                     imantigua.status = False
#                     imantigua.save()
#                     print(u"Desactiva antigua inscripcion -----------------------------")
#
#                 if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
#                                                        malla_id=mallanueva_id).exists():
#                     imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
#                     imnueva.save()
#                     print(u"Crea nueva inscripcion -----------------------------")
#
#                 equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
#                                                                             asignaturamalla__malla_id=mallaantigua_id).order_by(
#                     'asignaturamallasalto__nivelmalla__orden')
#                 for equivalencia in equivalencias:
#                     print(u"nueva - %s" % equivalencia.asignaturamallasalto)
#                     recordantiguo = inscripcion.recordacademico_set.filter(status=True,
#                                                                            asignaturamalla_id=equivalencia.asignaturamalla_id).first()
#
#                     if recordantiguo:
#                         print(u"anterior - %s" % equivalencia.asignaturamalla)
#                         print(u"Record antiguo: %s" % recordantiguo)
#                         recordnuevo = None
#                         recordantiguo.status = False
#                         recordantiguo.save(update_asignaturamalla=False)
#
#                         if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
#                             observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
#                             homologada = True
#
#
#
#                         else:
#                             observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
#                             homologada = False
#                         if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
#                                                               asignaturamalla=equivalencia.asignaturamallasalto).exists():
#
#                             recordnuevo = RecordAcademico(inscripcion=inscripcion,
#                                                           matriculas=recordantiguo.matriculas,
#                                                           asignaturamalla=equivalencia.asignaturamallasalto,
#                                                           asignatura=equivalencia.asignaturamallasalto.asignatura,
#                                                           asignaturaold_id=recordantiguo.asignatura.id,
#                                                           nota=recordantiguo.nota,
#                                                           asistencia=recordantiguo.asistencia,
#                                                           sinasistencia=recordantiguo.sinasistencia,
#                                                           fecha=recordantiguo.fecha,
#                                                           noaplica=recordantiguo.noaplica,
#                                                           aprobada=recordantiguo.aprobada,
#                                                           convalidacion=recordantiguo.convalidacion,
#                                                           pendiente=recordantiguo.pendiente,
#                                                           creditos=equivalencia.asignaturamallasalto.creditos,
#                                                           horas=equivalencia.asignaturamallasalto.horas,
#                                                           valida=recordantiguo.valida,
#                                                           validapromedio=recordantiguo.validapromedio,
#                                                           observaciones=observaciones,
#                                                           homologada=homologada,
#                                                           materiaregular=recordantiguo.materiaregular,
#                                                           materiacurso=None,
#                                                           completonota=recordantiguo.completonota,
#                                                           completoasistencia=recordantiguo.completoasistencia,
#                                                           fechainicio=recordantiguo.fechainicio,
#                                                           fechafin=recordantiguo.fechafin,
#                                                           suficiencia=recordantiguo.suficiencia,
#                                                           asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
#                                                           reverso=False)
#                             recordnuevo.save()
#                             print(u"Crea nuevo record %s" % recordnuevo)
#
#
#                         elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
#                                                             asignaturamalla=equivalencia.asignaturamallasalto):
#                             recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
#                                                                          asignaturamalla=equivalencia.asignaturamallasalto)[
#                                 0]
#                             recordnuevo.matriculas = recordantiguo.matriculas
#                             recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
#                             recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
#                             recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
#                             recordnuevo.nota = recordantiguo.nota
#                             recordnuevo.asistencia = recordantiguo.asistencia
#                             recordnuevo.sinasistencia = recordantiguo.sinasistencia
#                             recordnuevo.fecha = recordantiguo.fecha
#                             recordnuevo.noaplica = recordantiguo.noaplica
#                             recordnuevo.aprobada = recordantiguo.aprobada
#                             recordnuevo.convalidacion = recordantiguo.convalidacion
#                             recordnuevo.pendiente = recordantiguo.pendiente
#                             recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
#                             recordnuevo.horas = equivalencia.asignaturamallasalto.horas
#                             recordnuevo.valida = recordantiguo.valida
#                             recordnuevo.validapromedio = recordantiguo.validapromedio
#                             recordnuevo.observaciones = observaciones
#                             recordnuevo.homologada = homologada
#                             recordnuevo.materiaregular = recordantiguo.materiaregular
#                             recordnuevo.materiacurso = None
#                             recordnuevo.completonota = recordantiguo.completonota
#                             recordnuevo.completoasistencia = recordantiguo.completoasistencia
#                             recordnuevo.fechainicio = recordantiguo.fechainicio
#                             recordnuevo.fechafin = recordantiguo.fechafin
#                             recordnuevo.suficiencia = recordantiguo.suficiencia
#                             recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
#                             recordnuevo.reverso = False
#                             recordnuevo.save()
#
#                         if recordnuevo:
#                             historicos = HistoricoRecordAcademico.objects.filter(status=True,
#                                                                                  recordacademico=recordantiguo).update(
#                                 recordacademico=recordnuevo,
#                                 creditos=recordnuevo.creditos,
#                                 horas=recordnuevo.horas,
#                                 homologada=recordnuevo.homologada)
#                             respaldo = RespaldoRecordAcademico.objects.filter(status=True,
#                                                                               recordacademicooriginal=recordantiguo)
#
#                             if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
#                                 if not practicaspp:
#                                     if recordnuevo.aprobada:
#                                         profesor = None
#                                         if recordnuevo.materiaregular:
#                                             profesor = recordnuevo.materiaregular.profesor_principal()
#                                         elif recordnuevo.materiacurso:
#                                             profesor = recordnuevo.materiaregular.profesor()
#                                         if equivalencia.asignaturamallasalto_id == 10929:
#                                             itinerarioquinto = ItinerariosMalla.objects.get(status=True,
#                                                                                             malla_id=mallaantigua_id,
#                                                                                             nivel_id=5)
#                                             iquintonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                         malla_id=mallanueva_id,
#                                                                                         nivel_id=5)
#
#                                             practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
#                                                                                                            inscripcion=inscripcion,
#                                                                                                            culminada=True,
#                                                                                                            itinerariomalla=itinerarioquinto).exists()
#                                             if not practica:
#                                                 nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
#                                                                                                      inscripcion=inscripcion,
#                                                                                                      fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
#                                                                                                      numerohora=iquintonuevo.horas_practicas,
#                                                                                                      nivelmalla=iquintonuevo.nivel,
#                                                                                                      tiposolicitud=1,
#                                                                                                      estadosolicitud=2,
#                                                                                                      tipo=1,
#                                                                                                      itinerariomalla=iquintonuevo,
#                                                                                                      supervisor=profesor,
#                                                                                                      tutorunemi=profesor,
#                                                                                                      fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      tipoinstitucion=1,
#                                                                                                      sectoreconomico=6,
#                                                                                                      empresaempleadora_id=3,
#                                                                                                      culminada=True,
#                                                                                                      fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      lugarpractica_id=2,
#                                                                                                      observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                                                                      )
#                                                 nuevapractica.save()
#
#                                         if equivalencia.asignaturamallasalto_id == 10947:
#                                             itinerariosexto = ItinerariosMalla.objects.get(status=True,
#                                                                                            malla_id=mallaantigua_id,
#                                                                                            nivel_id=6)
#                                             isextonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                        malla_id=mallanueva_id,
#                                                                                        nivel_id=6)
#
#                                             practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
#                                                                                                            inscripcion=inscripcion,
#                                                                                                            culminada=True,
#                                                                                                            itinerariomalla=itinerariosexto).exists()
#                                             if not practica:
#                                                 nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
#                                                                                                      inscripcion=inscripcion,
#                                                                                                      fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
#                                                                                                      numerohora=isextonuevo.horas_practicas,
#                                                                                                      nivelmalla=isextonuevo.nivel,
#                                                                                                      tiposolicitud=1,
#                                                                                                      estadosolicitud=2,
#                                                                                                      tipo=1,
#                                                                                                      itinerariomalla=isextonuevo,
#                                                                                                      supervisor=profesor,
#                                                                                                      tutorunemi=profesor,
#                                                                                                      fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      tipoinstitucion=1,
#                                                                                                      sectoreconomico=6,
#                                                                                                      empresaempleadora_id=3,
#                                                                                                      culminada=True,
#                                                                                                      fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      lugarpractica_id=2,
#                                                                                                      observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                                                                      )
#                                                 nuevapractica.save()
#
#                                         if equivalencia.asignaturamallasalto_id == 10953:
#                                             itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
#                                                                                              malla_id=mallaantigua_id,
#                                                                                              nivel_id=7)
#                                             iseptimonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                          malla_id=mallanueva_id,
#                                                                                          nivel_id=7)
#
#                                             practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
#                                                                                                            inscripcion=inscripcion,
#                                                                                                            culminada=True,
#                                                                                                            itinerariomalla=itinerarioseptimo).exists()
#                                             if not practica:
#                                                 nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
#                                                                                                      inscripcion=inscripcion,
#                                                                                                      fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
#                                                                                                      numerohora=iseptimonuevo.horas_practicas,
#                                                                                                      nivelmalla=iseptimonuevo.nivel,
#                                                                                                      tiposolicitud=1,
#                                                                                                      estadosolicitud=2,
#                                                                                                      tipo=1,
#                                                                                                      itinerariomalla=iseptimonuevo,
#                                                                                                      supervisor=profesor,
#                                                                                                      tutorunemi=profesor,
#                                                                                                      fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      tipoinstitucion=1,
#                                                                                                      sectoreconomico=6,
#                                                                                                      empresaempleadora_id=3,
#                                                                                                      culminada=True,
#                                                                                                      fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                                      lugarpractica_id=2,
#                                                                                                      observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                                                                      )
#                                                 nuevapractica.save()
#
#                             if equivalencia.asignaturamallasalto_id in [10959, 10964]:
#                                 if not horasvinculacion:
#                                     if recordnuevo.aprobada:
#                                         totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
#                                         if totalhoras == 0:
#                                             a = 0
#                                             print('sin horas')
#                                         if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
#                                             horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
#                                             vinculacion = ParticipantesMatrices(status=True,
#                                                                                 matrizevidencia_id=2,
#                                                                                 proyecto_id=601,
#                                                                                 inscripcion=inscripcion,
#                                                                                 horas=horasfalta,
#                                                                                 registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                 registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
#                                                                                 estado=1
#                                                                                 )
#                                             vinculacion.save()
#
#                                         if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
#                                             horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
#                                             vinculacion = ParticipantesMatrices(status=True,
#                                                                                 matrizevidencia_id=2,
#                                                                                 proyecto_id=601,
#                                                                                 inscripcion=inscripcion,
#                                                                                 horas=horasfalta,
#                                                                                 registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
#                                                                                 registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
#                                                                                 estado=1
#                                                                                 )
#                                             vinculacion.save()
#
#                             if not respaldo.exists():
#                                 respaldorecord = RespaldoRecordAcademico(
#                                     recordacademicooriginal=recordantiguo,
#                                     recordacademiconuevo=recordnuevo
#                                 )
#                                 respaldorecord.save()
#                             else:
#                                 respaldorecord = respaldo[0]
#                                 respaldorecord.recordacademiconuevo = recordnuevo
#                                 respaldorecord.save()
#                             print(u"Record actualizado %s" % recordnuevo)
#
#
#                     else:
#                         hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
#                         fila += 1
#
#                 fila += 1
#
#                 time.sleep(3)
#
#             lin += 1
#
#         libdestino.save(output_folder + libre_origen)
#         print(output_folder + libre_origen)
#         print("Proceso finalizado. . .")
#
#
#     except Exception as ex:
#         noti = Notificacion(titulo='Error',
#                             cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
#                                 ex, sys.exc_info()[-1].tb_lineno),
#                             destinatario_id=29898, url="",
#                             prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                             tipo=2, en_proceso=False, error=True)
#         noti.save()
#
#         transaction.set_rollback(True)
#         print('error: %s' % ex)
#         hojadestino.write(fila, 3, str(ex))
#         fila += 1
#





# import openpyxl
# import xlwt
# from django.db import transaction
# from xlwt import easyxf
# from sga.models import *
#
# try:
#     libre_origen = '/reporte_TS.xls'
#     fuentecabecera = easyxf(
#         'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormal = easyxf(
#         'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#
#     output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
#     # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#     libdestino = xlwt.Workbook()
#     hojadestino = libdestino.add_sheet('Sheet1')
#     fil = 0
#     lin = 0
#     columnas = [(u"CEDULA", 6000),
#                 (u"APELLIDOS Y NOMBRES", 6000),
#                 (u"ITINERARO REPETIDO", 6000)
#                 ]
#     for col_num in range(len(columnas)):
#         hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#         hojadestino.col(col_num).width = columnas[col_num][1]
#     fila = 1
#
#     miarchivo = openpyxl.load_workbook(path_anexo)
#     # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")
#
#     ws = miarchivo.get_sheet_by_name("reporte")
#     worksheet = ws
#     carrera_id = 130
#     mallaantigua_id = 206
#     mallanueva_id = 485
#     for row in worksheet.iter_rows(min_row=0):
#         if lin >= 0:
#             currentValues, cadena = [], ''
#             for cell in row:
#                 cadena += str(cell.value) + ' '
#                 currentValues.append(str(cell.value))
#             identificacion = currentValues[0]
#
#             if not identificacion:
#                 break
#
#             matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
#                                                  inscripcion__persona__cedula=identificacion).first()
#             inscripcion = matricula.inscripcion
#             PPP = inscripcion.numero_horas_practicas_pre_profesionales()
#
#             for i in range(1, 10):
#
#                 itinerarioviejo = ItinerariosMalla.objects.get(status=True,
#                                                                malla_id=mallaantigua_id,
#                                                                nivel_id=i)
#                 itinerarionuevo = ItinerariosMalla.objects.get(status=True,
#                                                                malla_id=mallanueva_id,
#                                                                nivel_id=i)
#
#                 practicaanterior = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
#                                                                                        inscripcion=inscripcion,
#                                                                                        estadosolicitud__in=[
#                                                                                            1, 2, 4],
#                                                                                        itinerariomalla=itinerarioviejo).exists()
#                 practicanueva = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
#                                                                                     inscripcion=inscripcion,
#                                                                                     estadosolicitud__in=[
#                                                                                         1, 2, 4],
#                                                                                     itinerariomalla=itinerarionuevo).exists()
#
#                 if practicaanterior and practicanueva:
#                     hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
#                     hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
#                     hojadestino.write(fila, 2,
#                                       str(itinerarioviejo.nombre) + ' ' + str(itinerarioviejo.nivel) + ' - ' + str(
#                                           itinerarionuevo.nombre) + ' ' + str(itinerarionuevo.nivel), fuentenormal)
#                     fila += 1
#
#         lin += 1
#     libdestino.save(output_folder + libre_origen)
#     print("Proceso finalizado. . .")
# except Exception as ex:
#     noti = Notificacion(titulo='Error',
#                         cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
#                             ex, sys.exc_info()[-1].tb_lineno),
#                         destinatario_id=29898, url="",
#                         prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                         tipo=2, en_proceso=False, error=True)
#     noti.save()
#     print('error: %s' % ex)
#     hojadestino.write(fila, 3, str(ex))
#     fila += 1
#

# import openpyxl
# import xlwt
# from django.db import transaction
# from xlwt import easyxf
# from inno.funciones import *
#
# from sga.models import *
#
# try:
#     libre_origen = '/reporte_derecho_v.xls'
#     fuentecabecera = easyxf(
#         'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormal = easyxf(
#         'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#
#     output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
#     # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#     libdestino = xlwt.Workbook()
#     hojadestino = libdestino.add_sheet('Sheet1')
#     fil = 0
#     lin = 0
#     columnas = [(u"CEDULA", 6000),
#                 (u"APELLIDOS Y NOMBRES", 6000),
#                 (u"PRACTICAS", 6000),
#                 (u"VINCULACION", 6000)
#                 ]
#     for col_num in range(len(columnas)):
#         hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#         hojadestino.col(col_num).width = columnas[col_num][1]
#     fila = 1
#
#     miarchivo = openpyxl.load_workbook("derecholinea")
#     # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")
#
#     ws = miarchivo.get_sheet_by_name("Hoja1")
#     worksheet = ws
#     carrera_id = 126
#     # mallaantigua_id = 198
#     # mallanueva_id = 492
#     for row in worksheet.iter_rows(min_row=0):
#         if lin >= 0:
#             currentValues, cadena = [], ''
#             for cell in row:
#                 cadena += str(cell.value) + ' '
#                 currentValues.append(str(cell.value))
#             identificacion = currentValues[0]
#
#             if not identificacion:
#                 break
#
#             matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
#                                                  inscripcion__persona__cedula=identificacion).first()
#             inscripcion = matricula.inscripcion
#             hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
#             hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
#             practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
#             horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)
#
#             if practicaspp:
#                 hojadestino.write(fila, 2, 'SI', fuentenormal)
#             else:
#                 hojadestino.write(fila, 2, 'NO', fuentenormal)
#             if horasvinculacion:
#                 hojadestino.write(fila, 3, 'SI', fuentenormal)
#             else:
#                 hojadestino.write(fila, 3, 'NO', fuentenormal)
#
#
#
#         lin += 1
#     libdestino.save(output_folder + libre_origen)
#     print("Proceso finalizado. . .")
# except Exception as ex:
#     noti = Notificacion(titulo='Error',
#                         cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
#                                                                                        sys.exc_info()[-1].tb_lineno),
#                         destinatario_id=29898, url="",
#                         prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                         tipo=2, en_proceso=False, error=True)
#     noti.save()
#     transaction.set_rollback(True)
#     print('error: %s' % ex)
#     hojadestino.write(fila, 4, str(ex))
#     fila += 1


# import openpyxl
# import xlwt
# from django.db import transaction
# from xlwt import easyxf
#
# from inno.funciones import haber_cumplido_horas_creditos_practicas_preprofesionales, \
#     haber_cumplido_horas_creditos_vinculacion
# from sga.models import *
#
# try:
#     libre_origen = '/homologacion_idiomas_finales_3.xls'
#     fuentecabecera = easyxf(
#         'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormal = easyxf(
#         'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#
#     output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
#     # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#     libdestino = xlwt.Workbook()
#     hojadestino = libdestino.add_sheet('Sheet1')
#     fil = 0
#     columnas = [(u"CEDULA", 6000),
#                 (u"APELLIDOS Y NOMBRES", 6000),
#                 (u"OBSERVACIÓN", 6000),
#                 (u"HORAS PRACTICAS", 6000),
#                 (u"HORAS VINCULACION", 6000),
#                 (u"OBSERVACIÓN", 6000)
#                 ]
#     for col_num in range(len(columnas)):
#         hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#         hojadestino.col(col_num).width = columnas[col_num][1]
#     fila = 1
#     titulacion = 0
#
#     lin = 0
#     # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
#     miarchivo = openpyxl.load_workbook(path_anexo)
#     # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")
#
#     ws = miarchivo.get_sheet_by_name("prueba3")
#     worksheet = ws
#     c = 0
#     cont = 0
#     periodo_id = 224
#     carrera_id = 129
#     mallaantigua_id = 198
#     mallanueva_id = 492
#     sin_matricula = []
#
#     for row in worksheet.iter_rows(min_row=0):
#         if lin >= 0:
#             currentValues, cadena = [], ''
#             for cell in row:
#                 cadena += str(cell.value) + ' '
#                 currentValues.append(str(cell.value))
#             inscripcion = int(currentValues[0])
#
#             if not inscripcion:
#                 break
#
#             with transaction.atomic():
#                 matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
#                                                      inscripcion_id=inscripcion).first()
#                 if matricula:
#                     cont += 1
#                     matricula.pasoayuda = True
#                     matricula.save()
#                     print(u"%s - %s - %s" % (matricula, cont, inscripcion))
#                     inscripcion = matricula.inscripcion
#                     hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
#                     hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
#                     hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)
#
#                     practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
#                     horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)
#
#                     if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
#                                                        malla_id=mallaantigua_id).exists():
#                         imantigua = \
#                         InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
#                             0]
#                         imantigua.status = False
#                         imantigua.save()
#                         print(u"Desactiva antigua inscripcion -----------------------------")
#
#                     if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
#                                                            malla_id=mallanueva_id).exists():
#                         imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
#                         imnueva.save()
#                         print(u"Crea nueva inscripcion -----------------------------")
#
#                     equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
#                                                                                 asignaturamalla__malla_id=mallaantigua_id).order_by(
#                         'asignaturamallasalto__nivelmalla__orden')
#                     cont_asig_vinculacion_aprobadas = 0
#                     horasfalta = 0
#                     fechainicioitinerario = None
#                     fechafinitinerario = None
#                     for equivalencia in equivalencias:
#                         print(u"nueva - %s" % equivalencia.asignaturamallasalto)
#                         recordantiguo = inscripcion.recordacademico_set.filter(status=True,
#                                                                                asignaturamalla_id=equivalencia.asignaturamalla_id).first()
#
#                         if recordantiguo:
#                             print(u"anterior - %s" % equivalencia.asignaturamalla)
#                             print(u"Record antiguo: %s" % recordantiguo)
#                             recordnuevo = None
#                             recordantiguo.status = False
#                             recordantiguo.save(update_asignaturamalla=False)
#
#                             if equivalencia.asignaturamallasalto_id in [11018, 11021, 11020, 10990, 10991, 11005, 10993,
#                                                                         10997, 11007]:
#                                 observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
#                                 homologada = True
#
#
#                             else:
#                                 observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
#                                 homologada = False
#                             if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
#                                                                   asignaturamalla=equivalencia.asignaturamallasalto).exists():
#
#                                 recordnuevo = RecordAcademico(inscripcion=inscripcion,
#                                                               matriculas=recordantiguo.matriculas,
#                                                               asignaturamalla=equivalencia.asignaturamallasalto,
#                                                               asignatura=equivalencia.asignaturamallasalto.asignatura,
#                                                               asignaturaold_id=recordantiguo.asignatura.id,
#                                                               nota=recordantiguo.nota,
#                                                               asistencia=recordantiguo.asistencia,
#                                                               sinasistencia=recordantiguo.sinasistencia,
#                                                               fecha=recordantiguo.fecha,
#                                                               noaplica=recordantiguo.noaplica,
#                                                               aprobada=recordantiguo.aprobada,
#                                                               convalidacion=recordantiguo.convalidacion,
#                                                               pendiente=recordantiguo.pendiente,
#                                                               creditos=equivalencia.asignaturamallasalto.creditos,
#                                                               horas=equivalencia.asignaturamallasalto.horas,
#                                                               valida=recordantiguo.valida,
#                                                               validapromedio=recordantiguo.validapromedio,
#                                                               observaciones=observaciones,
#                                                               homologada=homologada,
#                                                               materiaregular=recordantiguo.materiaregular,
#                                                               materiacurso=None,
#                                                               completonota=recordantiguo.completonota,
#                                                               completoasistencia=recordantiguo.completoasistencia,
#                                                               fechainicio=recordantiguo.fechainicio,
#                                                               fechafin=recordantiguo.fechafin,
#                                                               suficiencia=recordantiguo.suficiencia,
#                                                               asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
#                                                               reverso=False)
#                                 recordnuevo.save()
#                                 print(u"Crea nuevo record %s" % recordnuevo)
#
#
#                             elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
#                                                                 asignaturamalla=equivalencia.asignaturamallasalto):
#                                 recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
#                                                                              asignaturamalla=equivalencia.asignaturamallasalto)[
#                                     0]
#                                 recordnuevo.matriculas = recordantiguo.matriculas
#                                 recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
#                                 recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
#                                 recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
#                                 recordnuevo.nota = recordantiguo.nota
#                                 recordnuevo.asistencia = recordantiguo.asistencia
#                                 recordnuevo.sinasistencia = recordantiguo.sinasistencia
#                                 recordnuevo.fecha = recordantiguo.fecha
#                                 recordnuevo.noaplica = recordantiguo.noaplica
#                                 recordnuevo.aprobada = recordantiguo.aprobada
#                                 recordnuevo.convalidacion = recordantiguo.convalidacion
#                                 recordnuevo.pendiente = recordantiguo.pendiente
#                                 recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
#                                 recordnuevo.horas = equivalencia.asignaturamallasalto.horas
#                                 recordnuevo.valida = recordantiguo.valida
#                                 recordnuevo.validapromedio = recordantiguo.validapromedio
#                                 recordnuevo.observaciones = observaciones
#                                 recordnuevo.homologada = homologada
#                                 recordnuevo.materiaregular = recordantiguo.materiaregular
#                                 recordnuevo.materiacurso = None
#                                 recordnuevo.completonota = recordantiguo.completonota
#                                 recordnuevo.completoasistencia = recordantiguo.completoasistencia
#                                 recordnuevo.fechainicio = recordantiguo.fechainicio
#                                 recordnuevo.fechafin = recordantiguo.fechafin
#                                 recordnuevo.suficiencia = recordantiguo.suficiencia
#                                 recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
#                                 recordnuevo.reverso = False
#                                 recordnuevo.save()
#
#                             if recordnuevo:
#                                 historicos = HistoricoRecordAcademico.objects.filter(status=True,
#                                                                                      recordacademico=recordantiguo).update(
#                                     recordacademico=recordnuevo,
#                                     creditos=recordnuevo.creditos,
#                                     horas=recordnuevo.horas,
#                                     homologada=recordnuevo.homologada)
#                                 respaldo = RespaldoRecordAcademico.objects.filter(status=True,
#                                                                                   recordacademicooriginal=recordantiguo)
#
#                                 if equivalencia.asignaturamallasalto_id in [11018, 11021, 11020, 10990, 10991, 11005,
#                                                                             10993, 10997, 11007]:
#                                     if not practicaspp or not horasvinculacion:
#                                         if recordnuevo.aprobada:
#                                             profesor = None
#                                             if recordnuevo.materiaregular:
#                                                 profesor = recordnuevo.materiaregular.profesor_principal()
#                                             elif recordnuevo.materiacurso:
#                                                 profesor = recordnuevo.materiaregular.profesor()
#                                             if equivalencia.asignaturamallasalto_id == 11018:
#
#                                                 itinerarioprimero = ItinerariosMalla.objects.get(status=True,
#                                                                                                  malla_id=mallaantigua_id,
#                                                                                                  nivel_id=1)
#                                                 iprimeronuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                              malla_id=mallanueva_id,
#                                                                                              nivel_id=1)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerarioprimero).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerarioprimero).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerarioprimero).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=iprimeronuevo.horas_practicas,
#                                                             nivelmalla=iprimeronuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=iprimeronuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerarioprimero).update(
#                                                     itinerario=iprimeronuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 11021:
#                                                 itinerariosegundo = ItinerariosMalla.objects.get(status=True,
#                                                                                                  malla_id=mallaantigua_id,
#                                                                                                  nivel_id=2)
#                                                 isegundonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                              malla_id=mallanueva_id,
#                                                                                              nivel_id=2)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerariosegundo).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerariosegundo).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerariosegundo).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=isegundonuevo.horas_practicas,
#                                                             nivelmalla=isegundonuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=isegundonuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerariosegundo).update(
#                                                     itinerario=isegundonuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 11020:
#                                                 itinerariotercero = ItinerariosMalla.objects.get(status=True,
#                                                                                                  malla_id=mallaantigua_id,
#                                                                                                  nivel_id=3)
#                                                 iterceronuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                              malla_id=mallanueva_id,
#                                                                                              nivel_id=3)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerariotercero).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerariotercero).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerariotercero).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=iterceronuevo.horas_practicas,
#                                                             nivelmalla=iterceronuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=iterceronuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerariotercero).update(
#                                                     itinerario=iterceronuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 10990:
#                                                 itinerariocuarto = ItinerariosMalla.objects.get(status=True,
#                                                                                                 malla_id=mallaantigua_id,
#                                                                                                 nivel_id=4)
#                                                 icuartonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                             malla_id=mallanueva_id,
#                                                                                             nivel_id=4)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerariocuarto).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerariocuarto).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerariocuarto).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=icuartonuevo.horas_practicas,
#                                                             nivelmalla=icuartonuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=icuartonuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerariocuarto).update(
#                                                     itinerario=icuartonuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 10991:
#                                                 itinerarioquinto = ItinerariosMalla.objects.get(status=True,
#                                                                                                 malla_id=mallaantigua_id,
#                                                                                                 nivel_id=5)
#                                                 iquintonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                             malla_id=mallanueva_id,
#                                                                                             nivel_id=5)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerarioquinto).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerarioquinto).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerarioquinto).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=iquintonuevo.horas_practicas,
#                                                             nivelmalla=iquintonuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=iquintonuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerarioquinto).update(
#                                                     itinerario=iquintonuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 11005:
#                                                 itinerariosexto = ItinerariosMalla.objects.get(status=True,
#                                                                                                malla_id=mallaantigua_id,
#                                                                                                nivel_id=6)
#                                                 isextonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                            malla_id=mallanueva_id,
#                                                                                            nivel_id=6)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerariosexto).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerariosexto).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerariosexto).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=isextonuevo.horas_practicas,
#                                                             nivelmalla=isextonuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=isextonuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerariosexto).update(
#                                                     itinerario=isextonuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 10993:
#                                                 ####################################### VINCULACION ###################################################
#                                                 if not horasvinculacion:
#                                                     if inscripcion.numero_horas_proyectos_vinculacion() < 80:
#                                                         horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
#                                                         vinculacion = ParticipantesMatrices(status=True,
#                                                                                             matrizevidencia_id=2,
#                                                                                             proyecto_id=601,
#                                                                                             inscripcion=inscripcion,
#                                                                                             horas=horasfalta,
#                                                                                             registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                                                             registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                                                             estado=1
#                                                                                             )
#                                                         vinculacion.save()
#                                                 ######################################################################################################
#                                                 itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
#                                                                                                  malla_id=mallaantigua_id,
#                                                                                                  nivel_id=7)
#                                                 iseptimonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                              malla_id=mallanueva_id,
#                                                                                              nivel_id=7)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerarioseptimo).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerarioseptimo).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerarioseptimo).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=iseptimonuevo.horas_practicas,
#                                                             nivelmalla=iseptimonuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=iseptimonuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerarioseptimo).update(
#                                                     itinerario=iseptimonuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 10997:
#                                                 itinerariooctavo = ItinerariosMalla.objects.get(status=True,
#                                                                                                 malla_id=mallaantigua_id,
#                                                                                                 nivel_id=8)
#                                                 ioctavonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                             malla_id=mallanueva_id,
#                                                                                             nivel_id=8)
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerariooctavo).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerariooctavo).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerariooctavo).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=ioctavonuevo.horas_practicas,
#                                                             nivelmalla=ioctavonuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=ioctavonuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerariooctavo).update(
#                                                     itinerario=ioctavonuevo)
#
#                                             if equivalencia.asignaturamallasalto_id == 11007:
#                                                 ####################################### VINCULACION ###################################################
#                                                 if not horasvinculacion:
#                                                     if inscripcion.numero_horas_proyectos_vinculacion() < 160:
#                                                         horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
#                                                         vinculacion = ParticipantesMatrices(status=True,
#                                                                                             matrizevidencia_id=2,
#                                                                                             proyecto_id=601,
#                                                                                             inscripcion=inscripcion,
#                                                                                             horas=horasfalta,
#                                                                                             registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                                                             registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                                                             estado=1
#                                                                                             )
#                                                         vinculacion.save()
#                                                 ######################################################################################################
#                                                 itinerarionoveno = ItinerariosMalla.objects.get(status=True,
#                                                                                                 malla_id=mallaantigua_id,
#                                                                                                 nivel_id=9)
#                                                 inovenonuevo = ItinerariosMalla.objects.get(status=True,
#                                                                                             malla_id=mallanueva_id,
#                                                                                             nivel_id=9)
#
#                                                 practica = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                     status=True,
#                                                     inscripcion=inscripcion,
#                                                     estadosolicitud__in=[
#                                                         1, 2, 4,
#                                                         5, 6],
#                                                     itinerariomalla=itinerarionoveno).exists()
#                                                 practicarechazada = False
#                                                 if not practica:
#                                                     practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
#                                                         status=True,
#                                                         inscripcion=inscripcion,
#                                                         estadosolicitud=3,
#                                                         itinerariomalla=itinerarionoveno).exists()
#
#                                                 if not practica or practicarechazada:
#                                                     if not PracticasPreprofesionalesInscripcion.objects.filter(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             actividad__itinerariomalla=itinerarionoveno).exists():
#                                                         nuevapractica = PracticasPreprofesionalesInscripcion(
#                                                             status=True,
#                                                             inscripcion=inscripcion,
#                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             numerohora=inovenonuevo.horas_practicas,
#                                                             nivelmalla=inovenonuevo.nivel,
#                                                             tiposolicitud=1,
#                                                             estadosolicitud=2,
#                                                             tipo=1,
#                                                             itinerariomalla=inovenonuevo,
#                                                             supervisor=profesor,
#                                                             tutorunemi=profesor,
#                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             tipoinstitucion=1,
#                                                             sectoreconomico=6,
#                                                             empresaempleadora_id=3,
#                                                             culminada=True,
#                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
#                                                             lugarpractica_id=2,
#                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
#                                                             )
#                                                         nuevapractica.save()
#
#                                                 ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
#                                                     status=True,
#                                                     itinerario=itinerarionoveno).update(
#                                                     itinerario=inovenonuevo)
#
#                                 if not respaldo.exists():
#                                     respaldorecord = RespaldoRecordAcademico(
#                                         recordacademicooriginal=recordantiguo,
#                                         recordacademiconuevo=recordnuevo
#                                     )
#                                     respaldorecord.save()
#                                 else:
#                                     respaldorecord = respaldo[0]
#                                     respaldorecord.recordacademiconuevo = recordnuevo
#                                     respaldorecord.save()
#                                 print(u"Record actualizado %s" % recordnuevo)
#
#                         else:
#                             hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
#                             fila += 1
#
#                     practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
#                     hojadestino.write(fila, 3, practicasppf, fuentenormal)
#                     horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
#                     hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
#
#                     fila += 1
#
#                     time.sleep(1)
#
#                 else:
#                     sin_matricula.append(inscripcion)
#
#         lin += 1
#
#     libdestino.save(output_folder + libre_origen)
#     print(output_folder + libre_origen)
#     print("Proceso finalizado. . .")
#     print(str(sin_matricula))
#
#
# except Exception as ex:
#     noti = Notificacion(titulo='Error',
#                         cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
#                         destinatario_id=29898, url="",
#                         prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                         tipo=2, en_proceso=False, error=True)
#     noti.save()
#     transaction.set_rollback(True)
#     print('error: %s' % ex)
#     hojadestino.write(fila, 3, str(ex))
#     fila += 1






# import openpyxl
# import xlwt
# from django.db import transaction
# from xlwt import easyxf
# import concurrent.futures
# from inno.funciones import haber_cumplido_horas_creditos_practicas_preprofesionales, \
#     haber_cumplido_horas_creditos_vinculacion
# from sga.models import *

import openpyxl
import xlwt
from django.db import transaction
from xlwt import easyxf
import concurrent.futures
from inno.funciones import haber_cumplido_horas_creditos_practicas_preprofesionales, \
    haber_cumplido_horas_creditos_vinculacion
from sga.models import *

try:
    libre_origen = '/homologacion_tics_rezagados_final.xls'
    fuentecabecera = easyxf(
        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
    fuentenormal = easyxf(
        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
    # liborigen = xlrd.open_workbook(output_folder + libre_origen)
    libdestino = xlwt.Workbook()
    hojadestino = libdestino.add_sheet('Sheet1')
    fil = 0
    columnas = [(u"CEDULA", 6000),
                (u"APELLIDOS Y NOMBRES", 6000),
                (u"OBSERVACIÓN", 6000),
                (u"HORAS PRACTICAS", 6000),
                (u"HORAS VINCULACION", 6000),
                (u"OBSERVACIÓN", 6000)
                ]
    for col_num in range(len(columnas)):
        hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
        hojadestino.col(col_num).width = columnas[col_num][1]
    fila = 1
    titulacion = 0

    lin = 0
    # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
    miarchivo = openpyxl.load_workbook('tecnologia.xlsx')
    # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

    ws = miarchivo.get_sheet_by_name("d1")
    worksheet = ws
    c = 0
    cont = 0
    periodo_id = 224
    carrera_id = 133
    mallaantigua_id = 202
    mallanueva_id = 478

    for row in worksheet.iter_rows(min_row=0):
        if lin >= 0:
            currentValues, cadena = [], ''
            for cell in row:
                cadena += str(cell.value) + ' '
                currentValues.append(str(cell.value))
            identificacion = currentValues[0]

            if not identificacion:
                break

            with transaction.atomic():
                inscripcion = Inscripcion.objects.filter(status=True, carrera__id=carrera_id,
                                                         pk=identificacion).first()

                cont += 1
                # matricula.pasoayuda = True
                # matricula.save()
                print(u"%s - %s" % (inscripcion, cont))
                # inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                # hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                   malla_id=mallaantigua_id).exists():
                    imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                            asignaturamalla__malla_id=mallaantigua_id).order_by(
                    'asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,materiaregular__asignaturamalla_id=equivalencia.asignaturamalla_id, asignaturamalla__isnull=True).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():

                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id if recordantiguo.asignaturamalla else None,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[
                                0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(
                                recordacademico=recordnuevo,
                                creditos=recordnuevo.creditos,
                                horas=recordnuevo.horas,
                                homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                              recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                fila += 1

                time.sleep(1)

        lin += 1

    libdestino.save(output_folder + libre_origen)
    print(output_folder + libre_origen)
    print("Proceso finalizado. . .")

except Exception as ex:
    noti = Notificacion(titulo='Error',
                        cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                       sys.exc_info()[
                                                                                           -1].tb_lineno),
                        destinatario_id=29898, url="",
                        prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                        tipo=2, en_proceso=False, error=True)
    noti.save()
    transaction.set_rollback(True)
    print('error: %s' % ex)
    hojadestino.write(fila, 3, str(ex))
    fila += 1