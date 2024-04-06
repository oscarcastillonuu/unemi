import json
import os
import io
import time
import threading
import zipfile
import csv
from decimal import ROUND_DOWN

import xlsxwriter
import xlwt
import pandas as pd
from django.contrib.admin.models import LogEntry
from django.core.files.base import ContentFile
from django.db.models.functions import ExtractYear, ExtractMonth, Concat
from django.db.models import F, Value
from django.template.loader import render_to_string
from unidecode import unidecode
from xlwt import *
from webpush import send_user_notification
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.forms import model_to_dict
from django.http import HttpResponse
from django.shortcuts import redirect
from openpyxl import load_workbook, workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
# from cronrelacionados import enviar_mensaje_bot_telegram
from bd.models import LogEntryLogin, PythonProcess
from core.firmar_documentos_ec import JavaFirmaEc
from inno.funciones import no_adeudar_institucion, haber_cumplido_horas_creditos_vinculacion, \
    haber_cumplido_horas_creditos_practicas_preprofesionales, haber_aprobado_modulos_computacion, \
    haber_aprobado_modulos_ingles, ficha_estudiantil_actualizada_completa, asignaturas_aprobadas_primero_ultimo_nivel, \
    asignaturas_aprobadas_primero_septimo_nivel, estar_matriculado_todas_asignaturas_ultimo_periodo_academico, \
    tiene_certificacion_segunda_lengua_sin_aprobar_director_carrera, \
    tiene_certificacion_segunda_lengua_aprobado_director_carrera, asignaturas_aprobadas_primero_penultimo_nivel
from inno.models import MateriaAsignadaPlanificacionSedeVirtualExamen
from investigacion.funciones import FORMATOS_CELDAS_EXCEL
from posgrado.models import Contrato, InscripcionCohorte, DetalleAprobacionContrato
from sagest.funciones import encrypt_id, generar_acta_constatacion_reportlab
from sagest.models import DistributivoPersona, Departamento, DistributivoPersonaHistorial, PersonaDepartamentoFirmas, \
    RegimenLaboral, NivelOcupacional, ModalidadLaboral, EstadoPuesto, \
    EstructuraProgramatica, EscalaOcupacional, DenominacionPuesto, PuestoAdicional, Pago, Rubro, TipoOtroRubro, \
    DetalleBajaActivo, PeriodoConstatacionAF, ActivoFijo, ConstatacionFisica, Factura, ArchivoFacturaEsigef, \
    ArchivoFacturaEsigefDetalle, CuentaContable, KardexInventario, Producto, RegistroDecimo, PazSalvo
from sga.commonviews import traerNotificaciones
from sga.funciones import generar_nombre, nivel_enletra_malla, paralelo_enletra_nivel, ingreso_total_hogar_rangos, \
    remover_caracteres_tildes_unicode, calculate_username, generar_usuario, notificacion3, notificacion,log
from sga.models import *
from settings import MEDIA_ROOT, MEDIA_URL, DEBUG, SITE_POPPLER, PROFESORES_GROUP_ID, EMAIL_INSTITUCIONAL_AUTOMATICO, \
    EMAIL_DOMAIN
from sga.templatetags.sga_extras import encrypt
from socioecon.models import FichaSocioeconomicaINEC
from sga.funcionesxhtml2pdf import conviert_html_to_pdf_name_save, conviert_html_to_pdfsavecertificados, \
    conviert_html_to_pdfsaveqr_generico, conviert_html_to_pdfsaveqrtitulo, convert_html_to_pdf, \
    conviert_html_to_pdfsave_generic_lotes
from django.db import transaction
from django.shortcuts import render
from pdf2image import convert_from_bytes
from num2words import num2words
import uuid
import pyqrcode
from core.firmar_documentos import obtener_posicion_y
from django.core.files import File as DjangoFile

def titulo_4_nivel(persona):
    try:
        titulo = ''
        if Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).exists():
            titulos = Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).order_by('-id')
            for idx, t in enumerate(titulos):
                titulo += t.titulo.nombre
                if idx < len(titulos) - 1:
                    titulo += ', '
        return titulo
    except Exception as ex:
        pass

def universidad_titulo_4_nivel(persona):
    try:
        titulo = ''
        if Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).exists():
            titulos = Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).order_by('-id')
            for idx, t in enumerate(titulos):
                if t.institucion:
                    titulo += t.institucion.nombre
                else:
                    titulo += 'NO REGISTRA'
                if idx < len(titulos) - 1:
                    titulo += ', '
        return titulo
    except Exception as ex:
        pass

def trabaja_en(persona):
    try:
        trabajo1 = 'NO REGISTRA'
        if PersonaSituacionLaboral.objects.filter(status=True, persona=persona).exists():
            trabajo = PersonaSituacionLaboral.objects.filter(status=True, persona=persona).first()
            if trabajo.lugartrabajo != '':
                trabajo1 = trabajo.lugartrabajo
            elif trabajo.negocio != '':
                trabajo1 = trabajo.negocio
        return trabajo1
    except Exception as ex:
        pass

def tiene_postulacion(persona):
    try:
        estado = 'NO'
        if InscripcionCohorte.objects.filter(status=True, inscripcionaspirante__persona=persona).exists():
            estado = 'SI'
        return estado
    except Exception as ex:
        pass

def tiene_postulacion_descripcion(persona):
    try:
        estado = ''
        if InscripcionCohorte.objects.filter(status=True, inscripcionaspirante__persona=persona).exists():
            postulaciones = InscripcionCohorte.objects.filter(status=True, inscripcionaspirante__persona=persona).order_by('-id')
            for idx, t in enumerate(postulaciones):
                estado += t.cohortes.__str__()
                if idx < len(estado) - 1:
                    estado += ', '
        return estado
    except Exception as ex:
        pass

def total_creditos_aprobadas(inscripcion):
    return null_to_numeric(inscripcion.recordacademico_set.filter(valida=True, status=True, asignaturamalla__isnull=False, noaplica=False, aprobada=True).aggregate(creditos=Sum('creditos'))['creditos'])

def total_creditos(inscripcion):
    inscripcionmalla = inscripcion.malla_inscripcion()
    return null_to_numeric(AsignaturaMalla.objects.filter(status=True, malla=inscripcionmalla.malla).aggregate(creditos=Sum('creditos'))['creditos'])

def num_segunda_mat(matricula):
    try:
        cantidad = 0
        if MateriaAsignada.objects.filter(status=True, matriculas=2, matricula=matricula).exists():
            cantidad = MateriaAsignada.objects.filter(status=True, matriculas=2, matricula=matricula).distinct().count()
        return cantidad
    except Exception as ex:
        pass

def num_tercera_mat(matricula):
    try:
        cantidad = 0
        if MateriaAsignada.objects.filter(status=True, matriculas=3, matricula=matricula).exists():
            cantidad = MateriaAsignada.objects.filter(status=True, matriculas=3, matricula=matricula).distinct().count()
        return cantidad
    except Exception as ex:
        pass

def total_horas_aprobadas(inscripcion):
    return null_to_numeric(inscripcion.recordacademico_set.filter(valida=True, status=True, asignaturamalla__isnull=False, noaplica=False, aprobada=True).aggregate(horas=Sum('horas'))['horas'])

def total_horas_ma(inscripcion):
    inscripcionmalla = inscripcion.malla_inscripcion()
    return null_to_numeric(AsignaturaMalla.objects.filter(status=True, malla=inscripcionmalla.malla).aggregate(horas=Sum('horas'))['horas'])

def es_extranjero(persona):
    try:
        estado = 'NO'
        if persona.paisnacimiento:
            if persona.paisnacimiento.id != 1:
                estado = 'SI'
        return estado
    except Exception as ex:
        pass

def inicio_primer_modulo(matricula):
    try:
        fecha = ''
        if MateriaAsignada.objects.filter(status=True, matricula=matricula).exists():
            ma = MateriaAsignada.objects.filter(status=True, matricula=matricula).first()
            fecha = ma.materia.inicio
        return fecha
    except Exception as ex:
        pass

def fin_ultimo_modulo(matricula):
    try:
        fecha = ''
        inscripcionmalla = matricula.inscripcion.malla_inscripcion()
        canti = AsignaturaMalla.objects.filter(status=True, malla=inscripcionmalla.malla).count()

        if MateriaAsignada.objects.filter(status=True, matricula=matricula, materia__cerrado=True).exists():
            cantima = MateriaAsignada.objects.filter(status=True, matricula=matricula, materia__cerrado=True).count()
            if cantima >= canti:
                ma = MateriaAsignada.objects.filter(status=True, matricula=matricula, materia__cerrado=True).order_by('-id').first()
                fecha = ma.materia.inicio
        return fecha
    except Exception as ex:
        pass

def periodo_matricula(inscripcion):
    try:
        periodo = None
        if MateriaAsignada.objects.filter(status=True, matricula__inscripcion=inscripcion, matricula__inscripcion__status=True).exists():
            mate = MateriaAsignada.objects.filter(status=True, matricula__inscripcion=inscripcion, matricula__inscripcion__status=True).first()
            return mate.matricula.nivel.periodo
        return periodo
    except Exception as ex:
        pass

def mecanismo_titulacion(inscripcion):
    try:
        tema = ''
        if TemaTitulacionPosgradoMatricula.objects.filter(status=True, matricula__inscripcion=inscripcion, mecanismotitulacionposgrado__status=True).exists():
            tema = TemaTitulacionPosgradoMatricula.objects.filter(status=True, matricula__inscripcion=inscripcion, mecanismotitulacionposgrado__status=True).order_by('-id').first()
            return tema.mecanismotitulacionposgrado.nombre
    except Exception as ex:
        pass

class reporte_matriculados_background(threading.Thread):

    def __init__(self, request, data, notiid):
        self.request = request
        self.data = data
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'titulacion')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "matriculados_proceso_titulacion_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'titulacion', nombre_archivo)

        try:
            tipo = ''
            per = ''

            borders = Borders()
            borders.left = 1
            borders.right = 1
            borders.top = 1
            borders.bottom = 1
            __author__ = 'Unemi'
            title = easyxf('font: name Arial, bold on , height 240; alignment: horiz centre')
            normal = easyxf('font: name Arial , height 150; alignment: horiz left')
            encabesado_tabla = easyxf('font: name Arial , bold on , height 150; alignment: horiz left')
            normalc = easyxf('font: name Arial , height 150; alignment: horiz center')
            subtema = easyxf('font: name Arial, bold on , height 180; alignment: horiz left')
            normalsub = easyxf('font: name Arial , height 180; alignment: horiz left')
            style1 = easyxf(num_format_str='D-MMM-YY')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            normal.borders = borders
            normalc.borders = borders
            normalsub.borders = borders
            subtema.borders = borders
            encabesado_tabla.borders = borders
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('exp_xls_post_part')

            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=Matriculados_Proceso_Titulación_Todos_Periodos ' + tipo + random.randint(
                1, 10000).__str__() + '.xls'
            # ws.write_merge(6, 6, 0, 1, 'FACULTAD', subtema)
            # ws.write_merge(7, 7, 0, 1, 'CARRERA', subtema)
            #
            # ws.write_merge(6, 6, 2, 6, grupo.facultad.nombre if grupo else alter.facultad.nombre, normalsub)
            # ws.write_merge(7, 7, 2, 6, carrera.nombre if carrera else  alter.carrera.nombre, normalsub)

            ws.col(0).width = 1000
            ws.col(1).width = 2500
            ws.col(2).width = 4000
            ws.col(3).width = 3000
            ws.col(4).width = 3000
            ws.col(5).width = 4000
            ws.col(6).width = 4000
            ws.col(7).width = 3200
            ws.col(8).width = 5000
            ws.col(9).width = 4000
            ws.col(10).width = 4000
            ws.col(11).width = 4000
            ws.col(12).width = 4000
            ws.col(13).width = 4000
            ws.col(14).width = 4000
            ws.col(15).width = 4000
            ws.col(16).width = 4000
            ws.col(17).width = 4000
            ws.col(18).width = 4000
            ws.col(19).width = 4000
            ws.col(20).width = 4000
            ws.col(21).width = 4000
            ws.col(22).width = 4000
            ws.col(23).width = 4000
            ws.col(24).width = 4000
            ws.col(25).width = 4000

            row_num = 0
            ws.write(row_num, 0, "Nº", encabesado_tabla)
            ws.write(row_num, 1, "FACULTAD", encabesado_tabla)
            ws.write(row_num, 2, "CARRERA", encabesado_tabla)
            ws.write(row_num, 3, u"PARALELO", encabesado_tabla)
            ws.write(row_num, 4, u"CEDULA", encabesado_tabla)
            ws.write(row_num, 5, u"PRIMER APELLIDO", encabesado_tabla)
            ws.write(row_num, 6, u"SEGUNDO APELLIDO", encabesado_tabla)
            ws.write(row_num, 7, u"NOMBRES", encabesado_tabla)
            ws.write(row_num, 8, u"MECANISMO TITULACIÓN", encabesado_tabla)
            ws.write(row_num, 9, u"ESTADO. MATRICULA", encabesado_tabla)
            ws.write(row_num, 10, u"Nº MATRICULA", encabesado_tabla)
            ws.write(row_num, 11, u"TUTOR", encabesado_tabla)
            ws.write(row_num, 12, u"INTEGRANTE GRUPO", encabesado_tabla)
            ws.write(row_num, 13, u"NUMERO DE HORAS TUTORIAS", encabesado_tabla)
            ws.write(row_num, 14, u"TEMA", encabesado_tabla)
            ws.write(row_num, 15, u"ESTADO FINAL PREVIO SUSTENTACION", encabesado_tabla)
            ws.write(row_num, 16, 'ESTADO PARA SUSTENTACION', encabesado_tabla)
            ws.write(row_num, 17, 'PRESIDENTE', encabesado_tabla)
            ws.write(row_num, 18, 'SECRETARIO', encabesado_tabla)
            ws.write(row_num, 19, 'INTEGRANTE', encabesado_tabla)
            ws.write(row_num, 20, u"PRUEBA TEÓRICA", encabesado_tabla)
            ws.write(row_num, 21, u"NOTA TRABAJO TITULACIÓN PROYECTO", encabesado_tabla)
            ws.write(row_num, 22, u"NOTA FINAL PERIODO", encabesado_tabla)
            ws.write(row_num, 23, u"ESTADO PRUEBA TEÓRICA", encabesado_tabla)
            ws.write(row_num, 24, u"ESTADO INVESTIGACIÓN", encabesado_tabla)
            ws.write(row_num, 25, u"PERIODOS", encabesado_tabla)
            ws.write(row_num, 26, u"FECHA SUSTENTACIÓN", encabesado_tabla)

            date_format = xlwt.XFStyle()
            date_format.num_format_str = 'yyyy/mm/dd'
            data = {}

            listamatriculados = MatriculaTitulacion.objects.select_related().filter(Q(status=True),
                                                                                    Q(alternativa__status=True),
                                                                                    (
                                                                                            Q(estado=1) | Q(
                                                                                        estado=9) | Q(
                                                                                        estado=10))).order_by(
                'alternativa__facultad', 'alternativa__carrera', 'alternativa',
                'inscripcion__persona__apellido1')
            row_num = 1
            i = 0
            for matriculados in listamatriculados:

                if matriculados.id == 1954:
                    c = matriculados.id

                campo0 = matriculados.id
                campo1 = matriculados.inscripcion.persona.apellido1
                campo2 = matriculados.inscripcion.persona.apellido2
                campo3 = matriculados.inscripcion.persona.nombres
                campo4 = matriculados.inscripcion.persona.sexo.nombre
                perfilinscripcion = PerfilInscripcion.objects.get(persona=matriculados.inscripcion.persona,
                                                                  status=True)
                campo5 = perfilinscripcion.raza.nombre
                campo6 = matriculados.inscripcion.persona.nacimiento.strftime('%Y/%m/%d')
                campo7 = matriculados.inscripcion.persona.pais.nombre
                campo8 = matriculados.inscripcion.persona.nacionalidad
                campo9 = matriculados.inscripcion.persona.pais.nombre
                campo10 = matriculados.inscripcion.persona.provincia.nombre
                campo11 = ""
                if not matriculados.inscripcion.persona.canton is None:
                    campo11 = matriculados.inscripcion.persona.canton.nombre

                campo12 = ''
                if not matriculados.inscripcion.fechainicioprimernivel is None:
                    campo12 = matriculados.inscripcion.fechainicioprimernivel.strftime('%Y/%m/%d')
                campo13 = ''
                if Egresado.objects.filter(inscripcion=matriculados.inscripcion, status=True).exists():
                    egresado = Egresado.objects.get(inscripcion=matriculados.inscripcion, status=True)
                    campo13 = egresado.fechaegreso.strftime('%Y/%m/%d')
                campo14 = matriculados.duracion_estudio()
                campo15 = ''
                campo16 = matriculados.inscripcion.especialidad.nombre if not matriculados.inscripcion.especialidad is None else ''
                campo17 = matriculados.inscripcion.colegio
                campo18 = ''
                listaestudios_previos = []
                if EstudioInscripcion.objects.filter(status=True,
                                                     persona=matriculados.inscripcion.persona).exists():
                    estudios = EstudioInscripcion.objects.filter(status=True,
                                                                 persona=matriculados.inscripcion.persona)
                    for estudios_pervios in estudios:
                        listaestudios_previos.append(estudios_pervios.carrera)  # estudiosde estudios previos'
                campo19 = ''  # carrera estudios previos'
                campo20 = ''  # tiempo de duracion de reconocimiento
                campo21 = ''
                campo22 = ''
                campo23 = ''
                campo24 = ''
                campo28 = ''
                campo36 = ''
                campo37 = ''
                if Graduado.objects.filter(inscripcion=matriculados.inscripcion, status=True).exists():
                    graduado = Graduado.objects.get(inscripcion=matriculados.inscripcion, status=True)
                    campo22 = graduado.fechaactagrado.strftime('%Y/%m/%d') if graduado.fechaactagrado else ''
                    campo23 = graduado.numeroactagrado if graduado.numeroactagrado else ''
                    campo24 = graduado.fecharefrendacion.strftime(
                        '%Y/%m/%d') if graduado.fecharefrendacion else ''
                    campo28 = graduado.promediotitulacion if graduado.promediotitulacion else ''
                    # campo36 = graduado.notafinal if graduado.notafinal else ''
                    campo36 = graduado.notagraduacion if graduado.notagraduacion else ''
                    campo37 = graduado.registro if graduado.registro else ''
                    # if matriculados.complexivodetallegrupo_set.filter(status=True):
                    #     notafinal = matriculados.notafinalcomplexivo() if matriculados.notafinalcomplexivo() else ''
                    # else:
                    #     notafinal = 0
                campo25 = matriculados.alternativa.tipotitulacion.nombre
                campo26 = ''  # link de tesis
                campo27 = matriculados.inscripcion.promedio_record()
                campo29 = ''  # nombre del rector
                campo30 = ''  # observaciones
                campo31 = matriculados.inscripcion.carrera.codigo
                campo32 = matriculados.inscripcion.persona.usuario.username
                campo33 = matriculados.inscripcion.carrera.nombre
                horas_paracticas = PracticasPreprofesionalesInscripcion.objects.filter(
                    inscripcion=matriculados.inscripcion, status=True, culminada=True)
                totalhoras = 0
                if horas_paracticas.exists():
                    for practicas in horas_paracticas:
                        if practicas.tiposolicitud == 3:
                            if practicas.horahomologacion:
                                totalhoras += practicas.horahomologacion
                        else:
                            totalhoras += practicas.numerohora
                campo34 = totalhoras
                lis = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True,
                                                           proyecto__status=True,
                                                           inscripcion_id=matriculados.inscripcion.id)
                horas_vinculacion = 0
                for vin in lis:
                    horas_vinculacion += vin.horas
                campo35 = horas_vinculacion
                campo38 = matriculados.inscripcion.persona.telefono
                campo39 = u'%s' % matriculados.inscripcion.persona.sangre if not matriculados.inscripcion.persona.sangre is None else ''
                campo40 = matriculados.inscripcion.persona.email
                campo41 = str(matriculados.inscripcion.mi_nivel())
                campo42 = ''
                if not matriculados.inscripcion.fechainicioconvalidacion is None:
                    campo42 = matriculados.inscripcion.fechainicioconvalidacion.strftime('%Y/%m/%d')
                campo43 = ''
                if not matriculados.inscripcion.fechainicioprimernivel is None:
                    campo43 = matriculados.inscripcion.fechainicioprimernivel.strftime('%Y/%m/%d')
                campo44 = str(matriculados.inscripcion.persona.estado_civil())
                campo45 = matriculados.inscripcion.persona.emailinst
                campo46 = 'NINGUNA'
                if not perfilinscripcion.tipodiscapacidad == None:
                    campo46 = perfilinscripcion.tipodiscapacidad.nombre
                campo47 = perfilinscripcion.porcientodiscapacidad
                campo48 = ''  # hogar
                campo49 = ''
                campo50 = ''
                lista_discapacidad = []
                datos_personal = PersonaDatosFamiliares.objects.filter(persona=matriculados.inscripcion.persona,
                                                                       status=True)
                if datos_personal.exists():
                    for familiardiscapasidad in datos_personal:
                        if familiardiscapasidad.tienediscapacidad:
                            lista_discapacidad.append(
                                [familiardiscapasidad.nombre, familiardiscapasidad.parentesco.nombre])
                campo51 = ''
                campo52 = ''
                campo53 = ''
                campo54 = ''
                campo55 = ''
                campo56 = ''
                campo57 = ''
                campo58 = ''
                campo59 = ''
                campo60 = ''
                campo61 = ''
                campo62 = ''
                campo63 = ''
                campo64 = ''
                campo65 = ''
                campo66 = ''
                campo67 = ''
                campo68 = ''
                campo69 = ''
                campo70 = ''
                campo71 = ''
                campo72 = ''
                campo73 = ''
                campo74 = ''
                campo75 = ''
                campo76 = ''
                campo77 = ''
                campo78 = ''
                campo79 = ''
                campo80 = ''
                campo81 = ''
                campo82 = ''
                campo83 = ''
                campo84 = ''
                campo85 = ''
                campo86 = ''
                campo87 = ''
                campo88 = ''
                campo89 = ''
                campo90 = ''
                campo91 = ''
                campo92 = ''
                campo93 = ''
                campo94 = ''
                campo95 = ''
                if FichaSocioeconomicaINEC.objects.filter(persona=matriculados.inscripcion.persona,
                                                          status=True).exists():
                    ficha = matriculados.inscripcion.persona.mi_ficha()
                    campo51 = ''
                    if not ficha.tipohogar is None:
                        campo51 = str(ficha.tipohogar.nombre)
                    campo52 = 'No'
                    if ficha.escabezafamilia:
                        campo52 = 'Si'
                    campo53 = 'No'
                    campo54 = ''
                    if ficha.esdependiente:
                        campo53 = 'Si'
                        campo54 = str(ficha.personacubregasto)
                    campo55 = ''
                    if not ficha.tipovivienda is None:
                        campo55 = ficha.tipovivienda.nombre
                    campo56 = ''
                    if not ficha.tipoviviendapro is None:
                        campo56 = ficha.tipoviviendapro.nombre
                    campo57 = ''
                    if not ficha.materialpared is None:
                        campo57 = ficha.materialpared.nombre
                    campo58 = ''
                    if not ficha.materialpiso is None:
                        campo58 = ficha.materialpiso.nombre
                    campo59 = ''
                    if not ficha.cantbannoducha is None:
                        campo59 = ficha.cantbannoducha.nombre
                    campo60 = ''
                    if not ficha.tiposervhig is None:
                        campo60 = ficha.tiposervhig.nombre
                    campo61 = 'No'
                    if ficha.tienetelefconv:
                        campo61 = 'Si'
                    campo62 = 'No'
                    if ficha.tienecocinahorno:
                        campo62 = 'Si'
                    campo63 = 'No'
                    if ficha.tienerefrig:
                        campo63 = 'Si'
                    campo64 = 'No'
                    if ficha.tienelavadora:
                        campo64 = 'Si'
                    campo65 = 'No'
                    if ficha.tienemusica:
                        campo65 = 'Si'
                    campo66 = ''
                    if not ficha.canttvcolor is None:
                        campo66 = ficha.canttvcolor.nombre
                    campo67 = ''
                    if not ficha.cantvehiculos is None:
                        campo67 = ficha.cantvehiculos.nombre
                    campo68 = 'No'
                    if ficha.tienesala:
                        campo68 = 'Si'
                    campo69 = 'No'
                    if ficha.tienecomedor:
                        campo69 = 'Si'
                    campo70 = 'No'
                    if ficha.tienecocina:
                        campo70 = 'Si'
                    campo71 = 'No'
                    if ficha.tienebanio:
                        campo71 = 'Si'
                    campo72 = 'No'
                    if ficha.tieneluz:
                        campo72 = 'Si'
                    campo73 = 'No'
                    if ficha.tieneagua:
                        campo73 = 'Si'
                    campo74 = 'No'
                    if ficha.tienealcantarilla:
                        campo74 = 'Si'
                    campo75 = 'No'
                    if ficha.tienetelefono:
                        campo75 = 'Si'
                    campo76 = 'No'
                    if ficha.tienetelefono:
                        campo76 = 'Si'
                        #    enfermedades erideritarias
                    campo77 = 'No'
                    if ficha.tienediabetes:
                        campo77 = 'Si'
                    campo78 = 'No'
                    if ficha.tienehipertencion:
                        campo78 = 'Si'
                    campo79 = 'No'
                    if ficha.tienecancer:
                        campo79 = 'Si'
                    campo80 = 'No'
                    if ficha.tienealzheimer:
                        campo80 = 'Si'
                    campo81 = 'No'
                    if ficha.tienevitiligo:
                        campo81 = 'Si'
                    campo82 = 'No'
                    if ficha.tienedesgastamiento:
                        campo82 = 'Si'
                    campo83 = 'No'
                    if ficha.tienepielblanca:
                        campo83 = 'Si'
                    campo84 = ''
                    if not ficha.otrasenfermedades is None:
                        campo84 = ficha.otrasenfermedades
                    campo85 = 'No'
                    if ficha.tienesida:
                        campo85 = 'Si'
                    campo86 = ''
                    if not ficha.enfermedadescomunes is None:
                        campo86 = ficha.enfermedadescomunes
                    campo87 = ''
                    if not ficha.niveljefehogar is None:
                        campo87 = ficha.niveljefehogar.nombre
                    campo88 = 'No'
                    if ficha.alguienafiliado:
                        campo88 = 'Si'
                    campo89 = 'No'
                    if ficha.alguienseguro:
                        campo89 = 'Si'
                    campo90 = ''
                    if not ficha.ocupacionjefehogar is None:
                        campo90 = ficha.ocupacionjefehogar.nombre
                    campo91 = 'No'
                    if ficha.tieneinternet:
                        campo91 = 'Si'
                    campo92 = 'No'
                    if ficha.tienedesktop:
                        campo92 = 'Si'
                    campo93 = 'No'
                    if ficha.tienelaptop:
                        campo93 = 'Si'
                    campo94 = ''  # genero
                    campo95 = 'No'
                    if matriculados.inscripcion.persona.lgtbi:
                        campo95 = 'Si'
                campo96 = 'No'
                if matriculados.estadogestacion_set.filter(estadogestacion=True, status=True).exists():
                    campo96 = 'Si'
                campo97 = 'REPROBADO' if matriculados.reprobo_examen_complexivo() else matriculados.get_estado_display()
                matriculados.reprobo_examen_complexivo()
                telefonoconv = ''
                if not matriculados.inscripcion.persona.telefono_conv == '':
                    telefonoconv = matriculados.inscripcion.persona.telefono_conv
                i += 1
                campo98 = ''
                campo99 = ''
                if matriculados.inscripcion.persona.fichasocioeconomicainec_set.exists():
                    campo98 = '%s' % matriculados.inscripcion.persona.fichasocioeconomicainec_set.all()[
                        0].grupoeconomico.nombre_corto()
                    campo99 = '%s' % matriculados.inscripcion.persona.fichasocioeconomicainec_set.all()[
                        0].grupoeconomico.id
                ws.write(row_num, 0, i, normal)
                # ws.write(row_num, 1, campo0, normal)
                ws.write(row_num, 1, matriculados.inscripcion.coordinacion.nombre, normal)
                ws.write(row_num, 2, matriculados.alternativa.carrera.nombre, normal)
                ws.write(row_num, 3, matriculados.alternativa.paralelo, normal)
                ws.write(row_num, 4, matriculados.inscripcion.persona.cedula, normal)
                ws.write(row_num, 5, campo1, normal)
                ws.write(row_num, 6, campo2, normal)
                ws.write(row_num, 7, campo3, normal)

                ws.write(row_num, 8, campo25, normal)

                # if len(lista_discapacidad) > 0:
                #     if not len(lista_discapacidad) == 1:
                #         separar = ""
                #         for lis in lista_discapacidad:
                #             campo49 = str(campo49) + separar + lis[0]
                #             campo50 = str(campo50) + separar + lis[1]
                #             separar = " - "
                #         ws.write(row_num, 54, campo49, normal)
                #         ws.write(row_num, 55, campo50, normal)
                #     else:
                #         for lista1 in lista_discapacidad:
                #             campo49 = lista1[0]
                #             campo50 = lista1[1]
                #             ws.write(row_num, 54, campo49, normal)
                #             ws.write(row_num, 55, campo50, normal)
                # else:
                #     ws.write(row_num, 54, campo49, normal)
                #     ws.write(row_num, 55, campo50, normal)
                campo100 = MatriculaTitulacion.objects.values('id').filter(
                    Q(inscripcion=matriculados.inscripcion),
                    (Q(estado=1) | Q(estado=9) | Q(
                        estado=10)),
                    fechainscripcion__lte=matriculados.fechainscripcion).count().__str__()
                campo101 = matriculados.fechainscripcion
                campo102 = ''
                campo103 = ''
                campo104 = ''
                campo105 = ''
                campo106 = ''
                integrante = ""
                if ComplexivoDetalleGrupo.objects.filter(matricula=matriculados, status=True):
                    # grupo = ComplexivoDetalleGrupo.objects.get(matricula=matriculados, status=True)
                    grupo = \
                        ComplexivoDetalleGrupo.objects.filter(matricula=matriculados, status=True).order_by(
                            '-id')[0]
                    campo104 = u"%s" % str(grupo.grupo.horas_totales_tutorias_grupo())
                    campo102 = u"%s" % grupo.grupo.tematica.tutor
                    campo106 = grupo.grupo.estado_propuesta().get_estado_display() if grupo.grupo.estado_propuesta() else ""
                    if grupo.grupo.subtema:
                        campo105 = grupo.grupo.subtema
                    for com in grupo.grupo.complexivodetallegrupo_set.filter(status=True,
                                                                             matricula__alternativa__grupotitulacion__periodogrupo=matriculados.alternativa.grupotitulacion.periodogrupo):
                        if matriculados != com.matricula:
                            campo103 = integrante + u"%s" % com.matricula
                    # col = 112
                    # tit = 0
                    # if ComplexivoPropuestaPractica.objects.filter(grupo=grupo.grupo, status=True).exists():
                    #     for pro in ComplexivoPropuestaPractica.objects.filter(grupo=grupo.grupo, status=True):
                    #         texto = ''
                    #         if pro.complexivopropuestapracticaarchivo_set.filter(tipo=1, status=True).exists():
                    #             fecha = pro.complexivopropuestapracticaarchivo_set.filter(tipo=1, status=True)[0].fecha.strftime('%Y-%m-%d')
                    #             archivo = 'Si'
                    #         texto = str(pro.fecharevision) + '-' + str(pro.observacion)
                    #         ws.write(row_num, tit + col, u"%s" % str(pro.get_estado_display()), normal)
                    #         col = col + 1
                    #         ws.write(row_num, tit + col, u"%s" % texto, normal)
                    #         col = col + 1

                campo122 = ''
                campo123 = ''
                campo107 = 'PENDIENTE PARA SUSTENTAR'
                if matriculados.cumplerequisitos == 2:
                    campo107 = 'APTO PARA SUSTENTAR'
                    campo122 = matriculados.fechavalidacumplerequisitos
                    campo123 = matriculados.horavalidacumplerequisitos
                    if campo122:
                        campo122 = matriculados.fechavalidacumplerequisitos
                    else:
                        campo122 = ""

                    if campo123:
                        campo123 = matriculados.horavalidacumplerequisitos
                    else:
                        campo123 = ""
                if matriculados.cumplerequisitos == 3:
                    campo107 = 'NO APTO PARA SUSTENTAR'
                alterna = AlternativaTitulacion.objects.get(pk=matriculados.alternativa_id)

                # # FICHAS
                # ficha = 0
                # if matriculados.inscripcion.persona.nombres and matriculados.inscripcion.persona.apellido1 and matriculados.inscripcion.persona.apellido2 and matriculados.inscripcion.persona.nacimiento and matriculados.inscripcion.persona.cedula and matriculados.inscripcion.persona.nacionalidad and matriculados.inscripcion.persona.email and matriculados.inscripcion.persona.estado_civil and matriculados.inscripcion.persona.sexo:
                #     data['datospersonales'] = True
                #     ficha += 1
                # if matriculados.inscripcion.persona.paisnacimiento and matriculados.inscripcion.persona.provincianacimiento and matriculados.inscripcion.persona.cantonnacimiento and matriculados.inscripcion.persona.parroquianacimiento:
                #     data['datosnacimientos'] = True
                #     ficha += 1
                # examenfisico = matriculados.inscripcion.persona.datos_examen_fisico()
                # if matriculados.inscripcion.persona.sangre and examenfisico.peso and examenfisico.talla:
                #     data['datosmedicos'] = True
                #     ficha += 1
                # if matriculados.inscripcion.persona.pais and matriculados.inscripcion.persona.provincia and matriculados.inscripcion.persona.canton and matriculados.inscripcion.persona.parroquia and matriculados.inscripcion.persona.direccion and matriculados.inscripcion.persona.direccion2 and matriculados.inscripcion.persona.num_direccion and matriculados.inscripcion.persona.telefono_conv or matriculados.inscripcion.persona.telefono:
                #     data['datosdomicilio'] = True
                #     ficha += 1
                # perfil = matriculados.inscripcion.persona.mi_perfil()
                # if perfil.raza:
                #     data['etnia'] = True
                #     ficha += 1
                # campo108 = 'DATOS PERSONALES INCOMPLETOS'
                # if ficha == 5:
                #     campo108 = 'DATOS PERSONALES COMPLETOS'

                campo109 = 'NO TIENE PRACTICAS PREPROFESIONALES'
                totalhoras = 0
                malla = matriculados.inscripcion.inscripcionmalla_set.filter(status=True)[0].malla
                practicaspreprofesionalesinscripcion = PracticasPreprofesionalesInscripcion.objects.filter(
                    inscripcion=matriculados.inscripcion, status=True, culminada=True)
                data['malla_horas_practicas'] = malla.horas_practicas
                tienepracticaspreprofesionales = False
                if practicaspreprofesionalesinscripcion.exists():
                    for practicas in practicaspreprofesionalesinscripcion:
                        if practicas.tiposolicitud == 3:
                            totalhoras += practicas.horahomologacion if practicas.horahomologacion else 0
                        else:
                            totalhoras += practicas.numerohora
                    if totalhoras >= malla.horas_practicas:
                        campo109 = 'TIENE PRACTICAS PREPROFESIONALES'
                        campo110 = 0
                        tienepracticaspreprofesionales = True
                else:
                    totalhoras = 0
                if not tienepracticaspreprofesionales:
                    horasfaltantes = malla.horas_practicas - totalhoras
                    campo110 = horasfaltantes

                total_materias_malla = malla.cantidad_materiasaprobadas()
                cantidad_materias_aprobadas_record = matriculados.inscripcion.recordacademico_set.filter(
                    aprobada=True, status=True,
                    asignatura__in=[x.asignatura for x in
                                    malla.asignaturamalla_set.filter(status=True)]).count()
                poraprobacion = round(cantidad_materias_aprobadas_record * 100 / total_materias_malla, 0)
                data['mi_nivel'] = nivel = matriculados.inscripcion.mi_nivel()
                inscripcionmalla = matriculados.inscripcion.malla_inscripcion()
                niveles_maximos = inscripcionmalla.malla.niveles_regulares

                totalfaltantesmalla = total_materias_malla - cantidad_materias_aprobadas_record
                campo112 = totalfaltantesmalla
                campo111 = 'MALLA IN COMPLETA'
                if poraprobacion >= 100:
                    campo111 = 'MALLA COMPLETA'
                    campo112 = 0

                campo113 = 'TIENE DEUDA'
                campo114 = matriculados.inscripcion.adeuda_a_la_fecha()
                if matriculados.inscripcion.adeuda_a_la_fecha() == 0:
                    campo113 = 'NO TIENE DEUDA'
                    campo114 = 0

                campo115 = 'INGLES INCOMPLETO'
                modulo_ingles = ModuloMalla.objects.filter(malla=malla, status=True).exclude(asignatura_id=782)
                numero_modulo_ingles = modulo_ingles.count()
                lista = []
                listaid = []
                for modulo in modulo_ingles:
                    if matriculados.inscripcion.estado_asignatura(modulo.asignatura) == NOTA_ESTADO_APROBADO:
                        lista.append(modulo.asignatura.nombre)
                        listaid.append(modulo.asignatura.id)
                data['modulo_ingles_aprobados'] = lista
                campo116 = modulo_ingles.exclude(asignatura_id__in=[int(i) for i in listaid]).count()
                if numero_modulo_ingles == len(listaid):
                    campo115 = 'INGLES COMPLETO'
                    campo116 = 0

                campo117 = 'NO TIENE VINCULACION'
                data['malla_horas_vinculacion'] = malla.horas_vinculacion
                horastotal = \
                    ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True,
                                                         proyecto__status=True,
                                                         inscripcion_id=matriculados.inscripcion.id).aggregate(
                        horastotal=Sum('horas'))['horastotal']
                horastotal = horastotal if horastotal else 0
                campo118 = malla.horas_vinculacion - horastotal
                if horastotal >= malla.horas_vinculacion:
                    campo117 = 'TIENE VINCULACION'
                    campo118 = 0

                campo119 = 'NO TIENE COMPUTACION'
                asignatura = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(
                    malla__id=32)
                record = RecordAcademico.objects.filter(inscripcion__id=matriculados.inscripcion.id,
                                                        asignatura__id__in=asignatura, aprobada=True)
                creditos_computacion = 0
                malla.creditos_computacion
                listconcreditos = []
                for comp in record:
                    listconcreditos.append(comp.asignatura.nombre)
                    creditos_computacion += comp.creditos
                campo121 = listconcreditos
                campo120 = malla.creditos_computacion - creditos_computacion
                if creditos_computacion >= malla.creditos_computacion:
                    campo119 = 'SI TIENE COMPUTACION'
                    campo120 = 0

                presidente = ""
                secretario = ""
                integrantedelegado = ""
                fechagrupo = ""
                horagrupo = ""
                lugargrupo = ""
                if ComplexivoDetalleGrupo.objects.filter(matricula=matriculados, status=True):
                    grupocomplexivo = \
                        ComplexivoDetalleGrupo.objects.filter(matricula=matriculados, status=True)[0]
                    if grupocomplexivo.grupo.presidentepropuesta:
                        presidente = u"%s" % grupocomplexivo.grupo.presidentepropuesta.persona

                    if grupocomplexivo.grupo.secretariopropuesta:
                        secretario = u"%s" % grupocomplexivo.grupo.secretariopropuesta.persona

                    if grupocomplexivo.grupo.delegadopropuesta:
                        integrantedelegado = u"%s" % grupocomplexivo.grupo.delegadopropuesta.persona

                    if grupocomplexivo.grupo.fechadefensa:
                        fechagrupo = u"%s" % grupocomplexivo.grupo.fechadefensa

                    if grupocomplexivo.grupo.horadefensa:
                        horagrupo = u"%s" % grupocomplexivo.grupo.horadefensa

                    if grupocomplexivo.grupo.lugardefensa:
                        lugargrupo = u"%s" % grupocomplexivo.grupo.lugardefensa

                codigoestado = 0
                nomestado = ''
                pexamen = 0
                if matriculados.alternativa.tiene_examen():
                    if ComplexivoExamenDetalle.objects.filter(status=True, matricula=matriculados).exists():
                        detalle = \
                            ComplexivoExamenDetalle.objects.filter(status=True,
                                                                   matricula=matriculados).order_by('-id')[
                                0]
                        nomestado = detalle.get_estado_display()
                        codigoestado = detalle.estado
                        pexamen = detalle.ponderacion()

                ppropuesta = 0
                ptotal = 0
                if ComplexivoGrupoTematica.objects.values('id').filter(status=True,
                                                                       complexivodetallegrupo__status=True,
                                                                       complexivodetallegrupo__matricula=matriculados).exists():
                    ppropuesta = matriculados.notapropuesta()
                if matriculados.alternativa.tipotitulacion.tipo == 1:
                    ptotal = ppropuesta
                if matriculados.alternativa.tipotitulacion.tipo == 2:
                    ptotal = matriculados.notafinalcomplexivoestado(codigoestado)
                estadotitulacion = matriculados.get_estadotitulacion_display()
                # ws.write(row_num, 149, lugargrupo, normal)
                # ws.write(row_num, 148, horagrupo, normal)
                # ws.write(row_num, 147, fechagrupo, date_format)
                ws.write(row_num, 19, integrantedelegado, normal)
                ws.write(row_num, 18, secretario, normal)
                ws.write(row_num, 17, presidente, normal)
                # ws.write(row_num, 143, str(campo121), normal)
                # ws.write(row_num, 142, campo120, normal)
                # ws.write(row_num, 141, campo119, normal)
                # if matriculados.inscripcion.exonerado_practias():
                #     ws.write(row_num, 140, 'EXONERADO', normal)
                # else:
                #     ws.write(row_num, 140, campo118, normal)
                # ws.write(row_num, 139, campo117, normal)
                # ws.write(row_num, 138, campo116, normal)
                # ws.write(row_num, 137, campo115, normal)
                # ws.write(row_num, 136, campo114, normal)
                # ws.write(row_num, 135, campo113, normal)
                # ws.write(row_num, 134, campo112, normal)
                # ws.write(row_num, 133, campo111, normal)
                # if matriculados.inscripcion.exonerado_practias():
                #     ws.write(row_num, 132, 'EXONERADO', normal)
                # else:
                #     ws.write(row_num, 132, campo110, normal)
                # ws.write(row_num, 131, campo109, normal)
                # ws.write(row_num, 130, u"%s" % campo123, normal)
                # ws.write(row_num, 129, campo122, date_format)
                ws.write(row_num, 16, campo107, normal)
                # ws.write(row_num, 56, campo51, normal)
                # ws.write(row_num, 57, campo52, normal)
                # ws.write(row_num, 58, campo53, normal)
                # ws.write(row_num, 59, campo54, normal)
                # ws.write(row_num, 60, campo55, normal)
                # ws.write(row_num, 61, campo56, normal)
                # ws.write(row_num, 62, campo57, normal)
                # ws.write(row_num, 63, campo58, normal)
                # ws.write(row_num, 64, campo59, normal)
                # ws.write(row_num, 65, campo60, normal)
                # ws.write(row_num, 66, campo61, normal)
                # ws.write(row_num, 67, campo62, normal)
                # ws.write(row_num, 68, campo63, normal)
                # ws.write(row_num, 69, campo64, normal)
                # ws.write(row_num, 70, campo65, normal)
                # ws.write(row_num, 71, campo66, normal)
                # ws.write(row_num, 72, campo67, normal)
                # ws.write(row_num, 73, campo68, normal)
                # ws.write(row_num, 74, campo69, normal)
                # ws.write(row_num, 75, campo70, normal)
                # ws.write(row_num, 76, campo71, normal)
                # ws.write(row_num, 77, campo72, normal)
                # ws.write(row_num, 78, campo73, normal)
                # ws.write(row_num, 79, campo74, normal)
                # ws.write(row_num, 80, campo75, normal)
                # ws.write(row_num, 81, campo76, normal)
                # ws.write(row_num, 82, campo77, normal)
                # ws.write(row_num, 83, campo78, normal)
                # ws.write(row_num, 84, campo79, normal)
                # ws.write(row_num, 85, campo80, normal)
                # ws.write(row_num, 86, campo81, normal)
                # ws.write(row_num, 87, campo82, normal)
                # ws.write(row_num, 88, campo83, normal)
                # ws.write(row_num, 89, campo84, normal)
                # ws.write(row_num, 90, campo85, normal)
                # ws.write(row_num, 91, campo86, normal)
                # ws.write(row_num, 92, campo87, normal)
                # ws.write(row_num, 93, campo88, normal)
                # ws.write(row_num, 94, campo89, normal)
                # ws.write(row_num, 95, campo90, normal)
                # ws.write(row_num, 96, campo91, normal)
                # ws.write(row_num, 97, campo92, normal)
                # ws.write(row_num, 98, campo93, normal)
                # ws.write(row_num, 99, campo94, normal)
                # ws.write(row_num, 100, campo95, normal)
                # ws.write(row_num, 101, campo96, normal)
                ws.write(row_num, 9, campo97, normal)
                # ws.write(row_num, 103, campo98, normal)
                # ws.write(row_num, 104, campo99, normal)
                ws.write(row_num, 10, campo100, normal)
                # ws.write(row_num, 106, campo101, date_format)
                ws.write(row_num, 11, campo102, normal)
                ws.write(row_num, 12, campo103, normal)
                ws.write(row_num, 13, campo104, normal)
                ws.write(row_num, 14, campo105, normal)
                ws.write(row_num, 15, campo106, normal)

                ws.write(row_num, 20, pexamen, normal)
                ws.write(row_num, 21, ppropuesta, normal)
                ws.write(row_num, 22, ptotal, normal)
                ws.write(row_num, 23, nomestado, normal)
                ws.write(row_num, 24, estadotitulacion, normal)
                ws.write(row_num, 25, matriculados.alternativa.grupotitulacion.periodogrupo.nombre, normal)
                ws.write(row_num, 26, fechagrupo)
                # ws.write(row_num, 112, campo107, normal)
                row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/titulacion/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Excel Matriculados en proceso de titulación',
                                    destinatario=pers, url="{}reportes/titulacion/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Matriculados en Proceso de Titulación',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/titulacion/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_pre_inscritos_ppp(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.periodo = periodo
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'practicasppp')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "reporte_practicas_ppp_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'practicasppp', nombre_archivo)

        try:
            columns = [
                (u"COD.", 2500),
                (u"MALLA", 20000),
                (u"FACULTAD", 20000),
                (u"CARRERA", 13000),
                (u"NIVEL", 3000),
                (u"PERIODO ACADEMICO", 16000),
                (u"CEDULA", 3000),
                (u"ESTUDIANTE", 12000),
                (u"SEXO", 3000),
                (u"CELULAR", 3000),
                (u"CONVENCIONAL", 5000),
                (u"EMAIL", 10000),
                (u"EMAIL INSTITUCIONAL", 10000),
                (u"DIRECCION DOMICILIARIA", 15000),
                (u"CANTON", 10000),
                (u"EGRESADO", 5000),
                (u"TIPO PRACTICAS", 8000),
                (u"TUTOR ACADÉMICO", 12000),
                (u"CED. TUTOR ACADÉMICO", 5000),
                (u"CORREO DEL TUTOR", 10000),
                (u"TUTOR PROFESIONAL", 12000),
                (u"SUPERVISOR", 12000),
                (u"CED. SUPERVISOR", 5000),
                (u"FECHA DESDE", 4000),
                (u"FECHA HASTA", 4000),
                (u"HORAS PRACTICAS", 6000),
                (u"HORAS HOMOLOGACIÓN", 6000),
                (u"INSTITUCION", 13000),
                (u"OTRA EMPRESA", 15000),
                (u"TIPO INSTITUCION", 6000),
                (u"DEPARTAMENTO", 15000),
                (u"SECTOR ECONOMICO", 6500),
                (u"USUARIO", 4000),
                (u"FECHA REGISTRO", 4000),
                (u"INSCRIPCION", 4000),
                (u"TIPO SOLICITUD", 6000),
                (u"ARCHIVO DE SOLICITUD", 6000),
                (u"ESTADO SOLICITUD", 6000),
                (u"PRACTICAS CULMINADAS", 7000),
                (u"EVIDENCIAS APROBADAS / TOTAL ", 10000),
                (u"EVIDENCIAS RECHAZADAS", 10000),
                (u"EVIDENCIAS SOLICITADAS", 10000),
                (u"EVIDENCIAS COMPLETAS", 10000),
                (u"FECHA EVIDENCIA", 5000),
                (u"FECHA ACTUALIZACION EVIDENCIA", 10000),
                (u"APROBACIÓN SOLICITUD", 12000),
                (u"OBSERVACIÓN", 30000),
                (u"VALIDACIÓN", 30000),
                (u"FECHA VALIDACIÓN", 6000),
                (u"ROTACION", 6000),
                (u"SESION", 6000),
                (u"EVALUA PROMEDIO PRÁCTICA", 10000),
                (u"PROMEDIO PRÁCTICA", 6000),
                (u"CONVENIO/ACUERDO", 6000),
                (u"EMPRESA CONVENIO/ACUERDO", 10000),
                (u"ASIGNACIÓN EMPRESA", 10000),
                (u"PAIS", 6000),
                (u"PROVINCIA", 6000),
                (u"LUGAR", 10000),
                (u"ITINERARIO", 10000),
                (u"NIVEL PRÁCTICA", 10000),
                (u"PERIODO DE EVIDENCIA", 20000),
                (u"TOTAL TUTORIAS", 10000),
                (u"ESTADO", 10000),
                (u"FECHA CULMINACIÓN TUTORIA", 10000),
                (u"FECHA CULMINACIÓN PRÁCTICAS", 10000),
                (u"PERIODO DE PREINSCRIPCIÓN", 20000),
                (u"F.NACIMIENTO", 10000),
                (u"EDAD", 10000),
                (u"TIPO DISCAPACIDAD", 10000),
                (u"% DISCAPACIDAD", 10000),
                (u"CARNET DISCAPACIDAD", 10000),
                (u"ETNIA", 10000),
            ]
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="reporte_practicas.xls"'
            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('HOJA_1')

            title = easyxf(
                'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
            ws.write_merge(0, 0, 0, 61, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)

            fuentecabecera = easyxf(
                'font: name Calibri, color-index black, bold on, height 200; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')

            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                ws.col(col_num).width = columns[col_num][1]

            font_style = easyxf(
                'font: name Calibri, color-index black, height 200; borders: left thin, right thin, top thin, bottom thin')
            date_format = font_style
            date_format.num_format_str = 'yyyy/mm/dd'

            row_num = 4
            for practicas in data:
                periodoinscripcion = ''
                campo52 = ''
                campo3 = ''
                campo4 = ''
                campo47 = ''
                campo56 = ''
                if practicas.inscripcion.nivelperiodo(self.periodo):
                    matriculado = practicas.inscripcion.nivelperiodo(self.periodo)
                    campo3 = matriculado.nivelmalla.nombre
                    campo4 = matriculado.nivel.periodo.nombre
                    campo47 = matriculado.nivel.sesion.nombre.__str__()
                persona_ = practicas.inscripcion.persona
                perfil_ = persona_.mi_perfil()
                campo1 = practicas.inscripcion.coordinacion.nombre
                campo2 = practicas.inscripcion.carrera.nombre_completo()
                campo5 = practicas.inscripcion.persona.cedula
                campo6 = practicas.inscripcion.persona.telefono
                campo7 = practicas.inscripcion.persona.telefono_conv
                campo8 = practicas.inscripcion.persona.email
                campo35 = practicas.inscripcion.persona.emailinst
                campo36 = ''
                if practicas.rotacionmalla:
                    campo36 = practicas.rotacionmalla.nombre
                if practicas.inscripcion.egresado():
                    campo9 = 'SI'
                else:
                    campo9 = 'NO'
                campo10 = practicas.get_tipo_display()
                campo15 = practicas.tutorunemi.persona.nombre_completo_inverso() if practicas.tutorunemi else ""
                cedtutor = practicas.tutorunemi.persona.cedula if practicas.tutorunemi else ""
                campo37 = practicas.tutorempresa if practicas.tutorempresa else ""
                campo41 = practicas.supervisor.persona.nombre_completo_inverso() if practicas.supervisor else ""
                cedsupervisor = practicas.supervisor.persona.cedula if practicas.supervisor else ""
                campo11 = practicas.fechadesde if practicas.fechadesde else ""
                campo12 = practicas.fechahasta if practicas.fechahasta else ""
                campo13 = practicas.numerohora
                campo14 = practicas.empresaempleadora.nombre if practicas.empresaempleadora else ""
                campo16 = practicas.get_tipoinstitucion_display()
                campo17 = practicas.get_sectoreconomico_display()
                if practicas.usuario_creacion:
                    campo18 = practicas.usuario_creacion.username
                else:
                    campo18 = ''
                campo19 = practicas.fecha_creacion
                campo20 = practicas.inscripcion.id
                campo21 = practicas.get_tiposolicitud_display()
                campo22 = practicas.get_estadosolicitud_display()
                campo23 = ''
                if InscripcionMalla.objects.filter(inscripcion=practicas.inscripcion, status=True).exists():
                    insmalla = \
                        InscripcionMalla.objects.select_related().filter(inscripcion=practicas.inscripcion,
                                                                         status=True)[0]
                    nommalla = 'HISTORICA'
                    mallamodalidad = ''
                    if insmalla.malla.vigente:
                        nommalla = 'VIGENTE'
                    if insmalla.malla.modalidad:
                        mallamodalidad = insmalla.malla.modalidad.nombre
                    campo23 = insmalla.malla.carrera.nombre + ' ' + MONTH_NAMES[
                        insmalla.malla.inicio.month - 1] + ' ' + str(
                        insmalla.malla.inicio.year) + ' ' + nommalla + ' ' + mallamodalidad
                if practicas.culminada:
                    campo24 = 'SI'
                else:
                    campo24 = 'NO'
                ultimafechaevidencia = DetalleEvidenciasPracticasPro.objects.select_related().filter(
                    inscripcionpracticas=practicas, status=True).order_by('-id')
                ultimafechaevidenciaact = DetalleEvidenciasPracticasPro.objects.select_related().filter(
                    inscripcionpracticas=practicas, status=True).order_by('-fecha_modificacion')
                if practicas.fechadesde:
                    campo25 = str(practicas.evidenciasaprobadas()) + ' / ' + str(practicas.totalevidencias())
                    if str(practicas.evidenciasaprobadas()) == str(practicas.totalevidencias()):
                        campo26 = 'SI'
                    else:
                        campo26 = 'NO'
                else:
                    campo25 = "0/0"
                    campo26 = ""
                campo27 = ''
                campo33 = ''
                if ultimafechaevidencia:
                    campo27 = ultimafechaevidencia[0].fecha_creacion if ultimafechaevidencia[
                        0].fecha_creacion else ''
                    campo33 = ultimafechaevidenciaact[0].fecha_modificacion if ultimafechaevidenciaact[
                        0].fecha_modificacion else ''
                campo28 = practicas.evidenciasreprobadas()
                campo29 = practicas.evidenciassolicitadas()
                campo30 = practicas.fechaaprueba if practicas.fechaaprueba else ""
                campo31 = practicas.horahomologacion if practicas.horahomologacion else ""
                campo32 = practicas.otraempresaempleadora if practicas.otraempresaempleadora else ""
                campo34 = practicas.observacion if practicas.observacion else ""
                campo38 = practicas.inscripcion.persona.sexo.nombre if practicas.inscripcion.persona.sexo else ""
                campo39 = practicas.validacion if practicas.validacion else " - "
                campo40 = practicas.fechavalidacion if practicas.fechavalidacion else "-"
                campo42 = practicas.inscripcion.persona.direccion_corta()
                campo43 = u'%s' % practicas.inscripcion.persona.canton.nombre if practicas.inscripcion.persona.canton else ''
                campo44 = ""
                if practicas.tutorunemi:
                    if practicas.tutorunemi.persona.emailinst:
                        campo44 = u'%s' % practicas.tutorunemi.persona.emailinst
                campo45 = ""
                if practicas.departamento:
                    campo45 = u'%s' % practicas.departamento
                if practicas.archivo:
                    campo46 = "SI"
                else:
                    campo46 = "NO"
                campo48 = ""
                if practicas.periodoppp:
                    if practicas.periodoppp.evaluarpromedio:
                        campo48 = u'%s' % practicas.total_promedio_nota_evidencia()
                campo49 = ""
                if practicas.periodoppp:
                    campo49 = "SI" if practicas.periodoppp.evaluarpromedio else "NO"
                if practicas.convenio:
                    campo50 = "CONVENIO"
                    campo52 = practicas.convenio.empresaempleadora.nombre
                elif practicas.acuerdo:
                    campo50 = "ACUERDO"
                    campo52 = practicas.acuerdo.empresa.nombre
                else:
                    campo50 = "NINGUNO"
                    campo52 = "NINGUNO"
                if practicas.asignacionempresapractica:
                    campo51 = practicas.asignacionempresapractica.nombre
                else:
                    campo51 = "SIN ASIGNAR"

                paispractica = practicas.lugarpractica.provincia.pais.nombre if practicas.lugarpractica else ''
                provinciapractica = practicas.lugarpractica.provincia.nombre if practicas.lugarpractica else ''
                lugarpractica = practicas.lugarpractica.nombre if practicas.lugarpractica else ''

                campo53 = ''
                if practicas.itinerariomalla:
                    campo53 = str(practicas.itinerariomalla.nombreitinerario())
                elif practicas.actividad:
                    if practicas.actividad.itinerariomalla:
                        campo53 = str(practicas.actividad.itinerariomalla)

                campo56 = ''
                periodoevidencia = ''
                if practicas.itinerariomalla:
                    periodoevidencia = str(practicas.periodoppp)
                if practicas.periodoppp:
                    periodoevidencia = str(practicas.periodoppp)
                if practicas.tiposolicitud != 3:
                    if practicas.preinscripcion:
                        if practicas.preinscripcion.preinscripcion:
                            if practicas.preinscripcion.preinscripcion.periodo:
                                periodoinscripcion = practicas.preinscripcion.preinscripcion.periodo
                ws.write(row_num, 0, str(practicas.id), font_style)
                ws.write(row_num, 1, campo23, font_style)
                ws.write(row_num, 2, campo1, font_style)
                ws.write(row_num, 3, campo2, font_style)
                ws.write(row_num, 4, campo3, font_style)
                ws.write(row_num, 5, campo4, font_style)
                ws.write(row_num, 6, campo5, font_style)
                ws.write(row_num, 7,
                         practicas.inscripcion.persona.apellido1 + ' ' + practicas.inscripcion.persona.apellido2 + ' ' + practicas.inscripcion.persona.nombres,
                         font_style)
                ws.write(row_num, 8, campo38, font_style)
                ws.write(row_num, 9, campo6, font_style)
                ws.write(row_num, 10, campo7, font_style)
                ws.write(row_num, 11, campo8, font_style)
                ws.write(row_num, 12, campo35, font_style)
                ws.write(row_num, 13, campo42, font_style)
                ws.write(row_num, 14, campo43, font_style)
                ws.write(row_num, 15, campo9, font_style)
                ws.write(row_num, 16, campo10, font_style)
                ws.write(row_num, 17, campo15, font_style)
                ws.write(row_num, 18, cedtutor, font_style)
                ws.write(row_num, 19, campo44, font_style)
                ws.write(row_num, 20, campo37, font_style)
                ws.write(row_num, 21, campo41, font_style)
                ws.write(row_num, 22, cedsupervisor, date_format)
                ws.write(row_num, 23, campo11, date_format)
                ws.write(row_num, 24, campo12, date_format)
                ws.write(row_num, 25, campo13, font_style)
                ws.write(row_num, 26, campo31, font_style)
                ws.write(row_num, 27, campo14, font_style)
                ws.write(row_num, 28, campo32, font_style)
                ws.write(row_num, 29, campo16, font_style)
                ws.write(row_num, 30, campo45, font_style)
                ws.write(row_num, 31, campo17, font_style)
                ws.write(row_num, 32, campo18, font_style)
                ws.write(row_num, 33, campo19, date_format)
                ws.write(row_num, 34, campo20, font_style)
                ws.write(row_num, 35, campo21, font_style)
                ws.write(row_num, 36, campo46, font_style)
                ws.write(row_num, 37, campo22, font_style)
                ws.write(row_num, 38, campo24, font_style)
                ws.write(row_num, 39, campo25, font_style)
                ws.write(row_num, 40, campo28, font_style)
                ws.write(row_num, 41, campo29, font_style)
                ws.write(row_num, 42, campo26, font_style)
                ws.write(row_num, 43, campo27, date_format)
                ws.write(row_num, 44, campo33, date_format)
                ws.write(row_num, 45, campo30, date_format)
                ws.write(row_num, 46, campo34, font_style)
                ws.write(row_num, 47, campo39, font_style)
                ws.write(row_num, 48, campo40, date_format)
                ws.write(row_num, 49, campo36, font_style)
                ws.write(row_num, 50, campo47, font_style)
                ws.write(row_num, 51, campo49, font_style)
                ws.write(row_num, 52, campo48, font_style)
                ws.write(row_num, 53, campo50, font_style)
                ws.write(row_num, 54, campo52, font_style)
                ws.write(row_num, 55, campo51, font_style)
                ws.write(row_num, 56, paispractica, font_style)
                ws.write(row_num, 57, provinciapractica, font_style)
                ws.write(row_num, 58, lugarpractica, font_style)
                ws.write(row_num, 59, campo53, font_style)
                ws.write(row_num, 60, practicas.nivelmalla.nombre if practicas.nivelmalla else '', font_style)
                ws.write(row_num, 61, periodoevidencia, font_style)
                ws.write(row_num, 62, practicas.cuenta_tutorias(), font_style)
                estadopractica = 'NO OBLIGATORIAS'
                if practicas.tipo == 1 or practicas.tipo == 2:
                    if practicas.aplicatutoria:
                        if practicas.culminatutoria:
                            estadopractica = 'CULMINADA'
                        else:
                            estadopractica = 'PENDIENTE'
                total_tutorias = practicas.practicastutoria_set.filter(status=True).count()
                ws.write(row_num, 63, estadopractica, font_style)
                fechaculminacion = ''
                if practicas.fechaculminacion:
                    fechaculminacion = str(practicas.fechaculminacion)
                ws.write(row_num, 64, fechaculminacion, font_style)
                fechaculminacionpracticas = ''
                if practicas.fechaculminacionpracticas:
                    fechaculminacionpracticas = str(practicas.fechaculminacionpracticas)
                ws.write(row_num, 65, fechaculminacionpracticas, font_style)
                ws.write(row_num, 66, periodoinscripcion.__str__(), font_style)
                ws.write(row_num, 67, persona_.nacimiento, font_style)
                ws.write(row_num, 68, persona_.edad(), font_style)
                if perfil_.tienediscapacidad:
                    ws.write(row_num, 69, perfil_.tipodiscapacidad.nombre if perfil_.tipodiscapacidad else '',
                             font_style)
                    ws.write(row_num, 70, perfil_.porcientodiscapacidad, font_style)
                    ws.write(row_num, 71, perfil_.carnetdiscapacidad, font_style)
                    ws.write(row_num, 72, perfil_.raza.__str__() if perfil_.raza else '', font_style)
                else:
                    ws.write(row_num, 69, '', font_style)
                    ws.write(row_num, 70, '', font_style)
                    ws.write(row_num, 71, '', font_style)
                    ws.write(row_num, 72, '', font_style)
                row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/practicasppp/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Excel Reporte de Practicas PreProfesionales',
                                    destinatario=pers,
                                    url="{}reportes/practicasppp/{}".format(MEDIA_URL, nombre_archivo), prioridad=1,
                                    app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Reporte de Practicas PreProfecionales',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/practicasppp/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_matrizestudiantes_anio_background(threading.Thread):

    def __init__(self, request, data, notiid, name_document):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.name_document = name_document
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        name_document = self.name_document
        nombre_archivo = name_document + "_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            # if 'anio' in request.GET:
            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            workbook = xlsxwriter.Workbook(directory)
            ws = workbook.add_worksheet('estudiantes')

            font_style2 = workbook.add_format(
                {'border': 1, 'text_wrap': True, 'align': 'center', 'font_size': 5, 'valign': 'vcenter',
                 'font_name': 'Century Gothic'})
            font_style = workbook.add_format(
                {'border': 1, 'text_wrap': True, 'bold': 1, 'align': 'center', 'font_size': 5, 'valign': 'vcenter',
                 'font_name': 'Century Gothic'})

            coordinacion = None
            carrera = None
            if 'idcoor' in request.GET:
                idcoor = request.GET['idcoor']
                if int(idcoor) > 0:
                    coordinacion = Coordinacion.objects.get(id=int(idcoor))
                    nombbre = u"Estudiantes %s" % coordinacion.nombre
                else:
                    nombbre = u"Estudiantes"
            else:
                nombbre = u"Estudiantes"

            if 'idcarr' in request.GET:
                idcarr = request.GET['idcarr']
                if idcarr != '':
                    if int(idcarr) > 0:
                        carrera = Carrera.objects.get(id=int(idcarr))
                        nombbre += u" Carrera %s" % carrera.nombre_completo()

            ws.write(0, 0, u"CODIGO_IES", font_style)
            ws.write(0, 1, u"CODIGO_CARRERA", font_style)
            ws.write(0, 2, u"CIUDAD_CARRERA", font_style)
            ws.write(0, 3, u"TIPO_IDENTIFICACION", font_style)
            ws.write(0, 4, u"IDENTIFICACION", font_style)
            ws.write(0, 5, u"PRIMER_APELLIDO", font_style)
            ws.write(0, 6, u"SEGUNDO_APELLIDO", font_style)
            ws.write(0, 7, u"NOMBRES", font_style)
            ws.write(0, 8, u"SEXO", font_style)
            ws.write(0, 9, u"FECHA_NACIMIENTO", font_style)
            ws.write(0, 10, u"PAIS_ORIGEN", font_style)
            ws.write(0, 11, u"DISCAPACIDAD", font_style)
            ws.write(0, 12, u"PORCENTAJE_DISCAPACIDAD", font_style)
            ws.write(0, 13, u"NUMERO_CONADIS", font_style)
            ws.write(0, 14, u"ETNIA", font_style)
            ws.write(0, 15, u"NACIONALIDAD", font_style)
            ws.write(0, 16, u"DIRECCION", font_style)
            ws.write(0, 17, u"EMAIL_PERSONAL", font_style)
            ws.write(0, 18, u"EMAIL_INSTITUCIONAL", font_style)
            ws.write(0, 19, u"FECHA_INICIO_PRIMER_NIVEL", font_style)
            ws.write(0, 20, u"FECHA_INGRESO_CONVALIDACION", font_style)
            ws.write(0, 21, u"PAIS_RESIDENCIA", font_style)
            ws.write(0, 22, u"PROVINCIA_RESIDENCIA", font_style)
            ws.write(0, 23, u"CANTON_RESIDENCIA", font_style)
            ws.write(0, 24, u"CELULAR", font_style)
            ws.write(0, 25, u"NIVEL_FORMACION_PADRE", font_style)
            ws.write(0, 26, u"NIVEL_FORMACION_MADRE", font_style)
            ws.write(0, 27, u"CANTIDAD_MIEMBROS_HOGAR", font_style)
            ws.write(0, 28, u"TIPO_COLEGIO", font_style)
            ws.write(0, 29, u"POLITICA_CUOTA", font_style)
            ws.write(0, 30, u"CARRERA", font_style)
            ws.write(0, 31, u"FACULTAD", font_style)
            ws.write(0, 32, u"PASAPORTE", font_style)
            ws.write(0, 33, u"LGBTI", font_style)
            ws.write(0, 34, u"ES EXTRANJERO", font_style)
            ws.write(0, 35, u"PROVINCIA_NACI", font_style)
            ws.write(0, 36, u"CANTON_NACI", font_style)

            cursor = connection.cursor()

            listaestudiante = f"""
                    SELECT 
                        mat.id, 
                        COUNT(mta.id) FILTER(WHERE i.coordinacion_id not in (7) AND asi.modulo) "num_modulo",
                        COUNT(mta.id) "num_total"
                    FROM
                        sga_inscripcion i
                        INNER JOIN sga_matricula mat ON mat.inscripcion_id=i.id
                        INNER JOIN sga_nivel niv ON niv.id=mat.nivel_id
                        INNER JOIN sga_nivelmalla nivm ON nivm.id=mat.nivelmalla_id
                        INNER JOIN sga_periodo peri ON peri.id=niv.periodo_id
                        INNER JOIN sga_materiaasignada mta ON mta.matricula_id=mat.id
                        LEFT JOIN sga_materia mt ON mt.id=mta.materia_id
                        LEFT  JOIN sga_asignatura asi ON asi.id=mt.asignatura_id
                    WHERE 
                        mat."status"
                        AND peri.anio={int(request.GET['anio'])}
                        and mat.estado_matricula in (2,3)
                    GROUP BY mat.id
                    HAVING 
                        COUNT(mta.id) FILTER(WHERE i.coordinacion_id not in (7) AND asi.modulo)=COUNT(mta.id)
            """
            cursor.execute(listaestudiante)
            results = cursor.fetchall()
            respuestas = []
            for per in results:
                respuestas.append(per[0])
            inscripciones = Inscripcion.objects.filter(
                matricula__status=True,
                matricula__estado_matricula__in=[2, 3],
                matricula__retiromatricula__isnull=True,
                matricula__nivel__periodo__anio=int(request.GET['anio'])).exclude(
                matricula__id__in=respuestas).distinct()
            if coordinacion:
                inscripciones = inscripciones.filter(carrera__coordinacion=coordinacion)
            if carrera:
                inscripciones = inscripciones.filter(carrera=carrera)
            row_num = 1

            for inscripcion in inscripciones:
                # Datos del estudiante
                tipoidentificacion = inscripcion.persona.tipo_documento()
                documento = inscripcion.persona.documento()
                eslgtbi = 'NO'
                if inscripcion.persona.lgtbi:
                    eslgtbi = 'SI'

                # Datos de perfil inscripcion
                tipo_discapacidad = 'NINGUNA'
                carnet_discapacidad = ''
                porcentaje_discapacidad = 0
                politica_cuota = 'NINGUNA'
                raza = 'NO REGISTRA'
                nacionalidad = 'NO APLICA'

                matriculas = inscripcion.matricula_set.filter(status=True,
                                                              retiromatricula__isnull=True,
                                                              nivel__periodo__anio=int(request.GET['anio'])).order_by(
                    'id')
                perfilinscripcion = inscripcion.persona.perfilinscripcion_set.filter(status=True).first()
                if perfilinscripcion is not None:
                    if perfilinscripcion.tienediscapacidad:
                        if perfilinscripcion.verificadiscapacidad:
                            tipo_discapacidad = perfilinscripcion.tipodiscapacidad.__str__()
                            porcentaje_discapacidad = perfilinscripcion.porcientodiscapacidad
                            carnet_discapacidad = perfilinscripcion.carnetdiscapacidad if perfilinscripcion.carnetdiscapacidad else carnet_discapacidad
                            politica_cuota = 'DISCAPACIDAD'

                    if perfilinscripcion.raza:
                        raza = perfilinscripcion.raza.__str__()
                        if perfilinscripcion.raza.id == 1:
                            politica_cuota = 'PUEBLOS Y NACIONALIDADES'
                            nacionalidad = perfilinscripcion.nacionalidadindigena.__str__()

                # Datos de familiares
                formacion_padre = 'NO REGISTRA'
                formacion_madre = 'NO REGISTRA'
                cantidad = 0
                persona_datos_familiares = inscripcion.persona.personadatosfamiliares_set.filter(status=True)
                if persona_datos_familiares.values_list('id', flat=True).exists():
                    cantidad = persona_datos_familiares.__len__()
                    parentesco_padre = persona_datos_familiares.filter(parentesco_id=1).first()
                    if parentesco_padre is not None:
                        formacion_padre = parentesco_padre.niveltitulacion.nombrecaces if parentesco_padre.niveltitulacion else formacion_padre
                    parentesco_madre = persona_datos_familiares.filter(parentesco_id=2).first()
                    if parentesco_madre is not None:
                        formacion_madre = parentesco_madre.niveltitulacion.nombrecaces if parentesco_madre.niveltitulacion else formacion_madre

                es_pre = False
                malla = inscripcion.mi_malla()
                carrera = inscripcion.carrera
                coordinacion = inscripcion.coordinacion

                primera_matricula = matriculas.first()
                segunda_matricula = matriculas.last()
                fecha = inscripcion.fechainicioprimernivel.strftime(
                    "%d/%m/%Y") if inscripcion.fechainicioprimernivel else ''
                if coordinacion:
                    if coordinacion.id == 9:
                        es_pre = True
                        fecha = primera_matricula.fecha.strftime("%d/%m/%Y") if primera_matricula is not None else ''

                codigo_carrera = '00098'
                if not es_pre:
                    codigo_carrera = malla.codigo if malla else ''

                if primera_matricula is not None:
                    primera_matricula_gruposocioeconomico = primera_matricula.matriculagruposocioeconomico_set.filter(
                        status=True).first()
                    if primera_matricula_gruposocioeconomico is not None:
                        if primera_matricula_gruposocioeconomico.gruposocioeconomico_id in [4, 5]:
                            politica_cuota = 'SOCIECONOMICA'

                tipo_colegio = 'NO REGISTRA'
                esextranjero = 'SI'
                if inscripcion.persona.paisnacimiento:
                    if inscripcion.persona.paisnacimiento.id==1:
                        esextranjero = 'NO'

                if inscripcion.unidadeducativa:
                    tipo_colegio = inscripcion.unidadeducativa.tipocolegio.nombre.upper() if inscripcion.unidadeducativa.tipocolegio else tipo_colegio
                ciudad_carrera = "MILAGRO"
                ws.write(row_num, 0, '1024', font_style2)
                ws.write(row_num, 1, codigo_carrera, font_style2)
                ws.write(row_num, 2, ciudad_carrera, font_style2)
                ws.write(row_num, 3, tipoidentificacion, font_style2)
                ws.write(row_num, 4, documento, font_style2)
                ws.write(row_num, 5, inscripcion.persona.apellido1, font_style2)
                ws.write(row_num, 6, inscripcion.persona.apellido2, font_style2)
                ws.write(row_num, 7, inscripcion.persona.nombres, font_style2)
                ws.write(row_num, 8, inscripcion.persona.sexo.nombre if inscripcion.persona.sexo_id else 'NO REGISTRA', font_style2)
                ws.write(row_num, 9, inscripcion.persona.nacimiento.strftime("%d/%m/%Y") if inscripcion.persona.nacimiento else '', font_style2)
                ws.write(row_num, 10,inscripcion.persona.paisnacimiento.nombre if inscripcion.persona.paisnacimiento_id else '', font_style2)
                ws.write(row_num, 11, tipo_discapacidad, font_style2)
                ws.write(row_num, 12, porcentaje_discapacidad, font_style2)
                ws.write(row_num, 13, carnet_discapacidad, font_style2)
                ws.write(row_num, 14, raza, font_style2)
                ws.write(row_num, 15, nacionalidad, font_style2)
                ws.write(row_num, 16, inscripcion.persona.direccion if inscripcion.persona.direccion2 else 'NO APLICA', font_style2)
                ws.write(row_num, 17, inscripcion.persona.email if inscripcion.persona.email else 'NO APLICA', font_style2)
                ws.write(row_num, 18, inscripcion.persona.emailinst if inscripcion.persona.emailinst else 'NO APLICA', font_style2)
                ws.write(row_num, 19, fecha, font_style2)
                ws.write(row_num, 20, inscripcion.fechainicioconvalidacion.strftime("%d/%m/%Y") if inscripcion.fechainicioconvalidacion else '', font_style2)
                ws.write(row_num, 21, inscripcion.persona.pais.nombre if inscripcion.persona.pais_id else 'NO APLICA', font_style2)
                ws.write(row_num, 22, inscripcion.persona.provincia.nombre if inscripcion.persona.provincia_id else 'NO APLICA', font_style2)
                ws.write(row_num, 23, inscripcion.persona.canton.nombre if inscripcion.persona.canton_id else 'NO APLICA',font_style2)
                ws.write(row_num, 24, inscripcion.persona.telefono if inscripcion.persona.telefono else '0000000000', font_style2)
                ws.write(row_num, 25, formacion_padre, font_style2)
                ws.write(row_num, 26, formacion_madre, font_style2)
                ws.write(row_num, 27, cantidad if cantidad else 0, font_style2)
                ws.write(row_num, 28, tipo_colegio, font_style2)
                ws.write(row_num, 29, politica_cuota, font_style2)
                ws.write(row_num, 30, carrera.__str__(), font_style2)
                ws.write(row_num, 31, coordinacion.__str__(), font_style2)
                ws.write(row_num, 32, inscripcion.persona.pasaporte.__str__(), font_style2)
                ws.write(row_num, 33, eslgtbi, font_style2)
                ws.write(row_num, 34, esextranjero, font_style2)
                ws.write(row_num, 35, inscripcion.persona.provincianacimiento.nombre if inscripcion.persona.provincianacimiento_id else '', font_style2)
                ws.write(row_num, 36, inscripcion.persona.cantonnacimiento.nombre if inscripcion.persona.cantonnacimiento_id else '', font_style2)
                row_num += 1
            workbook.close()

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo=name_document,
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": name_document,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_matrizestudiantesxfechaprimernivel_background(threading.Thread):

    def __init__(self, request, data, notiid, name_document):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.name_document = name_document
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        name_document = self.name_document
        nombre_archivo = name_document + "_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            if 'anio' in request.GET:
                __author__ = 'Unemi'
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                response = HttpResponse(content_type="application/ms-excel")
                coordinacion = None
                carrera = None
                if 'idcoor' in request.GET:
                    idcoor = request.GET['idcoor']
                    if int(idcoor) > 0:
                        coordinacion = Coordinacion.objects.get(id=int(idcoor))
                        nombbre = u"Estudiantes %s" % coordinacion.nombre
                    else:
                        nombbre = u"Estudiantes"
                else:
                    nombbre = u"Estudiantes"

                if 'idcarr' in request.GET:
                    idcarr = request.GET['idcarr']
                    if idcarr != '':
                        if int(idcarr) > 0:
                            carrera = Carrera.objects.get(id=int(idcarr))
                            nombbre += u" _ Carrera %s" % carrera.nombre_completo()

                response['Content-Disposition'] = 'attachment; filename=' + nombbre + ' ' + request.GET[
                    'anio'] + '-' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"CODIGO_IES", 6000),
                    (u"CODIGO_CARRERA", 6000),
                    (u"CIUDAD_CARRERA", 6000),
                    (u"TIPO_IDENTIFICACION", 6000),
                    (u"IDENTIFICACION", 6000),
                    (u"PRIMER_APELLIDO", 6000),
                    (u"SEGUNDO_APELLIDO", 6000),
                    (u"NOMBRES", 6000),
                    (u"SEXO", 6000),
                    (u"FECHA_NACIMIENTO", 6000),
                    (u"PAIS_ORIGEN", 6000),
                    (u"DISCAPACIDAD", 6000),
                    (u"PORCENTAJE_DISCAPACIDAD", 6000),
                    (u"NUMERO_CONADIS", 6000),
                    (u"ETNIA", 6000),
                    (u"NACIONALIDAD", 6000),
                    (u"DIRECCION", 6000),
                    (u"EMAIL_PERSONAL", 6000),
                    (u"EMAIL_INSTITUCIONAL", 6000),
                    (u"FECHA_INICIO_PRIMER_NIVEL", 6000),
                    (u"FECHA_INGRESO_CONVALIDACION", 6000),
                    (u"PAIS_RESIDENCIA", 6000),
                    (u"PROVINCIA_RESIDENCIA", 6000),
                    (u"CANTON_RESIDENCIA", 6000),
                    (u"CELULAR", 6000),
                    (u"NIVEL_FORMACION_PADRE", 6000),
                    (u"NIVEL_FORMACION_MADRE", 6000),
                    (u"CANTIDAD_MIEMBROS_HOGAR", 6000),
                    (u"TIPO_COLEGIO", 6000),
                    (u"POLITICA_CUOTA", 6000),
                    (u"CARRERA", 6000)
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                listainscriciones = Inscripcion.objects.filter(status=True, fechainicioprimernivel__year=int(
                    request.GET['anio'])).exclude(coordinacion__id__in=[6, 7, 8, 9]).order_by('persona').distinct()
                if coordinacion:
                    listainscriciones = listainscriciones.filter(carrera__coordinacion=coordinacion)

                if carrera:
                    listainscriciones = listainscriciones.filter(carrera=carrera)

                row_num = 1
                for r in listainscriciones:
                    tipoidentificacion = 'CEDULA' if r.persona.cedula else 'PASAPORTE'
                    nidentificacion = r.persona.cedula if r.persona.cedula else r.persona.pasaporte
                    tipodiscapacidad = 'NINGUNA'
                    carnetdiscapacidad = ''
                    porcientodiscapacidad = ''
                    nacionalidad = 'NO REGISTRA'
                    raza = 'NO REGISTRA'
                    politicacuota = 'NINGUNA'
                    if r.persona.perfilinscripcion_set.filter(status=True).exists():
                        pinscripcion = r.persona.perfilinscripcion_set.filter(status=True)[0]
                        if pinscripcion.tienediscapacidad:
                            if pinscripcion.tipodiscapacidad_id in [5, 1, 4, 8, 9, 7]:
                                tipodiscapacidad = u'%s' % pinscripcion.tipodiscapacidad
                                carnetdiscapacidad = pinscripcion.carnetdiscapacidad if pinscripcion.carnetdiscapacidad else ''
                                porcientodiscapacidad = pinscripcion.porcientodiscapacidad if pinscripcion.tienediscapacidad else 0
                                politicacuota = 'DISCAPACIDAD'
                        if pinscripcion.raza:
                            raza = pinscripcion.raza.nombre
                            if pinscripcion.raza.id == 1:
                                nacionalidad = u"%s" % pinscripcion.nacionalidadindigena
                                politicacuota = 'PUEBLOS Y NACIONALIDADES'
                    formacionpadre = ''
                    formacionmadre = ''
                    cantidad = 0
                    if r.persona.personadatosfamiliares_set.filter(status=True).exists():
                        if r.persona.personadatosfamiliares_set.filter(status=True, parentesco_id=1).exists():
                            formacionpadre = r.persona.personadatosfamiliares_set.filter(status=True, parentesco_id=1)[
                                0].niveltitulacion.nombrecaces if \
                                r.persona.personadatosfamiliares_set.filter(status=True, parentesco_id=1)[
                                    0].niveltitulacion else ''
                        if r.persona.personadatosfamiliares_set.filter(status=True, parentesco_id=2).exists():
                            formacionmadre = r.persona.personadatosfamiliares_set.filter(status=True, parentesco_id=2)[
                                0].niveltitulacion.nombrecaces if \
                                r.persona.personadatosfamiliares_set.filter(status=True, parentesco_id=2)[
                                    0].niveltitulacion else ''
                        cantidad = r.persona.personadatosfamiliares_set.filter(status=True).count()
                    codigocarrera = ''
                    espre = False
                    if r.coordinacion:
                        if r.coordinacion.id == 9:
                            espre = True
                    if espre:
                        codigocarrera = '00098'
                    else:
                        codigocarrera = r.mi_malla().codigo if r.mi_malla() else ''

                    if r.matricula_set.filter(nivel__periodo__anio=int(request.GET['anio'])):
                        matricula = \
                            r.matricula_set.filter(nivel__periodo__anio=int(request.GET['anio'])).order_by('id')[0]
                        if matricula.matriculagruposocioeconomico_set.filter(status=True).exists():
                            mat_gruposocioecono = matricula.matriculagruposocioeconomico_set.filter(status=True)[0]
                            if mat_gruposocioecono.gruposocioeconomico.id in [4, 5]:
                                politicacuota = 'SOCIOECONÓMICO'
                    # politicacuota = 'NINGUNA'
                    # if BecaAsignacion.objects.filter(status=True, solicitud__inscripcion=r, solicitud__estado=2, solicitud__periodo__anio=int(request.GET['anio']), solicitud__becatipo__id__in=[4, 7]).exists():
                    #     politicacuota = BecaAsignacion.objects.filter(status=True, solicitud__inscripcion=r, solicitud__estado=2, solicitud__periodo__anio=int(request.GET['anio']), solicitud__becatipo__id__in=[4, 7])[0]
                    #     politicacuota = politicacuota.solicitud.becatipo.nombre.upper()
                    ciudad_carrera = "MILAGRO"
                    ws.write(row_num, 0, '1024', font_style2)
                    ws.write(row_num, 1, codigocarrera, font_style2)
                    ws.write(row_num, 2, ciudad_carrera, font_style2)
                    ws.write(row_num, 3, tipoidentificacion if tipoidentificacion else '', font_style2)
                    ws.write(row_num, 4, nidentificacion if nidentificacion else '', font_style2)
                    ws.write(row_num, 5, r.persona.apellido1 if r.persona.apellido1 else '', font_style2)
                    ws.write(row_num, 6, r.persona.apellido2 if r.persona.apellido2 else '', font_style2)
                    ws.write(row_num, 7, r.persona.nombres if r.persona.nombres else '', font_style2)
                    ws.write(row_num, 8, r.persona.sexo.nombre if r.persona.sexo else '', font_style2)
                    ws.write(row_num, 9, r.persona.nacimiento.strftime("%d/%m/%Y") if r.persona.nacimiento else '',
                             font_style2)
                    ws.write(row_num, 10, r.persona.paisnacimiento.nombre if r.persona.paisnacimiento else '',
                             font_style2)
                    ws.write(row_num, 11, tipodiscapacidad if tipodiscapacidad else 'NINGUNA', font_style2)
                    ws.write(row_num, 12, porcientodiscapacidad if porcientodiscapacidad else 0, font_style2)
                    ws.write(row_num, 13, carnetdiscapacidad if carnetdiscapacidad else '', font_style2)
                    ws.write(row_num, 14, raza, font_style2)
                    ws.write(row_num, 15, nacionalidad, font_style2)
                    ws.write(row_num, 16, r.persona.direccion if r.persona.direccion2 else '', font_style2)
                    ws.write(row_num, 17, r.persona.email if r.persona.email else '', font_style2)
                    ws.write(row_num, 18, r.persona.emailinst if r.persona.emailinst else '', font_style2)
                    ws.write(row_num, 19,
                             r.fechainicioprimernivel.strftime("%d/%m/%Y") if r.fechainicioprimernivel else '',
                             font_style2)
                    ws.write(row_num, 20,
                             r.fechainicioconvalidacion.strftime("%d/%m/%Y") if r.fechainicioconvalidacion else '',
                             font_style2)
                    ws.write(row_num, 21, r.persona.pais.nombre if r.persona.pais else '', font_style2)
                    ws.write(row_num, 22, r.persona.provincia.nombre if r.persona.provincia else 'NO APLICA',
                             font_style2)
                    ws.write(row_num, 23, r.persona.canton.nombre if r.persona.canton else 'NO APLICA', font_style2)
                    ws.write(row_num, 24, r.persona.telefono if r.persona.telefono else '', font_style2)
                    ws.write(row_num, 25, formacionpadre if formacionpadre else 'NINGUNO', font_style2)
                    # ws.write(row_num, 24, 'NINGUNO', font_style2)
                    ws.write(row_num, 26, formacionmadre if formacionmadre else 'NINGUNO', font_style2)
                    # ws.write(row_num, 25, 'NINGUNO', font_style2)
                    ws.write(row_num, 27, cantidad if cantidad else 0, font_style2)
                    ws.write(row_num, 28, 'NO REGISTRA', font_style2)
                    ws.write(row_num, 29, politicacuota, font_style2)
                    ws.write(row_num, 30, r.carrera.nombre_completo() if r.carrera else '', font_style2)
                    row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo=name_document,
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": name_document,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_matriculaperiodoacademico_background(threading.Thread):

    def __init__(self, request, data, notiid):
        self.request = request
        self.data = data
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "reporte_matrices_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            if 'id' in request.GET:
                id_coor = int(request.GET['idfac'])
                periodo = Periodo.objects.get(pk=int(request.GET['id']))
                __author__ = 'Unemi'
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                response = HttpResponse(content_type="application/ms-excel")
                nomperiodo = u'%s' % periodo
                response['Content-Disposition'] = 'attachment; filename=Estudiantes ' + str(
                    periodo) + ' ' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"CODIGO_IES", 6000),
                    (u"CODIGO_CARRERA", 6000),
                    (u"CIUDAD_CARRERA", 6000),
                    (u"TIPO_IDENTIFICACION", 6000),
                    (u"IDENTIFICACION", 6000),
                    (u"TOTAL_CREDITOS_APROBADOS", 6000),
                    (u"CREDITOS_APROBADOS", 6000),
                    (u"TIPO_MATRICULA", 6000),
                    (u"PARALELO", 6000),
                    (u"NIVEL_ACADEMICO", 6000),
                    (u"DURACION_PERIODO_ACADEMICO", 6000),
                    (u"NUM_MATERIAS_SEGUNDA_MATRICULA", 6000),
                    (u"NUM_MATERIAS_TERCERA_MATRICULA", 6000),
                    (u"PERDIDA_GRATUIDAD", 6000),
                    (u"PENSION_DIFERENCIADA", 6000),
                    (u"PLAN_CONTINGENCIA", 6000),
                    (u"INGRESO_TOTAL_HOGAR", 6000),
                    (u"ORIGEN_RECURSOS_ESTUDIOS", 6000),
                    (u"TERMINO_PERIODO", 6000),
                    (u"TOTAL_HORAS_APROBADAS", 6000),
                    (u"HORAS_APROBADAS_PERIODO", 6000),
                    (u"MONTO_AYUDA_ECONOMICA", 6000),
                    (u"MONTO_CREDITO_EDUCATIVO", 6000),
                    (u"ESTADO", 6000),
                    (u"CARRERA", 6000),
                    (u"PASAPORTE", 6000),
                    (u"FACULTAD", 6000),
                    (u"MODALIDAD", 6000),
                    (u"BENEFICIARIO DE BECA", 6000),
                    (u"LGTBI", 6000),
                    (u"EXTRANJERO", 6000)
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                cursor = connection.cursor()
                listaestudiante = f"""
                        SELECT 
                            mat.id, 
                            COUNT(mta.id) FILTER(WHERE i.coordinacion_id not in (7) AND asi.modulo) "num_modulo",
                            COUNT(mta.id) "num_total"
                        FROM
                            sga_inscripcion i
                            INNER JOIN sga_matricula mat ON mat.inscripcion_id=i.id
                            INNER JOIN sga_nivel niv ON niv.id=mat.nivel_id
                            INNER JOIN sga_nivelmalla nivm ON nivm.id=mat.nivelmalla_id
                            INNER JOIN sga_periodo peri ON peri.id=niv.periodo_id
                            INNER JOIN sga_materiaasignada mta ON mta.matricula_id=mat.id
                            INNER JOIN sga_materia mt ON mt.id=mta.materia_id
                            INNER JOIN sga_asignatura asi ON asi.id=mt.asignatura_id
                        WHERE 
                            mat."status"
                            AND peri.anio={periodo.anio}
                            and mat.estado_matricula in (2,3)
                            AND mat.retiradomatricula = False
	                        AND mta.retiramateria = False
                        GROUP BY mat.id
                        HAVING 
                            COUNT(mta.id) FILTER(WHERE i.coordinacion_id not in (7) AND asi.modulo)=COUNT(mta.id)
                """
                cursor.execute(listaestudiante)
                results = cursor.fetchall()
                respuestas = []
                for per in results:
                    respuestas.append(per[0])
                matriculados = Matricula.objects.filter(nivel__periodo=periodo, inscripcion__coordinacion_id=id_coor, estado_matricula__in=[2, 3], status=True).exclude(pk__in=respuestas).exclude(retiromatricula__isnull=False)
                # matriculados = Matricula.objects.filter(nivel__periodo=periodo, estado_matricula__in=[2,3], status=True).exclude(pk__in=respuestas).exclude(retiromatricula__isnull=False)[:100]
                row_num = 1
                duracion = 0
                resta = periodo.fin - periodo.inicio
                wek, dias = divmod(resta.days, 7)
                for mat in matriculados:
                    espre = False
                    if mat.inscripcion.coordinacion:
                        if mat.inscripcion.coordinacion.id == 9:
                            espre = True
                    if espre:
                        codigocarrera = '00098'
                    else:
                        codigocarrera = mat.inscripcion.mi_malla().codigo if mat.inscripcion.mi_malla() else ''
                    carrera = mat.inscripcion.mi_malla().carrera.nombre_completo()
                    nompasaporte = mat.inscripcion.persona.pasaporte
                    nomfacultad = mat.inscripcion.coordinacion.nombre
                    nommodalidad = mat.inscripcion.modalidad.nombre
                    eslgtbi = 'NO'
                    if mat.inscripcion.persona.lgtbi:
                        eslgtbi = 'SI'
                    esextranjero = 'SI'
                    if mat.inscripcion.persona.paisnacimiento:
                        if mat.inscripcion.persona.paisnacimiento.id == 1:
                            esextranjero = 'NO'

                    benebeca = ''
                    alumnobeca = BecaAsignacion.objects.filter(status=True, solicitud__inscripcion=mat.inscripcion, solicitud__periodo=periodo, solicitud__becatipo_id__isnull=False).first()
                    if alumnobeca:
                        benebeca = alumnobeca.solicitud.becatipo.nombre
                    tipoidentificacion = 'CEDULA' if mat.inscripcion.persona.cedula else 'PASAPORTE'
                    nidentificacion = mat.inscripcion.persona.cedula if mat.inscripcion.persona.cedula else mat.inscripcion.persona.pasaporte
                    if mat.inscripcion.coordinacion.id == 9:
                        paralelo = 'NO APLICA'
                        nivel = 'NIVELACION'
                    else:
                        paralelo = nivel_enletra_malla(mat.nivelmalla.orden)
                        nivel = paralelo_enletra_nivel(mat.nivelmalla.orden)
                    numsegundamat = mat.materiaasignada_set.filter(status=True, matriculas=2).count()
                    numterceramat = mat.materiaasignada_set.filter(status=True, matriculas=3).count()
                    total_ingreso = sum([x.ingresomensual for x in
                                         mat.inscripcion.persona.personadatosfamiliares_set.filter(status=True)])
                    nombre_ingreso = ingreso_total_hogar_rangos(total_ingreso)
                    total_horasmat = mat.total_horas_matricula()
                    tipomatricula = ''
                    if mat.tipomatricula:
                        tipomatricula = mat.tipomatricula.nombre
                        if mat.tipomatricula.id == 1:
                            tipomatricula = 'ORDINARIA'
                        if mat.tipomatricula.id == 4:
                            tipomatricula = 'ESPECIAL'

                    ciudad_carrera = "MILAGRO"
                    estado = ""
                    terminado_per = ""
                    if mat.cantidad_materias_aprobadas(periodo) > 0:
                        estado = "APROBADO"
                        terminado_per = "SI"
                    else:
                        estado = "NO APROBADO"
                        terminado_per = "SI"
                    if RetiroMatricula.objects.filter(status=True, matricula=mat).exists():
                        estado = "RETIRADO"
                        terminado_per = "NO"
                    if FichaSocioeconomicaINEC.objects.values('id').filter(status=True,
                                                                           persona=mat.inscripcion.persona).exists():
                        ficha = FichaSocioeconomicaINEC.objects.filter(status=True,
                                                                       persona=mat.inscripcion.persona).order_by(
                            'id').last()
                        origen_recursos = u"%s" % ficha.personacubregasto if ficha.personacubregasto else "NO REGISTRA"
                    ws.write(row_num, 0, '1024', font_style2)
                    ws.write(row_num, 1, codigocarrera, font_style2)
                    ws.write(row_num, 2, ciudad_carrera, font_style2)
                    ws.write(row_num, 3, tipoidentificacion, font_style2)
                    ws.write(row_num, 4, nidentificacion, font_style2)
                    ws.write(row_num, 5, mat.inscripcion.total_creditos(), font_style2)
                    ws.write(row_num, 6, mat.total_creditos_matricula(), font_style2)
                    ws.write(row_num, 7, tipomatricula, font_style2)
                    ws.write(row_num, 8, paralelo, font_style2)
                    ws.write(row_num, 9, nivel, font_style2)
                    ws.write(row_num, 10, u"%s" % (wek), font_style2)
                    ws.write(row_num, 11, numsegundamat, font_style2)
                    ws.write(row_num, 12, numterceramat, font_style2)
                    ws.write(row_num, 13, 'SI' if mat.inscripcion.gratuidad else 'NO', font_style2)
                    ws.write(row_num, 14, 'NO', font_style2)
                    ws.write(row_num, 15, 'NO', font_style2)
                    ws.write(row_num, 16, nombre_ingreso, font_style2)
                    ws.write(row_num, 17, origen_recursos, font_style2)
                    ws.write(row_num, 18, u"%s" % (terminado_per), font_style2)
                    ws.write(row_num, 19, mat.inscripcion.total_horas(), font_style2)
                    ws.write(row_num, 20, total_horasmat, font_style2)
                    ws.write(row_num, 21, '0', font_style2)
                    ws.write(row_num, 22, '0', font_style2)
                    ws.write(row_num, 23, estado, font_style2)
                    ws.write(row_num, 24, carrera, font_style2)
                    ws.write(row_num, 25, nompasaporte, font_style2)
                    ws.write(row_num, 26, nomfacultad, font_style2)
                    ws.write(row_num, 27, nommodalidad, font_style2)
                    ws.write(row_num, 28, benebeca, font_style2)
                    ws.write(row_num, 29, eslgtbi, font_style2)
                    ws.write(row_num, 30, esextranjero, font_style2)
                    row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Excel Estudiantes - Matrícula Periodo Académico',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel Estudiantes - Matrícula Periodo Académico',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_matrizpracticaspreprofesionales_anio_background(threading.Thread):

    def __init__(self, request, data, notiid, name_document):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.name_document = name_document
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid

        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        name_document = self.name_document
        nombre_archivo = name_document + "_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            if 'anio' in request.GET:
                __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('exp_xls_post_part')
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = 'attachment; filename=Practicas- ' + request.GET[
                'anio'] + '-' + random.randint(1, 10000).__str__() + '.xls'
            columns = [
                (u"CODIGO_IES", 6000),
                (u"CODIGO_CARRERA", 6000),
                (u"CIUDAD_CARRERA", 6000),
                (u"TIPO_IDENTIFICACION", 6000),
                (u"IDENTIFICACION", 6000),
                (u"NOMBRE_INSTITUCION", 6000),
                (u"TIPO_INSTITUCION", 6000),
                (u"FECHA_INICIO", 6000),
                (u"FECHA_FIN", 6000),
                (u"NUMERO_HORAS", 6000),
                (u"CAMPO_ESPECIFICO", 6000),
                (u"IDENTIFICACION_DOCENTE_TUTOR", 6000),
                (u"TIPO", 6000)
            ]
            row_num = 0
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            listapracticas = PracticasPreprofesionalesInscripcion.objects.filter(status=True, culminada=True,
                                                                                 fechadesde__year=int(
                                                                                     request.GET['anio'])).exclude(
                inscripcion__coordinacion__id=1).distinct()
            row_num = 1
            for p in listapracticas.order_by('fechadesde'):
                nidentificacion = p.inscripcion.persona.cedula if p.inscripcion.persona.cedula else p.inscripcion.persona.pasaporte
                tipoidentificacion = 'CEDULA' if p.inscripcion.persona.cedula else 'PASAPORTE'
                identificaciondocentetutor = ''
                if p.tutorunemi:
                    identificaciondocentetutor = p.tutorunemi.persona.cedula if p.tutorunemi else p.tutorunemi.persona.pasaporte
                if p.tiposolicitud == 3:
                    numerohoras = p.horahomologacion if p.horahomologacion else ''
                else:
                    numerohoras = p.numerohora if p.numerohora else ''
                campo_Especifico = "NINGUNO"
                if p.inscripcion.carrera.malla_set.values('id').all().exists():
                    if p.inscripcion.carrera.malla_set.all()[0].campo_especifico:
                        campo_Especifico = u"%s" % p.inscripcion.carrera.malla_set.all()[0].campo_especifico.codigo
                ws.write(row_num, 0, '1024', font_style2)
                ws.write(row_num, 1, p.inscripcion.mi_malla().codigo if p.inscripcion.mi_malla() else '', font_style2)
                ws.write(row_num, 2, u"MILAGRO", font_style2)
                ws.write(row_num, 3, tipoidentificacion if tipoidentificacion else '', font_style2)
                ws.write(row_num, 4, nidentificacion if nidentificacion else '', font_style2)
                ws.write(row_num, 5, p.traer_empresa(), font_style2)
                ws.write(row_num, 6, p.get_tipoinstitucion_display() if p.tipoinstitucion else '', font_style2)
                ws.write(row_num, 7, p.fechadesde.strftime("%d-%m-%Y") if p.fechadesde else '', font_style2)
                ws.write(row_num, 8, p.fechahasta.strftime("%d-%m-%Y") if p.fechahasta else '', font_style2)
                ws.write(row_num, 9, numerohoras, font_style2)
                ws.write(row_num, 10, campo_Especifico, font_style2)
                ws.write(row_num, 11, identificaciondocentetutor, font_style2)
                ws.write(row_num, 12, p.get_tipo_display(), font_style2)
                row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo=name_document,
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": name_document,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_matrizmatrizgraduados_anio_background(threading.Thread):

    def __init__(self, request, data, notiid, name_document):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.name_document = name_document
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid

        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        name_document = self.name_document.strip()
        nombre_archivo = name_document + "_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            if 'anio' in request.GET:
                __author__ = 'Unemi'
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=Practicas- ' + request.GET[
                    'anio'] + '-' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"CODIGO_IES", 6000),
                    (u"CODIGO_CARRERA", 6000),
                    (u"CIUDAD_CARRERA", 6000),
                    (u"TIPO_IDENTIFICACION", 6000),
                    (u"IDENTIFICACION", 6000),
                    (u"PRIMER_APELLIDO", 6000),
                    (u"SEGUNDO_APELLIDO", 6000),
                    (u"NOMBRES", 6000),
                    (u"SEXO", 6000),
                    (u"FECHA_NACIMIENTO", 6000),
                    (u"PAIS_ORIGEN", 6000),
                    (u"DISCAPACIDAD", 6000),
                    (u"NUMERO_CONADIS", 6000),
                    (u"DIRECCION", 6000),
                    (u"EMAIL_PERSONAL", 6000),
                    (u"EMAIL_INSTITUCIONAL", 6000),
                    (u"FECHA_INICIO_PRIMER_NIVEL", 6000),
                    (u"FECHA_INGRESO_CONVALIDACION", 6000),
                    (u"FECHA_GRADUACION", 6000),
                    (u"MECANISMO_TITULACION", 6000),
                    (u"CARRERA", 6000),
                    (u"PASAPORTE", 6000),
                    (u"FACULTAD", 6000),
                    (u"ETNIA", 6000),
                    (u"LGTBI", 6000),
                    (u"ES EXTRANJERO", 6000),
                    (u"FECHA REFRENDACION", 6000)
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                coordinacion = None
                carrera = None
                if 'idcoor' in request.GET:
                    idcoor = request.GET['idcoor']
                    if int(idcoor) > 0:
                        coordinacion = Coordinacion.objects.get(id=int(idcoor))
                        nombbre = u"Graduados %s" % coordinacion.nombre
                    else:
                        nombbre = u"Graduados"
                else:
                    nombbre = u"Graduados"

                if 'idcarr' in request.GET:
                    idcarr = request.GET['idcarr']
                    if idcarr != '':
                        if int(idcarr) > 0:
                            carrera = Carrera.objects.get(id=int(idcarr))
                            nombbre += u" _ Carrera %s" % carrera.nombre_completo()

                graduados = Graduado.objects.filter(status=True, estadograduado=True,
                                                    fechagraduado__year=int(request.GET['anio']))
                if coordinacion:
                    graduados = graduados.filter(inscripcion__carrera__coordinacion=coordinacion)

                if carrera:
                    graduados = graduados.filter(inscripcion__carrera=carrera)

                row_num = 1
                for g in graduados:
                    nacionalidad = 'NO APLICA'
                    raza = 'NO REGISTRA'
                    nidentificacion = g.inscripcion.persona.cedula if g.inscripcion.persona.cedula else g.inscripcion.persona.pasaporte
                    tipoidentificacion = 'CEDULA' if g.inscripcion.persona.cedula else 'PASAPORTE'
                    tipodiscapacidad = 'NINGUNA'
                    carnetdiscapacidad = ''
                    if g.inscripcion.persona.perfilinscripcion_set.filter(status=True).exists():
                        pinscripcion = g.inscripcion.persona.perfilinscripcion_set.filter(status=True)[0]
                        tipodiscapacidad = pinscripcion.tipodiscapacidad.nombre if pinscripcion.tipodiscapacidad else 'NINGUNA'
                        carnetdiscapacidad = pinscripcion.carnetdiscapacidad if pinscripcion.carnetdiscapacidad else ''
                        if pinscripcion.raza:
                            raza = pinscripcion.raza.__str__()
                            if pinscripcion.raza.id == 1:
                                politica_cuota = 'PUEBLOS Y NACIONALIDADES'
                                nacionalidad = pinscripcion.nacionalidadindigena.__str__()

                    mecanismot = ''
                    if g.codigomecanismotitulacion:
                        if g.codigomecanismotitulacion.mecanismotitulacion:
                            mecanismot = g.codigomecanismotitulacion.mecanismotitulacion.nombre
                    ciudad_carrera = "MILAGRO"
                    eslgtbi = 'NO'
                    if g.inscripcion.persona.lgtbi:
                        eslgtbi = 'SI'
                    esextranjero = 'SI'
                    if g.inscripcion.persona.paisnacimiento:
                        if g.inscripcion.persona.paisnacimiento.id == 1:
                            esextranjero = 'NO'

                    ws.write(row_num, 0, '1024', font_style2)
                    ws.write(row_num, 1, g.inscripcion.mi_malla().codigo if g.inscripcion.mi_malla() else '', font_style2)
                    ws.write(row_num, 2, ciudad_carrera, font_style2)
                    ws.write(row_num, 3, tipoidentificacion, font_style2)
                    ws.write(row_num, 4, nidentificacion, font_style2)
                    ws.write(row_num, 5, g.inscripcion.persona.apellido1, font_style2)
                    ws.write(row_num, 6, g.inscripcion.persona.apellido2, font_style2)
                    ws.write(row_num, 7, g.inscripcion.persona.nombres, font_style2)
                    ws.write(row_num, 8, g.inscripcion.persona.sexo.nombre if g.inscripcion.persona.sexo else '', font_style2)
                    ws.write(row_num, 9, g.inscripcion.persona.nacimiento.strftime("%d/%m/%Y") if g.inscripcion.persona.nacimiento else '', font_style2)
                    ws.write(row_num, 10, g.inscripcion.persona.pais.nombre if g.inscripcion.persona.pais else '', font_style2)
                    ws.write(row_num, 11, tipodiscapacidad, font_style2)
                    ws.write(row_num, 12, carnetdiscapacidad, font_style2)
                    ws.write(row_num, 13, g.inscripcion.persona.direccion if g.inscripcion.persona.direccion else g.inscripcion.persona.direccion2, font_style2)
                    ws.write(row_num, 14, g.inscripcion.persona.email if g.inscripcion.persona.email else '', font_style2)
                    ws.write(row_num, 15, g.inscripcion.persona.emailinst if g.inscripcion.persona.emailinst else '', font_style2)
                    ws.write(row_num, 16, g.inscripcion.fechainicioprimernivel.strftime("%d/%m/%Y") if g.inscripcion.fechainicioprimernivel else '', font_style2)
                    ws.write(row_num, 17, g.inscripcion.fechainicioconvalidacion.strftime("%d/%m/%Y") if g.inscripcion.fechainicioconvalidacion else '', font_style2)
                    ws.write(row_num, 18, g.fechagraduado.strftime("%d/%m/%Y") if g.fechagraduado else '', font_style2)
                    ws.write(row_num, 19, mecanismot if mecanismot else 'TRABAJO TITULACIÓN ', font_style2)
                    ws.write(row_num, 20, g.inscripcion.carrera.nombre_completo(), font_style2)
                    ws.write(row_num, 21, g.inscripcion.persona.pasaporte.__str__(), font_style2)
                    ws.write(row_num, 22, g.inscripcion.coordinacion.nombre.__str__(), font_style2)
                    ws.write(row_num, 23, raza, font_style2)
                    ws.write(row_num, 24, eslgtbi, font_style2)
                    ws.write(row_num, 25, esextranjero, font_style2)
                    ws.write(row_num, 26, g.fecharefrendacion.strftime("%d/%m/%Y") if g.fecharefrendacion else '', font_style2)
                    row_num += 1
                wb.save(directory)
                usernotify = User.objects.get(pk=request.user.pk)
                pers = Persona.objects.get(usuario=usernotify)
                if notiid > 0:
                    noti = Notificacion.objects.get(pk=notiid)
                    noti.en_proceso = False
                    noti.cuerpo = 'Excel Listo'
                    noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                    noti.save()
                else:
                    noti = Notificacion(cuerpo='Reporte Listo', titulo=name_document,
                                        destinatario=pers,
                                        url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo), prioridad=1,
                                        app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                        tipo=2, en_proceso=False)
                    noti.save(request)

                send_user_notification(user=usernotify, payload={
                    "head": "Excel terminado",
                    "body": name_document,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_pre_matriculados_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.periodo = periodo
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'prematricula')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "reporte_planificacion_matricula{}.xlsx".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'prematricula', nombre_archivo)
        usernotify = User.objects.get(pk=request.user.pk)
        pers = Persona.objects.get(usuario=usernotify)

        try:
            columns = [
                (u"N°", 5),
                (u"FACULTAD", 50),
                (u"CARRERA", 50),
                (u"ESTUDIANTE", 40),
                (u"CEDULA", 13),
                (u"NIVEL", 30),
                (u"PERIODO ACADEMICO", 60),
                (u"MALLA", 100),
                (u"ASIGNATURA", 100),
                (u"JORNADA OFERTADA", 40),
                (u"JORNADA SUGERIDA", 40)
            ]
            workbook = xlsxwriter.Workbook(directory)
            ws = workbook.add_worksheet('DETALLE')

            formatocabeceracolumna = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': 'silver',
                'text_wrap': 1,
                'font_size': 10})

            formatocelda = workbook.add_format({
                'border': 1
            })

            formatotitulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 16})

            ws.merge_range(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)

            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])

            row_num = 4
            c = 1
            for asignatura in data:
                asigmalla = AsignaturaMalla.objects.filter(status=True, asignatura=asignatura.asignatura,
                                                           malla=asignatura.prematricula_set.first().inscripcion.malla_inscripcion().malla).first()
                alumno = asignatura.prematricula_set.first()
                campo1 = str(c)
                campo4 = str(alumno.inscripcion.carrera.coordinacion_set.first().nombre)
                campo5 = str(alumno.inscripcion.carrera.nombre_completo())
                campo2 = str(alumno.inscripcion)
                campo3 = str(alumno.inscripcion.persona.cedula)
                campo6 = str(asigmalla.nivelmalla)
                campo7 = str(alumno.periodo)
                campo8 = str(alumno.inscripcion.malla_inscripcion().malla)
                campo9 = str(asignatura.asignatura)
                campo10 = str(alumno.sesion)
                campo11 = str(asignatura.sesion)

                ws.write(row_num, 0, campo1, formatocelda)
                ws.write(row_num, 1, campo4, formatocelda)
                ws.write(row_num, 2, campo5, formatocelda)
                ws.write(row_num, 3, campo2, formatocelda)
                ws.write(row_num, 4, campo3, formatocelda)
                ws.write(row_num, 5, campo6, formatocelda)
                ws.write(row_num, 6, campo7, formatocelda)
                ws.write(row_num, 7, campo8, formatocelda)
                ws.write(row_num, 8, campo9, formatocelda)
                ws.write(row_num, 9, campo10, formatocelda)
                ws.write(row_num, 10, campo11, formatocelda)
                row_num += 1
                c += 1
            # segunda hoja para totales

            columns = [
                (u"FACULTAD", 50),
                (u"CARRERA", 50),
                (u"ASIGNATURA", 30),
                (u"NIVEL", 20),
                (u"JORNADA OFERTADA", 20),
                (u"JORNADA SUGERIDA", 20),
                (u"CANTIDAD", 20)
            ]
            ws = workbook.add_worksheet('TOTALES')
            ws.merge_range(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)

            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])

            row_num = 4
            c = 1
            # datatotal = data.annotate(cantidad=Count('id'))
            datatotal = data.values('asignatura_id', 'asignatura__nombre',
                                    'prematricula__inscripcion__coordinacion__nombre',
                                    'prematricula__inscripcion__carrera__nombre',
                                    'prematricula__inscripcion__inscripcionmalla__malla',
                                    'prematricula__sesion__nombre', 'sesion__nombre').annotate(
                cantidad=Count('asignatura_id')).order_by('asignatura_id')
            for total in datatotal:
                asigmalla = AsignaturaMalla.objects.filter(status=True, asignatura=total['asignatura_id'], malla=total[
                    'prematricula__inscripcion__inscripcionmalla__malla']).first()
                # alumno = total.prematricula_set.first()
                campo1 = str(total['prematricula__inscripcion__coordinacion__nombre'])
                campo2 = str(total['prematricula__inscripcion__carrera__nombre'])
                campo3 = str(total['asignatura__nombre'])
                campo4 = str(asigmalla.nivelmalla)
                campo5 = str(total['prematricula__sesion__nombre'])
                campo6 = str(total['sesion__nombre'])
                campo7 = str(total['cantidad'])
                ws.write(row_num, 0, campo1, formatocelda)
                ws.write(row_num, 1, campo2, formatocelda)
                ws.write(row_num, 2, campo3, formatocelda)
                ws.write(row_num, 3, campo4, formatocelda)
                ws.write(row_num, 4, campo5, formatocelda)
                ws.write(row_num, 5, campo6, formatocelda)
                ws.write(row_num, 6, campo7, formatocelda)
                row_num += 1
            workbook.close()

            response = HttpResponse(directory,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=%s' % nombre_archivo
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/prematricula/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Excel Reporte de Planificación de Matriculas',
                                    destinatario=pers,
                                    url="{}reportes/prematricula/{}".format(MEDIA_URL, nombre_archivo), prioridad=1,
                                    app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Reporte de Planificación de Matriculas',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/prematricula/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": "Su reporte ha sido terminado"
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))

            textoerror = 'Reporte Fallido - {} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.error = True
                noti.titulo = 'Reporte de Planificación de Matriculas falló en la ejecución'
                noti.cuerpo = textoerror
                # noti.url = "{}reportes/prematricula/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Fallido',
                                    titulo='Reporte de Planificación de Matriculas falló en la ejecución',
                                    destinatario=pers, prioridad=1, app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False,
                                    error=True)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte Fallido",
                "body": 'Reporte de Planificación de Matriculas a fallado',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": textoerror,
                "error": True
            }, ttl=500)


class reporte_encuesta_grupo_estudiante_background(threading.Thread):

    def __init__(self, request, notiid, encuesta):
        self.request = request
        self.encuesta = encuesta
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'encuestas')
        request, notiid, encuesta = self.request, self.notiid, self.encuesta
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        # nombre_archivo = "Resultado_encuesta{}.xls".format(random.randint(1, 10000).__str__())
        # directory = os.path.join(MEDIA_ROOT, 'reportes', 'encuestas', nombre_archivo)
        try:
            __author__ = 'Unemi'

            directorio = os.path.join(os.path.join(MEDIA_ROOT, 'reportes', 'encuestas'))
            try:
                os.stat(directorio)
            except:
                os.mkdir(directorio)
            nombrearchivo = "ENCUESTA_GRUPO_ESTUDIANTES" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"

            wb = xlsxwriter.Workbook(directorio + '/' + nombrearchivo)
            ws = wb.add_worksheet("datos")

            # Estilos de celda
            style0 = wb.add_format(
                {'font_name': 'Times New Roman', 'color': 'blue', 'bold': False, 'num_format': '#,##0.00'})
            style_nb = wb.add_format(
                {'font_name': 'Times New Roman', 'color': 'blue', 'bold': True, 'num_format': '#,##0.00'})
            style_sb = wb.add_format({'font_name': 'Times New Roman', 'color': 'blue', 'bold': True})
            title = wb.add_format({'font_name': 'Times New Roman', 'color': 'blue', 'bold': True, 'align': 'center'})
            style1 = wb.add_format({'num_format': 'dd-mmm-yy'})

            # Estilo de fuente
            font_style = wb.add_format({'bold': True})

            # Estilo de fuente 2
            font_style2 = wb.add_format()

            # wb = Workbook(encoding='utf-8')
            # ws = wb.add_sheet('datos')
            ws.merge_range(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = 'attachment; filename=datos_' + random.randint(1,
                                                                                             10000).__str__() + '.xls'
            preguntas = encuesta.preguntaencuestagrupoestudiantes_set.filter(status=True).order_by('orden')
            columns = [
                (u"Nº.", 2000),
                (u"ID", 2000),
                (u"CÉDULA", 3000),
                (u"ENCUESTADO", 9000),
            ]
            if encuesta.tipoperfil == 1:  # ALUMNO
                columns.append((u'CARRERA', 9000), )
                # solo para encuesta con y sin discapacidad
                if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23 or encuesta.id == 46 or encuesta.id == 47 or encuesta.id == 48:  # encuestas
                    columns.append((u'TIENE DISCAPACIDAD', 3000), )
                # fin
                # correo y movil
                if encuesta.id == 19 or encuesta.id == 46 or encuesta.id == 47 or encuesta.id == 48:  # encuesta
                    columns.append((u'CORREO PERSONAL', 3000), )
                    columns.append((u'CORREO INSTITUCIONAL', 3000), )
                    columns.append((u'TELEFONO MÓVIL', 3000), )
                    columns.append((u'TELEFONO FIJO', 3000), )
                    columns.append((u'PAIS', 3000), )
                    columns.append((u'PROVINCIA', 3000), )
                    columns.append((u'CANTON', 3000), )
                    columns.append((u'PARROQUIA', 3000), )
                    columns.append((u'CIUDAD', 3000), )
                    columns.append((u'CIUDADADELA', 3000), )
                    columns.append((u'CALLE PRINCIPAL', 3000), )
                    columns.append((u'CALLE SECUNDARIA', 3000), )
                # fin

            if encuesta.tipoperfil == 2:  # DOCENTE
                columns.append((u'Tipo de relación laboral', 9000), )
                columns.append((u'Tiempo de dedicación', 9000), )
                columns.append((u'Si es docente titular a qué categoría académica pertenece', 9000), )
                # columns.append((u'MODALIDAD CONTRATACIÓN', 9000), )
                columns.append((u'A qué Facultad pertenece', 9000), )
                columns.append(
                    (u'Las carreras en las que imparte docencia actualmente ¿De qué modalidad son? ', 9000), )
                # columns.append((u'MODALIDAD DE LA CARRERA QUE DESEARÍA TRABAJAR EN EL SEMESTRE 1S 2022', 9000), )

            if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
                columns.append((u'Tipo de relación laboral', 9000), )
                columns.append((u'Denominación del puesto', 9000), )
            if encuesta.tipoperfil == 4:  # PERSONAS
                columns.append((u'Tipo de relación laboral', 9000), )
                columns.append((u'Denominación del puesto', 9000), )

            for x in preguntas:
                columns.append((str(x.orden) + ") " + x.descripcion, 6000), )
                if x.tipo == 1:
                    if not x.esta_vacia():
                        columns.append((str(x.orden) + ") " + x.observacionporno, 6000), )
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                # set_column
            row_num = 4
            i = 0

            if encuesta.tipoperfil == 1:  # ALUMNO
                datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by(
                    'inscripcion__persona__apellido1')
            if encuesta.tipoperfil == 2:  # DOCENTE
                datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by(
                    'profesor__persona__apellido1')
            if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
                datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by(
                    'administrativo__persona__apellido1')
            if encuesta.tipoperfil == 4:  # ADMINISTRATIVO
                datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by(
                    'persona__apellido1')
            cout_register = datos.count()
            register_start = 0
            limit = 0
            for dato in datos:
                row_num += limit
                i += 1
                limit = 0
                ws.write(row_num, 0, i, font_style2)
                ws.write(row_num, 1, dato.id, font_style2)
                if encuesta.tipoperfil == 1:
                    ws.write(row_num, 2, dato.inscripcion.persona.cedula, font_style2)
                    ws.write(row_num, 3, dato.inscripcion.persona.nombre_completo_inverso(), font_style2)
                if encuesta.tipoperfil == 2:
                    ws.write(row_num, 2, dato.profesor.persona.cedula, font_style2)
                    ws.write(row_num, 3, dato.profesor.persona.nombre_completo_inverso(), font_style2)
                if encuesta.tipoperfil == 3:
                    ws.write(row_num, 2, dato.administrativo.persona.cedula, font_style2)
                    ws.write(row_num, 3, dato.administrativo.persona.nombre_completo_inverso(), font_style2)
                if encuesta.tipoperfil == 4:
                    ws.write(row_num, 2, dato.persona.cedula, font_style2)
                    ws.write(row_num, 3, dato.persona.nombre_completo_inverso(), font_style2)
                c = 4
                if encuesta.tipoperfil == 1:
                    ws.write(row_num, c, dato.inscripcion.carrera.__str__(),
                             font_style2) if not dato.inscripcion.carrera == None else ' '
                    c += 1

                # solo para encuesta con y sin discapacidad
                if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23 or encuesta.id == 46 or encuesta.id == 47 or encuesta.id == 48:  # encuesta
                    if dato.inscripcion.persona.mi_perfil().tienediscapacidad:
                        discapacidad = 'SI'
                    else:
                        discapacidad = 'NO'

                    ws.write(row_num, c, discapacidad, font_style2)
                    c += 1
                # fin

                if encuesta.id == 19 or encuesta.id == 46 or encuesta.id == 47 or encuesta.id == 48:
                    email = dato.inscripcion.persona.email
                    ws.write(row_num, c, email, font_style2)
                    c += 1
                    emailinst = dato.inscripcion.persona.emailinst
                    ws.write(row_num, c, emailinst, font_style2)
                    c += 1
                    telefono = dato.inscripcion.persona.telefono
                    ws.write(row_num, c, telefono, font_style2)
                    c += 1
                    telefono_conv = dato.inscripcion.persona.telefono_conv
                    ws.write(row_num, c, telefono_conv, font_style2)
                    c += 1

                    paiss = dato.inscripcion.persona.pais
                    ws.write(row_num, c, paiss.__str__(), font_style2)
                    c += 1

                    provincia = dato.inscripcion.persona.provincia
                    ws.write(row_num, c, provincia.__str__(), font_style2)
                    c += 1

                    canton = dato.inscripcion.persona.canton
                    ws.write(row_num, c, canton.__str__(), font_style2)
                    c += 1

                    parroquia = dato.inscripcion.persona.parroquia
                    ws.write(row_num, c, parroquia.__str__(), font_style2)
                    c += 1

                    ciudad = dato.inscripcion.persona.ciudad
                    ws.write(row_num, c, ciudad, font_style2)
                    c += 1

                    ciudadela = dato.inscripcion.persona.ciudadela
                    ws.write(row_num, c, ciudadela, font_style2)
                    c += 1

                    calleprincipal = dato.inscripcion.persona.direccion
                    ws.write(row_num, c, calleprincipal, font_style2)
                    c += 1
                    callesecundaria = dato.inscripcion.persona.direccion2
                    ws.write(row_num, c, callesecundaria, font_style2)
                    c += 1
                # fin

                if encuesta.tipoperfil == 2:
                    dt = ProfesorDistributivoHoras.objects.filter(status=True, periodo=126,
                                                                  profesor_id=dato.profesor.id).first()
                    ws.write(row_num, c,
                             dt.nivelcategoria.nombre if dt is not None and dt.nivelcategoria is not None else '',
                             font_style2)
                    c += 1
                    ws.write(row_num, c, dt.dedicacion.nombre if dt is not None and dt.dedicacion is not None else '',
                             font_style2)
                    c += 1
                    ws.write(row_num, c,
                             dt.categoria.nombre if dt is not None and dt.nivelcategoria is not None and dt.nivelcategoria.id == 1 else '',
                             font_style2)
                    c += 1
                    ws.write(row_num, c,
                             dt.coordinacion.nombre if dt is not None and dt.coordinacion is not None else '',
                             font_style2)
                    c += 1
                    w = 0
                    for m in dato.profesor.mis_materias(126).values_list('materia__nivel__modalidad__nombre',
                                                                         flat=True).distinct(
                        'materia__nivel__modalidad__nombre'):
                        ws.write(row_num + w, c, str(m), font_style2)
                        w += 1
                    if limit < w and w > 0:
                        limit = w - 1

                    c += 1
                if encuesta.tipoperfil == 3:
                    eDistributivoPersonas = DistributivoPersona.objects.filter(persona=dato.administrativo.persona,
                                                                               status=True, regimenlaboral_id=2,
                                                                               estadopuesto_id=1)
                    eDistributivoPersona = None
                    if eDistributivoPersonas.values("id").exists():
                        eDistributivoPersona = eDistributivoPersonas.first()
                    ws.write(row_num, c,
                             eDistributivoPersona.regimenlaboral.descripcion if eDistributivoPersona is not None else '',
                             font_style2)
                    c += 1
                    ws.write(row_num, c,
                             eDistributivoPersona.denominacionpuesto.descripcion if eDistributivoPersona is not None else '',
                             font_style2)
                    c += 1
                if encuesta.tipoperfil == 4:
                    eDistributivoPersonas = DistributivoPersona.objects.filter(persona=dato.persona, status=True,
                                                                               regimenlaboral_id=2, estadopuesto_id=1)
                    eDistributivoPersona = None
                    if eDistributivoPersonas.values("id").exists():
                        eDistributivoPersona = eDistributivoPersonas.first()
                    ws.write(row_num, c,
                             eDistributivoPersona.regimenlaboral.descripcion if eDistributivoPersona is not None else '',
                             font_style2)
                    c += 1
                    ws.write(row_num, c,
                             eDistributivoPersona.denominacionpuesto.descripcion if eDistributivoPersona is not None else '',
                             font_style2)
                    c += 1

                for x in preguntas:
                    respuesta = None
                    if x.tipo == 1:
                        respuesta = RespuestaPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                             inscripcionencuesta=dato)
                        if respuesta.values("id").exists():
                            respuesta = respuesta.first()
                            ws.write(row_num, c, respuesta.respuesta, font_style2)
                        else:
                            respuesta = None
                            ws.write(row_num, c, '', font_style2)
                        c += 1

                    if x.tipo == 2:
                        respuesta = dato.respuestarangoencuestagrupoestudiantes_set.filter(status=True,
                                                                                           pregunta=x).first() if dato.respuestarangoencuestagrupoestudiantes_set.filter(
                            status=True, pregunta=x).exists() else None
                        if respuesta is not None:
                            rango = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(pk=int(respuesta.opcionrango.id))
                            ws.write(row_num, c, rango.descripcion, font_style2)
                        else:
                            respuesta_pregunta = dato.respuestapreguntaencuestagrupoestudiantes_set.filter(status=True, pregunta=x).first()
                            if respuesta_pregunta:
                                rango=RangoPreguntaEncuestaGrupoEstudiantes.objects.get(id=int(respuesta_pregunta.respuesta))
                                ws.write(row_num, c,rango.valor, font_style2)
                            else:
                                ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 5:
                        respuesta = dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(status=True,
                                                                                                pregunta=x).first() if dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(
                            status=True, pregunta=x).exists() else None
                        if respuesta is not None:
                            try:
                                int(respuesta.respuesta)
                                # if encuesta.id == 25:
                                if encuesta.tipoperfil == 1:
                                    consulta = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True,
                                                                                                       pregunta=x,
                                                                                                       pk=respuesta.opcioncuadricula.pk,
                                                                                                       tipoopcion=2).first()
                                else:
                                    consulta = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True,
                                                                                                       pregunta=x,
                                                                                                       valor=respuesta.respuesta,
                                                                                                       tipoopcion=2).first()

                                if consulta == None:
                                    resp = 'Sin contestar'
                                else:
                                    # if encuesta.id == 25:
                                    if encuesta.tipoperfil == 1:
                                        resp = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True,
                                                                                                       pregunta=x,
                                                                                                       pk=respuesta.opcioncuadricula.pk,
                                                                                                       tipoopcion=2).first().descripcion
                                    else:
                                        resp = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True,
                                                                                                       pregunta=x,
                                                                                                       valor=respuesta.respuesta,
                                                                                                       tipoopcion=2).first().descripcion
                                    colurl = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True,
                                                                                                     pregunta=x,
                                                                                                     valor=respuesta.respuesta,
                                                                                                     tipoopcion=2).first().oparchivo if OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(
                                        status=True, pregunta=x, valor=respuesta.respuesta,
                                        tipoopcion=2).exists() else None

                                    if colurl:
                                        ws.write(row_num, c + 1,
                                                 str('https://sga.unemi.edu.ec' + respuesta.archivo.url))
                            except ValueError:
                                resp = respuesta.respuesta
                                if encuesta.tipoperfil == 4:
                                    resp = 'Tengo dudas'
                                    sintomas = respuesta.respuesta
                                    ws.write(row_num, c + 1, sintomas)
                                if encuesta.id == 29:
                                    if encuesta.tipoperfil == 1:
                                        resp = 'Otros'
                                        sintomas = respuesta.respuesta
                                        ws.write(row_num, c + 1, sintomas)

                            ws.write(row_num, c, resp, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 6:
                        respuesta = dato.respuestamultipleencuestagrupoestudiantes_set.filter(status=True,
                                                                                              pregunta=x) if dato.respuestamultipleencuestagrupoestudiantes_set.values(
                            'id').filter(status=True, pregunta=x).exists() else None
                        if respuesta is not None:
                            w = 0
                            for rmult in respuesta:
                                if rmult.opcionmultiple.opcotros:
                                    ws.write(row_num + w, c, rmult.respuesta, font_style2)
                                else:
                                    ws.write(row_num + w, c, rmult.opcionmultiple.descripcion, font_style2)
                                # row_num += 1
                                w += 1
                            if limit < w and w > 0:
                                limit = w - 1
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    else:
                        respuesta = RespuestaPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                             inscripcionencuesta=dato).first()
                        if respuesta is not None:
                            if x.tipo == 1:
                                if not x.esta_vacia():
                                    ws.write(row_num, c, respuesta.respuestaporno if respuesta.respuestaporno else "",
                                             font_style2)
                            else:
                                ws.write(row_num, c, respuesta.respuesta, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                row_num += 1
            # wb.save(directory)
            wb.close()
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/encuestas/{}".format(MEDIA_URL, nombrearchivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Resultados de Encuesta',
                                    destinatario=pers, url="{}reportes/encuestas/{}".format(MEDIA_URL, nombrearchivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Resultado de Encuesta',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/encuestas/{}".format(MEDIA_URL, nombrearchivo),
                "mensaje": 'Los resultados de la encuesta han sido generados con exito'
            }, ttl=500)
            return response
        except Exception as ex:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error en el reporte'
                noti.cuerpo = 'Ha ocurrido un error en al generar el reporte {} - Error en la linea {}'.format(ex,
                                                                                                               sys.exc_info()[
                                                                                                                   -1].tb_lineno)
                noti.url = ""
                noti.error = True
                noti.save()
            else:
                noti = Notificacion(titulo='Error en el reporte',
                                    cuerpo='Ha ocurrido un error en al generar el reporte {} - Error en la linea {}'.format(
                                        ex, sys.exc_info()[-1].tb_lineno),
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False, error=True)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Error en el reporte",
                "body": "Ha ocurrido un error en al generar el reporte {} - Error en la linea {}".format(ex,
                                                                                                         sys.exc_info()[
                                                                                                             -1].tb_lineno),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "mensaje": "Ha ocurrido un error en al generar el reporte"
            }, ttl=500)
            pass

        #
        # except Exception as ex:
        #     print(ex)
        #     print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
        #     textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_horarios_y_aulas_background(threading.Thread):

    def __init__(self, request, notiid, periodo):
        self.request = request
        self.periodo = periodo
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'encuestas')
        request, notiid, periodo = self.request, self.notiid, self.periodo
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "REPORTE_HORARIOS_Y_AULAS{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'encuestas', nombre_archivo)
        try:
            # directorio = os.path.join(os.path.join(MEDIA_ROOT, 'reportes', 'encuestas'))
            # try:
            #     os.stat(directorio)
            # except:
            #     os.mkdir(directorio)
            # nombrearchivo = "REPORTE_HORARIOS_Y_AULAS" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"

            # __author__ = 'Unemi'
            style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                              num_format_str='#,##0.00')
            style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            style1 = easyxf(num_format_str='D-MMM-YY')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False

            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet("datos")

            ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=exp_xls_post_part_' + random.randint(1,
                                                                                                    10000).__str__() + '.xls'
            columns = [
                (u"AULA", 6000),
                (u"CAPACIDAD AULA", 6000),
                (u"CUPO MATRICULA", 6000),
                (u"MATRICULADOS", 6000),
                (u"INSCRITOS", 6000),
                (u"DÍAS", 6000),
                (u"INICIO HORARIO", 6000),
                (u"FIN HORARIO", 6000),
                (u"HORARIOS", 6000),
                (u"ASIGNATURA", 6000),
                (u"MODALIDAD", 6000),
                (u"IDMATERIA", 6000),
                (u"JORNADA", 6000),
                (u"NIVEL", 6000),
                (u"CARRERA", 6000),
                (u"PARALELO", 6000),
                (u"DOCENTE", 6000),
                (u"TIPO PROFESOR", 6000),
                (u"CATEGORIA", 6000),
                (u"DEDICACIÓN", 6000),
                (u"PRINCIPAL", 6000),
                (u"FACULTAD", 6000),
                (u"TIPO MATERIA", 6000),
                (u"ID HORARIO", 4000),
                (u"INICIO MATERIA", 4000),
                (u"FIN MATERIA", 4000),
                (u"TIPO HORARIO", 4000),
                (u"ES PRÁCTICA", 4000),
                # (u"SILABO", 4000),
                # (u"TIENE GUIA PRÁCTICA", 4000),
                # (u"ID MATERIA", 4000),

            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            cursor = connections['sga_select'].cursor()
            # lista_json = []
            # data = {}

            sql = f""" 
                                    SET statement_timeout = '36000s';
                                    SELECT al.nombre AS Aula, al.capacidad AS capacidad_aula, mat.cupo AS cupo_matriculas, 
                                     (
                                    SELECT COUNT(*)
                                    FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                                    WHERE mat1.estado_matricula in (2,3) AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id AND ni1.periodo_id=ni.periodo_id AND mas1.materia_id=mat.id) AS Matriculados,
                                     (
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                                        WHERE mat1.estado_matricula=1 AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id 
                                        AND ni1.periodo_id=ni.periodo_id AND mas1.materia_id=mat.id) AS inscritos, 
                                     (CASE cl.dia WHEN 1 THEN 'LUNES' WHEN 2 THEN 'MARTES' WHEN 3 THEN 'MIERCOLES' WHEN 4 THEN 'JUEVES' WHEN 5 THEN 'VIERNES' WHEN 6 THEN 'SABADO' WHEN 7 THEN 'DOMINGO' END) AS dia, 
                                     cl.inicio, cl.fin, (tu.comienza|| '  ' || tu.termina) AS Horario, asi.nombre AS Asignatura, 
                                     (select CASE WHEN mod.modalidad = 1 THEN 'PRESENCIAL' else 'VIRTUAL' END) as modalidad,
                                     mat.id AS idmateria,
                                     ni.paralelo AS jornada, 
                                     nmall.nombre AS nivel, 
                                     ca.nombre AS Carrera, mat.paralelo AS paralelo, per.apellido1 ||' '|| per.apellido2 ||' '|| per.nombres AS docente,
                                     tipop.nombre AS tipo_profesor, cat.nombre AS categorizacion,ded.nombre AS dedicacion,
                                     (CASE pm.principal WHEN TRUE THEN 'SI' ELSE 'NO' END) AS principal, 
                                     cor.nombre AS facultad, 
                                     (CASE asimall.practicas WHEN TRUE THEN 'SI' ELSE 'NO' END) AS tipomateria, 
                                     cl.id AS Id, 
                                     mat.inicio AS Inicio_materia, 
                                     mat.fin AS Fin_materia, 
                                     mat.id AS id_materia,
                                     cl.tipohorario AS cl_tipohorario,
                                     asimall.practicas AS asimall_practicas
                                     /*(CASE 
                                        (SELECT count(*) 
                                                FROM sga_silabo AS sga_s 
                                                WHERE sga_s.materia_id = pm.materia_id AND sga_s.status=True AND sga_s.codigoqr=True
                                            ) 
                                        WHEN 0 THEN
                                            'NO'
                                        ELSE 
                                            'SI'
                                        END
                                    ) AS silabo,
                                    (SELECT 
                                        count(sga_tp.*) 
                                    FROM sga_tareapracticasilabosemanal AS sga_tp
                                    INNER JOIN sga_silabosemanal AS sga_ss ON sga_tp.silabosemanal_id = sga_ss.id
                                    INNER JOIN sga_silabo AS sga_s ON sga_ss.silabo_id = sga_s.id
                                    INNER JOIN sga_materia AS sga_m ON sga_m.id = sga_s.materia_id
                                    WHERE 
                                        sga_s.status= TRUE 
                                        AND sga_s.codigoqr= TRUE 
                                        AND sga_m.id = mat.id 
                                        AND sga_tp.estado_id!=3 
                                        AND sga_tp.status=True
                                    ) AS trabajos_practicos*/
                                    FROM sga_clase cl 
            						INNER JOIN sga_aula al ON cl.aula_id=al.id
            						INNER JOIN sga_materia mat ON cl.materia_id=mat.id
            						INNER JOIN sga_nivel ni ON mat.nivel_id=ni.id
            						INNER JOIN sga_asignatura asi ON mat.asignatura_id=asi.id
            						INNER JOIN sga_asignaturamalla asimall ON mat.asignaturamalla_id = asimall.id
            						INNER JOIN sga_nivelmalla nmall ON asimall.nivelmalla_id = nmall.id 
                                    INNER JOIN sga_malla mall ON asimall.malla_id = mall.id 
                                    INNER JOIN sga_carrera ca ON mall.carrera_id=ca.id
                                    INNER JOIN sga_turno tu ON tu.id=cl.turno_id 
                                    INNER JOIN sga_profesormateria pm ON pm.materia_id=mat.id
                                    INNER JOIN sga_tipoprofesor tipop ON tipop.id=pm.tipoprofesor_id
                                    INNER JOIN sga_profesor pr ON pr.id=pm.profesor_id AND cl.profesor_id=pr.id
                                    INNER JOIN sga_tiempodedicaciondocente ded ON pr.dedicacion_id=ded.id
                                    INNER JOIN sga_categorizaciondocente cat ON pr.categoria_id=cat.id
                                    INNER JOIN sga_persona per ON per.id=pr.persona_id
                                    INNER JOIN sga_coordinacion_carrera cc ON cc.carrera_id=ca.id
                           			INNER JOIN sga_coordinacion cor ON cor.id=cc.coordinacion_id 
                           			LEFT JOIN sga_detalleasignaturamallamodalidad mod ON asimall.id = mod.asignaturamalla_id 
                                    WHERE cl.activo= TRUE AND ni.periodo_id={periodo.id}
                                """
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4

            for r in results:
                i = 0
                campo1 = r[0]
                campo2 = r[1]
                campo3 = r[2]
                campo4 = r[3]
                campo5 = r[4]
                campo6 = r[5]
                campo7 = r[6]
                campo8 = r[7]
                campo9 = r[8]
                campo10 = r[9]
                campo11 = r[10]
                campo12 = r[11]
                campo13 = r[12]
                campo14 = r[13]
                campo15 = r[14]
                campo16 = r[15]
                campo17 = r[16]
                campo18 = r[17]
                campo19 = r[18]
                campo20 = r[19]
                campo21 = r[20]
                campo22 = r[21]
                campo23 = r[22]
                campo24 = r[23]
                campo25 = r[24]
                campo26 = r[25]
                campo27 = dict(TIPOHORARIO)[r[27]]
                campo28 = 'SI' if r[28] else 'NO'
                # campo29 = r[29]
                # campo30 = r[30]
                ws.write(row_num, 0, u"%s" % campo1, font_style2)
                ws.write(row_num, 1, u"%s" % campo2, font_style2)
                ws.write(row_num, 2, u"%s" % campo3, font_style2)
                ws.write(row_num, 3, u"%s" % campo4, font_style2)
                ws.write(row_num, 4, u"%s" % campo5, font_style2)
                ws.write(row_num, 5, u"%s" % campo6, style1)
                ws.write(row_num, 6, u"%s" % campo7, style1)
                ws.write(row_num, 7, u"%s" % campo8, font_style2)
                ws.write(row_num, 8, u"%s" % campo9, font_style2)
                ws.write(row_num, 9, u"%s" % campo10, font_style2)
                ws.write(row_num, 10, u"%s" % campo11, font_style2)
                ws.write(row_num, 11, u"%s" % campo12, font_style2)
                ws.write(row_num, 12, u"%s" % campo13, font_style2)
                ws.write(row_num, 13, u"%s" % campo14, font_style2)
                ws.write(row_num, 14, u"%s" % campo15, font_style2)
                ws.write(row_num, 15, u"%s" % campo16, font_style2)
                ws.write(row_num, 16, u"%s" % campo17, font_style2)
                ws.write(row_num, 17, u"%s" % campo18, font_style2)
                ws.write(row_num, 18, u"%s" % campo19, style1)
                ws.write(row_num, 19, u"%s" % campo20, style1)
                ws.write(row_num, 20, u"%s" % campo21, font_style2)
                ws.write(row_num, 21, u"%s" % campo22, font_style2)
                ws.write(row_num, 22, u"%s" % campo23, font_style2)
                ws.write(row_num, 23, u"%s" % campo24, font_style2)
                ws.write(row_num, 24, u"%s" % campo25, font_style2)
                ws.write(row_num, 25, u"%s" % campo26, font_style2)
                ws.write(row_num, 26, u"%s" % campo27, font_style2)
                ws.write(row_num, 27, u"%s" % campo28, font_style2)
                # ws.write(row_num, 28, u"%s" % campo29, font_style2)
                # ws.write(row_num, 29, u"%s" % campo30, font_style2)
                row_num += 1

            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/encuestas/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Resultados de Encuesta',
                                    destinatario=pers, url="{}reportes/encuestas/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Resultado de reporte de horarios y aulas',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/encuestas/{}".format(MEDIA_URL, nombre_archivo),
                "mensaje": 'Los resultados del reporte de horarios y aulas han sido generados con exito'
            }, ttl=500)
            return response
        except Exception as ex:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error en el reporte'
                noti.cuerpo = 'Ha ocurrido un error en al generar el reporte {} - Error en la linea {}'.format(ex,
                                                                                                               sys.exc_info()[
                                                                                                                   -1].tb_lineno)
                noti.url = ""
                noti.error = True
                noti.save()
            else:
                noti = Notificacion(titulo='Error en el reporte',
                                    cuerpo='Ha ocurrido un error en al generar el reporte {} - Error en la linea {}'.format(
                                        ex, sys.exc_info()[-1].tb_lineno),
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False, error=True)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Error en el reporte",
                "body": "Ha ocurrido un error en al generar el reporte {} - Error en la linea {}".format(ex,
                                                                                                         sys.exc_info()[
                                                                                                             -1].tb_lineno),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "mensaje": "Ha ocurrido un error en al generar el reporte"
            }, ttl=500)
            pass

        #
        # except Exception as ex:
        #     print(ex)
        #     print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
        #     textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_recurso_aprendizaje_background(threading.Thread):

    def __init__(self, request, notiid, periodo, coordinacion, codcarrera, tipo, cadenaparcial):
        self.request = request
        self.periodo = periodo
        self.coordinacion = coordinacion
        self.codcarrera = codcarrera
        self.tipo = tipo
        self.cadenaparcial = cadenaparcial
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        listado = []
        data = {}
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'recursos')
        request, notiid, periodo, coordinacion, codcarrera, tipo, cadenaparcial = self.request, self.notiid, self.periodo, self.coordinacion, self.codcarrera, self.tipo, self.cadenaparcial
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "Resultado_recursos{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'recursos', nombre_archivo)
        try:
            if tipo == 'excel':
                __author__ = 'Unemi'
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=Practicas- ' + random.randint(1,
                                                                                                      10000).__str__() + '.xls'
                columns = [
                    (u"N", 2500),
                    (u"FACULTAD", 6000),
                    (u"CARRERA", 6000),
                    (u"CEDULA", 3000),
                    (u"DOCENTE", 10000),
                    (u"TIPO PROFESOR", 6000),
                    (u"CATEGORIA", 6000),
                    (u"DEDICACION", 6000),
                    (u"SECCION", 6000),
                    (u"MODALIDAD", 6000),
                    (u"ASIGNATURA", 6000),
                    (u"NIVEL", 6000),
                    (u"PARALELO", 6000),
                    (u"% DE CUMPLIMIENTO RECURSOS DE APRENDIZAJE", 6000),
                    (u"% DE ACTIVIDAD Preparar, elaborar, aplicar y calificar exámenes, trabajos y prácticas", 6000)
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    # set_column

            data['periodo'] = periodo
            data['coordinacion'] = coordinacion
            data['fechaactual'] = datetime.now().date()
            data['id_fini'] = fechaini = periodo.inicio
            data['id_fin'] = fechafin = periodo.fin
            data['listadolineamiento'] = periodo.lineamientorecursoperiodo_set.filter(tipoprofesor_id__in=[1, 14],
                                                                                      nivelacion=False,
                                                                                      status=True).order_by(
                'tipoprofesor_id')
            data['listadocomponente'] = listadocomponente = periodo.evaluacioncomponenteperiodo_set.filter(
                parcial__in=cadenaparcial, nivelacion=False, status=True).order_by('parcial', 'componente_id')
            responsableccordinacion = '-'
            if coordinacion.responsable_periodo(periodo):
                responsableccordinacion = coordinacion.responsable_periodo(periodo).persona
            data['responsableccordinacion'] = responsableccordinacion
            # SE EXCLUYEN LAS CARRERAS DE ENFERMERIA Y NUTRICION EN LOS NIVELES MAYOR O IGUALES A 7MO
            if int(codcarrera) > 0:
                listadoprofesormateria = ProfesorMateria.objects.filter(tipoprofesor_id__in=[1, 14],
                                                                        materia__nivel__periodo=periodo,
                                                                        materia__asignaturamalla__malla__carrera_id=codcarrera,
                                                                        materia__status=True,
                                                                        materia__asignaturamalla__validarequisitograduacion=False,
                                                                        status=True).exclude(
                    materia_id__in=Materia.objects.values_list('id').filter(
                        asignaturamalla__malla__carrera_id__in=[1, 3, 111],
                        asignaturamalla__nivelmalla_id__in=[7, 8, 9], nivel__periodo=periodo, status=True)).exclude(
                    materia__modeloevaluativo_id=27).order_by('materia__asignaturamalla__malla__carrera_id',
                                                              'materia__asignaturamalla__nivelmalla_id')
            else:
                listadoprofesormateria = ProfesorMateria.objects.filter(tipoprofesor_id__in=[1, 14],
                                                                        materia__nivel__periodo=periodo,
                                                                        materia__asignaturamalla__malla__carrera__coordinacion=coordinacion,
                                                                        materia__asignaturamalla__validarequisitograduacion=False,
                                                                        materia__status=True, status=True).exclude(
                    materia_id__in=Materia.objects.values_list('id').filter(
                        asignaturamalla__malla__carrera_id__in=[1, 3, 111],
                        asignaturamalla__nivelmalla_id__in=[7, 8, 9], nivel__periodo=periodo, status=True)).exclude(
                    materia__modeloevaluativo_id=27).order_by('materia__asignaturamalla__malla__carrera_id',
                                                              'materia__asignaturamalla__nivelmalla_id')
            usucreacion = None
            cuenta = 0
            to_promedio_to = 0.0
            row_num = 1
            for profesormateria in listadoprofesormateria:
                cuenta = cuenta + 1
                if Silabo.objects.values("id").filter(materia=profesormateria.materia, status=True).exists():
                    mate = Silabo.objects.filter(materia=profesormateria.materia, status=True)[0]
                    listadocompendios = []
                    listadocompendiosmooc = []
                    listadodiapositiva = []
                    listadodiapositivamooc = []
                    listadoguiadocente = []
                    listadoguiadocentemooc = []
                    listadoguiaestudiante = []
                    listadoguiaestudiantemooc = []
                    listadomateriales = []
                    listadomaterialesmooc = []
                    listadotareas = []
                    listadoforos = []
                    listadotest = []
                    totalacd1 = 0
                    totalape1 = 0
                    totalaa1 = 0
                    totalmoodleacd = 0
                    totalmoodleaa = 0
                    totalmoodleape = 0

                    unidades = DetalleSilaboSemanalTema.objects.values_list(
                        'temaunidadresultadoprogramaanalitico__unidadresultadoprogramaanalitico__orden',
                        'silabosemanal__fechafinciosemana').filter(silabosemanal__silabo=mate,
                                                                   silabosemanal__parcial__in=cadenaparcial,
                                                                   status=True).order_by(
                        'temaunidadresultadoprogramaanalitico__unidadresultadoprogramaanalitico__orden',
                        '-silabosemanal__fechafinciosemana').distinct(
                        'temaunidadresultadoprogramaanalitico__unidadresultadoprogramaanalitico__orden')
                    cantidadunidades = unidades.count()
                    totalrecursos = 0
                    totalaprobadas = 0
                    totalaprobadasmoodle = 0
                    totalesporcentaje = 0
                    listalineamiento = LineamientoRecursoPeriodo.objects.filter(periodo=periodo,
                                                                                tipoprofesor_id=profesormateria.tipoprofesor.id,
                                                                                nivelacion=False, status=True)

                    # para sacar ACD
                    # test,exposicion y taller

                    testacd = TestSilaboSemanal.objects.filter(silabosemanal__silabo=mate,
                                                               silabosemanal__parcial__in=cadenaparcial, status=True)
                    tareaacd = TareaSilaboSemanal.objects.filter(silabosemanal__silabo=mate,
                                                                 silabosemanal__parcial__in=cadenaparcial,
                                                                 actividad_id__in=[2, 3], status=True)
                    totalacd1 = testacd.count() + tareaacd.count()

                    testacdmoodle = testacd.filter(estado_id=4)
                    tareaacdmoodle = tareaacd.filter(estado_id=4)
                    totalmoodleacd = testacdmoodle.count() + tareaacdmoodle.count()

                    componenteacdparcial1 = listadocomponente.filter(parcial=1, componente_id=1, status=True)[0]
                    testacdmoodlecom = testacd.filter(silabosemanal__parcial=1, estado_id=4).count()
                    tareaacdmoodlecom = tareaacd.filter(silabosemanal__parcial=1, estado_id=4).count()
                    totalmoodleacdcom = testacdmoodlecom + tareaacdmoodlecom
                    totalacdparciales = 0
                    if totalmoodleacdcom >= componenteacdparcial1.cantidad:
                        totalacdparcial1 = componenteacdparcial1.cantidad
                    else:
                        totalacdparcial1 = totalmoodleacdcom

                    if '2' in cadenaparcial:
                        componenteacdparcial2 = listadocomponente.filter(parcial=2, componente_id=1, status=True)[0]
                        testacdmoodlecom = testacd.filter(silabosemanal__parcial=2, estado_id=4).count()
                        tareaacdmoodlecom = tareaacd.filter(silabosemanal__parcial=2, estado_id=4).count()
                        totalmoodleacdcom = testacdmoodlecom + tareaacdmoodlecom
                        if totalmoodleacdcom >= componenteacdparcial2.cantidad:
                            totalacdparcial2 = componenteacdparcial2.cantidad
                        else:
                            totalacdparcial2 = totalmoodleacdcom
                    else:
                        totalacdparcial2 = 0
                    totalacdparciales = totalacdparcial1 + totalacdparcial2

                    # para sacar APE
                    # solo materias que son de practica
                    totalapeparciales = 0
                    if mate.materia.asignaturamalla.horasapeasistotal > 0:
                        totaleslineamientos = listadocomponente.count()
                        totalapeprimer = TareaPracticaSilaboSemanal.objects.filter(silabosemanal__silabo=mate,
                                                                                   silabosemanal__parcial__in=cadenaparcial,
                                                                                   status=True)
                        totalape1 = totalapeprimer.count()
                        totalape1moodle = totalapeprimer.filter(estado_id=4)
                        totalmoodleape = totalape1moodle.count()

                        componenteapeparcial1 = listadocomponente.filter(parcial=1, componente_id=2, status=True)[0]
                        totalmoodlapedcom = totalapeprimer.filter(silabosemanal__parcial=1, estado_id=4).count()

                        if totalmoodlapedcom >= componenteapeparcial1.cantidad:
                            totalapeparcial1 = componenteapeparcial1.cantidad
                        else:
                            totalapeparcial1 = totalmoodlapedcom

                        if '2' in cadenaparcial:
                            componenteapeparcial2 = listadocomponente.filter(parcial=2, componente_id=2, status=True)[0]
                            totalmoodlapedcom = totalapeprimer.filter(silabosemanal__parcial=2, estado_id=4).count()
                            if totalmoodlapedcom >= componenteapeparcial2.cantidad:
                                totalapeparcial2 = componenteapeparcial2.cantidad
                            else:
                                totalapeparcial2 = totalmoodlapedcom
                        else:
                            totalapeparcial2 = 0
                        totalapeparciales = totalapeparcial1 + totalapeparcial2
                    else:
                        totaleslineamientos = listadocomponente.filter(componente_id__in=[1, 3]).count()
                        totalape1 = '-'
                        totalmoodleape = '-'

                    # para sacar AA
                    # tareas,foros,trabajo de invesrtigacion,analisis de caso
                    foroaa = ForoSilaboSemanal.objects.filter(silabosemanal__silabo=mate,
                                                              silabosemanal__parcial__in=cadenaparcial, status=True)
                    tareaaa = TareaSilaboSemanal.objects.filter(silabosemanal__silabo=mate,
                                                                silabosemanal__parcial__in=cadenaparcial,
                                                                actividad_id__in=[5, 7, 8], status=True)
                    totalaa1 = foroaa.count() + tareaaa.count()

                    foroaamoodle = foroaa.filter(estado_id=4)
                    tareaaadmoodle = tareaaa.filter(estado_id=4)
                    totalmoodleaa = foroaamoodle.count() + tareaaadmoodle.count()

                    componenteaaparcial1 = listadocomponente.filter(parcial=1, componente_id=3, status=True)[0]
                    foroaamoodlecom = foroaa.filter(silabosemanal__parcial=1, estado_id=4).count()
                    tareaaamoodlecom = tareaaa.filter(silabosemanal__parcial=1, estado_id=4).count()
                    totalmoodleacdcom = foroaamoodlecom + tareaaamoodlecom
                    totalaaparciales = 0

                    if totalmoodleacdcom >= componenteaaparcial1.cantidad:
                        totalaaparcial1 = componenteaaparcial1.cantidad
                    else:
                        totalaaparcial1 = totalmoodleacdcom

                    if '2' in cadenaparcial:
                        componenteaaparcial2 = listadocomponente.filter(parcial=2, componente_id=3, status=True)[0]
                        foroaamoodlecom = foroaa.filter(silabosemanal__parcial=2, estado_id=4).count()
                        tareaaamoodlecom = tareaaa.filter(silabosemanal__parcial=2, estado_id=4).count()
                        totalmoodleacdcom = foroaamoodlecom + tareaaamoodlecom
                        if totalmoodleacdcom >= componenteaaparcial2.cantidad:
                            totalaaparcial2 = componenteaaparcial2.cantidad
                        else:
                            totalaaparcial2 = totalmoodleacdcom
                    else:
                        totalaaparcial2 = 0
                    totalaaparciales = totalaaparcial1 + totalaaparcial2

                    totallineamientostipodocente = 0
                    for linea in listalineamiento:
                        if linea.tiporecurso == 1:
                            totallineamientostipodocente = totallineamientostipodocente + (
                                    linea.cantidad * cantidadunidades)
                            totalamoodle = 0
                            listacompendio = CompendioSilaboSemanal.objects.values_list('id', 'fecha_creacion',
                                                                                        'silabosemanal_id',
                                                                                        'estado_id').filter(
                                fecha_creacion__lte=fechafin, silabosemanal__examen=False,
                                silabosemanal__parcial__in=cadenaparcial, silabosemanal__silabo__materia=mate.materia,
                                status=True)
                            for reccompendio in listacompendio:
                                if reccompendio[3] == 4:
                                    totalamoodle = totalamoodle + 1
                                unicompendio = DetalleSilaboSemanalTema.objects.values_list(
                                    'temaunidadresultadoprogramaanalitico__unidadresultadoprogramaanalitico__orden',
                                    flat=True).filter(silabosemanal__parcial__in=cadenaparcial,
                                                      silabosemanal_id=reccompendio[2])[0]
                                for comunidades in unidades:
                                    if unicompendio == comunidades[0] and reccompendio[1].date() > comunidades[1]:
                                        listadocompendios.append(reccompendio[0])
                                        if reccompendio[3] == 4:
                                            listadocompendiosmooc.append(reccompendio[0])
                            totalescompendio = listacompendio.count() - len(listadocompendios)
                            if totalescompendio >= (linea.cantidad * cantidadunidades):
                                totalaprobadas = totalaprobadas + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadas = totalaprobadas + totalescompendio
                            totalescompendiomoodle = totalamoodle - len(listadocompendiosmooc)
                            if totalescompendiomoodle >= (linea.cantidad * cantidadunidades):
                                totalaprobadasmoodle = totalaprobadasmoodle + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadasmoodle = totalaprobadasmoodle + totalescompendiomoodle

                        if linea.tiporecurso == 2:
                            totallineamientostipodocente = totallineamientostipodocente + (
                                    linea.cantidad * cantidadunidades)
                            totalamoodle = 0
                            listapresentacion = DiapositivaSilaboSemanal.objects.values_list('id', 'fecha_creacion',
                                                                                             'silabosemanal_id',
                                                                                             'estado_id').filter(
                                fecha_creacion__lte=fechafin, silabosemanal__parcial__in=cadenaparcial,
                                silabosemanal__examen=False, silabosemanal__silabo__materia=mate.materia, status=True)
                            for recpresentacion in listapresentacion:
                                if recpresentacion[3] == 4:
                                    totalamoodle = totalamoodle + 1
                                if DetalleSilaboSemanalTema.objects.values('id').filter(
                                        silabosemanal__parcial__in=cadenaparcial,
                                        silabosemanal_id=recpresentacion[2]).exists():
                                    unipresentacion = DetalleSilaboSemanalTema.objects.values_list(
                                        'temaunidadresultadoprogramaanalitico__unidadresultadoprogramaanalitico__orden',
                                        flat=True).filter(silabosemanal__parcial__in=cadenaparcial,
                                                          silabosemanal_id=recpresentacion[2])[0]
                                    for diaunidades in unidades:
                                        if unipresentacion == diaunidades[0] and recpresentacion[1].date() > \
                                                diaunidades[1]:
                                            listadodiapositiva.append(recpresentacion[0])
                                            if recpresentacion[3] == 4:
                                                listadodiapositivamooc.append(recpresentacion[0])
                            totalesdiapositiva = listapresentacion.count() - len(listadodiapositiva)
                            if totalesdiapositiva >= (linea.cantidad * cantidadunidades):
                                totalaprobadas = totalaprobadas + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadas = totalaprobadas + totalesdiapositiva

                            totalesdiapositivamoodle = totalamoodle - len(listadodiapositivamooc)
                            if totalesdiapositivamoodle >= (linea.cantidad * cantidadunidades):
                                totalaprobadasmoodle = totalaprobadasmoodle + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadasmoodle = totalaprobadasmoodle + totalesdiapositivamoodle

                        if linea.tiporecurso == 4:
                            totallineamientostipodocente = totallineamientostipodocente + (
                                    linea.cantidad * cantidadunidades)
                            totalamoodle = 0
                            listaguiaestudiante = GuiaEstudianteSilaboSemanal.objects.values_list('id',
                                                                                                  'fecha_creacion',
                                                                                                  'silabosemanal_id',
                                                                                                  'estado_id').filter(
                                fecha_creacion__lte=fechafin, silabosemanal__parcial__in=cadenaparcial,
                                silabosemanal__examen=False, silabosemanal__silabo__materia=mate.materia, status=True)
                            for recestudiante in listaguiaestudiante:
                                if recestudiante[3] == 4:
                                    totalamoodle = totalamoodle + 1
                                uniguiaestudiante = DetalleSilaboSemanalTema.objects.values_list(
                                    'temaunidadresultadoprogramaanalitico__unidadresultadoprogramaanalitico__orden',
                                    flat=True).filter(silabosemanal__parcial__in=cadenaparcial,
                                                      silabosemanal_id=recestudiante[2])[0]
                                for estunidades in unidades:
                                    if uniguiaestudiante == estunidades[0] and recestudiante[1].date() > estunidades[1]:
                                        listadoguiaestudiante.append(recestudiante[0])
                                        if recestudiante[3] == 4:
                                            listadoguiaestudiantemooc.append(recestudiante[0])
                            totalesguiaestudiante = listaguiaestudiante.count() - len(listadoguiaestudiante)

                            if totalesguiaestudiante >= (linea.cantidad * cantidadunidades):
                                totalaprobadas = totalaprobadas + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadas = totalaprobadas + totalesguiaestudiante

                            totalesguiaestudiantemoodle = totalamoodle - len(listadoguiaestudiantemooc)
                            if totalesguiaestudiantemoodle >= (linea.cantidad * cantidadunidades):
                                totalaprobadasmoodle = totalaprobadasmoodle + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadasmoodle = totalaprobadasmoodle + totalesguiaestudiantemoodle

                        if linea.tiporecurso == 11:
                            totallineamientostipodocente = totallineamientostipodocente + (
                                    linea.cantidad * cantidadunidades)
                            totalamoodle = 0
                            consultamateriales = MaterialAdicionalSilaboSemanal.objects.values_list('id',
                                                                                                    'fecha_creacion',
                                                                                                    'silabosemanal_id',
                                                                                                    'estado_id').filter(
                                fecha_creacion__lte=fechafin, silabosemanal__parcial__in=cadenaparcial,
                                silabosemanal__silabo__materia=mate.materia, silabosemanal__examen=False, status=True)
                            for recmateriales in consultamateriales:
                                if recmateriales[3] == 4:
                                    totalamoodle = totalamoodle + 1
                                unimateriales = DetalleSilaboSemanalTema.objects.values_list(
                                    'temaunidadresultadoprogramaanalitico__unidadresultadoprogramaanalitico__orden',
                                    flat=True).filter(silabosemanal__parcial__in=cadenaparcial,
                                                      silabosemanal_id=recmateriales[2])[0]
                                for materialunidades in unidades:
                                    if unimateriales == materialunidades[0] and recmateriales[1].date() > \
                                            materialunidades[1]:
                                        listadomateriales.append(recmateriales[0])
                                        if recmateriales[3] == 4:
                                            listadomaterialesmooc.append(recmateriales[0])
                            totalesmateriales = consultamateriales.count() - len(listadomateriales)
                            if totalesmateriales >= (linea.cantidad * cantidadunidades):
                                totalaprobadas = totalaprobadas + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadas = totalaprobadas + totalesmateriales

                            # totalesmoodle
                            totalesmaterialesmoodle = totalamoodle - len(listadomaterialesmooc)
                            if totalesmaterialesmoodle >= (linea.cantidad * cantidadunidades):
                                totalaprobadasmoodle = totalaprobadasmoodle + (linea.cantidad * cantidadunidades)
                            else:
                                totalaprobadasmoodle = totalaprobadasmoodle + totalesmaterialesmoodle

                    totalesrequisitos = totaleslineamientos + totallineamientostipodocente
                    totalesrequisitosaprobados = totalacdparciales + totalapeparciales + totalaaparciales + totalaprobadasmoodle

                    porcentajerequisitos = round((totalesrequisitosaprobados / totalesrequisitos) * 100, 2)

                    suma = listalineamiento.aggregate(valor=Sum('cantidad'))
                    if cantidadunidades > 0:
                        if suma['valor']:
                            totalrecursos = suma['valor'] * cantidadunidades
                            totalesporcentaje = round((totalaprobadas / totalrecursos) * 100, 2)
                        else:
                            totalrecursos = 0
                            totalesporcentaje = 0
                    else:
                        totalrecursos = 0
                    totalesfaltantes = len(listadocompendios) + len(listadodiapositiva) + len(listadoguiadocente) + \
                                       len(listadoguiaestudiante) + len(listadotareas) + len(listadoforos) + len(
                        listadotest)
                    # to_promedio_to += totalesporcentaje
                    to_promedio_to += porcentajerequisitos
                    numeroparcial = 2
                    listado.append([profesormateria, cantidadunidades, totalrecursos, totalaprobadas, totalesporcentaje,
                                    mate.id, totalesfaltantes, totalacd1, totalape1, totalaa1, totalmoodleacd,
                                    totalmoodleaa,
                                    totalmoodleape, totalaprobadasmoodle, porcentajerequisitos, int(totalacdparciales),
                                    totalapeparciales, int(totalaaparciales), int(numeroparcial)])
                    print(str(cuenta) + ' de ' + str(listadoprofesormateria.count()))
                    if tipo == 'excel':
                        distributivo = profesormateria.profesor.distributivohoras(periodo)
                        promedioactividadcalificar = 0
                        if distributivo.detalledistributivo_set.filter(criteriodocenciaperiodo__criterio_id=121,
                                                                       status=True):
                            detalle = \
                                distributivo.detalledistributivo_set.filter(criteriodocenciaperiodo__criterio_id=121,
                                                                            status=True)[0]
                            listadocadena = detalle.criteriodocenciaperiodo.horario_evidencia_moodlexmateria(
                                distributivo.profesor, distributivo.periodo.inicio, distributivo.periodo.fin,
                                profesormateria.materia)
                            listadocadena = listadocadena[-1]
                            if listadocadena[11] == 4:
                                promedioactividadcalificar = 'NO HAY ACTIVIDADES PLANIFICADAS'
                            else:
                                promedioactividadcalificar = listadocadena[10]
                            print(listadocadena[11])
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, coordinacion.nombre, font_style2)
                        ws.write(row_num, 2, str(profesormateria.materia.asignaturamalla.malla.carrera.nombre),
                                 font_style2)
                        ws.write(row_num, 3, profesormateria.profesor.persona.cedula, font_style2)
                        ws.write(row_num, 4, str(profesormateria.profesor.persona.nombre_completo_inverso()),
                                 font_style2)
                        ws.write(row_num, 5, profesormateria.tipoprofesor.nombre, font_style2)
                        ws.write(row_num, 6, distributivo.nivelcategoria.nombre if distributivo.nivelcategoria else "",
                                 font_style2)
                        ws.write(row_num, 7, distributivo.dedicacion.nombre, font_style2)
                        ws.write(row_num, 8, profesormateria.materia.nivel.sesion.nombre, font_style2)
                        ws.write(row_num, 9, profesormateria.materia.nivel.modalidad.nombre, font_style2)
                        ws.write(row_num, 10, profesormateria.materia.asignaturamalla.asignatura.nombre, font_style2)
                        ws.write(row_num, 11, profesormateria.materia.asignaturamalla.nivelmalla.nombre, font_style2)
                        ws.write(row_num, 12, profesormateria.materia.paralelo, font_style2)
                        ws.write(row_num, 13, porcentajerequisitos, font_style2)
                        ws.write(row_num, 14, promedioactividadcalificar, font_style2)
                        row_num += 1

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if tipo == 'excel':
                wb.save(directory)
                if notiid > 0:
                    noti = Notificacion.objects.get(pk=notiid)
                    noti.en_proceso = False
                    noti.cuerpo = 'Excel Listo'
                    noti.url = "{}reportes/recursos/{}".format(MEDIA_URL, nombre_archivo)
                    noti.save()
                else:
                    noti = Notificacion(cuerpo='Reporte Listo', titulo='mmm',
                                        destinatario=pers,
                                        url="{}reportes/recursos/{}".format(MEDIA_URL, nombre_archivo), prioridad=1,
                                        app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                        tipo=2, en_proceso=False)
                    noti.save(request)

                send_user_notification(user=usernotify, payload={
                    "head": "Excel terminado",
                    "body": 'mm',
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": "{}reportes/recursos/{}".format(MEDIA_URL, nombre_archivo),
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            else:
                nombrepersona = remover_caracteres_tildes_unicode(
                    remover_caracteres_especiales_unicode((pers.__str__()).replace(' ', '_')))
                aleatorio = random.randint(1, 100000)
                nombredocumento = 'CONSOLIDADO_{}_{}.pdf'.format(nombrepersona, aleatorio.__str__())
                nombredocumento1 = 'CONSOLIDADO_{}_{}'.format(nombrepersona, aleatorio.__str__())
                valida = conviert_html_to_pdf_name_save(
                    'aprobar_silabo/reporterecursoscomponente_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                        'listado': listado,
                        'total_promedio': (to_promedio_to / len(listado)),
                    }, nombredocumento1
                )
                if notiid > 0:
                    noti = Notificacion.objects.get(pk=notiid)
                    noti.en_proceso = False
                    noti.cuerpo = 'Reporte Listo'
                    noti.url = "{}qrcode/solicitudempresas/{}".format(MEDIA_URL, nombredocumento)
                    noti.save()
                else:
                    noti = Notificacion(cuerpo='Reporte Listo', titulo='Resultados de Encuesta',
                                        destinatario=pers,
                                        url="{}qrcode/solicitudempresas/{}".format(MEDIA_URL, nombredocumento),
                                        prioridad=1, app_label='SGA',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False)
                    noti.save(request)
                send_user_notification(user=usernotify, payload={
                    "head": "Reporte terminado",
                    "body": 'Reporte consolidado de cumplimiento de actividades y recursos de aprendizaje',
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": "{}qrcode/solicitudempresas/{}".format(MEDIA_URL, nombredocumento),
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Los resultados del cumplimineto de actividades han sido generados con exito'
                }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_cumplimiento_background(threading.Thread):

    def __init__(self, request, notiid, periodo, coordinacion, codcarrera, tipo, finicio, ffin):
        hoy = datetime.now()
        pastmonth = date(hoy.year, hoy.month, 1) - timedelta(days=1)
        fdaypastmonth = date(pastmonth.year, pastmonth.month, 1)
        self.request = request
        self.periodo = periodo
        self.coordinacion = coordinacion
        self.codcarrera = codcarrera
        self.tipo = tipo
        self.notiid = notiid
        self.finicio = finicio if finicio else fdaypastmonth
        self.ffin = ffin if ffin else pastmonth
        diasnolaborales = periodo.diasnolaborable_set.filter(status=True, activo=True).values_list('fecha', flat=True)
        self.diasnolaborales = []
        for dia in diasnolaborales:
            self.diasnolaborales.append(dia.strftime('%Y-%m-%d'))
        threading.Thread.__init__(self)

    def run(self):
        cuenta = 0
        listado = []
        data = {}
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'recursos')
        request, notiid, periodo, coordinacion, codcarrera, tipo, fechaini, fechafin, diasnolaborales = self.request, self.notiid, self.periodo, self.coordinacion, self.codcarrera, self.tipo, self.finicio, self.ffin, self.diasnolaborales
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        # nombre_archivo = "Resultado_recursos{}.xls".format(random.randint(1, 10000).__str__())
        # directory = os.path.join(MEDIA_ROOT, 'reportes', 'recursos', nombre_archivo)
        try:
            data['periodo'] = periodo
            data['coordinacion'] = coordinacion
            data['fechaactual'] = datetime.now().date()
            data['fechaini'] = fechaini
            data['fechafin'] = fechafin
            responsableccordinacion = '-'
            if coordinacion.responsable_periodo(periodo):
                responsableccordinacion = coordinacion.responsable_periodo(periodo).persona
            data['responsableccordinacion'] = responsableccordinacion
            if int(codcarrera) > 0:
                data['carrera'] = Carrera.objects.get(id=codcarrera)
                listadoprofesormateria = ProfesorMateria.objects.values_list('profesor_id').filter(
                    tipoprofesor_id__in=[1, 14], materia__nivel__periodo=periodo,
                    materia__asignaturamalla__malla__carrera_id=codcarrera, materia__status=True, status=True).exclude(
                    materia_id__in=Materia.objects.values_list('id').filter(
                        asignaturamalla__malla__carrera_id__in=[1, 3], asignaturamalla__nivelmalla_id__in=[7, 8, 9],
                        nivel__periodo=periodo, status=True)).order_by(
                    'materia__asignaturamalla__malla__carrera_id', 'materia__asignaturamalla__nivelmalla_id')
            else:
                listadoprofesormateria = ProfesorMateria.objects.values_list('profesor_id').filter(
                    tipoprofesor_id__in=[1, 14], materia__nivel__periodo=periodo, materia__status=True,
                    status=True).exclude(materia_id__in=Materia.objects.values_list('id').filter(
                    asignaturamalla__malla__carrera_id__in=[1, 3], asignaturamalla__nivelmalla_id__in=[7, 8, 9],
                    nivel__periodo=periodo, status=True)).order_by('materia__asignaturamalla__malla__carrera_id',
                                                                   'materia__asignaturamalla__nivelmalla_id')

            listadodistributivo = ProfesorDistributivoHoras.objects.filter(profesor_id__in=listadoprofesormateria,
                                                                           periodo=periodo, status=True,
                                                                           coordinacion=coordinacion).order_by(
                'profesor__persona__apellido1', 'profesor__persona__apellido2', 'profesor__persona__nombres')
            usucreacion = None

            to_promedio_to = 0.0
            row_num = 1
            for distri in listadodistributivo:
                if distri.profesor.id:
                    totcriterios = 0
                    totcriteriosgestion = 0
                    totcriteriosvinculacion = 0
                    totcriteriosinvestigacion = 0
                    asignaturas = ProfesorMateria.objects.filter(Q(profesor=distri.profesor),
                                                                 ((Q(desde__gte=fechaini) & Q(hasta__lte=fechafin)) |
                                                                  (Q(desde__lte=fechaini) & Q(hasta__gte=fechafin)) |
                                                                  (Q(desde__lte=fechafin) & Q(desde__gte=fechaini)) |
                                                                  (Q(hasta__gte=fechaini) & Q(
                                                                      hasta__lte=fechafin)))).distinct().order_by(
                        'desde',
                        'materia__asignatura__nombre')
                    totdocencia = [0, 0, 0]
                    totinvestigacion = [0, 0, 0]
                    totvinculacion = [0, 0, 0]
                    totgestion = [0, 0, 0]
                    actividadesdocencia = distri.detalle_horas_docencia(fechaini, fechafin)

                    for detalle in actividadesdocencia:
                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'impartirclase':
                            impartir = detalle.criteriodocenciaperiodo.totalimparticlase(distri.profesor, fechaini,
                                                                                         fechafin, asignaturas,
                                                                                         detalle.criteriodocenciaperiodo.periodo)
                            if impartir:
                                totcriterios += 1
                                totales = impartir[0]
                                totdocencia[0] += totales[0]
                                totdocencia[1] += totales[1] if totales[1] <= totales[0] else totales[0]
                                try:
                                    procentaje = round((totales[1] * 100) / totales[0], 2)
                                    totdocencia[2] += procentaje if procentaje <= 100 else 100
                                except ZeroDivisionError:
                                    totdocencia[2] += 0
                        # AGREGADO POR ERICK
                        # if detalle.criteriodocenciaperiodo.nombrehtmldocente == 'actividaddocente':
                        #     actividad = detalle.criteriodocenciaperiodo.horarios_actividaddocente_profesor(
                        #         distri.profesor, fechaini, fechafin)
                        #
                        #     if actividad:
                        #         totcriterios += 1
                        #         totdocencia[0] += actividad.totalplanificadas
                        #         totdocencia[1] += actividad.totalplanificadas
                        #         try:
                        #             procentaje = round(
                        #                 (actividad.totalplanificadas * 100) / actividad.totalplanificadas, 2)
                        #             totdocencia[2] += procentaje if procentaje <= 100 else 100
                        #         except ZeroDivisionError:
                        #             totdocencia[2] += 0

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'dirigirtitulogrado':
                            dirigir = detalle.criteriodocenciaperiodo.actividad_dirigir_titulaciongrado(distri.profesor,
                                                                                                        fechaini,
                                                                                                        fechafin)
                            if dirigir:
                                totcriterios += 1
                                totdocencia[0] += dirigir.totalactividadplanificada
                                totdocencia[
                                    1] += dirigir.totalejecutadas if dirigir.totalejecutadas <= dirigir.totalactividadplanificada else dirigir.totalactividadplanificada
                                try:
                                    procentaje = round(
                                        (dirigir.totalejecutadas * 100) / dirigir.totalactividadplanificada, 2)
                                    totdocencia[2] += procentaje if procentaje <= 100 else 100
                                except ZeroDivisionError:
                                    totdocencia[2] += 0

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'evidenciamoodle':
                            evidenciamoodle = detalle.criteriodocenciaperiodo.horario_evidencia_moodle(distri.profesor,
                                                                                                       fechaini,
                                                                                                       fechafin)
                            if evidenciamoodle:
                                totcriterios += 1
                                fstart = fechaini
                                planificadas = 0
                                while fstart <= fechafin:
                                    for turnodia in evidenciamoodle[0][0]:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                if evidenciamoodle[-1][11] == 4:
                                    totdocencia[2] += 100
                                    totdocencia[1] += planificadas

                                else:
                                    totdocencia[2] += evidenciamoodle[-1][10]
                                    totdocencia[1] += round((evidenciamoodle[-1][10] * planificadas) / 100, 2)

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'materialsilabo':
                            materialsilabo = detalle.criteriodocenciaperiodo.horarios_actividad_profesor(
                                distri.profesor, fechaini, fechafin)
                            if materialsilabo:
                                totcriterios += 1
                                planificadas = 0
                                fstart = fechaini
                                while fstart <= fechafin:
                                    for turnodia in materialsilabo[0][0]:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                porcentaje = round(materialsilabo[-1][3], 2) if round(materialsilabo[-1][3],
                                                                                      2) <= 100 else 100
                                totdocencia[1] += round((porcentaje * planificadas) / 100, 2)

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'cursonivelacion':
                            cursonivelacion = detalle.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(
                                distri.profesor, fechaini, fechafin)
                            if cursonivelacion:
                                totcriterios += 1
                                fstart = fechaini
                                planificadas = 0
                                # totdocencia[0] += planificar[-1][1]
                                # totdocencia[1] += planificar[-1][2] if planificar[-1][2] <= planificar[-1][1] else planificar[-1][1]
                                while fstart <= fechafin:
                                    for turnodia in cursonivelacion.claseactividad:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                porcentaje = 100 if cursonivelacion.totalactividades >= 1 else 0
                                totdocencia[1] += round((porcentaje * planificadas) / 100,
                                                        2) if porcentaje >= 1 or porcentaje <= 100 else 0
                                totdocencia[2] += porcentaje if porcentaje <= 100 or porcentaje >= 1 else 100

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'planificarcontenido':
                            planificar = detalle.criteriodocenciaperiodo.horarios_contenido_profesor(distri.profesor,
                                                                                                     fechaini, fechafin)
                            if planificar:
                                totcriterios += 1
                                fstart = fechaini
                                planificadas = 0
                                # totdocencia[0] += planificar[-1][1]
                                # totdocencia[1] += planificar[-1][2] if planificar[-1][2] <= planificar[-1][1] else planificar[-1][1]
                                while fstart <= fechafin:
                                    for turnodia in planificar[0][0]:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                porcentaje = round(planificar[-1][3], 2) if round(planificar[-1][3], 2) <= 100 else 100
                                totdocencia[1] += round((porcentaje * planificadas) / 100, 2)
                                try:
                                    totdocencia[2] += porcentaje if porcentaje <= 100 else 100
                                except ZeroDivisionError:
                                    totdocencia[2] += 0

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'tutoriaacademica':
                            tutorias = detalle.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(
                                distri.profesor, fechaini, fechafin)
                            if tutorias:
                                totcriterios += 1
                                tut = tutorias[0]
                                totdocencia[0] += tut[1]
                                totdocencia[1] += tut[2] if tut[2] <= tut[1] else tut[1]
                                try:
                                    procentaje = round((tut[2] * 100) / tut[1], 2)
                                    totdocencia[2] += procentaje if procentaje <= 100 else 100
                                except ZeroDivisionError:
                                    totdocencia[2] += 0

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientoplataforma':
                            seguimientoplataforma = detalle.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(
                                distri.profesor, fechaini, fechafin)
                            if seguimientoplataforma:
                                totcriterios += 1
                                planificadas = 0
                                fstart = fechaini
                                # totdocencia[0] += planificar[-1][1]
                                # totdocencia[1] += planificar[-1][2] if planificar[-1][2] <= planificar[-1][1] else planificar[-1][1]
                                while fstart <= fechafin:
                                    for turnodia in seguimientoplataforma[0][10]:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                porcentaje = round(seguimientoplataforma[-1][9], 2) if round(
                                    seguimientoplataforma[-1][9], 2) <= 100 else 100
                                totdocencia[1] += round((porcentaje * planificadas) / 100, 2)

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'criterioperiodoadmision':
                            criterioperiodoadmision = detalle.criteriodocenciaperiodo.horario_criterio_nivelacion(
                                distri.profesor, fechaini, fechafin)
                            if criterioperiodoadmision:
                                totcriterios += 1
                                planificadas = 0
                                fstart = fechaini
                                while fstart <= fechafin:
                                    for turnodia in criterioperiodoadmision[0][0]:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                porcentaje = round(criterioperiodoadmision[0][4], 2) if round(
                                    criterioperiodoadmision[0][4], 2) <= 100 else 100
                                totdocencia[1] += round((porcentaje * planificadas) / 100, 2)

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'clasesnivelacion':
                            asignaturas = distri.profesor.asignaturas_periodos_relacionado(
                                detalle.criteriodocenciaperiodo.periodosrelacionados.all(), fechaini, fechafin)
                            if asignaturas:
                                impartir = detalle.criteriodocenciaperiodo.totalimparticlase(distri.profesor, fechaini,
                                                                                             fechafin, asignaturas,
                                                                                             detalle.criteriodocenciaperiodo.periodo)
                                if impartir:
                                    totcriterios += 1
                                    totales = impartir[0]
                                    totdocencia[0] += totales[0]
                                    totdocencia[1] += totales[1] if totales[1] <= totales[0] else totales[0]
                                    try:
                                        procentaje = round((totales[1] * 100) / totales[0], 2)
                                        totdocencia[2] += procentaje if procentaje <= 100 else 100
                                    except ZeroDivisionError:
                                        totdocencia[2] += 0

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'nivelacioncarrera':
                            actividadnivelacioncarrera = detalle.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(
                                distri.profesor, fechaini, fechafin)
                            if actividadnivelacioncarrera:
                                totcriterios += 1
                                planificadas = 0
                                fstart = fechaini
                                while fstart <= fechafin:
                                    for turnodia in actividadnivelacioncarrera.claseactividad:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                porcentaje = 100
                                totdocencia[1] += round((porcentaje * planificadas) / 100, 2)

                        if detalle.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientotransversal':
                            seguimientotransversal = detalle.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(
                                distri.profesor, fechaini, fechafin)
                            if seguimientotransversal:
                                totcriterios += 1
                                planificadas = 0
                                fstart = fechaini
                                while fstart <= fechafin:
                                    for turnodia in seguimientotransversal[0][10]:
                                        if turnodia.dia == fstart.weekday() + 1 and str(fstart) not in diasnolaborales:
                                            planificadas += 1
                                    fstart += timedelta(days=1)
                                totdocencia[0] += planificadas
                                porcentaje = round(seguimientotransversal[-1][9], 2) if round(
                                    seguimientotransversal[-1][9], 2) <= 100 else 100
                                totdocencia[1] += round((porcentaje * planificadas) / 100, 2)

                    for actividad in distri.detalle_horas_investigacion():
                        if actividad.criterioinvestigacionperiodo:
                            if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadinvestigacion':
                                actividadgestion = actividad.criterioinvestigacionperiodo.horarios_actividadinvestigacion_profesor(
                                    distri.profesor, fechaini, fechafin)
                                if actividadgestion:
                                    totcriteriosinvestigacion += 1
                                    fstart = fechaini
                                    planificadas = 0
                                    while fstart <= fechafin:
                                        for turnodia in actividadgestion[0][3]:
                                            if turnodia.dia == fstart.weekday() + 1 and str(
                                                    fstart) not in diasnolaborales:
                                                planificadas += 1
                                        fstart += timedelta(days=1)
                                    totinvestigacion[0] += planificadas
                                    totinvestigacion[2] += 100
                                    totinvestigacion[1] += round((100 * planificadas) / 100, 2)
                                # actividadgestion = actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(
                                #      distri, fechaini, fechafin)
                                # if actividadgestion:
                                #     totcriteriosinvestigacion += 1
                                #     fstart = fechaini
                                #     planificadas = 0
                                #     while fstart <= fechafin:
                                #         for turnodia in actividadgestion[0][3]:
                                #             if turnodia.dia == fstart.weekday() + 1 and str(
                                #                     fstart) not in diasnolaborales:
                                #                 planificadas += 1
                                #         fstart += timedelta(days=1)
                                #     totinvestigacion[0] += planificadas
                                #     totinvestigacion[2] += 100
                                #     totinvestigacion[1] += round((100 * planificadas) / 100, 2)

                    for actividad in distri.detalle_horas_vinculacion():
                        if actividad.criteriodocenciaperiodo:
                            if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadvinculacion':
                                actividadnivelacioncarrera = actividad.criteriodocenciaperiodo.horarios_actividadvinculacion_profesor(
                                    distri.profesor, fechaini, fechafin)
                                if actividadnivelacioncarrera:
                                    totcriteriosvinculacion += 1
                                    fstart = fechaini
                                    planificadas = 0
                                    while fstart <= fechafin:
                                        for turnodia in actividadnivelacioncarrera.claseactividad:
                                            if turnodia.dia == fstart.weekday() + 1 and str(
                                                    fstart) not in diasnolaborales:
                                                planificadas += 1
                                        fstart += timedelta(days=1)
                                    totvinculacion[0] += planificadas
                                    totvinculacion[2] += 100
                                    totvinculacion[1] += round((100 * planificadas) / 100, 2)

                    for actividad in distri.detalle_horas_gestion():
                        if actividad.criteriogestionperiodo:
                            if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadgestion':
                                actividadgestion = actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(
                                    distri.profesor, fechaini, fechafin)
                                if actividadgestion:
                                    totcriteriosgestion += 1
                                    fstart = fechaini
                                    planificadas = 0
                                    while fstart <= fechafin:
                                        for turnodia in actividadgestion:
                                            if turnodia.dia == fstart.weekday() + 1 and str(
                                                    fstart) not in diasnolaborales:
                                                planificadas += 1
                                        fstart += timedelta(days=1)
                                    totgestion[0] += planificadas
                                    totgestion[2] += 100
                                    totgestion[1] += round((100 * planificadas) / 100, 2)

                    # Para sumar y calcular los porcentajes docencia
                    try:
                        # totdocencia[2] = round(totdocencia[2]/totcriterios, 2)
                        totdocencia[2] = round((totdocencia[1] * 100) / totdocencia[0], 2)
                    except ZeroDivisionError:
                        totdocencia[2] = 0

                    # Para sumar y calcular los porcentajes gestion
                    try:
                        totinvestigacion[2] = round((totinvestigacion[1] * 100) / totinvestigacion[0], 2)
                    except ZeroDivisionError:
                        totinvestigacion[2] = 0

                        # Para sumar y calcular los porcentajes vinculacion
                    try:
                        totvinculacion[2] = round((totvinculacion[1] * 100) / totvinculacion[0], 2)
                    except ZeroDivisionError:
                        totvinculacion[2] = 0

                    try:
                        totgestion[2] = round((totgestion[1] * 100) / totgestion[0], 2)

                    except ZeroDivisionError:
                        totgestion[2] = 0

                    cuenta = cuenta + 1
                    totalgeneral = totdocencia[0] + totinvestigacion[0] + totvinculacion[0] + totgestion[0]
                    cumplidogeneral = totdocencia[1] + totinvestigacion[1] + totvinculacion[1] + totgestion[1]
                    porcgeneral = round((cumplidogeneral * 100) / totalgeneral, 2)
                    listado.append([distri, totdocencia[0], totdocencia[1], totdocencia[2], totinvestigacion[0],
                                    totinvestigacion[1], totinvestigacion[2], totvinculacion[0],
                                    totvinculacion[1], totvinculacion[2], totgestion[0],
                                    totgestion[1], totgestion[2],
                                    round(totalgeneral, 2),
                                    round(porcgeneral, 2),
                                    round(100 - porcgeneral, 2),
                                    round(cumplidogeneral, 2),
                                    round(totalgeneral - cumplidogeneral, 2)])
                    print(str(cuenta) + ' de ' + str(listadodistributivo.count()))

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)

            nombrepersona = remover_caracteres_tildes_unicode(
                remover_caracteres_especiales_unicode((pers.__str__()).replace(' ', '_')))
            aleatorio = random.randint(1, 100000)
            nombredocumento = 'cumplimiento_{}_{}.pdf'.format(nombrepersona, aleatorio.__str__())
            nombredocumento1 = 'cumplimiento_{}_{}'.format(nombrepersona, aleatorio.__str__())
            valida = conviert_html_to_pdfsaveqr_generico(request,
                                                         'aprobar_silabo/reporteinformemensual_pdf.html',
                                                         {
                                                             'pagesize': 'A4',
                                                             'data': data,
                                                             'listado': listado,
                                                             # 'total_promedio': (to_promedio_to / len(listado)),
                                                         }, directory, nombredocumento1 + '.pdf'
                                                         )
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}/{}/{}".format(MEDIA_URL, 'reportes/recursos/', nombredocumento1 + '.pdf')
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Resultados de Encuesta',
                                    destinatario=pers,
                                    url="{}/{}/{}".format(MEDIA_URL, 'reportes/recursos/', nombredocumento1 + '.pdf'),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Reporte consolidado de cumplimiento de actividades y recursos de aprendizaje',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}/{}/{}".format(MEDIA_URL, 'reportes/recursos/', nombredocumento1 + '.pdf').replace("\n",
                                                                                                             " ").replace(
                    '\\', "/").replace("//", "/"),
                "btn_notificaciones": traerNotificaciones(request, data, pers).replace("\n", " ").replace('\\',
                                                                                                          "/").replace(
                    "//", "/"),
                "mensaje": 'Los resultados del cumplimineto de actividades han sido generados con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print(cuenta)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

class reporte_cumplimiento_background_v2(threading.Thread):
    def __init__(self, request, notiid, periodo, coordinacion, codcarrera, tipo, finicio, ffin):
        hoy = datetime.now()
        pastmonth = date(hoy.year, hoy.month, 1) - timedelta(days=1)
        fdaypastmonth = date(pastmonth.year, pastmonth.month, 1)
        self.request = request
        self.periodo = periodo
        self.coordinacion = coordinacion
        self.codcarrera = codcarrera
        self.tipo = tipo
        self.notiid = notiid
        self.finicio = finicio if finicio else fdaypastmonth
        self.ffin = ffin if ffin else pastmonth
        diasnolaborales = periodo.diasnolaborable_set.filter(status=True, activo=True).values_list('fecha', flat=True)
        self.diasnolaborales = []
        for dia in diasnolaborales:
            self.diasnolaborales.append(dia.strftime('%Y-%m-%d'))
        threading.Thread.__init__(self)

    def run(self):
        from sga.templatetags.sga_extras import listado_bitacora_docente, listado_colectivos_academicos
        cuenta = 0
        listado = []
        data = {}
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'recursos')
        request, notiid, periodo, coordinacion, codcarrera, tipo, fechaini, fechafin, diasnolaborales = self.request, self.notiid, self.periodo, self.coordinacion, self.codcarrera, self.tipo, self.finicio, self.ffin, self.diasnolaborales
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        try:
            data['periodo'] = periodo
            data['coordinacion'] = coordinacion
            data['fechaactual'] = datetime.now().date()
            data['fechaini'] = fechaini
            data['fechafin'] = fechafin
            responsableccordinacion = '-'
            if coordinacion.responsable_periodo(periodo):
                responsableccordinacion = coordinacion.responsable_periodo(periodo).persona
            data['responsableccordinacion'] = responsableccordinacion
            if int(codcarrera) > 0:
                data['carrera'] = Carrera.objects.get(id=codcarrera)
                listadoprofesormateria = ProfesorMateria.objects.values_list('profesor_id').filter(tipoprofesor_id__in=[1, 14], materia__nivel__periodo=periodo,materia__asignaturamalla__malla__carrera_id=codcarrera, materia__status=True, status=True).exclude(materia_id__in=Materia.objects.values_list('id').filter(asignaturamalla__malla__carrera_id__in=[1, 3], asignaturamalla__nivelmalla_id__in=[7, 8, 9],nivel__periodo=periodo, status=True)).order_by('materia__asignaturamalla__malla__carrera_id', 'materia__asignaturamalla__nivelmalla_id')
            else:
                listadoprofesormateria = ProfesorMateria.objects.values_list('profesor_id').filter(tipoprofesor_id__in=[1, 14], materia__nivel__periodo=periodo, materia__status=True,status=True).exclude(materia_id__in=Materia.objects.values_list('id').filter(asignaturamalla__malla__carrera_id__in=[1, 3], asignaturamalla__nivelmalla_id__in=[7, 8, 9],nivel__periodo=periodo, status=True)).order_by('materia__asignaturamalla__malla__carrera_id','materia__asignaturamalla__nivelmalla_id')

            listadodistributivo = ProfesorDistributivoHoras.objects.filter(profesor_id__in=listadoprofesormateria,periodo=periodo, status=True,coordinacion=coordinacion).order_by('profesor__persona__apellido1', 'profesor__persona__apellido2', 'profesor__persona__nombres')
            usucreacion = None

            to_promedio_to = 0.0
            row_num = 1

            finiinicio, ffinal = fechaini, fechafin
            fechainiresta = fechaini - timedelta(days=5)
            fechafinresta = fechafin - timedelta(days=5)
            finicresta = fechainiresta
            ffincresta = fechafinresta

            for distri in listadodistributivo:
                count, count1, count2, count3, count4 = 0, 0, 0, 0, 0
                totalporcentaje, totalhdocentes, totalhinvestigacion, totalhgestion, totalhvinculacion = 0, 0, 0, 0, 0
                lista_criterios, distributivo = [], distri
                listDocencia = []

                if distri.profesor.id:
                    totcriterios = 0
                    totcriteriosgestion = 0
                    totcriteriosvinculacion = 0
                    totcriteriosinvestigacion = 0
                    asignaturas = ProfesorMateria.objects.filter(Q(profesor=distri.profesor),((Q(desde__gte=fechaini) & Q(hasta__lte=fechafin)) |(Q(desde__lte=fechaini) & Q(hasta__gte=fechafin)) |(Q(desde__lte=fechafin) & Q(desde__gte=fechaini)) |(Q(hasta__gte=fechaini) & Q(hasta__lte=fechafin)))).distinct().order_by('desde','materia__asignatura__nombre')

                    __doc, __inv, __ges, __vin = [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]
                    if horasdocencia := distributivo.detalle_horas_docencia(finiinicio, ffinal):
                        listDocencia.append([0, 'ACTIVIDADES DE DOCENCIA'])
                        for actividad in horasdocencia:
                            if html := actividad.criteriodocenciaperiodo.nombrehtmldocente():
                                __doc[0] += actividad.horas
                                if html == 'impartirclase':
                                    profesormateria = ProfesorMateria.objects.filter(profesor=distributivo.profesor,materia__nivel__periodo=periodo,tipoprofesor__imparteclase=True,activo=True, materia__status=True).distinct().order_by('desde', 'materia__asignatura__nombre')
                                    if periodo.clasificacion == 1:
                                        asignaturas = profesormateria.filter(((Q(desde__gte=finiinicio) & Q(hasta__lte=ffinal)) | (Q(desde__lte=finiinicio) & Q(hasta__gte=ffinal)) | (Q(desde__lte=finiinicio) & Q(desde__gte=ffinal)) | (Q(hasta__gte=finiinicio) & Q(hasta__lte=ffinal)))).distinct().exclude(tipoprofesor_id__in=[15, 5]).order_by('desde','materia__asignatura__nombre')
                                    else:
                                        asignaturas = profesormateria.filter(((Q(desde__gte=finiinicio) & Q(hasta__lte=ffinal)) | (Q(desde__lte=finiinicio) & Q(hasta__gte=ffinal)) | (Q(desde__lte=finiinicio) & Q(desde__gte=ffinal)) | (Q(hasta__gte=finiinicio) & Q(hasta__lte=ffinal)))).exclude(tipoprofesor_id__in=[5]).distinct().order_by('desde','materia__asignatura__nombre')
                                    totalimpartir = actividad.criteriodocenciaperiodo.totalimparticlase(distributivo.profesor, finiinicio, ffinal, asignaturas, None, True)
                                    if totalimpartir[2]:
                                        count += 1
                                        totalhdocentes += totalimpartir[1]
                                    porcentaje = round(totalimpartir[1], 2) if round(totalimpartir[1], 2) <= 100 else 100
                                    __doc[1] += totalimpartir[0]
                                    __doc[2] += porcentaje

                                if html == 'evidenciamoodle':
                                    if not DEBUG:
                                        if listadoevidencias := actividad.criteriodocenciaperiodo.horario_evidencia_moodle(
                                                distributivo.profesor, finicresta, ffincresta, True):
                                            if listadoevidencias[2]:
                                                count += 1
                                                totalhdocentes += listadoevidencias[1]
                                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoevidencias[0], listadoevidencias[1]])
                                            __doc[1] += listadoevidencias[0]
                                            __doc[2] += listadoevidencias[1]

                                if html == 'materialsilabo':
                                    actividadhor = actividad.criteriodocenciaperiodo.horarios_actividad_profesor(
                                        distributivo.profesor, finiinicio, ffinal, True)
                                    count += 1
                                    totalhdocentes += actividadhor[1]
                                    __doc[1] += actividadhor[0]
                                    __doc[2] += actividadhor[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadhor[0], actividadhor[1]])

                                if html == 'cursonivelacion':
                                    actividadnivelacioncarrera = actividad.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(
                                        distributivo.profesor, finiinicio, ffinal)
                                    totitem4 = 0
                                    if actividadnivelacioncarrera:
                                        totitem4 += 100
                                        totalhdocentes += 100
                                        count += 1
                                        # inicio, fin = fechaini, fechafin
                                        # if a := actividad.actividaddetalledistributivo_set.filter(vigente=True, status=True).first():
                                        #     inicio, fin = a.desde, a.hasta
                                        #
                                        # claseactividad = actividad.claseactividad_set.filter(detalledistributivo__distributivo__profesor=profesor, status=True).values_list('dia', 'turno_id').order_by('inicio', 'dia', 'turno__comienza')
                                        # dt, end, step = date(fechafin.year, fechafin.month, 1), fechafin, timedelta(days=1)
                                        # _result = []
                                        # while dt <= end:
                                        #     if inicio <= dt <= fin:
                                        #         if not periodo.diasnolaborable_set.values('id').filter(status=True, fecha=dt, activo=True):
                                        #             _result += [dt.strftime('%Y-%m-%d') for dclase in claseactividad if dt.isocalendar()[2] == dclase[0]]
                                        #
                                        #     dt += step
                                        __doc[1] += actividadnivelacioncarrera['totalfechas']
                                        __doc[2] += 100

                                if html == 'planificarcontenido':
                                    contenidohor = actividad.criteriodocenciaperiodo.horarios_contenido_profesor(
                                        distributivo.profesor, finiinicio, ffinal, True)
                                    if contenidohor == 0:
                                        listDocencia.append([actividad.criteriodocenciaperiodo.id,
                                                             actividad.criteriodocenciaperiodo.criterio.nombre,
                                                             actividad.horas, '-', '-'])
                                    else:
                                        count += 1
                                        totalhdocentes += contenidohor[1]
                                        __doc[1] += contenidohor[0]
                                        __doc[2] += contenidohor[1]
                                        # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, contenidohor[0], contenidohor[1]])

                                if html == 'tutoriaacademica':
                                    tutoriasacademicas = actividad.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(
                                        distributivo.profesor, finiinicio, ffinal, True)
                                    count += 1
                                    totalhdocentes += tutoriasacademicas[1]
                                    __doc[1] += tutoriasacademicas[0]
                                    __doc[2] += tutoriasacademicas[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, tutoriasacademicas[0], tutoriasacademicas[1]])

                                if html == 'seguimientoplataforma':
                                    listadoseguimientos = actividad.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(
                                        distributivo.profesor, finiinicio, ffinal, True)
                                    if listadoseguimientos[2]:
                                        count += 1
                                        totalhdocentes += listadoseguimientos[1]
                                    __doc[1] += listadoseguimientos[0]
                                    __doc[2] += listadoseguimientos[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoseguimientos[0], listadoseguimientos[1]])

                                if html == 'nivelacioncarrera':
                                    actividadgestion = actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                                        distributivo, finiinicio, ffinal, True)
                                    count += 1
                                    totalhdocentes += actividadgestion[1]
                                    __doc[1] += actividadgestion[0]
                                    __doc[2] += actividadgestion[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])

                                if html == 'seguimientotransversal':
                                    listadoseguimientos = actividad.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(
                                        distributivo.profesor, finiinicio, ffinal, True)
                                    if listadoseguimientos[2]:
                                        count += 1
                                        totalhdocentes += listadoseguimientos[1]
                                    __doc[1] += listadoseguimientos[0]
                                    __doc[2] += listadoseguimientos[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoseguimientos[0], listadoseguimientos[1]])

                                if html == 'apoyovicerrectorado':
                                    actividadapoyo = actividad.criteriodocenciaperiodo.horarios_apoyo_profesor(distributivo.profesor, finiinicio, ffinal)
                                    totitem10 = 0
                                    if actividadapoyo:
                                        totitem10 += 100
                                        totalhdocentes += 100
                                        count += 1
                                        __doc[1] += actividadapoyo['totalplanificadas']
                                        __doc[2] += 100

                                if html == 'actividaddocente':
                                    actividaddocente1 = actividad.criteriodocenciaperiodo.horarios_actividaddocente_profesor(distributivo.profesor, finiinicio, ffinal, True)
                                    if actividaddocente1:
                                        count += 1
                                        totalhdocentes += 100
                                        __doc[1] += actividaddocente1[0]
                                        __doc[2] += float(actividaddocente1[1])
                                    else:
                                        count += 1
                                        totalhdocentes += 0
                                        __doc[2] += 0

                                if html == 'criterioperiodoadmision':
                                    actividaddocente1 = actividad.criteriodocenciaperiodo.horario_criterio_nivelacion(distributivo.profesor, finiinicio, ffinal, True)
                                    count += 1
                                    totalhdocentes += actividaddocente1[1]
                                    __doc[1] += actividaddocente1[0]
                                    __doc[2] += actividaddocente1[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividaddocente1[0], actividaddocente1[1]])

                                if html == 'actividadbitacora':
                                    actividaddocente1 = listado_bitacora_docente(0, actividad, ffinal, True)
                                    count += 1
                                    totalhdocentes += actividaddocente1[1]
                                    __doc[1] += actividaddocente1[0]
                                    __doc[2] += actividaddocente1[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividaddocente1[0], actividaddocente1[1]])

                                if html == 'aplicadorexamen':
                                    if temp := listado_colectivos_academicos(actividad, finiinicio, ffinal):
                                        count += 1
                                        __doc[1] += temp['planificadas']
                                        __doc[2] += temp['porcentajegeneral']

                        lista_criterios.append({'tipocriterio': 1, 'data': listDocencia})
                        listDocencia = []

                    if horasinvestigacion := distributivo.detalle_horas_investigacion():
                        listDocencia.append([0, 'ACTIVIDADES DE INVESTIGACIÓN'])
                        for actividad in horasinvestigacion:
                            if html := actividad.criterioinvestigacionperiodo.nombrehtmldocente():
                                __inv[0] += actividad.horas
                                if html == 'actividadinvestigacion':
                                    actividadgestion = actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(distributivo, finiinicio, ffinal, True)
                                    count1 += 1
                                    totalhinvestigacion += actividadgestion[1]
                                    __inv[1] += actividadgestion[0]
                                    __inv[2] += actividadgestion[1]
                                    # listDocencia.append([actividad.criterioinvestigacionperiodo.id, actividad.criterioinvestigacionperiodo.criterio.nombre, actividad.horas, '', actividadgestion[1]])

                                if html == 'actividadbitacora':
                                    actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                                    count1 += 1
                                    totalhinvestigacion += actividadgestion[1]
                                    __inv[1] += actividadgestion[0]
                                    __inv[2] += actividadgestion[1]
                                    # listDocencia.append([actividad.criterioinvestigacionperiodo.id, actividad.criterioinvestigacionperiodo.criterio.nombre, actividad.horas, hmes, actividadgestion[1]])

                        lista_criterios.append({'tipocriterio': 2, 'data': listDocencia})
                        listDocencia = []

                    if horasgestion := distributivo.detalle_horas_gestion(finiinicio, ffinal):
                        listDocencia.append([0, 'ACTIVIDADES DE GESTIÓN EDUCATIVA'])
                        for actividad in horasgestion:
                            if html := actividad.criteriogestionperiodo.nombrehtmldocente():
                                __ges[0] += actividad.horas
                                if html == 'actividadgestion':
                                    actividadgestion = actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(distributivo.profesor, finiinicio, ffinal, True)
                                    if actividadgestion:
                                        count2 += 1
                                        totalhgestion += 100
                                        __ges[1] += actividadgestion[0]
                                        __ges[2] += float(actividadgestion[1])
                                        # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])
                                    else:
                                        count2 += 1
                                        totalhgestion += 0
                                        __ges[1] += 0
                                        __ges[2] += 0
                                        # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre, actividad.horas, '-', '0.00'])

                                if html == 'actividadinformegestion':
                                    actividadgestion = actividad.criteriogestionperiodo.horarios_informesgestion_profesor(
                                        distributivo, finiinicio, ffinal, True)
                                    count2 += 1
                                    totalhgestion += actividadgestion[1]
                                    __ges[1] += actividadgestion[0]
                                    __ges[2] += float(actividadgestion[1])
                                    # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre,actividad.horas, '', actividadgestion[1]])

                                if html == 'actividadbitacora':
                                    actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                                    count2 += 1
                                    totalhgestion += actividadgestion[1]
                                    __ges[1] += actividadgestion[0]
                                    __ges[2] += float(actividadgestion[1])
                                    # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre,actividad.horas, actividadgestion[0], actividadgestion[1]])
                        lista_criterios.append({'tipocriterio': 3, 'data': listDocencia})
                        listDocencia = []

                    if horasvinculacion := distributivo.detalle_horas_vinculacion():
                        docVinculacion = {'tipo': 'Horas Vinculacion'}
                        listVinculacion = []
                        listDocencia.append([0, 'ACTIVIDADES DE VINCULACIÓN CON LA SOCIEDAD'])
                        for actividad in horasvinculacion:
                            if html := actividad.criteriodocenciaperiodo.nombrehtmldocente():
                                __vin[0] += actividad.horas
                                if html == 'actividadvinculacion':
                                    actividadgestion = actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                                        distributivo, finiinicio, ffinal, True)
                                    if actividadgestion:
                                        count3 += 1
                                        totalhvinculacion += actividadgestion[1]
                                        __vin[1] += actividadgestion[0]
                                        __vin[2] += actividadgestion[1]
                                        # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,actividadgestion[0], actividadgestion[1]])
                                    else:
                                        count3 += 1
                                        totalhvinculacion += 0
                                        __vin[1] += 0
                                        __vin[2] += 0
                                        # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,'-', '0.00'])

                                if html == 'actividadbitacora':
                                    actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                                    count3 += 1
                                    totalhvinculacion += actividadgestion[1]
                                    __vin[1] += actividadgestion[0]
                                    __vin[2] += actividadgestion[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])

                        lista_criterios.append({'tipocriterio': 4, 'data': listDocencia})
                        listDocencia = []

                    totalporcentaje = totalhdocentes + totalhinvestigacion + totalhgestion + totalhvinculacion
                    count4 = count + count1 + count2 + count3
                    total_porcentaje = round(totalporcentaje / count4 if count4 else totalporcentaje, 2)

                    cuenta = cuenta + 1
                    porcgeneral = total_porcentaje
                    totalgeneral = __doc[0] + __inv[0] + __vin[0] + __ges[0]
                    cumplidogeneral = __doc[1] + __inv[1] + __vin[1] + __ges[1]
                    listado.append([distri,
                                    __doc[0], __doc[1], "%.0f" % (__doc[2] / count) if count else 0,
                                    __inv[0], __inv[1], "%.0f" % (__inv[2] / count1) if count1 else 0,
                                    __ges[0], __ges[1], "%.0f" % (__ges[2] / count2) if count2 else 0,
                                    __vin[0], __vin[1], "%.0f" % (__vin[2] / count3) if count3 else 0,
                                    round(totalgeneral, 0),
                                    round(porcgeneral, 0),
                                    round(100 - porcgeneral, 0),
                                    round(cumplidogeneral, 0),
                                    round(totalgeneral - cumplidogeneral, 0)])

                    DEBUG and print(listado.__str__())

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)

            nombrepersona = remover_caracteres_tildes_unicode(
                remover_caracteres_especiales_unicode((pers.__str__()).replace(' ', '_')))
            aleatorio = random.randint(1, 100000)
            nombredocumento = 'cumplimiento_{}_{}.pdf'.format(nombrepersona, aleatorio.__str__())
            nombredocumento1 = 'cumplimiento_{}_{}'.format(nombrepersona, aleatorio.__str__())
            valida = conviert_html_to_pdfsaveqr_generico(request,
                                                         'aprobar_silabo/reporteinformemensual_pdf.html',
                                                         {
                                                             'pagesize': 'A4',
                                                             'data': data,
                                                             'listado': listado,
                                                             # 'total_promedio': (to_promedio_to / len(listado)),
                                                         }, directory, nombredocumento1 + '.pdf'
                                                         )
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}/{}/{}".format(MEDIA_URL, 'reportes/recursos/', nombredocumento1 + '.pdf')
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Resultados de Encuesta',
                                    destinatario=pers,
                                    url="{}/{}/{}".format(MEDIA_URL, 'reportes/recursos/', nombredocumento1 + '.pdf'),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Reporte consolidado de cumplimiento de actividades y recursos de aprendizaje',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}/{}/{}".format(MEDIA_URL, 'reportes/recursos/', nombredocumento1 + '.pdf').replace("\n",
                                                                                                             " ").replace(
                    '\\', "/").replace("//", "/"),
                "btn_notificaciones": traerNotificaciones(request, data, pers).replace("\n", " ").replace('\\',
                                                                                                          "/").replace(
                    "//", "/"),
                "mensaje": 'Los resultados del cumplimineto de actividades han sido generados con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print(cuenta)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

class reporte_estudiantesaprobadoreprobadosadmision_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "reporte_estudiantes_aprobados_reprobados_admision_{}.xls".format(
            random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('todos alumnos')
            ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_estudiantes_aprobados_reprobados_admision' + random.randint(
                1, 10000).__str__() + '.xls'
            columns = [
                (u"FACULTAD", 6000),
                (u"CARRERA", 6000),
                (u"PERIODO ACADÉMICO ACTUAL", 6000),
                (u"PERIODO ACADÉMICO ANTERIOR", 6000),
                (u"CÉDULA", 6000),
                (u"APELLIDO PATERNO", 6000),
                (u"APELLIDO MATERNO", 6000),
                (u"NOMBRES", 3000),
                (u"EMAIL PERSONAL", 3000),
                (u"EMAIL INSTITUCIONAL", 6000),
                (u"#CELULAR", 3000),
                (u"#TELÉFONO CONVENCIONAL", 6000),
                (u"GÉNERO", 6000),
                (u"ETNIA", 6000),
                (u"LGBTI", 3000),
                (u"DISCAPACIDAD", 4000),
                (u"TIPO DISCAPACIDAD", 7000),
                (u"PPL", 4000),
                (u"MODALIDAD", 3000),
                (u"TÉRMINOS", 3000),
                (u"Nro. ASIGNATURAS", 3000),
                (u"PROMEDIO FINAL", 3000),
                (u"ESTATUS FINAL", 3000),
                (u"OBSERVACIÓN", 3000),
                (u"ESTADO GRATUIDAD", 3000),
                (u"Nro. MATRICULAS", 3000),
                (u"ID MATRÍCULA", 3000),
                (u"ID INSCRIPCION", 3000),
                (u"CANT APROBADAS", 3000),
                (u"MATERIA1", 17000),
                (u"PARALELO", 3000),
                (u"NOTA1", 3000),
                (u"ID MATERIA", 3000),
                (u"ID CURSO MOODLE", 3000),
                (u"MATERIA2", 17000),
                (u"PARALELO", 3000),
                (u"NOTA2", 3000),
                (u"ID MATERIA", 3000),
                (u"ID CURSO MOODLE", 3000),
                (u"MATERIA3", 17000),
                (u"PARALELO", 3000),
                (u"NOTA3", 3000),
                (u"ID MATERIA", 3000),
                (u"ID CURSO MOODLE", 3000),
                (u"MATERIA4", 17000),
                (u"PARALELO", 3000),
                (u"NOTA4", 3000),
                (u"ID MATERIA", 3000),
                (u"ID CURSO MOODLE", 3000),
                (u"MATERIA5", 17000),
                (u"PARALELO", 3000),
                (u"NOTA5", 3000),
                (u"ID MATERIA", 3000),
                (u"ID CURSO MOODLE", 3000),

            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            # profesormaterias = ProfesorMateria.objects.filter(status=True, materia__nivel__periodo=periodo).exclude(materia__asignaturamalla__malla__carrera__coordinacion__id__in=[7, 9]).order_by('materia__asignaturamalla__malla__carrera__coordinacion', 'materia__asignaturamalla__malla__carrera', 'materia')

            periodoanterior = \
                Periodo.objects.filter(status=True, clasificacion=3, fin__lt=self.periodo.fin).order_by('-inicio')[0]
            # inscripciones = Inscripcion.objects.filter(status=True, carrera__coordinacion__id=9, matricula__nivel__periodo__id=self.periodo.id).order_by('carrera__nombre', 'persona__apellido1', 'persona__apellido2', 'persona__nombres')
            matriculas = Matricula.objects.filter(nivel__periodo=self.periodo,
                                                  inscripcion__carrera__coordinacion__id=ADMISION_ID,
                                                  status=True).exclude(retiradomatricula=True)
            # if DEBUG:
            #     matriculas = matriculas[:100]
            row_num = 4
            col_num = 20

            # for inscripcion in inscripciones:

            for matricula in matriculas:
                # if Matricula.objects.filter(inscripcion=inscripcion, nivel__periodo__id=self.periodo.id).exists():
                #     matricula = Matricula.objects.get(inscripcion=inscripcion, nivel__periodo__id=self.periodo.id)
                periodo_anterior_nombre = ""
                numero_matriculas = "1"
                matricula_anterior = Matricula.objects.filter(  # inscripcion__persona=matricula.inscripcion.persona,
                    inscripcion_id=matricula.inscripcion_id,
                    nivel__periodo__inicio__year=matricula.nivel.periodo.inicio.year,
                    status=True,
                    retiradomatricula=False).exclude(pk=matricula.id).order_by('-nivel__periodo__inicio').first()

                if matricula_anterior:
                    periodo_anterior_nombre = matricula_anterior.nivel.periodo.nombre if matricula_anterior.nivel.periodo else ""
                    numero_matriculas = "2"

                totalaprobadas = matricula.materiaasignada_set.filter(status=True, retiramateria=False,
                                                                      estado=1).count()
                tipodiscapacidad = 'NINGUNA'
                carnetdiscapacidad = ''
                porcientodiscapacidad = ''
                nacionalidad = 'NO APLICA'
                raza = 'NO REGISTRA'
                politicacuota = 'NINGUNA'
                discapacidad = 'NO'
                if matricula.inscripcion.persona.perfilinscripcion_set.filter(status=True).exists():
                    pinscripcion = matricula.inscripcion.persona.perfilinscripcion_set.filter(status=True)[0]
                    if pinscripcion.tienediscapacidad:
                        discapacidad = 'SI'
                        # if pinscripcion.tipodiscapacidad_id in [5, 1, 4, 8, 9, 7]:
                        tipodiscapacidad = u'%s' % pinscripcion.tipodiscapacidad
                        carnetdiscapacidad = pinscripcion.carnetdiscapacidad if pinscripcion.carnetdiscapacidad else ''
                        porcientodiscapacidad = pinscripcion.porcientodiscapacidad if pinscripcion.tienediscapacidad else 0
                        politicacuota = 'DISCAPACIDAD'
                if pinscripcion.raza:
                    raza = pinscripcion.raza.nombre
                    if pinscripcion.raza.id == 1:
                        nacionalidad = u"%s" % pinscripcion.nacionalidadindigena
                        politicacuota = 'PUEBLOS Y NACIONALIDADES'
                ws.write(row_num, 0, u"%s" % (
                    matricula.inscripcion.carrera.coordinacionvalida.alias if matricula.inscripcion.carrera.coordinacionvalida else 'S/N'),
                         font_style2)
                ws.write(row_num, 1, u"%s" % matricula.inscripcion.carrera.nombre, font_style2)
                ws.write(row_num, 2, u"%s" % matricula.nivel.periodo.nombre, font_style2)
                # ws.write(row_num, 3, u"%s" % periodo_anterior_nombre, font_style2)
                ws.write(row_num, 3, u"%s" % matricula.inscripcion.persona.cedula, font_style2)
                ws.write(row_num, 4, u"%s" % matricula.inscripcion.persona.apellido1, font_style2)
                ws.write(row_num, 5, u"%s" % matricula.inscripcion.persona.apellido2, font_style2)
                ws.write(row_num, 6, u"%s" % matricula.inscripcion.persona.nombres, font_style2)
                ws.write(row_num, 7,
                         u"%s" % matricula.inscripcion.persona.email if matricula.inscripcion.persona.email else "",
                         font_style2)
                ws.write(row_num, 8,
                         u"%s" % matricula.inscripcion.persona.emailinst if matricula.inscripcion.persona.emailinst else "",
                         font_style2)
                ws.write(row_num, 9,
                         u"%s" % matricula.inscripcion.persona.telefono if matricula.inscripcion.persona.telefono else "",
                         font_style2)
                ws.write(row_num, 10,
                         u"%s" % matricula.inscripcion.persona.telefono_conv if matricula.inscripcion.persona.telefono_conv else "",
                         font_style2)
                ws.write(row_num, 11,
                         u"%s" % matricula.inscripcion.persona.sexo.nombre if matricula.inscripcion.persona.sexo.nombre else "",
                         font_style2)
                ws.write(row_num, 12, u"%s" % raza, font_style2)
                ws.write(row_num, 13, u"%s" % "SI" if matricula.inscripcion.persona.lgtbi else "NO", font_style2)
                ws.write(row_num, 14, u"%s" % discapacidad, font_style2)
                ws.write(row_num, 15, u"%s" % tipodiscapacidad, font_style2)
                ws.write(row_num, 16, u"%s" % "SI" if matricula.inscripcion.persona.ppl else "NO", font_style2)
                ws.write(row_num, 17, u"%s" % matricula.inscripcion.modalidad, font_style2)
                ws.write(row_num, 18, u"%s" % "SI" if matricula.termino == True else "NO", font_style2)
                ws.write(row_num, 19, u"%s" % matricula.cantidad_materias(), font_style2)
                ws.write(row_num, 20, u"%s" % matricula.promedio_nota(), font_style2)
                ws.write(row_num, 21, u"%s" % "APROBADO" if matricula.materias_aprobadas_todas() else "REPROBADO",
                         font_style2)
                ws.write(row_num, 22, u"%s" % "Aprueba la Nivelación solo si en cada asignatura obtuvo como mínimo 70",
                         font_style2)
                ws.write(row_num, 23, u"%s" % matricula.inscripcion.estado_perdida_gratuidad(), font_style2)
                # matricula.inscripcion.matricula_set.filter(status=True, nivel__periodo_id=)
                ws.write(row_num, 24, u"%s" % numero_matriculas, font_style2)
                ws.write(row_num, 25, u"%s" % matricula.id, font_style2)
                ws.write(row_num, 26, u"%s" % matricula.inscripcion.id, font_style2)
                ws.write(row_num, 27, u"%s" % totalaprobadas, font_style2)
                materias = MateriaAsignada.objects.filter(matricula_id=matricula.id)
                # numcolu = 29
                # numcolu1 = 30
                # numcolu2 = 31
                # for materia in materias:
                #     ws.write(row_num, numcolu, u"%s" % materia.materia.nombre_mostrar_alias(), font_style2)
                #     ws.write(row_num, numcolu1, u"%s" % materia.notafinal, font_style2)
                #     ws.write(row_num, numcolu2, u"%s" % materia.id, font_style2)
                #     numcolu += 3
                #     numcolu1 += 3
                #     numcolu2 += 3
                ordenMaterias = [None, None, None, None, None]
                ordenMaterias[0] = materias.filter(status=True, materia__asignatura_id=4837).first()  # PROPEDÉUTICO
                ordenMaterias[1] = materias.filter(status=True, materia__asignatura_id__in=[969,
                                                                                            2688]).first()  # MATEMÁTICAS, ANATOMÍA
                ordenMaterias[2] = materias.filter(status=True,
                                                   materia__asignatura_id__in=[2677, 171]).first()  # Biología, ICA
                ordenMaterias[3] = materias.filter(status=True,
                                                   materia__asignatura_id=4881).first()  # Pensamiento Computacional
                ordenMaterias[4] = materias.filter(status=True, materia__asignatura_id=173).first()  # BIOQUÍMICA
                numcolu = 29
                numcolu1 = 30
                numcolu2 = 31
                numcolu3 = 32
                numcolu4 = 33
                for materia in ordenMaterias:
                    ws.write(row_num, numcolu,
                             u"%s" % materia.materia.nombre_mostrar_alias() if materia is not None else '', font_style2)
                    ws.write(row_num, numcolu1, u"%s" % materia.materia.paralelo if materia is not None else '',
                             font_style2)
                    ws.write(row_num, numcolu2, u"%s" % materia.notafinal if materia is not None else '', font_style2)
                    ws.write(row_num, numcolu3, u"%s" % materia.id if materia is not None else '', font_style2)
                    ws.write(row_num, numcolu4, u"%s" % materia.materia.idcursomoodle if materia else '', font_style2)
                    numcolu += 5
                    numcolu1 += 5
                    numcolu2 += 5
                    numcolu3 += 5
                    numcolu4 += 5
                row_num += 1
            wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel Estudiantes Admisión - aprobados y reprobados',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel Estudiantes Admisión - aprobados y reprobados',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


###revisar

class reporte_generalcapacitaciones_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_general_capacitacion_docente' + random.randint(1, 10000).__str__() + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente', nombre_archivo)

        try:
            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_general_capacitacion_docente' + random.randint(
                1, 10000).__str__() + '.xls'

            columns = [
                (u"APELLIDOS", 25),
                (u"NOMBRES", 25),
                (u"FACULTAD", 25),
                (u"CARRERA", 25),
                (u"MODALIDAD DE CONTRATO", 25),
                (u"CURSOS REALIZADOS", 25)

            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            periodo = CapPeriodoDocente.objects.get(pk=self.periodo.id)
            if periodo.id == 4:
                totalinscritos = CapCabeceraSolicitudDocente.objects.filter(
                    capeventoperiodo__periodo_id=self.periodo.id, capeventoperiodo__fechainicio__gte='2022-10-15',
                    capeventoperiodo__status=True,
                    status=True).distinct('participante__apellido1', 'participante__apellido2',
                                          'participante__nombres').order_by('participante')
            else:
                totalinscritos = CapCabeceraSolicitudDocente.objects.filter(
                    capeventoperiodo__periodo_id=self.periodo.id,
                    status=True).distinct('participante__apellido1', 'participante__apellido2',
                                          'participante__nombres').order_by('participante')

            row_num = 5
            for inscrito in totalinscritos:
                apellidos = inscrito.participante.apellido1 + ' ' + inscrito.participante.apellido2
                nombres = inscrito.participante.nombres
                facultad = inscrito.facultad if inscrito.facultad else 'NINGUNO'
                carrera = inscrito.carrera if inscrito.carrera else 'NINGUNO'

                if inscrito.participante.distributivopersona_set.filter(status=True).exists():
                    plantilla = inscrito.participante.distributivopersona_set.filter(status=True)[0]
                    modalidad = plantilla.modalidadlaboral
                else:
                    modalidad = 'NINGUNA'

                ws.write(row_num, 0, u'%s' % apellidos, font_style2)
                ws.write(row_num, 1, u'%s' % nombres, font_style2)
                ws.write(row_num, 2, u'%s' % facultad, font_style2)
                ws.write(row_num, 3, u'%s' % carrera, font_style2)
                ws.write(row_num, 4, u'%s' % modalidad, font_style2)
                if periodo.id == 4:
                    eventos = periodo.capeventoperiododocente_set.filter(status=True, fechainicio__gte='2022-10-15')
                else:
                    eventos = periodo.capeventoperiododocente_set.filter(status=True)
                numcolu = 5
                for evento in eventos:
                    if CapCabeceraSolicitudDocente.objects.filter(status=True,
                                                                  capeventoperiodo__periodo_id=self.periodo.id,
                                                                  participante=inscrito.participante.id,
                                                                  capeventoperiodo=evento).exists():
                        ws.write(row_num, numcolu, u"%s" % evento.capevento, font_style2)
                        numcolu += 1

                row_num += 1

            wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Excel reporte general de capacitaciones',
                                    destinatario=pers, url="{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel reporte general de capacitaciones',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_capacitaciones_aprobadas_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_capacitaciones_aprobadas' + random.randint(1, 10000).__str__() + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente', nombre_archivo)

        try:
            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_capacitaciones_aprobadas' + random.randint(
                1, 10000).__str__() + '.xls'

            columns = [
                (u"APELLIDOS", 25),
                (u"NOMBRES", 25),
                (u"FACULTAD", 25),
                (u"CARRERA", 25),
                (u"MODALIDAD DE CONTRATO", 25),
                (u"CURSOS APROBADOS", 25)

            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            periodo = CapPeriodoDocente.objects.get(pk=self.periodo.id)

            if periodo.id == 4:

                totalinscritos = CapCabeceraSolicitudDocente.objects.filter(
                    capeventoperiodo__periodo_id=self.periodo.id, capeventoperiodo__fechainicio__gte='2022-10-15',
                    capeventoperiodo__status=True,
                    status=True).distinct('participante__apellido1', 'participante__apellido2',
                                          'participante__nombres').order_by('participante')
            else:
                totalinscritos = CapCabeceraSolicitudDocente.objects.filter(
                    capeventoperiodo__periodo_id=self.periodo.id, capeventoperiodo__status=True,
                    status=True).distinct('participante__apellido1', 'participante__apellido2',
                                          'participante__nombres').order_by('participante')

            row_num = 5
            for inscrito in totalinscritos:
                apellidos = inscrito.participante.apellido1 + ' ' + inscrito.participante.apellido2
                nombres = inscrito.participante.nombres
                facultad = inscrito.facultad if inscrito.facultad else 'NINGUNO'
                carrera = inscrito.carrera if inscrito.carrera else 'NINGUNO'

                if inscrito.participante.distributivopersona_set.filter(status=True).exists():
                    plantilla = inscrito.participante.distributivopersona_set.filter(status=True)[0]
                    modalidad = plantilla.modalidadlaboral
                else:
                    modalidad = 'NINGUNA'

                ws.write(row_num, 0, u'%s' % apellidos, font_style2)
                ws.write(row_num, 1, u'%s' % nombres, font_style2)
                ws.write(row_num, 2, u'%s' % facultad, font_style2)
                ws.write(row_num, 3, u'%s' % carrera, font_style2)
                ws.write(row_num, 4, u'%s' % modalidad, font_style2)
                if periodo.id == 4:
                    eventos = periodo.capeventoperiododocente_set.filter(status=True, fechainicio__gte='2022-10-15')
                else:
                    eventos = periodo.capeventoperiododocente_set.filter(status=True)
                numcolu = 5
                for evento in eventos:
                    if CapCabeceraSolicitudDocente.objects.filter(status=True,
                                                                  capeventoperiodo__periodo_id=self.periodo.id,
                                                                  participante=inscrito.participante.id,
                                                                  capeventoperiodo=evento).exists():
                        inscritoevento = CapCabeceraSolicitudDocente.objects.get(status=True,
                                                                                 capeventoperiodo__periodo_id=self.periodo.id,
                                                                                 participante=inscrito.participante.id,
                                                                                 capeventoperiodo=evento)
                        if evento.tipoparticipacion_id == 1:
                            if inscritoevento.porciento_requerido_asistencia():
                                ws.write(row_num, numcolu, u"%s" % evento.capevento, font_style2)
                                numcolu += 1
                        elif evento.tipoparticipacion_id == 2:
                            if inscritoevento.calificacion_requerido_aprobacion():
                                ws.write(row_num, numcolu, u"%s" % evento.capevento, font_style2)
                                numcolu += 1

                        elif evento.tipoparticipacion_id == 3:
                            if inscrito.calificacion_requerido_aprobacion() and inscrito.porciento_requerido_asistencia():
                                ws.write(row_num, numcolu, u"%s" % evento.capevento, font_style2)
                                numcolu += 1

                row_num += 1

            wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Excel reporte_capacitaciones_aprobadas',
                                    destinatario=pers, url="{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel reporte capacitaciones aprobadas',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_generalcapacitaciones_inscritos_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_general_aprobados_capacitacion_docente' + random.randint(1, 10000).__str__() + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente', nombre_archivo)

        try:

            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_general_aprobados_capacitacion_docente' + random.randint(
                1, 10000).__str__() + '.xls'

            columns = [
                (u"EVENTO", 25),
                (u"INSCRITOS", 25),
                (u"APROBADOS", 25),
                (u"REPROBADOS", 25),
                (u"% CUMPLIMIENTO", 25)
            ]
            row_num = 4

            periodo = CapPeriodoDocente.objects.get(pk=self.periodo.id)

            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            if periodo.id == 4:
                eventos = periodo.capeventoperiododocente_set.filter(status=True, fechainicio__gte='2022-10-15')
            else:
                eventos = periodo.capeventoperiododocente_set.filter(status=True)

            row_num = 5
            for evento in eventos:
                nomevento = evento.capevento
                inscritos = evento.contar_inscripcion_evento_periodo()
                aprobados = evento.total_inscritos_aprobados()
                reprobados = evento.total_inscritos_reprobados()
                try:
                    cumplimiento = round(((aprobados * 100) / inscritos), 2)
                except ZeroDivisionError:
                    cumplimiento = 0

                ws.write(row_num, 0, u'%s' % nomevento, font_style2)
                ws.write(row_num, 1, u'%s' % inscritos, font_style2)
                ws.write(row_num, 2, u'%s' % aprobados, font_style2)
                ws.write(row_num, 3, u'%s' % reprobados, font_style2)
                ws.write(row_num, 4, u'%s' % cumplimiento, font_style2)

                row_num += 1

            wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel reporte general de inscritos en capacitaciones',
                                    destinatario=pers, url="{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel reporte general de inscritos en capacitaciones',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_generalcapacitaciones_facultad_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_general_aprobados_facultad_capacitacion_docente' + random.randint(1,
                                                                                                    10000).__str__() + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente', nombre_archivo)

        try:

            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_general_aprobados_facultad_capacitacion_docente' + random.randint(
                1, 10000).__str__() + '.xls'

            columns = [
                (u"FACULTAD", 30),
                (u"EVENTO", 50),
                (u"INSCRITOS", 25),
                (u"APROBADOS", 25),
                (u"REPROBADOS", 25),
                (u"% CUMPLIMIENTO", 25)

            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            periodo = CapPeriodoDocente.objects.get(pk=self.periodo.id)

            row_num = 5

            if periodo.id == 4:
                eventos = periodo.capeventoperiododocente_set.filter(status=True, fechainicio__gte='2022-10-15')
            else:
                eventos = periodo.capeventoperiododocente_set.filter(status=True)

            for evento in eventos:
                # departamentos = evento.total_departamentos()

                facultades = evento.listado_facultades()

                for facultad in facultades:

                    # else:
                    #     facultad = Departamento.objects.get(pk=departamento)
                    nomevento = evento.capevento
                    inscritos = evento.contar_inscripcion_evento_periodo_por_facultad2(facultad)
                    aprobados = evento.total_inscritos_aprobados_facultad2(facultad)
                    reprobados = evento.total_inscritos_reprobados_facultad2(facultad)
                    try:
                        cumplimiento = round(((aprobados * 100) / inscritos), 2)
                    except ZeroDivisionError:
                        cumplimiento = 0
                    #
                    #
                    if facultad is None:
                        facultad = 'NINGUNO'
                    ws.write(row_num, 0, u'%s' % facultad, font_style2)
                    ws.write(row_num, 1, u'%s' % nomevento, font_style2)
                    ws.write(row_num, 2, u'%s' % inscritos, font_style2)
                    ws.write(row_num, 3, u'%s' % aprobados, font_style2)
                    ws.write(row_num, 4, u'%s' % reprobados, font_style2)
                    ws.write(row_num, 5, u'%s' % cumplimiento, font_style2)

                    row_num += 1

            wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel reporte general de inscritos facultad en capacitaciones',
                                    destinatario=pers, url="{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel reporte general de inscritos facultad en capacitaciones',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_totalinscritos_facultad_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_total_inscritos_facultad_capacitacion_docente' + random.randint(1,
                                                                                                  10000).__str__() + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente', nombre_archivo)

        try:

            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_total_inscritos_facultad_capacitacion_docente' + random.randint(
                1, 10000).__str__() + '.xls'

            columns = [
                (u"FACULTAD", 30),
                (u"INSCRITOS", 25),
                (u"APROBADOS", 25),
                (u"REPROBADOS", 25),
                (u"% CUMPLIMIENTO", 25)

            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            periodo = CapPeriodoDocente.objects.get(pk=self.periodo.id)

            row_num = 5

            # eventos = periodo.capeventoperiododocente_set.filter(status=True)

            # for periodo in periodos:
            # departamentos = evento.total_departamentos()

            facultades = periodo.listado_facultades()

            for facultad in facultades:

                inscritos = periodo.contar_inscritos_periodo_facultad(facultad)
                aprobados = periodo.total_inscritos_aprobados_facultad2_periodo(facultad)
                reprobados = periodo.total_inscritos_reprobados_facultad2_periodo(facultad)
                try:
                    cumplimiento = round(((aprobados * 100) / inscritos), 2)
                except ZeroDivisionError:
                    cumplimiento = 0

                    #
                    #
                if facultad is None:
                    facultad = 'NINGUNO'
                ws.write(row_num, 0, u'%s' % facultad, font_style2)
                # ws.write(row_num, 1, u'%s' % nomevento, font_style2)
                ws.write(row_num, 1, u'%s' % inscritos, font_style2)
                ws.write(row_num, 2, u'%s' % aprobados, font_style2)
                ws.write(row_num, 3, u'%s' % reprobados, font_style2)
                ws.write(row_num, 4, u'%s' % cumplimiento, font_style2)

                row_num += 1

            wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel reporte total de inscritos facultad en capacitaciones',
                                    destinatario=pers, url="{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel reporte total de inscritos facultad en capacitaciones',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_inscritos_carrera_capdocente_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_inscritos_carrera_capdocente' + random.randint(1, 10000).__str__() + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente', nombre_archivo)

        try:

            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_inscritos_carrera_capdocente' + random.randint(
                1, 10000).__str__() + '.xls'

            columns = [
                (u"EVENTO", 50),
                (u"CARRERA", 50),
                (u"FACULTAD", 50),
                (u"INSCRITOS", 25),
                (u"APROBADOS", 25),
                (u"REPROBADOS", 25),
                (u"HORAS", 25)

            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            periodo = CapPeriodoDocente.objects.get(pk=self.periodo.id)

            row_num = 5

            if periodo.id == 4:
                eventos = periodo.capeventoperiododocente_set.filter(status=True, fechainicio__gte='2022-10-15')
            else:
                eventos = periodo.capeventoperiododocente_set.filter(status=True)

            for evento in eventos:

                carreras = evento.listado_carreras()

                for carrera in carreras:

                    facultad = 'NINGUNO'
                    nomevento = evento.capevento
                    inscritos = evento.contar_inscripcion_evento_periodo_por_carrera(carrera)
                    aprobados = evento.total_inscritos_aprobados_carrera(carrera)
                    reprobados = evento.total_inscritos_reprobados_carrera(carrera)
                    horas = evento.horas

                    if carrera:
                        fac = carrera.mi_coordinacion2()
                        facultad = Coordinacion.objects.get(status=True, pk=fac)

                    if carrera is None:
                        carrera = 'NINGUNO'
                    ws.write(row_num, 0, u'%s' % nomevento, font_style2)
                    ws.write(row_num, 1, u'%s' % carrera, font_style2)
                    ws.write(row_num, 2, u'%s' % facultad, font_style2)
                    ws.write(row_num, 3, u'%s' % inscritos, font_style2)
                    ws.write(row_num, 4, u'%s' % aprobados, font_style2)
                    ws.write(row_num, 5, u'%s' % reprobados, font_style2)
                    ws.write(row_num, 6, u'%s' % horas, font_style2)

                    row_num += 1

            wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel reporte general de inscritos por carrera en capacitaciones',
                                    destinatario=pers, url="{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel reporte general de inscritos carrera en capacitaciones',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/capdocente/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_cumplimiento_titulacion_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_cumplimiento_titulacion' + random.randint(1, 10000).__str__() + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 24, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_inscritos_carrera_capdocente' + random.randint(
                1, 10000).__str__() + '.xls'
            columns = [
                (u"MALLA", 29192),
                (u"FACULTAD", 16698),
                (u"CARRERA", 8957),
                (u"MATERIA", 16698),
                (u"NIVEL MATERIA", 8957),
                (u"MODALIDAD", 8957),
                (u"NIVEL MALLA", 4083),
                (u"CÉDULA", 3570),
                (u"ESTUDIANTE", 13045),
                (u"CORREO INSTITUCIONAL", 10720),
                (u"No adeudar valores o aranceles a la institución", 12059),
                (u"Haber cumplido con las horas o créditos de vinculación", 13045),
                (u"Haber cumplido con las horas o créditos de las prácticas pre profesionales", 17394),
                (u"Haber aprobado la suficiencia en computación", 10726),
                (u"Haber aprobado módulos o suficiencia de idioma inglés", 12784),
                (u"Haber aprobado módulos de suficiencia de segundo idioma (Ingles , Frances o Quichua)", 20000),
                (u"Ficha estudiantil actualizada y completa", 9189),
                (u"Asignaturas aprobadas del primer al penúltimo nivel", 12059),
                (u"Asignaturas aprobadas de primer al ultimo nivel", 11073),
                (u"Asignaturas aprobadas del primero al septimo nivel", 11887),
                (u"Estar matriculado todas las asignaturas del ultimo periodo academico", 16176),
                (u"Certificación segunda lengua sin aprobar director", 11508),
                (u"Certificación segunda lengua aprobado director", 11073),
                (u"Inscripcion", 3131),
                (u"Matricula", 3131)
            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            matriculas = Matricula.objects.filter(nivel__periodo_id=periodo.id,
                                                  estado_matricula__in=[2, 3],
                                                  status=True,
                                                  nivelmalla_id__in=[8, 9]
                                                  ).exclude(retiradomatricula=True).order_by('inscripcion__carrera')

            total = matriculas.count()

            row_num = 5

            for matricula in matriculas:
                materiaasignada = MateriaAsignada.objects.filter(status=True, matricula=matricula,
                                                                 materia__modeloevaluativo__id__in=[25, 26])
                # idins = matricula.inscripcion.persona.perfil_usuario_inscripcion_vigente()
                for matasig in materiaasignada:
                    inscripcion = matasig.matricula.inscripcion

                    matricula = matasig.matricula
                    # inscripcionid = matasig.matricula.inscripcion.id
                    nivelmalla = matasig.matricula.nivelmalla

                    malla = inscripcion.carrera.malla()

                    facultad = matasig.matricula.nivel.coordinacion()
                    carrera = inscripcion.carrera.nombre

                    persona = inscripcion.persona

                    cedula = persona.identificacion()
                    nombres = persona.nombre_completo_inverso()
                    correoinst = persona.emailinst
                    celular = persona.telefono
                    asignatura = matasig.materia.asignaturamalla
                    materia = matasig.materia.asignatura
                    modalidad = inscripcion.modalidad

                    requsitosasignatura = asignatura.requisitoingresounidadintegracioncurricular_set.filter(
                        obligatorio=True, status=True)
                    req1 = 'NO APLICA'
                    req2 = 'NO APLICA'
                    req3 = 'NO APLICA'
                    req4 = 'NO APLICA'
                    req5 = 'NO APLICA'
                    req6 = 'NO APLICA'
                    req7 = 'NO APLICA'
                    req8 = 'NO APLICA'
                    req9 = 'NO APLICA'
                    req10 = 'NO APLICA'
                    req11 = 'NO APLICA'
                    req12 = 'NO APLICA'
                    req13 = 'NO APLICA'

                    for requisito in requsitosasignatura:
                        if requisito.requisito_id == 1:
                            req1 = 'SI' if no_adeudar_institucion(inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 2:
                            req2 = 'SI' if haber_cumplido_horas_creditos_vinculacion(inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 3:
                            req3 = 'SI' if haber_cumplido_horas_creditos_practicas_preprofesionales(
                                inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 4:
                            req4 = 'SI' if haber_aprobado_modulos_computacion(inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 5:
                            req5 = 'SI' if haber_aprobado_modulos_ingles(inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 6:
                            req6 = 'SI' if ficha_estudiantil_actualizada_completa(inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 7:
                            req7 = 'SI' if asignaturas_aprobadas_primero_penultimo_nivel(inscripcion.id) else 'NO'



                        elif requisito.requisito_id == 8:
                            req8 = 'SI' if estar_matriculado_todas_asignaturas_ultimo_periodo_academico(
                                inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 9:
                            req9 = 'SI' if tiene_certificacion_segunda_lengua_sin_aprobar_director_carrera(
                                inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 10:
                            req10 = 'SI' if tiene_certificacion_segunda_lengua_aprobado_director_carrera(
                                inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 11:
                            req11 = 'SI' if haber_aprobado_modulos_ingles(
                                inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 12:
                            req12 = 'SI' if asignaturas_aprobadas_primero_ultimo_nivel(
                                inscripcion.id) else 'NO'

                        elif requisito.requisito_id == 13:
                            req13 = 'SI' if asignaturas_aprobadas_primero_septimo_nivel(
                                inscripcion.id) else 'NO'

                    ws.write(row_num, 0, u'%s' % malla, font_style2)
                    ws.write(row_num, 1, u'%s' % facultad, font_style2)
                    ws.write(row_num, 2, u'%s' % carrera, font_style2)
                    ws.write(row_num, 3, u'%s' % materia, font_style2)
                    ws.write(row_num, 4, u'%s' % asignatura.nivelmalla, font_style2)
                    ws.write(row_num, 5, u'%s' % modalidad, font_style2)
                    ws.write(row_num, 6, u'%s' % nivelmalla, font_style2)
                    ws.write(row_num, 7, u'%s' % cedula, font_style2)
                    ws.write(row_num, 8, u'%s' % nombres, font_style2)
                    ws.write(row_num, 9, u'%s' % correoinst, font_style2)
                    ws.write(row_num, 10, u'%s' % req1, font_style2)
                    ws.write(row_num, 11, u'%s' % req2, font_style2)
                    ws.write(row_num, 12, u'%s' % req3, font_style2)
                    ws.write(row_num, 13, u'%s' % req4, font_style2)
                    ws.write(row_num, 14, u'%s' % req5, font_style2)
                    ws.write(row_num, 15, u'%s' % req11, font_style2)
                    ws.write(row_num, 16, u'%s' % req6, font_style2)
                    ws.write(row_num, 17, u'%s' % req7, font_style2)
                    ws.write(row_num, 18, u'%s' % req12, font_style2)
                    ws.write(row_num, 19, u'%s' % req13, font_style2)
                    ws.write(row_num, 20, u'%s' % req8, font_style2)
                    ws.write(row_num, 21, u'%s' % req9, font_style2)
                    ws.write(row_num, 22, u'%s' % req10, font_style2)
                    ws.write(row_num, 23, u'%s' % inscripcion.id, font_style2)
                    ws.write(row_num, 24, u'%s' % matricula.id, font_style2)

                    row_num += 1
                wb.save(directory)

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Reporte cumplimiento titulacion',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)

                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel de reporte de estudiantes requisitos de titulacion',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error al generar el reporte'
                noti.cuerpo = str(ex)
                noti.url = ""
            else:
                noti = Notificacion(cuerpo=str(ex), titulo='Error al generar el reporte',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
            noti.save()
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_estudiantesaprobadoreprobadosadmision_matriz_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        import pandas as pd
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "reporte_estudiantes_aprobados_reprobados_admision_{}.xlsx".format(
            random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            matriculas = Matricula.objects.filter(nivel__periodo=self.periodo,
                                                  inscripcion__carrera__coordinacion__id=9,
                                                  status=True, termino=True).exclude(retiradomatricula=True).annotate(
                cantidad_asignaturas=Count('inscripcion__inscripcionmalla__malla__asignaturamalla__id',
                                           filter=Q(inscripcion__inscripcionmalla__malla__status=True,
                                                    inscripcion__inscripcionmalla__status=True
                                                    )
                                           )
            ).distinct()

            asignaturas_maxima = matriculas.aggregate(max_asignatura=Max('cantidad_asignaturas'))['max_asignatura']
            __author__ = 'Unemi'
            ahora = datetime.now()
            time_codigo = ahora.strftime('%Y%m%d_%H%M%S')
            name_file = f'reporte_excel_pedidos_online_{time_codigo}.xlsx'
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet("Listado Estudiantes")

            fuentecabecera = workbook.add_format({
                'align': 'center',
                'bg_color': 'silver',
                'border': 1,
                'bold': 1
            })

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            fuenteencabezado = workbook.add_format({
                'align': 'center',
                'bg_color': '#1C3247',
                'font_color': 'white',
                'border': 1,
                'font_size': 24,
                'bold': 1
            })

            columnas = [
                (u'FACULTAD', 10),
                (u'CARRERA', 80),
                (u'PERIODO ACADÉMICO ACTUAL', 80),
                #(u'ESTADO VERIFICACIÓN MATRIZ', 80),
                (u'CÉDULA', 15),
                (u'APELLIDO PATERNO', 40),
                (u'APELLIDO MATERNO', 40),
                (u'NOMBRES', 40),
                (u'EMAIL PERSONAL', 40),
                (u'EMAIL INSTITUCIONAL', 40),
                (u"#CELULAR", 30),
                (u"#TELÉFONO CONVENCIONAL", 40),
                (u"GÉNERO", 15),
                (u"ETNIA", 40),
                (u"LGBTI", 10),
                (u"DISCAPACIDAD", 50),
                (u"TIPO DISCAPACIDAD", 50),
                (u"PPL", 10),
                (u"MODALIDAD", 40),
                (u"TÉRMINOS", 10),
                (u"Nro. ASIGNATURAS", 40),
                (u"PROMEDIO FINAL", 40),
                (u"ESTATUS FINAL", 50),
                (u"OBSERVACIÓN", 50),
                (u"ESTADO GRATUIDAD", 50),
                (u"Nro. MATRICULAS", 50),
                (u"ID MATRÍCULA", 40),
                (u"ID INSCRIPCION", 40),
                (u"CANT APROBADAS", 40),
            ]

            for item in range(1, asignaturas_maxima + 1):
                columnas.extend([
                    (u"MATERIA%s" % item, 90),
                    (u"PARALELO%s" % item, 10),
                    (u"NOTA%s" % item, 10),
                    (u"ID MATERIA%s" % item, 10),
                    (u"ID CURSO MOODLE%s" % item, 10),
                ])

            ws.merge_range(0, 0, 0, columnas.__len__() - 1, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, columnas.__len__() - 1,
                           f'LISTADO DE ESTUDIANTES APROBADOS Y REPROBADOS DEL PERIODO {self.periodo}',
                           fuenteencabezado)

            row_num, numcolum = 2, 0

            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1

            row_num += 1
            periodo_anterior = Periodo.objects.filter(status=True,
                                                      clasificacion=3,
                                                      fin__lt=self.periodo.fin).order_by('-inicio').first()

            # matriz_periodo = self.periodo.subirmatrizinscripcion_set.filter(estado=2, status=True).first()
            # matriz_periodo_anterior = periodo_anterior.subirmatrizinscripcion_set.filter(estado=2,
            #                                                                              status=True).first() if periodo_anterior is not None else None

            columnas_excel = ['CARRERA_ID', 'CEDULA']
            # df_matriz_periodo = pd.read_excel(matriz_periodo.archivo, usecols=columnas_excel) if matriz_periodo else []
            # df_matriz_periodo_anterior = pd.read_excel(matriz_periodo_anterior.archivo,
            #                                            usecols=columnas_excel) if matriz_periodo_anterior else []

            if DEBUG:
                matriculas = matriculas[:100]

            for key, matricula in enumerate(matriculas):
                ePerfilInscripcion = matricula.inscripcion.persona.perfilinscripcion_set.filter(status=True).first()
                raza = 'NO REGISTRA'
                discapacidad = 'NO'
                tipodiscapacidad = 'NINGUNA'
                if ePerfilInscripcion is not None:
                    if ePerfilInscripcion.tienediscapacidad:
                        discapacidad = 'Si'
                        tipodiscapacidad = u'%s' % ePerfilInscripcion.tipodiscapacidad
                    if ePerfilInscripcion.raza:
                        raza = ePerfilInscripcion.raza.nombre

                # busqueda_matriz_actual = []
                # busqueda_matriz_anterior = []
                # if df_matriz_periodo.__len__() > 0 and df_matriz_periodo_anterior.__len__() > 0:
                #     busqueda_matriz_actual = df_matriz_periodo.loc[
                #         (df_matriz_periodo['CARRERA_ID'] == matricula.inscripcion.carrera_id) & (
                #                     df_matriz_periodo['CEDULA'] == matricula.inscripcion.persona.cedula)]
                #     busqueda_matriz_anterior = df_matriz_periodo_anterior.loc[
                #         (df_matriz_periodo_anterior['CARRERA_ID'] == matricula.inscripcion.carrera_id) & (
                #                     df_matriz_periodo_anterior['CEDULA'] == matricula.inscripcion.persona.cedula)]
                # elif df_matriz_periodo.__len__() > 0:
                #     busqueda_matriz_actual = df_matriz_periodo.loc[
                #         (df_matriz_periodo['CARRERA_ID'] == matricula.inscripcion.carrera_id) & (
                #                     df_matriz_periodo['CEDULA'] == matricula.inscripcion.persona.cedula)]
                # elif df_matriz_periodo.__len__() > 0:
                #     busqueda_matriz_anterior = df_matriz_periodo_anterior.loc[
                #         (df_matriz_periodo_anterior['CARRERA_ID'] == matricula.inscripcion.carrera_id) & (
                #                     df_matriz_periodo_anterior['CEDULA'] == matricula.inscripcion.persona.cedula)]
                #
                # texto_matriz = ""
                # numero_matriculas = "1"
                # if len(busqueda_matriz_actual) > 0 and len(busqueda_matriz_anterior) > 0:
                #     texto_matriz = f"SE ENCUENTRA EN  LAS 2 MATRICES => ({self.periodo.nombre}, {periodo_anterior.nombre})"
                # elif len(busqueda_matriz_actual) > 0:
                #     texto_matriz = f"SE ENCUENTRA EN  LA MATRIZ => {self.periodo.nombre}"
                # elif len(busqueda_matriz_anterior) > 0:
                #     numero_matriculas = "2"
                #     texto_matriz = f"SE ENCUENTRA EN  LA MATRIZ => {periodo_anterior.nombre}"
                # elif matriz_periodo_anterior is not None and matriz_periodo is not None:
                #     texto_matriz = f"EL REGISTRO NO SE ENCUENTRA EN NINGUNA MATRIZ => ({self.periodo.nombre}, {periodo_anterior.nombre})"

                totalaprobadas = matricula.materiaasignada_set.filter(status=True, retiramateria=False,
                                                                      estado=1).count()
                ws.write(row_num, 0, u"%s" % (
                    matricula.inscripcion.carrera.coordinacionvalida.alias if matricula.inscripcion.carrera.coordinacionvalida else 'S/N'),
                         formatoceldacenter)
                ws.write(row_num, 1, u"%s" % matricula.inscripcion.carrera.nombre, formatoceldacenter)
                ws.write(row_num, 2, u"%s" % matricula.nivel.periodo.nombre, formatoceldacenter)
                #ws.write(row_num, 3, u"%s" % texto_matriz, formatoceldacenter)
                ws.write(row_num, 4, u"%s" % matricula.inscripcion.persona.cedula, formatoceldacenter)
                ws.write(row_num, 5, u"%s" % matricula.inscripcion.persona.apellido1, formatoceldacenter)
                ws.write(row_num, 6, u"%s" % matricula.inscripcion.persona.apellido2, formatoceldacenter)
                ws.write(row_num, 7, u"%s" % matricula.inscripcion.persona.nombres, formatoceldacenter)
                ws.write(row_num, 8,
                         u"%s" % matricula.inscripcion.persona.email if matricula.inscripcion.persona.email else "",
                         formatoceldacenter)
                ws.write(row_num, 9,
                         u"%s" % matricula.inscripcion.persona.emailinst if matricula.inscripcion.persona.emailinst else "",
                         formatoceldacenter)
                ws.write(row_num, 10,
                         u"%s" % matricula.inscripcion.persona.telefono if matricula.inscripcion.persona.telefono else "",
                         formatoceldacenter)
                ws.write(row_num, 11,
                         u"%s" % matricula.inscripcion.persona.telefono_conv if matricula.inscripcion.persona.telefono_conv else "",
                         formatoceldacenter)
                ws.write(row_num, 12,
                         u"%s" % matricula.inscripcion.persona.sexo.nombre if matricula.inscripcion.persona.sexo.nombre else "",
                         formatoceldacenter)
                ws.write(row_num, 13, u"%s" % raza, formatoceldacenter)
                ws.write(row_num, 14, u"%s" % "SI" if matricula.inscripcion.persona.lgtbi else "NO", formatoceldacenter)
                ws.write(row_num, 15, u"%s" % discapacidad, formatoceldacenter)
                ws.write(row_num, 16, u"%s" % tipodiscapacidad, formatoceldacenter)
                ws.write(row_num, 17, u"%s" % "SI" if matricula.inscripcion.persona.ppl else "NO", formatoceldacenter)
                ws.write(row_num, 18, u"%s" % matricula.inscripcion.modalidad, formatoceldacenter)
                ws.write(row_num, 19, u"%s" % "SI" if matricula.termino == True else "NO", formatoceldacenter)
                ws.write(row_num, 20, u"%s" % matricula.cantidad_materias(), formatoceldacenter)
                ws.write(row_num, 21, u"%s" % matricula.promedio_nota(), formatoceldacenter)
                ws.write(row_num, 22, u"%s" % "APROBADO" if matricula.materias_aprobadas_todas() else "REPROBADO",
                         formatoceldacenter)
                ws.write(row_num, 23, u"%s" % "Aprueba la Nivelación solo si en cada asignatura obtuvo como mínimo 70",
                         formatoceldacenter)
                ws.write(row_num, 24, u"%s" % matricula.inscripcion.estado_perdida_gratuidad(), formatoceldacenter)
                ws.write(row_num, 25, u"%s" % "1" if matricula.cuposenescyt else "2", formatoceldacenter)
                ws.write(row_num, 26, u"%s" % matricula.id, formatoceldacenter)
                ws.write(row_num, 27, u"%s" % matricula.inscripcion.id, formatoceldacenter)
                ws.write(row_num, 28, u"%s" % totalaprobadas, formatoceldacenter)

                materiasasignadas = matricula.materiaasignada_set.filter(status=True)
                aplica_orden = False
                malla = matricula.inscripcion.mi_malla()
                if malla:
                    cantidad_asignaturas_sin_ordenar = malla.asignaturamalla_set.filter(status=True).__len__()
                    cantidad_asignaturas_ordenas = malla.asignaturamalla_set.filter(status=True,
                                                                                    asignaturamallaorden__status=True).__len__()
                    aplica_orden = cantidad_asignaturas_ordenas == cantidad_asignaturas_sin_ordenar

                numcolu = 29
                numcolu1 = 30
                numcolu2 = 31
                numcolu3 = 32
                numcolu4 = 33
                for item in range(asignaturas_maxima):
                    eMateriaAsignada = materiasasignadas[item] if item < materiasasignadas.__len__() else None
                    if aplica_orden:
                        eMateriaAsignada = materiasasignadas.filter(
                            materia__asignaturamalla__asignaturamallaorden__orden=item + 1,
                            materia__asignaturamalla__asignaturamallaorden__status=True).distinct().first()
                    ws.write(row_num, numcolu,
                             u"%s" % eMateriaAsignada.materia.nombre_mostrar_alias() if eMateriaAsignada is not None else '',
                             formatoceldacenter)
                    ws.write(row_num, numcolu1,
                             u"%s" % eMateriaAsignada.materia.paralelo if eMateriaAsignada is not None else '',
                             formatoceldacenter)
                    ws.write(row_num, numcolu2,
                             u"%s" % eMateriaAsignada.notafinal if eMateriaAsignada is not None else '',
                             formatoceldacenter)
                    ws.write(row_num, numcolu3, u"%s" % eMateriaAsignada.id if eMateriaAsignada is not None else '',
                             formatoceldacenter)
                    ws.write(row_num, numcolu4,
                             u"%s" % eMateriaAsignada.materia.idcursomoodle if eMateriaAsignada else '',
                             formatoceldacenter)
                    numcolu += 5
                    numcolu1 += 5
                    numcolu2 += 5
                    numcolu3 += 5
                    numcolu4 += 5

                # materias = MateriaAsignada.objects.filter(matricula_id=matricula.id)
                # ordenMaterias = [None, None, None, None, None]
                # ordenMaterias[0] = materias.filter(status=True, materia__asignatura_id=4837).first()  # PROPEDÉUTICO
                # ordenMaterias[1] = materias.filter(status=True, materia__asignatura_id__in=[969, 2688]).first()  # MATEMÁTICAS, ANATOMÍA
                # ordenMaterias[2] = materias.filter(status=True, materia__asignatura_id__in=[2677, 171]).first()  # Biología, ICA
                # ordenMaterias[3] = materias.filter(status=True, materia__asignatura_id=4881).first()  # Pensamiento Computacional
                # ordenMaterias[4] = materias.filter(status=True, materia__asignatura_id=173).first()  # BIOQUÍMICA
                # numcolu = 29
                # numcolu1 = 30
                # numcolu2 = 31
                # numcolu3 = 32
                # numcolu4 = 33
                # for materia in ordenMaterias:
                #     ws.write(row_num, numcolu, u"%s" % materia.materia.nombre_mostrar_alias() if materia is not None else '', formatoceldacenter)
                #     ws.write(row_num, numcolu1, u"%s" % materia.materia.paralelo if materia is not None else '', formatoceldacenter)
                #     ws.write(row_num, numcolu2, u"%s" % materia.notafinal if materia is not None else '', formatoceldacenter)
                #     ws.write(row_num, numcolu3, u"%s" % materia.id if materia is not None else '', formatoceldacenter)
                #     ws.write(row_num, numcolu4, u"%s" % materia.materia.idcursomoodle if materia else '', formatoceldacenter)
                #     numcolu += 5
                #     numcolu1 += 5
                #     numcolu2 += 5
                #     numcolu3 += 5
                #     numcolu4 += 5
                row_num += 1
                print(f"{key + 1}.- ", matricula.inscripcion.persona, matricula.inscripcion.persona.cedula)
            workbook.close()

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel Estudiantes Admisión - aprobados y reprobados',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel Estudiantes Admisión - aprobados y reprobados',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_tituloinsigniamasivo_background(threading.Thread):

    def __init__(self, request, notiid, graduados, data, IS_DEBUG):
        self.request = request
        self.notiid = notiid
        self.graduados = graduados
        self.data = data
        self.IS_DEBUG = IS_DEBUG
        threading.Thread.__init__(self)

    def run(self):
        request, notiid, graduados, data, IS_DEBUG = self.request, self.notiid, self.graduados, self.data, self.IS_DEBUG
        try:
            # funcion
            dominio_sistema = 'http://127.0.0.1:8000'
            if not IS_DEBUG:
                dominio_sistema = 'https://sga.unemi.edu.ec'
            data["DOMINIO_DEL_SISTEMA"] = dominio_sistema

            lista_correctos = []
            lista_errores = []
            for graduado in graduados:
                with transaction.atomic():
                    try:
                        persona_cargo_tercernivel = None
                        cargo = None
                        tamano = 0
                        firmauno = None
                        if DistributivoPersona.objects.filter(denominacionpuesto=113, estadopuesto__id=1,
                                                              status=True).exists():
                            firmauno = \
                                DistributivoPersona.objects.filter(denominacionpuesto=113, estadopuesto__id=1,
                                                                   status=True)[
                                    0]
                        data['firmauno'] = firmauno
                        firmatres = None
                        # Cambio de Secretaria General Velasco Neira Stefania desde el 25 de oct 2022 hasta el 12 de feb 2023 subrogante y desde el 14 de feb 2023 como titular, 13 feb 2023 Diana Pincay
                        firmatres_tipocargo = None
                        if graduado.fecharefrendacion >= datetime(2022, 10, 25, 0, 0, 0).date() and not (
                                graduado.fecharefrendacion == datetime(2023, 2, 13, 0, 0, 0).date()):
                            if DistributivoPersona.objects.filter(denominacionpuesto=502, estadopuesto__id=1,
                                                                  status=True).exists():
                                firmatres = \
                                    DistributivoPersona.objects.filter(denominacionpuesto=502, estadopuesto__id=1,
                                                                       status=True)[0]
                            if graduado.fecharefrendacion <= datetime(2023, 2, 12, 0, 0, 0).date():
                                firmatres_tipocargo = '(S)'
                        else:
                            if DistributivoPersonaHistorial.objects.filter(persona__id=26999, denominacionpuesto=502,
                                                                           status=True).exists():
                                firmatres = \
                                    DistributivoPersonaHistorial.objects.filter(persona__id=26999,
                                                                                denominacionpuesto=502,
                                                                                status=True)[0]
                                firmatres_tipocargo = '(E)'
                        # Tercera firma
                        data['firmatres_rubrica'] = FirmaPersona.objects.filter(status=True,
                                                                                persona=firmatres.persona).first().firma
                        data['firmatres_tipocargo'] = firmatres_tipocargo
                        # Tercera firma
                        data['firmatres'] = firmatres
                        # data['logoaval'] = graduado.capeventoperiodo.archivo
                        # data['elabora_persona'] = persona

                        # firmacertificado = None
                        # if PersonaDepartamentoFirmas.objects.filter(status=True, departamento=158).exists():
                        #     firmacertificado = PersonaDepartamentoFirmas.objects.filter(status=True, departamento=158).order_by('-id').first()
                        # data['firmacertificado'] = firmacertificado
                        # data['firmaimg'] = FirmaPersona.objects.filter(status=True,
                        #                                                persona=firmacertificado.personadepartamento).last()

                        # inicio firma vicerrector/director
                        firma_departamento = PersonaDepartamentoFirmas.objects.get(tipopersonadepartamento_id=1,
                                                                                   departamentofirma_id=1, status=True,
                                                                                   actualidad=True)
                        listafirmaspersonadepartamento = PersonaDepartamentoFirmas.objects.filter(
                            tipopersonadepartamento_id=1,
                            departamentofirma_id=1,
                            status=True)
                        for firma in listafirmaspersonadepartamento:
                            if firma.fechafin is not None and firma.fechainicio is not None:
                                if graduado.fecharefrendacion <= firma.fechafin and graduado.fecharefrendacion >= firma.fechainicio:
                                    firma_departamento = firma
                        data['firmadirector'] = firma_departamento
                        data['imgfirmadirector'] = firma_departamento.personadepartamento.firmapersona_set.filter(
                            status=True).order_by('-tipofirma').first()
                        # fin firma vicerrector/directo
                        if PersonaDepartamentoFirmas.objects.filter(actualidad=True, status=True).exists():
                            firmaizquierda = PersonaDepartamentoFirmas.objects.get(actualidad=True, status=True,
                                                                                   tipopersonadepartamento_id=2,
                                                                                   departamentofirma_id=1)
                        # if PersonaDepartamentoFirmas.objects.values('id').filter(status=True,
                        #                                                          fechafin__gte=evento.fechafin,
                        #                                                          fechainicio__lte=evento.fechafin,
                        #                                                          tipopersonadepartamento_id=2,
                        #                                                          departamentofirma_id=1).exists():
                        #     firmaizquierda = PersonaDepartamentoFirmas.objects.get(status=True,
                        #                                                            fechafin__gte=evento.fechafin,
                        #                                                            fechainicio__lte=evento.fechafin,
                        #                                                            tipopersonadepartamento_id=2,
                        #                                                            departamentofirma_id=1)
                        #
                        data['firmaizquierda'] = firmaizquierda
                        data['firmaimgizq'] = FirmaPersona.objects.filter(status=True,
                                                                          persona=firmaizquierda.personadepartamento).last()
                        # if evento.envionotaemail:
                        #     data['nota'] = evento.instructor_principal().extaer_notatotal(graduado.id)
                        # if DistributivoPersona.objects.filter(persona_id=persona, estadopuesto__id=PUESTO_ACTIVO_ID,
                        #                                       status=True).exists():
                        #     cargo = \
                        #         DistributivoPersona.objects.filter(persona_id=persona,
                        #                                            estadopuesto__id=PUESTO_ACTIVO_ID,
                        #                                            status=True)[0]
                        data['persona_cargo'] = cargo
                        # data['persona_cargo_titulo'] = titulo = persona.titulacion_principal_senescyt_registro()
                        # if not titulo == '':
                        #     persona_cargo_tercernivel = \
                        #         persona.titulacion_set.filter(titulo__nivel=3).order_by('-fechaobtencion')[
                        #             0] if titulo.titulo.nivel_id == 4 else None
                        data['persona_cargo_tercernivel'] = persona_cargo_tercernivel
                        data['graduado'] = graduado
                        # data['title'] = ''
                        data['fecha'] = None
                        fechagraduado = graduado.fecharefrendacion if graduado.fecharefrendacion else None
                        if fechagraduado:
                            mes = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto",
                                   "septiembre",
                                   "octubre", "noviembre", "diciembre"]
                            if fechagraduado.day > 1:
                                cadena = u"a los %s días" % (fechagraduado.day)
                            else:
                                cadena = u"al primer día"
                            data['fecha'] = u"San Francisco de Milagro, " + cadena + " del mes de %s de %s." % (
                                str(mes[fechagraduado.month - 1]), fechagraduado.year)
                        data['fecha_graduado_insignia'] = fechagraduado

                        # data['listado_contenido'] = listado = evento.contenido.split("\n") if evento.contenido else []
                        # if evento.objetivo.__len__() < 290:
                        #     if listado.__len__() < 21:
                        #         tamano = 120
                        #     elif listado.__len__() < 35:
                        #         tamano = 100
                        #     elif listado.__len__() < 41:
                        #         tamano = 70
                        data['controlar_bajada_logo'] = tamano
                        qrname = 'qr_titulo_' + str(graduado.id)
                        # folder = SITE_STORAGE + 'media/qrcode/evaluaciondocente/'
                        folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'titulos', 'qr'))
                        directory = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'titulos'))
                        # folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente')
                        rutapdf = folder + qrname + '.pdf'
                        rutaimg = folder + qrname + '.png'
                        try:
                            os.stat(directory)
                        except:
                            os.mkdir(directory)
                        if os.path.isfile(rutapdf):
                            os.remove(rutaimg)
                            os.remove(rutapdf)
                        # generar nombre html y url html
                        if not graduado.namehtmltitulo:
                            htmlname = "%s%s" % (uuid.uuid4().hex, '.html')
                        else:
                            htmlname = graduado.namehtmltitulo
                        urlname = "/media/qrcode/titulos/%s" % htmlname
                        rutahtml = SITE_STORAGE + urlname
                        if os.path.isfile(rutahtml):
                            os.remove(rutahtml)
                        # generar nombre html y url html
                        data['version'] = version = datetime.now().strftime('%Y%m%d_%H%M%S%f')
                        # url = pyqrcode.create('https://sga.unemi.edu.ec/media/qrcode/titulos/' + qrname + '.pdf')
                        url = pyqrcode.create(f'{dominio_sistema}/media/qrcode/titulos/{htmlname}?v={version}')
                        # url = pyqrcode.create(dominio_sistema + '/media/qrcode/titulos/' + htmlname)
                        imageqr = url.png(folder + qrname + '.png', 16, '#000000')
                        data['qrname'] = 'qr' + qrname
                        data['urlhtmlinsignia'] = urlhtmlinsignia = dominio_sistema + urlname
                        # data['urlhtmlinsigniaversion'] = f'{urlhtmlinsignia}?v={version}'
                        data['posgrado'] = u'DIRECCIÓN DE POSGRADO'
                        valida = conviert_html_to_pdfsaveqrtitulo(
                            'graduados/titulo_formatonuevo_pdf.html',
                            {'pagesize': 'A4', 'data': data},
                            qrname + '.pdf'
                        )
                        if valida:
                            # elimino codigo qr despues de pegarlo en el titulo titulo_formatonuevo_pdf.html
                            if os.path.isfile(rutaimg):
                                os.remove(rutaimg)
                            # generar portada del certificado
                            jpgname = f'{qrname}'
                            rutajpg = f'{SITE_STORAGE}/media/qrcode/titulos/{jpgname}.jpg'
                            if os.path.isfile(rutajpg):
                                os.remove(f'{rutajpg}')
                            with open(f'{SITE_STORAGE}/media/qrcode/titulos/{qrname}.pdf', mode='rb') as pdf:
                                images = convert_from_bytes(pdf.read(),
                                                            output_folder=f'{SITE_STORAGE}/media/qrcode/titulos/',
                                                            # first_page = True,
                                                            poppler_path=SITE_POPPLER,
                                                            fmt="jpg",
                                                            single_file=True,
                                                            thread_count=1,
                                                            # size=(507, 335), # tamaño optimizado
                                                            size=(711, 519),  # tamaño inicial
                                                            output_file=f'{jpgname}'
                                                            )
                            data['url_jpg'] = dominio_sistema + f'/media/qrcode/titulos/{jpgname}.jpg'
                            # generar portada del certificado
                            graduado.rutapdftitulo = 'qrcode/titulos/' + qrname + '.pdf'
                            # graduado.emailnotificado = True
                            # graduado.fecha_emailnotifica = datetime.now().date()
                            # graduado.persona_emailnotifica = persona
                            # graduado.save(request)
                            data['rutapdf'] = '/media/{}'.format(graduado.rutapdftitulo)  # ojo
                            #
                            data['idinsignia'] = htmlname[0:len(htmlname) - 5]

                            # consultar record academico
                            data['inscripcion'] = inscripcion = graduado.inscripcion
                            data['records'] = inscripcion.recordacademico_set.filter(status=True).order_by(
                                'asignaturamalla__nivelmalla', 'asignatura', 'fecha')
                            data['total_creditos'] = inscripcion.total_creditos()
                            data['total_creditos_malla'] = inscripcion.total_creditos_malla()
                            data['total_creditos_modulos'] = inscripcion.total_creditos_modulos()
                            data['total_creditos_otros'] = inscripcion.total_creditos_otros()
                            data['total_horas'] = inscripcion.total_horas()
                            data['promedio'] = inscripcion.promedio_record()
                            # data['aprobadas'] = inscripcion.recordacademico_set.filter(aprobada=True, valida=True).count()
                            # data['reprobadas'] = inscripcion.recordacademico_set.filter(aprobada=False, valida=True).count()
                            # data['reporte_0'] = obtener_reporte("record_alumno")
                            # consultar record academico
                            # crear html de titulo valido en la media  y guardar url en base
                            a = render(request, "graduados/titulovalido.html",
                                       {"data": data, 'institucion': 'UNIVERSIDAD ESTATAL DE MILAGRO',
                                        "remotenameaddr": 'sga.unemi.edu.ec'})
                            with open(SITE_STORAGE + urlname, "wb") as f:
                                f.write(a.content)
                            f.close()
                            # elimino portada de titulo despues de añadirla en el insignia titulovalido.html
                            # if os.path.isfile(data['ruta_jpg']):
                            #     os.remove(data['ruta_jpg'])

                            graduado.namehtmltitulo = htmlname
                            graduado.urlhtmltitulo = urlname
                            graduado.estadonotificacion = 2
                            # fin crear html en la media y guardar url en base
                            graduado.save(request)
                            # envio por correo
                            # correograduado = graduado.inscripcion.persona.emailpersonal()
                            # if IS_DEBUG:
                            #     correograduado = ['marianamadeleineas@gmail.com']
                            # asunto = u"INSIGNIA - " + str(graduado.inscripcion.carrera)
                            # send_html_mail(asunto, "emails/notificar_tituloinsignia_posgrado.html",
                            #                {'sistema': request.session['nombresistema'], 'graduado': graduado,
                            #                 'director': firmacertificado, 'urlhtmlinsignia': urlhtmlinsignia},
                            #                correograduado,
                            #                [], [graduado.rutapdftitulo],
                            #                cuenta=CUENTAS_CORREOS[0][1])
                            if not IS_DEBUG:
                                time.sleep(5)
                            lista_correctos.append(f'{graduado.inscripcion.persona.cedula}[{graduado.id}]')
                    except Exception as ex:
                        transaction.set_rollback(True)
                        lista_errores.append(
                            f'{graduado.inscripcion.persona.cedula}[{graduado.id}] error {str(ex)} on line {str(sys.exc_info()[-1].tb_lineno)}')
                        print(
                            f'{graduado.inscripcion.persona.cedula}[{graduado.id}] error {ex} on line {sys.exc_info()[-1].tb_lineno}\n')
            print(f'Titulos/Insignias correctos: {lista_correctos} - Lista errores: {lista_errores}')
            # Mensaje errores/correctos
            mensajeerrores = lista_errores[0] if lista_errores else None
            for error in lista_errores[1:]:
                mensajeerrores += f", {error}\n"
            mensajecorrectos = lista_correctos[0] if lista_correctos else None
            for correcto in lista_correctos[1:]:
                mensajecorrectos += f", {correcto}\n"
            # Para masivo enviar notificacion al sga del administrativo
            if len(lista_errores) == 0:
                titulonotificacion = f"Se han generado los títulos/insignias posgrado masivo exitosamente"
                cuerponotificacion = f"Se generó exitosamente el proceso. \nTotal correctos: {str(len(lista_correctos))} de {str(len(graduados))} (total graduados)."
                # Generados correctamente: {str(mensajecorrectos)}.
            else:
                titulonotificacion = f"Se han generado los títulos/insignias posgrado masivo con errores"
                cuerponotificacion = f"¡Error! No se generaron {str(len(lista_errores))} título(s): {mensajeerrores}. \nTotal correctos: {str(len(lista_correctos))} de {str(len(graduados))} (total graduados). Generados correctamente: {str(mensajecorrectos)}."
            # funcion
            # usernotify = User.objects.get(pk=request.user.pk)
            # pers = Persona.objects.get(usuario=usernotify)
            # Notifica el resultado del proceso webpush como notificacion en el sga
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                # noti.titulo = titulonotificacion
                noti.cuerpo = cuerponotificacion
                noti.url = f"/graduados"
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerponotificacion,
                                    titulo=titulonotificacion,
                                    destinatario=request.session['persona'],
                                    url=f"/graduados",
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)
            send_user_notification(user=request.user, payload={
                "head": f"Generación terminada {'con errores.' if lista_errores else 'exitosamente.'}",
                "body": f"Generación de Títulos/Insignias posgrado masivo terminado {'con errores.' if lista_errores else 'exitosamente.'}.",
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": f"/graduados",
                "btn_notificaciones": traerNotificaciones(request, data, request.session['persona']).replace("\n",
                                                                                                             " ").replace(
                    '\\', "/").replace("//", "/"),
                "mensaje": f"Los Títulos/Insignias han sido generados {'con errores. Ver detalles en el apartado Ver mis notificaciones.' if lista_errores else 'exitosamente.'}."
            }, ttl=500)
        except Exception as ex:
            print(f'Error {str(ex)} on line {str(sys.exc_info()[-1].tb_lineno)}')


class reporte_matriculados_nivelacion_background(threading.Thread):

    def __init__(self, request, data, notif, periodo):
        self.request = request
        self.data = data
        self.notif = notif
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notif, periodo = self.request, self.data, self.notif, self.periodo
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_matriculados_nivelacion' + random.randint(1, 10000).__str__() + '.xlsx'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:

            __author__ = 'Unemi'
            ahora = datetime.now()
            time_codigo = ahora.strftime('%Y%m%d_%H%M%S')
            name_file = f'reporte_excel_pedidos_online_{time_codigo}.xlsx'
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet("Listado Matriculados Nivelación")

            fuentecabecera = workbook.add_format({
                'align': 'center',
                'bg_color': 'silver',
                'border': 1,
                'bold': 1
            })

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            fuenteencabezado = workbook.add_format({
                'align': 'center',
                'bg_color': '#1C3247',
                'font_color': 'white',
                'border': 1,
                'font_size': 24,
                'bold': 1
            })

            columnas = [
                (u'CODIGO MATRICULA', 20),
                (u'PARALELO', 20),
                (u'NIVEL NOMBRE', 20),
                (u'NIVEL', 20),
                (u'CARRERA', 20),
                (u'MODALIDAD', 20),
                (u'SESIÓN', 20),
                (u'CODIGO SENECYT', 20),
                (u'MATERIA', 20),
                (u'CODIGO INSCRIPCIÓN', 20),
                (u'CEDULA', 20),
                (u'PASAPORTE', 20),
                (u'NOMBRES', 40),
                (u'APELLIDOS', 40),
                (u'EMAIL PERSONAL', 40),
                (u'EMAIL INSTITUCIONAL', 40),
                (u"#CELULAR", 30),
                (u'LGBTI', 80),
                (u'ETNIA', 80),
                (u'DISCAPACIDAD', 80),
                (u"TIPO DISCAPACIDAD", 40),
                (u"NACIONALIDAD", 40),
                (u'PAIS RESIDENCIA', 20),
                (u'PROVINCIA RESIDENCIA', 20),
                (u'CANTON RESIDENCIA', 20),
                (u'DIRECCIÓN', 20),
                (u'PAIS NACIMIENTO', 20),
                (u"PPL", 10),
                (u'RELIGION', 20),
                (u'FECHA MATRICULA', 20),
                (u'FECHA CREACIÓN', 20),
                # (u'USUARIO CREACIÓN', 20)

            ]
            ws.merge_range(0, 0, 0, columnas.__len__() - 1, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, columnas.__len__() - 1,
                           f'LISTADO DE ESTUDIANTES MATRICULADOS EN NIVELACIÓN EN PERIODO {self.periodo}',
                           fuenteencabezado)
            row_num, numcolum = 2, 0
            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1
            row_num += 1

            eMatriculados = Matricula.objects.filter(nivel__periodo=self.periodo,
                                                     inscripcion__carrera__coordinacion__id=9,
                                                     status=True, termino=True).exclude(retiradomatricula=True)
            if DEBUG:
                eMatriculados = eMatriculados[:100]
            for key, matricula in enumerate(eMatriculados):
                ePerfilInscripcion = matricula.inscripcion.persona.perfilinscripcion_set.filter(status=True).first()
                tipodiscapacidad = None
                if ePerfilInscripcion is not None:
                    if ePerfilInscripcion.tienediscapacidad:
                        tipodiscapacidad = u'%s' % ePerfilInscripcion.tipodiscapacidad
                eReligion = None
                if PersonaReligion.objects.filter(persona=matricula.inscripcion.persona).exists():
                    eReligion = PersonaReligion.objects.filter(persona=matricula.inscripcion.persona)[0]
                materias = MateriaAsignada.objects.filter(matricula=matricula, status=True)
                for materia in materias:
                    qsbaseinscripcionmalla = matricula.inscripcion.inscripcionmalla_set.filter(status=True).first()
                    if qsbaseinscripcionmalla:
                        senecytcod = qsbaseinscripcionmalla.malla.codigo
                    else:
                        senecytcod = 'SIN CODIGO'
                    persona_ = matricula.inscripcion.persona
                    apellidos = matricula.inscripcion.persona.apellido1 + ' ' + matricula.inscripcion.persona.apellido2
                    ws.write(row_num, 0, u"%s" % materia.id, formatoceldacenter)
                    ws.write(row_num, 1, u"%s" % materia.materia.paralelo, formatoceldacenter)
                    ws.write(row_num, 2, u"%s" % matricula.nivelmalla.nombre, formatoceldacenter)
                    ws.write(row_num, 3, u"%s" % matricula.nivelmalla.orden, formatoceldacenter)
                    ws.write(row_num, 4, u"%s" % matricula.inscripcion.carrera.nombre, formatoceldacenter)
                    ws.write(row_num, 5, u"%s" % matricula.inscripcion.modalidad, formatoceldacenter)
                    ws.write(row_num, 6,
                             u"%s" % matricula.inscripcion.sesion.__str__() if matricula.inscripcion.sesion else '',
                             formatoceldacenter)
                    ws.write(row_num, 7, u"%s" % senecytcod, formatoceldacenter)
                    ws.write(row_num, 8, u"%s" % materia.materia.asignatura.nombre, formatoceldacenter)
                    ws.write(row_num, 9, u"%s" % matricula.inscripcion.id, formatoceldacenter)
                    ws.write(row_num, 10, u"%s" % matricula.inscripcion.persona.cedula, formatoceldacenter)
                    ws.write(row_num, 11, u"%s" % matricula.inscripcion.persona.pasaporte, formatoceldacenter)
                    ws.write(row_num, 12, u"%s" % matricula.inscripcion.persona.nombres, formatoceldacenter)
                    ws.write(row_num, 13, u"%s" % apellidos, formatoceldacenter)
                    ws.write(row_num, 14,
                             u"%s" % matricula.inscripcion.persona.email if matricula.inscripcion.persona.email else "",
                             formatoceldacenter)
                    ws.write(row_num, 15,
                             u"%s" % matricula.inscripcion.persona.emailinst if matricula.inscripcion.persona.emailinst else "",
                             formatoceldacenter)
                    ws.write(row_num, 16,
                             u"%s" % matricula.inscripcion.persona.telefono if matricula.inscripcion.persona.telefono else "",
                             formatoceldacenter)
                    ws.write(row_num, 17, u"%s" % 'SI' if matricula.inscripcion.persona.lgtbi else "NO",
                             formatoceldacenter)
                    ws.write(row_num, 18, u"%s" % persona_.mi_perfil().raza if persona_.mi_perfil() else "NO",
                             formatoceldacenter)
                    ws.write(row_num, 19, u"%s" % 'SI' if persona_.tiene_discapasidad() else "NO", formatoceldacenter)
                    tipodiscapacidad = ''
                    if persona_.tiene_discapasidad():
                        tipodiscapacidad = persona_.tiene_discapasidad().first().tipodiscapacidad.nombre
                    ws.write(row_num, 20, u"%s" % tipodiscapacidad, formatoceldacenter)
                    ws.write(row_num, 21, u"%s" % persona_.nacionalidad, formatoceldacenter)
                    ws.write(row_num, 22,
                             u"%s" % matricula.inscripcion.persona.pais.nombre if matricula.inscripcion.persona.pais else "",
                             formatoceldacenter)
                    ws.write(row_num, 23,
                             u"%s" % matricula.inscripcion.persona.provincia.nombre if matricula.inscripcion.persona.provincia else "",
                             formatoceldacenter)
                    ws.write(row_num, 24,
                             u"%s" % matricula.inscripcion.persona.canton.nombre if matricula.inscripcion.persona.canton else "",
                             formatoceldacenter)
                    ws.write(row_num, 25,
                             u"%s" % matricula.inscripcion.persona.direccion if matricula.inscripcion.persona.canton else "",
                             formatoceldacenter)
                    ws.write(row_num, 26,
                             u"%s" % matricula.inscripcion.persona.paisnacimiento.nombre if matricula.inscripcion.persona.paisnacimiento else "",
                             formatoceldacenter)
                    ws.write(row_num, 27, u"%s" % "SI" if matricula.inscripcion.persona.ppl else "NO",
                             formatoceldacenter)
                    ws.write(row_num, 28, u"%s" % eReligion.credo.nombre if eReligion else "Sin Religion",
                             formatoceldacenter)
                    ws.write(row_num, 29, u"%s" % str(matricula.fecha), formatoceldacenter)
                    ws.write(row_num, 30, u"%s" % str(matricula.fecha_creacion), formatoceldacenter)
                    # ws.write(row_num, 31,
                    #          u"%s" % matricula.usuario_creacion.username if matricula.usuario_creacion else '',
                    #          formatoceldacenter)
                    row_num += 1
                    print(f"{key + 1}.- ", matricula.inscripcion.persona, matricula.inscripcion.persona.cedula)
            workbook.close()

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel Matriculados Nivelación',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel Estudiantes Matriculados Nivelación',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)


        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_matriculados_pregrado_background(threading.Thread):

    def __init__(self, request, data, notif, periodo):
        self.request = request
        self.data = data
        self.notif = notif
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notif, periodo = self.request, self.data, self.notif, self.periodo
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_matriculados_pregrado' + random.randint(1, 10000).__str__() + '.xlsx'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            eMatriculados = Matricula.objects.filter(nivel__periodo=periodo, status=True).exclude(
                inscripcion__coordinacion_id=9)

            __author__ = 'Unemi'
            ahora = datetime.now()
            time_codigo = ahora.strftime('%Y%m%d_%H%M%S')
            name_file = f'reporte_excel_pedidos_online_{time_codigo}.xlsx'
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet("Listado Matriculados Pregrado")

            fuentecabecera = workbook.add_format({
                'align': 'center',
                'bg_color': 'silver',
                'border': 1,
                'bold': 1
            })

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            fuenteencabezado = workbook.add_format({
                'align': 'center',
                'bg_color': '#1C3247',
                'font_color': 'white',
                'border': 1,
                'font_size': 24,
                'bold': 1
            })

            columnas = [
                (u'FACULTAD', 80),
                (u'CARRERA', 80),
                (u"MODALIDAD", 40),
                (u'CEDULA', 20),
                (u'ESTUDIANTE', 40),
                (u'SEXO', 20),
                (u'EMAIL PERSONAL', 40),
                (u'EMAIL INSTITUCIONAL', 40),
                (u"#CELULAR", 30),
                (u'PAIS', 20),
                (u'PROVINCIA', 20),
                (u'CANTON', 20),
                (u"TIPO DISCAPACIDAD", 50),
                (u"PPL", 10),

            ]
            ws.merge_range(0, 0, 0, columnas.__len__() - 1, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, columnas.__len__() - 1,
                           f'LISTADO DE ESTUDIANTES MATRICULADOS EN PREGRADO EN PERIODO {self.periodo}',
                           fuenteencabezado)

            row_num, numcolum = 2, 0

            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1

            if DEBUG:
                eMatriculados = eMatriculados[:10]
            row_num = 3
            for key, matricula in enumerate(eMatriculados):

                ePerfilInscripcion = matricula.inscripcion.persona.perfilinscripcion_set.filter(status=True).first()
                nombrecompleto = matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2 + " " + matricula.inscripcion.persona.nombres

                tipodiscapacidad = None
                if ePerfilInscripcion is not None:
                    if ePerfilInscripcion.tienediscapacidad:
                        tipodiscapacidad = u'%s' % ePerfilInscripcion.tipodiscapacidad

                ws.write(row_num, 0, u"%s" % matricula.nivel.coordinacion().nombre, formatoceldacenter)
                ws.write(row_num, 1, u"%s" % matricula.inscripcion.carrera.nombre, formatoceldacenter)
                ws.write(row_num, 2, u"%s" % matricula.inscripcion.modalidad, formatoceldacenter)
                ws.write(row_num, 3, u"%s" % matricula.inscripcion.persona.cedula, formatoceldacenter)
                ws.write(row_num, 4, u"%s" % nombrecompleto, formatoceldacenter)
                ws.write(row_num, 5,
                         u"%s" % matricula.inscripcion.persona.sexo.nombre if matricula.inscripcion.persona.sexo else "",
                         formatoceldacenter)
                ws.write(row_num, 6,
                         u"%s" % matricula.inscripcion.persona.email if matricula.inscripcion.persona.email else "",
                         formatoceldacenter)
                ws.write(row_num, 7,
                         u"%s" % matricula.inscripcion.persona.emailinst if matricula.inscripcion.persona.emailinst else "",
                         formatoceldacenter)
                ws.write(row_num, 8,
                         u"%s" % matricula.inscripcion.persona.telefono if matricula.inscripcion.persona.telefono else "",
                         formatoceldacenter)
                ws.write(row_num, 9,
                         u"%s" % matricula.inscripcion.persona.pais.nombre if matricula.inscripcion.persona.pais else "",
                         formatoceldacenter)
                ws.write(row_num, 10,
                         u"%s" % matricula.inscripcion.persona.provincia.nombre if matricula.inscripcion.persona.pais else "",
                         formatoceldacenter)
                ws.write(row_num, 11,
                         u"%s" % matricula.inscripcion.persona.canton.nombre if matricula.inscripcion.persona.pais else "",
                         formatoceldacenter)
                ws.write(row_num, 12, u"%s" % tipodiscapacidad if tipodiscapacidad else "", formatoceldacenter)
                ws.write(row_num, 13, u"%s" % "SI" if matricula.inscripcion.persona.ppl else "NO", formatoceldacenter)
                row_num += 1
                print(f"{key + 1}.- ", matricula.inscripcion.persona, matricula.inscripcion.persona.cedula)
            workbook.close()

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel Matriculados Pregrado',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel Estudiantes Matriculados Pregrado',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)


        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class masivo_datos_sistema_gobierno_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            numeroarchivo = data['numeroarchivo']
            archivo = Archivo.objects.get(id=data['id_archivo'])
            workbook = load_workbook(archivo.archivo.file.name)
            sheet = workbook.active
            datos = sheet.iter_rows()
            pers = Persona.objects.get(usuario_id=request.user.pk)
            # sheet = workbook.sheet_by_index(0)
            fechahistorial = datetime.now()
            for departamento in Departamento.objects.all():
                departamento.integrantes.clear()
            for elemento in DistributivoPersona.objects.all():
                distributivopersonahistorial = DistributivoPersonaHistorial(persona=elemento.persona,
                                                                            regimenlaboral=elemento.regimenlaboral,
                                                                            nivelocupacional=elemento.nivelocupacional,
                                                                            modalidadlaboral=elemento.modalidadlaboral,
                                                                            partidaindividual=elemento.partidaindividual,
                                                                            estadopuesto=elemento.estadopuesto,
                                                                            grado=elemento.grado,
                                                                            rmuescala=elemento.rmuescala,
                                                                            rmupuesto=elemento.rmupuesto,
                                                                            rmusobrevalorado=elemento.rmusobrevalorado,
                                                                            escalaocupacional=elemento.escalaocupacional,
                                                                            rucpatronal=elemento.rucpatronal,
                                                                            codigosucursal=elemento.codigosucursal,
                                                                            tipoidentificacion=elemento.tipoidentificacion,
                                                                            provincia=elemento.provincia,
                                                                            canton=elemento.canton,
                                                                            denominacionpuesto=elemento.denominacionpuesto,
                                                                            puestoadicinal=elemento.puestoadicinal,
                                                                            unidadorganica=elemento.unidadorganica,
                                                                            aporteindividual=elemento.aporteindividual,
                                                                            aportepatronal=elemento.aportepatronal,
                                                                            estructuraprogramatica=elemento.estructuraprogramatica,
                                                                            comisioservicios=elemento.comisioservicios,
                                                                            numeroarchivo=elemento.numeroarchivo,
                                                                            fechahistorial=fechahistorial)
                distributivopersonahistorial.save(request)
            DistributivoPersona.objects.filter(status=True).exclude(modalidadlaboral_id=6).delete()
            linea = 1

            for rowx in datos:
                if linea >= 2:
                    # cols = sheet.row_values(rowx)
                    cols = [str(cell.value) for cell in rowx]
                    # REGIMEN LABORAL
                    regimenlaboral = None
                    if not RegimenLaboral.objects.filter(codigo=str(cols[1]).strip().upper()).exists():
                        regimenlaboral = RegimenLaboral(codigo=str(cols[1]).strip(),
                                                        descripcion=str(cols[2]).strip())
                        regimenlaboral.save(request)
                    else:
                        regimenlaboral = RegimenLaboral.objects.filter(codigo=str(cols[1]).strip().upper())[0]
                    # NIVEL OCUPACIONAL
                    nivelocupacional = None
                    if not NivelOcupacional.objects.filter(codigo=str(cols[3]).strip().upper()).exists():
                        nivelocupacional = NivelOcupacional(codigo=str(cols[3]).strip(),
                                                            descripcion=str(cols[4]).strip())
                        nivelocupacional.save(request)
                    else:
                        nivelocupacional = NivelOcupacional.objects.filter(codigo=str(cols[3]).strip().upper())[0]
                    # MODALIDAD LABORAL
                    modalidadlaboral = None
                    if not ModalidadLaboral.objects.filter(codigo=str(cols[7]).strip().upper()).exists():
                        modalidadlaboral = ModalidadLaboral(codigo=str(cols[7]).strip(),
                                                            descripcion=str(cols[8]).strip())
                        modalidadlaboral.save(request)
                    else:
                        modalidadlaboral = ModalidadLaboral.objects.filter(codigo=str(cols[7]).strip().upper())[0]
                    partidaindividual = int(cols[9])
                    # ESTADO PUESTO
                    estadopuesto = None
                    if not EstadoPuesto.objects.filter(descripcion=cols[10].strip().upper()).exists():
                        estadopuesto = EstadoPuesto(descripcion=cols[10].strip())
                        estadopuesto.save(request)
                    else:
                        estadopuesto = EstadoPuesto.objects.filter(descripcion=cols[10].strip().upper())[0]
                    grado = int(cols[11])
                    # REMUNERACION ESCALA
                    rmuescala = 0
                    try:
                        rmuescala = float(cols[12].replace(',', ''))
                    except:
                        pass
                    # REMUNERACION PUESTO
                    rmupuesto = 0
                    try:
                        rmupuesto = float(cols[13].replace(',', ''))
                    except:
                        pass
                    # REMUNERACION SOBREVALORADO
                    rmusobrevalorado = 0
                    try:
                        rmusobrevalorado = float(cols[14].replace(',', ''))
                    except:
                        pass
                    # ESCALA OCUPACIONAL
                    escalaocupacional = None
                    if not EscalaOcupacional.objects.filter(codigo=str(cols[15]).strip().upper()).exists():
                        escalaocupacional = EscalaOcupacional(codigo=str(cols[15]).strip(),
                                                              descripcion=str(cols[16]).strip())
                        escalaocupacional.save(request)
                    else:
                        escalaocupacional = EscalaOcupacional.objects.filter(codigo=str(cols[15]).strip().upper())[0]
                    # RUC PATRONAL
                    rucpatronal = cols[17].strip()
                    # CODIGO SUCURSAL
                    codigosucursal = cols[18].strip()
                    # TIPO IDENTIFICACION
                    if cols[19].strip() == u'CÉDULA':
                        tipoidentificacion = 1
                    elif cols[19].strip() == u'PASAPORTE':
                        tipoidentificacion = 2
                    else:
                        tipoidentificacion = 0
                    # DENOMINACION DE PUESTO
                    denominacionpuesto = None
                    if not DenominacionPuesto.objects.filter(codigo=str(cols[25]).strip().upper()).exists():
                        denominacionpuesto = DenominacionPuesto(codigo=str(cols[25]).strip(),
                                                                descripcion=str(cols[26]).strip())
                        denominacionpuesto.save(request)
                    else:
                        denominacionpuesto = DenominacionPuesto.objects.get(codigo=str(cols[25]).strip().upper())
                        denominacionpuesto.descripcion = str(cols[26]).strip()
                        denominacionpuesto.save(request)
                        # denominacionpuesto = DenominacionPuesto.objects.filter(codigo=str(cols[25]).strip().upper())[0]
                    crearprofesor = False
                    if 'PROFESOR' in denominacionpuesto.descripcion:
                        crearprofesor = True
                    # DENOMINACION DE PUESTO ADICIONAL
                    numeropuestoadicional = ''
                    puestoadicional = None
                    try:
                        numeropuestoadicional = str(int(cols[27].strip().upper()))
                    except:
                        pass
                    if numeropuestoadicional:
                        if not PuestoAdicional.objects.filter(codigo=str(numeropuestoadicional)).exists():
                            puestoadicional = PuestoAdicional(codigo=str(numeropuestoadicional),
                                                              descripcion=str(cols[28]).strip())
                            puestoadicional.save(request)
                        else:
                            puestoadicional = PuestoAdicional.objects.filter(codigo=str(numeropuestoadicional))[0]
                    # DEPARTAMENTOS
                    departamento = None
                    if Departamento.objects.filter(nombre=str(cols[30]).strip().upper()).exists():
                        departamento = Departamento.objects.filter(nombre=str(cols[30]).strip())[0]
                        departamento.codigo = str(cols[29]).strip()
                        departamento.save(request)
                    else:
                        departamento = Departamento(nombre=str(cols[30]).strip(),
                                                    codigo=str(cols[29]).strip())
                        departamento.save(request)
                    # APORTE INDIVIDUAL
                    aporteindividual = 0
                    try:
                        aporteindividual = float(cols[31])
                    except:
                        pass
                    # APORTE PATRONAL
                    aportepatronal = 0
                    try:
                        aportepatronal = float(cols[32])
                    except:
                        pass
                    # ESTRUCTURA PROGRAMATICA
                    estructuraprogramatica = None
                    if not EstructuraProgramatica.objects.filter(codigo=str(cols[33]).strip().upper()).exists():
                        estructuraprogramatica = EstructuraProgramatica(codigo=str(cols[33]).strip(),
                                                                        descripcion=cols[34].strip())
                        estructuraprogramatica.save(request)
                    else:
                        estructuraprogramatica = \
                            EstructuraProgramatica.objects.filter(codigo=str(cols[33]).strip().upper())[0]
                    # COMISION DE SERVICIO
                    comisioservicios = cols[35].strip().upper()
                    # PERSONA
                    identificacion = str(cols[20]).strip().upper()
                    personaadmin = None
                    provincia = None
                    canton = None
                    if len(identificacion) > 0:
                        # PROVINCIA
                        if Provincia.objects.filter(nombre=str(cols[22]).strip().upper()).exists():
                            provincia = Provincia.objects.filter(nombre=str(cols[22]).strip())[0]
                        else:
                            provincia = Provincia(nombre=str(cols[22]).strip())
                            provincia.save()
                        # CANTON
                        if Canton.objects.filter(nombre=str(cols[23]).strip().upper()).exists():
                            canton = Canton.objects.filter(nombre=str(cols[23]).strip().upper())[0]
                            canton.codigo = str(cols[24]).strip()
                            canton.save()
                        else:
                            canton = Canton(nombre=str(cols[23]).strip(),
                                            codigo=str(cols[24]).strip())
                            canton.save()
                        if not Persona.objects.filter(Q(cedula=identificacion) | Q(pasaporte=identificacion)).exists():
                            nombreimportacion = str(cols[21]).strip()
                            if len(nombreimportacion.split(' ')) >= 3:
                                importacionnombres = ''
                                for x in nombreimportacion.split(' ')[2:]:
                                    importacionnombres += x + ' '
                                importacionnombres = importacionnombres.strip()
                                importacionapellido1 = nombreimportacion.split(' ')[0]
                                importacionapellido2 = nombreimportacion.split(' ')[1]
                            else:
                                importacionnombres = nombreimportacion.split(' ')[1]
                                importacionapellido1 = nombreimportacion.split(' ')[0]
                                importacionapellido2 = ''
                            personaadmin = Persona(nombres=importacionnombres,
                                                   apellido1=importacionapellido1,
                                                   apellido2=importacionapellido2,
                                                   cedula=identificacion if tipoidentificacion == 1 else '',
                                                   pasaporte=identificacion if tipoidentificacion == 2 else '',
                                                   nacimiento=datetime.now().date(),
                                                   sexo_id=SEXO_MASCULINO,
                                                   provincia=provincia,
                                                   canton=canton)
                            personaadmin.save()
                            administrativo = Administrativo(persona=personaadmin,
                                                            contrato='',
                                                            fechaingreso=datetime.now().date(),
                                                            activo=True)
                            administrativo.save()
                            username = calculate_username(personaadmin)
                            generar_usuario(personaadmin, username, variable_valor('ADMINISTRATIVOS_GROUP_ID'))
                            if EMAIL_INSTITUCIONAL_AUTOMATICO:
                                personaadmin.emailinst = username + '@' + EMAIL_DOMAIN
                            personaadmin.save()
                            personaadmin.crear_perfil(administrativo=administrativo)
                            personaadmin.mi_ficha()
                            personaadmin.mi_perfil()
                            personaadmin.datos_extension()
                            personaadmin.creacion_persona(request.session['nombresistema'], pers)
                        else:
                            personaadmin = \
                                Persona.objects.filter(Q(cedula=identificacion) | Q(pasaporte=identificacion))[0]
                            g = Group.objects.get(pk=variable_valor('ADMINISTRATIVOS_GROUP_ID'))
                            if not personaadmin.usuario:
                                username = calculate_username(personaadmin)
                                generar_usuario(personaadmin, username, g.id)
                            g.user_set.add(personaadmin.usuario)
                            g.save()
                        # CREA EL PROFESOR
                        if crearprofesor and not personaadmin.es_profesor():
                            profesor = Profesor(persona=personaadmin,
                                                activo=True,
                                                fechaingreso=datetime.now().date(),
                                                coordinacion=Coordinacion.objects.all()[0],
                                                dedicacion=TiempoDedicacionDocente.objects.all()[0])
                            profesor.save()
                            grupo = Group.objects.get(pk=PROFESORES_GROUP_ID)
                            grupo.user_set.add(personaadmin.usuario)
                            grupo.save()
                            personaadmin.crear_perfil(profesor=profesor)
                    distributivopersona = DistributivoPersona(persona=personaadmin,
                                                              regimenlaboral=regimenlaboral,
                                                              nivelocupacional=nivelocupacional,
                                                              modalidadlaboral=modalidadlaboral,
                                                              partidaindividual=partidaindividual,
                                                              estadopuesto=estadopuesto,
                                                              grado=grado,
                                                              rmuescala=rmuescala,
                                                              rmupuesto=rmupuesto,
                                                              rmusobrevalorado=rmusobrevalorado,
                                                              escalaocupacional=escalaocupacional,
                                                              rucpatronal=rucpatronal,
                                                              codigosucursal=codigosucursal,
                                                              tipoidentificacion=tipoidentificacion,
                                                              provincia=provincia,
                                                              canton=canton,
                                                              denominacionpuesto=denominacionpuesto,
                                                              puestoadicinal=puestoadicional,
                                                              unidadorganica=departamento,
                                                              aporteindividual=aporteindividual,
                                                              aportepatronal=aportepatronal,
                                                              estructuraprogramatica=estructuraprogramatica,
                                                              comisioservicios=comisioservicios,
                                                              numeroarchivo=numeroarchivo)
                    distributivopersona.save(request)
                    if personaadmin:
                        historial = personaadmin.distributivopersonahistorial_set.filter(fechahistorial=fechahistorial)
                        if historial:
                            lista = ['tic@unemi.edu.ec']

                            if historial[0].denominacionpuesto.id != distributivopersona.denominacionpuesto.id:
                                send_html_mail("Modificación de Persona", "emails/modificaciondenominacion.html",
                                               {'sistema': request.session['nombresistema'], 't': miinstitucion(),
                                                'd': personaadmin, 'persona': request.session['persona'],
                                                'denominacionantiguo': historial[0].denominacionpuesto.descripcion,
                                                'denominacionactual': distributivopersona.denominacionpuesto.descripcion},
                                               lista,
                                               [], cuenta=CUENTAS_CORREOS[4][1])
                            else:
                                if historial[0].regimenlaboral.id != distributivopersona.regimenlaboral.id:
                                    send_html_mail("Modificación de Persona", "emails/modificacionregimen.html",
                                                   {'sistema': request.session['nombresistema'], 't': miinstitucion(),
                                                    'd': personaadmin, 'persona': request.session['persona'],
                                                    'regimenantiguo': historial[0].regimenlaboral.descripcion,
                                                    'regimenactual': distributivopersona.regimenlaboral.descripcion},
                                                   lista, [],
                                                   cuenta=CUENTAS_CORREOS[4][1])
                                else:
                                    if historial[0].estadopuesto.id != distributivopersona.estadopuesto.id:
                                        send_html_mail("Modificación de Persona", "emails/modificacionestado.html",
                                                       {'sistema': request.session['nombresistema'],
                                                        't': miinstitucion(), 'd': personaadmin,
                                                        'persona': request.session['persona'],
                                                        'estadoantiguo': historial[0].estadopuesto.descripcion,
                                                        'estadoactual': distributivopersona.estadopuesto.descripcion},
                                                       lista, [],
                                                       cuenta=CUENTAS_CORREOS[4][1])

                        if len(identificacion) > 0:
                            departamento.integrantes.add(personaadmin)
                    # print(f'Fila {linea} de {sheet.max_row-1} procesada')
                linea += 1
            titulo = 'Importación finalizada.'
            cuerpo = 'Se guardaron exitosamente los registros cargados.'
            mensaje = f'Los registros cargados de {data["namefile"].lower()} se importaron correctamente al sistema.'
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = f"{request.path}"
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=f"{request.path}",
                                    prioridad=1, app_label='SAGEST',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": f"{request.path}",
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": mensaje
            }, ttl=500)
            lista_email = pers.lista_emails()
            datos_email = {'sistema': request.session['nombresistema'],
                           'fecha': datetime.now().date(),
                           'hora': datetime.now().time(),
                           'persona': pers,
                           'titulo': titulo,
                           'total': sheet.max_row - 1,
                           'url_': f'{request.path}',
                           'mensaje': mensaje}
            template = "emails/notificar_importacionregistros.html"
            send_html_mail(titulo, template, datos_email, lista_email, [], [],
                           cuenta=CUENTAS_CORREOS[0][1])
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_acceso_crai_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'crai')
        valido = True
        nombre_archivo = 'reporte_acceso_crai_' + str(random.randint(1, 10000)) + '.pdf'
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        try:
            pers = Persona.objects.get(usuario_id=request.user.pk)
            formato_date = "%Y-%m-%d"
            tipo, formato, servicio, finicio, ffin, filtros, tiporeporte = int(request.GET.get('tipo', 0)), \
                                                                           int(request.GET.get('formato', 1)), \
                                                                           int(request.GET.get('servicio', 0)), \
                                                                           request.GET.get('finicio', ''), \
                                                                           request.GET.get('ffin', ''), Q(status=True), \
                                                                           int(request.GET.get('tiporeporte', 1))
            if finicio:
                finicio = datetime.strptime(finicio, formato_date).date()
            if ffin:
                ffin = datetime.strptime(ffin, formato_date).date()
            if tipo > 0:
                if tipo == 1:
                    filtros = Q(status=True, regimenlaboral_id__isnull=True, inscripcion_id__isnull=False)
                elif tipo == 2:
                    filtros = Q(status=True, regimenlaboral_id=2, persona__isnull=False)
                elif tipo == 3:
                    filtros = Q(status=True, regimenlaboral_id=1, persona__isnull=False)
                elif tipo == 4:
                    filtros = Q(status=True, regimenlaboral_id=4, persona__isnull=False)
                elif tipo == 5:
                    filtros = Q(status=True, inscripcion__isnull=True, regimenlaboral_id__isnull=True)
            if servicio > 0:
                filtros = filtros & Q(tiposerviciocrai_id=servicio)
            if finicio:
                filtros = filtros & Q(fecha__gte=finicio)
            if ffin:
                filtros = filtros & Q(fecha__lte=ffin)

            visitas = RegistrarIngresoCrai.objects.filter(filtros).order_by('persona__apellido1')

            if formato == 1:
                data['total'] = visitas.count()
                data['desde'] = finicio
                data['hasta'] = ffin
                data['fechahoy'] = datetime.now().date()
                data['listado'] = visitas
                if tiporeporte == 1:
                    data['tipo'] = tipo
                    data['fechahoy'] = datetime.now().date()
                    valido = conviert_html_to_pdfsave_generic_lotes(
                        request,
                        'adm_crai/reportepdf.html',
                        {
                            'pagesize': 'A4 landscape',
                            'data': data,
                        },
                        directory, nombre_archivo
                    )
                else:
                    if tipo == 1:
                        data['servicio'] = servicio
                        template = 'adm_crai/reporte_estudiantes_pdf.html'
                    elif tipo == 5:
                        template = 'adm_crai/reporte_externos_pdf.html'
                    valido = conviert_html_to_pdfsave_generic_lotes(
                        request,
                        template,
                        {
                            'pagesize': 'A4 landscape',
                            'data': data,
                        },
                        directory,
                        nombre_archivo
                    )


            elif formato == 2:
                finicio, ffin = str(finicio.strftime("%d-%m-%Y")), str(ffin.strftime("%d-%m-%Y"))
                nombre_archivo = 'reporte_acceso_crai' + random.randint(1, 10000).__str__() + '.xls'
                directory = os.path.join(MEDIA_ROOT, 'reportes', 'crai', nombre_archivo)
                __author__ = 'Unemi'
                styrowD = easyxf(
                    'font: name Times New Roman, color-index black, bold off; borders: left thin, right thin, top thin, bottom thin')
                styrow = easyxf(
                    'font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                style_col = easyxf(
                    'font: name Times New Roman, color-index black, bold on; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                title = easyxf(
                    'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                title1 = easyxf(
                    'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                                    num_format_str='yy/mm/dd')
                style_hr = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                                  num_format_str='h:mm')

                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('Reporte Visita')
                ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write_merge(3, 3, 1, 1, 'PERIODO:  ', style_sb1)
                ws.write_merge(3, 3, 2, 2, 'DESDE  ' + finicio + ' HASTA  ' + ffin, style_sb)
                response = HttpResponse(content_type="application/ms-excel")
                if tiporeporte == 1:
                    ws.write_merge(1, 1, 0, 9, 'REPORTE DE NÚMERO DE USUARIOS ATENDIDOS', title1)
                    response[
                        'Content-Disposition'] = 'attachment; filename=ReporteVisitaEstudiante ' + random.randint(
                        1, 10000).__str__() + '.xls'

                    columns = [
                        (u"N°", 1100),
                        (u"FECHA", 4500),
                        (u"HORA INGRESO", 4500),
                        (u"HORA ESTIMADA DE SALIDA", 4500),
                        (u"HORA SALIDA", 4500),
                        (u"APELLIDOS Y NOMBRES", 10000),
                        (u"TIPO DE USUARIO", 10000),
                        (u"SERVICIO", 5000),
                        (u"ACTIVIDAD", 7000),
                        (u"N° IDENTIFICACIÓN", 4500),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]
                    # persona__estudioinscripcion__isnull=False
                    row_num = 6
                    cont = 1
                    for vis in visitas:
                        persona = vis.inscripcion.persona if vis.inscripcion else vis.persona
                        tipousuario = 'Externo'
                        if vis.inscripcion:
                            tipousuario = 'Estudiante'
                        elif vis.regimenlaboral:
                            if vis.regimenlaboral.id == 1:
                                tipousuario = 'Administrativo'
                            elif vis.regimenlaboral.id == 2:
                                tipousuario = 'Docente'
                            else:
                                tipousuario = 'Trabajador'
                        ws.write(row_num, 0, cont, styrow)
                        ws.write(row_num, 1, vis.fecha, style_date)
                        ws.write(row_num, 2, vis.horainicio, style_hr)
                        ws.write(row_num, 3, vis.horasalida, style_hr)
                        ws.write(row_num, 4, vis.horafin, style_hr)
                        ws.write(row_num, 5, str(persona.nombre_completo_inverso()), styrowD)
                        ws.write(row_num, 6, tipousuario, styrowD)
                        ws.write(row_num, 7, str(vis.tiposerviciocrai.descripcion), styrow)
                        ws.write(row_num, 8, str(vis.actividad), styrow)
                        ws.write(row_num, 9, persona.cedula, styrowD)
                        row_num += 1
                        cont += 1

                    row_num += 1
                    ws.write(row_num, 8, 'TOTAL', styrow)
                    ws.write(row_num, 9, len(visitas), styrow)
                    wb.save(directory)
                else:
                    if tipo == 1:
                        ws.write_merge(1, 1, 0, 9, 'REPORTE DE ESTUDIANTES ATENDIDOS', title1)
                        response[
                            'Content-Disposition'] = 'attachment; filename=ReporteVisitaEstudiante ' + random.randint(
                            1, 10000).__str__() + '.xls'

                        columns = [
                            (u"N°", 1100),
                            (u"N° CÉDULA", 4500),
                            (u"APELLIDOS Y NOMBRES", 10000),
                            (u"FACULTAD", 10000),
                            (u"CARRERA", 10000),
                            (u"FECHA", 4500),
                            (u"HORA INGRESO", 4500),
                            (u"HORA SALIDA", 4500),
                            (u"SERVICIO", 5000),
                            (u"LIBRO", 10000),
                            (u"ACTIVIDAD", 7000),
                        ]
                        row_num = 5
                        for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], style_col)
                            ws.col(col_num).width = columns[col_num][1]
                        # persona__estudioinscripcion__isnull=False
                        row_num = 6
                        cont = 1

                        for vis in visitas:
                            ws.write(row_num, 0, cont, styrow)
                            ws.write(row_num, 1, vis.inscripcion.persona.cedula, styrowD)
                            ws.write(row_num, 2, str(vis.inscripcion.persona.nombre_completo_inverso()), styrowD)
                            ws.write(row_num, 3, str(vis.inscripcion.coordinacion.nombre), styrow)
                            ws.write(row_num, 4, str(vis.inscripcion.carrera.nombre), styrow)
                            ws.write(row_num, 5, vis.fecha, style_date)
                            ws.write(row_num, 6, vis.horainicio, style_hr)
                            ws.write(row_num, 7, vis.horafin, style_hr)
                            ws.write(row_num, 8, str(vis.tiposerviciocrai.descripcion), styrow)
                            ws.write(row_num, 9, str(vis.librokoha.nombre) if vis.librokoha else ' ', styrow)
                            ws.write(row_num, 10, str(vis.actividad), styrow)
                            row_num += 1
                            cont += 1

                        row_num += 1
                        ws.write(row_num, 8, 'TOTAL', styrow)
                        ws.write(row_num, 9, len(visitas), styrow)
                        wb.save(directory)
                    elif tipo == 5:
                        ws.write_merge(1, 1, 0, 7, 'REPORTE DE ATENCIONES A USUARIOS EXTERNOS', title1)
                        response[
                            'Content-Disposition'] = 'attachment; filename=ReporteVisitaExternos ' + random.randint(1,
                                                                                                                    10000).__str__() + '.xls'

                        columns = [
                            (u"N°", 1100),
                            (u"N° CÉDULA", 4500),
                            (u"APELLIDOS Y NOMBRES", 10000),
                            (u"FECHA", 4500),
                            (u"HORA INGRESO", 4500),
                            (u"HORA SALIDA", 4500),
                            (u"SERVICIO", 5000),
                            (u"ACTIVIDAD", 5000),
                        ]
                        row_num = 5
                        for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], style_col)
                            ws.col(col_num).width = columns[col_num][1]

                        row_num = 6
                        cont = 1

                        for vis in visitas:
                            ws.write(row_num, 0, cont, styrow)
                            ws.write(row_num, 1, vis.persona.cedula, styrowD)
                            ws.write(row_num, 2, str(vis.persona.nombre_completo_inverso()), styrowD)
                            ws.write(row_num, 3, vis.fecha, style_date)
                            ws.write(row_num, 4, vis.horainicio, style_hr)
                            ws.write(row_num, 5, vis.horafin, style_hr)
                            ws.write(row_num, 6, str(vis.tiposerviciocrai.descripcion), styrow)
                            ws.write(row_num, 7, str(vis.actividad), styrow)
                            row_num += 1
                            cont += 1

                        row_num += 1
                        ws.write(row_num, 6, 'TOTAL', styrow)
                        ws.write(row_num, 7, len(visitas), styrow)
                        wb.save(directory)

            # NOTIFICACIÓN
            titulo = 'Reporte de acceso al crai generada exitosamente'
            cuerpo = 'Su reporte de acceso al crai se genero correctamente y puede ser descargada por la url proporcionada.'
            if not valido:
                titulo = 'Algo salio mal, el reporte de acceso al crai no se genero '
                cuerpo = 'Su reporte de acceso al crai no se genero correctamente.'

            url = "{}reportes/crai/{}".format(MEDIA_URL, nombre_archivo)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_cartera_vencida_general_rubro_version_final_background(threading.Thread):

    def __init__(self, request, data, idnotificacion, fechacorte):
        self.request = request
        self.data = data
        self.idnotificacion = idnotificacion
        self.fechacorte = fechacorte
        threading.Thread.__init__(self)

    def run(self):
        try:
            request, data, idnotificacion, fechacorte = self.request, self.data, self.idnotificacion, self.fechacorte
            usernotify = User.objects.get(pk=request.user.pk)
            personanotifica = Persona.objects.get(usuario=usernotify)
            tituloreporte = "Reporte Excel Cartera Vencida General - Detalle Rubros"

            directorio = os.path.join(os.path.join(MEDIA_ROOT, 'reportes', 'posgrado'))
            try:
                os.stat(directorio)
            except:
                os.mkdir(directorio)

            nombrearchivo = "CARTERA_VENCIDA_GENERAL_RUBROS_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"
            urlarchivo = MEDIA_URL + "reportes/posgrado/" + nombrearchivo

            # Crea un nuevo archivo de excel y le agrega una hoja
            workbook = xlsxwriter.Workbook(directorio + '/' + nombrearchivo)
            hojadestino = workbook.add_worksheet("Reporte")

            fcabeceracolumna = workbook.add_format(FORMATOS_CELDAS_EXCEL["cabeceracolumna"])
            fceldageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneral"])
            fceldageneralcent = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneralcent"])
            ftitulo2izq = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo2izq"])
            ftitulo3izq = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo3izq"])
            fceldafecha = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdafecha"])
            fceldamoneda = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdamoneda"])
            fceldamonedapie = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdamonedapie"])
            fceldanegritaizq = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritaizq"])
            fceldanegritageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritageneral"])

            hojadestino.merge_range(0, 0, 0, 15, 'UNIVERSIDAD ESTATAL DE MILAGRO', ftitulo2izq)
            hojadestino.merge_range(1, 0, 1, 15, 'DIRECCIÓN ADMINISTRATIVA Y FINANCIERA', ftitulo2izq)
            hojadestino.merge_range(2, 0, 2, 15, 'CARTERA VENCIDA GENERAL - DETALLE RUBROS', ftitulo2izq)
            hojadestino.merge_range(3, 0, 3, 15, 'FECHA DE CORTE: ' + str(fechacorte) + '', ftitulo2izq)
            hojadestino.merge_range(4, 0, 4, 15, 'FECHA DE DESCARGA DEL REPORTE (' + str(datetime.now().date()) + ')',
                                    ftitulo2izq)

            columnas = [
                (u"N°", 3),
                (u"PROGRAMA DE MAESTRÍA", 30),
                (u"COHORTE", 11),
                (u"PERIODO (INICIO-FIN)", 20),
                (u"ESTADO DE MAESTRÍA", 15),
                (u"PROVINCIA", 15),
                (u"CANTÓN", 15),
                (u"CÉDULA", 15),
                (u"ESTUDIANTE", 38),
                (u"N° CUOTAS VENCIDAS", 15),
                (u"ID RUBRO", 15),
                (u"FECHA VENCIMIENTO", 15),
                (u"DÍAS VENCIMIENTO", 15),
                (u"TOTAL PAGADO", 15),
                (u"VALOR VENCIDO", 15),
                (u"VALOR PENDIENTE", 15),
                (u"CATEGORÍA", 15),
                (u"RANGO DÍAS", 15),
                (u"ESTADO ESTUDIANTE", 15),
                (u"FECHA ÚLTIMA CUOTA", 15),
                (u"ESTADO FINANCIAMIENTO", 15)
            ]

            fila = 6
            for col_num in range(len(columnas)):
                hojadestino.write(fila, col_num, columnas[col_num][0], fcabeceracolumna)
                hojadestino.set_column(col_num, col_num, columnas[col_num][1])

            matriculas = Matricula.objects.filter(status=True, nivel__periodo__tipo__id__in=[3, 4]
                                                  # inscripcion__persona__cedula__in=cedulas,
                                                  # inscripcion__carrera__id__in=[173],
                                                  # nivel__periodo__id__in=*[143]
                                                  ).exclude(nivel__periodo__pk__in=[120, 128]).distinct().order_by(
                'inscripcion__persona__apellido1',
                'inscripcion__persona__apellido2',
                'inscripcion__persona__nombres')  # [:10]
            totalmatriculas = matriculas.count()

            secuencia = 0
            registros = 0
            totalvencido = 0
            totalpendiente = 0
            totalpagado = 0

            programas = []

            resumengeneral = {
                0: {'etiqueta': 'CARTERA VIGENTE', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0),
                    'pendiente': Decimal(0), 'antiguedad': ''},
                1: {'etiqueta': '1-30 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0),
                    'pendiente': Decimal(0), 'antiguedad': 'A'},
                2: {'etiqueta': '31-60 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0),
                    'pendiente': Decimal(0), 'antiguedad': 'B'},
                3: {'etiqueta': '61-90 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0),
                    'pendiente': Decimal(0), 'antiguedad': 'C'},
                4: {'etiqueta': '91-180 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0),
                    'pendiente': Decimal(0), 'antiguedad': 'D'},
                5: {'etiqueta': '181 DÍAS EN ADELANTE', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0),
                    'pendiente': Decimal(0), 'antiguedad': 'E'}
            }

            for matricula in matriculas:
                secuencia += 1

                alumno = matricula.inscripcion.persona.nombre_completo_inverso()
                personaalumno = matricula.inscripcion.persona
                rubrosalumno = matricula.rubros_maestria()
                ultimafechavence = ""

                # Verifico si el programa y cohorte existen en la lista de resumen
                existe = False
                indice = j = 0

                idprograma = matricula.inscripcion.carrera.id
                nombreprograma = matricula.inscripcion.carrera.nombre
                idperiodo = matricula.nivel.periodo.id
                numerocohorte = matricula.nivel.periodo.cohorte if matricula.nivel.periodo.cohorte else 0
                fechasperiodo = str(matricula.nivel.periodo.inicio) + " a " + str(matricula.nivel.periodo.fin)
                estadoprograma = "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO"

                for datoprograma in programas:
                    if datoprograma[0] == idprograma and datoprograma[1] == idperiodo:
                        # indice = j
                        existe = True
                        break

                    j += 1

                indice = j

                if not existe:
                    # Agrego el programa y cohorte a la lista de programas
                    resumen = {0: {'etiqueta': 'CARTERA VIGENTE', 'estudiantes': 0, 'pagado': Decimal(0),
                                   'vencido': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': ''},
                               1: {'etiqueta': '1-30 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0),
                                   'vencido': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'A'},
                               2: {'etiqueta': '31-60 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0),
                                   'vencido': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'B'},
                               3: {'etiqueta': '61-90 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0),
                                   'vencido': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'C'},
                               4: {'etiqueta': '91-180 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0),
                                   'vencido': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'D'},
                               5: {'etiqueta': '181 DÍAS EN ADELANTE', 'estudiantes': 0, 'pagado': Decimal(0),
                                   'vencido': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'E'}
                               }
                    datoprograma = [idprograma, idperiodo, nombreprograma, numerocohorte, fechasperiodo, estadoprograma,
                                    resumen]
                    programas.append(datoprograma)

                datos = matricula.rubros_maestria_vencidos_detalle_version_final(fechacorte)

                # No es retirado INICIO
                if not matricula.retirado_programa_maestria():
                    # Si hay rubros no vencidos: INICIO
                    if datos['rubrosnovencidos']:
                        for rubro_no_vencido in datos['rubrosnovencidos']:
                            fila += 1

                            # print(fila)

                            registros += 1

                            codigorubro = rubro_no_vencido[0]
                            valorpagado = rubro_no_vencido[3]
                            valorpendiente = rubro_no_vencido[4]
                            valorvencido = rubro_no_vencido[5]
                            fechavence = rubro_no_vencido[1]
                            diasvencidos = rubro_no_vencido[6]
                            pagosvencidos = 1 if valorvencido > 0 else 0

                            programas[indice][6][0]['estudiantes'] += 1
                            programas[indice][6][0]['pagado'] += valorpagado
                            programas[indice][6][0]['vencido'] += valorvencido
                            programas[indice][6][0]['pendiente'] += valorpendiente

                            resumengeneral[0]['estudiantes'] += 1
                            resumengeneral[0]['pagado'] += valorpagado
                            resumengeneral[0]['vencido'] += valorvencido
                            resumengeneral[0]['pendiente'] += valorpendiente

                            categoriaantiguedad = "VIGENTE"
                            rangodias = ""

                            hojadestino.write(fila, 0, registros, fceldageneral)
                            hojadestino.write(fila, 1, nombreprograma, fceldageneral)
                            hojadestino.write(fila, 2, numerocohorte, fceldageneralcent)
                            hojadestino.write(fila, 3, str(matricula.nivel.periodo.inicio) + " a " + str(
                                matricula.nivel.periodo.fin), fceldageneralcent)
                            hojadestino.write(fila, 4,
                                              "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO",
                                              fceldageneral)
                            hojadestino.write(fila, 5,
                                              personaalumno.provincia.nombre if personaalumno.provincia else '',
                                              fceldageneral)
                            hojadestino.write(fila, 6, personaalumno.canton.nombre if personaalumno.canton else '',
                                              fceldageneral)
                            hojadestino.write(fila, 7, personaalumno.identificacion(), fceldageneral)
                            hojadestino.write(fila, 8, alumno, fceldageneral)
                            hojadestino.write(fila, 9, pagosvencidos, fceldageneral)
                            hojadestino.write(fila, 10, codigorubro, fceldageneral)
                            hojadestino.write(fila, 11, fechavence, fceldafecha)
                            hojadestino.write(fila, 12, diasvencidos, fceldageneral)
                            hojadestino.write(fila, 13, valorpagado, fceldamoneda)
                            hojadestino.write(fila, 14, valorvencido, fceldamoneda)
                            hojadestino.write(fila, 15, valorpendiente, fceldamoneda)
                            hojadestino.write(fila, 16, categoriaantiguedad, fceldageneral)
                            hojadestino.write(fila, 17, rangodias, fceldageneral)
                            hojadestino.write(fila, 18, matricula.estado_inscripcion_maestria(), fceldageneral)

                            if rubrosalumno:
                                ultimafechavence = rubrosalumno.last().fechavence

                                if matricula.tiene_refinanciamiento_deuda_posgrado():
                                    estadofinanciamiento = "REFINANCIAMIENTO"
                                elif matricula.tiene_coactiva_posgrado():
                                    estadofinanciamiento = "COACTIVA"
                                elif datetime.now().date() <= ultimafechavence:
                                    estadofinanciamiento = "EN EJECUCIÓN"
                                else:
                                    estadofinanciamiento = "FINALIZADA"
                            else:
                                estadofinanciamiento = "NO TIENE RUBROS"

                            hojadestino.write(fila, 19, ultimafechavence, fceldafecha)
                            hojadestino.write(fila, 20, estadofinanciamiento, fceldageneral)

                            totalpagado += valorpagado
                            totalvencido += valorvencido
                            totalpendiente += valorpendiente

                    # Si hay rubros no vencidos: FIN

                    # Si hay rubros vencidos: INICIO
                    if datos['rubrosvencidos']:
                        for rubro_vencido in datos['rubrosvencidos']:
                            fila += 1

                            # print(fila)

                            registros += 1

                            codigorubro = rubro_vencido[0]
                            valorpagado = rubro_vencido[3]
                            valorpendiente = rubro_vencido[4]
                            valorvencido = rubro_vencido[5]
                            fechavence = rubro_vencido[1]
                            diasvencidos = rubro_vencido[6]
                            pagosvencidos = 1 if valorvencido > 0 else 0

                            categoriaantiguedad = ""

                            if diasvencidos <= 30:
                                programas[indice][6][1]['estudiantes'] += 1
                                programas[indice][6][1]['pagado'] += valorpagado
                                programas[indice][6][1]['vencido'] += valorvencido
                                programas[indice][6][1]['pendiente'] += valorpendiente

                                resumengeneral[1]['estudiantes'] += 1
                                resumengeneral[1]['pagado'] += valorpagado
                                resumengeneral[1]['vencido'] += valorvencido
                                resumengeneral[1]['pendiente'] += valorpendiente

                                categoriaantiguedad = "A"
                                rangodias = "1-30"
                            elif diasvencidos <= 60:
                                programas[indice][6][2]['estudiantes'] += 1
                                programas[indice][6][2]['pagado'] += valorpagado
                                programas[indice][6][2]['vencido'] += valorvencido
                                programas[indice][6][2]['pendiente'] += valorpendiente

                                resumengeneral[2]['estudiantes'] += 1
                                resumengeneral[2]['pagado'] += valorpagado
                                resumengeneral[2]['vencido'] += valorvencido
                                resumengeneral[2]['pendiente'] += valorpendiente

                                categoriaantiguedad = "B"
                                rangodias = "31-60"
                            elif diasvencidos <= 90:
                                programas[indice][6][3]['estudiantes'] += 1
                                programas[indice][6][3]['pagado'] += valorpagado
                                programas[indice][6][3]['vencido'] += valorvencido
                                programas[indice][6][3]['pendiente'] += valorpendiente

                                resumengeneral[3]['estudiantes'] += 1
                                resumengeneral[3]['pagado'] += valorpagado
                                resumengeneral[3]['vencido'] += valorvencido
                                resumengeneral[3]['pendiente'] += valorpendiente

                                categoriaantiguedad = "C"
                                rangodias = "61-90"
                            elif diasvencidos <= 180:
                                programas[indice][6][4]['estudiantes'] += 1
                                programas[indice][6][4]['pagado'] += valorpagado
                                programas[indice][6][4]['vencido'] += valorvencido
                                programas[indice][6][4]['pendiente'] += valorpendiente

                                resumengeneral[4]['estudiantes'] += 1
                                resumengeneral[4]['pagado'] += valorpagado
                                resumengeneral[4]['vencido'] += valorvencido
                                resumengeneral[4]['pendiente'] += valorpendiente

                                categoriaantiguedad = "D"
                                rangodias = "91-180"
                            else:
                                programas[indice][6][5]['estudiantes'] += 1
                                programas[indice][6][5]['pagado'] += valorpagado
                                programas[indice][6][5]['vencido'] += valorvencido
                                programas[indice][6][5]['pendiente'] += valorpendiente

                                resumengeneral[5]['estudiantes'] += 1
                                resumengeneral[5]['pagado'] += valorpagado
                                resumengeneral[5]['vencido'] += valorvencido
                                resumengeneral[5]['pendiente'] += valorpendiente

                                categoriaantiguedad = "E"
                                rangodias = "181 DÍAS EN ADELANTE"

                            hojadestino.write(fila, 0, registros, fceldageneral)
                            hojadestino.write(fila, 1, nombreprograma, fceldageneral)
                            hojadestino.write(fila, 2, numerocohorte, fceldageneralcent)
                            hojadestino.write(fila, 3, str(matricula.nivel.periodo.inicio) + " a " + str(
                                matricula.nivel.periodo.fin), fceldageneral)
                            hojadestino.write(fila, 4,
                                              "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO",
                                              fceldageneral)
                            hojadestino.write(fila, 5,
                                              personaalumno.provincia.nombre if personaalumno.provincia else '',
                                              fceldageneral)
                            hojadestino.write(fila, 6, personaalumno.canton.nombre if personaalumno.canton else '',
                                              fceldageneral)
                            hojadestino.write(fila, 7, personaalumno.identificacion(), fceldageneral)
                            hojadestino.write(fila, 8, alumno, fceldageneral)
                            hojadestino.write(fila, 9, pagosvencidos, fceldageneral)
                            hojadestino.write(fila, 10, codigorubro, fceldageneral)
                            hojadestino.write(fila, 11, fechavence, fceldafecha)
                            hojadestino.write(fila, 12, diasvencidos, fceldageneral)
                            hojadestino.write(fila, 13, valorpagado, fceldamoneda)
                            hojadestino.write(fila, 14, valorvencido, fceldamoneda)
                            hojadestino.write(fila, 15, valorpendiente, fceldamoneda)
                            hojadestino.write(fila, 16, categoriaantiguedad, fceldageneral)
                            hojadestino.write(fila, 17, rangodias, fceldageneral)
                            hojadestino.write(fila, 18, matricula.estado_inscripcion_maestria(), fceldageneral)

                            if rubrosalumno:
                                ultimafechavence = rubrosalumno.last().fechavence

                                if matricula.tiene_refinanciamiento_deuda_posgrado():
                                    estadofinanciamiento = "REFINANCIAMIENTO"
                                elif matricula.tiene_coactiva_posgrado():
                                    estadofinanciamiento = "COACTIVA"
                                elif datetime.now().date() <= ultimafechavence:
                                    estadofinanciamiento = "EN EJECUCIÓN"
                                else:
                                    estadofinanciamiento = "FINALIZADA"
                            else:
                                estadofinanciamiento = "NO TIENE RUBROS"

                            hojadestino.write(fila, 19, ultimafechavence, fceldafecha)
                            hojadestino.write(fila, 20, estadofinanciamiento, fceldageneral)

                            totalpagado += valorpagado
                            totalvencido += valorvencido
                            totalpendiente += valorpendiente

                    # Si hay rubros vencidos: FIN
                else:
                    registros += 1
                    fila += 1

                    # print(fila)

                    valorpagado = datos['totalpagado']
                    valorpendiente = datos['totalpendiente']
                    valorvencido = datos['totalvencido']
                    fechavence = datos['fechavence']
                    diasvencidos = datos['diasvencimiento']
                    pagosvencidos = 1 if valorvencido > 0 else 0

                    categoriaantiguedad = ""

                    if valorvencido > 0:
                        if diasvencidos <= 30:
                            programas[indice][6][1]['estudiantes'] += 1
                            programas[indice][6][1]['pagado'] += valorpagado
                            programas[indice][6][1]['vencido'] += valorvencido
                            programas[indice][6][1]['pendiente'] += valorpendiente

                            resumengeneral[1]['estudiantes'] += 1
                            resumengeneral[1]['pagado'] += valorpagado
                            resumengeneral[1]['vencido'] += valorvencido
                            resumengeneral[1]['pendiente'] += valorpendiente

                            categoriaantiguedad = "A"
                            rangodias = "1-30"
                        elif diasvencidos <= 60:
                            programas[indice][6][2]['estudiantes'] += 1
                            programas[indice][6][2]['pagado'] += valorpagado
                            programas[indice][6][2]['vencido'] += valorvencido
                            programas[indice][6][2]['pendiente'] += valorpendiente

                            resumengeneral[2]['estudiantes'] += 1
                            resumengeneral[2]['pagado'] += valorpagado
                            resumengeneral[2]['vencido'] += valorvencido
                            resumengeneral[2]['pendiente'] += valorpendiente

                            categoriaantiguedad = "B"
                            rangodias = "31-60"
                        elif diasvencidos <= 90:
                            programas[indice][6][3]['estudiantes'] += 1
                            programas[indice][6][3]['pagado'] += valorpagado
                            programas[indice][6][3]['vencido'] += valorvencido
                            programas[indice][6][3]['pendiente'] += valorpendiente

                            resumengeneral[3]['estudiantes'] += 1
                            resumengeneral[3]['pagado'] += valorpagado
                            resumengeneral[3]['vencido'] += valorvencido
                            resumengeneral[3]['pendiente'] += valorpendiente

                            categoriaantiguedad = "C"
                            rangodias = "61-90"
                        elif diasvencidos <= 180:
                            programas[indice][6][4]['estudiantes'] += 1
                            programas[indice][6][4]['pagado'] += valorpagado
                            programas[indice][6][4]['vencido'] += valorvencido
                            programas[indice][6][4]['pendiente'] += valorpendiente

                            resumengeneral[4]['estudiantes'] += 1
                            resumengeneral[4]['pagado'] += valorpagado
                            resumengeneral[4]['vencido'] += valorvencido
                            resumengeneral[4]['pendiente'] += valorpendiente

                            categoriaantiguedad = "D"
                            rangodias = "91-180"
                        else:
                            programas[indice][6][5]['estudiantes'] += 1
                            programas[indice][6][5]['pagado'] += valorpagado
                            programas[indice][6][5]['vencido'] += valorvencido
                            programas[indice][6][5]['pendiente'] += valorpendiente

                            resumengeneral[5]['estudiantes'] += 1
                            resumengeneral[5]['pagado'] += valorpagado
                            resumengeneral[5]['vencido'] += valorvencido
                            resumengeneral[5]['pendiente'] += valorpendiente

                            categoriaantiguedad = "E"
                            rangodias = "181 DÍAS EN ADELANTE"
                    else:
                        programas[indice][6][0]['estudiantes'] += 1
                        programas[indice][6][0]['pagado'] += valorpagado
                        programas[indice][6][0]['vencido'] += valorvencido
                        programas[indice][6][0]['pendiente'] += valorpendiente

                        resumengeneral[0]['estudiantes'] += 1
                        resumengeneral[0]['pagado'] += valorpagado
                        resumengeneral[0]['vencido'] += valorvencido
                        resumengeneral[0]['pendiente'] += valorpendiente

                        categoriaantiguedad = "VIGENTE"
                        rangodias = ""

                    hojadestino.write(fila, 0, registros, fceldageneral)
                    hojadestino.write(fila, 1, nombreprograma, fceldageneral)
                    hojadestino.write(fila, 2, numerocohorte, fceldageneralcent)
                    hojadestino.write(fila, 3,
                                      str(matricula.nivel.periodo.inicio) + " a " + str(matricula.nivel.periodo.fin),
                                      fceldageneral)
                    hojadestino.write(fila, 4,
                                      "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO",
                                      fceldageneral)
                    hojadestino.write(fila, 5, personaalumno.provincia.nombre if personaalumno.provincia else '',
                                      fceldageneral)
                    hojadestino.write(fila, 6, personaalumno.canton.nombre if personaalumno.canton else '',
                                      fceldageneral)
                    hojadestino.write(fila, 7, personaalumno.identificacion(), fceldageneral)
                    hojadestino.write(fila, 8, alumno, fceldageneral)
                    hojadestino.write(fila, 9, pagosvencidos, fceldageneral)
                    hojadestino.write(fila, 10, "S/N", fceldageneral)
                    hojadestino.write(fila, 11, fechavence, fceldafecha)
                    hojadestino.write(fila, 12, diasvencidos, fceldageneral)
                    hojadestino.write(fila, 13, valorpagado, fceldamoneda)
                    hojadestino.write(fila, 14, valorvencido, fceldamoneda)
                    hojadestino.write(fila, 15, valorpendiente, fceldamoneda)
                    hojadestino.write(fila, 16, categoriaantiguedad, fceldageneral)
                    hojadestino.write(fila, 17, rangodias, fceldageneral)
                    hojadestino.write(fila, 18, matricula.estado_inscripcion_maestria(), fceldageneral)

                    if rubrosalumno:
                        ultimafechavence = rubrosalumno.last().fechavence

                        if matricula.tiene_refinanciamiento_deuda_posgrado():
                            estadofinanciamiento = "REFINANCIAMIENTO"
                        elif matricula.tiene_coactiva_posgrado():
                            estadofinanciamiento = "COACTIVA"
                        elif datetime.now().date() <= ultimafechavence:
                            estadofinanciamiento = "EN EJECUCIÓN"
                        else:
                            estadofinanciamiento = "FINALIZADA"
                    else:
                        estadofinanciamiento = "NO TIENE RUBROS"

                    hojadestino.write(fila, 19, ultimafechavence, fceldafecha)
                    hojadestino.write(fila, 20, estadofinanciamiento, fceldageneral)

                    totalpagado += valorpagado
                    totalvencido += valorvencido
                    totalpendiente += valorpendiente

                # No es retirado FIN

                print(secuencia, " de ", totalmatriculas)

            fila += 1
            hojadestino.merge_range(fila, 0, fila, 12, "TOTALES", fceldanegritageneral)
            hojadestino.write(fila, 13, totalpagado, fceldamonedapie)
            hojadestino.write(fila, 14, totalvencido, fceldamonedapie)
            hojadestino.write(fila, 15, totalpendiente, fceldamonedapie)

            fila += 3
            hojadestino.merge_range(fila, 0, fila, 11, "RESUMEN DE CARTERA VENCIDA GENERAL", ftitulo3izq)

            fila += 1

            hojadestino.merge_range(fila, 0, fila, 1, "PERIODO DE VENCIMIENTO", fcabeceracolumna)
            hojadestino.write(fila, 2, "NRO.ESTUDIANTES", fcabeceracolumna)
            hojadestino.write(fila, 3, "VALOR PAGADO", fcabeceracolumna)
            hojadestino.write(fila, 4, "VALOR CARTERA VENCIDA", fcabeceracolumna)
            hojadestino.write(fila, 5, "VALOR PENDIENTE", fcabeceracolumna)
            hojadestino.write(fila, 6, "ANTIGUEDAD", fcabeceracolumna)

            fila += 1
            for i in resumengeneral:
                hojadestino.merge_range(fila, 0, fila, 1, resumengeneral[i]['etiqueta'], fceldageneral)
                hojadestino.write(fila, 2, resumengeneral[i]['estudiantes'], fceldageneral)
                hojadestino.write(fila, 3, resumengeneral[i]['pagado'], fceldamoneda)
                hojadestino.write(fila, 4, resumengeneral[i]['vencido'], fceldamoneda)
                hojadestino.write(fila, 5, resumengeneral[i]['pendiente'], fceldamoneda)
                hojadestino.write(fila, 6, resumengeneral[i]['antiguedad'], fceldageneralcent)
                fila += 1

            hojadestino.merge_range(fila, 0, fila, 1, "TOTAL", fceldanegritageneral)
            hojadestino.write(fila, 2, registros, fceldanegritageneral)
            hojadestino.write(fila, 3, totalpagado, fceldamonedapie)
            hojadestino.write(fila, 4, totalvencido, fceldamonedapie)
            hojadestino.write(fila, 5, totalpendiente, fceldamonedapie)

            fila += 3
            hojadestino.merge_range(fila, 0, fila, 11, "RESUMEN DE CARTERA VENCIDA POR PROGRAMA DE MAESTRÍA",
                                    ftitulo3izq)

            fila += 1
            hojadestino.write(fila, 0, "N°", fcabeceracolumna)
            hojadestino.write(fila, 1, "PROGRAMA", fcabeceracolumna)
            hojadestino.write(fila, 2, "COHORTE", fcabeceracolumna)
            hojadestino.write(fila, 3, "PERIODO (INICIO-FIN)", fcabeceracolumna)
            hojadestino.write(fila, 4, "ESTADO DE MAESTRIA", fcabeceracolumna)
            hojadestino.write(fila, 5, "PERIODO DE VENCIMIENTO", fcabeceracolumna)
            hojadestino.write(fila, 6, "NRO.ESTUDIANTES", fcabeceracolumna)
            hojadestino.write(fila, 7, "VALOR PAGADO", fcabeceracolumna)
            hojadestino.write(fila, 8, "VALOR CARTERA VENCIDA", fcabeceracolumna)
            hojadestino.write(fila, 9, "VALOR PENDIENTE", fcabeceracolumna)
            hojadestino.write(fila, 10, "ANTIGUEDAD", fcabeceracolumna)

            # Ordeno por programa y cohorte
            programas = sorted(programas, key=lambda programa: (programa[2], programa[3]))

            secresumen = 0
            for datoprograma in programas:
                fila += 1
                secresumen += 1
                hojadestino.merge_range(fila, 0, fila + 5, 0, secresumen, fceldageneralcent)
                hojadestino.merge_range(fila, 1, fila + 5, 1, datoprograma[2], fceldageneralcent)
                hojadestino.merge_range(fila, 2, fila + 5, 2, datoprograma[3], fceldageneralcent)
                hojadestino.merge_range(fila, 3, fila + 5, 3, datoprograma[4], fceldageneralcent)
                hojadestino.merge_range(fila, 4, fila + 5, 4, datoprograma[5], fceldageneralcent)

                tot_est_prog = tot_venc_prog = tot_pend_prog = tot_pag_prog = 0

                resumen = datoprograma[6]
                for i in resumen:
                    hojadestino.write(fila, 5, resumen[i]['etiqueta'], fceldageneral)
                    hojadestino.write(fila, 6, resumen[i]['estudiantes'], fceldageneral)
                    hojadestino.write(fila, 7, resumen[i]['pagado'], fceldamoneda)
                    hojadestino.write(fila, 8, resumen[i]['vencido'], fceldamoneda)
                    hojadestino.write(fila, 9, resumen[i]['pendiente'], fceldamoneda)
                    hojadestino.write(fila, 10, resumen[i]['antiguedad'], fceldageneralcent)

                    tot_est_prog += resumen[i]['estudiantes']
                    tot_pag_prog += resumen[i]['pagado']
                    tot_venc_prog += resumen[i]['vencido']
                    tot_pend_prog += resumen[i]['pendiente']

                    fila += 1

                hojadestino.merge_range(fila, 0, fila, 5,
                                        "TOTAL " + datoprograma[2] + " COHORTE " + str(datoprograma[3]),
                                        fceldanegritageneral)
                hojadestino.write(fila, 6, tot_est_prog, fceldanegritageneral)
                hojadestino.write(fila, 7, tot_pag_prog, fceldamonedapie)
                hojadestino.write(fila, 8, tot_venc_prog, fceldamonedapie)
                hojadestino.write(fila, 9, tot_pend_prog, fceldamonedapie)

            fila += 1
            hojadestino.merge_range(fila, 0, fila, 5, "TOTAL GENERAL", fceldanegritageneral)
            hojadestino.write(fila, 6, registros, fceldanegritageneral)
            hojadestino.write(fila, 7, totalpagado, fceldamonedapie)
            hojadestino.write(fila, 8, totalvencido, fceldamonedapie)
            hojadestino.write(fila, 9, totalpendiente, fceldamonedapie)

            workbook.close()

            # Notificar al usuario
            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.cuerpo = 'Reporte Excel finalizado'
                notificacion.url = urlarchivo
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo='Reporte Excel finalizado',
                    titulo=tituloreporte,
                    destinatario=personanotifica,
                    url=urlarchivo,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False
                )
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte Excel finalizado",
                "body": tituloreporte,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": urlarchivo,
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "El <b>{}</b> ha sido generado con éxito".format(tituloreporte)
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.error = True
                notificacion.titulo = "Error al generar {}".format(tituloreporte)
                notificacion.cuerpo = textoerror
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo="Error al generar Reporte Excel",
                    titulo="Error al generar {}".format(tituloreporte),
                    destinatario=personanotifica,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False,
                    error=True)
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error al generar Reporte Excel",
                "body": "Error al generar {}".format(tituloreporte),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "Error al generar <b>{}</b>. Error: {}".format(tituloreporte, textoerror),
                "error": True
            }, ttl=500)


class reporte_presupuesto_anual_background(threading.Thread):

    def __init__(self, request, data, idnotificacion, tiporeporte, anio, desde, hasta, listaporcentajes):
        self.request = request
        self.data = data
        self.idnotificacion = idnotificacion
        self.tiporeporte = tiporeporte
        self.anio = anio
        self.desde = desde
        self.hasta = hasta
        self.listaporcentajes = listaporcentajes
        threading.Thread.__init__(self)

    def run(self):
        try:
            request, data, idnotificacion, tiporeporte, anio, desde, hasta, listaporcentajes = self.request, self.data, self.idnotificacion, self.tiporeporte, self.anio, self.desde, self.hasta, self.listaporcentajes

            if tiporeporte == 1:
                fechadesde = datetime.strptime(anio + '-01-01', '%Y-%m-%d').date()
                fechahasta = datetime.strptime(anio + '-12-31', '%Y-%m-%d').date()
            else:
                fechadesde = datetime.strptime(desde, '%Y-%m-%d').date()
                fechahasta = datetime.strptime(hasta, '%Y-%m-%d').date()

            usernotify = User.objects.get(pk=request.user.pk)
            personanotifica = Persona.objects.get(usuario=usernotify)
            tituloreporte = "Reporte Excel Presupuesto Anual"

            directorio = os.path.join(os.path.join(MEDIA_ROOT, 'reportes', 'posgrado'))
            try:
                os.stat(directorio)
            except:
                os.mkdir(directorio)

            # nombrearchivo = "PROYECCIONANUAL_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"
            nombrearchivo = "PROYECCIONANUAL_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
            urlarchivo = MEDIA_URL + "reportes/posgrado/" + nombrearchivo

            anio = fechadesde.year

            style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            title = easyxf(
                'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
            title2 = easyxf(
                'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
            style1 = easyxf(num_format_str='DD/mm/YYYY')
            # style2 = easyxf(num_format_str='HH:mm')
            style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            fuentecabecera = easyxf(
                'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
            fuentenormal = easyxf(
                'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
            title3 = easyxf(
                'font: name Times New Roman, color-index black, bold on , height 200; alignment: horiz centre; borders: top thin')

            fuentemoneda = easyxf(
                'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                num_format_str=' "$" #,##0.00')

            fuentemonedafv = easyxf(
                'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour lime',
                num_format_str=' "$" #,##0.00')

            fuentemonedafn = easyxf(
                'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour light_orange',
                num_format_str=' "$" #,##0.00')

            fuentemonedaneg = easyxf(
                'font: name Verdana, color-index black, height 150, bold on; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour gray25',
                num_format_str=' "$" #,##0.00')

            fuentenormal2 = easyxf(
                'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
            fuentenormalnegcent2 = easyxf(
                'font: name Verdana, bold on, color-index black, height 150; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_colour gray25')
            fuentenormalneg2 = easyxf('font: name Verdana, bold on, color-index black, height 150')

            wb = Workbook(encoding='utf-8')
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = 'attachment; filename=presupuesto_anual_' + random.randint(1,
                                                                                                         10000).__str__() + '.xls'
            # nombre = "PROYECCIONANUAL_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
            filename = os.path.join(directorio, nombrearchivo)
            ruta = "media/postgrado/" + nombrearchivo
            ws = wb.add_sheet('PROY PRESU')
            ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            ws.write_merge(1, 1, 0, 10, 'DIRECCIÓN DE INVESTIGACIÓN Y POSGRADO', title2)
            ws.write_merge(2, 2, 0, 10, 'PROYECCION DEL PRESUPUESTO GENERAL ' + str(anio), title2)
            ws.write_merge(4, 4, 0, 0, u'INGRESOS', fuentenormalneg2)
            ws.write_merge(5, 5, 0, 0, u'PERIODO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 1, 1, u'PROGRAMAS EN EJECUCIÓN', fuentenormalnegcent2)
            ws.write_merge(5, 5, 2, 2, u'INGRESO BRUTO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 3, 3, u'RETIRADO%-MORA%-BECA%', fuentenormalnegcent2)
            ws.write_merge(5, 5, 4, 4, u'INGRESO NETO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 5, 5, u'PAGOS UNEMI', fuentenormalnegcent2)
            ws.write_merge(5, 5, 6, 6, u'PAGOS EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(5, 5, 7, 7, u'INGRESO PAGADO/TESORERIA', fuentenormalnegcent2)

            row_num = 6
            valor_total = 0
            ingreso_neto_total = 0
            valor_pagado_total = 0
            valor_pagado_total_unemi = 0
            valor_pagado_total_epunemi = 0

            valor_reprobado_total = 0
            valor_reprobado_total_unemi = 0
            valor_reprobado_total_epunemi = 0

            lp1 = Periodo.objects.values_list('id', flat=True).filter(nivel__matricula__rubro__fechavence__year=anio,
                                                                      tipo__id__in=[3, 4]).distinct().order_by('id')

            lp1 = Periodo.objects.values_list('id', flat=True).filter(
                nivel__matricula__rubro__fechavence__gte=fechadesde,
                nivel__matricula__rubro__fechavence__lte=fechahasta, tipo__id__in=[3, 4]).distinct().order_by('id')

            lp2 = Periodo.objects.values_list('id', flat=True).filter(nivel__matricula__rubro__pago__fecha__year=anio,
                                                                      tipo__id__in=[3, 4]).distinct().order_by('id')

            lp2 = Periodo.objects.values_list('id', flat=True).filter(
                nivel__matricula__rubro__pago__fecha__gte=fechadesde,
                nivel__matricula__rubro__pago__fecha__lte=fechahasta, tipo__id__in=[3, 4]).distinct().order_by('id')

            lista_periodos = (lp1 | lp2).distinct().order_by('id')

            lc1 = Carrera.objects.values_list('id', flat=True).filter(
                inscripcion__matricula__rubro__fechavence__year=anio,
                inscripcion__matricula__nivel__periodo_id__in=lista_periodos,
                inscripcion__matricula__nivel__periodo__tipo__id__in=[3, 4]).distinct().order_by('nombre')

            lc1 = Carrera.objects.values_list('id', flat=True).filter(
                inscripcion__matricula__rubro__fechavence__gte=fechadesde,
                inscripcion__matricula__rubro__fechavence__lte=fechahasta,
                inscripcion__matricula__nivel__periodo_id__in=lista_periodos,
                inscripcion__matricula__nivel__periodo__tipo__id__in=[3, 4]).distinct().order_by('nombre')

            lc2 = Carrera.objects.values_list('id', flat=True).filter(
                inscripcion__matricula__rubro__pago__fecha__year=anio,
                inscripcion__matricula__nivel__periodo_id__in=lista_periodos,
                inscripcion__matricula__nivel__periodo__tipo__id__in=[3, 4]).distinct().order_by('nombre')

            lc2 = Carrera.objects.values_list('id', flat=True).filter(
                inscripcion__matricula__rubro__pago__fecha__gte=fechadesde,
                inscripcion__matricula__rubro__pago__fecha__lte=fechahasta,
                inscripcion__matricula__nivel__periodo_id__in=lista_periodos,
                inscripcion__matricula__nivel__periodo__tipo__id__in=[3, 4]).distinct().order_by('nombre')

            lista_carreras = (lc1 | lc2).distinct().order_by('nombre')

            pagos = Pago.objects.filter(pagoliquidacion__isnull=True,

                                        fecha__gte=fechadesde,
                                        fecha__lte=fechahasta,

                                        rubro__matricula__nivel__periodo_id__in=lista_periodos,
                                        rubro__matricula__inscripcion__carrera_id__in=lista_carreras,
                                        status=True,
                                        rubro__status=True
                                        ).exclude(factura__valida=False).order_by('fecha',
                                                                                  'rubro__matricula__nivel__periodo_id',
                                                                                  'rubro__matricula__inscripcion__carrera')

            pagosunemi = pagos.filter(idpagoepunemi=0).annotate(anio=ExtractYear('fecha')).values_list('anio',
                                                                                                       'rubro__matricula__nivel__periodo_id',
                                                                                                       'rubro__matricula__inscripcion__carrera_id').annotate(
                tpagado=Sum('valortotal'))

            pagosepunemi = pagos.filter(~Q(idpagoepunemi=0)).annotate(anio=ExtractYear('fecha')).values_list('anio',
                                                                                                             'rubro__matricula__nivel__periodo_id',
                                                                                                             'rubro__matricula__inscripcion__carrera_id').annotate(
                tpagado=Sum('valortotal'))

            pagos = pagos.annotate(anio=ExtractYear('fecha')).values_list('anio',
                                                                          'rubro__matricula__nivel__periodo_id',
                                                                          'rubro__matricula__inscripcion__carrera_id').annotate(
                tpagado=Sum('valortotal'))

            for p in Periodo.objects.filter(pk__in=lista_periodos).order_by('id'):
                for carrera in Carrera.objects.filter(inscripcion__matricula__nivel__periodo=p,
                                                      pk__in=lista_carreras).distinct().order_by('nombre'):

                    porcentaje = 0
                    for lp in listaporcentajes:
                        if int(lp['periodo']) == p.id and int(lp['carrera']) == carrera.id:
                            porcentaje = Decimal(lp['porcentaje']).quantize(Decimal('.01'))
                            break

                    ws.write_merge(row_num, row_num, 0, 0, p.nombre, fuentenormal2)
                    ws.write_merge(row_num, row_num, 1, 1, carrera.nombre, fuentenormal2)
                    valor = Decimal(null_to_decimal(
                        Rubro.objects.filter(matricula__nivel__periodo=p, matricula__inscripcion__carrera=carrera,
                                             fechavence__year=anio, status=True).aggregate(valor=Sum('saldo'))['valor'],
                        2)).quantize(Decimal('.01'))

                    # Total generado rubros
                    totalproyectado = Decimal(
                        null_to_decimal(Rubro.objects.filter(matricula__nivel__periodo=p,
                                                             matricula__inscripcion__carrera=carrera,
                                                             status=True,
                                                             # fechavence__year=anio,

                                                             fechavence__gte=fechadesde,
                                                             fechavence__lte=fechahasta

                                                             ).aggregate(
                            valor=Sum('valor'))['valor'], 2)).quantize(Decimal('.01'))

                    # Total anulado rubros
                    totalanulado = Decimal(
                        null_to_decimal(Pago.objects.filter(rubro__matricula__nivel__periodo=p,
                                                            rubro__matricula__inscripcion__carrera=carrera,
                                                            status=True,
                                                            rubro__status=True,
                                                            # rubro__fechavence__year=anio,

                                                            rubro__fechavence__gte=fechadesde,
                                                            rubro__fechavence__lte=fechahasta,

                                                            factura__valida=False,
                                                            factura__status=True).aggregate(
                            valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))

                    # Total liquidado rubros
                    totalliquidado = Decimal(
                        null_to_decimal(Pago.objects.filter(rubro__matricula__nivel__periodo=p,
                                                            rubro__matricula__inscripcion__carrera=carrera,
                                                            status=True,
                                                            rubro__status=True,
                                                            # rubro__fechavence__year=anio,

                                                            rubro__fechavence__gte=fechadesde,
                                                            rubro__fechavence__lte=fechahasta,

                                                            pagoliquidacion__isnull=False,
                                                            pagoliquidacion__status=True).aggregate(
                            valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))

                    valor_bruto = totalproyectado - (totalanulado + totalliquidado)
                    valor = valor_bruto

                    totalpagado = 0
                    totalpagadounemi = 0
                    totalpagadoepunemi = 0

                    aniofila = 0
                    periodofila = 0
                    carrerafila = 0

                    for rp in pagos:
                        if rp[0] == anio and rp[1] == p.id and rp[2] == carrera.id:
                            totalpagado += Decimal(rp[3])
                            aniofila = rp[0]
                            periodofila = rp[1]
                            carrerafila = rp[2]
                        elif aniofila != 0 and aniofila != rp[0] and periodofila != rp[1] and carrerafila != rp[2]:
                            break

                    aniofila = 0
                    periodofila = 0
                    carrerafila = 0
                    for une in pagosunemi:
                        if une[0] == anio and une[1] == p.id and une[2] == carrera.id:
                            totalpagadounemi += Decimal(une[3])
                            aniofila = une[0]
                            periodofila = une[1]
                            carrerafila = une[2]
                        elif aniofila != 0 and aniofila != une[0] and periodofila != une[1] and carrerafila != une[2]:
                            break
                    aniofila = 0
                    periodofila = 0
                    carrerafila = 0
                    for epu in pagosepunemi:
                        if epu[0] == anio and epu[1] == p.id and epu[2] == carrera.id:
                            totalpagadoepunemi += Decimal(epu[3])
                            aniofila = epu[0]
                            periodofila = epu[1]
                            carrerafila = epu[2]
                        elif aniofila != 0 and aniofila != epu[0] and periodofila != epu[1] and carrerafila != epu[2]:
                            break

                    valor_pagado = totalpagado
                    valor_pagado_unemi = totalpagadounemi
                    valor_pagado_epunemi = totalpagadoepunemi

                    valor_descuento = Decimal(((valor * porcentaje) / 100)).quantize(Decimal('.01'))

                    ingreso_neto = valor - valor_descuento
                    valor_total += valor
                    valor_pagado_total += valor_pagado

                    valor_pagado_total_unemi += valor_pagado_unemi
                    valor_pagado_total_epunemi += valor_pagado_epunemi

                    ingreso_neto_total += ingreso_neto

                    ws.write_merge(row_num, row_num, 2, 2, valor, fuentemoneda)
                    ws.write_merge(row_num, row_num, 3, 3, valor_descuento, fuentemoneda)
                    ws.write_merge(row_num, row_num, 4, 4, ingreso_neto, fuentemoneda)
                    ws.write_merge(row_num, row_num, 5, 5, valor_pagado_unemi, fuentemonedafv)
                    ws.write_merge(row_num, row_num, 6, 6, valor_pagado_epunemi, fuentemonedafn)
                    ws.write_merge(row_num, row_num, 7, 7, valor_pagado, fuentemoneda)

                    row_num += 1
            ws.write_merge(row_num, row_num, 0, 0, u'SUBTOTAL DE EJECUCIÓN', fuentenormalneg2)
            ws.write_merge(row_num, row_num, 2, 2, valor_total, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 4, 4, ingreso_neto_total, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 5, 5, valor_pagado_total_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 6, 6, valor_pagado_total_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 7, 7, valor_pagado_total, fuentemonedaneg)

            # busco otros tipos rubros que tengan en la descripcion MAESTRIA y que sea de la partida 101
            # codigo_rubro = TipoOtroRubro.objects.values_list('id', flat=True).filter(partida__id=101, nombre__icontains='MAESTR', status=True)
            codigo_rubro = TipoOtroRubro.objects.values_list('id', flat=True).filter(tiporubro=1, status=True)
            # codigo_rubro = TipoOtroRubro.objects.values_list('id', flat=True).filter(pk__in=[2845,2957,2926,2958,3002,2915,2927,2989,2902,3009,2925,2870,2912,3010,3017,2956,3025], activo=True, status=True)
            # valornoadmitidos = Decimal(null_to_decimal(Rubro.objects.filter(fechavence__year__lte=anio, status=True, matricula__isnull=True, tipo__in=codigo_rubro).aggregate(valor=Sum('saldo'))['valor'],2)).quantize(Decimal('.01'))

            # valornoadmitidos = Decimal(null_to_decimal(Pago.objects.filter(rubro__matricula__isnull=True, rubro__tipo_id__in=codigo_rubro, status=True, rubro__status=True, fecha__year=anio).aggregate(valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))
            valornoadmitidos = Decimal(null_to_decimal(
                Pago.objects.filter(rubro__matricula__isnull=True, rubro__tipo_id__in=codigo_rubro, status=True,
                                    rubro__status=True, fecha__gte=fechadesde, fecha__lte=fechahasta).aggregate(
                    valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))

            # Rubros Adicionales Pagados
            pagosadicionales = Pago.objects.filter(
                ~Q(rubro__tipo__subtiporubro=1),
                rubro__tipo_id__in=codigo_rubro,
                pagoliquidacion__isnull=True,
                rubro__matricula__isnull=False,
                status=True,
                rubro__status=True,
                fecha__gte=fechadesde,
                fecha__lte=fechahasta
            ).exclude(factura__valida=False)

            # Obtengo el total pagado por rubros adicionales
            valoradicionales = Decimal(
                null_to_decimal(pagosadicionales.aggregate(valor=Sum('valortotal'))['valor'], 2)).quantize(
                Decimal('.01'))

            rubrosadicionales = Rubro.objects.values_list('tipo_id', 'tipo__nombre').filter(
                ~Q(tipo__subtiporubro=1),
                tipo_id__in=codigo_rubro,
                pago__pagoliquidacion__isnull=True,
                matricula__isnull=False,
                status=True,
                pago__status=True,
                pago__fecha__gte=fechadesde,
                pago__fecha__lte=fechahasta
            ).exclude(pago__factura__valida=False).order_by('tipo__nombre').distinct()

            row_num += 1
            ws.write_merge(row_num, row_num, 0, 0, u'NO ADMITIDOS', fuentenormalneg2)
            ws.write_merge(row_num, row_num, 7, 7, valornoadmitidos, fuentemonedaneg)

            row_num += 1
            ws.write_merge(row_num, row_num, 0, 0, u'RUBROS ADICIONALES', fuentenormalneg2)
            ws.write_merge(row_num, row_num, 7, 7, valoradicionales, fuentemonedaneg)

            row_num += 1
            ws.write_merge(row_num, row_num, 0, 0, u'TOTAL DE INGRESOS', fuentenormalneg2)
            ws.write_merge(row_num, row_num, 7, 7, valornoadmitidos + valor_pagado_total + valoradicionales,
                           fuentemonedaneg)

            columns = [
                (u"", 9000),
                (u"", 9000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000)
            ]
            row_num += 1
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style2)
                ws.col(col_num).width = columns[col_num][1]

            row_num += 1
            # LISTADO DE RUBROS NO ADMITIDOS
            ws.write_merge(row_num, row_num, 0, 0, u'NO ADMITIDOS', fuentenormalneg2)
            row_num += 1
            ws.write_merge(row_num, row_num, 0, 0, u'RUBRO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 1, 1, u'RUBRO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 2, 2, u'TOTAL UNEMI', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 3, 3, u'TOTAL EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 4, 4, u'TOTAL', fuentenormalnegcent2)

            valor_total = 0
            valor_total_unemi = 0
            valor_total_epunemi = 0
            row_num += 1

            # listadorubros = Rubro.objects.values_list('tipo_id', 'tipo__nombre').filter(pago__fecha__year=anio, status=True, matricula__isnull=True, tipo__in=codigo_rubro).distinct()
            listadorubros = Rubro.objects.values_list('tipo_id', 'tipo__nombre').filter(pago__fecha__gte=fechadesde,
                                                                                        pago__fecha__lte=fechahasta,
                                                                                        status=True,
                                                                                        matricula__isnull=True,
                                                                                        tipo__in=codigo_rubro).distinct()

            for lista in listadorubros:
                ws.write_merge(row_num, row_num, 0, 0, 'RUBRO', fuentenormal2)
                ws.write_merge(row_num, row_num, 1, 1, lista[1], fuentenormal2)

                # valor_pagado = Decimal(null_to_decimal(Pago.objects.filter(rubro__matricula__isnull=True, rubro__tipo_id=lista[0], status=True, rubro__status=True, fecha__year=anio).aggregate(valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))
                valor_pagado = Decimal(null_to_decimal(
                    Pago.objects.filter(rubro__matricula__isnull=True, rubro__tipo_id=lista[0], status=True,
                                        rubro__status=True, fecha__gte=fechadesde, fecha__lte=fechahasta).aggregate(
                        valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))

                # valor_pagado_unemi = Decimal(null_to_decimal(Pago.objects.filter(idpagoepunemi=0, rubro__matricula__isnull=True, rubro__tipo_id=lista[0], status=True, rubro__status=True, fecha__year=anio).aggregate(valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))
                valor_pagado_unemi = Decimal(null_to_decimal(
                    Pago.objects.filter(idpagoepunemi=0, rubro__matricula__isnull=True, rubro__tipo_id=lista[0],
                                        status=True, rubro__status=True, fecha__gte=fechadesde,
                                        fecha__lte=fechahasta).aggregate(valor=Sum('valortotal'))['valor'],
                    2)).quantize(
                    Decimal('.01'))

                # valor_pagado_epunemi = Decimal(null_to_decimal(Pago.objects.filter(~Q(idpagoepunemi=0), rubro__matricula__isnull=True, rubro__tipo_id=lista[0], status=True, rubro__status=True, fecha__year=anio).aggregate(valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))
                valor_pagado_epunemi = Decimal(null_to_decimal(
                    Pago.objects.filter(~Q(idpagoepunemi=0), rubro__matricula__isnull=True, rubro__tipo_id=lista[0],
                                        status=True, rubro__status=True, fecha__gte=fechadesde,
                                        fecha__lte=fechahasta).aggregate(valor=Sum('valortotal'))['valor'],
                    2)).quantize(
                    Decimal('.01'))

                ws.write_merge(row_num, row_num, 2, 2, valor_pagado_unemi, fuentemonedafv)
                ws.write_merge(row_num, row_num, 3, 3, valor_pagado_epunemi, fuentemonedafn)
                ws.write_merge(row_num, row_num, 4, 4, valor_pagado, fuentemoneda)
                valor_total += valor_pagado
                valor_total_unemi += valor_pagado_unemi
                valor_total_epunemi += valor_pagado_epunemi
                row_num += 1
            ws.write_merge(row_num, row_num, 0, 1, u'TOTAL NO ADMITIDOS', fuentenormalneg2)
            ws.write_merge(row_num, row_num, 2, 2, valor_total_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 3, 3, valor_total_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 4, 4, valor_total, fuentemonedaneg)

            columns = [
                (u"", 9000),
                (u"", 9000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000)]
            row_num += 1
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style2)
                ws.col(col_num).width = columns[col_num][1]
            row_num += 1

            # LISTADO DE RUBROS ADICIONALES
            ws.write_merge(row_num, row_num, 0, 0, u'RUBROS ADICIONALES', fuentenormalneg2)
            row_num += 1
            ws.write_merge(row_num, row_num, 0, 1, u'RUBRO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 2, 2, u'TOTAL UNEMI', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 3, 3, u'TOTAL EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 4, 4, u'TOTAL', fuentenormalnegcent2)

            valor_total = 0
            valor_total_unemi = 0
            valor_total_epunemi = 0
            row_num += 1
            for rubroadicional in rubrosadicionales:
                ws.write_merge(row_num, row_num, 0, 1, rubroadicional[1], fuentenormal2)
                valor_pagado_unemi = Decimal(null_to_decimal(
                    pagosadicionales.filter(idpagoepunemi=0, rubro__tipo_id=rubroadicional[0]).aggregate(
                        valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))
                valor_pagado_epunemi = Decimal(null_to_decimal(
                    pagosadicionales.filter(~Q(idpagoepunemi=0), rubro__tipo_id=rubroadicional[0]).aggregate(
                        valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))
                valor_pagado = valor_pagado_unemi + valor_pagado_epunemi
                ws.write_merge(row_num, row_num, 2, 2, valor_pagado_unemi, fuentemonedafv)
                ws.write_merge(row_num, row_num, 3, 3, valor_pagado_epunemi, fuentemonedafn)
                ws.write_merge(row_num, row_num, 4, 4, valor_pagado, fuentemoneda)
                valor_total += valor_pagado
                valor_total_unemi += valor_pagado_unemi
                valor_total_epunemi += valor_pagado_epunemi
                row_num += 1

            ws.write_merge(row_num, row_num, 0, 1, u'TOTAL RUBROS ADICIONALES', fuentenormalneg2)
            ws.write_merge(row_num, row_num, 2, 2, valor_total_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 3, 3, valor_total_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 4, 4, valor_total, fuentemonedaneg)

            # cartera vencida
            row_num += 2
            ws.write_merge(row_num, row_num, 0, 0, u'CARTERA VENCIDA AL ' + fechahasta.strftime("%d/%m/%Y"),
                           fuentenormalneg2)
            row_num += 1
            ws.write_merge(row_num, row_num, 0, 0, u'PERIODO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 1, 1, u'PROGRAMAS EN EJECUCIÓN', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 2, 2, u'TOTAL', fuentenormalnegcent2)
            valor_total = 0
            row_num += 1

            for p in Periodo.objects.filter(pk__in=lista_periodos).order_by('id'):
                for carrera in Carrera.objects.filter(inscripcion__matricula__nivel__periodo=p,
                                                      pk__in=lista_carreras).distinct().order_by('nombre'):
                    vencidoprograma = 0
                    ws.write_merge(row_num, row_num, 0, 0, p.nombre, fuentenormal2)
                    ws.write_merge(row_num, row_num, 1, 1, carrera.nombre, fuentenormal2)

                    matriculaspos = Matricula.objects.filter(status=True, inscripcion__carrera=carrera,
                                                             nivel__periodo=p).order_by('id')

                    for matriculap in matriculaspos:
                        vencidopersona = 0
                        datos = matriculap.rubros_maestria_vencidos_detalle_version_final(fechahasta)
                        if not matriculap.retirado_programa_maestria():
                            if datos['rubrosnovencidos']:  # Se considera a los no vencidos que tienen saldo negativo
                                for rubro_no_vencido in datos['rubrosnovencidos']:
                                    vencidopersona += rubro_no_vencido[5]

                            if datos['rubrosvencidos']:
                                for rubro_vencido in datos['rubrosvencidos']:
                                    vencidopersona += rubro_vencido[5]
                        else:
                            vencidopersona = datos['totalvencido']

                        vencidoprograma += vencidopersona

                    valor_total += vencidoprograma
                    ws.write_merge(row_num, row_num, 2, 2, vencidoprograma, fuentemoneda)
                    row_num += 1

            ws.write_merge(row_num, row_num, 0, 0, u'TOTAL CARTERA VENCIDA', fuentenormalneg2)
            ws.write_merge(row_num, row_num, 2, 2, valor_total, fuentemonedaneg)

            row_num += 2
            ws.write_merge(row_num, row_num, 0, 0, u'FECHA ARCHIVO  ' + str(datetime.now())[0:19], fuentenormalneg2)

            ws = wb.add_sheet('FLUJO INGRESO')
            ws.write_merge(0, 0, 0, 4, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            ws.write_merge(1, 1, 0, 4, 'DIRECCIÓN DE INVESTIGACIÓN Y POSGRADO', title2)
            ws.write_merge(2, 2, 0, 4, 'FLUJO DE INGRESOS VS GASTOS ' + str(anio), title2)

            ws.write_merge(4, 4, 0, 3, u'INGRESOS', fuentenormalneg2)
            ws.write_merge(4, 4, 4, 16, u'PROYECCIÓN', fuentenormalneg2)

            ws.write_merge(5, 6, 0, 0, u'PERIODO', fuentenormalnegcent2)
            ws.write_merge(5, 6, 1, 2, u'PROGRAMAS EN EJECUCIÓN', fuentenormalnegcent2)
            ws.write_merge(5, 6, 3, 3, u'INGR. MENSUAL', fuentenormalnegcent2)
            ws.write_merge(5, 6, 4, 4, u'ENERO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 5, 7, u'ENERO PAGADO', fuentenormalnegcent2)  # +4
            ws.write_merge(6, 6, 5, 5, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 6, 6, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 7, 7, u'PROGRAMA', fuentenormalnegcent2)  # +4
            ws.write_merge(5, 6, 8, 8, u'FEBRERO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 9, 11, u'FEBRERO PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 9, 9, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 10, 10, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 11, 11, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 12, 12, u'MARZO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 13, 15, u'MARZO PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 13, 13, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 14, 14, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 15, 15, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 16, 16, u'ABRIL PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 17, 19, u'ABRIL PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 17, 17, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 18, 18, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 19, 19, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 20, 20, u'MAYO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 21, 23, u'MAYO PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 21, 21, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 22, 22, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 23, 23, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 24, 24, u'JUNIO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 25, 27, u'JUNIO PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 25, 25, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 26, 26, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 27, 27, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 28, 28, u'JULIO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 29, 31, u'JULIO PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 29, 29, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 30, 30, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 31, 31, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 32, 32, u'AGOSTO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 33, 35, u'AGOSTO PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 33, 33, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 34, 34, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 35, 35, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 36, 36, u'SEPTIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 37, 39, u'SEPTIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 37, 37, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 38, 38, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 39, 39, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 40, 40, u'OCTUBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 41, 43, u'OCTUBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 41, 41, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 42, 42, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 43, 43, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 44, 44, u'NOVIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 45, 47, u'NOVIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 45, 45, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 46, 46, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 47, 47, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 48, 48, u'DICIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 5, 49, 51, u'DICIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(6, 6, 49, 49, u'PROG.UNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 50, 50, u'PROG.EPUNEMI', fuentenormalnegcent2)
            ws.write_merge(6, 6, 51, 51, u'PROGRAMA', fuentenormalnegcent2)
            ws.write_merge(5, 6, 52, 52, u'TOTAL PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(5, 6, 53, 53, u'TOTAL PAGADO PROGRAMA', fuentenormalnegcent2)

            row_num = 7
            valor_total_bruto = 0
            valor_mes1_pagado = 0
            valor_mes1_pagado_unemi = 0
            valor_mes1_pagado_epunemi = 0
            valor_mes2_pagado = 0
            valor_mes2_pagado_unemi = 0
            valor_mes2_pagado_epunemi = 0
            valor_mes3_pagado = 0
            valor_mes3_pagado_unemi = 0
            valor_mes3_pagado_epunemi = 0
            valor_mes4_pagado = 0
            valor_mes4_pagado_unemi = 0
            valor_mes4_pagado_epunemi = 0
            valor_mes5_pagado = 0
            valor_mes5_pagado_unemi = 0
            valor_mes5_pagado_epunemi = 0
            valor_mes6_pagado = 0
            valor_mes6_pagado_unemi = 0
            valor_mes6_pagado_epunemi = 0
            valor_mes7_pagado = 0
            valor_mes7_pagado_unemi = 0
            valor_mes7_pagado_epunemi = 0
            valor_mes8_pagado = 0
            valor_mes8_pagado_unemi = 0
            valor_mes8_pagado_epunemi = 0
            valor_mes9_pagado = 0
            valor_mes9_pagado_unemi = 0
            valor_mes9_pagado_epunemi = 0
            valor_mes10_pagado = 0
            valor_mes10_pagado_unemi = 0
            valor_mes10_pagado_epunemi = 0
            valor_mes11_pagado = 0
            valor_mes11_pagado_unemi = 0
            valor_mes11_pagado_epunemi = 0
            valor_mes12_pagado = 0
            valor_mes12_pagado_unemi = 0
            valor_mes12_pagado_epunemi = 0

            valor_total_neto = 0
            valor_mes1_neto = 0
            valor_mes2_neto = 0
            valor_mes3_neto = 0
            valor_mes4_neto = 0
            valor_mes5_neto = 0
            valor_mes6_neto = 0
            valor_mes7_neto = 0
            valor_mes8_neto = 0
            valor_mes9_neto = 0
            valor_mes10_neto = 0
            valor_mes11_neto = 0
            valor_mes12_neto = 0

            pagos = Pago.objects.filter(pagoliquidacion__isnull=True,

                                        fecha__gte=fechadesde,
                                        fecha__lte=fechahasta,

                                        rubro__matricula__nivel__periodo_id__in=lista_periodos,
                                        rubro__matricula__inscripcion__carrera_id__in=lista_carreras,
                                        status=True,
                                        rubro__status=True
                                        ).exclude(factura__valida=False).order_by('fecha',
                                                                                  'rubro__matricula__nivel__periodo_id',
                                                                                  'rubro__matricula__inscripcion__carrera')

            pagosunemi = pagos.filter(idpagoepunemi=0).annotate(anio=ExtractYear('fecha'),
                                                                mes=ExtractMonth('fecha')).values_list('anio', 'mes',
                                                                                                       'rubro__matricula__nivel__periodo_id',
                                                                                                       'rubro__matricula__inscripcion__carrera_id').annotate(
                tpagado=Sum('valortotal'))
            pagosepunemi = pagos.filter(~Q(idpagoepunemi=0)).annotate(anio=ExtractYear('fecha'),
                                                                      mes=ExtractMonth('fecha')).values_list('anio',
                                                                                                             'mes',
                                                                                                             'rubro__matricula__nivel__periodo_id',
                                                                                                             'rubro__matricula__inscripcion__carrera_id').annotate(
                tpagado=Sum('valortotal'))
            pagos = pagos.annotate(anio=ExtractYear('fecha'), mes=ExtractMonth('fecha')).values_list('anio', 'mes',
                                                                                                     'rubro__matricula__nivel__periodo_id',
                                                                                                     'rubro__matricula__inscripcion__carrera_id').annotate(
                tpagado=Sum('valortotal'))

            for p in Periodo.objects.filter(pk__in=lista_periodos).order_by('id'):
                for carrera in Carrera.objects.filter(inscripcion__matricula__nivel__periodo=p,
                                                      pk__in=lista_carreras).distinct().order_by('nombre'):

                    porcentaje = 0
                    for lp in listaporcentajes:
                        if int(lp['periodo']) == p.id and int(lp['carrera']) == carrera.id:
                            porcentaje = Decimal(lp['porcentaje']).quantize(Decimal('.01'))
                            break

                    ws.write_merge(row_num, row_num, 0, 0, p.nombre, fuentenormal2)
                    ws.write_merge(row_num, row_num, 1, 1, carrera.nombre, fuentenormal2)
                    mesdesde = 1
                    valor_mes_pagado = 0
                    valor_mes_pagado_une = 0
                    valor_mes_pagado_epu = 0
                    valor_mes_neto = 0
                    bandera_mes = 0
                    fila = 4
                    while mesdesde <= 12:
                        valor_bruto = 0
                        valor_pagado = 0

                        # Total generado rubros
                        totalproyectado = Decimal(
                            null_to_decimal(Rubro.objects.filter(matricula__nivel__periodo=p,
                                                                 matricula__inscripcion__carrera=carrera,
                                                                 status=True,
                                                                 fechavence__year=anio, fechavence__month=mesdesde,

                                                                 fechavence__gte=fechadesde, fechavence__lte=fechahasta

                                                                 ).aggregate(
                                valor=Sum('valor'))['valor'], 2)).quantize(Decimal('.01'))

                        # Total anulado rubros
                        totalanulado = Decimal(
                            null_to_decimal(Pago.objects.filter(rubro__matricula__nivel__periodo=p,
                                                                rubro__matricula__inscripcion__carrera=carrera,
                                                                status=True,
                                                                rubro__status=True,
                                                                rubro__fechavence__year=anio,
                                                                rubro__fechavence__month=mesdesde,

                                                                rubro__fechavence__gte=fechadesde,
                                                                rubro__fechavence__lte=fechahasta,

                                                                factura__valida=False,
                                                                factura__status=True).aggregate(
                                valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))

                        # Total liquidado rubros
                        totalliquidado = Decimal(
                            null_to_decimal(Pago.objects.filter(rubro__matricula__nivel__periodo=p,
                                                                rubro__matricula__inscripcion__carrera=carrera,
                                                                status=True,
                                                                rubro__status=True,
                                                                rubro__fechavence__year=anio,
                                                                rubro__fechavence__month=mesdesde,

                                                                rubro__fechavence__gte=fechadesde,
                                                                rubro__fechavence__lte=fechahasta,

                                                                pagoliquidacion__isnull=False,
                                                                pagoliquidacion__status=True).aggregate(
                                valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))

                        valor_bruto = totalproyectado - (totalanulado + totalliquidado)

                        totalpagado = 0
                        totalpagadoune = 0
                        totalpagadoepu = 0
                        aniofila = 0
                        mesfila = 0
                        periodofila = 0
                        carrerafila = 0

                        for rp in pagos:
                            if rp[0] == anio and rp[1] == mesdesde and rp[2] == p.id and rp[3] == carrera.id:
                                totalpagado += Decimal(rp[4])
                                aniofila = rp[0]
                                mesfila = rp[1]
                                periodofila = rp[2]
                                carrerafila = rp[3]
                            elif aniofila != 0 and aniofila != rp[0] and mesfila != rp[1] and periodofila != rp[
                                2] and carrerafila != rp[3]:
                                break

                        aniofila = 0
                        mesfila = 0
                        periodofila = 0
                        carrerafila = 0

                        for ru in pagosunemi:
                            if ru[0] == anio and ru[1] == mesdesde and ru[2] == p.id and ru[3] == carrera.id:
                                totalpagadoune += Decimal(ru[4])
                                aniofila = ru[0]
                                mesfila = ru[1]
                                periodofila = ru[2]
                                carrerafila = ru[3]
                            elif aniofila != 0 and aniofila != ru[0] and mesfila != ru[1] and periodofila != ru[
                                2] and carrerafila != ru[3]:
                                break

                        aniofila = 0
                        mesfila = 0
                        periodofila = 0
                        carrerafila = 0

                        for re in pagosepunemi:
                            if re[0] == anio and re[1] == mesdesde and re[2] == p.id and re[3] == carrera.id:
                                totalpagadoepu += Decimal(re[4])
                                aniofila = re[0]
                                mesfila = re[1]
                                periodofila = re[2]
                                carrerafila = re[3]
                            elif aniofila != 0 and aniofila != re[0] and mesfila != re[1] and periodofila != re[
                                2] and carrerafila != re[3]:
                                break

                        valor_pagado = totalpagado
                        valor_pagado_unemi = totalpagadoune
                        valor_pagado_epunemi = totalpagadoepu

                        valor_neto = Decimal(0).quantize(Decimal('.01'))
                        valor_originsal = valor_bruto
                        if valor_bruto > 0:
                            valor_descuento = Decimal(((valor_bruto * porcentaje) / 100)).quantize(Decimal('.01'))

                            valor_neto = valor_bruto - valor_descuento

                        if mesdesde == 1:
                            valor_mes1_pagado += valor_pagado
                            valor_mes1_pagado_unemi += valor_pagado_unemi
                            valor_mes1_pagado_epunemi += valor_pagado_epunemi
                            valor_mes1_neto += valor_neto
                        elif mesdesde == 2:
                            valor_mes2_pagado += valor_pagado
                            valor_mes2_pagado_unemi += valor_pagado_unemi
                            valor_mes2_pagado_epunemi += valor_pagado_epunemi
                            valor_mes2_neto += valor_neto
                        elif mesdesde == 3:
                            valor_mes3_pagado += valor_pagado
                            valor_mes3_pagado_unemi += valor_pagado_unemi
                            valor_mes3_pagado_epunemi += valor_pagado_epunemi
                            valor_mes3_neto += valor_neto
                        elif mesdesde == 4:
                            valor_mes4_pagado += valor_pagado
                            valor_mes4_pagado_unemi += valor_pagado_unemi
                            valor_mes4_pagado_epunemi += valor_pagado_epunemi
                            valor_mes4_neto += valor_neto
                        elif mesdesde == 5:
                            valor_mes5_pagado += valor_pagado
                            valor_mes5_pagado_unemi += valor_pagado_unemi
                            valor_mes5_pagado_epunemi += valor_pagado_epunemi
                            valor_mes5_neto += valor_neto
                        elif mesdesde == 6:
                            valor_mes6_pagado += valor_pagado
                            valor_mes6_pagado_unemi += valor_pagado_unemi
                            valor_mes6_pagado_epunemi += valor_pagado_epunemi
                            valor_mes6_neto += valor_neto
                        elif mesdesde == 7:
                            valor_mes7_pagado += valor_pagado
                            valor_mes7_pagado_unemi += valor_pagado_unemi
                            valor_mes7_pagado_epunemi += valor_pagado_epunemi
                            valor_mes7_neto += valor_neto
                        elif mesdesde == 8:
                            valor_mes8_pagado += valor_pagado
                            valor_mes8_pagado_unemi += valor_pagado_unemi
                            valor_mes8_pagado_epunemi += valor_pagado_epunemi
                            valor_mes8_neto += valor_neto
                        elif mesdesde == 9:
                            valor_mes9_pagado += valor_pagado
                            valor_mes9_pagado_unemi += valor_pagado_unemi
                            valor_mes9_pagado_epunemi += valor_pagado_epunemi
                            valor_mes9_neto += valor_neto
                        elif mesdesde == 10:
                            valor_mes10_pagado += valor_pagado
                            valor_mes10_pagado_unemi += valor_pagado_unemi
                            valor_mes10_pagado_epunemi += valor_pagado_epunemi
                            valor_mes10_neto += valor_neto
                        elif mesdesde == 11:
                            valor_mes11_pagado += valor_pagado
                            valor_mes11_pagado_unemi += valor_pagado_unemi
                            valor_mes11_pagado_epunemi += valor_pagado_epunemi
                            valor_mes11_neto += valor_neto
                        elif mesdesde == 12:
                            valor_mes12_pagado += valor_pagado
                            valor_mes12_pagado_unemi += valor_pagado_unemi
                            valor_mes12_pagado_epunemi += valor_pagado_epunemi
                            valor_mes12_neto += valor_neto
                        if valor_bruto > 0:
                            bandera_mes += 1

                        valor_mes_pagado += valor_pagado
                        valor_mes_pagado_une += valor_pagado_unemi
                        valor_mes_pagado_epu += valor_pagado_epunemi

                        valor_mes_neto += valor_neto
                        ws.write_merge(row_num, row_num, fila, fila, valor_neto, fuentemoneda)
                        fila += 1
                        ws.write_merge(row_num, row_num, fila, fila, valor_pagado_unemi, fuentemonedafv)
                        fila += 1
                        ws.write_merge(row_num, row_num, fila, fila, valor_pagado_epunemi, fuentemonedafn)
                        fila += 1
                        ws.write_merge(row_num, row_num, fila, fila, valor_pagado, fuentemoneda)
                        fila += 1

                        mesdesde += 1
                    ws.write_merge(row_num, row_num, 52, 52, valor_mes_neto, fuentemoneda)
                    ws.write_merge(row_num, row_num, 53, 53, valor_mes_pagado, fuentemoneda)
                    ws.write_merge(row_num, row_num, 2, 2, valor_mes_neto, fuentemoneda)
                    ws.write_merge(row_num, row_num, 3, 3, str(bandera_mes), fuentenormal2)

                    row_num += 1

            ws.write_merge(row_num, row_num, 4, 4, valor_mes1_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 5, 5, valor_mes1_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 6, 6, valor_mes1_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 7, 7, valor_mes1_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 8, 8, valor_mes2_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 9, 9, valor_mes2_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 10, 10, valor_mes2_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 11, 11, valor_mes2_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 12, 12, valor_mes3_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 13, 13, valor_mes3_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 14, 14, valor_mes3_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 15, 15, valor_mes3_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 16, 16, valor_mes4_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 17, 17, valor_mes4_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 18, 18, valor_mes4_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 19, 19, valor_mes4_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 20, 20, valor_mes5_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 21, 21, valor_mes5_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 22, 22, valor_mes5_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 23, 23, valor_mes5_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 24, 24, valor_mes6_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 25, 25, valor_mes6_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 26, 26, valor_mes6_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 27, 27, valor_mes6_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 28, 28, valor_mes7_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 29, 29, valor_mes7_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 30, 30, valor_mes7_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 31, 31, valor_mes7_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 32, 32, valor_mes8_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 33, 33, valor_mes8_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 34, 34, valor_mes8_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 35, 35, valor_mes8_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 36, 36, valor_mes9_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 37, 37, valor_mes9_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 38, 38, valor_mes9_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 39, 39, valor_mes9_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 40, 40, valor_mes10_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 41, 41, valor_mes10_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 42, 42, valor_mes10_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 43, 43, valor_mes10_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 44, 44, valor_mes11_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 45, 45, valor_mes11_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 46, 46, valor_mes11_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 47, 47, valor_mes11_pagado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 48, 48, valor_mes12_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 49, 49, valor_mes12_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 50, 50, valor_mes12_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 51, 51, valor_mes12_pagado, fuentemonedaneg)

            sumatotalproyectado = valor_mes1_neto + valor_mes2_neto + valor_mes3_neto + valor_mes4_neto + valor_mes5_neto + valor_mes6_neto + valor_mes7_neto + valor_mes8_neto + valor_mes9_neto + valor_mes10_neto + valor_mes11_neto + valor_mes12_neto
            sumatotalpagado = valor_mes1_pagado + valor_mes2_pagado + valor_mes3_pagado + valor_mes4_pagado + valor_mes5_pagado + valor_mes6_pagado + valor_mes7_pagado + valor_mes8_pagado + valor_mes9_pagado + valor_mes10_pagado + valor_mes11_pagado + valor_mes12_pagado

            ws.write_merge(row_num, row_num, 52, 52, sumatotalproyectado, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 53, 53, sumatotalpagado, fuentemonedaneg)

            columns = [
                (u"", 9000),
                (u"", 9000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000)]
            row_num += 2

            # for col_num in range(3:97):

            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style2)
                ws.col(col_num).width = columns[col_num][1]

            columns = [
                (u"", 9000),
                (u"", 9000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000),
                (u"", 5000)]
            row_num += 1
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style2)
                ws.col(col_num).width = columns[col_num][1]
            row_num += 1
            # cartera vencida

            ws.write_merge(row_num, row_num, 0, 0, u'NO ADMITIDOS', fuentenormalneg2)
            row_num += 1
            ws.write_merge(row_num, row_num, 0, 0, u'RUBROS', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 1, 1, u'RUBROS', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 2, 2, u'', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 3, 3, u'', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 4, 4, u'ENERO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 5, 5, u'ENERO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 6, 6, u'ENERO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 7, 7, u'ENERO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 8, 8, u'FEBRERO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 9, 9, u'FEBRERO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 10, 10, u'FEBRERO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 11, 11, u'FEBRERO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 12, 12, u'MARZO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 13, 13, u'MARZO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 14, 14, u'MARZO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 15, 15, u'MARZO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 16, 16, u'ABRIL PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 17, 17, u'ABRIL PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 18, 18, u'ABRIL PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 19, 19, u'ABRIL PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 20, 20, u'MAYO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 21, 21, u'MAYO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 22, 22, u'MAYO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 23, 23, u'MAYO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 24, 24, u'JUNIO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 25, 25, u'JUNIO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 26, 26, u'JUNIO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 27, 27, u'JUNIO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 28, 28, u'JULIO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 29, 29, u'JULIO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 30, 30, u'JULIO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 31, 31, u'JULIO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 32, 32, u'AGOSTO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 33, 33, u'AGOSTO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 34, 34, u'AGOSTO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 35, 35, u'AGOSTO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 36, 36, u'SEPTIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 37, 37, u'SEPTIEMBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 38, 38, u'SEPTIEMBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 39, 39, u'SEPTIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 40, 40, u'OCTUBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 41, 41, u'OCTUBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 42, 42, u'OCTUBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 43, 43, u'OCTUBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 44, 44, u'NOVIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 45, 45, u'NOVIEMBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 46, 46, u'NOVIEMBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 47, 47, u'NOVIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 48, 48, u'DICIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 49, 49, u'DICIEMBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 50, 50, u'DICIEMBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 51, 51, u'DICIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 52, 52, u'TOTAL PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 53, 53, u'TOTAL PAGADO', fuentenormalnegcent2)
            valor_total = 0
            valor_total_bruto = 0
            valor_mes1_pagado = 0
            valor_mes1_pagado_unemi = 0
            valor_mes1_pagado_epunemi = 0
            valor_mes2_pagado = 0
            valor_mes2_pagado_unemi = 0
            valor_mes2_pagado_epunemi = 0
            valor_mes3_pagado = 0
            valor_mes3_pagado_unemi = 0
            valor_mes3_pagado_epunemi = 0
            valor_mes4_pagado = 0
            valor_mes4_pagado_unemi = 0
            valor_mes4_pagado_epunemi = 0
            valor_mes5_pagado = 0
            valor_mes5_pagado_unemi = 0
            valor_mes5_pagado_epunemi = 0
            valor_mes6_pagado = 0
            valor_mes6_pagado_unemi = 0
            valor_mes6_pagado_epunemi = 0
            valor_mes7_pagado = 0
            valor_mes7_pagado_unemi = 0
            valor_mes7_pagado_epunemi = 0
            valor_mes8_pagado = 0
            valor_mes8_pagado_unemi = 0
            valor_mes8_pagado_epunemi = 0
            valor_mes9_pagado = 0
            valor_mes9_pagado_unemi = 0
            valor_mes9_pagado_epunemi = 0
            valor_mes10_pagado = 0
            valor_mes10_pagado_unemi = 0
            valor_mes10_pagado_epunemi = 0
            valor_mes11_pagado = 0
            valor_mes11_pagado_unemi = 0
            valor_mes11_pagado_epunemi = 0
            valor_mes12_pagado = 0
            valor_mes12_pagado_unemi = 0
            valor_mes12_pagado_epunemi = 0

            valor_total_neto = 0
            valor_mes1_neto = 0
            valor_mes2_neto = 0
            valor_mes3_neto = 0
            valor_mes4_neto = 0
            valor_mes5_neto = 0
            valor_mes6_neto = 0
            valor_mes7_neto = 0
            valor_mes8_neto = 0
            valor_mes9_neto = 0
            valor_mes10_neto = 0
            valor_mes11_neto = 0
            valor_mes12_neto = 0
            row_num += 1
            # valornoadmitidos = Decimal(null_to_decimal(Rubro.objects.filter(pago__factura__fecha__year=anio, pago__factura__valida=True, status=True, matricula__isnull=True, tipo__in=codigo_rubro).aggregate(valor=Sum('pago__subtotal0'))['valor'], 2)).quantize(Decimal('.01'))

            for lista in listadorubros:
                ws.write_merge(row_num, row_num, 0, 0, 'RUBRO', fuentenormal2)
                ws.write_merge(row_num, row_num, 1, 1, lista[1], fuentenormal2)
                # ws.write_merge(row_num, row_num, 2, 2, 0, font_style2)
                # ws.write_merge(row_num, row_num, 3, 3, 0, font_style2)
                mesdesde = 1
                valor_mes_pagado = 0
                valor_mes_neto = 0
                bandera_mes = 0

                fila = 4
                while mesdesde <= 12:
                    valor_bruto = 0
                    valor_pagado = 0

                    # valor_bruto = Decimal(null_to_decimal(Rubro.objects.filter(tipo_id=lista[0], fechavence__year=anio, matricula__isnull=True, fechavence__month=mesdesde, status=True).aggregate(valor=Sum('saldo'))['valor'], 2)).quantize(Decimal('.01'))
                    valor_bruto = Decimal(null_to_decimal(
                        Rubro.objects.filter(tipo_id=lista[0], fechavence__year=anio, matricula__isnull=True,
                                             fechavence__month=mesdesde, fechavence__gte=fechadesde,
                                             fechavence__lte=fechahasta, status=True).aggregate(valor=Sum('saldo'))[
                            'valor'], 2)).quantize(
                        Decimal('.01'))

                    # pagosotros = Pago.objects.filter(rubro__matricula__isnull=True, rubro__tipo_id=lista[0], status=True, rubro__status=True, fecha__year=anio, fecha__month=mesdesde)
                    pagosotros = Pago.objects.filter(rubro__matricula__isnull=True, rubro__tipo_id=lista[0],
                                                     status=True, rubro__status=True, fecha__year=anio,
                                                     fecha__month=mesdesde, fecha__gte=fechadesde,
                                                     fecha__lte=fechahasta)

                    valor_pagado = Decimal(
                        null_to_decimal(pagosotros.aggregate(valor=Sum('valortotal'))['valor'], 2)).quantize(
                        Decimal('.01'))
                    valor_pagado_unemi = Decimal(
                        null_to_decimal(pagosotros.filter(idpagoepunemi=0).aggregate(valor=Sum('valortotal'))['valor'],
                                        2)).quantize(Decimal('.01'))
                    valor_pagado_epunemi = Decimal(null_to_decimal(
                        pagosotros.filter(~Q(idpagoepunemi=0)).aggregate(valor=Sum('valortotal'))['valor'],
                        2)).quantize(Decimal('.01'))

                    valor_neto = Decimal(0).quantize(Decimal('.01'))

                    if mesdesde == 1:
                        valor_mes1_pagado += valor_pagado
                        valor_mes1_pagado_unemi += valor_pagado_unemi
                        valor_mes1_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 2:
                        valor_mes2_pagado += valor_pagado
                        valor_mes2_pagado_unemi += valor_pagado_unemi
                        valor_mes2_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 3:
                        valor_mes3_pagado += valor_pagado
                        valor_mes3_pagado_unemi += valor_pagado_unemi
                        valor_mes3_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 4:
                        valor_mes4_pagado += valor_pagado
                        valor_mes4_pagado_unemi += valor_pagado_unemi
                        valor_mes4_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 5:
                        valor_mes5_pagado += valor_pagado
                        valor_mes5_pagado_unemi += valor_pagado_unemi
                        valor_mes5_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 6:
                        valor_mes6_pagado += valor_pagado
                        valor_mes6_pagado_unemi += valor_pagado_unemi
                        valor_mes6_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 7:
                        valor_mes7_pagado += valor_pagado
                        valor_mes7_pagado_unemi += valor_pagado_unemi
                        valor_mes7_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 8:
                        valor_mes8_pagado += valor_pagado
                        valor_mes8_pagado_unemi += valor_pagado_unemi
                        valor_mes8_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 9:
                        valor_mes9_pagado += valor_pagado
                        valor_mes9_pagado_unemi += valor_pagado_unemi
                        valor_mes9_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 10:
                        valor_mes10_pagado += valor_pagado
                        valor_mes10_pagado_unemi += valor_pagado_unemi
                        valor_mes10_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 11:
                        valor_mes11_pagado += valor_pagado
                        valor_mes11_pagado_unemi += valor_pagado_unemi
                        valor_mes11_pagado_epunemi += valor_pagado_epunemi
                    elif mesdesde == 12:
                        valor_mes12_pagado += valor_pagado
                        valor_mes12_pagado_unemi += valor_pagado_unemi
                        valor_mes12_pagado_epunemi += valor_pagado_epunemi

                    if mesdesde == 1: valor_mes1_neto += valor_neto
                    if mesdesde == 2: valor_mes2_neto += valor_neto
                    if mesdesde == 3: valor_mes3_neto += valor_neto
                    if mesdesde == 4: valor_mes4_neto += valor_neto
                    if mesdesde == 5: valor_mes5_neto += valor_neto
                    if mesdesde == 6: valor_mes6_neto += valor_neto
                    if mesdesde == 7: valor_mes7_neto += valor_neto
                    if mesdesde == 8: valor_mes8_neto += valor_neto
                    if mesdesde == 9: valor_mes9_neto += valor_neto
                    if mesdesde == 10: valor_mes10_neto += valor_neto
                    if mesdesde == 11: valor_mes11_neto += valor_neto
                    if mesdesde == 12: valor_mes12_neto += valor_neto

                    if valor_bruto > 0:
                        bandera_mes += 1
                    valor_mes_pagado += valor_pagado
                    valor_mes_neto += valor_neto
                    ws.write_merge(row_num, row_num, fila, fila, valor_neto, fuentemoneda)
                    fila += 1
                    ws.write_merge(row_num, row_num, fila, fila, valor_pagado_unemi, fuentemonedafv)
                    fila += 1
                    ws.write_merge(row_num, row_num, fila, fila, valor_pagado_epunemi, fuentemonedafn)
                    fila += 1
                    ws.write_merge(row_num, row_num, fila, fila, valor_pagado, fuentemoneda)
                    fila += 1

                    mesdesde += 1

                ws.write_merge(row_num, row_num, 52, 52, valor_mes_neto, fuentemoneda)
                ws.write_merge(row_num, row_num, 53, 53, valor_mes_pagado, fuentemoneda)
                ws.write_merge(row_num, row_num, 2, 2, valor_mes_neto, fuentemoneda)
                ws.write_merge(row_num, row_num, 3, 3, str(bandera_mes), fuentenormal2)
                row_num += 1

            ws.write_merge(row_num, row_num, 4, 4, valor_mes1_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 5, 5, valor_mes1_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 6, 6, valor_mes1_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 7, 7, valor_mes1_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 8, 8, valor_mes2_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 9, 9, valor_mes2_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 10, 10, valor_mes2_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 11, 11, valor_mes2_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 12, 12, valor_mes3_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 13, 13, valor_mes3_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 14, 14, valor_mes3_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 15, 15, valor_mes3_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 16, 16, valor_mes4_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 17, 17, valor_mes4_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 18, 18, valor_mes4_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 19, 19, valor_mes4_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 20, 20, valor_mes5_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 21, 21, valor_mes5_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 22, 22, valor_mes5_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 23, 23, valor_mes5_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 24, 24, valor_mes6_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 25, 25, valor_mes6_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 26, 26, valor_mes6_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 27, 27, valor_mes6_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 28, 28, valor_mes7_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 29, 29, valor_mes7_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 30, 30, valor_mes7_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 31, 31, valor_mes7_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 32, 32, valor_mes8_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 33, 33, valor_mes8_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 34, 34, valor_mes8_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 35, 35, valor_mes8_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 36, 36, valor_mes9_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 37, 37, valor_mes9_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 38, 38, valor_mes9_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 39, 39, valor_mes9_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 40, 40, valor_mes10_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 41, 41, valor_mes10_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 42, 42, valor_mes10_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 43, 43, valor_mes10_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 44, 44, valor_mes11_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 45, 45, valor_mes11_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 46, 46, valor_mes11_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 47, 47, valor_mes11_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 48, 48, valor_mes12_neto, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 49, 49, valor_mes12_pagado_unemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 50, 50, valor_mes12_pagado_epunemi, fuentemonedaneg)
            ws.write_merge(row_num, row_num, 51, 51, valor_mes12_pagado, fuentemonedaneg)

            ws.write_merge(row_num, row_num, 52, 52,
                           valor_mes1_neto + valor_mes2_neto + valor_mes3_neto + valor_mes4_neto + valor_mes5_neto + valor_mes6_neto + valor_mes7_neto + valor_mes8_neto + valor_mes9_neto + valor_mes10_neto + valor_mes11_neto + valor_mes12_neto,
                           fuentemonedaneg)
            ws.write_merge(row_num, row_num, 53, 53,
                           valor_mes1_pagado + valor_mes2_pagado + valor_mes3_pagado + valor_mes4_pagado + valor_mes5_pagado + valor_mes6_pagado + valor_mes7_pagado + valor_mes8_pagado + valor_mes9_pagado + valor_mes10_pagado + valor_mes11_pagado + valor_mes12_pagado,
                           fuentemonedaneg)

            row_num += 3
            ws.write_merge(row_num, row_num, 0, 0, u'RUBROS ADICIONALES', fuentenormalneg2)
            row_num += 1

            ws.write_merge(row_num, row_num, 0, 3, u'RUBROS', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 4, 4, u'ENERO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 5, 5, u'ENERO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 6, 6, u'ENERO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 7, 7, u'ENERO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 8, 8, u'FEBRERO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 9, 9, u'FEBRERO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 10, 10, u'FEBRERO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 11, 11, u'FEBRERO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 12, 12, u'MARZO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 13, 13, u'MARZO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 14, 14, u'MARZO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 15, 15, u'MARZO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 16, 16, u'ABRIL PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 17, 17, u'ABRIL PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 18, 18, u'ABRIL PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 19, 19, u'ABRIL PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 20, 20, u'MAYO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 21, 21, u'MAYO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 22, 22, u'MAYO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 23, 23, u'MAYO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 24, 24, u'JUNIO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 25, 25, u'JUNIO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 26, 26, u'JUNIO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 27, 27, u'JUNIO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 28, 28, u'JULIO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 29, 29, u'JULIO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 30, 30, u'JULIO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 31, 31, u'JULIO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 32, 32, u'AGOSTO PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 33, 33, u'AGOSTO PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 34, 34, u'AGOSTO PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 35, 35, u'AGOSTO PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 36, 36, u'SEPTIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 37, 37, u'SEPTIEMBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 38, 38, u'SEPTIEMBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 39, 39, u'SEPTIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 40, 40, u'OCTUBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 41, 41, u'OCTUBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 42, 42, u'OCTUBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 43, 43, u'OCTUBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 44, 44, u'NOVIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 45, 45, u'NOVIEMBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 46, 46, u'NOVIEMBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 47, 47, u'NOVIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 48, 48, u'DICIEMBRE PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 49, 49, u'DICIEMBRE PAG.UNE', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 50, 50, u'DICIEMBRE PAG.EPU', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 51, 51, u'DICIEMBRE PAGADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 52, 52, u'TOTAL PROYECTADO', fuentenormalnegcent2)
            ws.write_merge(row_num, row_num, 53, 53, u'TOTAL PAGADO', fuentenormalnegcent2)

            totalesadicionales = [[Decimal(0.0) for i in range(3)] for j in range(12)]

            row_num += 1

            for rubroadicional in rubrosadicionales:
                ws.write_merge(row_num, row_num, 0, 3, rubroadicional[1], fuentenormal2)
                col = 4
                totalrubroadicional = Decimal(0)
                for nmes in range(1, 13):
                    ws.write_merge(row_num, row_num, col, col, 0, fuentemoneda)

                    valor_adicional_mes = Decimal(null_to_decimal(
                        pagosadicionales.filter(rubro__tipo_id=rubroadicional[0], fecha__month=nmes).aggregate(
                            valor=Sum('valortotal'))['valor'], 2)).quantize(Decimal('.01'))
                    valor_adicional_mes_une = Decimal(null_to_decimal(
                        pagosadicionales.filter(idpagoepunemi=0, rubro__tipo_id=rubroadicional[0],
                                                fecha__month=nmes).aggregate(valor=Sum('valortotal'))['valor'],
                        2)).quantize(Decimal('.01'))
                    valor_adicional_mes_epu = Decimal(null_to_decimal(
                        pagosadicionales.filter(~Q(idpagoepunemi=0), rubro__tipo_id=rubroadicional[0],
                                                fecha__month=nmes).aggregate(valor=Sum('valortotal'))['valor'],
                        2)).quantize(Decimal('.01'))

                    ws.write_merge(row_num, row_num, col + 1, col + 1, valor_adicional_mes_une, fuentemonedafv)
                    ws.write_merge(row_num, row_num, col + 2, col + 2, valor_adicional_mes_epu, fuentemonedafn)
                    ws.write_merge(row_num, row_num, col + 3, col + 3, valor_adicional_mes, fuentemoneda)
                    col += 4

                    totalesadicionales[nmes - 1][0] += valor_adicional_mes_une
                    totalesadicionales[nmes - 1][1] += valor_adicional_mes_epu
                    totalesadicionales[nmes - 1][2] += valor_adicional_mes
                    totalrubroadicional += valor_adicional_mes

                ws.write_merge(row_num, row_num, col, col, 0, fuentemoneda)
                ws.write_merge(row_num, row_num, col + 1, col + 1, totalrubroadicional, fuentemoneda)
                row_num += 1

            col = 4
            for nmes in range(1, 13):
                ws.write_merge(row_num, row_num, col, col, 0, fuentemonedaneg)
                ws.write_merge(row_num, row_num, col + 1, col + 1, totalesadicionales[nmes - 1][0], fuentemonedaneg)
                ws.write_merge(row_num, row_num, col + 2, col + 2, totalesadicionales[nmes - 1][1], fuentemonedaneg)
                ws.write_merge(row_num, row_num, col + 3, col + 3, totalesadicionales[nmes - 1][2], fuentemonedaneg)
                col += 4

            ws.write_merge(row_num, row_num, col, col, 0, fuentemonedaneg)
            ws.write_merge(row_num, row_num, col + 1, col + 1, valoradicionales, fuentemonedaneg)

            wb.save(filename)

            # Notificar al usuario
            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.cuerpo = 'Reporte Excel finalizado'
                notificacion.url = urlarchivo
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo='Reporte Excel finalizado',
                    titulo=tituloreporte,
                    destinatario=personanotifica,
                    url=urlarchivo,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False
                )
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte Excel finalizado",
                "body": tituloreporte,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": urlarchivo,
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "El <b>{}</b> ha sido generado con éxito".format(tituloreporte)
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.error = True
                notificacion.titulo = "Error al generar {}".format(tituloreporte)
                notificacion.cuerpo = textoerror
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo="Error al generar Reporte Excel",
                    titulo="Error al generar {}".format(tituloreporte),
                    destinatario=personanotifica,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False,
                    error=True)
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error al generar Reporte Excel",
                "body": "Error al generar {}".format(tituloreporte),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "Error al generar <b>{}</b>. Error: {}".format(tituloreporte, textoerror),
                "error": True
            }, ttl=500)


class descarga_masica_documentos_familiares_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            periodo = data['periodo']
            pers = Persona.objects.get(usuario_id=request.user.pk)
            personasgastos = periodo.personasgasto()
            directory = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)
            name_zip = f'docsfamiliares_{periodo.anio}.zip'
            url = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos', name_zip)
            fantasy_zip = zipfile.ZipFile(url, 'w')
            for pg in personasgastos:
                cont = 0
                carpeta = unidecode(pg.persona.nombre_completo_minus())
                acta = Archivo.objects.filter(tipo_id=19, persona=pg.persona, aprobado=True).order_by('id').last()
                if acta and acta.archivo:
                    cont += 1
                    url_file = acta.archivo.url
                    name_file = carpeta.replace(" ", "_")
                    ext = url_file[url_file.rfind("."):].lower()
                    ruta_archivo_zip = os.path.join(carpeta, f'Acta_Proyeccion_{name_file}{ext}')
                    fantasy_zip.write(acta.archivo.path, ruta_archivo_zip)
                for f in pg.persona.familiares():
                    name_file = f.parentesco.nombre.title() + '_' + unidecode(f.nombre.title()).replace(' ', '_')
                    if f.cedulaidentidad or f.cartaconsentimiento or f.archivoautorizado or f.ceduladiscapacidad:
                        cont += 1
                        if f.cedulaidentidad:
                            url_file = f.cedulaidentidad.url
                            ext = url_file[url_file.rfind("."):].lower()
                            ruta_archivo_zip = os.path.join(carpeta, f'{cont}.Cedula_{name_file}{ext}')
                            fantasy_zip.write(f.cedulaidentidad.path, ruta_archivo_zip)
                        if f.cartaconsentimiento:
                            url_file = f.cartaconsentimiento.url
                            ext = url_file[url_file.rfind("."):].lower()
                            ruta_archivo_zip = os.path.join(carpeta, f'{cont}.Carta_Consentimiento_{name_file}{ext}')
                            fantasy_zip.write(f.cartaconsentimiento.path, ruta_archivo_zip)
                        if f.archivoautorizado:
                            url_file = f.archivoautorizado.url
                            ext = url_file[url_file.rfind("."):].lower()
                            ruta_archivo_zip = os.path.join(carpeta, f'{cont}.Autorizacion_{name_file}{ext}')
                            fantasy_zip.write(f.archivoautorizado.path, ruta_archivo_zip)
                        if f.ceduladiscapacidad:
                            url_file = f.ceduladiscapacidad.url
                            ext = url_file[url_file.rfind("."):].lower()
                            ruta_archivo_zip = os.path.join(carpeta, f'{cont}.Carnet_Discapacidad_{name_file}{ext}')
                            fantasy_zip.write(f.ceduladiscapacidad.path, ruta_archivo_zip)

            fantasy_zip.close()
            # NOTIFICACIÓN
            titulo = f'Archivo .zip de familiares del periodo {periodo.anio} generado con éxito.'
            cuerpo = f'Archivo .zip generado con éxito'
            url = f"{MEDIA_URL}talento_humano/docs_comprimidos/{name_zip}"
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su archivo .zip ha sido generado con éxito.'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

class descarga_masiva_actas_firmadas_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            pers = Persona.objects.get(usuario_id=request.user.pk)
            personasdecimo = RegistroDecimo.objects.filter(activo=True, status=True, estado=2)
            directory = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)
            name_zip = f'docsdecimo_{datetime.now().year}.zip'
            url = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos', name_zip)
            fantasy_zip = zipfile.ZipFile(url, 'w')
            for pd in personasdecimo:
                cont = 0
                carpeta = unidecode(pd.persona.nombre_completo_minus())
                if pd.archivo:
                    cont += 1
                    url_file = pd.archivo.url
                    name_file = carpeta.replace(" ", "_")
                    ext = url_file[url_file.rfind("."):].lower()
                    ruta_archivo_zip = os.path.join(carpeta, f'Acta_decimo_{name_file}{ext}')
                    fantasy_zip.write(pd.archivo.path, ruta_archivo_zip)
            fantasy_zip.close()
            # NOTIFICACIÓN
            titulo = f'Archivo .zip de décimo del año {datetime.now().year} generado con éxito.'
            cuerpo = f'Archivo .zip generado con éxito'
            url = f"{MEDIA_URL}talento_humano/docs_comprimidos/{name_zip}"
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su archivo .zip ha sido generado con éxito.'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class descarga_cargas_gastosp_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            periodo = data['periodo']
            fechainicio = data['fechainicio']
            fechafin = data['fechafin']
            todos = data['todos']
            pers = Persona.objects.get(usuario_id=request.user.pk)
            personasgastos = periodo.personasgasto()
            nombre_archivo = 'reporte_cargas' + random.randint(1, 10000).__str__() + '.xlsx'
            directory = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'reportes')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)
            directory = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'reportes', nombre_archivo)

            filtro = Q(status=True, actagenerada=True, aplicaproyeccion=True, personafamiliar__isnull=False)
            if not todos:
                filtro = filtro & (Q(fecha_creacion__lte=fechainicio,
                                     fecha_creacion__gte=fechafin) | \
                                   Q(fecha_modificacion__lte=fechainicio,
                                     fecha_modificacion__gte=fechafin))

            __author__ = 'Unemi'
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('cargas')
            ws.set_column(0, 20, 50)
            formatoceldagris = workbook.add_format(
                {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
            formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})
            ws.write('A1', 'IDENTIFICACIÓN', formatoceldagris)
            ws.write('B1', 'SERVIDOR', formatoceldagris)
            ws.write('C1', 'CARGO', formatoceldagris)
            ws.write('D1', 'CANT. CARGAS', formatoceldagris)
            ws.write('E1', 'MENORES 5 AÑOS', formatoceldagris)
            i = 2
            for per in personasgastos:
                familiares = per.persona.personadatosfamiliares_set.filter(filtro)
                if familiares:
                    ws.write('A%s' % i, str(per.persona.identificacion()), formatoceldaleft)
                    ws.write('B%s' % i, str(per.persona.nombre_completo_inverso_2()), formatoceldaleft)
                    ws.write('C%s' % i, str(per.persona.mi_cargo_actual().denominacionpuesto), formatoceldaleft)
                    ws.write('D%s' % i, str(familiares.count()), formatoceldaleft)
                    total_menor = 0
                    for fam in familiares:
                        if datetime.now().year - fam.nacimiento.year < 5:
                            total_menor = total_menor + 1
                    ws.write('E%s' % i, str(total_menor), formatoceldaleft)
                    i += 1
            workbook.close()
            # NOTIFICACIÓN
            url = f"{MEDIA_URL}talento_humano/reportes/{nombre_archivo}"
            titulo = f'Reporte de cargas familiares del periodo {periodo.anio} generado con éxito.'
            cuerpo = f'Archivo excel generado con éxito'

            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='sga-sagest',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su archivo excel ha sido generado con éxito.'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class descarga_activos_bajas_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            anioinicio = data['anioinicio']
            aniofin = data['aniofin']
            todos = data['todos']
            pers = Persona.objects.get(usuario_id=request.user.pk)
            nombre_archivo = 'reporte_baja' + random.randint(1, 10000).__str__() + '.xlsx'
            directory = os.path.join(SITE_STORAGE, 'media', 'activos')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)

            directory = os.path.join(SITE_STORAGE, 'media', 'activos', 'reportes')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)

            directory = os.path.join(SITE_STORAGE, 'media', 'activos', 'reportes', nombre_archivo)

            filtro = Q(status=True, seleccionado=True)
            if not todos:
                # filtro = filtro & (Q(fecha_creacion__year__gte = anioinicio,
                #            fecha_creacion__year__lte = aniofin) | \
                #          Q(fecha_modificacion__year__gte=anioinicio,
                #            fecha_modificacion__year__lte=aniofin))
                filtro = filtro & Q(codigobaja__fecha__year__range=(anioinicio, aniofin))
            bajas = DetalleBajaActivo.objects.filter(filtro)

            __author__ = 'Unemi'
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('bajas')
            ws.set_column(1, 20, 50)
            formatoceldagris = workbook.add_format(
                {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
            formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})
            ws.write('A1', 'NRO', formatoceldagris)
            ws.write('B1', 'BIEN', formatoceldagris)
            ws.write('C1', 'COD. GOBIERNO', formatoceldagris)
            ws.write('D1', 'COD. INTERNO', formatoceldagris)
            ws.write('E1', 'UBICACIÓN', formatoceldagris)
            ws.write('F1', 'USUARIO', formatoceldagris)
            ws.write('G1', 'MARCA', formatoceldagris)
            ws.write('H1', 'MODELO', formatoceldagris)
            ws.write('I1', 'SERIE', formatoceldagris)
            ws.write('J1', 'ESTADO', formatoceldagris)
            ws.write('K1', 'FECHA DE INGRESO', formatoceldagris)
            ws.write('L1', 'FECHA BAJA', formatoceldagris)
            ws.write('M1', 'VALOR CONTABLE', formatoceldagris)
            ws.write('N1', 'OFICIO', formatoceldagris)
            ws.write('O1', 'FECHA OFICIO', formatoceldagris)
            ws.write('P1', 'OBSERVACIÓN', formatoceldagris)
            i = 2
            cont = 1
            for baja in bajas:
                ws.write('A%s' % i, str(cont), formatoceldaleft)
                ws.write('B%s' % i, str(baja.activo.observacion), formatoceldaleft)
                ws.write('C%s' % i, str(baja.activo.codigogobierno), formatoceldaleft)
                ws.write('D%s' % i, str(baja.activo.codigointerno), formatoceldaleft)
                ws.write('E%s' % i, str(baja.activo.ubicacion), formatoceldaleft)
                ws.write('F%s' % i, str(baja.activo.responsable), formatoceldaleft)
                ws.write('G%s' % i, str(baja.activo.marca), formatoceldaleft)
                ws.write('H%s' % i, str(baja.activo.modelo), formatoceldaleft)
                ws.write('I%s' % i, str(baja.activo.serie), formatoceldaleft)
                ws.write('J%s' % i, str(baja.activo.estado), formatoceldaleft)
                ws.write('K%s' % i, str(baja.activo.fechaingreso), formatoceldaleft)
                ws.write('L%s' % i, str(baja.codigobaja.fecha), formatoceldaleft)
                ws.write('M%s' % i, str(baja.activo.valorlibros), formatoceldaleft)
                ws.write('N%s' % i, str(baja.codigobaja.oficio), formatoceldaleft)
                ws.write('O%s' % i, str(baja.codigobaja.fechaoficio), formatoceldaleft)
                ws.write('P%s' % i, str(baja.codigobaja.observacion), formatoceldaleft)
                cont += 1
                i += 1

            workbook.close()
            # NOTIFICACIÓN
            url = f"{MEDIA_URL}activos/reportes/{nombre_archivo}"
            titulo = f'Reporte de bajas generado con éxito.'
            cuerpo = f'Archivo excel generado con éxito'

            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='sga-sagest',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su archivo excel ha sido generado con éxito.'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class actualizar_visible_horario_masivo(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            periodo = data['periodo']
            accion = data['accion']
            es_pregrado = data.get('es_pregrado', False)
            es_admision = data.get('es_admision', False)
            mostrar = True if accion == 1 else False
            pers = Persona.objects.get(usuario_id=request.user.pk)
            filtro = Q(status=True) & Q(retiradomatricula=False) & Q(bloqueomatricula=False) & Q(
                nivel__periodo_id=periodo)
            if es_admision:
                filtro = filtro & Q(inscripcion__coordinacion_id__in=[9])
            elif es_pregrado:
                filtro = filtro & Q(inscripcion__coordinacion_id__in=[1, 2, 3, 4, 5, 12])
            matriculas = Matricula.objects.filter(filtro)
            if matriculas:
                for eMatricula in matriculas:
                    MateriaAsignada.objects.filter(matricula=eMatricula).update(visiblehorarioexamen=mostrar)
                    eMatricula.delete_cache()
                    eInscripcion = eMatricula.inscripcion
                    eInscripcion.delete_cache()
                    ePersona = eInscripcion.persona
                    ePersona.delete_cache()
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = 'Materias actualizadas'
                noti.en_proceso = False
                noti.cuerpo = 'Todas las materias fueron actualizadas'
                noti.url = f"{request.path}"
                noti.save()
            else:
                noti = Notificacion(cuerpo='Materias actualizadas',
                                    titulo='Todas las materias fueron actualizadas',
                                    destinatario=pers, url=f"{request.path}",
                                    prioridad=1, app_label='SAGEST',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": 'Materias actualizadas',
                "body": 'Todas las materias fueron actualizadas',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": f"{request.path}",
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Todas las materias fueron actualizadas',
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class limpiar_cache_masivo(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            periodo = data['periodo']
            es_pregrado = data.get('es_pregrado', False)
            es_admision = data.get('es_admision', False)
            pers = Persona.objects.get(usuario_id=request.user.pk)
            filtro = Q(status=True) & Q(retiradomatricula=False) & Q(bloqueomatricula=False) & Q(
                nivel__periodo_id=periodo)
            if es_admision:
                filtro = filtro & Q(inscripcion__coordinacion_id__in=[9])
            elif es_pregrado:
                filtro = filtro & Q(inscripcion__coordinacion_id__in=[1, 2, 3, 4, 5, 12])
            matriculas = Matricula.objects.filter(filtro)
            if matriculas:
                for eMatricula in matriculas:
                    eMatricula.delete_cache()
                    eInscripcion = eMatricula.inscripcion
                    eInscripcion.delete_cache()
                    ePersona = eInscripcion.persona
                    ePersona.delete_cache()
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = 'Limpieza de cache de pregrado' if es_pregrado else 'Limpieza de cache de admisión'
                noti.en_proceso = False
                noti.cuerpo = 'La limpieza de cache de pregrado fue realizada' if es_pregrado else 'La limpieza de cache de admisión fue realizada'
                noti.url = f"{request.path}"
                noti.save()
            else:
                noti = Notificacion(
                    cuerpo='La limpieza de cache de pregrado fue realizada' if es_pregrado else 'La limpieza de cache de admisión fue realizada',
                    titulo='Limpieza de cache de pregrado' if es_pregrado else 'Limpieza de cache de admisión',
                    destinatario=pers,
                    url=f"{request.path}",
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": 'Limpieza de cache de pregrado' if es_pregrado else 'Limpieza de cache de admisión',
                "body": 'La limpieza de cache de pregrado fue realizada' if es_pregrado else 'La limpieza de cache de admisión fue realizada',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": f"{request.path}",
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'La limpieza de cache de pregrado fue realizada' if es_pregrado else 'La limpieza de cache de admisió fue realizada',
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_persona_sin_examen(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        pers = Persona.objects.get(usuario_id=request.user.pk)
        # directory = os.path.join(MEDIA_ROOT, 'reportes', 'horarios')
        directory = os.path.join(SITE_STORAGE, 'media', 'reportes', 'horarios')
        nombre_archivo = 'reporte_personas_sin_examen_' + str(random.randint(1, 1000000000000)) + '.xls'
        ePeriodo = request.session['periodo']
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'horarios', nombre_archivo)
            if not 'inputSede' in request.GET:
                raise NameError(u'Parametro de sede no encontrado')
            if not 'inputFechaInicio' in request.GET:
                raise NameError(u'Debe seleccionar un afecha de inicio')
            if not 'inputFechaFin' in request.GET:
                raise NameError(u'Debe seleccionar un afecha de inicio')
            fechainicio = convertir_fecha_invertida(request.GET['inputFechaInicio'])
            fechafin = convertir_fecha_invertida(request.GET['inputFechaFin'])
            if fechafin < fechainicio:
                raise NameError(u'La fecha de fin debe ser mayor a la fecha de inicio')
            ids = []
            materias = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                    aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=ePeriodo,
                                                                                    aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha__range=[
                                                                                        fechainicio, fechafin]).values(
                'idtestmoodle', 'materiaasignada__matricula__inscripcion__persona__usuario__username', 'id')
            if 'inputSede' in request.GET:
                sede = request.GET['inputSede']
                if int(sede) > 1:
                    materias = materias.filter(aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede)

            if DEBUG:
                materias = materias[:100]
            for materia in materias:
                cursor = connections['db_moodle_virtual'].cursor()
                sql_ = f"SELECT quizat.id, quiz.id, quiz.name, course.fullname, us.username, TO_TIMESTAMP(quizat.timestart), TO_TIMESTAMP(quizat.timefinish) FROM mooc_quiz_attempts quizat " \
                       f"inner join mooc_quiz quiz on quiz.id = quizat.quiz " \
                       f"inner join mooc_course course on course.id = quiz.course " \
                       f"inner join mooc_user us on us.id = quizat.userid " \
                       f"where quiz.id = '{materia['idtestmoodle']}' and us.username='{materia['materiaasignada__matricula__inscripcion__persona__usuario__username']}';"

                cursor.execute(sql_)
                results = cursor.fetchall()
                if not len(results):
                    ids.append(materia['id'])
            if not ids:
                raise NameError('No existe personas sin rendir examen')
            __author__ = 'Unemi'
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            response = HttpResponse(content_type="application/ms-excel")
            response[
                'Content-Disposition'] = 'attachment; filename=reporte_total_inscritos_facultad_capacitacion_docente' + random.randint(
                1, 10000).__str__() + '.xls'

            columns = [
                (u"Alumno", 9000),
                (u"Materia", 9000),
                (u"Sede", 7000),
                (u"Carrera", 7000),
                (u"Fecha", 5000),
                (u"Turno", 5000)

            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            row_num = 5
            for alumno in MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(id__in=ids):
                ws.write(row_num, 0,
                         u'%s' % alumno.materiaasignada.matricula.inscripcion.persona.nombre_completo_minus(),
                         font_style2)
                ws.write(row_num, 1, u'%s' % alumno.materiaasignada.materia.asignatura.nombre, font_style2)
                ws.write(row_num, 2, u'%s' % alumno.aulaplanificacion.turnoplanificacion.fechaplanificacion.sede,
                         font_style2)
                ws.write(row_num, 3, u'%s' % alumno.materiaasignada.matricula.inscripcion.carrera.nombre, font_style2)
                ws.write(row_num, 4, u'%s' % str(alumno.aulaplanificacion.turnoplanificacion.fechaplanificacion.fecha),
                         font_style2)
                ws.write(row_num, 5,
                         u'%s' % '{} - {}'.format(alumno.aulaplanificacion.turnoplanificacion.horainicio.__str__(),
                                                  alumno.aulaplanificacion.turnoplanificacion.horafin.__str__()),
                         font_style2)

                row_num += 1
            wb.save(directory)

            url = "{}/{}/{}".format(MEDIA_URL, 'reportes/horarios/', nombre_archivo)

            if notif:
                notif.titulo = 'Reporte termiando'
                notif.en_proceso = False
                notif.cuerpo = 'Reporte de personas que no rindieron examen fue terminado'
                notif.url = url
                notif.save()
            else:
                noti = Notificacion(cuerpo='Reporte de personas que no rindieron examen fue terminado',
                                    titulo='Reporte termiando',
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": 'Reporte termiando',
                "body": 'Reporte de personas que no rindieron examen fue terminado',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notif:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = 'Error en el reporte'
                noti.en_proceso = False
                noti.cuerpo = textoerror
                noti.url = ''
                noti.save()
            else:
                noti = Notificacion(cuerpo=textoerror,
                                    titulo='Error en el reporte',
                                    destinatario=pers, url='',
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)


class reporte_general_porcentaje_cumplimiento_background(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        from sga.templatetags.sga_extras import nombremes
        from sga.adm_criteriosactividadesdocente import reporte_porcentaje_cumplimiento, \
            reporte_porcentaje_general_cumplimiento

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'cumplimientodocente')
        request, data, notiid = self.request, self.data, self.notiid

        os.makedirs(directory, exist_ok=True)

        nombre_archivo = generar_nombre("reporte_general_porcentaje_cumplimiento_", '') + '.xls'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'cumplimientodocente', nombre_archivo)

        now = datetime.now()
        m = datetime(now.year, int(request.GET['m']), 1) if request.GET.get('m') else now
        mes = nombremes(m).__str__().upper()

        data = reporte_porcentaje_general_cumplimiento(request)
        criterios = [(1, 'DOCENCIA'), (2, 'INVESTIGACIÓN'), (3, 'GESTIÓN'), (4, 'VINCULACIÓN')]

        try:
            __author__ = 'Unemi'
            titulo = easyxf(
                'font: name Verdana, color-index black, bold on , height 350; alignment: horiz centre;borders: left thin, right thin, top thin, bottom thin')
            titulo2 = easyxf(
                'font: name Verdana, color-index black, bold on , height 250; alignment: horiz centre;borders: left thin, right thin, top thin, bottom thin')
            encabesado_tabla = easyxf(
                'font: name Verdana , bold on , height 150; alignment: horiz left;borders: left thin, right thin, top thin, bottom thin;pattern: pattern solid, fore_colour gray25')

            fuentecabecera = easyxf(
                'font: name Verdana, bold on, color-index black, height 150; alignment: vert distributed, horiz left; borders: left thin, right thin, top thin, bottom thin')
            fuentenormal = easyxf(
                'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
            fuentenormalred = easyxf(
                'font: name Verdana, color-index red, height 150; borders: left thin, right thin, top thin, bottom thin')
            fuentenumeroentero = easyxf(
                'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('PORCENTAJES')
            ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
            ws.write_merge(1, 1, 0, 6, 'PORCENTAJE DE CUMPLIMIENTO DOCENTE EN EL MES DE %s' % mes, titulo2)
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

            ws.col(0).width = 2000
            ws.col(1).width = 20500
            ws.col(2).width = 10000
            ws.col(3).width = 10000
            ws.col(4).width = 5000
            ws.col(5).width = 10000
            ws.col(6).width = 10000

            row_num = 2

            ws.write(row_num, 0, "Nº", encabesado_tabla)
            ws.write(row_num, 1, "DOCENTE", encabesado_tabla)
            ws.write(row_num, 2, "FACULTAD", encabesado_tabla)
            ws.write(row_num, 3, "CARRERA", encabesado_tabla)
            ws.write(row_num, 4, "DEDICACIÓN", encabesado_tabla)
            ws.write(row_num, 5, "CARGO", encabesado_tabla)
            ws.write(row_num, 6, "% CUMPLIMIENTO", encabesado_tabla)

            date_format = xlwt.XFStyle()
            date_format.num_format_str = 'yyyy/mm/dd'
            c = 1
            row_num = 3

            ABR_CRITERIOS = {1: 'DOC', 2: 'INV', 3: 'GES', 4: 'VIN'}
            for d in data:
                _puesto = ''
                ws.write(row_num, 0, "%s" % c, fuentecabecera)
                ws.write(row_num, 1, "%s" % d[0], fuentecabecera)
                ws.write(row_num, 2, "%s" % d[3].coordinacion_set.filter(status=True).first() if d[3] else '',
                         fuentecabecera)
                ws.write(row_num, 3, "%s" % d[3], fuentecabecera)
                ws.write(row_num, 4, "%s" % d[4].dedicacion, fuentecabecera)
                if puesto := d[0].persona.mis_cargos_vigente().first(): _puesto = puesto.denominacionpuesto.descripcion
                ws.write(row_num, 5, "%s" % _puesto, fuentecabecera)
                ws.write(row_num, 6, "%.2f" % d[1], fuentecabecera)
                row_num += 1

                if d[1] < 100:
                    for v in d[2]:
                        criterio, tipo, porcentaje = v.get('criterio').criterio, v.get('tipo'), v.get('porcentaje')
                        ws.write_merge(row_num, row_num, 1, 6,
                                       "    %s - %s - %.2f" % (ABR_CRITERIOS[tipo], criterio, porcentaje), fuentenormal)
                        row_num += 1

                c += 1
            wb.save(directory)

            user = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=user)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/cumplimientodocente/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Excel porcentaje de cumplimiento de actividades mensuales del docente',
                                    destinatario=pers,
                                    url="{}reportes/cumplimientodocente/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=user, payload={
                "head": "Excel terminado",
                "body": 'Excel porcentaje de cumplimiento de actividades mensuales del docente',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/cumplimientodocente/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito.'
            }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


##ES PDF NO EXCEL
class reporte_general_seguimiento_silabo(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.funcionesxhtml2pdf import convert_html_to_pdf
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'seguimientosilabo')
        request, data, notiid = self.request, self.data, self.notiid
        os.makedirs(directory, exist_ok=True)
        nombre_archivo = generar_nombre("reporte_general_seguimiento_silabo_", '') + '.pdf'
        try:
            if 'exportar_seguimiento_pdf' in request.GET:
                pdfgen = convert_html_to_pdf(
                    'adm_seguimientosilabo/generarinforme_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    },
                    nombre_archivo,
                    directory
                )
            user = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=user)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'PDF listo'
                noti.url = "{}reportes/seguimientosilabo/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo', titulo='PDF seguimiento al sílabo del docente',
                                    destinatario=pers,
                                    url="{}reportes/seguimientosilabo/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=user, payload={
                "head": "PDF generado",
                "body": 'PDF seguimiento al sílabo del docente',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/seguimientosilabo/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito.'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_reportealumnos_background(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):

        from sga.templatetags.sga_extras import nombremes
        from sga.adm_criteriosactividadesdocente import reporte_porcentaje_cumplimiento, \
            reporte_porcentaje_general_cumplimiento

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
        request, data, notiid = self.request, self.data, self.notiid

        os.makedirs(directory, exist_ok=True)

        nombre_archivo = generar_nombre("reporte_alumnos_matriculados_", '') + '.xlsx'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)

        now = datetime.now()
        m = datetime(now.year, int(request.GET['m']), 1) if request.GET.get('m') else now
        # mes = nombremes(m).__str__().upper()

        try:
            __author__ = 'Unemi'
            # Aplicar el objeto de Formato personalizado al estilo
            wb = openxl.Workbook()
            ws = wb.active
            style_title = openxlFont(name='Arial', size=16, bold=True)
            style_cab = openxlFont(name='Arial', size=10, bold=True)
            alinear = alin(horizontal="center", vertical="center")
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + '-' + random.randint(1,
                                                                                                              10000).__str__() + '.xlsx'
            ws.merge_cells('A1:H1')
            ws.merge_cells('A2:H2')
            ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO'
            ws['A2'] = 'REPORTE DE ALUMNOS MATRICULADOS'
            celda1 = ws['A2']
            celda1.font = style_title
            celda1.alignment = alinear
            columns = ["FACULTAD", "CARRERA", "SECCION", "NIVEL", 'ASIGNATURA', "CEDULA ALUMN.", "ALUMNO",
                       "SEXO", "LGTBI", "GRUPO SOCIOECONOMICO", "ESTADO CIVIL", 'ETNIA', "DISCAPACIDAD", "PORCENTAJE",
                       "CARNET", "VERIFICADO", "BECA", "NOTA FINAL", 'ASISTENCIA', "ESTADO", "CEDULA DOC.",
                       "DOCENTE", "SEXO", "TIPO DOCENTE", "DEDICACION DOCENTE", 'TELEFONO', "EMAIL", "EMAIL INST",
                       "CIUDAD", "DIRECCION", "PARALELO", "VECES DE MATRICULA", 'CREDITOS', "TIPO MATRICULA",
                       "GRATUIDAD",
                       "CARRERA DE LA ASIGNATURA", "CODIGO MATERIA"
                       ]

            row_num = 3
            for col_num in range(0, len(columns)):
                celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                celda.font = style_cab
            row_num = 5

            cursor = connections['sga_select'].cursor()
            sql = "SELECT DISTINCT " \
                  "   sga_coordinacion.nombre as sga_coordinacion_nombre, " \
                  "   sga_carrera.nombre AS sga_carrera_nombre, " \
                  "   sga_sesion.nombre AS sga_sesion_nombre, " \
                  "   sga_nivelmalla.nombre as sga_nivelmalla_nombre, " \
                  "   sga_asignatura.nombre as sga_asignatura_nombre, " \
                  "   sga_persona.apellido1 || ' ' || sga_persona.apellido2  || ' ' ||  sga_persona.nombres AS sga_persona_nombres, " \
                  "   sga_materiaasignada.notafinal as notafinal, " \
                  "   sga_materiaasignada.asistenciafinal as asistenciafinal, " \
                  "   (select p.apellido1 || ' ' || p.apellido2 || ' ' || p.nombres from sga_persona p where p.id=sga_profesor.persona_id) as docente, " \
                  "   sga_tipoestado.nombre as estado, " \
                  "   sga_nivelmalla.id as sga_nivelmalla_id," \
                  "   sga_persona.telefono, sga_persona.email, sga_persona.emailinst, sga_persona.ciudad, " \
                  "   sga_persona.direccion || ', ' || sga_persona.direccion2 || ' #:' || sga_persona.num_direccion || ' , SECTOR:' || sga_persona.sector as direccion, sga_materia.paralelo, sga_materiaasignada.matriculas, sga_asignaturamalla.creditos," \
                  "   (select p.cedula from sga_persona p where p.id=sga_profesor.persona_id) as cedula, " \
                  "   (select s.nombre from sga_persona p INNER JOIN sga_sexo s on s.id=p.sexo_id where p.id=sga_profesor.persona_id) as sexo, " \
                  "   (select tp.nombre from sga_tipoprofesor tp where tp.id=sga_profesormateria.tipoprofesor_id) as tipoprofesor, " \
                  "   (select de.nombre from sga_tiempodedicaciondocente de where de.id=sga_profesor.dedicacion_id) as dedicacion, " \
                  "   sga_persona.cedula, (select s.nombre from sga_sexo s where s.id=sga_persona.sexo_id ) as sexoalu , sga_persona.lgtbi, " \
                  "   (select gr.nombre from socioecon_fichasocioeconomicainec f inner join socioecon_gruposocioeconomico gr on gr.id=f.grupoeconomico_id where persona_id=sga_persona.id) as socioeco, " \
                  "   (select r.nombre from sga_perfilinscripcion pa inner join sga_raza r on r.id=pa.raza_id where persona_id=sga_persona.id) as etnia, " \
                  "   (select d.nombre from sga_perfilinscripcion pa inner join sga_discapacidad d on d.id=pa.tipodiscapacidad_id where persona_id=sga_persona.id) as discapacidad, " \
                  "   (select pe.porcientodiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as porcentaje, " \
                  "   (select pe.carnetdiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as carnetdiscapacidad, " \
                  "   (select pe.verificadiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as verificadiscapacidad, " \
                  "   (select be.id from sga_inscripcionbecario be where be.inscripcion_id=sga_inscripcion.id) as tienebeca, " \
                  "   (select pe.nombre from med_personaextension es inner join sga_personaestadocivil pe on pe.id=es.estadocivil_id where es.persona_id=sga_persona.id) as estadocivil, " \
                  " (case gru.tipomatricula when 1 then 'REGULAR'  when 2 then 'IREGULAR' else '' end) as tipomatricula, " \
                  " (case gru.estado_gratuidad when 1 then 'GRATUIDAD COMPLETA'  when 2 then 'GRATUIDAD PARCIAL' when 3 then 'PERDIDA DE GRATUIDAD' else '' end) as estado_gratuidad," \
                  "   (select carr.nombre from sga_asignaturamalla asig inner join sga_malla mall on mall.id=asig.malla_id inner join sga_carrera carr on carr.id=mall.carrera_id where asig.id=sga_asignaturamalla.id) as Carrera_de_la_asignatura," \
                  "sga_materia.id as id " \
                  " FROM sga_persona sga_persona " \
                  "      RIGHT OUTER JOIN sga_inscripcion sga_inscripcion ON sga_persona.id = sga_inscripcion.persona_id " \
                  "     INNER JOIN sga_matricula sga_matricula ON sga_matricula.inscripcion_id=sga_inscripcion.id " \
                  " left JOIN sga_matriculagruposocioeconomico gru ON gru.matricula_id=sga_matricula.id " \
                  "    inner join sga_materiaasignada sga_materiaasignada on sga_materiaasignada.matricula_id=sga_matricula.id " \
                  "     inner join sga_materia sga_materia on sga_materia.id=sga_materiaasignada.materia_id " \
                  "     LEFT join sga_profesormateria on sga_profesormateria.materia_id=sga_materia.id " \
                  "     LEFT join sga_profesor on sga_profesor.id=sga_profesormateria.profesor_id " \
                  "     inner join sga_asignatura sga_asignatura on sga_asignatura.id=sga_materia.asignatura_id " \
                  "      inner join sga_asignaturamalla sga_asignaturamalla on sga_asignaturamalla.id=sga_materia.asignaturamalla_id " \
                  "    inner join sga_nivel sga_nivel ON sga_nivel.id=sga_matricula.nivel_id and sga_nivel.periodo_id= '" + str(
                self.periodo.id) + "' " \
                                   "      inner join sga_nivelmalla sga_nivelmalla on sga_nivelmalla.id=sga_asignaturamalla.nivelmalla_id " \
                                   "     INNER JOIN sga_carrera sga_carrera ON sga_inscripcion.carrera_id = sga_carrera.id " \
                                   "     INNER JOIN sga_coordinacion_carrera on sga_coordinacion_carrera.carrera_id=sga_carrera.id " \
                                   "     INNER JOIN sga_coordinacion on sga_coordinacion.id=sga_coordinacion_carrera.coordinacion_id " \
                                   "     INNER JOIN sga_modalidad sga_modalidad ON sga_inscripcion.modalidad_id = sga_modalidad.id " \
                                   "     INNER JOIN sga_sesion sga_sesion ON sga_inscripcion.sesion_id = sga_sesion.id " \
                                   "     inner join sga_tipoestado on sga_tipoestado.id=sga_materiaasignada.estado_id " \
                                   "    where sga_matricula.estado_matricula in (2,3) and sga_matricula.id not in (select ret.matricula_id from sga_retiromatricula ret) and sga_profesormateria.activo=True and sga_profesormateria.id = (select pm.id from sga_profesormateria pm inner join sga_profesor profe on pm.profesor_id=profe.id inner join sga_persona p on p.id=profe.persona_id  where pm.materia_id=sga_materiaasignada.materia_id and pm.status=True and pm.activo=True and (pm.tipoprofesor_id in (1, 3, 8, 14, 16)) order by pm.tipoprofesor_id LIMIT 1) order by sga_carrera.nombre,sga_nivelmalla.id,sga_sesion.nombre,sga_persona_nombres"
            cursor.execute(sql)
            results = cursor.fetchall()
            for r in results:
                campo1 = u"%s" % r[0]
                campo2 = u"%s" % r[1]
                campo3 = u"%s" % r[2]
                campo4 = u"%s" % r[3]
                campo5 = u"%s" % r[4]
                campo6 = u"%s" % r[5]
                campo7 = float(r[6]) if r[6] else 0
                campo8 = int(r[7]) if r[7] else 0
                campo9 = u"%s" % r[8]
                campo10 = u"%s" % r[9]
                campo33 = u"%s" % r[33]
                campo11 = u"%s" % r[11]
                campo12 = u"%s" % r[12]
                campo13 = u"%s" % r[13]
                campo14 = u"%s" % r[14]
                campo15 = u"%s" % r[15]
                campo16 = u"%s" % r[16]
                campo17 = int(r[17]) if r[17] else 0
                campo18 = float(r[18]) if r[18] else 0
                campo19 = u"%s" % r[19]
                campo20 = u"%s" % r[20]
                campo21 = u"%s" % r[21]
                campo22 = u"%s" % r[22]
                campo23 = u"%s" % r[23]
                campo24 = u"%s" % r[24]
                campo25 = "SI" if r[25] else "NO"
                campo26 = u"%s" % r[26]
                campo27 = u"%s" % r[27]
                campo28 = u"%s" % r[28] if r[28] else "NINGUNA"
                campo29 = float(r[29]) if r[29] else 0
                campo30 = u"%s" % r[30]
                campo31 = "SI" if r[31] else "NO"
                campo32 = "SI" if r[31] else "NO"
                campo34 = u"%s" % r[34]
                campo35 = u"%s" % r[35]
                campo36 = u"%s" % r[36]
                campo37 = int(r[37]) if r[37] else 0
                #  ws.cell(row=row_num, column=1, value=numero)
                ws.cell(row=row_num, column=1, value=str(campo1))
                ws.cell(row=row_num, column=2, value=str(campo2))
                ws.cell(row=row_num, column=3, value=str(campo3))
                ws.cell(row=row_num, column=4, value=str(campo4))
                ws.cell(row=row_num, column=5, value=str(campo5))
                ws.cell(row=row_num, column=6, value=str(campo23))
                ws.cell(row=row_num, column=7, value=str(campo6))
                ws.cell(row=row_num, column=8, value=str(campo24))
                ws.cell(row=row_num, column=9, value=str(campo25))
                ws.cell(row=row_num, column=10, value=str(campo26))
                ws.cell(row=row_num, column=11, value=str(campo33))
                ws.cell(row=row_num, column=12, value=str(campo27))
                ws.cell(row=row_num, column=13, value=str(campo28))
                ws.cell(row=row_num, column=14, value=campo29)
                ws.cell(row=row_num, column=15, value=str(campo30))
                ws.cell(row=row_num, column=16, value=str(campo31))
                ws.cell(row=row_num, column=17, value=str(campo32))
                ws.cell(row=row_num, column=18, value=campo7)
                ws.cell(row=row_num, column=19, value=campo8)
                ws.cell(row=row_num, column=20, value=str(campo10))
                ws.cell(row=row_num, column=21, value=str(campo19))
                ws.cell(row=row_num, column=22, value=str(campo9))
                ws.cell(row=row_num, column=23, value=str(campo20))
                ws.cell(row=row_num, column=24, value=str(campo21))
                ws.cell(row=row_num, column=25, value=str(campo22))
                ws.cell(row=row_num, column=26, value=str(campo11))
                ws.cell(row=row_num, column=27, value=str(campo12))
                ws.cell(row=row_num, column=28, value=str(campo13))
                ws.cell(row=row_num, column=29, value=str(campo14))
                ws.cell(row=row_num, column=30, value=str(campo15))
                ws.cell(row=row_num, column=31, value=str(campo16))
                ws.cell(row=row_num, column=32, value=campo17)
                ws.cell(row=row_num, column=33, value=campo18)
                ws.cell(row=row_num, column=34, value=str(campo34))
                ws.cell(row=row_num, column=35, value=str(campo35))
                ws.cell(row=row_num, column=36, value=str(campo36))
                ws.cell(row=row_num, column=37, value=campo37)

                row_num += 1
            wb.save(directory)

            user = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=user)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo', titulo='Reporte de asignaturas y estudiantes',
                                    destinatario=pers, url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=user, payload={
                "head": "Reporte terminado",
                "body": 'Reporte de asignaturas y estudiantes',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito.'
            }, ttl=500)

        except Exception as ex:
            user = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=user)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte - {} - Linea:{} '.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(
                    cuerpo='Error en el reporte - {} - Linea:{} '.format(ex, sys.exc_info()[-1].tb_lineno),
                    titulo='Error en reporte',
                    destinatario=pers, url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=user, payload={
                "head": "Error en el reporte",
                "body": 'Su reporte tuvo un error {}'.format(ex),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte tuvo un error {}'.format(ex)
            }, ttl=500)


class reporte_transversales_background(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        row_num = 0
        materia = None
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos_transversales')
        request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo

        os.makedirs(directory, exist_ok=True)

        nombre_archivo = generar_nombre("reporte_alumnos_matriculados_transversales", '') + '.xlsx'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)

        now = datetime.now()
        m = datetime(now.year, int(request.GET['m']), 1) if request.GET.get('m') else now
        # mes = nombremes(m).__str__().upper()

        try:
            __author__ = 'Unemi'
            # Aplicar el objeto de Formato personalizado al estilo
            wb = openxl.Workbook()
            ws = wb.active
            style_title = openxlFont(name='Arial', size=16, bold=True)
            style_cab = openxlFont(name='Arial', size=10, bold=True)
            alinear = alin(horizontal="center", vertical="center")
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + '-' + random.randint(1,
                                                                                                              10000).__str__() + '.xlsx'
            ws.merge_cells('A1:L1')
            ws.merge_cells('A2:L2')
            ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO'
            ws['A2'] = 'REPORTE DE ALUMNOS MATRICULADOS MATERIAS TRANSVERSALES'
            celda1 = ws['A2']
            celda1.font = style_title
            celda1.alignment = alinear
            columns = ["ALUMNO", "EMAIL", "EMAIL INST", "CARRERA", "SECCION", "NIVEL", "PARALELO", "ASIGNATURA", "N1",
                       "N2", "N RE", "NOTA FINAL"]

            row_num = 3
            for col_num in range(0, len(columns)):
                celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                celda.font = style_cab
            row_num = 5
            materiasasignadas = MateriaAsignada.objects.filter(status=True, materia__status=True,
                                                               materia__nivel__periodo=periodo,
                                                               materia__modeloevaluativo_id=27,
                                                               materia__asignaturamalla__transversal=True).values('id',
                                                                                                                  'matricula__inscripcion__persona__email',
                                                                                                                  'matricula__inscripcion__persona__emailinst',
                                                                                                                  'matricula__inscripcion__carrera__nombre',
                                                                                                                  'matricula__inscripcion__carrera__modalidad',
                                                                                                                  'materia__asignaturamalla__nivelmalla_id',
                                                                                                                  'materia__paralelo',
                                                                                                                  'materia__asignatura__nombre',
                                                                                                                  'notafinal').annotate(
                alumno=Concat(F('matricula__inscripcion__persona__apellido1'), Value(' '),
                              F('matricula__inscripcion__persona__apellido2'), Value(' '),
                              F('matricula__inscripcion__persona__nombres'))).distinct()
            for materia in materiasasignadas:
                ws.cell(row=row_num, column=1, value=str(materia['alumno']))
                ws.cell(row=row_num, column=2, value=str(materia['matricula__inscripcion__persona__email']))
                ws.cell(row=row_num, column=3, value=str(materia['matricula__inscripcion__persona__emailinst']))
                ws.cell(row=row_num, column=4, value=str(materia['matricula__inscripcion__carrera__nombre']))
                ws.cell(row=row_num, column=5,
                        value=MODALIDAD_CARRERA[int(materia['matricula__inscripcion__carrera__modalidad'])][1])
                ws.cell(row=row_num, column=6, value=str(materia['materia__asignaturamalla__nivelmalla_id']))
                ws.cell(row=row_num, column=7, value=str(materia['materia__paralelo']))
                ws.cell(row=row_num, column=8, value=str(materia['materia__asignatura__nombre']))
                colum = 9
                for n in EvaluacionGenerica.objects.filter(status=True, materiaasignada_id=int(materia['id'])).exclude(
                        detallemodeloevaluativo_id=124).order_by('detallemodeloevaluativo__nombre'):
                    ws.cell(row=row_num, column=colum, value=str(n.valor))
                    colum += 1
                ws.cell(row=row_num, column=colum, value=str(materia['notafinal']))
                row_num += 1
            wb.save(directory)

            user = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=user)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo', titulo='Reporte de asignaturas y estudiantes',
                                    destinatario=pers, url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=user, payload={
                "head": "Reporte terminado",
                "body": 'Reporte de asignaturas transversales y estudiantes',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito.'
            }, ttl=500)

        except Exception as ex:
            user = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=user)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte - {} - Linea:{} '.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(
                    cuerpo='Error en el reporte - {} - Linea:{} '.format(ex, sys.exc_info()[-1].tb_lineno),
                    titulo='Error en reporte',
                    destinatario=pers, url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=user, payload={
                "head": "Error en el reporte",
                "body": 'Su reporte tuvo un error {}'.format(ex),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte tuvo un error {}'.format(ex)
            }, ttl=500)


class reporte_activos_constatados_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            nombre_archivo = generar_nombre("reporte_constatacion", '') + '.xlsx'
            url = ''
            directory = os.path.join(SITE_STORAGE, 'media', 'activos')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)
            directory = os.path.join(SITE_STORAGE, 'media', 'activos', 'reportes')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)

            ruta = f'{directory}\\{nombre_archivo}'
            pers = Persona.objects.get(usuario_id=request.user.pk)
            periodo = PeriodoConstatacionAF.objects.get(id=encrypt_id(request.GET['id_obj']))
            constatacion, item, responsables, constatadores, estado = request.GET.get('constatacion', ''), \
                                                                      request.GET.get('item', ''), \
                                                                      request.GET.getlist('responsable', ''), \
                                                                      request.GET.getlist('constatador', ''), \
                                                                      request.GET.get('estado', '')
            filtro, filtro_c, filtro_excl = Q(status=True), Q(status=True), []

            if estado:
                filtro = filtro & Q(estado_id=estado)

            if responsables:
                filtro = filtro & Q(responsable_id__in=responsables)

            if constatacion:
                constatacion = int(constatacion)
                if constatacion == 1:
                    if constatadores:
                        filtro_c = filtro_c & Q(responsable_id__in=constatadores)
                    if item:
                        if int(item) == 1:
                            filtro_c = filtro_c & Q(encontraro=True)
                        elif int(item) == 2:
                            filtro_c = filtro_c & Q(enuso=True)
                        elif int(item) == 3:
                            filtro_c = filtro_c & Q(requieretraspaso=True)
                        elif int(item) == 4:
                            filtro_c = filtro_c & Q(requierebaja=True)
                    ids_a = periodo.detalle_constatacion().filter(filtro_c).values_list('activo_id', flat=True)
                    filtro = filtro & Q(id__in=ids_a)
                elif constatacion == 2:
                    ids_a = periodo.detalle_constatacion().values_list('activo_id', flat=True)
                    filtro_excl = ids_a

            activos = ActivoFijo.objects.select_related().filter(filtro).exclude(id__in=filtro_excl).order_by('-id')

            columnas = ['Código Gob',
                        'Código Int',
                        'Activo',
                        'Constatado',
                        'Responsable',
                        'Constatador',
                        'Estado',
                        'En uso',
                        'Encontrado',
                        'Requiere Traspaso',
                        'Requiere dar baja',
                        ]
            # Crear el DataFrame con los nombres de columna personalizados
            df = pd.DataFrame(columns=columnas)

            # Llenar el DataFrame con los datos del queryset
            for activo in activos:
                constatado = activo.activo_constatado(periodo)
                fila = {
                    'Código Gob': activo.codigogobierno,
                    'Código Int': activo.codigointerno,
                    'Activo': activo.descripcion,
                    'Constatado': 'Si' if constatado else 'No',
                    'Responsable': activo.responsable.nombre_completo_minus(),
                    'Constatador': constatado.responsable.nombre_completo_minus() if constatado else '',
                    'Estado': activo.estado,
                    'En uso': 'Si' if constatado and constatado.enuso else 'No',
                    'Encontrado': 'Si' if constatado and constatado.encontrado else 'No',
                    'Requiere Traspaso': 'Si' if constatado and constatado.requieretraspaso else 'No',
                    'Requiere dar baja': 'Si' if constatado and constatado.requieredarbaja else 'No'
                }
                fila_df = pd.DataFrame([fila])
                df = pd.concat([df, fila_df], ignore_index=True)

            # Crear una respuesta HTTP
            # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # response['Content-Disposition'] = 'attachment; filename=reporte_excel.xlsx'

            # Escribir el DataFrame en el archivo Excel y adjuntarlo a la respuesta
            df.to_excel(ruta, index=False)

            # NOTIFICACIÓN
            titulo = 'Reporte de constatación de activos generada exitosamente'
            cuerpo = 'Su reporte de constatación de activos se genero correctamente y puede ser descargada por la url proporcionada.'
            url = "{}activos/reportes/{}".format(MEDIA_URL, nombre_archivo)

        except Exception as ex:
            titulo = 'Algo salio mal, el reporte constatación de activos no se genero'
            cuerpo = 'Su reporte de constatación de activos no se genero'
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
        if notif > 0:
            noti = Notificacion.objects.get(pk=notif)
            noti.titulo = titulo
            noti.en_proceso = False
            noti.cuerpo = cuerpo
            noti.url = url
            noti.save()
        else:
            noti = Notificacion(cuerpo=cuerpo,
                                titulo=titulo,
                                destinatario=pers, url=url,
                                prioridad=1, app_label='sga-sagest',
                                fecha_hora_visible=datetime.now() + timedelta(days=1),
                                tipo=2, en_proceso=False)
            noti.save(request)

        send_user_notification(user=request.user, payload={
            "head": titulo,
            "body": cuerpo,
            "action": "notificacion",
            "timestamp": time.mktime(datetime.now().timetuple()),
            "url": url,
            "btn_notificaciones": traerNotificaciones(request, data, pers),
            "mensaje": titulo
        }, ttl=500)


class reporte_activos_constatados_openxl_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            nombre_archivo = generar_nombre("reporte_constatacion", '') + '.xlsx'
            url = ''
            directory = os.path.join(SITE_STORAGE, 'media', 'activos', 'reportes')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)

            directory = os.path.join(MEDIA_ROOT, 'activos', 'reportes', nombre_archivo)
            pers = Persona.objects.get(usuario_id=request.user.pk)
            periodo = PeriodoConstatacionAF.objects.get(id=encrypt_id(request.GET['id_obj']))
            constatacion, item, responsables, constatadores, estado = request.GET.get('constatacion', ''), \
                                                                      request.GET.get('item', ''), \
                                                                      request.GET.getlist('responsable', ''), \
                                                                      request.GET.getlist('constatador', ''), \
                                                                      request.GET.get('estado', '')
            filtro, filtro_c, filtro_excl = Q(status=True, statusactivo=1), Q(status=True), []

            if estado:
                filtro = filtro & Q(estado_id=estado)

            if responsables:
                filtro = filtro & Q(responsable_id__in=responsables)

            if constatacion:
                constatacion = int(constatacion)
                if constatacion == 1:
                    if constatadores:
                        filtro_c = filtro_c & Q(responsable_id__in=constatadores)
                    if item:
                        if int(item) == 1:
                            filtro_c = filtro_c & Q(encontraro=True)
                        elif int(item) == 2:
                            filtro_c = filtro_c & Q(enuso=True)
                        elif int(item) == 3:
                            filtro_c = filtro_c & Q(requieretraspaso=True)
                        elif int(item) == 4:
                            filtro_c = filtro_c & Q(requierebaja=True)
                    ids_a = periodo.detalle_constatacion().filter(filtro_c).values_list('activo_id', flat=True)
                    filtro = filtro & Q(id__in=ids_a)
                elif constatacion == 2:
                    ids_a = periodo.detalle_constatacion().values_list('activo_id', flat=True)
                    filtro_excl = ids_a

            __author__ = 'Unemi'
            # Aplicar el objeto de Formato personalizado al estilo
            wb = openxl.Workbook()
            ws = wb.active
            style_title = openxlFont(name='Arial', size=16, bold=True)
            style_cab = openxlFont(name='Arial', size=10, bold=True)
            alinear = alin(horizontal="center", vertical="center")
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + '-' + random.randint(1,
                                                                                                              10000).__str__() + '.xlsx'
            ws.merge_cells('A1:Q1')
            ws.merge_cells('A2:Q2')
            ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO'
            ws['A2'] = 'REPORTE DE ACTIVOS CONSTATADOS'
            celda1 = ws['A1']
            celda1.font = style_title
            celda1.alignment = alinear
            celda2 = ws['A2']
            celda2.font = style_title
            celda2.alignment = alinear
            columns = ['Código Gob',
                       'Código Int',
                       'Activo',
                       'Catalogo',
                       'Tipo de bien',
                       'Constatado',
                       'Responsable',
                       'Estado',
                       'Constatador',
                       'En uso',
                       'Encontrado',
                       'Requiere Traspaso',
                       'Requiere dar baja',
                       'Observación',
                       'Fecha constatación',
                       'Ubicación',
                       'Baja',
                       ]
            row_num = 3
            for col_num in range(0, len(columns)):
                celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                celda.font = style_cab
            row_num = 4
            activos = ActivoFijo.objects.select_related().filter(filtro).exclude(id__in=filtro_excl).order_by('-id')

            # Llenar el DataFrame con los datos del queryset
            for activo in activos:
                constatado = activo.activo_constatado(periodo)
                ws.cell(row=row_num, column=1, value=activo.codigogobierno)
                ws.cell(row=row_num, column=2, value=activo.codigointerno)
                ws.cell(row=row_num, column=3, value=activo.descripcion)
                ws.cell(row=row_num, column=4, value=activo.catalogo.descripcion if activo.catalogo else '')
                ws.cell(row=row_num, column=5, value=str(activo.catalogo.tipobien) if activo.catalogo else '')
                ws.cell(row=row_num, column=6, value='Si' if constatado else 'No')
                ws.cell(row=row_num, column=7, value=activo.responsable.nombre_completo_minus() if activo.responsable else 'S/R')
                ws.cell(row=row_num, column=8, value=str(activo.estado))
                if constatado:
                    ws.cell(row=row_num, column=9, value=constatado.responsable.nombre_completo_minus())
                    ws.cell(row=row_num, column=10, value='Si' if constatado.enuso else 'No')
                    ws.cell(row=row_num, column=11, value='Si' if constatado.encontrado else 'No')
                    ws.cell(row=row_num, column=12, value='Si' if constatado.requieretraspaso else 'No')
                    ws.cell(row=row_num, column=13, value='Si' if constatado.requieredarbaja else 'No')
                    ws.cell(row=row_num, column=14, value=constatado.observacion)
                    ws.cell(row=row_num, column=15, value=constatado.fecha_creacion)
                    ws.cell(row=row_num, column=16, value=str(constatado.ubicacionbienes))
                ws.cell(row=row_num, column=17, value=activo.get_statusactivo_display())
                row_num += 1
            wb.save(directory)

            # NOTIFICACIÓN
            titulo = 'Reporte de constatación de activos generada exitosamente'
            cuerpo = 'Su reporte de constatación de activos se genero correctamente y puede ser descargada por la url proporcionada.'
            url = "{}activos/reportes/{}".format(MEDIA_URL, nombre_archivo)

        except Exception as ex:
            titulo = 'Algo salio mal, el reporte constatación de activos no se genero'
            cuerpo = 'Su reporte de constatación de activos no se genero'
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
        if notif > 0:
            noti = Notificacion.objects.get(pk=notif)
            noti.titulo = titulo
            noti.en_proceso = False
            noti.cuerpo = cuerpo
            noti.url = url
            noti.save()
        else:
            noti = Notificacion(cuerpo=cuerpo,
                                titulo=titulo,
                                destinatario=pers, url=url,
                                prioridad=1, app_label='sga-sagest',
                                fecha_hora_visible=datetime.now() + timedelta(days=1),
                                tipo=2, en_proceso=False)
            noti.save(request)

        send_user_notification(user=request.user, payload={
            "head": titulo,
            "body": cuerpo,
            "action": "notificacion",
            "timestamp": time.mktime(datetime.now().timetuple()),
            "url": url,
            "btn_notificaciones": traerNotificaciones(request, data, pers),
            "mensaje": titulo
        }, ttl=500)


class reporte_historico_estado_activos_por_constatacion_background(threading.Thread):

    def __init__(self, request, data, idnotificacion, fechadesde, fechahasta, codigo):
        self.request = request
        self.data = data
        self.idnotificacion = idnotificacion
        self.fechadesde = fechadesde
        self.fechahasta = fechahasta
        self.codigo = codigo
        threading.Thread.__init__(self)

    def run(self):
        try:
            request, data, idnotificacion, fechadesde, fechahasta, codigo = self.request, self.data, self.idnotificacion, self.fechadesde, self.fechahasta, self.codigo
            usernotify = User.objects.get(pk=request.user.pk)
            personanotifica = Persona.objects.get(usuario=usernotify)
            tituloreporte = "Reporte Excel Histórico Estado de Activos por Constataciones Física"

            directorio = os.path.join(os.path.join(MEDIA_ROOT, 'activos', 'reportes'))
            try:
                os.stat(directorio)
            except:
                os.mkdir(directorio)

            nombrearchivo = "HISTORICO_ESTADO_ACTIVOS_CONSTATACIONES_" + datetime.now().strftime(
                '%Y%m%d_%H%M%S') + ".xlsx"
            urlarchivo = MEDIA_URL + "activos/reportes/" + nombrearchivo

            # Crea un nuevo archivo de excel y le agrega una hoja
            workbook = xlsxwriter.Workbook(directorio + '/' + nombrearchivo)
            hojadestino = workbook.add_worksheet("Reporte")

            fcabeceracolumna = workbook.add_format(FORMATOS_CELDAS_EXCEL["cabeceracolumna"])
            fceldageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneral"])
            fceldageneralcent = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneralcent"])
            ftitulo2izq = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo2izq"])
            ftitulo3izq = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo3izq"])
            fceldafecha = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdafecha"])
            fceldamoneda = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdamoneda"])
            fceldamonedapie = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdamonedapie"])
            fceldanegritaizq = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritaizq"])
            fceldanegritageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritageneral"])

            hojadestino.merge_range(0, 0, 0, 15, 'UNIVERSIDAD ESTATAL DE MILAGRO', ftitulo2izq)
            hojadestino.merge_range(1, 0, 1, 15, 'DIRECCIÓN ADMINISTRATIVA Y FINANCIERA', ftitulo2izq)
            hojadestino.merge_range(2, 0, 2, 15,
                                    'REPORTE HISTÓRICO DE ESTADO DE ACTIVOS FIJOS POR CONSTATACIONES FÍSICAS',
                                    ftitulo2izq)
            hojadestino.merge_range(3, 0, 3, 15,
                                    'FECHA DE CORTE: DEL ' + str(fechadesde) + ' AL ' + str(fechahasta) + ' ',
                                    ftitulo2izq)
            hojadestino.merge_range(4, 0, 4, 15, 'FECHA DE DESCARGA DEL REPORTE (' + str(datetime.now().date()) + ')',
                                    ftitulo2izq)

            columnas = [
                (u"N°", 3),
                (u"FECHA CONSTATACIÓN", 15),
                (u"N°CONSTATACIÓN", 11),
                (u"USUARIO BIENES", 38),
                (u"UBICACIÓN", 38),
                (u"C.GOBIERNO", 15),
                (u"C.INTERNO", 15),
                (u"ACTIVO", 38),
                (u"SERIE", 15),
                (u"MODELO", 15),
                (u"MARCA", 15),
                (u"ENCONTRADO", 15),
                (u"EN USO", 15),
                (u"R.TRASPASO", 15),
                (u"ESTADO", 15),
                (u"PERTENECE A USUARIO", 15)
            ]

            fila = 6
            for col_num in range(len(columnas)):
                hojadestino.write(fila, col_num, columnas[col_num][0], fcabeceracolumna)
                hojadestino.set_column(col_num, col_num, columnas[col_num][1])

            fila = 7

            if not codigo:
                constataciones = ConstatacionFisica.objects.filter(status=True, estado=2,
                                                                   fechainicio__range=[fechadesde, fechahasta],
                                                                   fechafin__range=[fechadesde, fechahasta]).order_by(
                    'numero')
            else:
                constataciones = ConstatacionFisica.objects.filter(
                    Q(detalleconstatacionfisica__activo__codigogobierno=codigo) | Q(
                        detalleconstatacionfisica__activo__codigointerno=codigo), status=True, estado=2,
                    fechainicio__range=[fechadesde, fechahasta], fechafin__range=[fechadesde, fechahasta]).order_by(
                    'numero')

            secuencia = 0

            for constatacion in constataciones:
                secuencia += 1
                print(secuencia)
                hojadestino.write(fila, 0, secuencia, fceldageneral)
                hojadestino.write(fila, 1, constatacion.fechainicio, fceldafecha)
                hojadestino.write(fila, 2, constatacion.numero, fceldageneral)
                hojadestino.write(fila, 3, constatacion.usuariobienes.nombre_completo_inverso(), fceldageneral)
                hojadestino.write(fila, 4, constatacion.ubicacionbienes.nombre, fceldageneral)

                secudeta = 0
                # Activos que pertenecen al usuario
                if not codigo:
                    detalles = constatacion.detalle_constatacion().filter(perteneceusuario=True).order_by('id')
                else:
                    detalles = constatacion.detalle_constatacion().filter(
                        Q(activo__codigogobierno=codigo) | Q(activo__codigointerno=codigo),
                        perteneceusuario=True).order_by('id')

                for detalle in detalles:
                    secudeta += 1
                    if secudeta > 1:
                        secuencia += 1
                        # print(secuencia)
                        hojadestino.write(fila, 0, secuencia, fceldageneral)
                        hojadestino.write(fila, 1, constatacion.fechainicio, fceldafecha)
                        hojadestino.write(fila, 2, constatacion.numero, fceldageneral)
                        hojadestino.write(fila, 3, constatacion.usuariobienes.nombre_completo_inverso(), fceldageneral)
                        hojadestino.write(fila, 4, constatacion.ubicacionbienes.nombre, fceldageneral)

                    hojadestino.write(fila, 5, detalle.activo.codigogobierno, fceldageneral)
                    hojadestino.write(fila, 6, detalle.activo.codigointerno, fceldageneral)
                    hojadestino.write(fila, 7, detalle.activo.descripcion, fceldageneral)
                    hojadestino.write(fila, 8, detalle.activo.serie, fceldageneral)
                    hojadestino.write(fila, 9, detalle.activo.modelo, fceldageneral)
                    hojadestino.write(fila, 10, detalle.activo.marca, fceldageneral)
                    hojadestino.write(fila, 11, "SI" if detalle.encontrado else "NO", fceldageneral)
                    hojadestino.write(fila, 12, "SI" if detalle.enuso else "NO", fceldageneral)
                    hojadestino.write(fila, 13, "SI" if detalle.requieretraspaso else "NO", fceldageneral)
                    hojadestino.write(fila, 14, detalle.estadoactual.nombre, fceldageneral)
                    hojadestino.write(fila, 15, "PERTENECE", fceldageneral)

                    fila += 1

                # Activos que NO pertenecen al usuario
                if not codigo:
                    detalles = constatacion.detalle_constatacion().filter(perteneceusuario=False).order_by('id')
                else:
                    detalles = constatacion.detalle_constatacion().filter(
                        Q(activo__codigogobierno=codigo) | Q(activo__codigointerno=codigo),
                        perteneceusuario=False).order_by('id')

                for detalle in detalles:
                    secudeta += 1
                    if secudeta > 1:
                        secuencia += 1
                        # print(secuencia)
                        hojadestino.write(fila, 0, secuencia, fceldageneral)
                        hojadestino.write(fila, 1, constatacion.fechainicio, fceldafecha)
                        hojadestino.write(fila, 2, constatacion.numero, fceldageneral)
                        hojadestino.write(fila, 3, constatacion.usuariobienes.nombre_completo_inverso(), fceldageneral)
                        hojadestino.write(fila, 4, constatacion.ubicacionbienes.nombre, fceldageneral)

                    hojadestino.write(fila, 5, detalle.activo.codigogobierno, fceldageneral)
                    hojadestino.write(fila, 6, detalle.activo.codigointerno, fceldageneral)
                    hojadestino.write(fila, 7, detalle.activo.descripcion, fceldageneral)
                    hojadestino.write(fila, 8, detalle.activo.serie, fceldageneral)
                    hojadestino.write(fila, 9, detalle.activo.modelo, fceldageneral)
                    hojadestino.write(fila, 10, detalle.activo.marca, fceldageneral)
                    hojadestino.write(fila, 11, "SI" if detalle.encontrado else "NO", fceldageneral)
                    hojadestino.write(fila, 12, "SI" if detalle.enuso else "NO", fceldageneral)
                    hojadestino.write(fila, 13, "SI" if detalle.requieretraspaso else "NO", fceldageneral)
                    hojadestino.write(fila, 14, detalle.estadoactual.nombre, fceldageneral)
                    hojadestino.write(fila, 15, "NO PERTENECE", fceldageneral)

                    fila += 1

            workbook.close()

            # Notificar al usuario
            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.cuerpo = 'Reporte Excel finalizado'
                notificacion.url = urlarchivo
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo='Reporte Excel finalizado',
                    titulo=tituloreporte,
                    destinatario=personanotifica,
                    url=urlarchivo,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False
                )
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte Excel finalizado",
                "body": tituloreporte,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": urlarchivo,
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "El <b>{}</b> ha sido generado con éxito".format(tituloreporte)
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.error = True
                notificacion.titulo = "Error al generar {}".format(tituloreporte)
                notificacion.cuerpo = textoerror
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo="Error al generar Reporte Excel",
                    titulo="Error al generar {}".format(tituloreporte),
                    destinatario=personanotifica,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False,
                    error=True)
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error al generar Reporte Excel",
                "body": "Error al generar {}".format(tituloreporte),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "Error al generar <b>{}</b>. Error: {}".format(tituloreporte, textoerror),
                "error": True
            }, ttl=500)


class reporte_cartera_vencida_general_vs_cobros_rubro_version_final_background(threading.Thread):

    def __init__(self, request, data, idnotificacion, fechacorte, fechapago):
        self.request = request
        self.data = data
        self.idnotificacion = idnotificacion
        self.fechacorte = fechacorte
        self.fechapago = fechapago
        threading.Thread.__init__(self)

    def run(self):
        try:
            request, data, idnotificacion, fechacorte, fechapago = self.request, self.data, self.idnotificacion, self.fechacorte, self.fechapago
            usernotify = User.objects.get(pk=request.user.pk)
            personanotifica = Persona.objects.get(usuario=usernotify)
            tituloreporte = "Reporte Excel Cartera Vencida General vs Cobros - Detalle Rubros"

            directorio = os.path.join(os.path.join(MEDIA_ROOT, 'reportes', 'posgrado'))
            try:
                os.stat(directorio)
            except:
                os.mkdir(directorio)

            nombrearchivo = "CARTERA_VENCIDA_GENERAL_VS_COBRO_RUBROS_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"
            urlarchivo = MEDIA_URL + "reportes/posgrado/" + nombrearchivo

            # Crea un nuevo archivo de excel y le agrega una hoja
            workbook = xlsxwriter.Workbook(directorio + '/' + nombrearchivo)
            hojadestino = workbook.add_worksheet("Reporte")

            fcabeceracolumna = workbook.add_format(FORMATOS_CELDAS_EXCEL["cabeceracolumna"])
            fceldageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneral"])
            fceldageneralcent = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneralcent"])
            ftitulo2izq = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo2izq"])
            ftitulo3izq = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo3izq"])
            fceldafecha = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdafecha"])
            fceldamoneda = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdamoneda"])
            fceldamonedapie = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdamonedapie"])
            fceldanegritaizq = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritaizq"])
            fceldanegritageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritageneral"])

            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

            fechasigdiacorte = fechacorte + relativedelta(days=1)

            textofechacorte = str(fechacorte.day) + "-" + meses[fechacorte.month - 1][:3].upper() + "-" + str(fechacorte.year)
            textofechapago = str(fechapago.day) + "-" + meses[fechapago.month - 1][:3].upper() + "-" + str(fechapago.year)
            textosigdiacorte = str(fechasigdiacorte.day) + "-" + meses[fechasigdiacorte.month - 1][:3].upper() + "-" + str(fechasigdiacorte.year)

            hojadestino.merge_range(0, 0, 0, 15, 'UNIVERSIDAD ESTATAL DE MILAGRO', ftitulo2izq)
            hojadestino.merge_range(1, 0, 1, 15, 'DIRECCIÓN ADMINISTRATIVA Y FINANCIERA', ftitulo2izq)
            hojadestino.merge_range(2, 0, 2, 15, 'CARTERA VENCIDA GENERAL VS COBROS - DETALLES RUBROS', ftitulo2izq)
            hojadestino.merge_range(3, 0, 3, 15, 'FECHA DE CORTE CARTERA VENCIDA: ' + str(fechacorte) + '', ftitulo2izq)
            hojadestino.merge_range(4, 0, 4, 15, 'FECHA DE CORTE DE COBROS: ' + str(fechapago), ftitulo2izq)
            hojadestino.merge_range(5, 0, 5, 15, 'FECHA DE DESCARGA DEL REPORTE (' + str(datetime.now().date()) + ')', ftitulo2izq)

            fila = 7

            columnas = [
                (u"N°", 3),
                (u"PROGRAMA DE MAESTRÍA", 30),
                (u"COHORTE", 11),
                (u"PERIODO (INICIO-FIN)", 20),
                (u"ESTADO DE MAESTRÍA", 15),
                (u"PROVINCIA", 15),
                (u"CANTÓN", 15),
                (u"CÉDULA", 15),
                (u"ESTUDIANTE", 38),
                (u"N° CUOTAS VENCIDAS", 15),
                (u"ID RUBRO", 15),
                (u"FECHA VENCIMIENTO", 15),
                (u"DÍAS VENCIMIENTO", 15),
                (u"VALOR VENCIDO", 15),
                (u"CATEGORÍA", 15),
                (u"RANGO DÍAS", 15),
                (u"FECHA DE PAGO", 15),
                (u"VALOR COBRADO", 15),
                (u"DIAS DE COBRO", 15),
                (u"TOTAL PAGADO", 15),
                (u"VALOR PENDIENTE", 15),
                (u"ESTADO ESTUDIANTE", 15),
                (u"FECHA ÚLTIMA CUOTA", 15),
                (u"ESTADO FINANCIAMIENTO", 15)
            ]

            for col_num in range(len(columnas)):
                hojadestino.write(fila, col_num, columnas[col_num][0], fcabeceracolumna)
                hojadestino.set_column(col_num, col_num, columnas[col_num][1])

            matriculas = Matricula.objects.filter(status=True, nivel__periodo__tipo__id__in=[3, 4]
                                                  # inscripcion__persona__cedula__in=cedulas,
                                                  # inscripcion__carrera__id__in=[173],
                                                  # nivel__periodo__id__in=[143]
                                                  ).exclude(nivel__periodo__pk__in=[120, 128]).distinct().order_by('inscripcion__persona__apellido1',
                                                                                                                   'inscripcion__persona__apellido2',
                                                                                                                   'inscripcion__persona__nombres')  # [:10]
            totalmatriculas = matriculas.count()

            print("Generación de archivo en proceso...")

            secuencia = 0
            registros = 0
            totalvencido = 0
            totalpendiente = 0
            totalcobrocartera = 0
            totalpagado = 0

            programas = []

            SUMANV = SUMAV = 0

            resumengeneral = {0: {'etiqueta': 'CARTERA VIGENTE', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'vencidoencurso': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': ''},
                              1: {'etiqueta': '1-30 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'vencidoencurso': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'A'},
                              2: {'etiqueta': '31-60 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'vencidoencurso': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'B'},
                              3: {'etiqueta': '61-90 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'vencidoencurso': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'C'},
                              4: {'etiqueta': '91-180 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'vencidoencurso': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'D'},
                              5: {'etiqueta': '181 DÍAS EN ADELANTE', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'vencidoencurso': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'E'}
                              }

            for matricula in matriculas:
                secuencia += 1

                alumno = matricula.inscripcion.persona.nombre_completo_inverso()
                personaalumno = matricula.inscripcion.persona
                rubrosalumno = matricula.rubros_maestria()
                ultimafechavence = ""

                # Verifico si el programa y cohorte existen en la lista de resumen
                existe = False
                indice = j = 0

                idprograma = matricula.inscripcion.carrera.id
                nombreprograma = matricula.inscripcion.carrera.nombre
                idperiodo = matricula.nivel.periodo.id
                numerocohorte = matricula.nivel.periodo.cohorte if matricula.nivel.periodo.cohorte else 0
                fechasperiodo = str(matricula.nivel.periodo.inicio) + " a " + str(matricula.nivel.periodo.fin)
                estadoprograma = "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO"

                for datoprograma in programas:
                    if datoprograma[0] == idprograma and datoprograma[1] == idperiodo:
                        # indice = j
                        existe = True
                        break

                    j += 1

                indice = j

                if not existe:
                    # Agrego el programa y cohorte a la lista de programas
                    resumen = {0: {'etiqueta': 'CARTERA VIGENTE', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': ''},
                               1: {'etiqueta': '1-30 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'A'},
                               2: {'etiqueta': '31-60 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'B'},
                               3: {'etiqueta': '61-90 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'C'},
                               4: {'etiqueta': '91-180 DÍAS', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'D'},
                               5: {'etiqueta': '181 DÍAS EN ADELANTE', 'estudiantes': 0, 'pagado': Decimal(0), 'vencido': Decimal(0), 'cobroscartera': Decimal(0), 'pendiente': Decimal(0), 'antiguedad': 'E'}
                               }
                    datoprograma = [idprograma, idperiodo, nombreprograma, numerocohorte, fechasperiodo, estadoprograma, resumen]
                    programas.append(datoprograma)

                datos = matricula.rubros_maestria_vencidos_vs_pagos_detalle_version_final(fechacorte, fechapago)

                if not matricula.retirado_programa_maestria():
                    # SIi hay rubros no vencidos
                    if datos['rubrosnovencidos']:
                        for rubro_no_vencido in datos['rubrosnovencidos']:
                            fila += 1

                            registros += 1

                            codigorubro = rubro_no_vencido[0]
                            valorpagado = rubro_no_vencido[3]
                            valorpendiente = rubro_no_vencido[4]
                            valorvencido = 0
                            valorvencidoencurso = 0
                            fechavence = rubro_no_vencido[1]
                            diasvencidos = rubro_no_vencido[6]
                            pagosvencidos = 1 if valorvencido > 0 or valorvencidoencurso > 0 else 0
                            fechacobrocartera = rubro_no_vencido[7]
                            valorcobrocartera = rubro_no_vencido[8]
                            diascobro = rubro_no_vencido[9]

                            categoriaantiguedad = ""

                            programas[indice][6][0]['estudiantes'] += 1
                            programas[indice][6][0]['pagado'] += valorpagado
                            programas[indice][6][0]['vencido'] += valorvencido
                            programas[indice][6][0]['cobroscartera'] += valorcobrocartera
                            programas[indice][6][0]['pendiente'] += valorpendiente

                            resumengeneral[0]['estudiantes'] += 1
                            resumengeneral[0]['pagado'] += valorpagado
                            resumengeneral[0]['vencido'] += valorvencido

                            resumengeneral[0]['vencidoencurso'] = 0
                            resumengeneral[0]['cobroscartera'] += valorcobrocartera
                            resumengeneral[0]['pendiente'] += valorpendiente

                            categoriaantiguedad = "VIGENTE"
                            rangodias = ""

                            hojadestino.write(fila, 0, registros, fceldageneral)
                            hojadestino.write(fila, 1, nombreprograma, fceldageneral)
                            hojadestino.write(fila, 2, numerocohorte, fceldageneralcent)
                            hojadestino.write(fila, 3, str(matricula.nivel.periodo.inicio) + " a " + str(matricula.nivel.periodo.fin), fceldageneralcent)
                            hojadestino.write(fila, 4, "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO", fceldageneral)
                            hojadestino.write(fila, 5, personaalumno.provincia.nombre if personaalumno.provincia else '', fceldageneral)
                            hojadestino.write(fila, 6, personaalumno.canton.nombre if personaalumno.canton else '', fceldageneral)
                            hojadestino.write(fila, 7, personaalumno.identificacion(), fceldageneral)
                            hojadestino.write(fila, 8, alumno, fceldageneral)
                            hojadestino.write(fila, 9, pagosvencidos, fceldageneral)
                            hojadestino.write(fila, 10, codigorubro, fceldageneral)
                            hojadestino.write(fila, 11, fechavence, fceldafecha)
                            hojadestino.write(fila, 12, diasvencidos, fceldageneral)
                            hojadestino.write(fila, 13, valorvencido, fceldamoneda)
                            hojadestino.write(fila, 14, categoriaantiguedad, fceldageneral)
                            hojadestino.write(fila, 15, rangodias, fceldageneral)
                            hojadestino.write(fila, 16, fechacobrocartera, fceldafecha)
                            hojadestino.write(fila, 17, valorcobrocartera, fceldamoneda)
                            hojadestino.write(fila, 18, diascobro if valorcobrocartera > 0 else '', fceldageneral)
                            hojadestino.write(fila, 19, valorpagado, fceldamoneda)
                            hojadestino.write(fila, 20, valorpendiente, fceldamoneda)
                            hojadestino.write(fila, 21, matricula.estado_inscripcion_maestria(), fceldageneral)

                            if rubrosalumno:
                                ultimafechavence = rubrosalumno.last().fechavence

                                if matricula.tiene_refinanciamiento_deuda_posgrado():
                                    estadofinanciamiento = "REFINANCIAMIENTO"
                                elif matricula.tiene_coactiva_posgrado():
                                    estadofinanciamiento = "COACTIVA"
                                elif datetime.now().date() <= ultimafechavence:
                                    estadofinanciamiento = "EN EJECUCIÓN"
                                else:
                                    estadofinanciamiento = "FINALIZADA"
                            else:
                                estadofinanciamiento = "NO TIENE RUBROS"

                            hojadestino.write(fila, 22, ultimafechavence, fceldafecha)
                            hojadestino.write(fila, 23, estadofinanciamiento, fceldageneral)

                            totalpagado += valorpagado
                            totalvencido += valorvencido
                            totalpendiente += valorpendiente
                            totalcobrocartera += valorcobrocartera

                    # SIi hay rubros no vencidos

                    # Si hay rubros vencidos: INICIO
                    if datos['rubrosvencidos']:
                        for rubro_vencido in datos['rubrosvencidos']:
                            fila += 1
                            registros += 1

                            codigorubro = rubro_vencido[0]
                            valorpagado = rubro_vencido[3]
                            valorpendiente = rubro_vencido[4]
                            valorvencido = rubro_vencido[5] if rubro_vencido[10] == 'NO' else 0
                            valorvencidoencurso = rubro_vencido[5] if rubro_vencido[10] == 'SI' else 0

                            fechavence = rubro_vencido[1]
                            diasvencidos = rubro_vencido[6]
                            pagosvencidos = 1 if valorvencido > 0 or valorvencidoencurso > 0 else 0
                            fechacobrocartera = rubro_vencido[7]
                            valorcobrocartera = rubro_vencido[8]
                            diascobro = rubro_vencido[9]

                            categoriaantiguedad = ""

                            if diasvencidos <= 30:
                                programas[indice][6][1]['estudiantes'] += 1
                                programas[indice][6][1]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][1]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][1]['vencido'] += valorvencidoencurso

                                programas[indice][6][1]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][1]['pendiente'] += valorpendiente

                                resumengeneral[1]['estudiantes'] += 1
                                resumengeneral[1]['pagado'] += valorpagado
                                resumengeneral[1]['vencido'] += valorvencido
                                resumengeneral[1]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[1]['cobroscartera'] += valorcobrocartera
                                resumengeneral[1]['pendiente'] += valorpendiente

                                # print(rubro_vencido[10], valorvencido, valorvencidoencurso)
                                SUMANV += valorvencido
                                SUMAV += valorvencidoencurso
                                # print(resumengeneral[1]['vencidoencurso'])

                                categoriaantiguedad = "A"
                                rangodias = "1-30"
                            elif diasvencidos <= 60:
                                programas[indice][6][2]['estudiantes'] += 1
                                programas[indice][6][2]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][2]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][2]['vencido'] += valorvencidoencurso

                                programas[indice][6][2]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][2]['pendiente'] += valorpendiente

                                resumengeneral[2]['estudiantes'] += 1
                                resumengeneral[2]['pagado'] += valorpagado
                                resumengeneral[2]['vencido'] += valorvencido
                                resumengeneral[2]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[2]['cobroscartera'] += valorcobrocartera
                                resumengeneral[2]['pendiente'] += valorpendiente

                                categoriaantiguedad = "B"
                                rangodias = "31-60"
                            elif diasvencidos <= 90:
                                programas[indice][6][3]['estudiantes'] += 1
                                programas[indice][6][3]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][3]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][3]['vencido'] += valorvencidoencurso

                                programas[indice][6][3]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][3]['pendiente'] += valorpendiente

                                resumengeneral[3]['estudiantes'] += 1
                                resumengeneral[3]['pagado'] += valorpagado
                                resumengeneral[3]['vencido'] += valorvencido
                                resumengeneral[3]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[3]['cobroscartera'] += valorcobrocartera
                                resumengeneral[3]['pendiente'] += valorpendiente

                                categoriaantiguedad = "C"
                                rangodias = "61-90"
                            elif diasvencidos <= 180:
                                programas[indice][6][4]['estudiantes'] += 1
                                programas[indice][6][4]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][4]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][4]['vencido'] += valorvencidoencurso

                                programas[indice][6][4]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][4]['pendiente'] += valorpendiente

                                resumengeneral[4]['estudiantes'] += 1
                                resumengeneral[4]['pagado'] += valorpagado
                                resumengeneral[4]['vencido'] += valorvencido
                                resumengeneral[4]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[4]['cobroscartera'] += valorcobrocartera
                                resumengeneral[4]['pendiente'] += valorpendiente

                                categoriaantiguedad = "D"
                                rangodias = "91-180"
                            else:
                                programas[indice][6][5]['estudiantes'] += 1
                                programas[indice][6][5]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][5]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][5]['vencido'] += valorvencidoencurso

                                programas[indice][6][5]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][5]['pendiente'] += valorpendiente

                                resumengeneral[5]['estudiantes'] += 1
                                resumengeneral[5]['pagado'] += valorpagado
                                resumengeneral[5]['vencido'] += valorvencido
                                resumengeneral[5]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[5]['cobroscartera'] += valorcobrocartera
                                resumengeneral[5]['pendiente'] += valorpendiente

                                categoriaantiguedad = "E"
                                rangodias = "181 DÍAS EN ADELANTE"

                            hojadestino.write(fila, 0, registros, fceldageneral)
                            hojadestino.write(fila, 1, nombreprograma, fceldageneral)
                            hojadestino.write(fila, 2, numerocohorte, fceldageneralcent)
                            hojadestino.write(fila, 3, str(matricula.nivel.periodo.inicio) + " a " + str(matricula.nivel.periodo.fin), fceldageneralcent)
                            hojadestino.write(fila, 4, "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO", fceldageneral)
                            hojadestino.write(fila, 5, personaalumno.provincia.nombre if personaalumno.provincia else '', fceldageneral)
                            hojadestino.write(fila, 6, personaalumno.canton.nombre if personaalumno.canton else '', fceldageneral)
                            hojadestino.write(fila, 7, personaalumno.identificacion(), fceldageneral)
                            hojadestino.write(fila, 8, alumno, fceldageneral)
                            hojadestino.write(fila, 9, pagosvencidos, fceldageneral)
                            hojadestino.write(fila, 10, codigorubro, fceldageneral)
                            hojadestino.write(fila, 11, fechavence, fceldafecha)
                            hojadestino.write(fila, 12, diasvencidos, fceldageneral)

                            if valorvencido > 0:
                                hojadestino.write(fila, 13, valorvencido, fceldamoneda)
                            else:
                                hojadestino.write(fila, 13, valorvencidoencurso, fceldamoneda)

                            hojadestino.write(fila, 14, categoriaantiguedad, fceldageneral)
                            hojadestino.write(fila, 15, rangodias, fceldageneral)
                            hojadestino.write(fila, 16, fechacobrocartera, fceldafecha)
                            hojadestino.write(fila, 17, valorcobrocartera, fceldamoneda)
                            hojadestino.write(fila, 18, diascobro if valorcobrocartera > 0 else '', fceldageneral)
                            hojadestino.write(fila, 19, valorpagado, fceldamoneda)
                            hojadestino.write(fila, 20, valorpendiente, fceldamoneda)
                            hojadestino.write(fila, 21, matricula.estado_inscripcion_maestria(), fceldageneral)

                            if rubrosalumno:
                                ultimafechavence = rubrosalumno.last().fechavence

                                if matricula.tiene_refinanciamiento_deuda_posgrado():
                                    estadofinanciamiento = "REFINANCIAMIENTO"
                                elif matricula.tiene_coactiva_posgrado():
                                    estadofinanciamiento = "COACTIVA"
                                elif datetime.now().date() <= ultimafechavence:
                                    estadofinanciamiento = "EN EJECUCIÓN"
                                else:
                                    estadofinanciamiento = "FINALIZADA"
                            else:
                                estadofinanciamiento = "NO TIENE RUBROS"

                            hojadestino.write(fila, 22, ultimafechavence, fceldafecha)
                            hojadestino.write(fila, 23, estadofinanciamiento, fceldageneral)

                            totalpagado += valorpagado

                            if valorvencido > 0:
                                totalvencido += valorvencido
                            else:
                                totalvencido += valorvencidoencurso

                            totalpendiente += valorpendiente
                            totalcobrocartera += valorcobrocartera

                    # Si hay rubros vencidos: FIN

                    # Si hay rubros vencimiento en curso: INICIO
                    if datos['rubrosvencimientocurso']:
                        for rubro_vencido in datos['rubrosvencimientocurso']:
                            fila += 1
                            registros += 1

                            codigorubro = rubro_vencido[0]
                            valorpagado = rubro_vencido[3]
                            valorpendiente = rubro_vencido[4]
                            valorvencido = rubro_vencido[5] if rubro_vencido[10] == 'NO' else 0
                            valorvencidoencurso = rubro_vencido[5] if rubro_vencido[10] == 'SI' else 0
                            fechavence = rubro_vencido[1]
                            diasvencidos = rubro_vencido[6]
                            pagosvencidos = 1 if valorvencido > 0 or valorvencidoencurso > 0 else 0
                            fechacobrocartera = rubro_vencido[7]
                            valorcobrocartera = rubro_vencido[8]
                            diascobro = rubro_vencido[9]

                            categoriaantiguedad = ""

                            if diasvencidos <= 30:
                                programas[indice][6][1]['estudiantes'] += 1
                                programas[indice][6][1]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][1]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][1]['vencido'] += valorvencidoencurso

                                programas[indice][6][1]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][1]['pendiente'] += valorpendiente

                                resumengeneral[1]['estudiantes'] += 1
                                resumengeneral[1]['pagado'] += valorpagado
                                resumengeneral[1]['vencido'] += valorvencido
                                resumengeneral[1]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[1]['cobroscartera'] += valorcobrocartera
                                resumengeneral[1]['pendiente'] += valorpendiente

                                # print(rubro_vencido[10], valorvencido, valorvencidoencurso)
                                SUMANV += valorvencido
                                SUMAV += valorvencidoencurso
                                # print(resumengeneral[1]['vencidoencurso'])

                                categoriaantiguedad = "A"
                                rangodias = "1-30"
                            elif diasvencidos <= 60:
                                programas[indice][6][2]['estudiantes'] += 1
                                programas[indice][6][2]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][2]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][2]['vencido'] += valorvencidoencurso

                                programas[indice][6][2]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][2]['pendiente'] += valorpendiente

                                resumengeneral[2]['estudiantes'] += 1
                                resumengeneral[2]['pagado'] += valorpagado
                                resumengeneral[2]['vencido'] += valorvencido
                                resumengeneral[2]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[2]['cobroscartera'] += valorcobrocartera
                                resumengeneral[2]['pendiente'] += valorpendiente

                                categoriaantiguedad = "B"
                                rangodias = "31-60"
                            elif diasvencidos <= 90:
                                programas[indice][6][3]['estudiantes'] += 1
                                programas[indice][6][3]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][3]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][3]['vencido'] += valorvencidoencurso

                                programas[indice][6][3]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][3]['pendiente'] += valorpendiente

                                resumengeneral[3]['estudiantes'] += 1
                                resumengeneral[3]['pagado'] += valorpagado
                                resumengeneral[3]['vencido'] += valorvencido
                                resumengeneral[3]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[3]['cobroscartera'] += valorcobrocartera
                                resumengeneral[3]['pendiente'] += valorpendiente

                                categoriaantiguedad = "C"
                                rangodias = "61-90"
                            elif diasvencidos <= 180:
                                programas[indice][6][4]['estudiantes'] += 1
                                programas[indice][6][4]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][4]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][4]['vencido'] += valorvencidoencurso

                                programas[indice][6][4]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][4]['pendiente'] += valorpendiente

                                resumengeneral[4]['estudiantes'] += 1
                                resumengeneral[4]['pagado'] += valorpagado
                                resumengeneral[4]['vencido'] += valorvencido
                                resumengeneral[4]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[4]['cobroscartera'] += valorcobrocartera
                                resumengeneral[4]['pendiente'] += valorpendiente

                                categoriaantiguedad = "D"
                                rangodias = "91-180"
                            else:
                                programas[indice][6][5]['estudiantes'] += 1
                                programas[indice][6][5]['pagado'] += valorpagado

                                if valorvencido > 0:
                                    programas[indice][6][5]['vencido'] += valorvencido
                                else:
                                    programas[indice][6][5]['vencido'] += valorvencidoencurso

                                programas[indice][6][5]['cobroscartera'] += valorcobrocartera
                                programas[indice][6][5]['pendiente'] += valorpendiente

                                resumengeneral[5]['estudiantes'] += 1
                                resumengeneral[5]['pagado'] += valorpagado
                                resumengeneral[5]['vencido'] += valorvencido
                                resumengeneral[5]['vencidoencurso'] += valorvencidoencurso
                                resumengeneral[5]['cobroscartera'] += valorcobrocartera
                                resumengeneral[5]['pendiente'] += valorpendiente

                                categoriaantiguedad = "E"
                                rangodias = "181 DÍAS EN ADELANTE"

                            hojadestino.write(fila, 0, registros, fceldageneral)
                            hojadestino.write(fila, 1, nombreprograma, fceldageneral)
                            hojadestino.write(fila, 2, numerocohorte, fceldageneralcent)
                            hojadestino.write(fila, 3, str(matricula.nivel.periodo.inicio) + " a " + str(matricula.nivel.periodo.fin), fceldageneralcent)
                            hojadestino.write(fila, 4, "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO", fceldageneral)
                            hojadestino.write(fila, 5, personaalumno.provincia.nombre if personaalumno.provincia else '', fceldageneral)
                            hojadestino.write(fila, 6, personaalumno.canton.nombre if personaalumno.canton else '', fceldageneral)
                            hojadestino.write(fila, 7, personaalumno.identificacion(), fceldageneral)
                            hojadestino.write(fila, 8, alumno, fceldageneral)
                            hojadestino.write(fila, 9, pagosvencidos, fceldageneral)
                            hojadestino.write(fila, 10, codigorubro, fceldageneral)
                            hojadestino.write(fila, 11, fechavence, fceldafecha)
                            hojadestino.write(fila, 12, diasvencidos, fceldageneral)

                            if valorvencido > 0:
                                hojadestino.write(fila, 13, valorvencido, fceldamoneda)
                            else:
                                hojadestino.write(fila, 13, valorvencidoencurso, fceldamoneda)

                            hojadestino.write(fila, 14, categoriaantiguedad, fceldageneral)
                            hojadestino.write(fila, 15, rangodias, fceldageneral)
                            hojadestino.write(fila, 16, fechacobrocartera, fceldafecha)
                            hojadestino.write(fila, 17, valorcobrocartera, fceldamoneda)
                            hojadestino.write(fila, 18, diascobro if valorcobrocartera > 0 else '', fceldageneral)
                            hojadestino.write(fila, 19, valorpagado, fceldamoneda)
                            hojadestino.write(fila, 20, valorpendiente, fceldamoneda)
                            hojadestino.write(fila, 21, matricula.estado_inscripcion_maestria(), fceldageneral)

                            if rubrosalumno:
                                ultimafechavence = rubrosalumno.last().fechavence

                                if matricula.tiene_refinanciamiento_deuda_posgrado():
                                    estadofinanciamiento = "REFINANCIAMIENTO"
                                elif matricula.tiene_coactiva_posgrado():
                                    estadofinanciamiento = "COACTIVA"
                                elif datetime.now().date() <= ultimafechavence:
                                    estadofinanciamiento = "EN EJECUCIÓN"
                                else:
                                    estadofinanciamiento = "FINALIZADA"
                            else:
                                estadofinanciamiento = "NO TIENE RUBROS"

                            hojadestino.write(fila, 22, ultimafechavence, fceldafecha)
                            hojadestino.write(fila, 23, estadofinanciamiento, fceldageneral)

                            totalpagado += valorpagado

                            if valorvencido > 0:
                                totalvencido += valorvencido
                            else:
                                totalvencido += valorvencidoencurso

                            totalpendiente += valorpendiente
                            totalcobrocartera += valorcobrocartera

                    # Si hay rubros vencidos: FIN
                else:
                    # RETIRADO
                    registros += 1
                    fila += 1

                    valorpagado = datos['totalpagado']
                    valorpendiente = datos['totalpendiente']
                    valorvencido = datos['totalvencido']
                    fechavence = datos['fechavence']
                    diasvencidos = datos['diasvencimiento']
                    pagosvencidos = 1 if valorvencido > 0 else 0
                    fechacobrocartera = ""
                    valorcobrocartera = 0
                    diascobro = 0

                    categoriaantiguedad = ""

                    if valorvencido > 0:
                        if diasvencidos <= 30:
                            programas[indice][6][1]['estudiantes'] += 1
                            programas[indice][6][1]['pagado'] += valorpagado
                            programas[indice][6][1]['vencido'] += valorvencido
                            programas[indice][6][1]['cobroscartera'] += valorcobrocartera
                            programas[indice][6][1]['pendiente'] += valorpendiente

                            resumengeneral[1]['estudiantes'] += 1
                            resumengeneral[1]['pagado'] += valorpagado
                            resumengeneral[1]['vencido'] += valorvencido
                            resumengeneral[1]['vencidoencurso'] += 0
                            resumengeneral[1]['cobroscartera'] += valorcobrocartera
                            resumengeneral[1]['pendiente'] += valorpendiente

                            categoriaantiguedad = "A"
                            rangodias = "1-30"
                        elif diasvencidos <= 60:
                            programas[indice][6][2]['estudiantes'] += 1
                            programas[indice][6][2]['pagado'] += valorpagado
                            programas[indice][6][2]['vencido'] += valorvencido
                            programas[indice][6][2]['cobroscartera'] += valorcobrocartera
                            programas[indice][6][2]['pendiente'] += valorpendiente

                            resumengeneral[2]['estudiantes'] += 1
                            resumengeneral[2]['pagado'] += valorpagado
                            resumengeneral[2]['vencido'] += valorvencido
                            resumengeneral[2]['vencidoencurso'] += 0
                            resumengeneral[2]['cobroscartera'] += valorcobrocartera
                            resumengeneral[2]['pendiente'] += valorpendiente

                            categoriaantiguedad = "B"
                            rangodias = "31-60"
                        elif diasvencidos <= 90:
                            programas[indice][6][3]['estudiantes'] += 1
                            programas[indice][6][3]['pagado'] += valorpagado
                            programas[indice][6][3]['vencido'] += valorvencido
                            programas[indice][6][3]['cobroscartera'] += valorcobrocartera
                            programas[indice][6][3]['pendiente'] += valorpendiente

                            resumengeneral[3]['estudiantes'] += 1
                            resumengeneral[3]['pagado'] += valorpagado
                            resumengeneral[3]['vencido'] += valorvencido
                            resumengeneral[3]['vencidoencurso'] += 0
                            resumengeneral[3]['cobroscartera'] += valorcobrocartera
                            resumengeneral[3]['pendiente'] += valorpendiente

                            categoriaantiguedad = "C"
                            rangodias = "61-90"
                        elif diasvencidos <= 180:
                            programas[indice][6][4]['estudiantes'] += 1
                            programas[indice][6][4]['pagado'] += valorpagado
                            programas[indice][6][4]['vencido'] += valorvencido
                            programas[indice][6][4]['cobroscartera'] += valorcobrocartera
                            programas[indice][6][4]['pendiente'] += valorpendiente

                            resumengeneral[4]['estudiantes'] += 1
                            resumengeneral[4]['pagado'] += valorpagado
                            resumengeneral[4]['vencido'] += valorvencido
                            resumengeneral[4]['vencidoencurso'] += 0
                            resumengeneral[4]['cobroscartera'] += valorcobrocartera
                            resumengeneral[4]['pendiente'] += valorpendiente

                            categoriaantiguedad = "D"
                            rangodias = "91-180"
                        else:
                            programas[indice][6][5]['estudiantes'] += 1
                            programas[indice][6][5]['pagado'] += valorpagado
                            programas[indice][6][5]['vencido'] += valorvencido
                            programas[indice][6][5]['cobroscartera'] += valorcobrocartera
                            programas[indice][6][5]['pendiente'] += valorpendiente

                            resumengeneral[5]['estudiantes'] += 1
                            resumengeneral[5]['pagado'] += valorpagado
                            resumengeneral[5]['vencido'] += valorvencido
                            resumengeneral[5]['vencidoencurso'] += 0
                            resumengeneral[5]['cobroscartera'] += valorcobrocartera
                            resumengeneral[5]['pendiente'] += valorpendiente

                            categoriaantiguedad = "E"
                            rangodias = "181 DÍAS EN ADELANTE"
                    else:
                        programas[indice][6][0]['estudiantes'] += 1
                        programas[indice][6][0]['pagado'] += valorpagado
                        programas[indice][6][0]['vencido'] += valorvencido
                        programas[indice][6][0]['cobroscartera'] += valorcobrocartera
                        programas[indice][6][0]['pendiente'] += valorpendiente

                        resumengeneral[0]['estudiantes'] += 1
                        resumengeneral[0]['pagado'] += valorpagado
                        resumengeneral[0]['vencido'] += valorvencido
                        resumengeneral[0]['vencidoencurso'] += 0
                        resumengeneral[0]['cobroscartera'] += valorcobrocartera
                        resumengeneral[0]['pendiente'] += valorpendiente

                        categoriaantiguedad = "VIGENTE"
                        rangodias = ""

                    hojadestino.write(fila, 0, registros, fceldageneral)
                    hojadestino.write(fila, 1, nombreprograma, fceldageneral)
                    hojadestino.write(fila, 2, numerocohorte, fceldageneralcent)
                    hojadestino.write(fila, 3, str(matricula.nivel.periodo.inicio) + " a " + str(matricula.nivel.periodo.fin), fceldageneralcent)
                    hojadestino.write(fila, 4, "EN EJECUCIÓN" if datetime.now().date() <= matricula.nivel.periodo.fin else "FINALIZADO", fceldageneral)
                    hojadestino.write(fila, 5, personaalumno.provincia.nombre if personaalumno.provincia else '', fceldageneral)
                    hojadestino.write(fila, 6, personaalumno.canton.nombre if personaalumno.canton else '', fceldageneral)
                    hojadestino.write(fila, 7, personaalumno.identificacion(), fceldageneral)
                    hojadestino.write(fila, 8, alumno, fceldageneral)
                    hojadestino.write(fila, 9, pagosvencidos, fceldageneral)
                    hojadestino.write(fila, 10, "S/N", fceldageneral)
                    hojadestino.write(fila, 11, fechavence, fceldafecha)
                    hojadestino.write(fila, 12, diasvencidos, fceldageneral)
                    hojadestino.write(fila, 13, valorvencido, fceldamoneda)
                    hojadestino.write(fila, 14, categoriaantiguedad, fceldageneral)
                    hojadestino.write(fila, 15, rangodias, fceldageneral)
                    hojadestino.write(fila, 16, fechacobrocartera, fceldafecha)
                    hojadestino.write(fila, 17, valorcobrocartera, fceldamoneda)
                    hojadestino.write(fila, 18, diascobro if valorcobrocartera > 0 else '', fceldageneral)
                    hojadestino.write(fila, 19, valorpagado, fceldamoneda)
                    hojadestino.write(fila, 20, valorpendiente, fceldamoneda)
                    hojadestino.write(fila, 21, matricula.estado_inscripcion_maestria(), fceldageneral)

                    if rubrosalumno:
                        ultimafechavence = rubrosalumno.last().fechavence

                        if matricula.tiene_refinanciamiento_deuda_posgrado():
                            estadofinanciamiento = "REFINANCIAMIENTO"
                        elif matricula.tiene_coactiva_posgrado():
                            estadofinanciamiento = "COACTIVA"
                        elif datetime.now().date() <= ultimafechavence:
                            estadofinanciamiento = "EN EJECUCIÓN"
                        else:
                            estadofinanciamiento = "FINALIZADA"
                    else:
                        estadofinanciamiento = "NO TIENE RUBROS"

                    hojadestino.write(fila, 22, ultimafechavence, fceldafecha)
                    hojadestino.write(fila, 23, estadofinanciamiento, fceldageneral)

                    totalpagado += valorpagado
                    totalvencido += valorvencido
                    totalpendiente += valorpendiente
                    totalcobrocartera += valorcobrocartera

                print(secuencia, " de ", totalmatriculas)

            fila += 1
            hojadestino.merge_range(fila, 0, fila, 12, "TOTALES", fceldanegritageneral)
            hojadestino.write(fila, 13, totalvencido, fceldamonedapie)
            hojadestino.write(fila, 17, totalcobrocartera, fceldamonedapie)
            hojadestino.write(fila, 19, totalpagado, fceldamonedapie)
            hojadestino.write(fila, 20, totalpendiente, fceldamonedapie)

            fila += 3
            hojadestino.merge_range(fila, 0, fila, 11, "RESUMEN DE CARTERA VENCIDA GENERAL", ftitulo3izq)

            fila += 1

            hojadestino.merge_range(fila, 0, fila, 1, "PERIODO DE VENCIMIENTO", fcabeceracolumna)
            hojadestino.write(fila, 2, "NRO.ESTUDIANTES", fcabeceracolumna)
            hojadestino.write(fila, 3, "VALOR PAGADO", fcabeceracolumna)
            hojadestino.write(fila, 4, "VALOR CARTERA VENCIDA AL " + textofechacorte, fcabeceracolumna)
            hojadestino.write(fila, 5, "VALOR CARTERA VENCIDA DEL " + textosigdiacorte + " AL " + textofechapago, fcabeceracolumna)
            hojadestino.write(fila, 6, "COBROS CARTERA AL " + textofechapago, fcabeceracolumna)
            hojadestino.write(fila, 7, "VALOR CARTERA VENCIDA AL " + textofechapago, fcabeceracolumna)
            hojadestino.write(fila, 8, "VALOR PENDIENTE", fcabeceracolumna)
            hojadestino.write(fila, 9, "ANTIGUEDAD", fcabeceracolumna)

            fila += 1
            totalvencidoresgen = totalvencidoencurso = totalsaldofinal = 0

            for i in resumengeneral:
                hojadestino.merge_range(fila, 0, fila, 1, resumengeneral[i]['etiqueta'], fceldageneral)
                hojadestino.write(fila, 2, resumengeneral[i]['estudiantes'], fceldageneral)
                hojadestino.write(fila, 3, resumengeneral[i]['pagado'], fceldamoneda)
                hojadestino.write(fila, 4, resumengeneral[i]['vencido'], fceldamoneda)
                hojadestino.write(fila, 5, resumengeneral[i]['vencidoencurso'], fceldamoneda)
                hojadestino.write(fila, 6, resumengeneral[i]['cobroscartera'], fceldamoneda)

                totalvencidoresgen += resumengeneral[i]['vencido']
                totalvencidoencurso += resumengeneral[i]['vencidoencurso']

                saldofinal = (resumengeneral[i]['vencido'] + resumengeneral[i]['vencidoencurso']) - resumengeneral[i]['cobroscartera']
                totalsaldofinal += saldofinal

                hojadestino.write(fila, 7, saldofinal, fceldamoneda)

                hojadestino.write(fila, 8, resumengeneral[i]['pendiente'], fceldamoneda)
                hojadestino.write(fila, 9, resumengeneral[i]['antiguedad'], fceldageneralcent)
                fila += 1

            hojadestino.merge_range(fila, 0, fila, 1, "TOTAL", fceldanegritageneral)
            hojadestino.write(fila, 2, registros, fceldanegritageneral)
            hojadestino.write(fila, 3, totalpagado, fceldamonedapie)
            hojadestino.write(fila, 4, totalvencidoresgen, fceldamonedapie)
            hojadestino.write(fila, 5, totalvencidoencurso, fceldamonedapie)
            hojadestino.write(fila, 6, totalcobrocartera, fceldamonedapie)
            hojadestino.write(fila, 7, totalsaldofinal, fceldamonedapie)
            hojadestino.write(fila, 8, totalpendiente, fceldamonedapie)

            fila += 3
            hojadestino.merge_range(fila, 0, fila, 11, "RESUMEN DE CARTERA VENCIDA POR PROGRAMA DE MAESTRÍA", ftitulo3izq)

            fila += 1
            hojadestino.write(fila, 0, "N°", fcabeceracolumna)
            hojadestino.write(fila, 1, "PROGRAMA", fcabeceracolumna)
            hojadestino.write(fila, 2, "COHORTE", fcabeceracolumna)
            hojadestino.write(fila, 3, "PERIODO (INICIO-FIN)", fcabeceracolumna)
            hojadestino.write(fila, 4, "ESTADO DE MAESTRIA", fcabeceracolumna)
            hojadestino.write(fila, 5, "PERIODO DE VENCIMIENTO", fcabeceracolumna)
            hojadestino.write(fila, 6, "NRO.ESTUDIANTES", fcabeceracolumna)
            hojadestino.write(fila, 7, "VALOR PAGADO", fcabeceracolumna)
            hojadestino.write(fila, 8, "VALOR CARTERA VENCIDA", fcabeceracolumna)
            hojadestino.write(fila, 9, "COBROS CARTERA A FECHA DE CORTE", fcabeceracolumna)
            hojadestino.write(fila, 10, "VALOR PENDIENTE", fcabeceracolumna)
            hojadestino.write(fila, 11, "ANTIGUEDAD", fcabeceracolumna)

            # Ordeno por programa y cohorte
            programas = sorted(programas, key=lambda programa: (programa[2], programa[3]))

            secresumen = 0
            for datoprograma in programas:
                fila += 1
                secresumen += 1
                hojadestino.merge_range(fila, 0, fila + 5, 0, secresumen, fceldageneralcent)
                hojadestino.merge_range(fila, 1, fila + 5, 1, datoprograma[2], fceldageneralcent)
                hojadestino.merge_range(fila, 2, fila + 5, 2, datoprograma[3], fceldageneralcent)
                hojadestino.merge_range(fila, 3, fila + 5, 3, datoprograma[4], fceldageneralcent)
                hojadestino.merge_range(fila, 4, fila + 5, 4, datoprograma[5], fceldageneralcent)

                tot_est_prog = tot_venc_prog = tot_pend_prog = tot_pag_prog = tot_cobrocart_prog = 0

                resumen = datoprograma[6]
                for i in resumen:
                    hojadestino.write(fila, 5, resumen[i]['etiqueta'], fceldageneral)
                    hojadestino.write(fila, 6, resumen[i]['estudiantes'], fceldageneral)
                    hojadestino.write(fila, 7, resumen[i]['pagado'], fceldamoneda)
                    hojadestino.write(fila, 8, resumen[i]['vencido'], fceldamoneda)
                    hojadestino.write(fila, 9, resumen[i]['cobroscartera'], fceldamoneda)
                    hojadestino.write(fila, 10, resumen[i]['pendiente'], fceldamoneda)
                    hojadestino.write(fila, 11, resumen[i]['antiguedad'], fceldageneralcent)

                    tot_est_prog += resumen[i]['estudiantes']
                    tot_pag_prog += resumen[i]['pagado']
                    tot_venc_prog += resumen[i]['vencido']
                    tot_pend_prog += resumen[i]['pendiente']
                    tot_cobrocart_prog += resumen[i]['cobroscartera']

                    fila += 1

                hojadestino.merge_range(fila, 0, fila, 5, "TOTAL " + datoprograma[2] + " COHORTE " + str(datoprograma[3]), fceldanegritageneral)
                hojadestino.write(fila, 6, tot_est_prog, fceldanegritageneral)
                hojadestino.write(fila, 7, tot_pag_prog, fceldamonedapie)
                hojadestino.write(fila, 8, tot_venc_prog, fceldamonedapie)
                hojadestino.write(fila, 9, tot_cobrocart_prog, fceldamonedapie)
                hojadestino.write(fila, 10, tot_pend_prog, fceldamonedapie)

            fila += 1
            hojadestino.merge_range(fila, 0, fila, 5, "TOTAL GENERAL", fceldanegritageneral)
            hojadestino.write(fila, 6, registros, fceldanegritageneral)
            hojadestino.write(fila, 7, totalpagado, fceldamonedapie)
            hojadestino.write(fila, 8, totalvencido, fceldamonedapie)
            hojadestino.write(fila, 9, totalcobrocartera, fceldamonedapie)
            hojadestino.write(fila, 10, totalpendiente, fceldamonedapie)

            workbook.close()

            # Notificar al usuario
            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.cuerpo = 'Reporte Excel finalizado'
                notificacion.url = urlarchivo
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo='Reporte Excel finalizado',
                    titulo=tituloreporte,
                    destinatario=personanotifica,
                    url=urlarchivo,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False
                )
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte Excel finalizado",
                "body": tituloreporte,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": urlarchivo,
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "El <b>{}</b> ha sido generado con éxito".format(tituloreporte)
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.error = True
                notificacion.titulo = "Error al generar {}".format(tituloreporte)
                notificacion.cuerpo = textoerror
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo="Error al generar Reporte Excel",
                    titulo="Error al generar {}".format(tituloreporte),
                    destinatario=personanotifica,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False,
                    error=True)
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error al generar Reporte Excel",
                "body": "Error al generar {}".format(tituloreporte),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "Error al generar <b>{}</b>. Error: {}".format(tituloreporte, textoerror),
                "error": True
            }, ttl=500)


class reporte_generar_csv_facturas_esigef_background(threading.Thread):

    def __init__(self, request, data, idnotificacion, tiporeporte, fechadesde, fechahasta):
        self.request = request
        self.data = data
        self.idnotificacion = idnotificacion
        self.tiporeporte = tiporeporte
        self.fechadesde = fechadesde
        self.fechahasta = fechahasta
        threading.Thread.__init__(self)

    def run(self):
        try:
            request, data, idnotificacion, tiporeporte, fechadesde, fechahasta = self.request, self.data, self.idnotificacion, self.tiporeporte, self.fechadesde, self.fechahasta
            usernotify = User.objects.get(pk=request.user.pk)
            personanotifica = Persona.objects.get(usuario=usernotify)
            tituloreporte = "Reporte CSV Facturas Emitidas para subir al eSIGEF"

            directorio = os.path.join(os.path.join(MEDIA_ROOT, 'reportes', 'tesoreria'))
            print(directorio)
            try:
                os.stat(directorio)
            except:
                os.mkdir(directorio)

            nombrearchivo = "facturas_emitidas_subir_esigef_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".csv"

            # Consulto las facturas
            if tiporeporte == 1:
                facturas = Factura.objects.filter(status=True, fecha__gte=datetime.strptime('2023-08-01', "%Y-%m-%d").date(), enviadacliente=True, archivofacturaesigefdetalle__isnull=True).order_by('id')
            else:
                facturas = Factura.objects.filter(status=True, fecha__range = [fechadesde, fechahasta], enviadacliente=True, archivofacturaesigefdetalle__isnull=True).order_by('id')

            # Creo el registro de archivo csv
            archivoesigef = ArchivoFacturaEsigef(
                fecha=datetime.now().date(),
                cantidad=facturas.count()
            )
            archivoesigef.save(request)

            # Crear el archivo csv con las facturas
            with open(directorio + '/' + nombrearchivo, 'w', newline='') as csvfile:
                spamwriter = csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
                spamwriter.writerow(['RUC Receptor',
                                     'Nombre del Receptor',
                                     'Tipo de Comprobante',
                                     'No. Autorizacion',
                                     'Nro. Serie - Establecimiento',
                                     'Nro. Serie - Pto Emision',
                                     'Nro. Comprobante',
                                     'Fecha Comprobante',
                                     'Tipo de Venta',
                                     'Base Impuesto Tarifa IVA 0%',
                                     'Base Impuesto Tarifa 12%',
                                     'Base Impuesto No Objeto IVA',
                                     'Monto IVA'])

                for factura in facturas:
                    spamwriter.writerow([factura.identificacion,
                                         elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(factura.nombre))),
                                         '01',
                                         factura.autorizacion,
                                         factura.puntoventa.establecimiento,
                                         factura.puntoventa.puntoventa,
                                         str(factura.numero).zfill(9),
                                         factura.fecha.strftime("%d/%m/%Y"),
                                         '403' if factura.subtotal_base0 > 0 else '401',
                                         factura.subtotal_base0,
                                         factura.subtotal_base_iva,
                                         '0.00',
                                         factura.total_iva])

                    # Creo el registro del detalle de archivo csv
                    detallearchivo = ArchivoFacturaEsigefDetalle(
                        archivoesigef=archivoesigef,
                        factura=factura
                    )
                    detallearchivo.save(request)


            archivo = SITE_STORAGE + '/media/reportes/tesoreria/' + nombrearchivo

            # Aperturo el archivo generado
            with open(archivo, 'rb') as f:
                data = f.read()

            buffer = io.BytesIO()
            buffer.write(data)
            pdfcopia = buffer.getvalue()
            buffer.seek(0)
            buffer.close()

            # Extraigo el contenido
            archivocopiado = ContentFile(pdfcopia)
            archivocopiado.name = nombrearchivo

            # Asignar el archivo al registro archivo csv
            archivoesigef.archivo = archivocopiado
            archivoesigef.save(request)

            # Borro el archivo creado en /media/reportes/tesoreria/
            os.remove(archivo)

            urlarchivo = archivoesigef.archivo.url

            # Notificar al usuario
            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.cuerpo = 'Reporte Excel finalizado'
                notificacion.url = urlarchivo
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo='Reporte Excel finalizado',
                    titulo=tituloreporte,
                    destinatario=personanotifica,
                    url=urlarchivo,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False
                )
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Reporte Excel finalizado",
                "body": tituloreporte,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": urlarchivo,
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "El <b>{}</b> ha sido generado con éxito".format(tituloreporte)
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.error = True
                notificacion.titulo = "Error al generar {}".format(tituloreporte)
                notificacion.cuerpo = textoerror
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo="Error al generar Reporte Excel",
                    titulo="Error al generar {}".format(tituloreporte),
                    destinatario=personanotifica,
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False,
                    error=True)
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error al generar Reporte Excel",
                "body": "Error al generar {}".format(tituloreporte),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "Error al generar <b>{}</b>. Error: {}".format(tituloreporte, textoerror),
                "error": True
            }, ttl=500)

class reporte_auditoria_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            nombre_archivo = generar_nombre("reporte_auditoria", '') + '.xlsx'
            url = ''
            directory = os.path.join(SITE_STORAGE, 'media', 'auditoria')
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)

            # Inicializo cabecera de  excel
            __author__ = 'Unemi'
            wb = openxl.Workbook()
            ws = wb.active
            style_title = openxlFont(name='Arial', size=14, bold=True)
            style_cab = openxlFont(name='Arial', size=10, bold=True)
            alinear = alin(horizontal="center", vertical="center")
            response = HttpResponse(content_type="application/ms-excel")
            response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + '-' + random.randint(1, 10000).__str__() + '.xlsx'
            ws.merge_cells('A1:E1')
            ws.merge_cells('A2:E2')
            ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO'
            ws['A2'] = 'REPORTE DE AUDITORIA'
            celda1 = ws['A1']
            celda1.font = style_title
            celda1.alignment = alinear
            celda2 = ws['A2']
            celda2.font = style_title
            celda2.alignment = alinear
            columns = ['Fecha',
                       'Hora',
                       'Acción',
                       'Usuario',
                       'Mensaje',
                       ]
            row_num = 3
            for col_num in range(0, len(columns)):
                celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                celda.font = style_cab
            row_num = 4

            #Obtencion de directorio y varibles requeridas
            directory = os.path.join(MEDIA_ROOT, 'auditoria', nombre_archivo)
            pers = Persona.objects.get(usuario_id=request.user.pk)
            baseDate = datetime.today()
            year = request.POST['year'] if 'year' in request.POST and request.POST['year'] else baseDate.year
            month = request.POST['month'] if 'month' in request.POST and request.POST['month'] else baseDate.month
            data['idi'] = request.POST['id']
            persona = data['persona']

            # Consultas y agrupacion de listas
            logs = LogEntry.objects.filter(Q(change_message__icontains=persona.__str__()) | Q(user=persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
            logs1 = LogEntryBackup.objects.filter(Q(change_message__icontains=persona.__str__()) | Q(user=persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
            logs2 = LogEntryBackupdos.objects.filter(Q(change_message__icontains=persona.__str__()) | Q(user=persona.usuario),
                                                     action_time__year=year).exclude(user__is_superuser=True)
            logs3 = LogEntryLogin.objects.filter(user=persona.usuario, action_time__year=year).exclude(user__is_superuser=True, action_app=2)
            if int(month):
                logs = logs.filter(action_time__month=month)
                logs1 = logs1.filter(action_time__month=month)
                logs2 = logs2.filter(action_time__month=month)
                logs3 = logs3.filter(action_time__month=month)
            logslist0 = list(logs.values_list("action_time", "action_flag", "change_message", "user__username"))
            logslist1 = list(logs1.values_list("action_time", "action_flag", "change_message", "user__username"))
            logslist2 = list(logs2.values_list("action_time", "action_flag", "change_message", "user__username"))
            logslist = logslist0 + logslist1 + logslist2
            accioneslogin = {1:'Exitoso', 2:'Fallido'}
            acciones = {1:'Agregar', 2:'Editar', 3:'Eliminar'}
            for xItem in logslist:
                accion = acciones[xItem[1]] if xItem[1] in acciones else 'Otro'
                ws.cell(row=row_num, column=1, value=xItem[0].date())
                ws.cell(row=row_num, column=2, value= xItem[0].time())
                ws.cell(row=row_num, column=3, value=accion)
                ws.cell(row=row_num, column=4, value= xItem[3])
                ws.cell(row=row_num, column=5, value=xItem[2])
                row_num += 1
            for xItem in list(logs3.values_list("action_time", "action_flag", "change_message", "user__username", "id")):
                l = LogEntryLogin.objects.get(pk=xItem[4])
                accion = accioneslogin[xItem[1]] if xItem[1] in accioneslogin else 'Desconocido'
                ws.cell(row=row_num, column=1, value=xItem[0].date())
                ws.cell(row=row_num, column=2, value=xItem[0].time())
                ws.cell(row=row_num, column=3, value=accion)
                ws.cell(row=row_num, column=4, value=xItem[3])
                ws.cell(row=row_num, column=5, value=l.get_data_message())
                row_num += 1
            wb.save(directory)

            # NOTIFICACIÓN
            titulo = 'Reporte de auditoria generado exitosamente'
            cuerpo = 'Su reporte de auditoria se genero correctamente y puede ser descargada por la url proporcionada.'
            url = "{}auditoria/{}".format(MEDIA_URL, nombre_archivo)

        except Exception as ex:
            titulo = 'Algo salio mal, reporte de auditoria no generado'
            cuerpo = 'Su reporte de auditoria no fue generado'
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
        if notif > 0:
            noti = Notificacion.objects.get(pk=notif)
            noti.titulo = titulo
            noti.en_proceso = False
            noti.cuerpo = cuerpo
            noti.url = url
            noti.save()
        else:
            noti = Notificacion(cuerpo=cuerpo,
                                titulo=titulo,
                                destinatario=pers, url=url,
                                prioridad=1, app_label='sga-sagest',
                                fecha_hora_visible=datetime.now() + timedelta(days=1),
                                tipo=2, en_proceso=False)
            noti.save(request)

        send_user_notification(user=request.user, payload={
            "head": titulo,
            "body": cuerpo,
            "action": "notificacion",
            "timestamp": time.mktime(datetime.now().timetuple()),
            "url": url,
            "btn_notificaciones": traerNotificaciones(request, data, pers),
            "mensaje": titulo
        }, ttl=500)


class reporte_distributivo_asignaturas_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notif, periodo = self.request, self.data, self.notiid, self.periodo
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'reporte_distributivo_asignaturas' + random.randint(1, 10000).__str__() + '.xlsx'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:

            __author__ = 'Unemi'
            ahora = datetime.now()
            time_codigo = ahora.strftime('%Y%m%d_%H%M%S')
            name_file = f'reporte_excel_pedidos_online_{time_codigo}.xlsx'
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet("Listado Matriculados Nivelación")

            fuentecabecera = workbook.add_format({
                'align': 'center',
                'bg_color': 'silver',
                'border': 1,
                'bold': 1
            })

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            fuenteencabezado = workbook.add_format({
                'align': 'center',
                'bg_color': '#1C3247',
                'font_color': 'white',
                'border': 1,
                'font_size': 24,
                'bold': 1
            })

            columnas = [
                (u"FACULTAD", 60),
                (u"CARRERA", 60),
                (u"MALLA", 60),
                (u"SECCIÓN", 60),
                (u"NIVEL", 60),
                (u"PARALELO", 60),
                (u"ASIGNATURA", 60),
                (u"TEORICA PRACTICA", 60),
                (u"CUPO", 60),
                (u"MATRICULADOS", 60),
                (u"INSCRITOS", 60),  #
                (u"TOTAL MATRICULADOS", 60),
                (u"DOCENTE", 60),
                (u"CEDULA", 60),
                # (u"USUARIO", 6000),
                (u"AFINIDAD", 60),
                (u"HORAS SEMANALES", 40),
                (u"MALLA (HORAS PRESENCIALES SEMANALES)", 40),
                (u"TIPO", 40),
                (u"CORREO PERSONAL", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"TIPO PROFESOR", 40),
                (u"PROFESOR DESDE", 40),
                (u"PROFESOR HASTA", 40),
                (u"DEDICACION", 50),
                (u"CATEGORIA", 40),
                (u"INICIO MATERIA", 40),
                (u"FIN MATERIA", 40),
                (u"FIN ASISTENCIA", 40),
                # (u"ID", 4000),
                (u"TELEFONO", 60),
                (u"MODELO EVALUATIVO", 60),
                (u"IDMATERIA", 25),
                (u"ACEPTACION", 25),
                (u"OBSERVACION ACEPTACION", 20),
                (u"HORARIO FECHA ACEPTACION", 20),
                (u"HORARIO ACEPTACION", 20),
                (u"HORARIO OBSERVACION ACEPTACION", 20),
                (u"IDMOODLE", 25),
                (u"PROVINCIA", 25),
                (u"CIUDAD", 25),
                (u"CANT ACEPTADOS", 60),
                (u"MODALIDAD IMPARTICION", 60),
                (u"ES PRÁCTICA", 40),
                (u"SILABO", 40),
                (u"TIENE GUIA PRÁCTICA", 40),
                (u"CERRADA", 40),
            ]
            ws.merge_range(0, 0, 0, columnas.__len__() - 1, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, columnas.__len__() - 1, f'DISTRIBUTIVO DE ASIGNATURAS',fuenteencabezado)
            row_num, numcolum = 2, 0
            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1
            row_num += 1

            cursor = connections['default'].cursor()
            sql = f"""SET statement_timeout='20000 s';
                                        SELECT sga_coordinacion.nombre AS Facultad, 
                                            sga_carrera.nombre AS Carrera, sga_sesion.nombre AS Seccion, 
                                            sga_nivelmalla.nombre AS Nivel, sga_materia.paralelo AS Paralelo, 
                                            sga_materia.id AS Idmateria, sga_asignatura.nombre AS Asignatura, 
                                            sga_persona.apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres AS Docente, 
                                            sga_profesormateria.hora AS sga_profesormateria_hora, (CASE sga_profesormateria.principal WHEN TRUE THEN 'PRINCIPAL' ELSE 'PRACTICA' END) AS Tipo, sga_persona.cedula, (
                                        SELECT u.username
                                        FROM auth_user u
                                        WHERE u.id=sga_persona.usuario_id), sga_persona.email, sga_persona.emailinst, sga_materia.cupo AS cupo, (
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada ma, sga_matricula mat1
                                        WHERE ma.matricula_id=mat1.id AND ma.status=True AND mat1.status=True AND mat1.estado_matricula in (2,3) AND ma.materia_id=sga_materia.id AND ma.id NOT in (
                                        SELECT mr.materiaasignada_id
                                        FROM sga_materiaasignadaretiro mr)) AS nmatriculados, sga_tipoprofesor.nombre AS Tipoprofesor,
                                         sga_profesormateria.desde AS desde, sga_profesormateria.hasta AS hasta, (
                                        SELECT ti.nombre
                                        FROM sga_profesordistributivohoras dis,sga_tiempodedicaciondocente ti
                                        WHERE dis.dedicacion_id=ti.id AND dis.profesor_id=sga_profesor.id AND periodo_id={periodo} AND dis.status= TRUE) AS dedicacion, (
                                        SELECT ca.nombre
                                        FROM sga_profesordistributivohoras dis,sga_categorizaciondocente ca
                                        WHERE dis.categoria_id=ca.id AND dis.profesor_id=sga_profesor.id AND dis.periodo_id={periodo} AND dis.status= TRUE) AS categoria, 
                                        (CASE sga_asignaturamalla.practicas WHEN TRUE THEN 'SI' ELSE 'NO' END) AS tipomateria, 
                                        (CASE sga_profesormateria.afinidad WHEN TRUE THEN 'SI' ELSE 'NO' END) AS afinidad, 
                                        sga_materia.inicio AS inicio, sga_materia.fin AS fin, sga_materia.fechafinasistencias AS finasistencia, sga_materia.id AS id, sga_persona.telefono_conv AS telefonoconv, sga_persona.telefono AS telefono, EXTRACT(YEAR
                                        FROM sga_malla.inicio) AS anio, (
                                        SELECT modelo.nombre
                                        FROM sga_modeloevaluativo modelo
                                        WHERE modelo.id = sga_materia.modeloevaluativo_id) AS modeloevaluativo, sga_asignaturamalla.horaspresencialessemanales AS horaspresencialessemanales, sga_profesormateria.aceptarmateria AS aceptarmateria, sga_profesormateria.aceptarmateriaobs AS aceptarmateriaobs, sga_profesormateria.fecha_horario AS fecha_horario, sga_profesormateria.aceptarhorario AS aceptarhorario, sga_profesormateria.aceptarhorarioobs AS aceptarhorarioobs, sga_materia.idcursomoodle AS idcursomoodle, prov.nombre AS provincia, cant.nombre AS ciudad
                                        ,(
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada ma2, sga_matricula mat1
                                        WHERE ma2.matricula_id=mat1.id AND mat1.termino= TRUE AND ma2.materia_id=sga_materia.id AND ma2.id NOT in (
                                        SELECT mr.materiaasignada_id
                                        FROM sga_materiaasignadaretiro mr)) AS nmatriculados_acpta_termino,
                                        (
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                                        WHERE mat1.estado_matricula=1 AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id 
                                        AND ni1.periodo_id=sga_nivel.periodo_id AND mas1.materia_id=sga_materia.id) AS inscritos,
                                        (select CASE WHEN sga_detalleasignaturamallamodalidad.modalidad = 1 THEN 'PRESENCIAL' else 'VIRTUAL' END) as modalidad,
                                        (CASE sga_asignaturamalla.practicas 
                                        WHEN True THEN 'SI' ELSE 'NO' END) AS Malla_Practicas,
                                        (CASE 
                                            (SELECT count(*) 
                                                    FROM sga_silabo AS sga_s 
                                                    WHERE sga_s.materia_id = sga_materia.id AND sga_s.status=True AND sga_s.codigoqr=True
                                                ) 
                                            WHEN 0 THEN
                                                'NO'
                                            ELSE 
                                                'SI'
                                            END
                                        ) AS silabo,
                                        (SELECT 
                                            count(sga_tp.*) 
                                        FROM sga_tareapracticasilabosemanal AS sga_tp
                                        INNER JOIN sga_silabosemanal AS sga_ss ON sga_tp.silabosemanal_id = sga_ss.id
                                        INNER JOIN sga_silabo AS sga_s ON sga_ss.silabo_id = sga_s.id
                                        INNER JOIN sga_materia AS sga_m ON sga_m.id = sga_s.materia_id
                                        WHERE 
                                            sga_s.status= TRUE 
                                            AND sga_s.codigoqr= TRUE 
                                            AND sga_m.id = sga_materia.id 
                                            AND sga_tp.estado_id!=3 
                                            AND sga_tp.status=True
                                        ) AS trabajos_practicos,
                                        sga_materia.cerrado
                                        FROM public.sga_materia sga_materia
                                        LEFT JOIN public.sga_profesormateria sga_profesormateria ON sga_materia.id = sga_profesormateria.materia_id AND sga_profesormateria.status= TRUE AND sga_profesormateria.activo= TRUE
                                        LEFT JOIN public.sga_profesor sga_profesor ON sga_profesor.id = sga_profesormateria.profesor_id AND sga_profesor.status= TRUE
                                        LEFT JOIN public.sga_tipoprofesor sga_tipoprofesor ON sga_tipoprofesor.id = sga_profesormateria.tipoprofesor_id AND sga_tipoprofesor.status= TRUE
                                        LEFT JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id AND sga_persona.status= TRUE
                                        LEFT JOIN sga_provincia prov ON prov.id=sga_persona.provincia_id
                                        LEFT JOIN sga_canton cant ON cant.id=sga_persona.canton_id
                                        INNER JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id AND sga_nivel.status= TRUE
                                        INNER JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id AND sga_asignatura.status= TRUE
                                        INNER JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id AND sga_asignaturamalla.status= TRUE
                                        LEFT JOIN public.sga_detalleasignaturamallamodalidad sga_detalleasignaturamallamodalidad ON sga_asignaturamalla.id = sga_detalleasignaturamallamodalidad.asignaturamalla_id 
                                        INNER JOIN public.sga_nivelmalla sga_nivelmalla ON sga_asignaturamalla.nivelmalla_id = sga_nivelmalla.id AND sga_nivelmalla.status= TRUE
                                        INNER JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id AND sga_malla.status= TRUE
                                        INNER JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id AND sga_carrera.status= TRUE
                                        INNER JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id
                                        INNER JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id AND sga_coordinacion.status= TRUE
                                        INNER JOIN public.sga_sesion sga_sesion ON sga_nivel.sesion_id = sga_sesion.id
                                        INNER JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id
                                        WHERE sga_periodo.id = {periodo} AND sga_materia.status= TRUE
                                        ORDER BY sga_coordinacion.nombre, sga_carrera.nombre, sga_sesion.nombre, sga_nivelmalla.nombre,sga_materia.paralelo,sga_asignatura.nombre
                                """
            cursor.execute(sql)
            results = cursor.fetchall()

            for r in results:
                campo1 = r[0].__str__()
                campo2 = r[1].__str__()
                campo3 = r[2].__str__()
                campo4 = r[3].__str__()
                campo5 = r[4].__str__()
                campo6 = r[5]
                campo7 = r[6].__str__()
                campo8 = r[7].__str__()
                campo9 = r[8]
                campo10 = r[9].__str__()
                campo11 = r[10].__str__()
                # campo12 = r[11]
                campo13 = r[12].__str__()
                campo14 = r[13].__str__()
                campo15 = int(r[14])
                campo16 = int(r[15])
                campo17 = r[16].__str__()
                campo18 = r[17].__str__()
                campo19 = r[18].__str__()
                campo20 = r[19].__str__()
                campo21 = r[20].__str__()
                campo22 = r[21].__str__()
                campo23 = r[22].__str__()
                campo24 = r[23].__str__()
                campo25 = r[24].__str__()
                campo26 = r[25].__str__()
                # campo27 = r[26]
                campo28 = r[27].__str__() + " - " + r[28].__str__()
                campo29 = r[29].__str__()
                campo30 = r[30].__str__()
                campo31 = r[31].__str__()
                if r[33] == None or r[33] == '':
                    campo32 = ''
                    campo33 = ''
                else:
                    campo32 = 'NO'
                    if r[32]:
                        campo32 = 'SI'
                    campo33 = r[33]
                campo34 = r[34]
                if r[36] == None or r[36] == '':
                    campo35 = ''
                    campo36 = ''
                else:
                    campo35 = 'NO'
                    if r[35]:
                        campo35 = 'SI'
                    campo36 = r[36]
                campo37 = r[37]
                campo38 = r[38]
                campo39 = r[39]
                campo_cant_acapta = r[40]
                inscritos = int(r[41])
                campo40 = r[42].__str__()
                campo41 = r[43]
                campo42 = r[44]
                campo43 = r[45]
                campo46 = r[46]
                totalMatriculados = inscritos + int(campo16)
                ws.write(row_num, 0, campo1, formatoceldacenter)
                ws.write(row_num, 1, campo2, formatoceldacenter)
                ws.write(row_num, 2, campo29, formatoceldacenter)
                ws.write(row_num, 3, campo3, formatoceldacenter)
                ws.write(row_num, 4, campo4, formatoceldacenter)
                ws.write(row_num, 5, campo5, formatoceldacenter)
                ws.write(row_num, 6, campo7, formatoceldacenter)
                ws.write(row_num, 7, campo22, formatoceldacenter)
                ws.write(row_num, 8, campo15, formatoceldacenter)
                ws.write(row_num, 9, campo16, formatoceldacenter)
                ws.write(row_num, 10, inscritos, formatoceldacenter)
                ws.write(row_num, 11, totalMatriculados, formatoceldacenter)

                ws.write(row_num, 12, campo8, formatoceldacenter)
                ws.write(row_num, 13, campo11, formatoceldacenter)
                # ws.write(row_num, 12, campo12, font_style2)
                ws.write(row_num, 14, campo23, formatoceldacenter)
                ws.write(row_num, 15, campo9, formatoceldacenter)
                ws.write(row_num, 16, campo9, formatoceldacenter)
                ws.write(row_num, 17, campo10, formatoceldacenter)
                ws.write(row_num, 18, campo13, formatoceldacenter)
                ws.write(row_num, 19, campo14, formatoceldacenter)
                ws.write(row_num, 20, campo17, formatoceldacenter)
                ws.write(row_num, 21, campo18, formatoceldacenter)
                ws.write(row_num, 22, campo19, formatoceldacenter)
                ws.write(row_num, 23, campo20, formatoceldacenter)
                ws.write(row_num, 24, campo21, formatoceldacenter)
                ws.write(row_num, 25, campo24, formatoceldacenter)
                ws.write(row_num, 26, campo25, formatoceldacenter)
                ws.write(row_num, 27, campo26, formatoceldacenter)
                # ws.write(row_num, 25, campo27, font_style2)
                ws.write(row_num, 28, campo28, formatoceldacenter)
                ws.write(row_num, 29, campo30, formatoceldacenter)
                ws.write(row_num, 30, campo6, formatoceldacenter)
                ws.write(row_num, 31, campo32, formatoceldacenter)
                ws.write(row_num, 32, campo33, formatoceldacenter)
                ws.write(row_num, 33, str(campo34), formatoceldacenter)
                ws.write(row_num, 34, campo35, formatoceldacenter)
                ws.write(row_num, 35, campo36, formatoceldacenter)
                ws.write(row_num, 36, campo37, formatoceldacenter)
                ws.write(row_num, 37, campo38, formatoceldacenter)
                ws.write(row_num, 38, campo39, formatoceldacenter)
                ws.write(row_num, 39, campo_cant_acapta, formatoceldacenter)
                ws.write(row_num, 40, campo40, formatoceldacenter)
                ws.write(row_num, 41, campo41, formatoceldacenter)
                ws.write(row_num, 42, campo42, formatoceldacenter)
                ws.write(row_num, 43, campo43, formatoceldacenter)
                ws.write(row_num, 44, campo46, formatoceldacenter)
                row_num += 1
            workbook.close()

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel Distributivo de asignaturas',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel Distributivo de asignaturas',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)


        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Excel Distributivo de asignaturas',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error en el reporte",
                "body": 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)


class reporte_criterios_actividades_formacion_docente_background(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notif, periodo = self.request, self.data, self.notiid, self.periodo
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = 'listado_actividaddocente_materia' + random.randint(1, 10000).__str__() + '.xlsx'
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:

            __author__ = 'Unemi'
            ahora = datetime.now()
            time_codigo = ahora.strftime('%Y%m%d_%H%M%S')
            name_file = f'listado_actividaddocente_materia{time_codigo}.xlsx'
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet("Listado")

            fuentecabecera = workbook.add_format({
                'align': 'center',
                'bg_color': 'silver',
                'border': 1,
                'bold': 1
            })

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            fuenteencabezado = workbook.add_format({
                'align': 'center',
                'bg_color': '#1C3247',
                'font_color': 'white',
                'border': 1,
                'font_size': 24,
                'bold': 1
            })

            columnas = [
                (u"CARRERA", 30),
                (u"MALLA", 30),
                (u"SECCIÓN", 30),
                (u"NIVEL", 30),
                (u"PARALELO", 30),
                (u"ACTIVIDADES", 60),
                (u"SUBACTIVIDADES", 60),
                (u"SUBACTIVIDADES FECHA DESDE", 60),
                (u"SUBACTIVIDADES FECHA FIN", 60),
                (u"SUBACTIVIDADES HORAS", 60),  #
                (u"MATERIAS", 60),
                (u"HORAS", 60),
                (u"APELLIDOS Y NOMBRES", 60),
                (u"CRITERIO", 60),
                (u"CRITERIO", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"CATEGORIZACIÓN", 40),
                (u"TIPO", 40),
                (u"DEDICACION", 40),
                (u"FACULTAD", 40),
                (u"TIPO PROFESOR", 40),
                (u"TITULO TERCER NIVEL", 40),
                (u"TITULO MASTER", 40),
                (u"TITULO PHD", 50),
                (u"ETNIA", 40),
                (u"SEXO", 40),
                (u"LGTBI", 40)
            ]
            ws.merge_range(0, 0, 0, columnas.__len__() - 1, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, columnas.__len__() - 1, f'DISTRIBUTIVO DE ACTIVIDADES Y ASIGNATURAS',fuenteencabezado)
            row_num, numcolum = 2, 0
            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1
            row_num += 1

            cursor = connections['default'].cursor()
            sql = f"""
                        SELECT  case cri.tipo when 1 then 'Docencia' ELSE 'Vinculación' END as criterio,coor.nombre as facultad,per.apellido1, per.apellido2 , per.nombres as docente,  
                        null,null as nivel,null, cri.nombre as actividad, actdetdis.nombre as subactividades, actdetdis.desde as subactividades_fechadesde, 
                        actdetdis.hasta as subactividades_fechahasta, actdetdis.horas as subactividades_horas, null as ateria,detdis.horas,us.username, 
                        td.nombre, cat.nombre as categoria,null,per.cedula,per.emailinst,us.username,  null as tipoprofesor,  
                        master.nombremaster,phd.cantidad as phd,
                        etn.nombre as etnia, 
                        sex.nombre as sexo,per.lgtbi, tercernivel.cantidad AS tercernivel ,pt1.nombre as tipoprofesor,0,0  
                        from sga_profesordistributivohoras dis
                        join sga_detalledistributivo detdis on dis.id=detdis.distributivo_id
                        join sga_criteriodocenciaperiodo critd on detdis.criteriodocenciaperiodo_id=critd.id
                        join sga_tiempodedicaciondocente td on td.id=dis.dedicacion_id
                        join sga_criteriodocencia cri on critd.criterio_id=cri.id
                        join sga_profesor pro on dis.profesor_id=pro.id
                        join sga_persona per on pro.persona_id=per.id
                        join auth_user us on per.usuario_id=us.id
                        join sga_coordinacion coor on dis.coordinacion_id=coor.id
                        join sga_actividaddetalledistributivo actdetdis on actdetdis.criterio_id = detdis.id
                        join sga_categorizaciondocente cat on cat.id=dis.categoria_id
                        join sga_profesortipo pt1 on pt1.id=dis.nivelcategoria_id
                        join sga_sexo sex on sex.id=per.sexo_id
                        left join (select perfil.persona_id, raza.nombre from sga_perfilinscripcion perfil
                        join sga_raza raza on perfil.raza_id=raza.id
                        where perfil.status=True) etn on etn.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre),',') as cantidad
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id=1
                        group by tc.persona_id) phd on phd.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre),',') as nombremaster
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id in(2, 5)
                        group by tc.persona_id) master on master.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre || '-' || insti.nombre),',') as cantidad
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        join sga_institucioneducacionsuperior insti on insti.id=tc.institucion_id 
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id in(2, 5)
                        group by tc.persona_id) tercernivel on tercernivel.persona_id=per.id
                        where detdis.criteriodocenciaperiodo_id is not null 
                        and dis.periodo_id={periodo} 
                        and cri.id not in (15,16,17,18) 
                        UNION ALL
                        SELECT  'Investigacion' as criterio,coor.nombre as facultad,per.apellido1, per.apellido2 , per.nombres as docente,  
                        null,null as nivel,null, cri.nombre as actividad, actdetdis.nombre as subactividades, actdetdis.desde as subactividades_fechadesde, 
                        actdetdis.hasta as subactividades_fechahasta, actdetdis.horas as subactividades_horas, null as ateria,detdis.horas,us.username, 
                        td.nombre, cat.nombre as categoria,null,per.cedula,per.emailinst,us.username,  null as tipoprofesor,  
                        master.nombremaster,phd.cantidad as phd,
                        etn.nombre as etnia, 
                        sex.nombre as sexo,per.lgtbi, tercernivel.cantidad AS tercernivel ,pt1.nombre as tipoprofesor,0,0  
                        from sga_profesordistributivohoras dis
                        join sga_detalledistributivo detdis on dis.id=detdis.distributivo_id
                        join sga_criterioinvestigacionperiodo critd on detdis.criterioinvestigacionperiodo_id=critd.id
                        join sga_tiempodedicaciondocente td on td.id=dis.dedicacion_id
                        join sga_criterioinvestigacion cri on critd.criterio_id=cri.id
                        join sga_profesor pro on dis.profesor_id=pro.id
                        join sga_persona per on pro.persona_id=per.id
                        join auth_user us on per.usuario_id=us.id
                        join sga_coordinacion coor on dis.coordinacion_id=coor.id
                        join sga_actividaddetalledistributivo actdetdis on actdetdis.criterio_id = detdis.id
                        join sga_categorizaciondocente cat on cat.id=dis.categoria_id
                        join sga_profesortipo pt1 on pt1.id=dis.nivelcategoria_id
                        join sga_sexo sex on sex.id=per.sexo_id
                        left join (select perfil.persona_id, raza.nombre from sga_perfilinscripcion perfil
                        join sga_raza raza on perfil.raza_id=raza.id
                        where perfil.status=True) etn on etn.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre),',') as cantidad
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id=1
                        group by tc.persona_id) phd on phd.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre),',') as nombremaster
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id in(2, 5)
                        group by tc.persona_id) master on master.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre || '-' || insti.nombre),',') as cantidad
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        join sga_institucioneducacionsuperior insti on insti.id=tc.institucion_id 
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id in(2, 5)
                        group by tc.persona_id) tercernivel on tercernivel.persona_id=per.id
                        where detdis.criterioinvestigacionperiodo_id is not null 
                        and dis.periodo_id={periodo} 
                        UNION ALL
                        SELECT  'Gestion' as criterio,coor.nombre as facultad,per.apellido1, per.apellido2 , per.nombres as docente,  
                        null,null as nivel,null, cri.nombre as actividad, actdetdis.nombre as subactividades, actdetdis.desde as subactividades_fechadesde, 
                        actdetdis.hasta as subactividades_fechahasta, actdetdis.horas as subactividades_horas, null as ateria,detdis.horas,us.username, 
                        td.nombre, cat.nombre as categoria,null,per.cedula,per.emailinst,us.username,  null as tipoprofesor,  
                        master.nombremaster,phd.cantidad as phd,
                        etn.nombre as etnia, 
                        sex.nombre as sexo,per.lgtbi, tercernivel.cantidad AS tercernivel ,pt1.nombre as tipoprofesor,0,0  
                        from sga_profesordistributivohoras dis
                        join sga_detalledistributivo detdis on dis.id=detdis.distributivo_id
                        join sga_criteriogestionperiodo critd on detdis.criteriogestionperiodo_id=critd.id
                        join sga_tiempodedicaciondocente td on td.id=dis.dedicacion_id
                        join sga_criteriogestion cri on critd.criterio_id=cri.id
                        join sga_profesor pro on dis.profesor_id=pro.id
                        join sga_persona per on pro.persona_id=per.id
                        join auth_user us on per.usuario_id=us.id
                        join sga_coordinacion coor on dis.coordinacion_id=coor.id
                        join sga_actividaddetalledistributivo actdetdis on actdetdis.criterio_id = detdis.id
                        join sga_categorizaciondocente cat on cat.id=dis.categoria_id
                        join sga_profesortipo pt1 on pt1.id=dis.nivelcategoria_id
                        join sga_sexo sex on sex.id=per.sexo_id
                        left join (select perfil.persona_id, raza.nombre from sga_perfilinscripcion perfil
                        join sga_raza raza on perfil.raza_id=raza.id
                        where perfil.status=True) etn on etn.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre),',') as cantidad
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id=1
                        group by tc.persona_id) phd on phd.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre),',') as nombremaster
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id in(2, 5)
                        group by tc.persona_id) master on master.persona_id=per.id
                        left join (select tc.persona_id, array_to_string(array_agg(tit.nombre || '-' || insti.nombre),',') as cantidad
                        from sga_titulacion tc 
                        join sga_titulo tit on tc.titulo_id=tit.id
                        join sga_institucioneducacionsuperior insti on insti.id=tc.institucion_id 
                        where  tc.status and tc."status"=True and tit.nivel_id=4 and tit.grado_id in(2, 5)
                        group by tc.persona_id) tercernivel on tercernivel.persona_id=per.id
                        where detdis.criteriogestionperiodo_id is not null 
                        and dis.periodo_id={periodo} 
                        UNION ALL
                        SELECT 'Materias' as criterio,coor.nombre as facultad,per.apellido1 , per.apellido2 , per.nombres as docente,   
                        mat.paralelo,nmalla.nombre as nivel,carr.nombre || ' ' || carr.mencion as carreras, null as actividad, null as subactividades, 
                        null as subactividades_fechadesde, null as subactividades_fechahasta, null as subactividades_horas, 
                        asig.nombre as materia,pmat.hora,us.username ,td.nombre,
                        (select cat.nombre from sga_categorizaciondocente cat, sga_profesordistributivohoras dist  
                        where cat.id=dist.categoria_id and dist.periodo_id={periodo}  and dist.profesor_id=pro.id ),ses.nombre as sesion,  
                        per.cedula,per.emailinst,us.username,tipro.nombre as tipoprofesor,
                        array_to_string(array_agg(CASE WHEN tit.nivel_id = 4 AND tit.grado_id IN (2, 5) THEN tit.nombre || '-' || ins.nombre END), ',') as master,
                        array_to_string(array_agg(CASE WHEN tit.nivel_id = 4 AND tit.grado_id = 1 THEN tit.nombre END), ',') as phd,
                        (select (raza.nombre) from sga_perfilinscripcion perfil,sga_raza raza where perfil.raza_id=raza.id 
                        and perfil.persona_id=per.id and perfil.status=True)  as etnia,   
                        (select nombre from sga_sexo sexo where sexo.id=per.sexo_id ) as sexo,per.lgtbi,  
                        array_to_string(array_agg(CASE WHEN tit.nivel_id = 3 THEN tit.nombre || '-' || ins.nombre END), ',') as tercernivel,  
                        (select pt1.nombre from sga_profesortipo pt1, sga_profesordistributivohoras dist 
                        where pt1.id=dist.nivelcategoria_id and dist.periodo_id={periodo}  and dist.profesor_id=pro.id) as tipoprofesor,
                        extract(year from malla.inicio) as anio,0  
                        FROM sga_profesormateria pmat
                        JOIN sga_materia mat ON pmat.materia_id=mat.id  
                        JOIN sga_nivel niv ON mat.nivel_id=niv.id
                        JOIN sga_profesor pro ON pmat.profesor_id=pro.id
                        JOIN sga_persona per ON pro.persona_id=per.id
                        JOIN sga_asignaturamalla asimalla ON mat.asignaturamalla_id=asimalla.id 
                        JOIN sga_malla malla ON asimalla.malla_id=malla.id   
                        JOIN sga_carrera carr ON malla.carrera_id=carr.id 
                        JOIN sga_asignatura asig ON asimalla.asignatura_id=asig.id
                        JOIN sga_nivelmalla nmalla ON asimalla.nivelmalla_id=nmalla.id
                        JOIN auth_user us ON per.usuario_id=us.id 
                        JOIN sga_coordinacion_carrera corcar ON corcar.carrera_id=carr.id 
                        JOIN sga_coordinacion coor ON corcar.coordinacion_id=coor.id 
                        JOIN sga_tiempodedicaciondocente td ON pro.dedicacion_id=td.id 
                        JOIN sga_sesion ses ON niv.sesion_id=ses.id 
                        join sga_tipoprofesor tipro ON pmat.tipoprofesor_id=tipro.id
                        LEFT JOIN sga_titulacion tc ON tc.persona_id = per.id
                        LEFT JOIN sga_titulo tit ON tc.titulo_id = tit.id
                        LEFT JOIN sga_institucioneducacionsuperior ins ON ins.id = tc.institucion_id
                        WHERE niv.periodo_id={periodo}
                        AND tipro.id not in (4)
                        GROUP BY 
                        per.id,coor.nombre,per.apellido1,per.apellido2,per.nombres, mat.paralelo,nmalla.nombre,
                        us.username,td.nombre,carr.nombre,carr.mencion,asig.nombre,pmat.hora,pro.id,ses.nombre,
                        tipro.nombre,malla.inicio,per.cedula,per.cedula,per.emailinst,per.lgtbi
                        """
            cursor.execute(sql)
            results = cursor.fetchall()

            for per in results:
                ws.write(row_num, 0, per[7], formatoceldacenter)
                ws.write(row_num, 1, per[30] if per[30] != 0 else '', formatoceldacenter)
                ws.write(row_num, 2, per[18], formatoceldacenter)
                ws.write(row_num, 3, per[6], formatoceldacenter)
                ws.write(row_num, 4, per[5], formatoceldacenter)
                ws.write(row_num, 5, per[8], formatoceldacenter)
                ws.write(row_num, 6, per[9], formatoceldacenter)
                ws.write(row_num, 7, per[10].strftime("%d/%m/%Y") if per[10] else '', formatoceldacenter)
                ws.write(row_num, 8, per[11].strftime("%d/%m/%Y") if per[11] else '', formatoceldacenter)
                ws.write(row_num, 9, per[12], formatoceldacenter)
                ws.write(row_num, 10, per[13], formatoceldacenter)
                ws.write(row_num, 11, per[14], formatoceldacenter)
                ws.write(row_num, 12, per[2] + ' ' + per[3] + ' ' + per[4], formatoceldacenter)
                ws.write(row_num, 13, per[0], formatoceldacenter)
                ws.write(row_num, 14, per[19], formatoceldacenter)
                ws.write(row_num, 15, per[21] + '@unemi.edu.ec', formatoceldacenter)
                ws.write(row_num, 16, per[17], formatoceldacenter)
                ws.write(row_num, 17, per[29], formatoceldacenter)
                ws.write(row_num, 18, per[16], formatoceldacenter)
                ws.write(row_num, 19, per[1], formatoceldacenter)
                ws.write(row_num, 20, per[22], formatoceldacenter)
                ws.write(row_num, 21, per[28], formatoceldacenter)
                ws.write(row_num, 22, per[23], formatoceldacenter)
                ws.write(row_num, 23, per[24], formatoceldacenter)
                ws.write(row_num, 24, per[25], formatoceldacenter)
                ws.write(row_num, 25, per[26], formatoceldacenter)
                row_num += 1
            workbook.close()

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel actividad docente por materia',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel actividad docente por materia',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)


        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Excel actividad docente por materia',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error en el reporte",
                "body": 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte no ha sido generado'
            }, ttl=500)


class reporte_matrizprofesoresdistributivo_background(threading.Thread):

    def __init__(self, request, data, notiid):
        self.request = request
        self.data = data
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):

        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
        request, data, notiid = self.request, self.data, self.notiid
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)

        nombre_archivo = "reporte_matrices_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)

        try:
            if 'id' in request.GET:
                periodo = Periodo.objects.get(pk=int(request.GET['id']))
                __author__ = 'Unemi'
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                # ws.write_merge(0, 0, 0, 15, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response[
                    'Content-Disposition'] = 'attachment; filename=Profesores_distribucionh ' + str(
                    periodo) + ' ' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"CODIGO_IES", 6000),
                    (u"TIPO_IDENTIFICACION", 6000),
                    (u"IDENTIFICACION", 6000),
                    (u"NUMERO_DE_DOCUMENTO", 6000),
                    (u"FECHA_INICIO", 3000),
                    (u"FECHA_FIN", 3000),
                    (u"ACCION_NUMERO_DE_DOCUMENTO", 6000),
                    (u"FECHA", 3000),
                    (u"HORAS_CLASE", 6000),
                    (u"HORAS_TUTORIA", 6000),
                    (u"HORAS_ADMINISTRATIVAS", 6000),
                    (u"HORAS_INVESTIGACION", 6000),
                    (u"HORAS_VINCULACION", 6000),
                    (u"HORAS_OTRAS_ACTIVIDADES", 6000),
                    (u"HORAS_CLASE_NIVEL_TECNICO", 6000),
                    (u"HORAS_CLASE_TERCER_NIVEL", 6000),
                    (u"HORAS_CLASE_CUARTO_NIVEL", 6000),
                    (u"CALIFICACION_ACTIVIDADES_DOCENCIA", 6000),
                    (u"CALIFICACION_ACTIVIDADES_INVESTIGACION", 6000),
                    (u"CALIFICACION_ACTIVIDADES_DIRECCION_GESTION_ACADEMICA", 6000),
                    (u"APELLIDOS Y NOMBRES", 7000),
                    (u"NIVEL CATEGORIA", 6000),
                    (u"CATEGORIA", 6000),
                    (u"DEDICACION", 6000)
                ]
                row_num = 0
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                cursor = connection.cursor()
                sql = "select tabla1.CODIGO_IES , (case substr(per.cedula,1,2) when 'VS' then 'PASAPORTE' else 'CEDULA' end) as TIPO_IDENTIFICACION , per.cedula as IDENTIFICACION,  " \
                      "COALESCE((select pc.numerodocumento from sagest_personacontratos pc, sga_periodo p where pc.persona_id=per.id and pc.fechainicio>=p.inicio and pc.fechainicio<=p.fin and p.id=%s ORDER BY pc.fechainicio desc LIMIT 1),'') as NUMERO_DOCUMENTO, " \
                      "(select pc.fechainicio from sagest_personacontratos pc, sga_periodo p where pc.persona_id=per.id and pc.fechainicio>=p.inicio and pc.fechainicio<=p.fin and p.id=%s ORDER BY pc.fechainicio desc LIMIT 1) as FINI_DOCUMENTO, " \
                      "(select pc.fechafin from sagest_personacontratos pc, sga_periodo p where pc.persona_id=per.id and pc.fechainicio>=p.inicio and pc.fechainicio<=p.fin and p.id=%s ORDER BY pc.fechainicio desc LIMIT 1) as FFIN_DOCUMENTO, " \
                      "COALESCE((SELECT pacc.numerodocumento FROM sagest_personaacciones pacc, sga_periodo p WHERE pacc.persona_id=per.id AND pacc.fecharige<=p.fin AND p.id=%s ORDER BY pacc.fecharige desc LIMIT 1),'') AS NUMERO_DOCUMENTO_ACCION," \
                      "(SELECT pacc.fecharige FROM sagest_personaacciones pacc, sga_periodo p WHERE pacc.persona_id=per.id AND pacc.fecharige<=p.fin AND p.id=%s ORDER BY pacc.fecharige desc LIMIT 1) AS FECHA_DOCUMENTO_ACCION," \
                      "tabla1.HORA_CLASE,tabla1.HORA_TUTORIA,tabla1.HORA_ADMINISTRACION,tabla1.HORA_INVESTIGACION,tabla1.HORA_VINCULACION,  " \
                      "(case when COALESCE((select tdd.horas from sga_profesordistributivohoras pdh, sga_tiempodedicaciondocente tdd, sga_profesor pro where pdh.periodo_id=%s and tdd.id=pdh.dedicacion_id and pro.id=pdh.profesor_id and pro.persona_id=per.id LIMIT 1),0) - tabla1.HORAS_OTRAS_ACTIVIDADES >= 0 then COALESCE((select tdd.horas from sga_profesordistributivohoras pdh, sga_tiempodedicaciondocente tdd, sga_profesor pro where pdh.periodo_id=%s and tdd.id=pdh.dedicacion_id and pro.id=pdh.profesor_id and pro.persona_id=per.id),0) - tabla1.HORAS_OTRAS_ACTIVIDADES else 0 end) as HORAS_OTRAS_ACTIVIDADES, " \
                      "0 as HORAS_CLASE_NIVEL_TECNICO, " \
                      "0 as HORAS_CLASE_TERCER_NIVEL, " \
                      "0 as HORAS_CLASE_CUARTO_NIVEL, " \
                      "COALESCE((select round((rfa.resultado_docencia*100/5),2) from sga_resumenfinalevaluacionacreditacion rfa where rfa.id=(select distinct ra.id  " \
                      "from sga_profesordistributivohoras dh  " \
                      "left join sga_respuestaevaluacionacreditacion rea on dh.profesor_id=rea.profesor_id and rea.proceso_id=%s  " \
                      "left join sga_resumenfinalevaluacionacreditacion ra on ra.distributivo_id=dh.id  " \
                      "left join sga_coordinacion coor on coor.id=dh.coordinacion_id  " \
                      "left join sga_carrera car on car.id=dh.carrera_id  " \
                      "left join sga_profesor pro on pro.id=dh.profesor_id  " \
                      "left join sga_persona per1 on per1.id=pro.persona_id and per1.id=per.id  " \
                      "where dh.periodo_id=%s and per1.real=True and ra.promedio_docencia_hetero notnull order by 1)),0) as CALIFICACION_ACTIVIDADES_DOCENCIA,  " \
                      "COALESCE((select round((rfa.resultado_investigacion*100/5),2) from sga_resumenfinalevaluacionacreditacion rfa where rfa.id=(select distinct ra.id  " \
                      "from sga_profesordistributivohoras dh  " \
                      "left join sga_respuestaevaluacionacreditacion rea on dh.profesor_id=rea.profesor_id and rea.proceso_id=%s  " \
                      "left join sga_resumenfinalevaluacionacreditacion ra on ra.distributivo_id=dh.id  " \
                      "left join sga_coordinacion coor on coor.id=dh.coordinacion_id  " \
                      "left join sga_carrera car on car.id=dh.carrera_id  " \
                      "left join sga_profesor pro on pro.id=dh.profesor_id  " \
                      "left join sga_persona per1 on per1.id=pro.persona_id and per1.id=per.id  " \
                      "where dh.periodo_id=%s and per1.real=True and ra.promedio_docencia_hetero notnull order by 1)),0) as CALIFICACION_ACTIVIDADES_INVESTIGACION,  " \
                      "COALESCE((select round((rfa.resultado_gestion*100/5),2) from sga_resumenfinalevaluacionacreditacion rfa where rfa.id=(select distinct ra.id from sga_profesordistributivohoras dh  " \
                      "left join sga_respuestaevaluacionacreditacion rea on dh.profesor_id=rea.profesor_id and rea.proceso_id=%s  " \
                      "left join sga_resumenfinalevaluacionacreditacion ra on ra.distributivo_id=dh.id  " \
                      "left join sga_coordinacion coor on coor.id=dh.coordinacion_id  " \
                      "left join sga_carrera car on car.id=dh.carrera_id  " \
                      "left join sga_profesor pro on pro.id=dh.profesor_id  " \
                      "left join sga_persona per1 on per1.id=pro.persona_id and per1.id=per.id  " \
                      "where dh.periodo_id=%s and per1.real=True and ra.promedio_docencia_hetero notnull order by 1)),0) as CALIFICACION_ACTIVIDADES_DIRECCION_GESTION_ACADEMICA,per.apellido1||' '||per.apellido2||' '||per.nombres as nompersona,per.id " \
                      "from  (select  tabla.id, '1024' as CODIGO_IES ,  " \
                      "sum(tabla.horas_clases) as HORA_CLASE, " \
                      "sum(tabla.horas_tutorias) as HORA_TUTORIA, " \
                      "sum(tabla.horas_administracion) as HORA_ADMINISTRACION, " \
                      "sum(tabla.horas_investigacion) as HORA_INVESTIGACION, " \
                      "sum(tabla.horas_vinculacion) as HORA_VINCULACION, " \
                      "sum(tabla.horas_clases)+sum(tabla.horas_tutorias)+sum(tabla.horas_administracion)+sum(tabla.horas_investigacion)+sum(tabla.horas_vinculacion) as HORAS_OTRAS_ACTIVIDADES " \
                      "from  (select pe.id,  " \
                      "sum((case cd.tipocriterioactividad when 1 then dd.horas else 0 end)) as horas_administracion, " \
                      "sum((case cd.tipocriterioactividad when 2 then dd.horas else 0 end)) as horas_clases, " \
                      "sum((case cd.tipocriterioactividad when 3 then dd.horas else 0 end)) as horas_tutorias, " \
                      "sum((case cd.tipocriterioactividad when 4 then dd.horas else 0 end)) as horas_vinculacion, " \
                      "sum((case cd.tipocriterioactividad when 5 then dd.horas else 0 end)) as horas_investigacion " \
                      "from sga_profesordistributivohoras pdh, sga_detalledistributivo dd, sga_criteriodocenciaperiodo cdp,  " \
                      "sga_criteriodocencia cd, sga_profesor p, sga_persona pe " \
                      "where pdh.periodo_id in (%s) and dd.distributivo_id=pdh.id and cdp.id=dd.criteriodocenciaperiodo_id " \
                      "and cdp.criterio_id=cd.id and cdp.periodo_id=pdh.periodo_id and p.id=pdh.profesor_id and pe.id=p.persona_id " \
                      "GROUP BY pe.id " \
                      "union " \
                      "select pe.id,  " \
                      "sum((case cd.tipocriterioactividad when 1 then dd.horas else 0 end)) as horas_administracion, " \
                      "sum((case cd.tipocriterioactividad when 2 then dd.horas else 0 end)) as horas_clases, " \
                      "sum((case cd.tipocriterioactividad when 3 then dd.horas else 0 end)) as horas_tutorias, " \
                      "sum((case cd.tipocriterioactividad when 4 then dd.horas else 0 end)) as horas_vinculacion, " \
                      "sum((case cd.tipocriterioactividad when 5 then dd.horas else 0 end)) as horas_investigacion " \
                      "from sga_profesordistributivohoras pdh, sga_detalledistributivo dd, sga_criterioinvestigacionperiodo cdp,  " \
                      "sga_criterioinvestigacion cd, sga_profesor p, sga_persona pe " \
                      "where pdh.periodo_id in (%s) and dd.distributivo_id=pdh.id and cdp.id=dd.criterioinvestigacionperiodo_id " \
                      "and cdp.criterio_id=cd.id and cdp.periodo_id=pdh.periodo_id and p.id=pdh.profesor_id and pe.id=p.persona_id " \
                      "GROUP BY pe.id " \
                      "union " \
                      "select pe.id,  " \
                      "sum((case cd.tipocriterioactividad when 1 then dd.horas else 0 end)) as horas_administracion, " \
                      "sum((case cd.tipocriterioactividad when 2 then dd.horas else 0 end)) as horas_clases, " \
                      "sum((case cd.tipocriterioactividad when 3 then dd.horas else 0 end)) as horas_tutorias, " \
                      "sum((case cd.tipocriterioactividad when 4 then dd.horas else 0 end)) as horas_vinculacion, " \
                      "sum((case cd.tipocriterioactividad when 5 then dd.horas else 0 end)) as horas_investigacion " \
                      "from sga_profesordistributivohoras pdh, sga_detalledistributivo dd, sga_criteriogestionperiodo cdp,  " \
                      "sga_criteriogestion cd, sga_profesor p, sga_persona pe " \
                      "where pdh.periodo_id in (%s) and dd.distributivo_id=pdh.id and cdp.id=dd.criteriogestionperiodo_id " \
                      "and cdp.criterio_id=cd.id and cdp.periodo_id=pdh.periodo_id and p.id=pdh.profesor_id and pe.id=p.persona_id " \
                      "GROUP BY pe.id) as tabla GROUP BY tabla.id) as tabla1 " \
                      "inner join sga_persona per on tabla1.id=per.id  " % (
                          periodo.id, periodo.id, periodo.id, periodo.id, periodo.id, periodo.id,
                          periodo.id, periodo.id, periodo.id, periodo.id, periodo.id, periodo.id,
                          periodo.id, periodo.id, periodo.id, periodo.id)
                cursor.execute(sql)
                results = cursor.fetchall()
                row_num = 1
                for r in results:
                    nivelcategoria = ''
                    categoria = ''
                    dedicacion = ''
                    distri = ProfesorDistributivoHoras.objects.get(profesor__persona__id=int(r[21]),
                                                                   periodo__id=periodo.id, status=True)
                    if distri.nivelcategoria:
                        nivelcategoria = distri.nivelcategoria.nombre
                    if distri.categoria:
                        categoria = distri.categoria.nombre
                    if distri.dedicacion:
                        dedicacion = distri.dedicacion.nombre
                    ws.write(row_num, 0, r[0] if r[0] else '', font_style2)
                    ws.write(row_num, 1, r[1] if r[1] else '', font_style2)
                    ws.write(row_num, 2, str(r[2]) if r[2] else '', font_style2)
                    ws.write(row_num, 3, r[3] if r[3] else '', font_style2)
                    ws.write(row_num, 4, r[4], date_format)
                    ws.write(row_num, 5, r[5], date_format)
                    ws.write(row_num, 6, r[6] if r[6] else '', font_style2)
                    ws.write(row_num, 7, r[7], date_format)
                    ws.write(row_num, 8, r[8], font_style2)
                    ws.write(row_num, 9, r[9], font_style2)
                    ws.write(row_num, 10, r[10], font_style2)
                    ws.write(row_num, 11, r[11], font_style2)
                    ws.write(row_num, 12, r[12], font_style2)
                    ws.write(row_num, 13, r[13], font_style2)
                    ws.write(row_num, 14, r[14], font_style2)
                    ws.write(row_num, 15, r[15], font_style2)
                    ws.write(row_num, 16, r[16], font_style2)
                    ws.write(row_num, 17, r[17], font_style2)
                    ws.write(row_num, 18, r[18], font_style2)
                    ws.write(row_num, 19, r[19], font_style2)
                    ws.write(row_num, 20, r[20], font_style2)
                    ws.write(row_num, 21, nivelcategoria, font_style2)
                    ws.write(row_num, 22, categoria, font_style2)
                    ws.write(row_num, 23, dedicacion, font_style2)
                    row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Generación de reporte de excel listo', titulo='Excel - Distribucion Horas Periodo Academico',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Excel terminado",
                "body": 'Excel - Distribucion Horas Periodo Academico',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su reporte ha sido generado con exito'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)


class reporte_corte_inventario_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'inventario')
        valido = True

        idnotificacion = notif
        usernotify = User.objects.get(pk=request.user.pk)
        personanotifica = Persona.objects.get(usuario=usernotify)

        nombre_archivo = 'reporte_corte_inventario' + str(random.randint(1, 10000)) + '.pdf'
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        try:
            pers = Persona.objects.get(usuario_id=request.user.pk)
            formato_date = "%Y-%m-%d"
            formato, finicio, ffin= int(request.GET.get('formato', 1)), \
                                    request.GET.get('finicio', ''), \
                                    request.GET.get('ffin', '')
            if finicio:
                finicio = datetime.strptime(finicio, formato_date)
            if ffin:
                ffin = datetime.strptime(ffin, formato_date)

            tituloreporte = "Reporte Excel Corte de Inventario" if formato == 2 else "Reporte PDF Corte de Inventario"

            cursor = connections['default'].cursor()
            sql =f""" 
                    SET statement_timeout = 0;
                    SELECT
                         sagest_producto."id" AS sagest_producto_id,
                         sagest_producto."usuario_creacion_id" AS sagest_producto_usuario_creacion_id,
                         sagest_producto."fecha_creacion" AS sagest_producto_fecha_creacion,
                         sagest_producto."usuario_modificacion_id" AS sagest_producto_usuario_modificacion_id,
                         sagest_producto."fecha_modificacion" AS sagest_producto_fecha_modificacion,
                         sagest_producto."codigo" AS sagest_producto_codigo,
                         sagest_producto."descripcion" AS sagest_producto_descripcion,
                         sagest_producto."unidadmedida_id" AS sagest_producto_unidadmedida_id,
                         sagest_producto."tipoproducto_id" AS sagest_producto_tipoproducto_id,
                         sagest_producto."cuenta_id" AS sagest_producto_cuenta_id,
                         sagest_producto."alias" AS sagest_producto_alias,
                         sagest_producto."codigobarra" AS sagest_producto_codigobarra,
                         sagest_producto."minimo" AS sagest_producto_minimo,
                         sagest_producto."maximo" AS sagest_producto_maximo,
                         sagest_producto."status" AS sagest_producto_status,
                         sagest_cuentacontable."id" AS sagest_cuentacontable_id,
                         sagest_cuentacontable."cuenta" AS sagest_cuentacontable_cuenta,
                         sagest_cuentacontable."descripcion" AS sagest_cuentacontable_descripcion,
                         coalesce(saldo(sagest_producto."id"::int4, '{finicio}'::date), 0) AS saldoinicio,
                         coalesce(saldofinal(sagest_producto."id"::int4, '{ffin}'::date), 0) AS saldofin,
                         coalesce(entrada(sagest_producto."id"::int4, '{finicio}'::date, '{ffin}'::date), 0) AS entrada,
                         coalesce(salida(sagest_producto."id"::int4, '{finicio}'::date, '{ffin}'::date), 0) AS salida,
                         coalesce(saldo_monetario(sagest_producto."id"::INT4, '{finicio}'::date), 0) AS saldoiniciomoneda,
                         round(coalesce(saldo_monetario(sagest_producto."id"::int4, '{finicio}'::date), 0),6) AS saldoiniciomoneda_2,
                         round(coalesce(saldofinal_monetario(sagest_producto."id"::int4, '{ffin}'::date), 0),6) AS saldofinmoneda,
                         round(coalesce(saldofinal_monetario(sagest_producto."id"::int4, '{ffin}'::date), 0), 6) AS saldofinmoneda2,
                         round(coalesce(entrada_monetaria(sagest_producto."id"::int4, '{finicio}'::date, '{ffin}'::date), 0),6) AS entradamoneda,
                         round(coalesce(entrada_monetaria(sagest_producto."id"::int4, '{finicio}'::date, '{ffin}'::date), 0),6) AS entradamoneda_2,
                         round(coalesce(salida_monetaria(sagest_producto."id"::int4, '{finicio}'::date, '{ffin}'::date), 0),6) AS salidamoneda,
                         round(coalesce(salida_monetaria(sagest_producto."id"::int4, '{finicio}'::date, '{ffin}'::date), 0),6) AS salidamoneda_2
                    FROM
                         "public"."sagest_cuentacontable" sagest_cuentacontable RIGHT OUTER JOIN "public"."sagest_producto" sagest_producto ON sagest_cuentacontable."id" = sagest_producto."cuenta_id"
                    WHERE coalesce(saldo(sagest_producto."id"::INT4, '{finicio}'::date), 0) > 0 
                    OR  coalesce(saldofinal(sagest_producto."id"::INT4,  '{ffin}'::date), 0) > 0 
                    OR  coalesce(entrada(sagest_producto."id"::int4, '{finicio}'::DATE, '{ffin}'::date), 0) > 0 
                    OR   coalesce(salida(sagest_producto."id"::int4, '{finicio}'::date, '{ffin}'::date), 0) > 0
                    ORDER BY
                         sagest_cuentacontable."id" ASC,
                         sagest_producto."codigo" ASC,
                         sagest_producto."descripcion" ASC
                  """
            cursor.execute(sql)
            results = cursor.fetchall()
            # Creamos un diccionario para almacenar los datos agrupados por el índice 17
            datos_agrupados = {}
            # Iteramos sobre cada tupla de datos
            for dato in results:
                # Obtenemos el valor en el índice 17
                key = dato[16]

                # Si aún no hay una lista para ese valor, la creamos
                if key not in datos_agrupados:
                    datos_agrupados[key] = { 'numerocuenta':  dato[16],
                                             'cuenta':  dato[17],
                                             'datos':[]}

                # Creamos un diccionario con los datos y lo agregamos a la lista correspondiente
                datos_diccionario = {
                    'codigo':f'{dato[16]}.{dato[5]}',
                    'producto':dato[6],
                    'saldoinicio':dato[18].quantize(Decimal('0.00'), rounding=ROUND_DOWN),
                    'saldofin':dato[19].quantize(Decimal('0.00'), rounding=ROUND_DOWN),
                    'entrada':dato[20].quantize(Decimal('0.00'), rounding=ROUND_DOWN),
                    'salida':dato[21].quantize(Decimal('0.00'), rounding=ROUND_DOWN),
                    'saldoiniciomoneda': dato[23].quantize(Decimal('0.0000'), rounding=ROUND_DOWN),
                    'saldofinmoneda':dato[24].quantize(Decimal('0.0000'), rounding=ROUND_DOWN),
                    'entradamoneda':dato[26].quantize(Decimal('0.0000'), rounding=ROUND_DOWN),
                    'salidamoneda':dato[28].quantize(Decimal('0.0000'), rounding=ROUND_DOWN),
                }
                datos_agrupados[key]['datos'].append(datos_diccionario)
            for clave, valor in datos_agrupados.items():
                valor['suma_saldoiniciomoneda'] = sum(item['saldoiniciomoneda'] for item in valor['datos'])
                valor['suma_entradamoneda'] = sum(item['entradamoneda'] for item in valor['datos'])
                valor['suma_salidamoneda'] = sum(item['salidamoneda'] for item in valor['datos'])
                valor['suma_saldofinmoneda'] = sum(item['saldofinmoneda'] for item in valor['datos'])
            total_saldoiniciomoneda = sum(item['suma_saldoiniciomoneda'] for item in datos_agrupados.values())
            total_entradamoneda = sum(item['suma_entradamoneda'] for item in datos_agrupados.values())
            total_salidamoneda = sum(item['suma_salidamoneda'] for item in datos_agrupados.values())
            total_saldofinmoneda = sum(item['suma_saldofinmoneda'] for item in datos_agrupados.values())
            total_saldoletras = num2words(total_saldofinmoneda, lang='es')
            if formato == 1:
                data['total_saldoiniciomoneda'] = total_saldoiniciomoneda
                data['total_entradamoneda'] = total_entradamoneda
                data['total_salidamoneda'] = total_salidamoneda
                data['total_saldofinmoneda'] = total_saldofinmoneda
                data['total'] = len(datos_agrupados)
                data['totalsaldoletras'] = total_saldoletras.upper()
                data['username'] = request.user.username
                data['desde'] = finicio
                data['hasta'] = ffin
                data['fechahoy'] = datetime.now().date()
                data['listado'] = datos_agrupados
                valido = conviert_html_to_pdfsave_generic_lotes(
                    request,
                    'adm_inventarios/corte_inventario.html',
                    {
                        'pagesize': 'A4 landscape',
                        'data': data,
                    },
                    directory, nombre_archivo
                )

                # NOTIFICACIÓN
                titulo = 'Reporte de corte de inventario generada exitosamente'
                cuerpo = 'Su reporte de corte de inventario se genero correctamente y puede ser descargada por la url proporcionada.'
                if not valido:
                    titulo = 'Algo salio mal, el reporte de corte de inventario no se genero '
                    cuerpo = 'Su reporte de corte de inventario no se genero correctamente.'

                url = "{}reportes/inventario/{}".format(MEDIA_URL, nombre_archivo)
                if notif > 0:
                    noti = Notificacion.objects.get(pk=notif)
                    noti.titulo = titulo
                    noti.en_proceso = False
                    noti.cuerpo = cuerpo
                    noti.url = url
                    noti.save()
                else:
                    noti = Notificacion(cuerpo=cuerpo,
                                        titulo=titulo,
                                        destinatario=pers, url=url,
                                        prioridad=1, app_label='SGA-SAGEST', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                        tipo=2, en_proceso=False)
                    noti.save(request)

                send_user_notification(user=request.user, payload={
                    "head": titulo,
                    "body": cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)

            elif formato == 2:
                resultados = [{'etiqueta': 'TOTAL SAL. ANT', 'valor': total_saldoiniciomoneda},
                              {'etiqueta': 'TOTAL ING:', 'valor': total_entradamoneda},
                              {'etiqueta': 'TOTAL EGRE: ', 'valor': total_salidamoneda},
                              {'etiqueta': 'TOTAL SALDOS:', 'valor': total_saldofinmoneda},
                              {'etiqueta': 'DIFERENCIA POR REDONDEO:', 'valor': 0.0000}]

                directorio = os.path.join(os.path.join(MEDIA_ROOT, 'reportes', 'inventario'))
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                nombrearchivo = "REPORTE_CORTE_INVENTARIO_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"
                urlarchivo = MEDIA_URL + "reportes/inventario/" + nombrearchivo

                # Crea un nuevo archivo de excel y le agrega una hoja
                workbook = xlsxwriter.Workbook(directorio + '/' + nombrearchivo)
                hojadestino = workbook.add_worksheet("Reporte")

                fcabeceracolumna = workbook.add_format(FORMATOS_CELDAS_EXCEL["cabeceracolumna"])
                fceldageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneral"])
                fceldanumerodecimal = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanumerodecimal"])
                fceldanumerodecimal4dec = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanumerodecimal4dec"])
                fceldageneralneg = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneralneg"])
                ftitulo2 = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo2"])
                ftitulo3 = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo3"])
                fceldanumerodecimal4decpie = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanumerodecimal4decpie"])
                fceldanegritaizq = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritaizq"])
                fceldanegritageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdanegritageneral"])
                ftextonegrita = workbook.add_format(FORMATOS_CELDAS_EXCEL["textonegrita"])

                hojadestino.merge_range(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', ftitulo2)
                hojadestino.merge_range(1, 0, 1, 10, 'DIRECCIÓN ADMINISTRATIVA', ftitulo2)
                hojadestino.merge_range(2, 0, 2, 10, 'SECCIÓN PROVEDURÍA', ftitulo2)
                hojadestino.merge_range(3, 0, 3, 10, 'CORTE DE INVENTARIO', ftitulo2)
                hojadestino.merge_range(4, 0, 4, 10, 'DESDE: ' + str(finicio)[0:10] + ' HASTA: ' + str(ffin)[0:10], ftitulo2)

                fila = 6

                columnas = [
                    (u"Sal.", 15),
                    (u"Ingres.", 15),
                    (u"Egres.", 15),
                    (u"Saldo", 15),
                    (u"Sal.Ant.", 15),
                    (u"Ingreso", 15),
                    (u"Egreso", 15),
                    (u"Saldo", 15)
                ]

                hojadestino.merge_range(fila, 0, fila + 1, 0, 'Cuenta', fcabeceracolumna)
                hojadestino.merge_range(fila, 1, fila + 1, 1, 'Producto', fcabeceracolumna)
                hojadestino.merge_range(fila, 2, fila, 5, 'Unidades', fcabeceracolumna)
                hojadestino.merge_range(fila, 6, fila, 9, 'Moneda', fcabeceracolumna)
                hojadestino.set_column(0, 0, 14)
                hojadestino.set_column(1, 1, 47)

                fila += 1
                col_num_cab = 0
                for col_num in range(2, 10):
                    hojadestino.write(fila, col_num, columnas[col_num_cab][0], fcabeceracolumna)
                    hojadestino.set_column(col_num, col_num, columnas[col_num_cab][1])
                    col_num_cab += 1

                fila += 1
                # Llenar el DataFrame con los datos del queryset
                for clave, da in datos_agrupados.items():
                    hojadestino.write(fila, 0, da['numerocuenta'], fceldageneralneg)
                    hojadestino.merge_range(fila, 1, fila, 9, da['cuenta'], fceldageneralneg)
                    fila += 1

                    for dato in da['datos']:
                        hojadestino.write(fila, 0, dato['codigo'], fceldageneral)
                        hojadestino.write(fila, 1, dato['producto'], fceldageneral)
                        hojadestino.write(fila, 2, dato['saldoinicio'], fceldanumerodecimal)
                        hojadestino.write(fila, 3, dato['entrada'], fceldanumerodecimal)
                        hojadestino.write(fila, 4, dato['salida'], fceldanumerodecimal)
                        hojadestino.write(fila, 5, dato['saldofin'], fceldanumerodecimal)
                        hojadestino.write(fila, 6, dato['saldoiniciomoneda'], fceldanumerodecimal4dec)
                        hojadestino.write(fila, 7, dato['entradamoneda'], fceldanumerodecimal4dec)
                        hojadestino.write(fila, 8, dato['salidamoneda'], fceldanumerodecimal4dec)
                        hojadestino.write(fila, 9, dato['saldofinmoneda'], fceldanumerodecimal4dec)
                        fila += 1

                    hojadestino.merge_range(fila, 0, fila, 5, "SUBTOTAL", fceldageneralneg)
                    hojadestino.write(fila, 6, da['suma_saldoiniciomoneda'], fceldanumerodecimal4decpie)
                    hojadestino.write(fila, 7, da['suma_entradamoneda'], fceldanumerodecimal4decpie)
                    hojadestino.write(fila, 8, da['suma_salidamoneda'], fceldanumerodecimal4decpie)
                    hojadestino.write(fila, 9, da['suma_saldofinmoneda'], fceldanumerodecimal4decpie)
                    fila += 1

                for resultado in resultados:
                    hojadestino.merge_range(fila, 5, fila, 7, resultado['etiqueta'], fceldageneralneg)
                    hojadestino.merge_range(fila, 8, fila, 9, resultado['valor'], fceldanumerodecimal4dec)
                    fila += 1

                hojadestino.merge_range(fila, 0, fila, 9, "SON: " + total_saldoletras.upper(), ftextonegrita)

                workbook.close()

                # Notificar al usuario
                if idnotificacion > 0:
                    notificacion = Notificacion.objects.get(pk=idnotificacion)
                    notificacion.en_proceso = False
                    notificacion.cuerpo = 'Reporte Excel finalizado'
                    notificacion.url = urlarchivo
                    notificacion.save()
                else:
                    notificacion = Notificacion(
                        cuerpo='Reporte Excel finalizado',
                        titulo=tituloreporte,
                        destinatario=personanotifica,
                        url=urlarchivo,
                        prioridad=1,
                        app_label='SGA-SAGEST',
                        fecha_hora_visible=datetime.now() + timedelta(days=1),
                        tipo=2,
                        en_proceso=False
                    )
                    notificacion.save(request)

                send_user_notification(user=usernotify, payload={
                    "head": "Reporte Excel finalizado",
                    "body": tituloreporte,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": urlarchivo,
                    "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                    "mensaje": "El <b>{}</b> ha sido generado con éxito".format(tituloreporte)
                }, ttl=500)

                #####################################


                # nombre_archivo = 'reporte_corte_inventario' + random.randint(1, 10000).__str__() + '.xls'
                # directory = os.path.join(MEDIA_ROOT, 'reportes', 'inventario', nombre_archivo)
                # __author__ = 'Unemi'
                # # Aplicar el objeto de Formato personalizado al estilo
                # wb = openxl.Workbook()
                # ws = wb.active
                # style_title = openxlFont(name='Arial', size=16, bold=True)
                # style_cab = openxlFont(name='Arial', size=10, bold=True)
                # alinear = alin(horizontal="center", vertical="center")
                # response = HttpResponse(content_type="application/ms-excel")
                # response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + '-' + random.randint(1,
                #                                                                                                   10000).__str__() + '.xlsx'
                # ws.merge_cells('A1:J1')
                # ws.merge_cells('A2:J2')
                # ws.merge_cells('A3:J3')
                # ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO'
                # ws['A2'] = 'DEPARTAMENTO ADMINISTRATIVO'
                # ws['A3'] = 'SECCIÓN PROVEDURÍA | CORTE DE INVENTARIO'
                # celda1 = ws['A1']
                # celda1.font = style_title
                # celda1.alignment = alinear
                # celda2 = ws['A2']
                # celda2.font = style_title
                # celda2.alignment = alinear
                # celda3 = ws['A3']
                # celda3.font = style_title
                # celda3.alignment = alinear
                #
                # ws.merge_cells('C4:F4')
                # ws.merge_cells('G4:J4')
                # columns1 = ['Cuenta',
                #            'Producto',
                #            'Unidades',
                #            'Moneda'
                #            ]
                # columns2 = ['Sal.',
                #            'Ingres.',
                #            'Egres.',
                #            'Saldo',
                #            'Sal.Ant.',
                #            'Ingreso',
                #            'Egreso',
                #            'Saldo',
                #            ]
                # row_num = 4
                # for col_num in range(0, len(columns1)):
                #     if col_num == 4:
                #         celda = ws.cell(row=row_num, column=(7), value=columns1[col_num])
                #     else:
                #         celda = ws.cell(row=row_num, column=(col_num + 1), value=columns1[col_num])
                #     celda.font = style_cab
                # row_num = 5
                # for col_num in range(0, len(columns2)):
                #     celda = ws.cell(row=row_num, column=(col_num + 1), value=columns2[col_num])
                #     celda.font = style_cab
                # row_num = 6
                # # Llenar el DataFrame con los datos del queryset
                # for clave, da in datos_agrupados.items():
                #     ws.merge_cells(f'B{row_num}:J{row_num}')
                #     ws.cell(row=row_num, column=1, value=da['numerocuenta'])
                #     ws.cell(row=row_num, column=2, value=da['cuenta'])
                #     row_num += 1
                #     for dato in da['datos']:
                #         ws.cell(row=row_num, column=1, value=dato['codigo'])
                #         ws.cell(row=row_num, column=2, value=dato['producto'])
                #         ws.cell(row=row_num, column=3, value=dato['saldoinicio'])
                #         ws.cell(row=row_num, column=4, value=dato['entrada'])
                #         ws.cell(row=row_num, column=5, value=dato['salida'])
                #         ws.cell(row=row_num, column=6, value=dato['saldofin'])
                #         ws.cell(row=row_num, column=7, value=dato['saldoiniciomoneda'])
                #         ws.cell(row=row_num, column=8, value=dato['entradamoneda'])
                #         ws.cell(row=row_num, column=9, value=dato['salidamoneda'])
                #         ws.cell(row=row_num, column=10, value=dato['saldofinmoneda'])
                #         row_num += 1
                #     ws.merge_cells(f'A{row_num}:F{row_num}')
                #     ws.cell(row=row_num, column=1, value='SUBTOTAL')
                #     ws.cell(row=row_num, column=7, value=da['suma_saldoiniciomoneda'])
                #     ws.cell(row=row_num, column=8, value=da['suma_entradamoneda'])
                #     ws.cell(row=row_num, column=9, value=da['suma_salidamoneda'])
                #     ws.cell(row=row_num, column=10, value=da['suma_saldofinmoneda'])
                #     row_num += 1
                #
                # for key, r in resultados.items():
                #     ws.merge_cells(f'A{row_num}:E{row_num}')
                #     ws.merge_cells(f'F{row_num}:H{row_num}')
                #     ws.merge_cells(f'I{row_num}:J{row_num}')
                #     ws.cell(row=row_num, column=6, value=key)
                #     ws.cell(row=row_num, column=9, value=r)
                #     row_num += 1
                # wb.save(directory)


        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

            if idnotificacion > 0:
                notificacion = Notificacion.objects.get(pk=idnotificacion)
                notificacion.en_proceso = False
                notificacion.error = True
                notificacion.titulo = "Error al generar {}".format(tituloreporte)
                notificacion.cuerpo = textoerror
                notificacion.save()
            else:
                notificacion = Notificacion(
                    cuerpo="Error al generar Reporte Excel" if formato == 2 else "Error al generar Reporte PDF",
                    titulo="Error al generar {}".format(tituloreporte),
                    destinatario=personanotifica,
                    prioridad=1,
                    app_label='SGA-SAGEST',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=False,
                    error=True)
                notificacion.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Error al generar Reporte Excel" if formato == 2 else "Error al generar Reporte PDF",
                "body": "Error al generar {}".format(tituloreporte),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, personanotifica),
                "mensaje": "Error al generar <b>{}</b>. Error: {}".format(tituloreporte, textoerror),
                "error": True
            }, ttl=500)


class ejecutar_procesos_background(threading.Thread):

    def __init__(self, request, data, notiid):
        self.request = request
        self.data = data
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        request, data, notiid = self.request, self.data, self.notiid
        try:
            proceso = PythonProcess.objects.get(id=data['id'])
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            d = locals()
            code = proceso.code
            if proceso.anexo:
                d['path_anexo'] = proceso.anexo.path
            globalvar = globals()
            globalvar['funcionlocal'] = locals()
            globalvar['request'] = request
            tablas_criticas = ['AuditoriaNota', 'sga_auditorianotas', 'sga_materiaasignada', 'MateriaAsignada', 'sga_evaluaciongenerica', 'EvaluacionGenerica',
                               'sga_asistencialeccion', 'AsistenciaLeccion', 'sga_asistencialeccion_historico', 'AsistenciaLeccion_Historico',
                               'sga_asistenciamoodle', 'AsistenciaMoodle']
            if proceso.tipo == 1:
                if not request.user.pk == 29882:
                    for tabla in tablas_criticas:
                        if tabla in proceso.code:
                            raise NameError('Lo sentimos, este codigo no se puede ejecutar')
                exec(proceso.code, globalvar, d)
            else:
                if not proceso.archivo:
                    raise NameError('Archivo no encontrado')
                with open(proceso.archivo.path, 'r') as archivo:
                    contenido = archivo.read()
                if not request.user.pk == 29882:
                    for tabla in tablas_criticas:
                        if tabla in contenido:
                            raise NameError('Lo sentimos, este codigo no se puede ejecutar')
                exec(contenido, globalvar, d)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                titulo = 'Proceso terminado sin errores'
                noti.cuerpo = 'El proceso de python terminó satisfactoriamente'
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='El proceso de python terminó satisfactoriamente', titulo='Proceso terminado sin errores',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=usernotify, payload={
                "head": "Proceso terminado",
                "body": 'El proceso de python terminó satisfactoriamente',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "btn_notificaciones": json.dumps(traerNotificaciones(request, data, pers)),
                "mensaje": 'El proceso de python terminó satisfactoriamente'
            }, ttl=500)
        except Exception as ex:
            proceso = PythonProcess.objects.get(id=data['id'])
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            textoerror = f'Proceso: {proceso.nombre} {ex} Linea:{sys.exc_info()[-1].tb_lineno}'
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.titulo = f'Proceso no terminado'
                noti.en_proceso = False
                noti.cuerpo = f'Error al ejecutar el proceso: {proceso.nombre} --- {textoerror}'
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error al ejecutar el proceso {}'.format(textoerror), titulo='Proceso no terminado',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            send_user_notification(user=usernotify, payload={
                "head": f"Proceso {proceso.nombre} terminado",
                "body": 'El proceso de python terminó satisfactoriamente',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "btn_notificaciones": json.dumps(traerNotificaciones(request, data, pers)),
                "mensaje": f'El proceso de python {proceso.nombre} terminó satisfactoriamente'
            }, ttl=500)


class firmar_contratos_posgrado_vicerrectorado_background(threading.Thread):

    def __init__(self, request, notiid, firma, passfirma, contratosselect,personasesion):
        self.request = request
        self.firma = firma
        self.passfirma = passfirma
        self.contratosselect = contratosselect
        self.personasesion = personasesion
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        try:
            __author__ = 'Unemi'
            request, notiid, firma, passfirma, contratosselect,personasesion = self.request, self.notiid, self.firma,self.passfirma,self.contratosselect,self.personasesion

            bandera = False
            p12 = None
            listainscripcion = []
            nombresmae = ''
            conterrornombre = 0
            conteoerror = 0
            if not FirmaPersona.objects.filter(status=True, persona=personasesion, tipofirma=3).exists():
                objetofirma = FirmaPersona(
                    persona=personasesion,
                    tipofirma=3, firma=firma
                )
                objetofirma.save(request)
            palabras = 'Dr. Edwuin Jesús Carrasquero Rodríguez'
            for con in contratosselect:
                objetofirma = FirmaPersona.objects.filter(status=True, persona=personasesion, tipofirma=3).last()
                firmaarchivo = objetofirma.firma
                bytes_certificado = firmaarchivo.read()
                extension_certificado = os.path.splitext(firmaarchivo.name)[1][1:]

                contrato = Contrato.objects.get(pk=con)
                if not contrato.respaldoarchivocontrato:
                    contrato.respaldoarchivocontrato = contrato.archivocontrato
                contrato.contratobackground=True
                contrato.save(request)
                documento_a_firmar = contrato.respaldoarchivocontrato
                # obtener la posicion xy de la firma del doctor en el pdf
                print(u"%s" % contrato)
                y, numpaginafirma = obtener_posicion_y(documento_a_firmar.url, palabras)
                # FIN obtener la posicion y
                if y:
                    datau = JavaFirmaEc(
                        archivo_a_firmar=documento_a_firmar, archivo_certificado=bytes_certificado,
                        extension_certificado=extension_certificado,
                        password_certificado=passfirma,
                        page=numpaginafirma, reason='Contrato legalizado', lx=115, ly=y).sign_and_get_content_bytes()
                    if datau:
                        generar_archivo_firmado = io.BytesIO()
                        generar_archivo_firmado.write(datau)
                        generar_archivo_firmado.seek(0)
                        extension = documento_a_firmar.name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        nombrefile_ = remover_caracteres_tildes_unicode(
                            remover_caracteres_especiales_unicode(documento_a_firmar.name)).replace('-', '_').replace(
                            '.pdf', '')
                        _name = 'cp_contratopago_' + str(contrato.inscripcion.id)
                        file_obj = DjangoFile(generar_archivo_firmado,
                                              name=f"{remover_caracteres_especiales_unicode(_name)}_firmado.pdf")
                        contrato.archivocontrato = file_obj
                        contrato.contratolegalizado = True
                        contrato.save(request)
                        detalleevidencia = DetalleAprobacionContrato(contrato_id=contrato.id)
                        detalleevidencia.save(request)
                        detalleevidencia.observacion = 'Contrato legalizado'
                        detalleevidencia.archivocontrato = file_obj
                        detalleevidencia.persona = personasesion
                        detalleevidencia.estado_aprobacion = 2
                        detalleevidencia.fecha_aprobacion = datetime.now()
                        detalleevidencia.save(request)
                        log(u'Masivo Firmó Documento: {}'.format(nombrefile_), request, "add")
                        listainscripcion.append(contrato.inscripcion.id)
                        nombresmae += '%s, ' % contrato.inscripcion.inscripcionaspirante

                        integrante = InscripcionCohorte.objects.get(status=True, pk=contrato.inscripcion.id)

                        asunto = u"CONTRATO APROBADO Y LEGALIZADO"
                        observacion = f'Se le comunica que el contrato de {integrante.formapagopac.descripcion} del admitido {integrante.inscripcionaspirante.persona} con cédula {integrante.inscripcionaspirante.persona.cedula} ha sido aprobado y legalizado. Por favor, dar seguimiento en la siguiente fase.'
                        para = integrante.asesor.persona
                        perfiu = integrante.asesor.perfil_administrativo()

                        notificacion3(asunto, observacion, para, None,
                                      '/comercial?s=' + integrante.inscripcionaspirante.persona.cedula,
                                      integrante.pk, 1,
                                      'sga', InscripcionCohorte, perfiu, request)

                        finan = Persona.objects.get(status=True, pk=24145)
                        para2 = finan
                        perfiu2 = finan.perfilusuario_administrativo()
                        url = ''
                        if integrante.formapagopac.id == 2:
                            url = '/comercial?s=' + integrante.inscripcionaspirante.persona.cedula
                        else:
                            url = '/comercial?action=prospectoscontado&s=' + integrante.inscripcionaspirante.persona.cedula

                        notificacion3(asunto, observacion, para2, None,
                                      url,
                                      integrante.pk, 1,
                                      'sga', InscripcionCohorte, perfiu2, request)
                    else:
                        contrato.contratobackground = False
                        conteoerror += 1
                        if contrato.contratolegalizado:
                            contrato.contratolegalizado = False
                        contrato.estado = 3
                        contrato.save(request)
                        detalleevidencia = DetalleAprobacionContrato(contrato_id=contrato.id)
                        detalleevidencia.save(request)
                        detalleevidencia.observacion = 'Documento con inconsistencia en la firma del Estudiante'
                        detalleevidencia.persona = personasesion
                        detalleevidencia.estado_aprobacion = 3
                        detalleevidencia.fecha_aprobacion = datetime.now()
                        detalleevidencia.save(request)

                        cuerpo = ('Contrato de pago con inconsistencia en la Firma del Estudiante: %s - %s' % (
                            contrato.inscripcion.inscripcionaspirante.persona.cedula,
                            contrato.inscripcion.inscripcionaspirante))
                        personanotificar = Persona.objects.get(pk=24145)
                        notificacion('Revisión de documento de Contrato de pago',
                                     cuerpo, personanotificar, None,
                                     '/comercial?action=evidenciacontrato&idcohorte=%s&aspirante=%s' % (
                                         encrypt(contrato.inscripcion.cohortes.id), encrypt(contrato.inscripcion.id)),
                                     contrato.pk, 1, 'sga', Contrato, request)

                        cuerpo = (
                                'Documento con inconsistencia en la firma del ESTUDIANTE %s. Enviado a comercialización para su revisión.' % contrato.inscripcion.inscripcionaspirante)
                        notificacion('Firma electrónica SGA',
                                     cuerpo, personasesion, None, '',
                                     contrato.pk, 1, 'sga', Contrato, request)
                else:
                    conteoerror += 1
                    conterrornombre += 1
                    if contrato.contratolegalizado:
                        contrato.contratolegalizado = False
                    contrato.estado = 3
                    contrato.contratobackground = False
                    contrato.save(request)
                    detalleevidencia = DetalleAprobacionContrato(contrato_id=contrato.id)
                    detalleevidencia.save(request)
                    detalleevidencia.observacion = 'El nombre del dr. en la firma del contrato no es el correcto.'
                    detalleevidencia.persona = personasesion
                    detalleevidencia.estado_aprobacion = 3
                    detalleevidencia.fecha_aprobacion = datetime.now()
                    detalleevidencia.save(request)
                time.sleep(2)
            if listainscripcion:
                cuerpo = ('Documentos firmados con éxito: %s' % nombresmae)
                notificacion('Firma electrónica SGA', cuerpo, personasesion, None,
                             '/firmardocumentosposgrado?action=firmaelectronicacontratos', None, 1, 'sga', Contrato,
                             request)
                if conteoerror > 0:
                    messages.success(request, f'Documentos firmados con éxito. %s' % (
                        'Existieron %s contratos con inconsistencia que no fueron firmados. Enviados a comercialización: %s' % (
                            conteoerror, conterrornombre) if conterrornombre > 0 else ''))
                else:
                    messages.success(request, f'Documentos firmados con éxito')
            else:
                if conteoerror > 0:
                    messages.warning(request, f'%s' % (
                        'Existieron %s contrato(s) con inconsistencia que no fueron firmados. Enviados a comercialización.' % conteoerror if conteoerror > 0 else ''))

            # -----------------------------------------------------------------------------------------
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Documentos firmados con éxito'
                noti.url = "/firmardocumentosposgrado?action=firmaelectronicacontratos&s=&estadof=2"
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Resultados de Encuesta',
                                    destinatario=pers, url="/firmardocumentosposgrado?action=firmaelectronicacontratos&s=&estadof=2",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Documentos firmados",
                "body": 'Contratos firmados',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "/firmardocumentosposgrado?action=firmaelectronicacontratos&s=&estadof=2",
                "mensaje": 'Los documentos de contratos fueron firmados'
            }, ttl=500)
            return True
        except Exception as ex:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error al firmar'
                noti.cuerpo = 'Ha ocurrido un error en al firmar {} - Error en la linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.error = True
                noti.save()
            else:
                noti = Notificacion(titulo='Error al firmar reportes',
                                    cuerpo='Ha ocurrido un error en al firmar contratos {} - Error en la linea {}'.format(
                                        ex, sys.exc_info()[-1].tb_lineno),
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False, error=True)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Error al firmar contratos",
                "body": "Ha ocurrido un error al firmar contratos {} - Error en la linea {}".format(ex,
                                                                                                    sys.exc_info()[
                                                                                                        -1].tb_lineno),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "mensaje": "Ha ocurrido un error en al firmar contratos"
            }, ttl=500)
            pass


class reporte_estudiantes_egresados_faltante_ingles_background(threading.Thread):

    def __init__(self, request, data):
        self.request = request
        self.data = data
        threading.Thread.__init__(self)

    def run(self):
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
            request, data = self.request, self.data
            ePeriodo = data.get('ePeriodo', None)
            eNotificacion = data.get('eNotificacion', None)
            fecha = data.get('fecha', None)
            eCoordinaciones = data.get('eCoordinaciones', Coordinacion.objects.none())
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)

        # nombre_archivo = 'reporte_alumnos_egresado_faltante_ingles' + random.randint(1, 10000).__str__() + '.xlsx'
        # directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)
        #
        # try:
        #
        #     __author__ = 'Unemi'
        #     ahora = datetime.now()
        #     time_codigo = ahora.strftime('%Y%m%d_%H%M%S')
        #     name_file = f'reporte_excel_pedidos_online_{time_codigo}.xlsx'
        #     output = io.BytesIO()
        #     workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
        #     ws = workbook.add_worksheet("Listado Matriculados Nivelación")
        #
        #     fuentecabecera = workbook.add_format({
        #         'align': 'center',
        #         'bg_color': 'silver',
        #         'border': 1,
        #         'bold': 1
        #     })
        #
        #     formatoceldacenter = workbook.add_format({
        #         'border': 1,
        #         'valign': 'vcenter',
        #         'align': 'center'})
        #
        #     formatoceldacenter = workbook.add_format({
        #         'border': 1,
        #         'valign': 'vcenter',
        #         'align': 'center'})
        #
        #     fuenteencabezado = workbook.add_format({
        #         'align': 'center',
        #         'bg_color': '#1C3247',
        #         'font_color': 'white',
        #         'border': 1,
        #         'font_size': 24,
        #         'bold': 1
        #     })
        #
        #     columnas = [
        #                 (u"FACULTAD", 60),
        #                 (u"CARRERA", 60),
        #                 (u"MALLA", 60),
        #                 (u"SECCIÓN", 60),
        #                 (u"NIVEL", 60),
        #                 (u"PARALELO", 60),
        #                 (u"ASIGNATURA", 60),
        #                 (u"TEORICA PRACTICA", 60),
        #                 (u"CUPO", 60),
        #                 (u"MATRICULADOS", 60),
        #                 (u"INSCRITOS", 60),  #
        #                 (u"TOTAL MATRICULADOS", 60),
        #                 (u"DOCENTE", 60),
        #                 (u"CEDULA", 60),
        #                 # (u"USUARIO", 6000),
        #                 (u"AFINIDAD", 60),
        #                 (u"HORAS SEMANALES", 40),
        #                 (u"MALLA (HORAS PRESENCIALES SEMANALES)", 40),
        #                 (u"TIPO", 40),
        #                 (u"CORREO PERSONAL", 40),
        #                 (u"CORREO INSTITUCIONAL", 40),
        #                 (u"TIPO PROFESOR", 40),
        #                 (u"PROFESOR DESDE", 40),
        #                 (u"PROFESOR HASTA", 40),
        #                 (u"DEDICACION", 50),
        #                 (u"CATEGORIA", 40),
        #                 (u"INICIO MATERIA", 40),
        #                 (u"FIN MATERIA", 40),
        #                 (u"FIN ASISTENCIA", 40),
        #                 # (u"ID", 4000),
        #                 (u"TELEFONO", 60),
        #                 (u"MODELO EVALUATIVO", 60),
        #                 (u"IDMATERIA", 25),
        #                 (u"ACEPTACION", 25),
        #                 (u"OBSERVACION ACEPTACION", 20),
        #                 (u"HORARIO FECHA ACEPTACION", 20),
        #                 (u"HORARIO ACEPTACION", 20),
        #                 (u"HORARIO OBSERVACION ACEPTACION", 20),
        #                 (u"IDMOODLE", 25),
        #                 (u"PROVINCIA", 25),
        #                 (u"CIUDAD", 25),
        #                 (u"CANT ACEPTADOS", 60),
        #                 (u"MODALIDAD IMPARTICION", 60),
        #                 (u"ES PRÁCTICA", 40),
        #                 (u"SILABO", 40),
        #                 (u"TIENE GUIA PRÁCTICA", 40),
        #                 (u"CERRADA", 40),
        #             ]
        #     ws.merge_range(0, 0, 0, columnas.__len__() - 1, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
        #     ws.merge_range(1, 0, 1, columnas.__len__() - 1, f'DISTRIBUTIVO DE ASIGNATURAS',fuenteencabezado)
        #     row_num, numcolum = 2, 0
        #     for col_name in columnas:
        #         ws.write(row_num, numcolum, col_name[0], fuentecabecera)
        #         ws.set_column(numcolum, numcolum, col_name[1])
        #         numcolum += 1
        #     row_num += 1
        #
        #     cursor = connections['default'].cursor()
        #     sql = f"""SET statement_timeout='20000 s';
        #                                 SELECT sga_coordinacion.nombre AS Facultad,
        #                                     sga_carrera.nombre AS Carrera, sga_sesion.nombre AS Seccion,
        #                                     sga_nivelmalla.nombre AS Nivel, sga_materia.paralelo AS Paralelo,
        #                                     sga_materia.id AS Idmateria, sga_asignatura.nombre AS Asignatura,
        #                                     sga_persona.apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres AS Docente,
        #                                     sga_profesormateria.hora AS sga_profesormateria_hora, (CASE sga_profesormateria.principal WHEN TRUE THEN 'PRINCIPAL' ELSE 'PRACTICA' END) AS Tipo, sga_persona.cedula, (
        #                                 SELECT u.username
        #                                 FROM auth_user u
        #                                 WHERE u.id=sga_persona.usuario_id), sga_persona.email, sga_persona.emailinst, sga_materia.cupo AS cupo, (
        #                                 SELECT COUNT(*)
        #                                 FROM sga_materiaasignada ma, sga_matricula mat1
        #                                 WHERE ma.matricula_id=mat1.id AND ma.status=True AND mat1.status=True AND mat1.estado_matricula in (2,3) AND ma.materia_id=sga_materia.id AND ma.id NOT in (
        #                                 SELECT mr.materiaasignada_id
        #                                 FROM sga_materiaasignadaretiro mr)) AS nmatriculados, sga_tipoprofesor.nombre AS Tipoprofesor,
        #                                  sga_profesormateria.desde AS desde, sga_profesormateria.hasta AS hasta, (
        #                                 SELECT ti.nombre
        #                                 FROM sga_profesordistributivohoras dis,sga_tiempodedicaciondocente ti
        #                                 WHERE dis.dedicacion_id=ti.id AND dis.profesor_id=sga_profesor.id AND periodo_id={periodo} AND dis.status= TRUE) AS dedicacion, (
        #                                 SELECT ca.nombre
        #                                 FROM sga_profesordistributivohoras dis,sga_categorizaciondocente ca
        #                                 WHERE dis.categoria_id=ca.id AND dis.profesor_id=sga_profesor.id AND dis.periodo_id={periodo} AND dis.status= TRUE) AS categoria,
        #                                 (CASE sga_asignaturamalla.practicas WHEN TRUE THEN 'SI' ELSE 'NO' END) AS tipomateria,
        #                                 (CASE sga_profesormateria.afinidad WHEN TRUE THEN 'SI' ELSE 'NO' END) AS afinidad,
        #                                 sga_materia.inicio AS inicio, sga_materia.fin AS fin, sga_materia.fechafinasistencias AS finasistencia, sga_materia.id AS id, sga_persona.telefono_conv AS telefonoconv, sga_persona.telefono AS telefono, EXTRACT(YEAR
        #                                 FROM sga_malla.inicio) AS anio, (
        #                                 SELECT modelo.nombre
        #                                 FROM sga_modeloevaluativo modelo
        #                                 WHERE modelo.id = sga_materia.modeloevaluativo_id) AS modeloevaluativo, sga_asignaturamalla.horaspresencialessemanales AS horaspresencialessemanales, sga_profesormateria.aceptarmateria AS aceptarmateria, sga_profesormateria.aceptarmateriaobs AS aceptarmateriaobs, sga_profesormateria.fecha_horario AS fecha_horario, sga_profesormateria.aceptarhorario AS aceptarhorario, sga_profesormateria.aceptarhorarioobs AS aceptarhorarioobs, sga_materia.idcursomoodle AS idcursomoodle, prov.nombre AS provincia, cant.nombre AS ciudad
        #                                 ,(
        #                                 SELECT COUNT(*)
        #                                 FROM sga_materiaasignada ma2, sga_matricula mat1
        #                                 WHERE ma2.matricula_id=mat1.id AND mat1.termino= TRUE AND ma2.materia_id=sga_materia.id AND ma2.id NOT in (
        #                                 SELECT mr.materiaasignada_id
        #                                 FROM sga_materiaasignadaretiro mr)) AS nmatriculados_acpta_termino,
        #                                 (
        #                                 SELECT COUNT(*)
        #                                 FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
        #                                 WHERE mat1.estado_matricula=1 AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id
        #                                 AND ni1.periodo_id=sga_nivel.periodo_id AND mas1.materia_id=sga_materia.id) AS inscritos,
        #                                 (select CASE WHEN sga_detalleasignaturamallamodalidad.modalidad = 1 THEN 'PRESENCIAL' else 'VIRTUAL' END) as modalidad,
        #                                 (CASE sga_asignaturamalla.practicas
        #                                 WHEN True THEN 'SI' ELSE 'NO' END) AS Malla_Practicas,
        #                                 (CASE
        #                                     (SELECT count(*)
        #                                             FROM sga_silabo AS sga_s
        #                                             WHERE sga_s.materia_id = sga_materia.id AND sga_s.status=True AND sga_s.codigoqr=True
        #                                         )
        #                                     WHEN 0 THEN
        #                                         'NO'
        #                                     ELSE
        #                                         'SI'
        #                                     END
        #                                 ) AS silabo,
        #                                 (SELECT
        #                                     count(sga_tp.*)
        #                                 FROM sga_tareapracticasilabosemanal AS sga_tp
        #                                 INNER JOIN sga_silabosemanal AS sga_ss ON sga_tp.silabosemanal_id = sga_ss.id
        #                                 INNER JOIN sga_silabo AS sga_s ON sga_ss.silabo_id = sga_s.id
        #                                 INNER JOIN sga_materia AS sga_m ON sga_m.id = sga_s.materia_id
        #                                 WHERE
        #                                     sga_s.status= TRUE
        #                                     AND sga_s.codigoqr= TRUE
        #                                     AND sga_m.id = sga_materia.id
        #                                     AND sga_tp.estado_id!=3
        #                                     AND sga_tp.status=True
        #                                 ) AS trabajos_practicos,
        #                                 sga_materia.cerrado
        #                                 FROM public.sga_materia sga_materia
        #                                 LEFT JOIN public.sga_profesormateria sga_profesormateria ON sga_materia.id = sga_profesormateria.materia_id AND sga_profesormateria.status= TRUE AND sga_profesormateria.activo= TRUE
        #                                 LEFT JOIN public.sga_profesor sga_profesor ON sga_profesor.id = sga_profesormateria.profesor_id AND sga_profesor.status= TRUE
        #                                 LEFT JOIN public.sga_tipoprofesor sga_tipoprofesor ON sga_tipoprofesor.id = sga_profesormateria.tipoprofesor_id AND sga_tipoprofesor.status= TRUE
        #                                 LEFT JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id AND sga_persona.status= TRUE
        #                                 LEFT JOIN sga_provincia prov ON prov.id=sga_persona.provincia_id
        #                                 LEFT JOIN sga_canton cant ON cant.id=sga_persona.canton_id
        #                                 INNER JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id AND sga_nivel.status= TRUE
        #                                 INNER JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id AND sga_asignatura.status= TRUE
        #                                 INNER JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id AND sga_asignaturamalla.status= TRUE
        #                                 LEFT JOIN public.sga_detalleasignaturamallamodalidad sga_detalleasignaturamallamodalidad ON sga_asignaturamalla.id = sga_detalleasignaturamallamodalidad.asignaturamalla_id
        #                                 INNER JOIN public.sga_nivelmalla sga_nivelmalla ON sga_asignaturamalla.nivelmalla_id = sga_nivelmalla.id AND sga_nivelmalla.status= TRUE
        #                                 INNER JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id AND sga_malla.status= TRUE
        #                                 INNER JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id AND sga_carrera.status= TRUE
        #                                 INNER JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id
        #                                 INNER JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id AND sga_coordinacion.status= TRUE
        #                                 INNER JOIN public.sga_sesion sga_sesion ON sga_nivel.sesion_id = sga_sesion.id
        #                                 INNER JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id
        #                                 WHERE sga_periodo.id = {periodo} AND sga_materia.status= TRUE
        #                                 ORDER BY sga_coordinacion.nombre, sga_carrera.nombre, sga_sesion.nombre, sga_nivelmalla.nombre,sga_materia.paralelo,sga_asignatura.nombre
        #                         """
        #     cursor.execute(sql)
        #     results = cursor.fetchall()
        #
        #     for r in results:
        #         campo1 = r[0].__str__()
        #         campo2 = r[1].__str__()
        #         campo3 = r[2].__str__()
        #         campo4 = r[3].__str__()
        #         campo5 = r[4].__str__()
        #         campo6 = r[5]
        #         campo7 = r[6].__str__()
        #         campo8 = r[7].__str__()
        #         campo9 = r[8]
        #         campo10 = r[9].__str__()
        #         campo11 = r[10].__str__()
        #         # campo12 = r[11]
        #         campo13 = r[12].__str__()
        #         campo14 = r[13].__str__()
        #         campo15 = int(r[14])
        #         campo16 = int(r[15])
        #         campo17 = r[16].__str__()
        #         campo18 = r[17].__str__()
        #         campo19 = r[18].__str__()
        #         campo20 = r[19].__str__()
        #         campo21 = r[20].__str__()
        #         campo22 = r[21].__str__()
        #         campo23 = r[22].__str__()
        #         campo24 = r[23].__str__()
        #         campo25 = r[24].__str__()
        #         campo26 = r[25].__str__()
        #         # campo27 = r[26]
        #         campo28 = r[27].__str__() + " - " + r[28].__str__()
        #         campo29 = r[29].__str__()
        #         campo30 = r[30].__str__()
        #         campo31 = r[31].__str__()
        #         if r[33] == None or r[33] == '':
        #             campo32 = ''
        #             campo33 = ''
        #         else:
        #             campo32 = 'NO'
        #             if r[32]:
        #                 campo32 = 'SI'
        #             campo33 = r[33]
        #         campo34 = r[34]
        #         if r[36] == None or r[36] == '':
        #             campo35 = ''
        #             campo36 = ''
        #         else:
        #             campo35 = 'NO'
        #             if r[35]:
        #                 campo35 = 'SI'
        #             campo36 = r[36]
        #         campo37 = r[37]
        #         campo38 = r[38]
        #         campo39 = r[39]
        #         campo_cant_acapta = r[40]
        #         inscritos = int(r[41])
        #         campo40 = r[42].__str__()
        #         campo41 = r[43]
        #         campo42 = r[44]
        #         campo43 = r[45]
        #         campo46 = r[46]
        #         totalMatriculados = inscritos + int(campo16)
        #         ws.write(row_num, 0, campo1, formatoceldacenter)
        #         ws.write(row_num, 1, campo2, formatoceldacenter)
        #         ws.write(row_num, 2, campo29, formatoceldacenter)
        #         ws.write(row_num, 3, campo3, formatoceldacenter)
        #         ws.write(row_num, 4, campo4, formatoceldacenter)
        #         ws.write(row_num, 5, campo5, formatoceldacenter)
        #         ws.write(row_num, 6, campo7, formatoceldacenter)
        #         ws.write(row_num, 7, campo22, formatoceldacenter)
        #         ws.write(row_num, 8, campo15, formatoceldacenter)
        #         ws.write(row_num, 9, campo16, formatoceldacenter)
        #         ws.write(row_num, 10, inscritos, formatoceldacenter)
        #         ws.write(row_num, 11, totalMatriculados, formatoceldacenter)
        #
        #         ws.write(row_num, 12, campo8, formatoceldacenter)
        #         ws.write(row_num, 13, campo11, formatoceldacenter)
        #         # ws.write(row_num, 12, campo12, font_style2)
        #         ws.write(row_num, 14, campo23, formatoceldacenter)
        #         ws.write(row_num, 15, campo9, formatoceldacenter)
        #         ws.write(row_num, 16, campo9, formatoceldacenter)
        #         ws.write(row_num, 17, campo10, formatoceldacenter)
        #         ws.write(row_num, 18, campo13, formatoceldacenter)
        #         ws.write(row_num, 19, campo14, formatoceldacenter)
        #         ws.write(row_num, 20, campo17, formatoceldacenter)
        #         ws.write(row_num, 21, campo18, formatoceldacenter)
        #         ws.write(row_num, 22, campo19, formatoceldacenter)
        #         ws.write(row_num, 23, campo20, formatoceldacenter)
        #         ws.write(row_num, 24, campo21, formatoceldacenter)
        #         ws.write(row_num, 25, campo24, formatoceldacenter)
        #         ws.write(row_num, 26, campo25, formatoceldacenter)
        #         ws.write(row_num, 27, campo26, formatoceldacenter)
        #         # ws.write(row_num, 25, campo27, font_style2)
        #         ws.write(row_num, 28, campo28, formatoceldacenter)
        #         ws.write(row_num, 29, campo30, formatoceldacenter)
        #         ws.write(row_num, 30, campo6, formatoceldacenter)
        #         ws.write(row_num, 31, campo32, formatoceldacenter)
        #         ws.write(row_num, 32, campo33, formatoceldacenter)
        #         ws.write(row_num, 33, str(campo34), formatoceldacenter)
        #         ws.write(row_num, 34, campo35, formatoceldacenter)
        #         ws.write(row_num, 35, campo36, formatoceldacenter)
        #         ws.write(row_num, 36, campo37, formatoceldacenter)
        #         ws.write(row_num, 37, campo38, formatoceldacenter)
        #         ws.write(row_num, 38, campo39, formatoceldacenter)
        #         ws.write(row_num, 39, campo_cant_acapta, formatoceldacenter)
        #         ws.write(row_num, 40, campo40, formatoceldacenter)
        #         ws.write(row_num, 41, campo41, formatoceldacenter)
        #         ws.write(row_num, 42, campo42, formatoceldacenter)
        #         ws.write(row_num, 43, campo43, formatoceldacenter)
        #         ws.write(row_num, 44, campo46, formatoceldacenter)
        #         row_num += 1
        #     workbook.close()
        #
        #     usernotify = User.objects.get(pk=request.user.pk)
        #     pers = Persona.objects.get(usuario=usernotify)
        #     if notif > 0:
        #         noti = Notificacion.objects.get(pk=notif)
        #         noti.en_proceso = False
        #         noti.cuerpo = 'Excel Listo'
        #         noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
        #         noti.save()
        #     else:
        #         noti = Notificacion(cuerpo='Reporte Listo',
        #                             titulo='Excel Distributivo de asignaturas',
        #                             destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
        #                             prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
        #                             tipo=2, en_proceso=False)
        #         noti.save(request)
        #
        #     send_user_notification(user=usernotify, payload={
        #         "head": "Excel terminado",
        #         "body": 'Excel Distributivo de asignaturas',
        #         "action": "notificacion",
        #         "timestamp": time.mktime(datetime.now().timetuple()),
        #         "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
        #         "btn_notificaciones": traerNotificaciones(request, data, pers),
        #         "mensaje": 'Su reporte ha sido generado con exito'
        #     }, ttl=500)

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

class reporte_estudiantes_unemi(threading.Thread):
    def __init__(self, request, notiid):
        self.request = request
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        request, notiid = self.request, self.notiid
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'ventasu')
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        nombre_archivo = "REPORTE_GRADUADOS_UNEMI{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'ventasu', nombre_archivo)
        try:
            workbook = xlsxwriter.Workbook(directory)
            ws = workbook.add_worksheet('resultados')

            ws.set_column(0, 0, 10)
            ws.set_column(1, 1, 15)
            ws.set_column(2, 2, 15)
            ws.set_column(3, 3, 35)
            ws.set_column(4, 4, 25)
            ws.set_column(5, 5, 40)
            ws.set_column(6, 6, 40)
            ws.set_column(7, 7, 40)
            ws.set_column(8, 8, 15)
            ws.set_column(9, 9, 15)
            ws.set_column(10, 10, 45)
            ws.set_column(11, 11, 35)
            ws.set_column(12, 12, 30)
            ws.set_column(13, 13, 40)
            ws.set_column(14, 14, 35)
            ws.set_column(15, 15, 20)
            ws.set_column(16, 16, 45)

            formatotitulo_filtros = workbook.add_format(
                {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

            formatoceldacab = workbook.add_format(
                {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247',
                 'font_color': 'white'})
            formatoceldaleft = workbook.add_format(
                {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

            formatoceldaleft2 = workbook.add_format(
                {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter',
                 'border': 1, 'bold': 1})

            formatoceldaleft3 = workbook.add_format(
                {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})

            decimalformat = workbook.add_format(
                {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter',
                 'border': 1})

            decimalformat2 = workbook.add_format(
                {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter',
                 'border': 1, 'bold': 1})
            carrera = request.GET.getlist('carrera[]')
            anio = 0

            if 'carrera' in request.GET:
                carrera = request.GET['carrera']
            if 'anio' in request.GET:
                anio = int(request.GET['anio'])

            ws.write(2, 0, 'N°', formatoceldacab)
            ws.write(2, 1, 'FECHA GRADUACIÓN', formatoceldacab)
            ws.write(2, 2, 'CÉDULA', formatoceldacab)
            ws.write(2, 3, 'CARRERA', formatoceldacab)
            ws.write(2, 4, 'MODALIDAD', formatoceldacab)
            ws.write(2, 5, 'NOMBRES', formatoceldacab)
            ws.write(2, 6, 'EMAIL PERSONAL', formatoceldacab)
            ws.write(2, 7, 'EMAIL CORPORATIVO- INSTITUCIONAL', formatoceldacab)
            ws.write(2, 8, 'TELÉFONO 1', formatoceldacab)
            ws.write(2, 9, 'TELÉFONO 2', formatoceldacab)
            ws.write(2, 10, 'TITULO TERCER NIVEL OBTENIDO', formatoceldacab)
            ws.write(2, 11, 'INSTITUCION DONDE OBTUVO EL 3ER NIVEL', formatoceldacab)
            ws.write(2, 12, 'TITULO  CUARTO NIVEL OBTENIDO', formatoceldacab)
            ws.write(2, 13, 'INSTITUCION DONDE OBTUVO EL 4TO NIVEL', formatoceldacab)
            ws.write(2, 14, 'TRABAJA EN', formatoceldacab)
            ws.write(2, 15, '¿TIENE POSTULACIÓN POSGRADO?', formatoceldacab)
            ws.write(2, 16, 'POSTULACIONES', formatoceldacab)

            filtro = Q(status=True)

            if carrera != "":
                if carrera[0] != "0":
                    filtro = filtro & Q(inscripcion__carrera__id__in=carrera)

            if anio > 0:
                filtro = filtro & Q(fechagraduado__year=anio)

            graduados = Graduado.objects.filter(filtro).exclude(inscripcion__carrera__coordinacion__id=7).order_by(
                'inscripcion__carrera__nombre')

            filas_recorridas = 4
            cont = 1

            for graduado in graduados:
                titulo = ''
                if graduado.inscripcion.carrera.titulootorga:
                    titulo = graduado.inscripcion.carrera.titulootorga
                elif graduado.inscripcion.carrera.tituloobtenido:
                    titulo = graduado.inscripcion.carrera.tituloobtenido
                elif graduado.nombretitulo:
                    titulo = graduado.nombretitulo
                elif graduado.inscripcion.carrera.tituloobtenidohombre:
                    titulo = graduado.inscripcion.carrera.tituloobtenidohombre
                elif graduado.inscripcion.carrera.tituloobtenidomujer:
                    titulo = graduado.inscripcion.carrera.tituloobtenidomujer

                ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
                ws.write('B%s' % filas_recorridas, str(graduado.fechagraduado), formatoceldaleft)
                ws.write('C%s' % filas_recorridas, str(graduado.inscripcion.persona.cedula), formatoceldaleft)
                ws.write('D%s' % filas_recorridas, str(graduado.inscripcion.carrera), formatoceldaleft)
                ws.write('E%s' % filas_recorridas, str(graduado.inscripcion.carrera.get_modalidad_display()), formatoceldaleft)
                ws.write('F%s' % filas_recorridas, str(graduado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
                ws.write('G%s' % filas_recorridas, str(graduado.inscripcion.persona.email if graduado.inscripcion.persona.email else 'NO REGISTRA'), formatoceldaleft)
                ws.write('H%s' % filas_recorridas, str(graduado.inscripcion.persona.emailinst if graduado.inscripcion.persona.email else 'NO REGISTRA'), formatoceldaleft)
                ws.write('I%s' % filas_recorridas, str(graduado.inscripcion.persona.telefono), formatoceldaleft)
                ws.write('J%s' % filas_recorridas, str(graduado.inscripcion.persona.telefono2), formatoceldaleft)
                ws.write('K%s' % filas_recorridas, str(titulo), formatoceldaleft)
                ws.write('L%s' % filas_recorridas, str('UNIVERSIDAD ESTATAL DE MILAGRO'), formatoceldaleft)
                ws.write('M%s' % filas_recorridas, str(titulo_4_nivel(graduado.inscripcion.persona)), formatoceldaleft)
                ws.write('N%s' % filas_recorridas, str(universidad_titulo_4_nivel(graduado.inscripcion.persona)), formatoceldaleft)
                ws.write('O%s' % filas_recorridas, str(trabaja_en(graduado.inscripcion.persona)), formatoceldaleft)
                ws.write('P%s' % filas_recorridas, str(tiene_postulacion(graduado.inscripcion.persona)), formatoceldaleft)
                ws.write('Q%s' % filas_recorridas, str(tiene_postulacion_descripcion(graduado.inscripcion.persona)), formatoceldaleft)

                filas_recorridas += 1
                cont += 1

            workbook.close()

            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/ventasu/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Resultados de Encuesta',
                                    destinatario=pers, url="{}reportes/ventasu/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Resultado de reporte graduados UNEMI',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}reportes/ventasu/{}".format(MEDIA_URL, nombre_archivo),
                "mensaje": 'Los resultados del reporte de graduados UNEMI han sido generados con exito'
            }, ttl=500)
        except Exception as ex:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error en el reporte'
                noti.cuerpo = 'Ha ocurrido un error en al generar el reporte {} - Error en la linea {}'.format(ex,
                                                                                                               sys.exc_info()[
                                                                                                                   -1].tb_lineno)
                noti.url = ""
                noti.error = True
                noti.save()
            else:
                noti = Notificacion(titulo='Error en el reporte',
                                    cuerpo='Ha ocurrido un error en al generar el reporte {} - Error en la linea {}'.format(
                                        ex, sys.exc_info()[-1].tb_lineno),
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False, error=True)
                noti.save(request)
            data = {}
            send_user_notification(user=usernotify, payload={
                "head": "Error en el reporte",
                "body": "Ha ocurrido un error en al generar el reporte {} - Error en la linea {}".format(ex,
                                                                                                         sys.exc_info()[
                                                                                                             -1].tb_lineno),
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "",
                "mensaje": "Ha ocurrido un error en al generar el reporte"
            }, ttl=500)
            pass

class reporte_estudiantes_matricula_posgrado(threading.Thread):
    def __init__(self, request, notiid):
        self.request = request
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        request, notiid = self.request, self.notiid
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matricespos')
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        nombre_archivo = "ESTUDIANTES_MATRICULA_POS_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matricespos', nombre_archivo)
        try:
            workbook = xlsxwriter.Workbook(directory)
            ws = workbook.add_worksheet('resultados_es_mat_pos')

            ws.set_column(0, 0, 20)
            ws.set_column(1, 1, 25)
            ws.set_column(2, 2, 25)
            ws.set_column(3, 3, 20)
            ws.set_column(4, 4, 20)
            ws.set_column(5, 5, 20)
            ws.set_column(6, 6, 45)
            ws.set_column(7, 7, 40)
            ws.set_column(8, 8, 35)
            ws.set_column(9, 9, 25)
            ws.set_column(10, 10, 20)
            ws.set_column(11, 11, 30)
            ws.set_column(12, 12, 25)
            ws.set_column(13, 13, 30)
            ws.set_column(14, 14, 40)
            ws.set_column(15, 15, 40)
            ws.set_column(16, 16, 30)
            ws.set_column(17, 17, 30)
            ws.set_column(18, 18, 30)
            ws.set_column(19, 19, 30)
            ws.set_column(20, 20, 30)
            ws.set_column(21, 21, 25)
            ws.set_column(22, 22, 20)
            ws.set_column(23, 23, 15)
            ws.set_column(24, 24, 20)
            ws.set_column(25, 25, 20)
            ws.set_column(26, 26, 30)
            ws.set_column(27, 27, 30)

            formatotitulo_filtros = workbook.add_format(
                {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

            formatoceldacab = workbook.add_format(
                {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247',
                 'font_color': 'white'})
            formatoceldaleft = workbook.add_format(
                {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

            formatoceldaleft2 = workbook.add_format(
                {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
                 'bold': 1})

            formatoceldaleft3 = workbook.add_format(
                {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})

            decimalformat = workbook.add_format(
                {'num_format': '#,##0.0', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

            decimalformat2 = workbook.add_format(
                {'num_format': '#,##0.0', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
                 'bold': 1})

            maestria = request.GET.getlist('maestria[]')
            anio = request.GET.getlist('anio[]')
            periodo = request.GET.getlist('periodo[]')

            if 'maestria' in request.GET:
                maestria = request.GET['maestria']

            if 'anio' in request.GET:
                anio = request.GET['anio']

            if 'periodo' in request.GET:
                periodo = request.GET['periodo']

            ws.merge_range('A1:AB1', 'Estudiantes - Matrícula Periodo Académico (Posgrado)', formatotitulo_filtros)

            ws.write(1, 0, 'CODIGO_IES', formatoceldacab)
            ws.write(1, 1, 'CODIGO CARRERA', formatoceldacab)
            ws.write(1, 2, 'CIUDAD CARRERA', formatoceldacab)
            ws.write(1, 3, 'TIPO IDENTIFICACION', formatoceldacab)
            ws.write(1, 4, 'NÚMERO_CÉDULA', formatoceldacab)
            ws.write(1, 5, 'NÚMERO_PASAPORTE', formatoceldacab)
            ws.write(1, 6, 'NOMBRES', formatoceldacab)
            ws.write(1, 7, 'CARRERA', formatoceldacab)
            ws.write(1, 8, 'PERIODO', formatoceldacab)
            ws.write(1, 9, 'FECHA_INICIO_PERIODO', formatoceldacab)
            ws.write(1, 10, 'FECHA_FIN_PERIODO', formatoceldacab)
            ws.write(1, 11, 'TOTAL_CREDITOS_APROBADOS', formatoceldacab)
            ws.write(1, 12, 'CREDITOS_APROBADOS', formatoceldacab)
            ws.write(1, 13, 'NIVEL ACADEMICO', formatoceldacab)
            ws.write(1, 14, 'NUM_MATERIAS_SEGUNDA_MATRICULA', formatoceldacab)
            ws.write(1, 15, 'NUM_MATERIAS_TERCERA_MATRICULA', formatoceldacab)
            ws.write(1, 16, 'PERDIDA_GRATUIDAD', formatoceldacab)
            ws.write(1, 17, 'TOTAL_HORAS_APROBADAS', formatoceldacab)
            ws.write(1, 18, 'HORAS APROBADAS PERIODO', formatoceldacab)
            ws.write(1, 19, 'MONTO_AYUDA_ECONOMICA', formatoceldacab)
            ws.write(1, 20, 'MONTO_CREDITO_EDUCATIVO', formatoceldacab)
            ws.write(1, 21, 'ESTADO', formatoceldacab)
            ws.write(1, 22, 'MODALIDAD', formatoceldacab)
            ws.write(1, 23, 'GLBTI', formatoceldacab)
            ws.write(1, 24, 'EXTRANJERO', formatoceldacab)
            ws.write(1, 25, 'FECHA_MATRICULA', formatoceldacab)
            ws.write(1, 26, 'FECHA_INICIO_PRIMER_MODULO', formatoceldacab)
            ws.write(1, 27, 'FECHA_FIN_ÚLTIMO_MODULO', formatoceldacab)

            filtro = Q(status=True, inscripcion__carrera__coordinacion__id=7, retiradomatricula=False, inscripcion__status=True)

            if maestria != "":
                if maestria[0] != "0":
                    filtro = filtro & Q(inscripcion__carrera__id__in=maestria)

            if anio != "":
                if anio[0] != "0":
                    filtro = filtro & Q(fecha__year__in=anio)

            if periodo != "":
                if periodo[0] != "0":
                    filtro = filtro & Q(nivel__periodo__in=periodo)

            matriculados = Matricula.objects.filter(filtro).order_by('fecha')

            filas_recorridas = 3
            cont = 1
            for matriculado in matriculados:
                ws.write('A%s' % filas_recorridas, str(1024), formatoceldaleft)
                ws.write('B%s' % filas_recorridas, str(matriculado.inscripcion.carrera.codigo if matriculado.inscripcion.carrera.codigo else ''), formatoceldaleft)
                ws.write('C%s' % filas_recorridas, str('MILAGRO'), formatoceldaleft)
                ws.write('D%s' % filas_recorridas, str(matriculado.inscripcion.persona.tipo_identificacion_completo()), formatoceldaleft)
                ws.write('E%s' % filas_recorridas, str(matriculado.inscripcion.persona.cedula if matriculado.inscripcion.persona.cedula else ''), formatoceldaleft)
                ws.write('F%s' % filas_recorridas, str(matriculado.inscripcion.persona.pasaporte if matriculado.inscripcion.persona.pasaporte else ''), formatoceldaleft)
                ws.write('G%s' % filas_recorridas, str(matriculado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
                ws.write('H%s' % filas_recorridas, str(f'{matriculado.inscripcion.carrera.nombre} CON MENCIÓN EN {matriculado.inscripcion.carrera.mencion}' if matriculado.inscripcion.carrera.mencion else matriculado.inscripcion.carrera.nombre), formatoceldaleft)
                ws.write('I%s' % filas_recorridas, str(matriculado.nivel.periodo), formatoceldaleft)
                ws.write('J%s' % filas_recorridas, str(matriculado.nivel.periodo.inicio), formatoceldaleft)
                ws.write('K%s' % filas_recorridas, str(matriculado.nivel.periodo.fin), formatoceldaleft)
                ws.write('L%s' % filas_recorridas, total_creditos_aprobadas(matriculado.inscripcion), decimalformat)
                ws.write('M%s' % filas_recorridas, total_creditos(matriculado.inscripcion), decimalformat)
                ws.write('N%s' % filas_recorridas, str('1,2'), formatoceldaleft)
                ws.write('O%s' % filas_recorridas, num_segunda_mat(matriculado), decimalformat)
                ws.write('P%s' % filas_recorridas, num_tercera_mat(matriculado), decimalformat)
                ws.write('Q%s' % filas_recorridas, str('SI'), formatoceldaleft)
                ws.write('R%s' % filas_recorridas, total_horas_aprobadas(matriculado.inscripcion), decimalformat)
                ws.write('S%s' % filas_recorridas, total_horas_ma(matriculado.inscripcion), decimalformat)
                ws.write('T%s' % filas_recorridas, 0, decimalformat)
                ws.write('U%s' % filas_recorridas, 0, decimalformat)
                ws.write('V%s' % filas_recorridas, str('NO APLICA'), formatoceldaleft)
                ws.write('W%s' % filas_recorridas, str(matriculado.inscripcion.carrera.get_modalidad_display()), formatoceldaleft)
                ws.write('X%s' % filas_recorridas, str('SI' if matriculado.inscripcion.persona.lgtbi else 'NO'), formatoceldaleft)
                ws.write('Y%s' % filas_recorridas, str(es_extranjero(matriculado.inscripcion.persona)), formatoceldaleft)
                ws.write('Z%s' % filas_recorridas, str(matriculado.fecha if matriculado.fecha else ''), formatoceldaleft)
                ws.write('AA%s' % filas_recorridas, str(inicio_primer_modulo(matriculado)), formatoceldaleft)
                ws.write('AB%s' % filas_recorridas, str(fin_ultimo_modulo(matriculado)), formatoceldaleft)

                filas_recorridas += 1
                cont += 1

            workbook.close()
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Estudiantes - Matrícula Periodo Académico (Posgrado)',
                                    destinatario=pers, url="{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            # data = {}
            # send_user_notification(user=usernotify, payload={
            #     "head": "Reporte terminado",
            #     "body": 'Estudiantes - Matrícula Periodo Académico (Posgrado)',
            #     "action": "notificacion",
            #     "timestamp": time.mktime(datetime.now().timetuple()),
            #     "url": "{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo),
            #     "mensaje": 'Los resultados del reporte Estudiantes - Matrícula Periodo Académico (Posgrado) han sido generados con exito'
            # }, ttl=500)

        except Exception as ex:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error en el reporte'
                noti.cuerpo = 'Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.'.format(ex,
                                                                                                               sys.exc_info()[
                                                                                                                   -1].tb_lineno)
                noti.url = ""
                noti.error = True
                noti.save()
            else:
                noti = Notificacion(titulo='Error en el reporte',
                                    cuerpo='Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.'.format(
                                        ex, sys.exc_info()[-1].tb_lineno),
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False, error=True)
                noti.save(request)
            # data = {}
            # send_user_notification(user=usernotify, payload={
            #     "head": "Error en el reporte",
            #     "body": "Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.".format(ex,
            #                                                                                              sys.exc_info()[
            #                                                                                                  -1].tb_lineno),
            #     "action": "notificacion",
            #     "timestamp": time.mktime(datetime.now().timetuple()),
            #     "url": "",
            #     "mensaje": "Ha ocurrido un error en al generar el reporte"
            # }, ttl=500)
            pass

class reporte_estudiantes_posgrado(threading.Thread):
    def __init__(self, request, notiid):
        self.request = request
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        request, notiid = self.request, self.notiid
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matricespos')
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        nombre_archivo = "ESTUDIANTES_POSGRADO_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matricespos', nombre_archivo)
        try:
            workbook = xlsxwriter.Workbook(directory)
            ws = workbook.add_worksheet('listado_estudiantes_posgrado')
            ws.set_column(0, 0, 20)
            ws.set_column(1, 1, 25)
            ws.set_column(2, 2, 25)
            ws.set_column(3, 3, 20)
            ws.set_column(4, 4, 20)
            ws.set_column(5, 5, 20)
            ws.set_column(6, 6, 45)
            ws.set_column(7, 7, 40)
            ws.set_column(8, 8, 35)
            ws.set_column(9, 9, 25)
            ws.set_column(10, 10, 20)
            ws.set_column(11, 11, 20)
            ws.set_column(12, 12, 25)
            ws.set_column(13, 13, 30)
            ws.set_column(14, 14, 30)
            ws.set_column(15, 15, 25)
            ws.set_column(16, 16, 25)
            ws.set_column(17, 17, 25)
            ws.set_column(18, 18, 20)
            ws.set_column(19, 19, 30)
            ws.set_column(20, 20, 30)
            ws.set_column(21, 21, 30)
            ws.set_column(22, 22, 30)
            ws.set_column(23, 23, 20)
            ws.set_column(24, 24, 20)
            ws.set_column(25, 25, 30)
            ws.set_column(26, 26, 30)
            ws.set_column(27, 27, 30)
            ws.set_column(28, 28, 20)
            ws.set_column(29, 29, 20)
            ws.set_column(30, 30, 30)
            ws.set_column(31, 31, 30)
            ws.set_column(32, 32, 30)
            ws.set_column(33, 33, 30)

            formatotitulo_filtros = workbook.add_format(
                {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

            formatoceldacab = workbook.add_format(
                {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247',
                 'font_color': 'white'})
            formatoceldaleft = workbook.add_format(
                {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

            formatoceldaleft2 = workbook.add_format(
                {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
                 'bold': 1})

            formatoceldaleft3 = workbook.add_format(
                {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})

            decimalformat = workbook.add_format(
                {'num_format': '#,##0.0', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

            decimalformat2 = workbook.add_format(
                {'num_format': '#,##0.0', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
                 'bold': 1})

            maestria = request.GET.getlist('maestria[]')
            anio = request.GET.getlist('anio[]')
            periodo = request.GET.getlist('periodo[]')

            if 'maestria' in request.GET:
                maestria = request.GET['maestria']

            if 'anio' in request.GET:
                anio = request.GET['anio']

            if 'periodo' in request.GET:
                periodo = request.GET['periodo']

            ws.merge_range('A1:AH1', 'Estudiantes Posgrado', formatotitulo_filtros)

            ws.write(1, 0, 'CODIGO_IES', formatoceldacab)
            ws.write(1, 1, 'CODIGO CARRERA', formatoceldacab)
            ws.write(1, 2, 'CIUDAD CARRERA', formatoceldacab)
            ws.write(1, 3, 'TIPO IDENTIFICACION', formatoceldacab)
            ws.write(1, 4, 'NÚMERO_CÉDULA', formatoceldacab)
            ws.write(1, 5, 'NÚMERO_PASAPORTE', formatoceldacab)
            ws.write(1, 6, 'NOMBRES', formatoceldacab)
            ws.write(1, 7, 'CARRERA', formatoceldacab)
            ws.write(1, 8, 'PERIODO', formatoceldacab)
            ws.write(1, 9, 'FECHA_INICIO_PERIODO', formatoceldacab)
            ws.write(1, 10, 'FECHA_FIN_PERIODO', formatoceldacab)
            ws.write(1, 11, 'SEXO', formatoceldacab)
            ws.write(1, 12, 'PAIS ORIGEN', formatoceldacab)
            ws.write(1, 13, 'DISCAPACIDAD', formatoceldacab)
            ws.write(1, 14, 'PORCENTAJE DISCAPACIDAD', formatoceldacab)
            ws.write(1, 15, 'NUMERO CONADIS', formatoceldacab)
            ws.write(1, 16, 'ETNIA', formatoceldacab)
            ws.write(1, 17, 'TIPO_INDÍGENA', formatoceldacab)
            ws.write(1, 18, 'NACIONALIDAD', formatoceldacab)
            ws.write(1, 19, 'EMAIL INSTITUCIONAL', formatoceldacab)
            ws.write(1, 20, 'FECHA INICIO PRIMER NIVEL', formatoceldacab)
            ws.write(1, 21, 'FECHA INGRESO CONVALIDACION', formatoceldacab)
            ws.write(1, 22, 'PAIS RESIDENCIA', formatoceldacab)
            ws.write(1, 23, 'PROVINCIA RESIDENCIA', formatoceldacab)
            ws.write(1, 24, 'CANTON RESIDENCIA', formatoceldacab)
            ws.write(1, 25, 'TIPO COLEGIO', formatoceldacab)
            ws.write(1, 26, 'POLITICA CUOTA', formatoceldacab)
            ws.write(1, 27, 'NOMBRE_PROGRAMA', formatoceldacab)
            ws.write(1, 28, 'MODALIDAD', formatoceldacab)
            ws.write(1, 29, 'GLBTI', formatoceldacab)
            ws.write(1, 30, 'EXTRANJERO', formatoceldacab)
            ws.write(1, 31, 'FECHA_MATRICULA', formatoceldacab)
            ws.write(1, 32, 'FECHA_INICIO_PRIMER_MODULO', formatoceldacab)
            ws.write(1, 33, 'FECHA_FIN_ÚLTIMO_MODULO', formatoceldacab)

            filtro = Q(status=True, inscripcion__carrera__coordinacion__id=7, retiradomatricula=False, inscripcion__status=True)

            if maestria != "":
                if maestria[0] != "0":
                    filtro = filtro & Q(inscripcion__carrera__id__in=maestria)

            if anio != "":
                if anio[0] != "0":
                    filtro = filtro & Q(fecha__year__in=anio)

            if periodo != "":
                if periodo[0] != "0":
                    filtro = filtro & Q(nivel__periodo__in=periodo)

            matriculados = Matricula.objects.filter(filtro).order_by('fecha')

            filas_recorridas = 3
            cont = 1
            for matriculado in matriculados:
                ws.write('A%s' % filas_recorridas, str(1024), formatoceldaleft)
                ws.write('B%s' % filas_recorridas, str(matriculado.inscripcion.carrera.codigo if matriculado.inscripcion.carrera.codigo else ''), formatoceldaleft)
                ws.write('C%s' % filas_recorridas, str('MILAGRO'), formatoceldaleft)
                ws.write('D%s' % filas_recorridas, str(matriculado.inscripcion.persona.tipo_identificacion_completo()), formatoceldaleft)
                ws.write('E%s' % filas_recorridas, str(matriculado.inscripcion.persona.cedula if matriculado.inscripcion.persona.cedula else ''), formatoceldaleft)
                ws.write('F%s' % filas_recorridas, str(matriculado.inscripcion.persona.pasaporte if matriculado.inscripcion.persona.pasaporte else ''), formatoceldaleft)
                ws.write('G%s' % filas_recorridas, str(matriculado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
                ws.write('H%s' % filas_recorridas, str(f'{matriculado.inscripcion.carrera.nombre} CON MENCIÓN EN {matriculado.inscripcion.carrera.mencion}' if matriculado.inscripcion.carrera.mencion else matriculado.inscripcion.carrera.nombre), formatoceldaleft)
                ws.write('I%s' % filas_recorridas, str(matriculado.nivel.periodo), formatoceldaleft)
                ws.write('J%s' % filas_recorridas, str(matriculado.nivel.periodo.inicio), formatoceldaleft)
                ws.write('K%s' % filas_recorridas, str(matriculado.nivel.periodo.fin), formatoceldaleft)
                ws.write('L%s' % filas_recorridas, str(matriculado.inscripcion.persona.sexo.nombre if matriculado.inscripcion.persona.sexo else ''), formatoceldaleft)
                ws.write('M%s' % filas_recorridas, str(matriculado.inscripcion.persona.paisnacimiento if matriculado.inscripcion.persona.paisnacimiento else ''), formatoceldaleft)
                ws.write('N%s' % filas_recorridas, str(matriculado.inscripcion.persona.mi_perfil().tipodiscapacidad.nombre if matriculado.inscripcion.persona.mi_perfil().tienediscapacidad and matriculado.inscripcion.persona.mi_perfil().tipodiscapacidad else ''), formatoceldaleft)
                ws.write('O%s' % filas_recorridas, str(f'{matriculado.inscripcion.persona.mi_perfil().porcientodiscapacidad}%' if matriculado.inscripcion.persona.mi_perfil().tienediscapacidad and matriculado.inscripcion.persona.mi_perfil().porcientodiscapacidad else '0%'), formatoceldaleft)
                ws.write('P%s' % filas_recorridas, str(matriculado.inscripcion.persona.mi_perfil().carnetdiscapacidad if matriculado.inscripcion.persona.mi_perfil().tienediscapacidad and matriculado.inscripcion.persona.mi_perfil().carnetdiscapacidad else ''), formatoceldaleft)
                ws.write('Q%s' % filas_recorridas, str(matriculado.inscripcion.persona.mi_perfil().raza.nombre if matriculado.inscripcion.persona.mi_perfil().raza else ''), formatoceldaleft)
                ws.write('R%s' % filas_recorridas, str(matriculado.inscripcion.persona.mi_perfil().nacionalidadindigena.nombre if matriculado.inscripcion.persona.mi_perfil().nacionalidadindigena else ''), formatoceldaleft)
                ws.write('S%s' % filas_recorridas, str(matriculado.inscripcion.persona.nacionalidad if matriculado.inscripcion.persona.nacionalidad else ''), formatoceldaleft)
                ws.write('T%s' % filas_recorridas, str(matriculado.inscripcion.persona.emailinst), formatoceldaleft)
                ws.write('U%s' % filas_recorridas, str(matriculado.inscripcion.fechainicioprimernivel if matriculado.inscripcion.fechainicioprimernivel else ''), formatoceldaleft)
                ws.write('V%s' % filas_recorridas, '', formatoceldaleft)
                ws.write('W%s' % filas_recorridas, str(matriculado.inscripcion.persona.pais.nombre if matriculado.inscripcion.persona.pais else ''), formatoceldaleft)
                ws.write('X%s' % filas_recorridas, str(matriculado.inscripcion.persona.provincia.nombre if matriculado.inscripcion.persona.provincia else ''), formatoceldaleft)
                ws.write('Y%s' % filas_recorridas, str(matriculado.inscripcion.persona.canton.nombre if matriculado.inscripcion.persona.canton else ''), formatoceldaleft)
                ws.write('Z%s' % filas_recorridas, str(matriculado.inscripcion.unidadeducativa.tipocolegio.nombre if matriculado.inscripcion.unidadeducativa else ''), formatoceldaleft)
                ws.write('AA%s' % filas_recorridas, str(''), formatoceldaleft)
                ws.write('AB%s' % filas_recorridas, str(f'{matriculado.inscripcion.carrera.nombre} CON MENCIÓN EN {matriculado.inscripcion.carrera.mencion}' if matriculado.inscripcion.carrera.mencion else matriculado.inscripcion.carrera.nombre), formatoceldaleft)
                ws.write('AC%s' % filas_recorridas, str(matriculado.inscripcion.carrera.get_modalidad_display()), formatoceldaleft)
                ws.write('AD%s' % filas_recorridas, str('SI' if matriculado.inscripcion.persona.lgtbi else 'NO'), formatoceldaleft)
                ws.write('AE%s' % filas_recorridas, str(es_extranjero(matriculado.inscripcion.persona)), formatoceldaleft)
                ws.write('AF%s' % filas_recorridas, str(matriculado.fecha if matriculado.fecha else ''), formatoceldaleft)
                ws.write('AG%s' % filas_recorridas, str(inicio_primer_modulo(matriculado)), formatoceldaleft)
                ws.write('AH%s' % filas_recorridas, str(fin_ultimo_modulo(matriculado)), formatoceldaleft)

                filas_recorridas += 1
                cont += 1

            workbook.close()
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Estudiantes Posgrado',
                                    destinatario=pers, url="{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            # data = {}
            # send_user_notification(user=usernotify, payload={
            #     "head": "Reporte terminado",
            #     "body": 'Estudiantes Posgrado',
            #     "action": "notificacion",
            #     "timestamp": time.mktime(datetime.now().timetuple()),
            #     "url": "{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo),
            #     "mensaje": 'Los resultados del reporte Estudiantes Posgrado han sido generados con exito'
            # }, ttl=500)

        except Exception as ex:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error en el reporte'
                noti.cuerpo = 'Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.'.format(ex,
                                                                                                               sys.exc_info()[
                                                                                                                   -1].tb_lineno)
                noti.url = ""
                noti.error = True
                noti.save()
            else:
                noti = Notificacion(titulo='Error en el reporte',
                                    cuerpo='Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.'.format(
                                        ex, sys.exc_info()[-1].tb_lineno),
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False, error=True)
                noti.save(request)
            # data = {}
            # send_user_notification(user=usernotify, payload={
            #     "head": "Error en el reporte",
            #     "body": "Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.".format(ex,
            #                                                                                              sys.exc_info()[
            #                                                                                                  -1].tb_lineno),
            #     "action": "notificacion",
            #     "timestamp": time.mktime(datetime.now().timetuple()),
            #     "url": "",
            #     "mensaje": "Ha ocurrido un error en al generar el reporte"
            # }, ttl=500)
            pass

class reporte_graduados_posgrado(threading.Thread):
    def __init__(self, request, notiid):
        self.request = request
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        request, notiid = self.request, self.notiid
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matricespos')
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        nombre_archivo = "ESTUDIANTES_GRADUADOS_POS_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'matricespos', nombre_archivo)
        try:
            workbook = xlsxwriter.Workbook(directory)
            ws = workbook.add_worksheet('graduados_posgrado')
            ws.set_column(0, 0, 20)
            ws.set_column(1, 1, 25)
            ws.set_column(2, 2, 25)
            ws.set_column(3, 3, 20)
            ws.set_column(4, 4, 20)
            ws.set_column(5, 5, 20)
            ws.set_column(6, 6, 45)
            ws.set_column(7, 7, 40)
            ws.set_column(8, 8, 35)
            ws.set_column(9, 9, 25)
            ws.set_column(10, 10, 20)
            ws.set_column(11, 11, 25)
            ws.set_column(12, 12, 30)
            ws.set_column(13, 13, 30)
            ws.set_column(14, 14, 30)
            ws.set_column(15, 15, 40)
            ws.set_column(16, 16, 40)
            ws.set_column(17, 17, 40)
            ws.set_column(18, 18, 25)
            ws.set_column(19, 19, 35)
            ws.set_column(20, 20, 40)
            ws.set_column(21, 21, 25)
            ws.set_column(22, 22, 30)
            ws.set_column(23, 23, 30)
            ws.set_column(24, 24, 25)
            ws.set_column(25, 25, 25)

            formatotitulo_filtros = workbook.add_format(
                {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

            formatoceldacab = workbook.add_format(
                {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247',
                 'font_color': 'white'})
            formatoceldaleft = workbook.add_format(
                {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

            formatoceldaleft2 = workbook.add_format(
                {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
                 'bold': 1})

            formatoceldaleft3 = workbook.add_format(
                {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})

            decimalformat = workbook.add_format(
                {'num_format': '#,##0.0', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

            decimalformat2 = workbook.add_format(
                {'num_format': '#,##0.0', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
                 'bold': 1})

            maestria = request.GET.getlist('maestria[]')
            anio =  request.GET.getlist('anio[]')

            if 'maestria' in request.GET:
                maestria = request.GET['maestria']

            if 'anio' in request.GET:
                anio = request.GET['anio']

            ws.merge_range('A1:Z1', 'Graduados Posgrado', formatotitulo_filtros)

            ws.write(1, 0, 'CODIGO_IES', formatoceldacab)
            ws.write(1, 1, 'CODIGO CARRERA', formatoceldacab)
            ws.write(1, 2, 'CIUDAD CARRERA', formatoceldacab)
            ws.write(1, 3, 'TIPO IDENTIFICACION', formatoceldacab)
            ws.write(1, 4, 'NÚMERO_CÉDULA', formatoceldacab)
            ws.write(1, 5, 'NÚMERO_PASAPORTE', formatoceldacab)
            ws.write(1, 6, 'NOMBRES', formatoceldacab)
            ws.write(1, 7, 'CARRERA', formatoceldacab)
            ws.write(1, 8, 'PERIODO', formatoceldacab)
            ws.write(1, 9, 'FECHA_INICIO_PERIODO', formatoceldacab)
            ws.write(1, 10, 'FECHA_FIN_PERIODO', formatoceldacab)
            ws.write(1, 11, 'FECHA GRADUACION', formatoceldacab)
            ws.write(1, 12, 'MECANISMO TITULACION', formatoceldacab)
            ws.write(1, 13, 'FECHA DE REFRENDACIÓN', formatoceldacab)
            ws.write(1, 14, 'MODALIDAD', formatoceldacab)
            ws.write(1, 15, 'PAIS RESIDENCIA', formatoceldacab)
            ws.write(1, 16, 'PROVINCIA RESIDENCIA', formatoceldacab)
            ws.write(1, 17, 'CANTON RESIDENCIA', formatoceldacab)
            ws.write(1, 18, 'GLBTI', formatoceldacab)
            ws.write(1, 19, 'DISCAPACIDAD', formatoceldacab)
            ws.write(1, 20, 'PORCENTAJE DISCAPACIDAD', formatoceldacab)
            ws.write(1, 21, 'NUMERO CONADIS', formatoceldacab)
            ws.write(1, 22, 'ETNIA', formatoceldacab)
            ws.write(1, 23, 'TIPO_INDÍGENA', formatoceldacab)
            ws.write(1, 24, 'EXTRANJERO', formatoceldacab)
            ws.write(1, 25, 'SEXO', formatoceldacab)

            filtro = Q(status=True, inscripcion__carrera__coordinacion__id=7, inscripcion__status=True)

            if maestria != "":
                if maestria[0] != "0":
                    filtro = filtro & Q(inscripcion__carrera__id__in=maestria)

            if anio != "":
                if anio[0] != "0":
                    filtro = filtro & Q(fechagraduado__year__in=anio)

            graduados = Graduado.objects.filter(filtro).order_by('fechagraduado')

            filas_recorridas = 3
            cont = 1
            for graduado in graduados:
                ws.write('A%s' % filas_recorridas, str(1024), formatoceldaleft)
                ws.write('B%s' % filas_recorridas, str(graduado.inscripcion.carrera.codigo if graduado.inscripcion.carrera.codigo else ''), formatoceldaleft)
                ws.write('C%s' % filas_recorridas, str('MILAGRO'), formatoceldaleft)
                ws.write('D%s' % filas_recorridas, str(graduado.inscripcion.persona.tipo_identificacion_completo()), formatoceldaleft)
                ws.write('E%s' % filas_recorridas, str(graduado.inscripcion.persona.cedula if graduado.inscripcion.persona.cedula else ''), formatoceldaleft)
                ws.write('F%s' % filas_recorridas, str(graduado.inscripcion.persona.pasaporte if graduado.inscripcion.persona.pasaporte else ''), formatoceldaleft)
                ws.write('G%s' % filas_recorridas, str(graduado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
                ws.write('H%s' % filas_recorridas, str(f'{graduado.inscripcion.carrera.nombre} CON MENCIÓN EN {graduado.inscripcion.carrera.mencion}' if graduado.inscripcion.carrera.mencion else graduado.inscripcion.carrera.nombre), formatoceldaleft)
                ws.write('I%s' % filas_recorridas, str(periodo_matricula(graduado.inscripcion) if periodo_matricula(graduado.inscripcion) is not None else ''), formatoceldaleft)
                ws.write('J%s' % filas_recorridas, str(periodo_matricula(graduado.inscripcion).inicio if periodo_matricula(graduado.inscripcion) is not None else ''), formatoceldaleft)
                ws.write('K%s' % filas_recorridas, str(periodo_matricula(graduado.inscripcion).fin if periodo_matricula(graduado.inscripcion) is not None else ''), formatoceldaleft)
                ws.write('L%s' % filas_recorridas, str(graduado.fechagraduado if graduado.fechagraduado else graduado.fecha_creacion.date()), formatoceldaleft)
                ws.write('M%s' % filas_recorridas, str(mecanismo_titulacion(graduado.inscripcion)), formatoceldaleft)
                ws.write('N%s' % filas_recorridas, str(graduado.fecharefrendacion if graduado.fecharefrendacion else ''), formatoceldaleft)
                ws.write('O%s' % filas_recorridas, str(graduado.inscripcion.carrera.get_modalidad_display()), formatoceldaleft)
                ws.write('P%s' % filas_recorridas, str(graduado.inscripcion.persona.pais.nombre if graduado.inscripcion.persona.pais else ''), formatoceldaleft)
                ws.write('Q%s' % filas_recorridas, str(graduado.inscripcion.persona.provincia.nombre if graduado.inscripcion.persona.provincia else ''), formatoceldaleft)
                ws.write('R%s' % filas_recorridas, str(graduado.inscripcion.persona.canton.nombre if graduado.inscripcion.persona.canton else ''), formatoceldaleft)
                ws.write('S%s' % filas_recorridas, str('SI' if graduado.inscripcion.persona.lgtbi else 'NO'), formatoceldaleft)
                ws.write('T%s' % filas_recorridas, str(graduado.inscripcion.persona.mi_perfil().tipodiscapacidad.nombre if graduado.inscripcion.persona.mi_perfil().tienediscapacidad and graduado.inscripcion.persona.mi_perfil().tipodiscapacidad else ''), formatoceldaleft)
                ws.write('U%s' % filas_recorridas, str(f'{graduado.inscripcion.persona.mi_perfil().porcientodiscapacidad}%' if graduado.inscripcion.persona.mi_perfil().tienediscapacidad and graduado.inscripcion.persona.mi_perfil().porcientodiscapacidad else '0%'), formatoceldaleft)
                ws.write('V%s' % filas_recorridas, str(graduado.inscripcion.persona.mi_perfil().carnetdiscapacidad if graduado.inscripcion.persona.mi_perfil().tienediscapacidad and graduado.inscripcion.persona.mi_perfil().carnetdiscapacidad else ''), formatoceldaleft)
                ws.write('W%s' % filas_recorridas, str(graduado.inscripcion.persona.mi_perfil().raza.nombre if graduado.inscripcion.persona.mi_perfil().raza else ''), formatoceldaleft)
                ws.write('X%s' % filas_recorridas, str(graduado.inscripcion.persona.mi_perfil().nacionalidadindigena.nombre if graduado.inscripcion.persona.mi_perfil().nacionalidadindigena else ''), formatoceldaleft)
                ws.write('Y%s' % filas_recorridas, str(es_extranjero(graduado.inscripcion.persona)), formatoceldaleft)
                ws.write('Z%s' % filas_recorridas, str(graduado.inscripcion.persona.sexo.nombre if graduado.inscripcion.persona.sexo else ''), formatoceldaleft)

                filas_recorridas += 1
                cont += 1

            workbook.close()
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Graduados Posgrado',
                                    destinatario=pers, url="{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            # data = {}
            # send_user_notification(user=usernotify, payload={
            #     "head": "Reporte terminado",
            #     "body": 'Graduados Posgrado',
            #     "action": "notificacion",
            #     "timestamp": time.mktime(datetime.now().timetuple()),
            #     "url": "{}reportes/matricespos/{}".format(MEDIA_URL, nombre_archivo),
            #     "mensaje": 'Los resultados del reporte Graduados Posgrado han sido generados con exito'
            # }, ttl=500)

        except Exception as ex:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.titulo = 'Error en el reporte'
                noti.cuerpo = 'Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.'.format(ex,
                                                                                                               sys.exc_info()[
                                                                                                                   -1].tb_lineno)
                noti.url = ""
                noti.error = True
                noti.save()
            else:
                noti = Notificacion(titulo='Error en el reporte',
                                    cuerpo='Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.'.format(
                                        ex, sys.exc_info()[-1].tb_lineno),
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False, error=True)
                noti.save(request)
            # data = {}
            # send_user_notification(user=usernotify, payload={
            #     "head": "Error en el reporte",
            #     "body": "Ha ocurrido un error en al generar el reporte {} - Error en la linea {}. Por favor, intente otra vez.".format(ex,
            #                                                                                              sys.exc_info()[
            #                                                                                                  -1].tb_lineno),
            #     "action": "notificacion",
            #     "timestamp": time.mktime(datetime.now().timetuple()),
            #     "url": "",
            #     "mensaje": "Ha ocurrido un error en al generar el reporte"
            # }, ttl=500)
            pass

class descarga_masiva_requisitos_pazsalvo_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            pazsalvos = data['pazsalvos']
            pers = Persona.objects.get(usuario_id=request.user.pk)
            directory_p = os.path.join(MEDIA_ROOT, 'talento_humano')
            directory = os.path.join(directory_p, 'docs_comprimidos')
            os.makedirs(directory_p, exist_ok=True)
            os.makedirs(directory, exist_ok=True)
            name_zip = generar_nombre("Requisitos de Paz y Salvo", '')+'.zip'
            url = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos', name_zip)
            fantasy_zip = zipfile.ZipFile(url, 'w')

            for idp, ps in enumerate(pazsalvos):
                certificado = ps.documento()
                historial = certificado.ultimo_archivo() if certificado else None
                fecha = ps.fecha.strftime('%d-%m-%Y')
                carpeta = f'{idp + 1}.' + unidecode(ps.persona.nombre_completo_minus()) + f' - {fecha}'
                if historial:
                    name_file = unidecode(f'Certificado de Paz y Salvo {ps.persona.usuario} {ps.id}')
                    url_file = historial.archivo.url
                    ext = url_file[url_file.rfind("."):].lower()
                    ruta_archivo_zip = os.path.join(carpeta, f'0.{name_file}{ext}')
                    fantasy_zip.write(historial.archivo.path, ruta_archivo_zip)

                for idx, r in enumerate(ps.documentos_subidos()):
                    if r.archivo:
                        name_file = unidecode(r.requisito.nombre.replace(" ", "_"))
                        url_file = r.archivo.url
                        ext = url_file[url_file.rfind("."):].lower()
                        ruta_archivo_zip = os.path.join(carpeta, f'{idx+1}.{name_file}{ext}')
                        fantasy_zip.write(r.archivo.path, ruta_archivo_zip)
            fantasy_zip.close()
            # NOTIFICACIÓN
            titulo = f'Archivo .zip de requisitos de certificados de paz y salvo generado con éxito.'
            cuerpo = f'Archivo .zip generado con éxito'
            url = f"{MEDIA_URL}talento_humano/docs_comprimidos/{name_zip}"
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='sga-sagest', fecha_hora_visible=datetime.now() + timedelta(days=2),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su archivo .zip ha sido generado con éxito.'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

class descarga_masiva_requisitos_ingreso_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        request, data, notif = self.request, self.data, self.notif
        try:
            ePersonasPeriodo = data['ePersonasPeriodo']
            pers = Persona.objects.get(usuario_id=request.user.pk)
            directory_p = os.path.join(MEDIA_ROOT, 'talento_humano')
            directory = os.path.join(directory_p, 'docs_comprimidos')
            os.makedirs(directory_p, exist_ok=True)
            os.makedirs(directory, exist_ok=True)
            name_zip = generar_nombre("Requisitos de ingreso", '')+'.zip'
            url = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos', name_zip)
            fantasy_zip = zipfile.ZipFile(url, 'w')
            idsRequisito = request.POST.getlist('requisito', [])
            for idp, ps in enumerate(ePersonasPeriodo):
                carpeta = f'{idp + 1}.' + unidecode(ps.persona.nombre_completo_minus()) + f' - {ps.periodotthh.id}'
                requisitos = ps.documentos_subidos().filter(Q(requisito_id__in=idsRequisito)).order_by('-requisito_id') if idsRequisito else ps.documentos_subidos()
                for idx, r in enumerate(requisitos):
                    if r.archivo:
                        name_file = unidecode(r.requisito.nombre.replace(" ", "_"))
                        url_file = r.archivo.url
                        ext = url_file[url_file.rfind("."):].lower()
                        ruta_archivo_zip = os.path.join(carpeta, f'{idx+1}.{name_file}{ext}')
                        fantasy_zip.write(r.archivo.path, ruta_archivo_zip)
            fantasy_zip.close()
            # NOTIFICACIÓN
            titulo = f'Archivo .zip de requisitos de ingreso generado con éxito.'
            cuerpo = f'Archivo .zip generado con éxito'
            url = f"{MEDIA_URL}talento_humano/docs_comprimidos/{name_zip}"
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='sga-sagest', fecha_hora_visible=datetime.now() + timedelta(days=2),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            send_user_notification(user=request.user, payload={
                "head": titulo,
                "body": cuerpo,
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": url,
                "btn_notificaciones": traerNotificaciones(request, data, pers),
                "mensaje": 'Su archivo .zip ha sido generado con éxito.'
            }, ttl=500)
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

class generar_acta_constatacion_background(threading.Thread):

    def __init__(self, request, data, notif):
        self.request = request
        self.data = data
        self.notif = notif
        threading.Thread.__init__(self)

    def run(self):
        from sagest.funciones import generar_acta_constatacion
        from sagest.models import ActaConstatacion
        request, data, notif = self.request, self.data, self.notif
        url = ''
        try:
            eConstatacion = data['eConstatacion']
            persona = data['persona']
            pers = Persona.objects.get(usuario_id=request.user.pk)
            pdf = generar_acta_constatacion_reportlab(request, eConstatacion)
            acta = ActaConstatacion(constatacion=eConstatacion,
                                    persona=persona,
                                    archivo=pdf)
            acta.save(request)
            log(u'Creo historial de acta de constatación: %s' % acta, request, "add")
            # NOTIFICACIÓN
            titulo = f'Acta de constatación de {eConstatacion.usuariobienes} creada exitosamente'
            cuerpo = f'Se creo correctamente el acta de constatación de {eConstatacion.usuariobienes}'
            url = f"/af_activofijo?action=constatacionesusuario&id={encrypt(eConstatacion.periodo.id)}&s={eConstatacion.usuariobienes.cedula}"
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.titulo = titulo
                noti.en_proceso = False
                noti.cuerpo = cuerpo
                noti.url = url
                noti.save()
            else:
                noti = Notificacion(cuerpo=cuerpo,
                                    titulo=titulo,
                                    destinatario=pers, url=url,
                                    prioridad=1, app_label='sga-sagest', fecha_hora_visible=datetime.now() + timedelta(days=2),
                                    tipo=2, en_proceso=False)
                noti.save(request)


        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            titulo = f'Error al generar acta de constatación de {eConstatacion.usuariobienes}'
            cuerpo = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)

        send_user_notification(user=request.user, payload={
                        "head": titulo,
                        "body": cuerpo,
                        "action": "notificacion",
                        "timestamp": time.mktime(datetime.now().timetuple()),
                        "url": url,
                        "btn_notificaciones": traerNotificaciones(request, data, pers),
                        "mensaje": 'Erro al generar acta de constatación'
                    }, ttl=500)