#!/usr/bin/env python

import os
import sys
import time

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
sys.path.append(your_djangoproject_home)
from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

import warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')
print(u"Inicio")

from helpdesk.models import *
from sagest import models
from django.db import transaction
from colorama import Back, init
from sga.models import *
from sga.funciones import variable_valor, notificacion2
from investigacion.models import *
import xlwt
import openpyxl
from xlwt import *
from settings import MEDIA_ROOT, BASE_DIR

def progress_bar(progress, total):
    percent = 100 * (progress / float(total))
    bar = (Back.GREEN+' '+Back.RESET) * int(percent) + '-' * (100 - int(percent))
    print(f"\r|{bar}| {percent:.2f}%", end="\r")


def corregir_secuencial(**kwargs):
    tables = []
    app = kwargs.pop('app', 'helpdesk') # Si no se especifica el parametro 'app' al llamar a la funcion se usa 'helpdesk'
    cursor = connections['default'].cursor()
    sql = f"SELECT array_to_string(ARRAY(SELECT TABLE_NAME FROM information_schema.tables WHERE table_schema = 'public' AND table_name like '%{app}%'), ',') AS {app}_tables;"
    cursor.execute(sql)
    results = cursor.fetchall()
    if results: tables = results[0][0].split(',')
    for table in tables:
        try:
            sql=f"SELECT pg_get_serial_sequence('{table}', 'id');"
            cursor.execute(sql)
            serial_sequence = cursor.fetchall()[0][0]
            sql=f"SELECT SETVAL('{serial_sequence}', (select max(id) from {table}));"
            cursor.execute(sql)
            #print(f"Serial sequence of <<{table}>> updated.")
        except Exception as ex:
            print(ex.__str__())


@transaction.atomic()
def migrar_registros_helpdesk():
    try:
        val = None
        data = models.HdTipoIncidente.objects.all().values().order_by('id')
        instances = [HdTipoIncidente(**item) for item in data if not HdTipoIncidente.objects.filter(pk=item['id']).exists()]
        HdTipoIncidente.objects.bulk_create(instances)

        for val in models.HdGrupo.objects.all().exclude(id__in=HdGrupo.objects.values_list('id', flat=True)).order_by('id'):
            with transaction.atomic():
                try:
                    HdGrupo.objects.create(pk=val.pk, nombre=val.nombre, descripcion=val.descripcion, tipoincidente_id=val.tipoincidente.id, fecha_creacion=val.fecha_creacion, usuario_creacion=val.usuario_creacion, fecha_modificacion=val.fecha_modificacion, usuario_modificacion=val.usuario_modificacion)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        for val in models.HdDetalle_Grupo.objects.all().exclude(id__in=HdDetalle_Grupo.objects.values_list('id', flat=True)).order_by('id'):
            with transaction.atomic():
                try:
                    HdDetalle_Grupo.objects.create(pk=val.pk, grupo_id=val.grupo.id, persona=val.persona, responsable=val.responsable, estado=val.estado, fecha_creacion=val.fecha_creacion, usuario_creacion=val.usuario_creacion, fecha_modificacion=val.fecha_modificacion, usuario_modificacion=val.usuario_modificacion)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        data = models.HdCategoria.objects.all().values().order_by('id')
        instances = [HdCategoria(**item) for item in data if not HdCategoria.objects.filter(pk=item['id']).exists()]
        HdCategoria.objects.bulk_create(instances)

        data = models.HdSubCategoria.objects.all().values().order_by('id')
        instances = [HdSubCategoria(**item) for item in data if not HdSubCategoria.objects.filter(pk=item['id']).exists()]
        HdSubCategoria.objects.bulk_create(instances)

        data = models.HdUrgencia.objects.all().values().order_by('id')
        instances = [HdUrgencia(**item) for item in data if not HdUrgencia.objects.filter(pk=item['id']).exists()]
        HdUrgencia.objects.bulk_create(instances)

        data = models.HdImpacto.objects.all().values().order_by('id')
        instances = [HdImpacto(**item) for item in data if not HdImpacto.objects.filter(pk=item['id']).exists()]
        HdImpacto.objects.bulk_create(instances)

        data = models.HdPrioridad.objects.all().values().order_by('id')
        instances = [HdPrioridad(**item) for item in data if not HdPrioridad.objects.filter(pk=item['id']).exists()]
        HdPrioridad.objects.bulk_create(instances)

        data = models.HdUrgencia_Impacto_Prioridad.objects.all().values().order_by('id')
        instances = [HdUrgencia_Impacto_Prioridad(**item) for item in data if not HdUrgencia_Impacto_Prioridad.objects.filter(pk=item['id']).exists()]
        HdUrgencia_Impacto_Prioridad.objects.bulk_create(instances)

        data = models.HdDetalle_SubCategoria.objects.all().values().order_by('id')
        instances = [HdDetalle_SubCategoria(**item) for item in data if not HdDetalle_SubCategoria.objects.filter(pk=item['id']).exists()]
        HdDetalle_SubCategoria.objects.bulk_create(instances)

        data = models.HdMedioReporte.objects.all().values().order_by('id')
        instances = [HdMedioReporte(**item) for item in data if not HdMedioReporte.objects.filter(pk=item['id']).exists()]
        HdMedioReporte.objects.bulk_create(instances)

        data = models.HdEstado.objects.all().values().order_by('id')
        instances = [HdEstado(**item) for item in data if not HdEstado.objects.filter(pk=item['id']).exists()]
        HdEstado.objects.bulk_create(instances)

        data = models.HdTipoIncidente.objects.all().values().order_by('id')
        instances = [HdTipoIncidente(**item) for item in data if not HdTipoIncidente.objects.filter(pk=item['id']).exists()]
        HdTipoIncidente.objects.bulk_create(instances)

        data = models.HdCausas.objects.all().values().order_by('id')
        instances = [HdCausas(**item) for item in data if not HdCausas.objects.filter(pk=item['id']).exists()]
        HdCausas.objects.bulk_create(instances)

        data = models.OrdenTrabajo.objects.all().values().order_by('id')
        instances = [OrdenTrabajo(**item) for item in data if not OrdenTrabajo.objects.filter(pk=item['id']).exists()]
        OrdenTrabajo.objects.bulk_create(instances)

        cont, val = 0, None
        incidentes_sagest = models.HdIncidente.objects.filter(tipoincidente__id=2).exclude(pk__in=HdIncidente.objects.values_list('id', flat=True)).order_by('id')
        for val in incidentes_sagest:
            with transaction.atomic():
                try:
                    HdIncidente.objects.create(pk=val.pk,
                                               asunto=val.asunto,
                                               persona=val.persona,
                                               departamento=val.departamento,
                                               descripcion=val.descripcion,
                                               subcategoria_id=val.subcategoria.id if val.subcategoria else None,
                                               detallesubcategoria_id=val.detallesubcategoria.id if val.detallesubcategoria else None,
                                               activo=val.activo,
                                               fechareporte=val.fechareporte,
                                               horareporte=val.horareporte,
                                               medioreporte_id=val.medioreporte.id if val.medioreporte else None,
                                               director=val.director,
                                               archivo=val.archivo,
                                               estado_id=val.estado.id if val.estado else None,
                                               tipoincidente_id=val.tipoincidente.id if val.tipoincidente else None,
                                               ubicacion=val.ubicacion,
                                               causa_id=val.causa.id if val.causa else None,
                                               responsableactivofijo=val.responsableactivofijo,
                                               realizoencuesta=val.realizoencuesta,
                                               revisionequipoexterno=val.revisionequipoexterno,
                                               revisionequiposincodigo=val.revisionequiposincodigo,
                                               serie=val.serie,
                                               ordentrabajo_id=val.ordentrabajo.id if val.ordentrabajo else None,
                                               tercerapersona=val.tercerapersona,
                                               tipousuario=val.tipousuario,
                                               concodigo=val.concodigo,
                                               activosincodigo=val.activosincodigo,
                                               fecha_creacion=val.fecha_creacion,
                                               usuario_creacion=val.usuario_creacion,
                                               usuario_modificacion=val.usuario_modificacion,
                                               fecha_modificacion=val.fecha_modificacion,
                                               status=val.status)

                    cont += 1
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        proceso_sagest = models.HdProceso.objects.all().exclude(pk__in=HdProceso.objects.values_list('id', flat=True)).order_by('id')
        for val in proceso_sagest:
            with transaction.atomic():
                try:
                    if not HdProceso.objects.filter(pk=val.pk):
                        HdProceso.objects.create(id=val.pk, nombre=val.nombre, fecha_creacion=val.fecha_creacion, usuario_creacion=val.usuario_creacion, usuario_modificacion=val.usuario_modificacion, fecha_modificacion=val.fecha_modificacion, status=val.status)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        for val in models.HdEstado_Proceso.objects.all().exclude(id__in=HdEstado_Proceso.objects.values_list('id', flat=True)).order_by('id'):
            with transaction.atomic():
                try:
                    HdEstado_Proceso.objects.create(id=val.pk, nombre=val.nombre, proceso_id=val.proceso.id, detalle=val.detalle, fecha_creacion=val.fecha_creacion, usuario_creacion=val.usuario_creacion, usuario_modificacion=val.usuario_modificacion, fecha_modificacion=val.fecha_modificacion, status=val.status)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        cont, val = 0, None
        detalle_sagest = models.HdDetalle_Incidente.objects.filter(Q(incidente__tipoincidente__id=2) | Q(incidente_id__isnull=True)).exclude(pk__in=HdDetalle_Incidente.objects.values_list('id', flat=True)).order_by('id')
        for val in detalle_sagest:
            with transaction.atomic():
                try:
                    HdDetalle_Incidente.objects.create(pk=val.pk,
                                                       incidente_id=val.incidente.id if val.incidente else None,
                                                       agente_id=val.agente.id if val.agente else None,
                                                       responsable=val.responsable,
                                                       grupo_id=val.grupo.id if val.grupo else None,
                                                       resolucion=val.resolucion,
                                                       fecharesolucion=val.fecharesolucion,
                                                       horaresolucion=val.horaresolucion,
                                                       estadoasignacion=val.estadoasignacion,
                                                       estadoproceso_id=val.estadoproceso.id if val.estadoproceso else None,
                                                       estado_id=val.estado.id if val.estado else None,
                                                       fecha_creacion=val.fecha_creacion,
                                                       usuario_creacion=val.usuario_creacion,
                                                       usuario_modificacion=val.usuario_modificacion,
                                                       fecha_modificacion=val.fecha_modificacion,
                                                       status=val.status)
                    cont += 1
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        data = models.HdDirector.objects.all().values().order_by('id')
        instances = [HdDirector(**item) for item in data if not HdDirector.objects.filter(pk=item['id']).exists()]
        HdDirector.objects.bulk_create(instances)

        corregir_secuencial()
    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex.__str__())


def migrar_tipoprofesor_sga_to_postulaciondip(**kwargs):
    try:
        from postulaciondip.models import TipoDocente, Convocatoria
        lista = list(Convocatoria.objects.values('id', 'tipodocente').all())

        with transaction.atomic():
            for tp in TipoProfesor.objects.values('id', 'nombre', 'abreviatura').all():
                if not TipoDocente.objects.values('id').filter(pk=tp['id']).exists():
                    try:
                        TipoDocente(**tp).save()
                    except Exception as ex:
                        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                        transaction.set_rollback(True)
                        print(f"ERROR: {ex.__str__()} - {linea_error}")

        with transaction.atomic():
            for cv in lista:
                convocatoria = Convocatoria.objects.get(id=cv['id'])
                if not convocatoria.tipodocente:
                    try:
                        convocatoria.tipodocente_id = cv['tipodocente']
                        convocatoria.save()
                    except Exception as ex:
                        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                        transaction.set_rollback(True)
                        print(f"ERROR: {ex.__str__()} - {linea_error}")

        corregir_secuencial(app='postulaciondip')
    except Exception as ex:
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        print(f"ERROR: {ex.__str__()} - {linea_error}")


def barrido_por_estado_director_proyecto():
    try:
        parse_funcion_criterio = lambda x: {1: 55, 2: 56, 3: 57}[x] if x in (1, 2, 3) else None
        proyectos = ProyectoInvestigacion.objects.filter(status=True, estado=37).exclude(cerrado=True)
        count = 0
        for proyecto in proyectos:
            try:
                for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(criterio__criterioinvestigacionperiodo__criterio__id=55, criterio__distributivo__profesor=proyecto.profesor, hasta__month__in=[5, 6, 7, 8, 9, 10, 11, 12], hasta__year=2023, estadoaprobacion__in=[2, 4], status=True):
                    for integrante in proyecto.integrantes_proyecto().exclude(funcion=1):
                        if distributivo := DetalleDistributivo.objects.filter(criterioinvestigacionperiodo__criterio__id=parse_funcion_criterio(integrante.funcion), distributivo__profesor=integrante.profesor, distributivo__periodo=evidencia.criterio.distributivo.periodo, status=True).first():
                            for _evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(criterio=distributivo, hasta__month=evidencia.hasta.month, hasta__year=evidencia.hasta.year, status=True):
                                if not _evidencia.estadoaprobacion == evidencia.estadoaprobacion:
                                    _evidencia.desde = evidencia.desde
                                    _evidencia.hasta = evidencia.hasta
                                    _evidencia.actividad = evidencia.actividad
                                    _evidencia.aprobado = evidencia.aprobado
                                    _evidencia.archivo = evidencia.archivo
                                    _evidencia.usuarioaprobado = evidencia.usuarioaprobado
                                    _evidencia.fechaaprobado = evidencia.fechaaprobado
                                    _evidencia.estadoaprobacion = evidencia.estadoaprobacion
                                    _evidencia.archivofirmado = evidencia.archivofirmado
                                    _evidencia.save()

                                    _evidencia.anexoevidenciaactividad_set.filter(status=True).update(status=False)
                                    for anexo in evidencia.anexoevidenciaactividad_set.filter(status=True): AnexoEvidenciaActividad(evidencia=_evidencia, observacion=anexo.observacion, archivo=anexo.archivo).save()

                                    _evidencia.evidenciaactividadaudi_set.filter(status=True).update(status=False)
                                    for anexo in evidencia.evidenciaactividadaudi_set.filter(status=True): EvidenciaActividadAudi(evidencia=_evidencia, archivo=anexo.archivo).save()

                                    if h := HistorialAprobacionEvidenciaActividad.objects.filter(evidencia=evidencia, estadoaprobacion__in=[2, 4], status=True).order_by('-id').first():
                                        HistorialAprobacionEvidenciaActividad(evidencia=_evidencia, aprobacionpersona=h.aprobacionpersona, observacion=h.observacion, fechaaprobacion=h.fechaaprobacion, estadoaprobacion=h.estadoaprobacion).save()

                                    count += 1
                                    print(f"{count}.- INFORME DEL MES {_evidencia.hasta.month} - {integrante.persona} migrado correctamente... {evidencia.estadoaprobacion}/{_evidencia.estadoaprobacion}")
            except Exception as ex:
                print(f"Error en la migración del proyecto {proyecto.pk} {proyecto}. Exception {ex.__str__()}")
    except Exception as ex:
        pass


# Eliminar evidencia de Julio de Tutor prácticas pre-profesionales (internado rotativo) - salud
def eliminar_evidencia():
    try:
        print('Eliminando registros...')
        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(criterio__criteriodocenciaperiodo__criterio__id=167, criterio__criteriodocenciaperiodo__periodo__id=177,  hasta__month__in=[7, 8], desde__month__in=[7, 8], desde__year=2023, hasta__year=2023, status=True):
            print(f"{evidencia.criterio.distributivo} - {evidencia.criterio.criteriodocenciaperiodo.criterio} eliminado...")
            evidencia.delete()
    except Exception as ex:
        pass


def barrido_por_evidencia_director_proyecto_investigacion():
    try:
        parse_funcion_criterio = lambda x: {1: 55, 2: 56, 3: 57}[x] if x in (1, 2, 3) else None
        _count0 = 0
        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(pk__in=variable_valor('ID_EVIDENCIA_PROYECTO_INVESTIGACION')):
            _count0 += 1
            try:
                if proyecto := ProyectoInvestigacion.objects.filter(profesor=evidencia.criterio.distributivo.profesor, estado__id=37, status=True).exclude(cerrado=True).first():
                    _count = 0
                    for integrante in proyecto.integrantes_proyecto().exclude(funcion=1):
                        _count += 1
                        try:
                            _evidencia = None
                            if distributivo := DetalleDistributivo.objects.filter(criterioinvestigacionperiodo__criterio__id=parse_funcion_criterio(integrante.funcion), distributivo__profesor=integrante.profesor, distributivo__periodo=evidencia.criterio.distributivo.periodo, status=True).first():
                                _evidencia = EvidenciaActividadDetalleDistributivo.objects.filter(criterio=distributivo, hasta__month=evidencia.hasta.month, hasta__year=evidencia.hasta.year, status=True).first()
                                if _evidencia:
                                    _evidencia.desde = evidencia.desde
                                    _evidencia.hasta = evidencia.hasta
                                    _evidencia.actividad = evidencia.actividad
                                    _evidencia.aprobado = evidencia.aprobado
                                    _evidencia.archivo = evidencia.archivo
                                    _evidencia.usuarioaprobado = evidencia.usuarioaprobado
                                    _evidencia.fechaaprobado = evidencia.fechaaprobado
                                    _evidencia.estadoaprobacion = evidencia.estadoaprobacion
                                    _evidencia.archivofirmado = evidencia.archivofirmado
                                else:
                                    _evidencia = EvidenciaActividadDetalleDistributivo(desde=evidencia.desde, hasta=evidencia.hasta, actividad=evidencia.actividad, aprobado=evidencia.aprobado, archivo=evidencia.archivo, usuarioaprobado=evidencia.usuarioaprobado, fechaaprobado=evidencia.fechaaprobado, estadoaprobacion=evidencia.estadoaprobacion, archivofirmado=evidencia.archivofirmado, criterio=distributivo)
                                _evidencia.save()
                                _evidencia.anexoevidenciaactividad_set.filter(status=True).delete()
                                for anexo in evidencia.anexoevidenciaactividad_set.filter(status=True): AnexoEvidenciaActividad(evidencia=_evidencia, observacion=anexo.observacion, archivo=anexo.archivo, fecha_creacion=anexo.fecha_creacion).save()
                                _evidencia.evidenciaactividadaudi_set.filter(status=True).delete()
                                for anexo in evidencia.evidenciaactividadaudi_set.filter(status=True): EvidenciaActividadAudi(evidencia=_evidencia, archivo=anexo.archivo, fecha_creacion=anexo.fecha_creacion).save()
                                _evidencia.historialaprobacionevidenciaactividad_set.filter(status=True).delete()
                                for anexo in evidencia.historialaprobacionevidenciaactividad_set.filter(status=True): HistorialAprobacionEvidenciaActividad(evidencia=_evidencia, fecha_creacion=anexo.fecha_creacion, aprobacionpersona=anexo.aprobacionpersona, observacion=anexo.observacion, fechaaprobacion=anexo.fechaaprobacion, estadoaprobacion=anexo.estadoaprobacion).save()
                                print(f"Evidencia {_count0}.- {integrante.persona} migrado correctamente. {_count}/{proyecto.integrantes_proyecto().exclude(funcion=1).values('id').count() - 1}")
                        except Exception as ex1:
                            print(f"{ex1.__str__()}")
            except Exception as ex2:
                print(f"Error en la migración de la evidencia {evidencia.pk} {evidencia}. Exception {ex2.__str__()}")
    except Exception as ex3:
        print(f"{ex3.__str__()}")


# Cambiar de estado PENDIENTE a estado APROBADA las evidencias correspondientes a los meses de abril, mayo, junio y julio de las siguientes actividades
def actualiza_estado_evidencia_criterios_varios():
    try:
        from sga.models import CriterioDocencia, CriterioInvestigacion, CriterioVinculacion

        CRITERIOS_DOCENCIA = [131, 166, 132, 19, 133, 168]
        CRITERIOS_INVESTIGACION = [59]
        CRITERIOS_VINCULACION = [4]

        # for x in CriterioDocencia.objects.filter(pk__in=CRITERIOS_DOCENCIA): print(x.nombre)
        #print(CriterioInvestigacion.objects.filter(pk__in=CRITERIOS_INVESTIGACION))


        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(Q(Q(criterio__criteriodocenciaperiodo__criterio__id__in=CRITERIOS_DOCENCIA) |
                                                                                Q(criterio__criterioinvestigacionperiodo__criterio__id__in=CRITERIOS_INVESTIGACION) |
                                                                                Q(criterio__criteriovinculacionperiodo__criterio__id__in=CRITERIOS_VINCULACION) |
                                                                                Q(criterio__criteriogestionperiodo__isnull=False)) &
                                                                              Q(estadoaprobacion__in=(1, 4),
                                                                                desde__year=2023, hasta__year=2023,
                                                                                criterio__distributivo__periodo__id=177,
                                                                                hasta__month__in=[4, 5, 6, 7, 8], status=True)):
            if not evidencia.archivofirmado:
                if criterio := evidencia.criterio.criteriodocenciaperiodo:
                    print(f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                if criterio := evidencia.criterio.criterioinvestigacionperiodo:
                    print(f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                if criterio := evidencia.criterio.criteriovinculacionperiodo:
                    print(f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                if criterio := evidencia.criterio.criteriogestionperiodo:
                    print(f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                evidencia.estadoaprobacion = 2
                evidencia.save()
    except Exception as ex:
        pass


# actualiza_estado_evidencia_criterios_varios()
# barrido_por_estado_director_proyecto()

def reporte_porcentaje_cumplimiento():
    periodo = Periodo.objects.get(pk=177)
    # profesor = Profesor.objects.get(persona__cedula='0919615906') # pk=278

    print(periodo.__str__())
    #print(profesor.__str__())

    try:
        response = []
        distributivos = ProfesorDistributivoHoras.objects.filter(periodo=periodo, activo=True, status=True)
        _count, forloop = distributivos.count(), 0
        for distributivo in distributivos:
            profesor = distributivo.profesor
            d = distributivo.profesor.informe_actividades_mensual_docente_v4(periodo, '01-08-2023', '31-08-2023', 'FACULTAD')
            fini, ffin, asignaturas = date(2023, 8, 1), date(2023, 8, 31), d.get('asignaturas')
            criteriodocenciaperiodo = d.get('criteriodocenciaperiodo')
            _data = []
            subtotal, total, numeroactividades = 0, 0, 0
            forloop += 1
            for actividad in distributivo.detalle_horas_docencia(fini, ffin):
                flag, _data = False, {}
                if actividad.criteriodocenciaperiodo:
                    if actividad.criteriodocenciaperiodo.htmldocente:
                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividaddocente':
                            if actividaddocente := actividad.criteriodocenciaperiodo.horarios_actividaddocente_profesor(profesor, fini, ffin):
                                total += 100
                                flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'impartirclase':
                        totalimpartir = actividad.criteriodocenciaperiodo.totalimparticlase(distributivo.profesor, fini, ffin, asignaturas)
                        if totalimpartir[0][0]:
                            _value = totalimpartir[0][2]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                            flag = True

                    # Evidencia Moodle
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'evidenciamoodle':
                        if listadoevidencias := actividad.criteriodocenciaperiodo.horario_evidencia_moodle(profesor, d.get('finicresta'), d.get('ffincresta')):
                            evidencia = listadoevidencias[-1]
                            if not evidencia[11] == 4:
                                _value = evidencia[10]
                                total += _value
                                if evidencia[10] < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                                flag = True

                    # Actividades Moodle
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'materialsilabo':
                        if actividadhor := actividad.criteriodocenciaperiodo.horarios_actividad_profesor(profesor, fini, ffin):
                            if actividadhor[-1][3]:
                                _value = actividadhor[-1][3]
                                total += _value
                                if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                                flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'cursonivelacion':
                        if actividadnivelacioncarrera := actividad.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(profesor, fini, ffin):
                            total += 100
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'planificarcontenido':
                        if contenidohor := actividad.criteriodocenciaperiodo.horarios_contenido_profesor(profesor, fini, ffin):
                            evidencia = contenidohor[-1]
                            _value = evidencia[3]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'tutoriaacademica':
                        if tutoriasacademicas := actividad.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(profesor, fini, ffin):
                            _value = tutoriasacademicas[0][3]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientoplataforma':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            _value = s[9]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'nivelacioncarrera':
                        if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(distributivo, fini, ffin):
                            if actigestion := actividadgestion.listadoevidencias:
                                if temp := list(filter(lambda x: x[0] == 2, actigestion)):
                                    _value = temp[0][2]
                                    total += _value
                                    if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                                    flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientotransversal':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            _value = s[9]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'apoyovicerrectorado':
                        if actividadapoyo := actividad.criteriodocenciaperiodo.horarios_apoyo_profesor(profesor, fini, ffin):
                            total += 100
                            flag = True

                if flag:
                    numeroactividades += 1
                    # print(f"DOC - {numeroactividades}.- {'%.2f' % total}  {actividad.criteriodocenciaperiodo.criterio}")
                    subtotal = total

            for actividad in distributivo.detalle_horas_investigacion():
                flag = False
                if actividad.criterioinvestigacionperiodo:
                    if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadinvestigacion':
                        if actividadgestion := actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(distributivo, fini, ffin):
                            if actividadgestion.listadoevidencias:
                                for actigestion in actividadgestion.listadoevidencias:
                                    if actigestion[0] == 2:
                                        _value = actigestion[2]
                                        total += _value
                                        if _value < 100: _data = {'tipo': 2, 'criterio': actividad.criterioinvestigacionperiodo, 'porcentaje': _value}
                                        flag = True

                if flag :
                    numeroactividades += 1
                    # print(f"INV - {numeroactividades}.- {'%.2f' % total}  {actividad.criterioinvestigacionperiodo.criterio}")
                    subtotal = total

            for actividad in distributivo.detalle_horas_gestion(fini, ffin):
                flag = False
                if actividad.criteriogestionperiodo:
                    if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadgestion':
                        if actividadgestion := actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(profesor, fini, ffin):
                            total += 100
                            flag = True

                if flag :
                    numeroactividades += 1
                    # print(f"GES - {numeroactividades}.- {'%.2f' % total}  {actividad.criteriogestionperiodo.criterio}")
                    subtotal = total

            for actividad in distributivo.detalle_horas_vinculacion():
                flag = False
                if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadvinculacion':
                    if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(distributivo, fini, ffin):
                        if actividadgestion.listadoevidencias:
                            for actigestion in actividadgestion.listadoevidencias:
                                if actigestion[0] == 2:
                                    _value = actigestion[2]
                                    total += _value
                                    if _value < 100: _data = {'tipo': 4, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value}
                                    flag = True

                if flag :
                    numeroactividades += 1
                    # print(f"VIN - {numeroactividades}.- {total}  {actividad.criteriodocenciaperiodo.criterio}")
                    subtotal = total

            _porcent = 0

            try:
                _porcent = total / numeroactividades
            except ZeroDivisionError as ex:
                pass

            response.append([profesor, _porcent, _data])
            print([profesor, _porcent, _data])
        return response
    except Exception as ex:
        print(ex.__str__())


def get_promedio():
    try:
        import calendar
        periodo = Periodo.objects.get(pk=177)
        distributivo = ProfesorDistributivoHoras.objects.filter(profesor__persona__usuario__username='psalasp', periodo=periodo, activo=True, status=True).first()
        now = datetime.now().date()
        fini, ffin = date(now.year, 8, 1), date(now.year, 8, calendar.monthrange(now.year, 8)[1])
        profesor = distributivo.profesor
        _data = []
        subtotal, total, numeroactividades = 0, 0, 0
        finicresta = fini - timedelta(days=5)
        ffincresta = ffin - timedelta(days=5)

        if distributivo.detalledistributivo_set.values('id').filter(criteriodocenciaperiodo__criterio__tipo=1, status=True).exists():
            for actividad in distributivo.detalle_horas_docencia(fini, ffin):
                if h := actividad.actividaddetalledistributivofecha(fini, ffin):
                    if not h.horas:
                        continue
                flag, html = False, actividad.criteriodocenciaperiodo.nombrehtmldocente()
                if actividad.criteriodocenciaperiodo:
                    if actividad.criteriodocenciaperiodo.htmldocente():
                        if html == 'actividaddocente':
                            if actividaddocente := actividad.criteriodocenciaperiodo.horarios_actividaddocente_profesor(profesor, fini, ffin):
                                if actividaddocente.totalplanificadas:
                                    total += 100
                                    flag = True

                    if html == 'impartirclase':
                        profesormateria = ProfesorMateria.objects.filter(profesor=profesor, materia__nivel__periodo=periodo, tipoprofesor__imparteclase=True, activo=True, materia__status=True).distinct().order_by('desde', 'materia__asignatura__nombre')
                        if periodo.clasificacion == 1:
                            asignaturas = profesormateria.filter(((Q(desde__gte=fini) & Q(hasta__lte=ffin)) | (Q(desde__lte=fini) & Q(hasta__gte=ffin)) | (Q(desde__lte=ffin) & Q(desde__gte=fini)) | (Q(hasta__gte=fini) & Q(hasta__lte=ffin)))).distinct().exclude(tipoprofesor_id=15, materia__modeloevaluativo_id__in=[27]).order_by('desde', 'materia__asignatura__nombre')
                        else:
                            asignaturas = profesormateria.filter(((Q(desde__gte=fini) & Q(hasta__lte=ffin)) | (Q(desde__lte=fini) & Q(hasta__gte=ffin)) | (Q(desde__lte=ffin) & Q(desde__gte=fini)) | (Q(hasta__gte=fini) & Q(hasta__lte=ffin)))).distinct().order_by('desde', 'materia__asignatura__nombre')

                        totalimpartir = actividad.criteriodocenciaperiodo.totalimparticlase(distributivo.profesor, fini, ffin, asignaturas)

                        if totalimpartir[0][0]:
                            _value = totalimpartir[0][2]
                            total += _value
                            if _value < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                            flag = True

                    if html == 'evidenciamoodle':
                        if listadoevidencias := actividad.criteriodocenciaperiodo.horario_evidencia_moodle(profesor, finicresta, ffincresta):
                            evidencia = listadoevidencias[-1]
                            if not evidencia[11] == 4 and evidencia[1]:
                                _value = evidencia[10]
                                total += _value
                                if evidencia[10] < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'materialsilabo':
                        if actividadhor := actividad.criteriodocenciaperiodo.horarios_actividad_profesor(profesor, fini, ffin):
                            if not actividadhor.__len__() <= 2:
                                _value = 0
                                for acti in actividadhor:
                                    if not acti[4] == 1:
                                        _value = acti[3]
                                if _value:
                                    total += _value
                                    if _value < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                    flag = True

                    if html == 'cursonivelacion':
                        if actividadnivelacioncarrera := actividad.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(profesor, fini, ffin):
                            total += 100

                            flag = True

                    if html == 'planificarcontenido':
                        if contenidohor := actividad.criteriodocenciaperiodo.horarios_contenido_profesor(profesor, fini, ffin):
                            _value = 0
                            for x in contenidohor:
                                if not x[6] == 3 and x[4]: _value = x[3]
                            if _value:
                                total += _value
                                if _value < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'tutoriaacademica':
                        if tutoriasacademicas := actividad.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(profesor, fini, ffin):
                            if tutoriasacademicas[0][1]:
                                _value = tutoriasacademicas[0][3]
                                total += _value
                                if _value < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'seguimientoplataforma':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            if s[11]:
                                _value = s[9]
                                total += _value
                                if _value < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'nivelacioncarrera':
                        if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(distributivo, fini, ffin):
                            if actividadgestion.totalplanificadas:
                                if actigestion := actividadgestion.listadoevidencias:
                                    if temp := list(filter(lambda x: x[0] == 2, actigestion)):
                                        _value = temp[0][2]
                                        total += _value
                                        if _value < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                        flag = True

                    if html == 'seguimientotransversal':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            if s[11]:
                                _value = s[9]
                                total += _value
                                if _value < 100: _data.append({'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'apoyovicerrectorado':
                        if actividadapoyo := actividad.criteriodocenciaperiodo.horarios_apoyo_profesor(profesor, fini, ffin):
                            if actividadapoyo.totalplanificadas:
                                total += 100
                                flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        if distributivo.detalledistributivo_set.values('id').filter(criterioinvestigacionperiodo__isnull=False).exists():
            for actividad in distributivo.detalle_horas_investigacion():
                flag = False
                if actividad.criterioinvestigacionperiodo:
                    if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadinvestigacion':
                        if actividadgestion := actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(distributivo, fini, ffin):
                            if actividadgestion.listadoevidencias and actividadgestion[0][1]:
                                for actigestion in actividadgestion.listadoevidencias:
                                    if actividadgestion.totalplanificadas:
                                        if actigestion[0] == 2:
                                            _value = actigestion[2]
                                            total += _value
                                            if _value < 100: _data.append({'tipo': 2, 'criterio': actividad.criterioinvestigacionperiodo, 'porcentaje': _value})
                                            flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        if distributivo.detalledistributivo_set.values('id').filter(criteriogestionperiodo__isnull=False).exists():
            for actividad in distributivo.detalle_horas_gestion(fini, ffin):
                flag = False
                if actividad.criteriogestionperiodo:
                    if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadgestion':
                        if actividadgestion := actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(profesor, fini, ffin):
                            if actividadgestion.totalplanificadas:
                                total += 100
                                flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        if distributivo.detalledistributivo_set.values('id').filter(criteriodocenciaperiodo__isnull=False, criteriodocenciaperiodo__criterio__tipo=2).exists():
            for actividad in distributivo.detalle_horas_vinculacion():
                flag = False
                if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadvinculacion':
                    if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(distributivo, fini, ffin):
                        if actividadgestion.listadoevidencias and actividadgestion.totalplanificadas:
                            for actigestion in actividadgestion.listadoevidencias:
                                if actigestion[0] == 2:
                                    _value = actigestion[2]
                                    total += _value
                                    if _value < 100: _data.append({'tipo': 4, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                    flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        _porcent = 0
        try:
            _porcent = total / numeroactividades
        except ZeroDivisionError as ex:
            pass

        return [profesor, _porcent, _data, distributivo.carrera]
    except Exception as ex:
        pass


def reporte_docentes_admision():
    coordinacion = Coordinacion.objects.get(pk=9)
    periodo = Periodo.objects.get(pk=224)
    _ids_materia = []
    for nivel in Nivel.objects.filter(nivellibrecoordinacion__coordinacion=coordinacion, periodo=periodo):
        _ids_materia += nivel.materia_set.filter(status=True).values_list('id', flat=True).distinct()

    exclude = Coordinacion.objects.filter(status=True).values_list('id', flat=True).exclude(pk=9)
    _exclude = Materia.objects.filter(nivel__nivellibrecoordinacion__coordinacion__in=exclude, nivel__periodo=periodo).values_list('id', flat=True)

    _ids = ProfesorMateria.objects.values_list('profesor', flat=True).filter(materia__in=_ids_materia, materia__nivel__periodo=periodo).exclude(materia__id__in=_exclude)
    profesores = Profesor.objects.filter(id__in=_ids, status=True)
    for i, pm in enumerate(profesores):
        print("%s %s" % (i+1, pm))


def migrar_evidencia_director_grupo_investigacion():
    from investigacion.models import GrupoInvestigacion, GrupoInvestigacionIntegrante
    from sga.models import EvidenciaActividadDetalleDistributivo
    from sga.templatetags.sga_extras import nombremes
    from django.http import HttpResponse

    CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION = 58
    MONTH = 9

    try:

        criterio = CriterioInvestigacion.objects.get(pk=CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION)
        os.makedirs(os.path.join(BASE_DIR, 'media'), exist_ok=True)
        os.makedirs(os.path.join(BASE_DIR, 'media', 'backup_arreglo_jefferson'), exist_ok=True)

        now = datetime.now()

        nombre_archivo = f'reporte_evidencias_eliminadas_{now.strftime("%Y%m%d_%H%M%S")}.xls'
        directory = os.path.join(os.path.join(BASE_DIR, 'media', 'backup_arreglo_jefferson'), nombre_archivo)

        __author__ = 'Unemi'
        fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        titulo2 = easyxf('font: name Verdana, color-index black, bold on , height 250; alignment: horiz centre;borders: left thin, right thin, top thin, bottom thin')
        encabesado_tabla = easyxf('font: name Verdana , bold on , height 150; alignment: horiz left;borders: left thin, right thin, top thin, bottom thin;pattern: pattern solid, fore_colour gray25')
        encabesado_tabla_center = easyxf('font: name Verdana , bold on , height 150; alignment: horiz centre;borders: left thin, right thin, top thin, bottom thin;pattern: pattern solid, fore_colour gray25')

        font_style = XFStyle()
        font_style.font.bold = True
        font_style2 = XFStyle()
        font_style2.font.bold = False
        wb = Workbook(encoding='utf-8')
        ws = wb.add_sheet('PORCENTAJES')
        ws.write_merge(0, 0, 0, 4, 'LISTADO DE EVIDENCIAS ELIMINADAS EN EL MES DE %s' % nombremes(MONTH).upper(), titulo2)
        ws.write_merge(1, 1, 0, 4, '%s' % criterio.nombre, titulo2)
        response = HttpResponse(content_type="application/ms-excel")
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        ws.col(0).width = 2000
        ws.col(1).width = 20500
        ws.col(2).width = 20500
        ws.col(3).width = 10000
        ws.col(4).width = 20500

        row_num = 2

        ws.write(row_num, 0, "Nº", encabesado_tabla)
        ws.write(row_num, 1, "DOCENTE", encabesado_tabla)
        ws.write(row_num, 2, "ACTIVIDAD", encabesado_tabla)
        ws.write(row_num, 3, "FECHA ELIMINACION", encabesado_tabla)
        ws.write(row_num, 4, "PERIODO", encabesado_tabla)

        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'yyyy/mm/dd'
        c = 1
        row_num += 1

        for grupo in GrupoInvestigacion.objects.filter(vigente=True, status=True):
            if director := grupo.director():
                if detalledistributivo := DetalleDistributivo.objects.filter(criterioinvestigacionperiodo__criterio__id=CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION, distributivo__profesor__persona=director.persona, distributivo__periodo_id=224, distributivo__activo=True, status=True).first():
                    if evidencia := EvidenciaActividadDetalleDistributivo.objects.filter(criterio=detalledistributivo, hasta__month=MONTH, hasta__year=2023, status=True).first():
                        evidencia.grupoinvestigacion=grupo
                        evidencia.save()

                        # Reporte eliminados
                        try:
                            for integrante in grupo.grupoinvestigacionintegrante_set.filter(status=True).exclude(funcion=1):
                                distributivo = DetalleDistributivo.objects.filter(criterioinvestigacionperiodo__criterio__id=CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION, distributivo__profesor__persona=integrante.persona, distributivo__periodo=detalledistributivo.distributivo.periodo, distributivo__activo=True, status=True).first()
                                if distributivo:

                                    # Eliminacion de evidencias pre cargadas
                                    for e in EvidenciaActividadDetalleDistributivo.objects.filter(criterio=distributivo, hasta__month=evidencia.hasta.month, hasta__year=evidencia.hasta.year, status=True):
                                        ws.write(row_num, 0, f"{c}" , fuentenormal)
                                        ws.write(row_num, 1, f"{e.criterio.distributivo.profesor.persona}", fuentenormal)
                                        ws.write(row_num, 2, f"{e.actividad}", fuentenormal)
                                        ws.write(row_num, 3, f"{now.strftime('%Y%m%d_%H%M%S')}", fuentenormal)
                                        ws.write(row_num, 4, f"{evidencia.criterio.distributivo.periodo.nombre}", fuentenormal)
                                        row_num += 1
                                        c += 1

                                        e.delete()

                                    wb.save(directory)
                                    #--------------------------------------------------

                                    _evidencia = grupo.evidenciaactividaddetalledistributivo_set.filter(criterio=distributivo, hasta__month=evidencia.hasta.month, hasta__year=evidencia.hasta.year, status=True).first()
                                    if _evidencia:
                                        _evidencia.actividaddetalledistributivo = evidencia.actividaddetalledistributivo
                                        _evidencia.desde = evidencia.desde
                                        _evidencia.hasta = evidencia.hasta
                                        _evidencia.actividad = evidencia.actividad
                                        _evidencia.aprobado = evidencia.aprobado
                                        _evidencia.archivo = evidencia.archivo
                                        _evidencia.usuarioaprobado = evidencia.usuarioaprobado
                                        _evidencia.fechaaprobado = evidencia.fechaaprobado
                                        _evidencia.estadoaprobacion = evidencia.estadoaprobacion
                                        _evidencia.archivofirmado = evidencia.archivofirmado
                                        _evidencia.grupoinvestigacion = evidencia.grupoinvestigacion
                                    else:
                                        _dict = EvidenciaActividadDetalleDistributivo.objects.filter(pk=evidencia.pk).values()[0]
                                        _dict.pop('id')
                                        _evidencia = EvidenciaActividadDetalleDistributivo(**_dict)
                                        _evidencia.criterio = distributivo

                                    _evidencia.save()

                                    _evidencia.anexoevidenciaactividad_set.filter(status=True).delete()
                                    for anexo in evidencia.anexoevidenciaactividad_set.filter(status=True):
                                        a = AnexoEvidenciaActividad(evidencia=_evidencia, observacion=anexo.observacion, archivo=anexo.archivo)
                                        a.save()

                                    _evidencia.evidenciaactividadaudi_set.filter(status=True).delete()
                                    for anexo in evidencia.evidenciaactividadaudi_set.filter(status=True):
                                        a = EvidenciaActividadAudi(evidencia=_evidencia, archivo=anexo.archivo)
                                        a.save()

                                    _evidencia.historialaprobacionevidenciaactividad_set.filter(status=True).delete()
                                    for anexo in evidencia.historialaprobacionevidenciaactividad_set.filter(status=True):
                                        model = HistorialAprobacionEvidenciaActividad(evidencia=_evidencia, aprobacionpersona=anexo.aprobacionpersona, observacion=anexo.observacion, fechaaprobacion=anexo.fechaaprobacion, estadoaprobacion=anexo.estadoaprobacion)
                                        model.save()
                                else:
                                    if DetalleDistributivo.objects.filter(distributivo__profesor__persona=integrante.persona, distributivo__periodo=detalledistributivo.distributivo.periodo, distributivo__activo=True, status=True).values('id').exists():
                                        gnro = "a" if integrante.persona.es_mujer() else "o"
                                        msj = (f"""Estimad{gnro} {integrante.persona.__str__().lower().title()}, usted se encuentra asociad{gnro} al grupo de investigación "{director.grupo.nombre}" como <b>{integrante.get_funcion_display().lower().title()}</b> pero no cuenta con el criterio<b>{detalledistributivo.criterioinvestigacionperiodo.criterio.nombre.lower().title()}</b> en su distributivo de horas.<br><br> Por favor comuníquese con su director de carrera.""")
                                        notificacion2("Problemas en el distributivo del docente", msj, integrante.persona, None, 'notificacion', CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION, 1, 'sga', CriterioInvestigacion)

                        except Exception as ex:
                            print(ex.__str__())
        print("%s" % directory)
    except Exception as ex:
        print(ex.__str__())

#  migrar_evidencia_director_grupo_investigacion()

def actualizar_fecha_fin_bitacora_diciembre_v2():
    try:
        from inno.models import InformeMensualDocente
        from sga.templatetags.sga_extras import encrypt

        libre_origen = '/actualizacion_bitacora_diciembre_2023.xls'
        fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000), (u"APELLIDOS Y NOMBRES", 6000), (u"CORREO", 6000)]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        lin = 0
        miarchivo = openpyxl.load_workbook("media/importar_permisos_loes_diciembre_2023.xlsx")
        lista = miarchivo.get_sheet_by_name('Hoja1')
        a = 0
        for filas in lista.rows:
            if f"{filas[0].value}".isdigit():
                ci = f"{filas[0].value}"
                if persona := Persona.objects.filter(cedula=ci if ci.__len__() == 10 else f"0{ci}").first():
                    if filas[4].value.day > 15:
                        for detalle in DetalleDistributivo.objects.filter(Q(criteriodocenciaperiodo__llenarbitacora=True) | Q(criterioinvestigacionperiodo__llenarbitacora=True) | Q(criteriogestionperiodo__llenarbitacora=True)).filter(distributivo__periodo=224, distributivo__profesor__persona=persona).order_by('distributivo__profesor__persona__apellido1', 'distributivo__profesor__persona__apellido2', 'distributivo__profesor__persona__nombres'):
                            if bitacora := detalle.bitacoraactividaddocente_set.filter(fechaini__month=12, fechaini__year=2023, profesor__persona=persona, status=True).first():
                                if not InformeMensualDocente.objects.filter(distributivo=detalle.distributivo, fechafin__month=12, fechafin__year=2023, estado__in=(2, 3, 4)).exists():
                                    a += 1
                                    bitacora.fechafin = filas[4].value
                                    bitacora.save()

                                    cri = ''
                                    if detalle.criteriodocenciaperiodo: cri = detalle.criteriodocenciaperiodo.criterio
                                    if detalle.criterioinvestigacionperiodo: cri = detalle.criterioinvestigacionperiodo.criterio
                                    if detalle.criteriogestionperiodo: cri = detalle.criteriogestionperiodo.criterio

                                    notificacion2("Actualización de la bitácora de actividades",
                                                  f"Estimad{'a' if persona.es_mujer() else 'o'} docente, se informa que se actualizó la fecha fin de la bitácora de diciembre 2023 para la actividad <b>{cri}</b> en función de la fecha de inicio de sus vacaciones. Por favor registrar los {bitacora.fechafin.day - 15} días faltantes.", persona, None, f'/pro_cronograma?action=detallebitacora&idbitacora={encrypt(bitacora.pk)}', 1, 1, 'sga', DetalleDistributivo)

                                    # Crea archivo excel
                                    hojadestino.write(fila, 0, "%s" % persona.cedula, fuentenormal)
                                    hojadestino.write(fila, 1, "%s" % persona, fuentenormal)
                                    hojadestino.write(fila, 2, "%s" % persona.emailinst, fuentenormal)

                                    fila += 1
                                    lin += 1
                                    print(f'{a}. Se actualizó la fecha fin {filas[4].value} en la bitácora del/a docente {persona}')
        libdestino.save(output_folder + libre_origen)
        notificacion2("Proceso finalizado", f"Se generó el archivo con los resultados en la ruta: <a href='{output_folder + libre_origen}'>{libre_origen}</a>", Persona.objects.get(cedula='0606274652'), None, '', 1, 1, 'sga', DetalleDistributivo)
        print(output_folder + libre_origen)
    except Exception as ex:
        pass


def actualizar_fecha_fin_bitacora_diciembre():
    try:
        from sagest.models import PermisoInstitucionalDetalle
        from inno.models import InformeMensualDocente

        BitacoraActividadDocente.objects.filter(fechafin__month=12, fechafin__year=2023, fechafin__day=31, status=True).update(status=False)

        count = 0
        detalle = DetalleDistributivo.objects.filter(Q(criteriodocenciaperiodo__llenarbitacora=True) | Q(criterioinvestigacionperiodo__llenarbitacora=True) | Q(criteriogestionperiodo__llenarbitacora=True)).filter(distributivo__periodo=224).order_by('distributivo__profesor__persona__apellido1', 'distributivo__profesor__persona__apellido2', 'distributivo__profesor__persona__nombres')
        for d in detalle:
            persona = d.distributivo.profesor.persona
            if permiso := PermisoInstitucionalDetalle.objects.filter(permisoinstitucional__regimenlaboral_id=2, fechainicio__month=12, fechainicio__year=2023, status=True, permisoinstitucional__status=True, permisoinstitucional__estadosolicitud__in=(3,5), permisoinstitucional__tipopermiso_id=24, permisoinstitucional__solicita=persona).first():
                if permiso.fechainicio.day > 15:
                    if bitacora := BitacoraActividadDocente.objects.filter(criterio=d, profesor__persona=persona, status=True).first():
                        if not InformeMensualDocente.objects.filter(distributivo=d.distributivo, fechafin__month=12, fechafin__year=2023, estado__in=(2, 3, 4)).exists():
                            bitacora.fechafin = permiso.fechainicio
                            bitacora.save()
                            count += 1

                            # Envío de correo
                            print(f'{count}. Se actualizó la fecha fin {permiso.fechainicio} en la bitácora del/a docente {persona}')
    except Exception as ex:
        pass


# actualizar_fecha_fin_bitacora_diciembre_v2()

def actualizar_actividad_internado_rotativo():
    from sga.models import Persona, DetalleDistributivo, CriterioDocenciaPeriodo, ClaseActividadEstado
    persona = Persona.objects.get(pk=37121)
    try:
        periodo = 306
        if not CriterioDocenciaPeriodo.objects.filter(periodo=periodo, criterio=167):
            item = CriterioDocenciaPeriodo.objects.filter(periodo=224, criterio=167).values()[0]
            item.pop('id')
            item['periodo_id'] = periodo
            CriterioDocenciaPeriodo(**item).save()
        count = 0
        criteriodocenciaperiodo = CriterioDocenciaPeriodo.objects.filter(periodo=periodo, criterio=167).first()
        listtorollback = []
        for detalledistributivo in DetalleDistributivo.objects.filter(criteriodocenciaperiodo__criterio=133, criteriodocenciaperiodo__periodo=periodo, distributivo__activo=True, status=True, distributivo__status=True)[0:1]:
            distributivo = detalledistributivo.distributivo
            distributivo.bloqueardistributivo = False
            distributivo.save()

            uncheck = False
            # Para que no les notifique el cambio de actividad
            if claseactividadestado := ClaseActividadEstado.objects.filter(profesor=distributivo.profesor, periodo=periodo, estadosolicitud=2, status=True).first():
                claseactividadestado.estadosolicitud = 1
                claseactividadestado.save()
                uncheck = True

            detalledistributivo.criteriodocenciaperiodo = criteriodocenciaperiodo
            detalledistributivo.save()
            detalledistributivo.actualiza_padre()

            if actividaddetalledistributivo := detalledistributivo.actividaddetalledistributivo_set.filter(vigente=True).first():
                actividaddetalledistributivo.nombre = f"{criteriodocenciaperiodo.criterio.nombre}".strip()
                actividaddetalledistributivo.save()

            # Cambiar la actividad planificada en el horario de actividades
            # ClaseActividad.objects.filter(detalledistributivo=detalledistributivo, activo=True, status=True)

            if uncheck:
                claseactividadestado.estadosolicitud = 2
                ClaseActividad.objects.filter(detalledistributivo__distributivo=distributivo, activo=True, status=True).update(estadosolicitud=2)
                claseactividadestado.save()

            distributivo.bloqueardistributivo = True
            distributivo.save()
            count += 1
            listtorollback.append(detalledistributivo.pk)
        notificacion2(f"Lista para revertir en caso de error {count}", "%s" % listtorollback, persona, None,'notificacion', persona.pk, 1, 'sga', Persona)
    except Exception as ex:
        linea_error = '{}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Problemas en el arreglo", f"ERROR: {ex.__str__()} - {linea_error=}", persona, None, 'notificacion', persona.pk, 1, 'sga', Persona)


def actualizar_fecha_actividades():
    from inno.models import InformeMensualDocente
    with transaction.atomic():
        try:
            if DEBUG:
                path_anexo = "media/docentes_y_tecnicos_periodo_224.xlsx"

            miarchivo = openpyxl.load_workbook(path_anexo)
            lista = miarchivo.get_sheet_by_name('OCASIONALES')
            count = 0
            nuevocriteriodocencia = 1187
            subactividad = 'APLICAR EXAMENES PARCIALES O FINALES'
            for filas in lista.rows:
                cedula, periodo, modelexample = f"{filas[0].value}", 224, None
                if cedula.isdigit():
                    porcentaje = float(filas[2].value.replace('%', ''))
                    if not porcentaje == 100:
                        if distributivo := ProfesorDistributivoHoras.objects.filter(profesor__persona__cedula=cedula, periodo=periodo, activo=True, status=True).first():
                            if not InformeMensualDocente.objects.filter(distributivo=distributivo, fechafin__year=2024, fechafin__month=1, status=True).values('id').exists():
                                distributivo.bloqueardistributivo=False
                                distributivo.save()
                                horas = 0

                                uncheck = False
                                if claseactividadestado := ClaseActividadEstado.objects.filter(profesor=distributivo.profesor, periodo=periodo, estadosolicitud=2, status=True).first():
                                    claseactividadestado.estadosolicitud = 1
                                    claseactividadestado.save()
                                    uncheck = True

                                ddistributivo = DetalleDistributivo.objects.filter(distributivo=distributivo, status=True)
                                pks, iddoc, idinv, idges = [], [], [], []

                                if doc := filas[5].value:
                                    iddoc = f"{doc}".strip().split(',')

                                if inv := filas[6].value:
                                    idinv = f"{inv}".strip().split(',')

                                if ges := filas[7].value:
                                    idges = f"{ges}".strip().split(',')

                                if d := ddistributivo.filter(Q(criteriodocenciaperiodo__in=iddoc) | Q(criteriogestionperiodo__in=idges) | Q(criterioinvestigacionperiodo__in=idinv)):
                                    pks += list(d.values_list('id', flat=True))

                                if detallesdistributivodocenciagestion := DetalleDistributivo.objects.filter(pk__in=pks): #.filter(Q(criteriodocenciaperiodo=1009) | Q(criteriogestionperiodo=1039))
                                    for dd in detallesdistributivodocenciagestion:
                                        if act := dd.actividaddetalledistributivo_set.filter(status=True, vigente=True).first():
                                            act.hasta, act.vigente = date(2023, 12, 31), False
                                            act.save()
                                            act.actualiza_padre()
                                            horas += act.horas

                                    if nuevodetalle := DetalleDistributivo.objects.filter(status=True, distributivo=distributivo, criteriodocenciaperiodo=nuevocriteriodocencia).first():
                                        nuevodetalle.horas=horas
                                    else:
                                        nuevodetalle = DetalleDistributivo(distributivo=distributivo,
                                                                           criteriodocenciaperiodo_id=nuevocriteriodocencia,
                                                                           criterioinvestigacionperiodo=None,
                                                                           criteriogestionperiodo=None,
                                                                           criteriovinculacionperiodo=None,
                                                                           horas=horas,
                                                                           ponderacion_horas=0,
                                                                           es_admision=False)
                                    nuevodetalle.save()

                                    desde, hasta = date(2024,1, 1), date(2024,1,31)
                                    if act := nuevodetalle.actividaddetalledistributivo_set.filter(status=True, vigente=True).first():
                                        act.desde=desde
                                        act.hasta=hasta
                                    else:
                                        act = ActividadDetalleDistributivo(criterio=nuevodetalle,
                                                                           nombre=nuevodetalle.criteriodocenciaperiodo.criterio.nombre,
                                                                           desde=desde,
                                                                           hasta=hasta,
                                                                           horas=horas,
                                                                           vigente=True)
                                    act.save()
                                    act.actualiza_padre()

                                    # Cambiar la actividad planificada en el horario de actividades
                                    ClaseActividad.objects.filter(detalledistributivo__in=detallesdistributivodocenciagestion.values_list('id', flat=True), activo=True, status=True).update(detalledistributivo=nuevodetalle, actividaddetallehorario=act, tipodistributivo=1)

                                    if uncheck:
                                        claseactividadestado.estadosolicitud=2
                                        ClaseActividad.objects.filter(detalledistributivo__distributivo=distributivo, activo=True, status=True).update(estadosolicitud=2)
                                        claseactividadestado.save()

                                    distributivo.bloqueardistributivo = True
                                    distributivo.save()
                                    count += 1
                                    print(f"{count}.- {distributivo.profesor.persona}")


        except Exception as ex:
            linea_error = '{}'.format(sys.exc_info()[-1].tb_lineno)
            transaction.set_rollback(True)
            print(f"ERROR: {ex.__str__()} - {linea_error=}")
            persona = Persona.objects.get(cedula='0606274652')
            notificacion2("Problemas en el arreglo", f"ERROR: {ex.__str__()} - {linea_error=}", persona, None, 'notificacion', persona.pk, 1, 'sga', Persona)


def generar_informe_cumplimiento():
    from sga.templatetags.sga_extras import listado_bitacora_docente
    try:
        fechaini, fechafin = date(2024, 1, 1), date(2024, 1, 31)
        periodo = 224
        listadoprofesormateria = ProfesorDistributivoHoras.objects.filter(carrera_id__isnull=False, status=True, periodo=periodo, activo=True).order_by('profesor__persona__apellido1', 'profesor__persona__apellido2', 'profesor__persona__nombres')
        resumen = []
        criterios = (
            (1, 'DOCENCIA'),
            (2, 'INVESTIGACIÓN'),
            (3, 'GESTIÓN'),
            (4, 'VINCULACIÓN'),
            (5, 'TOTAL'),
        )
        for distributivo in listadoprofesormateria[10:15]:
            periodo = distributivo.periodo
            profesor = distributivo.profesor
            count, count1, count2, count3, count4 = 0, 0, 0, 0, 0
            totalporcentaje, totalhdocentes, totalhinvestigacion, totalhgestion, totalhvinculacion = 0, 0, 0, 0, 0
            fechames = datetime.now().date()
            now = datetime.now()
            yearini = now.year
            year = now.year
            dayini = 1
            dia = int(now.day)
            fini, ffin = fechaini, fechafin
            fechainiresta = fini - timedelta(days=5)
            fechafinresta = ffin - timedelta(days=5)
            finicresta = fechainiresta
            ffincresta = fechafinresta

            print(f"Calculando: {profesor} {fini} - {ffin}")

            finiinicio = fechaini
            ffinal = fechafin

            adicional_lista = []
            lista_criterios = []
            listDocencia = []
            # print(f'--------------DOCENTES--------------')
            __doc, __inv, __ges, __vin = [0,0,0], [0,0,0], [0,0,0], [0,0,0]
            if horasdocencia := distributivo.detalle_horas_docencia(finiinicio, ffinal):
                dicDocencia = {'tipo': 'Horas Docencia'}
                # listDocencia = []
                listDocencia.append([0, 'ACTIVIDADES DE DOCENCIA'])
                for actividad in horasdocencia:
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente():
                        __doc[0] += actividad.horas
                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'impartirclase':
                            profesormateria = ProfesorMateria.objects.filter(profesor=distributivo.profesor, materia__nivel__periodo=periodo, tipoprofesor__imparteclase=True, activo=True, materia__status=True).distinct().order_by('desde', 'materia__asignatura__nombre')
                            if periodo.clasificacion == 1:
                                asignaturas = profesormateria.filter(((Q(desde__gte=finiinicio) & Q(hasta__lte=ffinal)) | (Q(desde__lte=finiinicio) & Q(hasta__gte=ffinal)) | (Q(desde__lte=finiinicio) & Q(desde__gte=ffinal)) | (Q(hasta__gte=finiinicio) & Q(hasta__lte=ffinal)))).distinct().exclude(tipoprofesor_id__in=[15, 5]).order_by('desde','materia__asignatura__nombre')
                            else:
                                asignaturas = profesormateria.filter(((Q(desde__gte=finiinicio) & Q(hasta__lte=ffinal)) | (Q(desde__lte=finiinicio) & Q(hasta__gte=ffinal)) | (Q(desde__lte=finiinicio) & Q(desde__gte=ffinal)) | (Q(hasta__gte=finiinicio) & Q(hasta__lte=ffinal)))).exclude(tipoprofesor_id__in=[5]).distinct().order_by('desde','materia__asignatura__nombre')
                            totalimpartir = actividad.criteriodocenciaperiodo.totalimparticlase(distributivo.profesor,finiinicio, ffinal,asignaturas, None, True)
                            if totalimpartir[2]:
                                count += 1
                                totalhdocentes += totalimpartir[1]
                            __doc[1] += totalimpartir[0]
                            __doc[2] += totalimpartir[1]

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'evidenciamoodle':
                            if not DEBUG:
                                if listadoevidencias := actividad.criteriodocenciaperiodo.horario_evidencia_moodle(distributivo.profesor, finicresta, ffincresta, True):
                                    if listadoevidencias[2]:
                                        count += 1
                                        totalhdocentes += listadoevidencias[1]
                                    #listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoevidencias[0], listadoevidencias[1]])
                                    __doc[1] += listadoevidencias[0]
                                    __doc[2] += listadoevidencias[1]

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'materialsilabo':
                            actividadhor = actividad.criteriodocenciaperiodo.horarios_actividad_profesor(distributivo.profesor, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += actividadhor[1]
                            __doc[1] += actividadhor[0]
                            __doc[2] += actividadhor[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadhor[0], actividadhor[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'cursonivelacion':
                            actividadnivelacioncarrera = actividad.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(distributivo.profesor, finiinicio, ffinal)
                            totitem4 = 0
                            if actividadnivelacioncarrera:
                                totitem4 += 100
                                totalhdocentes += 100
                                count += 1
                                __doc[1] += actividad.horas
                                __doc[2] += 100

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'planificarcontenido':
                            contenidohor = actividad.criteriodocenciaperiodo.horarios_contenido_profesor(distributivo.profesor, finiinicio, ffinal, True)
                            if contenidohor == 0:
                                listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,'-', '-'])
                            else:
                                count += 1
                                totalhdocentes += contenidohor[1]
                                __doc[1] += contenidohor[0]
                                __doc[2] += contenidohor[1]
                                #listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, contenidohor[0], contenidohor[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'tutoriaacademica':
                            tutoriasacademicas = actividad.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(distributivo.profesor, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += tutoriasacademicas[1]
                            __doc[1] += tutoriasacademicas[0]
                            __doc[2] += tutoriasacademicas[1]
                            #listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, tutoriasacademicas[0], tutoriasacademicas[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientoplataforma':
                            listadoseguimientos = actividad.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(distributivo.profesor, finiinicio, ffinal, True)
                            if listadoseguimientos[2]:
                                count += 1
                                totalhdocentes += listadoseguimientos[1]
                            __doc[1] += listadoseguimientos[0]
                            __doc[2] += listadoseguimientos[1]
                            #listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoseguimientos[0], listadoseguimientos[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'nivelacioncarrera':
                            actividadgestion = actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(distributivo, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += actividadgestion[1]
                            __doc[1] += actividadgestion[0]
                            __doc[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientotransversal':
                            listadoseguimientos = actividad.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(distributivo.profesor, finiinicio, ffinal, True)
                            if listadoseguimientos[2]:
                                count += 1
                                totalhdocentes += listadoseguimientos[1]
                            __doc[1] += listadoseguimientos[0]
                            __doc[2] += listadoseguimientos[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoseguimientos[0], listadoseguimientos[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'apoyovicerrectorado':
                            actividadapoyo = actividad.criteriodocenciaperiodo.horarios_apoyo_profesor(distributivo.profesor, finiinicio, ffinal)
                            totitem10 = 0
                            if actividadapoyo:
                                totitem10 += 100
                                totalhdocentes += 100
                                count += 1
                                __doc[1] += claseactividad['totalplanificadas']
                                __doc[2] += 100

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividaddocente':
                            actividaddocente1 = actividad.criteriodocenciaperiodo.horarios_actividaddocente_profesor(distributivo.profesor, finiinicio, ffinal, True)
                            if actividaddocente1:
                                count += 1
                                totalhdocentes += 100
                                __doc[1] += actividaddocente1[0]
                                __doc[2] += actividaddocente1[1]
                            else:
                                count += 1
                                totalhdocentes += 0
                                __doc[2] += 0

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'criterioperiodoadmision':
                            actividaddocente1 = actividad.criteriodocenciaperiodo.horario_criterio_nivelacion(distributivo.profesor, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += actividaddocente1[1]
                            __doc[1] += actividaddocente1[0]
                            __doc[2] += actividaddocente1[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividaddocente1[0], actividaddocente1[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividaddocente1 = listado_bitacora_docente(0, actividad, ffinal, True)
                            count += 1
                            totalhdocentes += actividaddocente1[1]
                            __doc[1] += actividaddocente1[0]
                            __doc[2] += actividaddocente1[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividaddocente1[0], actividaddocente1[1]])

                adicional_lista.append(listDocencia)
                lista_criterios.append({'tipocriterio': 1, 'data': listDocencia})
                listDocencia = []

            if horasinvestigacion := distributivo.detalle_horas_investigacion():
                docInvestigacion = {'tipo': 'Horas Investigación'}
                listDocencia.append([0, 'ACTIVIDADES DE INVESTIGACIÓN'])
                listInvestigacion = []
                for actividad in horasinvestigacion:
                    if actividad.criterioinvestigacionperiodo.nombrehtmldocente():
                        __inv[0] += actividad.horas
                        if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadinvestigacion':
                            actividadgestion = actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(distributivo, finiinicio, ffinal, True)
                            count1 += 1
                            totalhinvestigacion += actividadgestion[1]
                            __inv[1] += actividadgestion[0]
                            __inv[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criterioinvestigacionperiodo.id, actividad.criterioinvestigacionperiodo.criterio.nombre, actividad.horas, '', actividadgestion[1]])

                        if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                            count1 += 1
                            totalhinvestigacion += actividadgestion[1]
                            __inv[1] += actividadgestion[0]
                            __inv[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criterioinvestigacionperiodo.id, actividad.criterioinvestigacionperiodo.criterio.nombre, actividad.horas, hmes, actividadgestion[1]])
                lista_criterios.append({'tipocriterio': 2, 'data': listDocencia})
                listDocencia = []

            if horasgestion := distributivo.detalle_horas_gestion(finiinicio, ffinal):
                docGestion = {'tipo': 'Horas Gestión'}
                listGestion = []
                listDocencia.append([0, 'ACTIVIDADES DE GESTIÓN EDUCATIVA'])
                for actividad in horasgestion:
                    if actividad.criteriogestionperiodo.nombrehtmldocente():
                        __ges[0] += actividad.horas
                        if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadgestion':
                            actividadgestion = actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(distributivo.profesor, finiinicio, ffinal, True)
                            if actividadgestion:
                                count2 += 1
                                totalhgestion += 100
                                __ges[1] += actividadgestion[0]
                                __ges[2] += float(actividadgestion[1])
                                # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])
                            else:
                                count2 += 1
                                totalhgestion += 0
                                __ges[1] += 0
                                __ges[2] += 0
                                # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre, actividad.horas, '-', '0.00'])
                        if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadinformegestion':
                            actividadgestion = actividad.criteriogestionperiodo.horarios_informesgestion_profesor(distributivo, finiinicio, ffinal, True)
                            count2 += 1
                            totalhgestion += actividadgestion[1]
                            __ges[1] += actividadgestion[0]
                            __ges[2] += float(actividadgestion[1])
                            #listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre,actividad.horas, '', actividadgestion[1]])

                        if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                            count2 += 1
                            totalhgestion += actividadgestion[1]
                            __ges[1] += actividadgestion[0]
                            __ges[2] += float(actividadgestion[1])
                            # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre,actividad.horas, actividadgestion[0], actividadgestion[1]])
                lista_criterios.append({'tipocriterio': 3, 'data': listDocencia})
                listDocencia = []

            if horasvinculacion := distributivo.detalle_horas_vinculacion():
                docVinculacion = {'tipo': 'Horas Vinculacion'}
                listVinculacion = []
                listDocencia.append([0, 'ACTIVIDADES DE VINCULACIÓN CON LA SOCIEDAD'])
                for actividad in horasvinculacion:
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente():
                        __vin[0] += actividad.horas
                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadvinculacion':
                            actividadgestion = actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(distributivo, finiinicio, ffinal, True)
                            if actividadgestion:
                                count3 += 1
                                totalhvinculacion += actividadgestion[1]
                                __vin[1] += actividadgestion[0]
                                __vin[2] += actividadgestion[1]
                                # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,actividadgestion[0], actividadgestion[1]])
                            else:
                                count3 += 1
                                totalhvinculacion += 0
                                __vin[1] += 0
                                __vin[2] += 0
                                # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,'-', '0.00'])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                            count3 += 1
                            totalhvinculacion += actividadgestion[1]
                            __vin[1] += actividadgestion[0]
                            __vin[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])

                lista_criterios.append({'tipocriterio': 4, 'data': listDocencia})
                listDocencia = []

            totalporcentaje = totalhdocentes + totalhinvestigacion + totalhgestion + totalhvinculacion
            count4 = count + count1 + count2 + count3
            total_porcentaje = round(totalporcentaje / count4 if count4 else totalporcentaje, 2)
            listDocencia.append(['total', total_porcentaje])
            lista_criterios.append({'tipocriterio': 5, 'data': total_porcentaje})
            resumen.append({'profesor': profesor, 'data': lista_criterios})
        print(resumen.__str__())
    except Exception as ex:
        pass


#generar_informe_cumplimiento()

actualizar_actividad_internado_rotativo()