#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from sga.models import *
from sagest.models import *
from sga.funciones import log, salvaRubros
# from balcon.models import *
from moodle import moodle
from certi.models import *
from posgrado.models import InscripcionCohorte, AsesorComercial, HistorialAsesor, \
    EvidenciaRequisitosAspirante, RequisitosMaestria, CohorteMaestria, ConfigFinanciamientoCohorte, Contrato, DetalleEvidenciaRequisitosAspirante, \
    DetalleAprobacionContrato, DetalleAprobacionFormaPago, VentasProgramaMaestria, MaestriasAdmision, Requisito, AsesorMeta
from sga.commonviews import ficha_socioeconomica
from Moodle_Funciones import crearhtmlphpmoodle
from inno.models import TipoFormaPagoPac
from settings import MEDIA_ROOT, MEDIA_URL
import xlsxwriter
import warnings
from django.http import HttpResponse
from django.db.models.query_utils import Q
import io
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

print(u"********************Arreglo***************************")
def imprimirreporte():
    try:
        __author__ = 'Unemi'

        output = io.BytesIO()
        name_document = 'reporte_plantilla'
        nombre_archivo = name_document + "_11.xlsx"
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
        workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
        ws = workbook.add_worksheet('plantilla')
        ws.set_column(0, 0, 10)
        ws.set_column(1, 1, 15)
        ws.set_column(2, 2, 25)
        ws.set_column(3, 3, 25)
        ws.set_column(4, 4, 15)
        ws.set_column(5, 5, 25)
        ws.set_column(6, 6, 25)
        ws.set_column(7, 7, 15)
        ws.set_column(8, 8, 15)

        formatotitulo_filtros = workbook.add_format(
            {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

        formatoceldacab = workbook.add_format(
            {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247',
             'font_color': 'white'})
        formatoceldaleft = workbook.add_format(
            {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

        formatoceldaleft2 = workbook.add_format(
            {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
             'bold': 1})

        formatoceldaleft3 = workbook.add_format(
            {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})

        decimalformat = workbook.add_format(
            {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

        decimalformat2 = workbook.add_format(
            {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
             'bold': 1})

        ws.write(2, 0, 'num', formatoceldacab)
        ws.write(2, 1, 'identification', formatoceldacab)
        ws.write(2, 2, 'first_name', formatoceldacab)
        ws.write(2, 3, 'last_name', formatoceldacab)
        ws.write(2, 4, 'alias', formatoceldacab)
        ws.write(2, 5, 'role', formatoceldacab)
        ws.write(2, 6, 'gender', formatoceldacab)
        ws.write(2, 7, 'birth_date', formatoceldacab)
        ws.write(2, 8, '¿tiene foto?', formatoceldacab)

        filtro = Q(status=True)

        plantillas = DistributivoPersona.objects.filter(filtro).order_by('id')

        filas_recorridas = 4
        c = 0
        for plantilla in plantillas:
            rol = ''
            if plantilla.denominacionpuesto.id in [580, 61, 561, 50, 736, 564, 966, 819, 821, 34, 553, 973, 40, 722, 47,
                                                   735, 568, 741, 681, 540, 678, 463, 952, 954, 45, 578, 720, 945, 827,
                                                   550, 559, 739, 721, 51, 967, 832, 554, 788, 402, 378, 278, 959, 794,
                                                   752, 962, 825, 814, 732, 900, 839, 946, 841, 792, 724, 969, 757, 793,
                                                   731, 755, 686, 971, 751, 16, 958, 772, 577, 530, 899, 733, 661, 659,
                                                   669, 964, 759, 754, 750, 833, 532, 30, 471, 807, 517, 956, 710, 526,
                                                   762, 790, 761, 789, 824, 970, 691, 791, 636, 955, 826, 383, 901, 902,
                                                   965, 766, 76, 616, 713, 515, 143, 548, 740, 823, 822, 968, 169, 398,
                                                   397, 118, 120, 717, 799, 947, 648, 646, 137, 150, 805, 144, 132, 829,
                                                   838, 149, 645, 642, 649, 644, 803, 782, 650, 647, 695, 138, 699, 831,
                                                   702, 703, 804, 781, 944, 801, 800, 817, 816, 809, 780, 779, 492, 778,
                                                   719, 815, 811, 728, 718, 598, 716, 612, 611, 376, 508, 510, 776, 87,
                                                   469, 83, 599, 602, 709, 586, 711, 786, 961, 569, 725, 963, 960, 768,
                                                   573, 747, 769, 810, 743, 734, 957, 953, 565, 787, 773, 742, 840, 567,
                                                   834, 802, 72, 66, 478, 176, 535, 74, 113, 502, 972, 796, 795, 797]:
                rol = 'ADMINISTRATIVO'
            elif plantilla.denominacionpuesto.id in [101, 99, 360, 812, 359, 715, 104, 95, 400, 504, 102, 97, 190, 189,
                                                     362, 399, 466]:
                rol = 'DOCENTE'
            elif plantilla.denominacionpuesto.id in [622, 836, 623, 109, 533, 107, 619, 4, 2, 3, 111, 639, 584, 828,
                                                     582, 108, 730, 536, 624, 837, 693, 506, 638, 637, 640, 621, 618,
                                                     78, 188, 365, 547, 556, 641]:
                rol = 'TRABAJADOR'

            sexo = ''
            if plantilla.persona.sexo.id == 1:
                sexo = 'MUJER'
            elif plantilla.persona.sexo.id == 2:
                sexo = 'HOMBRE'

            ws.write('A%s' % filas_recorridas, str(plantilla.id), formatoceldaleft)
            ws.write('B%s' % filas_recorridas, str(plantilla.persona.cedula), formatoceldaleft)
            ws.write('C%s' % filas_recorridas, str(plantilla.persona.nombres), formatoceldaleft)
            ws.write('D%s' % filas_recorridas, str(plantilla.persona.apellido1 + ' ' + plantilla.persona.apellido2), formatoceldaleft)
            ws.write('E%s' % filas_recorridas, str(''), formatoceldaleft)
            ws.write('F%s' % filas_recorridas, str(rol), formatoceldaleft)
            ws.write('G%s' % filas_recorridas, str(sexo), formatoceldaleft)
            ws.write('H%s' % filas_recorridas, str(plantilla.persona.nacimiento.strftime("%d/%m/%Y")), formatoceldaleft)
            ws.write('I%s' % filas_recorridas, str('SI' if plantilla.persona.foto() else 'NO'), formatoceldaleft)

            filas_recorridas += 1
            c += 1
            print(f'{c}/{plantillas.count()} - {plantilla.persona}')

        workbook.close()
        # output.seek(0)
        # fecha_hora_actual = datetime.now().date()
        # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
        response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % name_document
        url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
        print(url_file)
    except Exception as ex:
        pass

imprimirreporte()

# print(u"********************COHORTES BACHILLERATO CIENCIAS NATURALES***************************")
# try:
#     cohortes = CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=236).exclude(pk=192)
#     maestriaadmision = MaestriasAdmision.objects.get(status=True, carrera__id=267)
#     valorexamen = 0
#     for cohorte in cohortes:
#         print(f'Cohorte a duplicar: {cohorte}')
#         periodo = None
#         if cohorte.id == 162:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE I 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE I 2023')
#         elif cohorte.id == 179:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE II 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE II 2023')
#
#         if periodo:
#             if not CohorteMaestria.objects.filter(status=True, periodoacademico=periodo).exists():
#                 programamaestria = CohorteMaestria(periodoacademico=periodo,
#                                                     coordinador=cohorte.coordinador,
#                                                     maestriaadmision=maestriaadmision,
#                                                     descripcion=cohorte.descripcion,
#                                                     modalidad=cohorte.modalidad,
#                                                     alias=cohorte.alias,
#                                                     numerochorte=cohorte.numerochorte,
#                                                     cupodisponible=cohorte.cupodisponible,
#                                                     cantidadgruposexamen=cohorte.cantidadgruposexamen,
#                                                     fechainiciocohorte=cohorte.fechainiciocohorte,
#                                                     fechafincohorte=cohorte.fechafincohorte,
#                                                     fechainicioinsp=cohorte.fechainicioinsp,
#                                                     fechafininsp=cohorte.fechafininsp,
#                                                     fechainiciorequisito=cohorte.fechainiciorequisito,
#                                                     fechafinrequisito=cohorte.fechafinrequisito,
#                                                     fechafinrequisitobeca=cohorte.fechafinrequisitobeca,
#                                                     fechainicioexamen=cohorte.fechainicioexamen,
#                                                     fechafinexamen=cohorte.fechafinexamen,
#                                                     notaminimaexa=cohorte.notaminimaexa,
#                                                     notaminimatest=cohorte.notaminimatest,
#                                                     valorexamen=valorexamen,
#                                                     valortramite=cohorte.valortramite,
#                                                     activo=cohorte.activo,
#                                                     minutosrango=cohorte.minutosrango,
#                                                     cantidadgruposentrevista=cohorte.cantidadgruposentrevista,
#                                                     totaladmitidoscohorte=cohorte.totaladmitidoscohorte)
#                 programamaestria.save()
#
#                 print(f'Nueva cohorte creada: {programamaestria}')
#
#                 if cohorte.tienecostomatricula:
#                     programamaestria.tienecostomatricula = cohorte.tienecostomatricula
#                     programamaestria.valormatricula = cohorte.valormatricula
#                 else:
#                     programamaestria.tienecostomatricula = False
#                     programamaestria.valormatricula = 0
#                 if cohorte.tienecostototal:
#                     programamaestria.tienecostototal = cohorte.tienecostototal
#                     programamaestria.valorprograma = cohorte.valorprograma
#                     # if f.cleaned_data['tipootrorubro']:
#                     #     cohorte.tiporubro_id = f.cleaned_data['tipootrorubro']
#                 else:
#                     programamaestria.tienecostototal = False
#                     programamaestria.valorprograma = 0
#                     programamaestria.tiporubro_id = None
#                 if cohorte.tiporubro:
#                     programamaestria.tiporubro_id = cohorte.tiporubro.id
#
#                 programamaestria.valorprogramacertificado = cohorte.valorprogramacertificado
#                 programamaestria.fechavencerubro = cohorte.fechavencerubro
#                 programamaestria.fechainiordinaria = cohorte.fechainiordinaria
#                 programamaestria.fechafinordinaria = cohorte.fechafinordinaria
#                 programamaestria.fechainiextraordinaria = cohorte.fechainiextraordinaria
#                 programamaestria.fechafinextraordinaria = cohorte.fechafinextraordinaria
#                 programamaestria.presupuestobeca = cohorte.presupuestobeca
#                 programamaestria.save()
#
#                 print(f'Rubro configurado')
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohorte)
#                 for requisitom in requisitos:
#                     if not RequisitosMaestria.objects.filter(cohorte=programamaestria, requisito=requisitom.requisito,
#                                                              status=True).exists():
#                         requisitomaestria = RequisitosMaestria(cohorte=programamaestria,
#                                                                requisito=requisitom.requisito)
#                         requisitomaestria.save()
#
#                 print(f'Requisitos configurado')
#
#                 tiposfinancia = ConfigFinanciamientoCohorte.objects.filter(status=True, cohorte=cohorte)
#                 for tipo in tiposfinancia:
#                     instance = ConfigFinanciamientoCohorte(descripcion=tipo.descripcion,
#                                                             cohorte_id=programamaestria.id,
#                                                             valormatricula=tipo.valormatricula,
#                                                             valorarancel=tipo.valorarancel,
#                                                             valortotalprograma=tipo.valortotalprograma,
#                                                             porcentajeminpagomatricula=tipo.porcentajeminpagomatricula,
#                                                             maxnumcuota=tipo.maxnumcuota,
#                                                             fecha=tipo.fecha)
#                     instance.save()
#
#                 print(f'Tipos de financiamiento creados')
#
#                 asesores = AsesorMeta.objects.filter(status=True, cohorte=cohorte)
#
#                 for asesorm in asesores:
#                     if not AsesorMeta.objects.filter(status=True, cohorte=cohorte, asesor=asesorm.asesor).exists():
#                         asesormeta = AsesorMeta(asesor=asesorm.asesor,
#                                                 cohorte=programamaestria,
#                                                 fecha_inicio_meta=programamaestria.fechainicioinsp,
#                                                 fecha_fin_meta=programamaestria.fechafininsp,
#                                                 meta=0)
#                         asesormeta.save()
#
#                 print(f'Asesor Meta creado')
#
#     print(u"********************COHORTES BACHILLERATO CIENCIAS SOCIALES***************************")
#     cohortes = CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=236).exclude(pk=192)
#     maestriaadmision = MaestriasAdmision.objects.get(status=True, carrera__id=268)
#     valorexamen = 0
#     for cohorte in cohortes:
#         print(f'Cohorte a duplicar: {cohorte}')
#         periodo = None
#         if cohorte.id == 162:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE I 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE I 2023')
#         elif cohorte.id == 179:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE II 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE II 2023')
#
#         if periodo:
#             if not CohorteMaestria.objects.filter(status=True, periodoacademico=periodo).exists():
#                 programamaestria = CohorteMaestria(periodoacademico=periodo,
#                                                     coordinador=cohorte.coordinador,
#                                                     maestriaadmision=maestriaadmision,
#                                                     descripcion=cohorte.descripcion,
#                                                     modalidad=cohorte.modalidad,
#                                                     alias=cohorte.alias,
#                                                     numerochorte=cohorte.numerochorte,
#                                                     cupodisponible=cohorte.cupodisponible,
#                                                     cantidadgruposexamen=cohorte.cantidadgruposexamen,
#                                                     fechainiciocohorte=cohorte.fechainiciocohorte,
#                                                     fechafincohorte=cohorte.fechafincohorte,
#                                                     fechainicioinsp=cohorte.fechainicioinsp,
#                                                     fechafininsp=cohorte.fechafininsp,
#                                                     fechainiciorequisito=cohorte.fechainiciorequisito,
#                                                     fechafinrequisito=cohorte.fechafinrequisito,
#                                                     fechafinrequisitobeca=cohorte.fechafinrequisitobeca,
#                                                     fechainicioexamen=cohorte.fechainicioexamen,
#                                                     fechafinexamen=cohorte.fechafinexamen,
#                                                     notaminimaexa=cohorte.notaminimaexa,
#                                                     notaminimatest=cohorte.notaminimatest,
#                                                     valorexamen=valorexamen,
#                                                     valortramite=cohorte.valortramite,
#                                                     activo=cohorte.activo,
#                                                     minutosrango=cohorte.minutosrango,
#                                                     cantidadgruposentrevista=cohorte.cantidadgruposentrevista,
#                                                     totaladmitidoscohorte=cohorte.totaladmitidoscohorte)
#                 programamaestria.save()
#
#                 print(f'Nueva cohorte creada: {programamaestria}')
#
#                 if cohorte.tienecostomatricula:
#                     programamaestria.tienecostomatricula = cohorte.tienecostomatricula
#                     programamaestria.valormatricula = cohorte.valormatricula
#                 else:
#                     programamaestria.tienecostomatricula = False
#                     programamaestria.valormatricula = 0
#                 if cohorte.tienecostototal:
#                     programamaestria.tienecostototal = cohorte.tienecostototal
#                     programamaestria.valorprograma = cohorte.valorprograma
#                     # if f.cleaned_data['tipootrorubro']:
#                     #     cohorte.tiporubro_id = f.cleaned_data['tipootrorubro']
#                 else:
#                     programamaestria.tienecostototal = False
#                     programamaestria.valorprograma = 0
#                     programamaestria.tiporubro_id = None
#                 if cohorte.tiporubro:
#                     programamaestria.tiporubro_id = cohorte.tiporubro.id
#
#                 programamaestria.valorprogramacertificado = cohorte.valorprogramacertificado
#                 programamaestria.fechavencerubro = cohorte.fechavencerubro
#                 programamaestria.fechainiordinaria = cohorte.fechainiordinaria
#                 programamaestria.fechafinordinaria = cohorte.fechafinordinaria
#                 programamaestria.fechainiextraordinaria = cohorte.fechainiextraordinaria
#                 programamaestria.fechafinextraordinaria = cohorte.fechafinextraordinaria
#                 programamaestria.presupuestobeca = cohorte.presupuestobeca
#                 programamaestria.save()
#
#                 print(f'Rubro configurado')
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohorte)
#                 for requisitom in requisitos:
#                     if not RequisitosMaestria.objects.filter(cohorte=programamaestria, requisito=requisitom.requisito,
#                                                              status=True).exists():
#                         requisitomaestria = RequisitosMaestria(cohorte=programamaestria,
#                                                                requisito=requisitom.requisito)
#                         requisitomaestria.save()
#
#                 print(f'Requisitos configurado')
#
#                 tiposfinancia = ConfigFinanciamientoCohorte.objects.filter(status=True, cohorte=cohorte)
#                 for tipo in tiposfinancia:
#                     instance = ConfigFinanciamientoCohorte(descripcion=tipo.descripcion,
#                                                             cohorte_id=programamaestria.id,
#                                                             valormatricula=tipo.valormatricula,
#                                                             valorarancel=tipo.valorarancel,
#                                                             valortotalprograma=tipo.valortotalprograma,
#                                                             porcentajeminpagomatricula=tipo.porcentajeminpagomatricula,
#                                                             maxnumcuota=tipo.maxnumcuota,
#                                                             fecha=tipo.fecha)
#                     instance.save()
#
#                 print(f'Tipos de financiamiento creados')
#
#                 asesores = AsesorMeta.objects.filter(status=True, cohorte=cohorte)
#
#                 for asesorm in asesores:
#                     if not AsesorMeta.objects.filter(status=True, cohorte=cohorte, asesor=asesorm.asesor).exists():
#                         asesormeta = AsesorMeta(asesor=asesorm.asesor,
#                                                 cohorte=programamaestria,
#                                                 fecha_inicio_meta=programamaestria.fechainicioinsp,
#                                                 fecha_fin_meta=programamaestria.fechafininsp,
#                                                 meta=0)
#                         asesormeta.save()
#
#                 print(f'Asesor Meta creado')
# except Exception as ex:
#     pass

# print(u"UNIDADES CERTIFICADORAS")
# try:
#     certificados = Certificado.objects.filter(status=True, visible=True, servicio__isnull=False).exclude(tipo_origen=1)
#     responsable = Persona.objects.get(pk=127492)
#     for certificado in certificados:
#         if CertificadoUnidadCertificadora.objects.filter(certificado=certificado, responsable__id=29110).exists():
#             unidad = CertificadoUnidadCertificadora.objects.get(certificado=certificado, responsable__id=29110)
#             unidad.responsable = responsable
#             unidad.responsable_titulo = 'Abg. Edison Sempertegui Henriquez'
#             unidad.responsable_denominacion = 'Secretario General (S)'
#             unidad.save()
#             print(f'{certificado} - {certificado.codigo} - Responsable: {unidad.responsable} - Titulo: {unidad.responsable_titulo} - Cargo: {unidad.responsable_denominacion}')
# except Exception as ex:
#     pass
# try:
#     cohortes = CohorteMaestria.objects.filter(status=True, procesoabierto=False, maestriaadmision__carrera__id=215).exclude(pk=194)
#     for cohorte in cohortes:
#         if CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=266, descripcion=cohorte.descripcion, procesoabierto=False).exists():
#             cohortenueva = CohorteMaestria.objects.get(status=True, maestriaadmision__carrera__id=266, descripcion=cohorte.descripcion, procesoabierto=False)
#             postulantes = InscripcionCohorte.objects.filter(status=True, cohortes=cohorte, id__in=[]).exclude(cohortes__id__in=[194, 205]).order_by('id')
#             c = 0
#             for postulante in postulantes:
#                 postulante.cohortes = cohortenueva
#                 postulante.save()
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#                 for reqma in requisitos:
#                     if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                         evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                         evi.requisitos = reqma
#                         evi.save()
#
#                 if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#                     rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#                     for rubro in rubros:
#                         rubro.cohortemaestria = cohortenueva
#                         rubro.tipo = cohorte.tiporubro
#                         rubro.save()
#                 if postulante.Configfinanciamientocohorte:
#                     if ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                         finan = ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                         postulante.Configfinanciamientocohorte = finan
#                         postulante.save()
#                 c += 1
#                 print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
# except Exception as ex:
#     pass
# try:
#     c = 0
#     cohortenueva = CohorteMaestria.objects.get(status=True, pk=206)
#     postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=192, itinerario=1).order_by('id')
#     for postulante in postulantes:
#         postulante.cohortes = cohortenueva
#         postulante.save()
#
#         requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#         for reqma in requisitos:
#             if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#         if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#             rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#             tipo = TipoOtroRubro.objects.get(pk=3566)
#             for rubro in rubros:
#                 rubro.cohortemaestria = cohortenueva
#                 rubro.tipo = tipo
#                 rubro.save()
#         if postulante.Configfinanciamientocohorte:
#             if ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                 finan = ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                 postulante.Configfinanciamientocohorte = finan
#                 postulante.save()
#         c += 1
#         print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
#
#     print('*************************+++CIENCIAS SOCIALES - ACTIVOS*************************')
#     c = 0
#     cohortenueva = CohorteMaestria.objects.get(status=True, pk=207)
#     postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=192, itinerario=3).order_by('id')
#     for postulante in postulantes:
#         postulante.cohortes = cohortenueva
#         postulante.save()
#
#         requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#         for reqma in requisitos:
#             if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#         if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#             rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#             tipo = TipoOtroRubro.objects.get(pk=3565)
#             for rubro in rubros:
#                 rubro.cohortemaestria = cohortenueva
#                 rubro.tipo = tipo
#                 rubro.save()
#         if postulante.Configfinanciamientocohorte:
#             if ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                 finan = ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                 postulante.Configfinanciamientocohorte = finan
#                 postulante.save()
#         c += 1
#         print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
# except Exception as ex:
#     pass
# def promedio_re(inscripcion):
#     return round(null_to_numeric(
#         inscripcion.recordacademico_set.filter(validapromedio=True, aprobada=True, status=True, asignaturamalla__isnull=False).exclude(asignaturamalla__nivelmalla=inscripcion.mi_nivel().nivel).aggregate(
#             promedio=Avg('nota'))['promedio']), 2)
#
#
# def imprimirreporte():
#     try:
#         estado = ''
#
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_mejores_estudiantes_facultad'
#         nombre_archivo = name_document + "_11.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 35)
#         ws.set_column(3, 3, 45)
#         ws.set_column(4, 4, 25)
#         ws.set_column(5, 5, 15)
#         ws.set_column(6, 6, 45)
#         ws.set_column(7, 7, 15)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '#,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Idmat', formatoceldacab)
#         ws.write(2, 2, 'Coordinacion', formatoceldacab)
#         ws.write(2, 3, 'Carrera', formatoceldacab)
#         ws.write(2, 4, 'Nivel', formatoceldacab)
#         ws.write(2, 5, 'Cédula', formatoceldacab)
#         ws.write(2, 6, 'Maestrante', formatoceldacab)
#         ws.write(2, 7, 'Promedio', formatoceldacab)
#
#         filas_recorridas = 4
#         cont = 1
#         conta = 1
#
#         diccionario_promedios = {}
#         matriculados = Matricula.objects.filter(status=True, inscripcion__status=True,
#                                                 inscripcion__carrera__status=True, retiradomatricula=False,
#                                                 nivel__periodo__id=224, nivelmalla__id__gte=4, inscripcion__carrera__coordinacion__id__in=[5]).distinct()
#         for matriculado in matriculados:
#             canasignaturasnivel = AsignaturaMalla.objects.filter(status=True, malla=matriculado.inscripcion.malla_inscripcion().malla,
#                                                          nivelmalla=matriculado.inscripcion.mi_nivel().nivel).distinct().count()
#
#             idasignaturasnivel = AsignaturaMalla.objects.filter(status=True, malla=matriculado.inscripcion.malla_inscripcion().malla,
#                                                          nivelmalla=matriculado.inscripcion.mi_nivel().nivel).values_list('id', flat=True)
#
#             canmateriasnivel = MateriaAsignada.objects.filter(status=True, matricula=matriculado,
#                                                                  materia__asignaturamalla__id__in=idasignaturasnivel).distinct().count()
#
#             if not HistoricoRecordAcademico.objects.filter(status=True, inscripcion=matriculado.inscripcion, aprobada=False).exists():
#                 if canasignaturasnivel == canmateriasnivel:
#                     nombre = matriculado.id
#                     numero = promedio_re(matriculado.inscripcion)
#                     diccionario_promedios[nombre] = numero
#                 else:
#                     print(f'{conta} / {matriculados.count()}')
#                     conta += 1
#             else:
#                 print(f'{conta} / {matriculados.count()}')
#                 conta += 1
#         diccionario_ordenado = dict(sorted(diccionario_promedios.items(), key=lambda item: item[1], reverse=True))
#         lista = []
#         for nombre in diccionario_ordenado.keys():
#             if len(lista) <= 5:
#                 lista.append(nombre)
#
#         for lis in lista:
#             matriculado = Matricula.objects.get(status=True, pk=lis)
#             ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(matriculado.id), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(matriculado.inscripcion.carrera.coordinacion_carrera().nombre), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(matriculado.inscripcion.carrera), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(matriculado.inscripcion.mi_nivel().nivel), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(matriculado.inscripcion.persona.cedula), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(matriculado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, promedio_re(matriculado.inscripcion), decimalformat)
#
#             filas_recorridas += 1
#             print(f'{cont} / {len(lista)}')
#             cont += 1
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()

# def imprimirreporte():
#     try:
#         estado = ''
#
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_mariela'
#         nombre_archivo = name_document + "_6.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 15)
#         ws.set_column(3, 3, 30)
#         ws.set_column(4, 4, 30)
#         ws.set_column(5, 5, 40)
#         ws.set_column(6, 6, 30)
#         ws.set_column(7, 7, 30)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Id', formatoceldacab)
#         ws.write(2, 2, 'Cédula', formatoceldacab)
#         ws.write(2, 3, 'Postulante', formatoceldacab)
#         ws.write(2, 4, 'Maestría', formatoceldacab)
#         ws.write(2, 5, 'Cohorte', formatoceldacab)
#         ws.write(2, 6, 'Fecha de postulación', formatoceldacab)
#         ws.write(2, 7, 'Estado', formatoceldacab)
#         ws.write(2, 8, '¿Activo?', formatoceldacab)
#
#         postulantes = InscripcionCohorte.objects.filter(inscripcionaspirante__persona__cedula__in=[''])
#
#         filas_recorridas = 4
#         cont = 1
#
#         for postulante in postulantes:
#             if postulante.tiene_matricula_cohorte():
#                 if postulante.retirado_matricula():
#                     estado = 'Retirado'
#                 else:
#                     estado = 'Matriculado'
#             elif postulante.inscripcion:
#                 estado = 'Inscrito'
#             elif postulante.estado_aprobador == 1:
#                 estado = 'En proceso'
#             elif postulante.estado_aprobador == 2:
#                 estado = 'Admitido'
#             elif postulante.estado_aprobador == 3:
#                 estado = 'Rechazado'
#
#             ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(postulante.id), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.identificacion()), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.nombre_completo_inverso()), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(postulante.cohortes.maestriaadmision.descripcion), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(postulante.cohortes.descripcion), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(postulante.fecha_creacion.date()), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, str(estado), formatoceldaleft)
#             ws.write('I%s' % filas_recorridas, str('Si' if postulante.status else 'No'), formatoceldaleft)
#
#             filas_recorridas += 1
#             cont += 1
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()
    # print("Migracion de leads de asesores antiguos Merino-Cruz")
    # antiguo = AsesorComercial.objects.get(status=True, pk=35)
    # nuevo = AsesorComercial.objects.get(status=True, pk=38)
    # idventas = VentasProgramaMaestria.objects.filter(status=True, asesor=antiguo).values_list('inscripcioncohorte__id', flat=True)
    # leads = InscripcionCohorte.objects.filter(status=True, asesor=antiguo, vendido=False).exclude(id__in=idventas)
    # c=0
    # for lead in leads:
    #     lead.asesor = nuevo
    #     lead.tiporespuesta = None
    #     lead.save()
    #
    #     histoanti = HistorialAsesor.objects.get(inscripcion_id=lead.id, fecha_fin=None)
    #     histoanti.fecha_fin = lead.fecha_modificacion
    #     histoanti.save()
    #     histo = HistorialAsesor(inscripcion_id=lead.id, fecha_inicio=lead.fecha_modificacion,
    #                             fecha_fin=None, asesor=lead.asesor, observacion='MASIVO - LEAD DE ASESOR INACTIVO')
    #     histo.save()
    #     c += 1
    #     print(f'{c} - Lead: {lead} - asesor {lead.asesor}')
    #
    # print("Migracion de leads de asesores antiguos Morocho-Hurtado")
    # antiguo = AsesorComercial.objects.get(status=True, pk=36)
    # nuevo = AsesorComercial.objects.get(status=True, pk=1)
    # idventas = VentasProgramaMaestria.objects.filter(status=True, asesor=antiguo).values_list('inscripcioncohorte__id', flat=True)
    # leads = InscripcionCohorte.objects.filter(status=True, asesor=antiguo, vendido=False).exclude(id__in=idventas)
    # c=0
    # for lead in leads:
    #     lead.asesor = nuevo
    #     lead.tiporespuesta = None
    #     lead.save()
    #
    #     histoanti = HistorialAsesor.objects.get(inscripcion_id=lead.id, fecha_fin=None)
    #     histoanti.fecha_fin = lead.fecha_modificacion
    #     histoanti.save()
    #     histo = HistorialAsesor(inscripcion_id=lead.id, fecha_inicio=lead.fecha_modificacion,
    #                             fecha_fin=None, asesor=lead.asesor, observacion='MASIVO - LEAD DE ASESOR INACTIVO')
    #     histo.save()
    #     c += 1
    #     print(f'{c} - Lead: {lead} - asesor {lead.asesor}')
    #
#     print("Migracion de leads de contabilidad")
#
#     cohorte2023 = CohorteMaestria.objects.get(id=173, status=True)
#     cont = 0
#     matriculados = []
#     fvence = datetime.strptime('2023-08-20', '%Y-%m-%d').date()
#
#     postu = InscripcionCohorte.objects.filter(status=True, cohortes__id=173)
#     for pos in postu:
#         if pos.tiene_matricula_cohorte():
#             matriculados.append(pos.id)
#
#     postulantes = InscripcionCohorte.objects.filter(estado_aprobador__in=[1, 2], cohortes__id=173, id__in=[40927,
# 24432,
# 52095,
# 28902,
# 44155,
# 53998,
# 62225,
# 29845,
# 33277,
# 55254,
# 51213,
# 52569,
# 53130,
# 55278,
# 58843,
# 54975,
# 45196,
# 56783,
# 24831,
# 58856,
# 43843,
# 61242,
# 50707,
# 21696,
# 48453,
# 55023,
# 23475,
# 55297,
# 22098,
# 39161,
# 21843,
# 55374,
# 62271,
# 56323,
# 56322,
# 22043,
# 22069,
# 49510,
# 22097,
# 52407,
# 60995,
# 21874,
# 22260,
# 22656,
# 22935,
# 50565,
# 22732,
# 33604,
# 21971,
# 22463,
# 22975,
# 22473,
# 31704,
# 50571,
# 56367,
# 63819,
# 34027,
# 31839,
# 63822,
# 39237,
# 33536,
# 32495,
# 52182,
# 22918,
# 64368,
# 36840,
# 36325,
# 36337,
# 36700,
# 32694,
# 36788,
# 52426,
# 49113,
# 52189,
# 46922,
# 49624,
# 49630,
# 56976,
# 37404,
# 35767,
# 21915,
# 37235,
# 27694,
# 43571,
# 52445,
# 58839,
# 31186,
# 45391,
# 22902,
# 21925,
# 34522,
# 57336,
# 57317,
# 39428,
# 21934,
# 26408,
# 35302,
# 37364,
# 37681,
# 49519,
# 56366,
# 33409,
# 26772,
# 38272,
# 56287,
# 22208,
# 22766,
# 52273,
# 33513,
# 53683,
# 22756,
# 45797,
# 36062,
# 53739,
# 42727,
# 23373,
# 27733,
# 34302,
# 34347,
# 28246,
# 34455,
# 44652,
# 49643,
# 57179,
# 21942,
# 44226,
# 22247,
# 27352,
# 21784,
# 29078,
# 21919,
# 61918,
# 24208,
# 37198,
# 40664,
# 42420,
# 21717,
# 31182,
# 55750,
# 44331,
# 43175,
# 44888,
# 34657,
# 31835,
# 56237,
# 52715,
# 58680,
# 32476,
# 56303,
# 32522,
# 55936,
# 34943,
# 34963,
# 58737,
# 41219,
# 50232,
# 60960,
# 36558,
# 52097,
# 52135,
# 52155,
# 52303,
# 53085,
# 54211,
# 54227,
# 54748,
# 54756,
# 54886,
# 54926,
# 55009,
# 55010,
# 55027,
# 55143,
# 39845,
# 56263,
# 56500,
# 56544,
# 42123,
# 34795,
# 53983,
# 44894,
# 48387,
# 55506,
# 58938,
# 58914,
# 58960,
# 60840,
# 60849,
# 61410,
# 61778,
# 35733,
# 53981,
# 54717,
# 56167,
# 59881,
# 54950,
# 55144,
# 44859,
# 55564,
# 61680,
# 59349,
# 58736,
# 59973,
# 58263,
# 44508,
# 60004,
# 46950,
# 60713,
# 53289,
# 57987,
# 60576,
# 62924,
# 60973,
# 36535,
# 62968,
# 63414,
# 63029,
# 48878,
# 40012,
# 49201,
# 63091,
# 57145,
# 40218,
# 37678,
# 61542,
# 50143,
# 50275,
# 50298,
# 50308,
# 50452,
# 50689,
# 50842,
# 51237,
# 52479,
# 52894,
# 64948,
# 54313,
# 54653,
# 54738,
# 44301,
# 65071,
# 55081,
# 55087,
# 65335,
# 65596,
# 57140,
# 59511,
# 61080,
# 61545,
# 64832,
# 65836,
# 65856,
# 22145,
# 22156,
# 22196,
# 22968,
# 24269,
# 24459,
# 25996,
# 26065,
# 26729,
# 26883,
# 27410,
# 27416,
# 38066,
# 50436,
# 52616,
# 53014,
# 55864,
# 55866,
# 63617,
# 63355,
# 21650,
# 21687,
# 21688,
# 21710,
# 21728,
# 21792,
# 21812,
# 21821,
# 21831,
# 21832,
# 21859,
# 21867,
# 21869,
# 21873,
# 21883,
# 21900,
# 21902,
# 21912,
# 21913,
# 21918,
# 21929,
# 21931,
# 21939,
# 21947,
# 21964,
# 21969,
# 21981,
# 21985,
# 21987,
# 22013,
# 22036,
# 22122,
# 22126,
# 22238,
# 22269,
# 22283,
# 22356,
# 22413,
# 22607,
# 22710,
# 22712,
# 23021,
# 23025,
# 23029,
# 23250,
# 23379,
# 23452,
# 23478,
# 23580,
# 23835,
# 24054,
# 24484,
# 24921,
# 25030,
# 24121,
# 25195,
# 25461,
# 25503,
# 27082,
# 27087,
# 28115,
# 28518,
# 28910,
# 29052,
# 29079,
# 29136,
# 29323,
# 29522,
# 29714,
# 31195,
# 31214,
# 31661,
# 31694,
# 32183,
# 32195,
# 32304,
# 32456,
# 32519,
# 32773,
# 32786,
# 32830,
# 32841,
# 33289,
# 33643,
# 34528,
# 35359,
# 36015,
# 36857,
# 37472,
# 38106,
# 38912,
# 39104,
# 39116,
# 39163,
# 39173,
# 39968,
# 40100,
# 40220,
# 40924,
# 41643,
# 41864,
# 42681,
# 44682,
# 44791,
# 45330,
# 45393,
# 45711,
# 45978,
# 46192,
# 47452,
# 47502,
# 47589,
# 47600,
# 48078,
# 56611,
# 56662,
# 56682,
# 56776,
# 56777,
# 57235,
# 58023,
# 58032,
# 58132,
# 58424,
# 58639,
# 58657,
# 59126,
# 59261,
# 59332,
# 59345,
# 59410,
# 59774,
# 59776,
# 59779,
# 60170,
# 60309,
# 60391,
# 60592,
# 61878,
# 61947,
# 62324,
# 62409,
# 62410,
# 63001,
# 63102,
# 63233,
# 63240,
# 63415,
# 63477,
# 63674,
# 63882,
# 63901,
# 64038,
# 64056,
# 64327,
# 64400,
# 65195,
# 65986,
# 66020]).exclude(id__in=matriculados)
#     conta = 0
#     contado = 0
#     for postulante in postulantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
#         for lis in listarequisitos:
#             #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
#             if lis in [970]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1375)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#                     deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
#                     if deta.fecha_aprobacion.date() < fvence:
#                         deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.estado_aprobacion = 3
#                         deta.save()
#             #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
#             if lis in [967]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1372)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #COPIA A COLOR DE CEDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
#             if lis in [968]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1373)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CERTIFICADO DE EXPERIENCIA PROFESIONAL
#             if lis in [966]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1371)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#             #HOJA DE VIDA.
#             if lis in [969]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1374)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#         evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[1375, 1372, 1374, 1371, 1373])
#         if not evidema.count() >= 5:
#             postulante.estado_aprobador = 1
#             postulante.save()
#             conta += 1
#         else:
#             if evidema.count() >= 5 and postulante.estado_aprobador == 1:
#                 postulante.estado_aprobador = 2
#                 postulante.save()
#             contado += 1
#
#         if postulante.formapagopac and postulante.formapagopac.id == 2:
#             listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
#             for lis in listarequisitosfi:
#                 if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
#                     requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                               cohorte_id=cohorte2023,
#                                                               status=True)[0]
#                     lis.requisitos = requi
#                     lis.save()
#
#                     if lis.requisitos.requisito.id in [54, 62]:
#                         deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).order_by('-id').first()
#                         if deta.fecha_aprobacion.date() < fvence:
#                             deta.estado_aprobacion = 3
#                             deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.save()
#
#                         if DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).count() > 1:
#                             detalles = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).exclude(id=deta.id)
#                             for det in detalles:
#                                 det.status = False
#                                 det.save()
#
#         if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(status=True, id=crubro.id)
#                 rubro.status = False
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (rubro.idrubroepunemi)
#                     cursor.execute(sql)
#                     tienerubropagos = cursor.fetchone()
#
#                     if tienerubropagos is None:
#                         sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         postulante.cohortes_id = cohorte2023.id
#
#         if postulante.tiporespuesta:
#             if postulante.tiporespuesta.id not in [4, 2]:
#                 if postulante.status == False:
#                     postulante.status =True
#
#         postulante.tiporespuesta = None
#         postulante.save()
#
#         cont += 1
#
#         estado = 0
#         if postulante.estado_aprobador == 1:
#             estado = 'EN PROCESO'
#         elif postulante.estado_aprobador == 2:
#             estado = 'ADMITIDO'
#         elif postulante.estado_aprobador == 3:
#             estado = 'RECHAZADO'
#
#         print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado, 'Canti: ', postulante.total_evidencias())
#
#     # print("Migración de tiene modalidad")
#     #
#     # co = 0
#     # postulantes = InscripcionCohorte.objects.filter(status=True, fecha_creacion__year__in=[2022,2023])
#     # for postulante in postulantes:
#     #     if DetalleAprobacionFormaPago.objects.filter(status=True, inscripcion=postulante).count() > 1:
#     #         postulante.puedeeditarmp = False
#     #     else:
#     #         postulante.puedeeditarmp = True
#     #     postulante.save()
#     #     co += 1
#     #     print(co, 'Lead:', ' ', postulante, 'estado:', ' ', postulante.puedeeditarmp)
# except Exception as ex:
#     pass
#
    # idcontratos = Contrato.objects.filter(status=True).values_list('inscripcion__id', flat=True).order_by('inscripcion__id')
    # postulantes = InscripcionCohorte.objects.filter(status=True, id__in=idcontratos, aceptado=False)
    # c = 0
    # for postulante in postulantes:
    #     if postulante.tiene_contrato_subido() == 2:
    #         postulante.aceptado = True
    #         postulante.save()
    #         observacion = ''
    #         if postulante.formapagopac.id == 1:
    #             observacion = 'Aceptó modalidad de pago por contado - Proceso masivo'
    #         else:
    #             observacion = 'Aceptó modalidad de pago diferido - Proceso masivo'
    #
    #         deta = DetalleAprobacionFormaPago(inscripcion_id=postulante.id,
    #                                           formapagopac=postulante.formapagopac,
    #                                           estadoformapago=1,
    #                                           observacion=observacion,
    #                                           persona=postulante.inscripcionaspirante.persona)
    #         deta.save()
    #
    #         c += 1
    #         print(f'{c} - Postulante: {postulante.inscripcionaspirante.persona}, - Cohorte: {postulante.cohortes}')
    #
    # print("Matriculados antiguos")
    # postulantes = InscripcionCohorte.objects.filter(status=True, aceptado=False)

    # c = 0
    # for postulante in postulantes:
    #     if postulante.tiene_matricula_cohorte():
    #         postulante.aceptado = True
    #         postulante.save()
    #
    #         c += 1
    #         print(f'{c} - Postulante: {postulante.inscripcionaspirante.persona}, - Cohorte: {postulante.cohortes}')



# try:
#     print("Contratos")
#     filtro = Q(status = True, cohortes__maestriaadmision__carrera__coordinacion__id = 7, formapagopac__id = 2, estado_aprobador = 2)
#     filtro = filtro & (~Q(contrato__archivocontrato__exact='')) & (Q(contrato__estado=1))
#     postulantes = InscripcionCohorte.objects.filter(filtro).order_by('-fecha_creacion')
#     c=0
#     for postulante in postulantes:
#         contrato = Contrato.objects.get(status=True, inscripcion=postulante)
#         evi = DetalleAprobacionContrato.objects.filter(status=True, contrato_id=contrato.id, espagare=False).order_by('-id').first()
#         if contrato.estado != evi.estado_aprobacion:
#             contrato.estado = evi.estado_aprobacion
#             contrato.save()
#             c += 1
#             print(f'N°{c} - Postulante: {postulante}')
#
#     filtros = Q(status = True, cohortes__maestriaadmision__carrera__coordinacion__id = 7, formapagopac__id = 2, estado_aprobador = 2)
#     filtros = filtros & (~Q(contrato__archivopagare__exact='')) & (Q(contrato__estadopagare=1))
#     postulantes = InscripcionCohorte.objects.filter(filtros).order_by('-fecha_creacion')
#
#     print("Pagarés")
#     for postulante in postulantes:
#         contrato = Contrato.objects.get(status=True, inscripcion=postulante)
#         evi = DetalleAprobacionContrato.objects.filter(status=True, contrato_id=contrato.id, espagare=True).order_by('-id').first()
#         if contrato.estadopagare != evi.estado_aprobacion:
#             contrato.estadopagare = evi.estado_aprobacion
#             contrato.save()
#
#             print(f'Postulante: {postulante}')
#
# except Exception as ex:
#     pass
# try:
#     cohorte2023 = CohorteMaestria.objects.get(id=161, status=True)
#     cont = 0
#     matriculados = []
#     fvence = datetime.strptime('2023-02-05', '%Y-%m-%d').date()
#
#     postu = InscripcionCohorte.objects.filter(status=True, cohortes__id=143)
#     for pos in postu:
#         if pos.tiene_matricula_cohorte():
#             matriculados.append(pos.id)
#
#     postulantes = InscripcionCohorte.objects.filter(status=True, estado_aprobador__in=[1, 2], cohortes__id=143).exclude(id__in=matriculados)
#     conta = 0
#     contado = 0
#     for postulante in postulantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
#         for lis in listarequisitos:
#             #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
#             if lis in [899]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1189)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#                     deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
#                     if deta.fecha_aprobacion.date() < fvence:
#                         deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.estado_aprobacion = 3
#                         deta.save()
#             #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
#             if lis in [900]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1190)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #COPIA A COLOR DE CEDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
#             if lis in [898]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1188)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CERTIFICADO DE EXPERIENCIA DE ENSEÑANZA DE INGLÉS  MÍNIMO 1 AÑO
#             if lis in [901]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1202)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#             #CERTIFICADO DE FORMACIÓN PEDAGÓGICA Y/O METODOLÓGICA DE INGLÉS - DURACIÓN 120 HORAS
#             if lis in [903]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1192)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CERTIFICADO DE NIVEL DE SUFICIENCIA DE INGLÉS B1 DE ACUERDO AL MARCO COMÚN EUROPEO DE REFERENCIAS PARA LENGUAS (MCER)
#             if lis in [902]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1191)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CARTA DE INTENCIÓN DE INGRESO AL PROGRAMA EN INGLÉS
#             if lis in [904]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1193)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CARTA DE COMPROMISO
#             if lis in [1044]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1220)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#             #CERTIFICADO LABORAL CON EXPERIENCIA DOCENTE MÍNIMO 1 AÑO
#             # if lis in [893]:
#             #     reqma = RequisitosMaestria.objects.get(status=True, pk=1045)
#             #     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#             #     evi.requisitos = reqma
#             #     evi.save()
#             #
#             #     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#             #         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#             #         for evid in evidences:
#             #             evid.status = False
#             #             evid.save()
#
#         evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[1189, 1190, 1188, 1202, 1192, 1191, 1193, 1220])
#         if not evidema.count() >= 7:
#             postulante.estado_aprobador = 1
#             postulante.save()
#             conta += 1
#         else:
#             if evidema.count() >= 7 and postulante.estado_aprobador == 1:
#                 postulante.estado_aprobador = 2
#                 postulante.save()
#             contado += 1
#
#         if postulante.formapagopac and postulante.formapagopac.id == 2:
#             listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
#             for lis in listarequisitosfi:
#                 if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
#                     requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                               cohorte_id=cohorte2023,
#                                                               status=True)[0]
#                     lis.requisitos = requi
#                     lis.save()
#
#                     if lis.requisitos.requisito.id in [54, 62]:
#                         deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).order_by('-id').first()
#                         if deta.fecha_aprobacion.date() < fvence:
#                             deta.estado_aprobacion = 3
#                             deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.save()
#
#                         if DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).count() > 1:
#                             detalles = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).exclude(id=deta.id)
#                             for det in detalles:
#                                 det.status = False
#                                 det.save()
#
#         if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(status=True, id=crubro.id)
#                 rubro.status = False
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (rubro.idrubroepunemi)
#                     cursor.execute(sql)
#                     tienerubropagos = cursor.fetchone()
#
#                     if tienerubropagos is None:
#                         sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         postulante.cohortes_id = cohorte2023.id
#         # if postulante.tiporespuesta:
#         #     postulante.tiporespuesta = 1
#         postulante.save()
#         cont += 1
#
#         estado = 0
#         if postulante.estado_aprobador == 1:
#             estado = 'EN PROCESO'
#         elif postulante.estado_aprobador == 2:
#             estado = 'ADMITIDO'
#         elif postulante.estado_aprobador == 3:
#             estado = 'RECHAZADO'
#
#         print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado, 'Canti: ', postulante.total_evidencias())
# except Exception as ex:
#     print('error: %s' % ex)
# def traer_datos_query():
#     try:
#         sql="""
#                 SELECT coordinacion.nombre AS facultad, carrera.nombre AS carrera,
#                 (persona.apellido1||' '||persona.apellido2||' '||persona.nombres) AS estudiante,
#                 sesion.nombre AS seccion,
#                 (ARRAY_TO_STRING(array(SELECT DISTINCT sesion0.nombre FROM sga_materia materia0
#                 INNER JOIN sga_nivel nivel0 ON materia0.nivel_id = nivel0.id
#                 INNER JOIN sga_sesion sesion0 ON nivel0.sesion_id = sesion0.id
#                 INNER JOIN sga_asignaturamalla asignaturamalla0 ON materia0.asignaturamalla_id = asignaturamalla0.id
#                 INNER JOIN sga_malla malla0 ON asignaturamalla0.malla_id = malla0.id
#                 WHERE malla0.carrera_id=carrera.id AND nivel0.periodo_id=224
#                 AND materia0."status" ORDER BY sesion0.nombre), ', ')) AS sesiones,
#                 inscripcion.id AS inscripcion_id,
#                 sesion.id AS seccion_id,
#                 (ARRAY_TO_STRING(array(SELECT DISTINCT sesion0.id FROM sga_materia materia0
#                 INNER JOIN sga_nivel nivel0 ON materia0.nivel_id = nivel0.id
#                 INNER JOIN sga_sesion sesion0 ON nivel0.sesion_id = sesion0.id
#                 INNER JOIN sga_asignaturamalla asignaturamalla0 ON materia0.asignaturamalla_id = asignaturamalla0.id
#                 INNER JOIN sga_malla malla0 ON asignaturamalla0.malla_id = malla0.id
#                 WHERE malla0.carrera_id=carrera.id AND nivel0.periodo_id=224
#                 AND materia0."status" ORDER BY sesion0.id), ', ')) AS sesiones_id
#                 FROM sga_matricula matricula
#                 INNER JOIN sga_inscripcion inscripcion ON matricula.inscripcion_id = inscripcion.id
#                 INNER JOIN sga_nivel nivel ON matricula.nivel_id = nivel.id
#                 INNER JOIN sga_carrera carrera ON inscripcion.carrera_id = carrera.id
#                 INNER JOIN sga_coordinacion_carrera corcar ON corcar.carrera_id= carrera.id
#                 INNER JOIN sga_coordinacion coordinacion ON coordinacion.id = corcar.coordinacion_id
#                 INNER JOIN sga_sesion sesion ON inscripcion.sesion_id = sesion.id
#                 INNER JOIN sga_persona persona ON inscripcion.persona_id = persona.id
#                 WHERE nivel.periodo_id = 177
#                 AND inscripcion."status"
#                 AND matricula.retiradomatricula = FALSE
#                 AND matricula."status"
#                 AND coordinacion.id IN (1, 2, 3, 4, 5)
#                 ORDER BY coordinacion.nombre, carrera.nombre
#         """
#         cursor = connections['default'].cursor()
#         cursor.execute(sql)
#         rows_effected = cursor.rowcount
#         listado = cursor.fetchall()
#         campos = [desc[0] for desc in cursor.description]
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion')
#         try:
#             os.stat(directory)
#         except:
#             os.mkdir(directory)
#         name_document = 'reporte'
#         nombre_archivo = name_document + "_1.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#
#         __author__ = 'Unemi'
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         fuentecabecera = workbook.add_format({
#             'align': 'center',
#             'bg_color': 'silver',
#             'border': 1,
#             'bold': 1
#         })
#
#         formatoceldacenter = workbook.add_format({
#             'border': 1,
#             'valign': 'vcenter',
#             'align': 'center'})
#
#         row_num, numcolum = 0, 0
#
#         for col_num in campos:
#             ws.write(row_num, numcolum, col_num, fuentecabecera)
#             ws.set_column(row_num, numcolum, 40)
#             numcolum += 1
#         row_num += 1
#         for lis in listado:
#             colum_num = 0
#             for l in lis:
#                 ws.write(row_num, colum_num, l, formatoceldacenter)
#                 ws.set_column(row_num, numcolum, 40)
#                 colum_num += 1
#             row_num += 1
#
#         workbook.close()
#         response = HttpResponse(directory,
#                                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         #
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# traer_datos_query()
#
# print("Finaliza")

# print(u"********************POSTULANTES CON EVIDENCIAS COMPLETAS GENERAL***************************")
#
# try:
#     c = 0
#     postulantes = InscripcionCohorte.objects.filter(status=True, todosubido=False)
#     for postulante in postulantes:
#         requistosmaestria = RequisitosMaestria.objects.filter(status=True, cohorte=postulante.cohortes,
#                                                               requisito__claserequisito__clasificacion__id=1,
#                                                               obligatorio=True).values_list('id', flat=True)
#         cont = 0
#         estado = False
#         for requisto in requistosmaestria:
#             if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante,
#                                                            requisitos_id=requisto,
#                                                            requisitos__requisito__claserequisito__clasificacion__id=1).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante,
#                                                                   requisitos_id=requisto,
#                                                                   requisitos__requisito__claserequisito__clasificacion__id=1).order_by(
#                     '-id').first()
#                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by(
#                     '-id').first()
#                 if deta.estadorevision == 1 or deta.estado_aprobacion == 2:
#                     cont += 1
#
#         if cont == requistosmaestria.count():
#             postulante.todosubido = True
#             postulante.save()
#             c += 1
#
#             print(
#                 f"N: {c} - Postulante: {postulante.inscripcionaspirante.persona} - Cohorte: {postulante.cohortes} - Cantidad: {cont} - Asesor: {postulante.asesor.persona if postulante.asesor else'No registra'}")
#
# except Exception as ex:
#     print(ex)

#     c = 0
#     lista_postu = []
#     postulantes = InscripcionCohorte.objects.filter(status=True, estado_aprobador=1, todosubido=True)
#     for postulante in postulantes:
#         requistosmaestria = RequisitosMaestria.objects.filter(status=True, cohorte=postulante.cohortes, requisito__claserequisito__clasificacion__id=1, obligatorio=True).values_list('id', flat=True)
#         estado = False
#         for requisto in requistosmaestria:
#             if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos_id=requisto, requisitos__requisito__claserequisito__clasificacion__id=1).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos_id=requisto, requisitos__requisito__claserequisito__clasificacion__id=1).order_by('-id').first()
#                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id').first()
#                 if deta.estado_aprobacion == 3:
#                     lista_postu.append(postulante.id)
#                     break
#
#     postulante_rec = InscripcionCohorte.objects.filter(status=True, id__in=lista_postu)
#     for po in postulante_rec:
#         po.tienerechazo = True
#         po.todosubido = False
#         po.preaprobado=False
#         po.save()
#         c += 1
#         print(f"N: {c} - Postulante: {po.inscripcionaspirante.persona} - Cohorte: {po.cohortes} - Asesor: {po.asesor.persona}")
#
# except Exception as ex:
#     print(ex)



# traer_datos_query()
# try:
#     # postulantes = InscripcionCohorte.objects.filter(status=True, asesor__id__in=[24, 25])
#     # listado_pagos = []
#     # for postulante in postulantes:
#     #     listado_fechas = []
#     #     if postulante.comprobante_subido():
#     #         listado_fechas.append(postulante.get_comprobante_subido().fecha_creacion.date())
#     #     if postulante.comprobante_subido_epunemi():
#     #         listado_fechas.append(postulante.get_comprobante_subido_epunemi())
#     #     if postulante.total_pagado_rubro_cohorte() > 0:
#     #         listado_fechas.append(postulante.fecha_primer_pago())
#     #
#     #     if len(listado_fechas) > 0:
#     #         listado_pagos.append(postulante.id)
#     #
#     # leads = InscripcionCohorte.objects.filter(status=True, asesor__id__in=[24, 25]).exclude(id__in=listado_pagos)
#     # asesores = [23, 20, 22, 21, 6]
#     # c = a = 0
#     # observacion = 'ASIGNACIÓN MASIVA'
#     # for lead in leads:
#     #     if AsesorComercial.objects.filter(status=True, pk=int(asesores[c])).exists():
#     #         asesor = AsesorComercial.objects.get(status=True, pk=int(asesores[c]))
#     #         asesoranti = lead.asesor
#     #         lead.asesor = asesor
#     #         lead.estado_asesor = 2
#     #         lead.tiporespuesta_id = None
#     #         lead.save()
#     #         if not asesoranti:
#     #             histo = HistorialAsesor(inscripcion=lead, fecha_inicio=lead.fecha_modificacion,
#     #                             fecha_fin=None, asesor=lead.asesor, observacion=observacion)
#     #             histo.save()
#     #         else:
#     #             histoanti = HistorialAsesor.objects.get(inscripcion_id=lead.id, fecha_fin=None)
#     #             histoanti.fecha_fin = lead.fecha_modificacion
#     #             histoanti.save()
#     #             histo = HistorialAsesor(inscripcion_id=lead.id, fecha_inicio=lead.fecha_modificacion,
#     #                                     fecha_fin=None, asesor=lead.asesor, observacion=observacion)
#     #             histo.save()
#     #
#     #     a += 1
#     #     if c == 4:
#     #         c = 0
#     #     else:
#     #         c += 1
#     #
#     #     print(f"{a} - Lead: {lead.inscripcionaspirante.persona} - Cohorte: {lead.cohortes} - Asesor: {lead.asesor.persona}")

# cohorte2023 = CohorteMaestria.objects.get(id=148, status=True)
    # cont = 0
    # matriculados = []
    #
    # postu = InscripcionCohorte.objects.filter(status=True, cohortes__id=141)
    # for pos in postu:
    #     if pos.tiene_matricula_cohorte():
    #         matriculados.append(pos.id)
    #
    # postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=141).exclude(id__in=matriculados)
    # conta = 0
    # contado = 0
    # for postulante in postulantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
    #     for lis in listarequisitos:
    #         #COPIA A COLOR CÉDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
    #         if lis in [879, 355, 412]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1003)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             if lis in [355, 412]:
    #                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #                 deta.observacion = ''
    #                 deta.observacion_aprobacion = ''
    #                 deta.estado_aprobacion = 1
    #                 deta.save()
    #         #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
    #         if lis in [877]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1001)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #             deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #             deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #             deta.estado_aprobacion = 3
    #             deta.save()
    #         #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
    #         if lis in [878, 358, 415]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1002)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             if lis in [358, 415]:
    #                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #                 deta.observacion = ''
    #                 deta.observacion_aprobacion = ''
    #                 deta.estado_aprobacion = 1
    #                 deta.save()
    #
    #         #CERTIFICADO LABORAL CON EXPERIENCIA DOCENTE MÍNIMO 1 AÑO
    #         if lis in [882, 398, 416]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1004)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             if lis in [398, 416]:
    #                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #                 deta.observacion = ''
    #                 deta.observacion_aprobacion = ''
    #                 deta.estado_aprobacion = 1
    #                 deta.save()
    #
    #     evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[1003, 1001, 1002, 1004])
    #     if not evidema.count() == 4:
    #         postulante.estado_aprobador = 1
    #         postulante.save()
    #         conta += 1
    #     else:
    #         if evidema.count() == 4 and postulante.estado_aprobador == 1:
    #             postulante.estado_aprobador = 2
    #             postulante.save()
    #         contado += 1
    #
    #     if postulante.formapagopac and postulante.formapagopac.id == 2:
    #         listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
    #         for lis in listarequisitosfi:
    #             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
    #                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                           cohorte_id=cohorte2023,
    #                                                           status=True)[0]
    #                 lis.requisitos = requi
    #                 lis.save()
    #
    #                 if lis.requisitos.requisito.id in [54, 62]:
    #                     deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).order_by('-id').first()
    #                     deta.estado_aprobacion = 3
    #                     deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #                     deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #                     deta.save()
    #
    #                     if DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).count() > 1:
    #                         detalles = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).exclude(id=deta.id)
    #                         for det in detalles:
    #                             det.status = False
    #                             det.save()
    #
    #     if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
    #         aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
    #         for crubro in aspiranteconrubro:
    #             rubro = Rubro.objects.get(status=True, id=crubro.id)
    #             rubro.status = False
    #             rubro.save()
    #
    #             if rubro.idrubroepunemi != 0:
    #                 cursor = connections['epunemi'].cursor()
    #                 sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (rubro.idrubroepunemi)
    #                 cursor.execute(sql)
    #                 tienerubropagos = cursor.fetchone()
    #
    #                 if tienerubropagos is None:
    #                     sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     postulante.cohortes_id = cohorte2023.id
    #     # if postulante.tiporespuesta:
    #     #     postulante.tiporespuesta = 1
    #     postulante.save()
    #     cont += 1
    #
    #     estado = 0
    #     if postulante.estado_aprobador == 1:
    #         estado = 'EN PROCESO'
    #     elif postulante.estado_aprobador == 2:
    #         estado = 'ADMITIDO'
    #     elif postulante.estado_aprobador == 3:
    #         estado = 'RECHAZADO'
    #
    #     print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado, 'FP: ', postulante.formapagopac.descripcion)
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# miarchivo = openpyxl.load_workbook("Listado_de_postulantes_atendidos_asesores_nuevo.xlsx")
    # lista = miarchivo.get_sheet_by_name('resultados')
    # totallista = lista.rows
    # a=0
    # c=0
    # for filas in totallista:
    #     a += 1
    #     if a > 1:
    #         idinscripcioncohorte = int(filas[0].value)
    #         cedula = str((filas[9].value)).replace(".", "")
    #         observacion = 'MIGRAXIÓN MASIVA DE LEADS ATENDIDOS DESDE EL USUARIO DE AMY TORRES'
    #         if idinscripcioncohorte != 'None':
    #             if AsesorComercial.objects.filter(status=True, persona__cedula=cedula).exists():
    #                 asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
    #                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
    #                     inscrito = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
    #                     asesoranti = inscrito.asesor
    #                     inscrito.asesor = asesor
    #                     inscrito.estado_asesor = 2
    #                     inscrito.save()
    #                     if not asesoranti:
    #                         histo = HistorialAsesor(inscripcion=inscrito, fecha_inicio=inscrito.fecha_modificacion,
    #                                         fecha_fin=None, asesor=inscrito.asesor, observacion=observacion)
    #                         histo.save()
    #                     else:
    #                         if asesoranti.id != inscrito.asesor.id:
    #                             histoanti = HistorialAsesor.objects.get(inscripcion=inscrito, fecha_fin=None)
    #                             histoanti.fecha_fin = inscrito.fecha_modificacion
    #                             histoanti.save()
    #                             histo = HistorialAsesor(inscripcion=inscrito, fecha_inicio=inscrito.fecha_modificacion,
    #                                             fecha_fin=None, asesor=inscrito.asesor, observacion=observacion)
    #                             histo.save()
    #                         print(f"{c} - Lead: {inscrito.inscripcionaspirante.persona} - Cohorte: {inscrito.cohortes} - Asesor: {inscrito.asesor.persona}")
    #                 c += 1
    #             else:
    #                 print(u"Este asesor no existe")
    #
    #         else:
    #             print(u"FINALIZADO")
    #             break


# try:
#     cohorte2023 = CohorteMaestria.objects.get(id=146, status=True)
#     cont = 0
#     postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=132).exclude(id__in=[32878, 35002, 34618, 32954, 32930,
#     36003, 33010, 31422, 33446, 36732, 31371, 32550, 32756, 31831, 33876, 31245, 33414, 32540, 35297, 34569, 32090, 32940, 35463, 33286, 31907,
#     33862, 33380, 25451, 32719, 32762, 34370, 34213, 31911, 34840, 31343, 34885, 34141, 32236, 34662, 34514, 35827, 33056, 28499, 35704, 34102,
#     33285, 35333, 29922, 31363, 32948, 34816, 29378, 34892, 22703, 31992, 33265, 35226, 33907, 34753, 34336, 34179, 33125, 33908, 36037, 34841,
#     35323, 34409, 32931, 34591, 34654, 31367, 34037, 31395, 35152, 31518, 36468, 34344, 25715, 27667, 34294, 33011, 32721, 34387, 32706, 33984,
#     34614, 32553, 29461, 34107, 32695, 35165, 36557, 36384, 37240, 32693, 38226, 37875, 37576, 36807, 37280, 37782, 35520, 33887, 36944, 37610,
#     35770, 37600, 38874, 31530, 31533, 37433, 35356, 34534, 39170, 37241, 39024, 38806, 39044, 38505, 37869, 38722, 39432, 38297, 39043, 37184,
#     39144, 22685, 39642, 35362, 39976, 35495, 38293, 35788, 37608, 37248, 35364, 38152, 38153, 38530, 39752, 37894, 40245, 40148, 40262, 38996,
#     40467, 39379, 37502, 39581, 25209, 36546, 37868, 37344, 39529, 40284, 38010, 24868, 39935, 36808, 23746, 39578, 40533, 39377, 40594, 22241,
#     40788, 40890, 40110, 34671, 36551, 28990, 31175, 34540, 33186, 31767, 35891, 38525, 35379, 32802, 33123, 23142, 32812, 34597, 34743, 36245,
#     29569, 36343, 36259, 39771, 41300, 29675, 40837, 40217, 39260, 40912, 40696, 41092, 35983, 22217, 23083, 38465, 38110, 40636, 38285, 38934,
#     39182, 39776, 41120, 38489, 25555, 23736, 41665, 41615, 41713, 41355, 38314, 40911, 31969, 35582, 41566, 33474, 33352, 26906, 40969, 41736,
#     34232, 40817, 41762, 40637, 34529, 37416, 41504, 40255, 42665, 42675, 42839, 41476, 43549, 26010])
#     conta = 0
#     contado = 0
#     for postulante in postulantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
#         for lis in listarequisitos:
#             #COPIA A COLOR DE CEDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
#             if lis in [550, 732]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=968)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#             #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
#             if lis in [551, 735]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=970)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
#                 deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                 deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                 deta.estado_aprobacion = 3
#                 deta.save()
#             #HOJA DE VIDA.
#             if lis in [552, 733]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=969)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#             #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
#             if lis in [553, 747]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=967)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#             #CERTIFICADO LABORAL MINIMO 1 AÑO DE EXPERIENCIA.
#             if lis in [569, 746]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=966)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#         evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[966, 967, 968, 969, 970])
#         if not evidema.count() == 4:
#             postulante.estado_aprobador = 1
#             postulante.save()
#             conta += 1
#         else:
#             if evidema.count() == 4 and postulante.estado_aprobador == 1:
#                 postulante.estado_aprobador = 2
#                 postulante.save()
#             contado += 1
#
#         if postulante.formapagopac and postulante.formapagopac.id == 2:
#             listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
#             for lis in listarequisitosfi:
#                 if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
#                     requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                               cohorte_id=cohorte2023,
#                                                               status=True)[0]
#                     lis.requisitos = requi
#                     lis.save()
#
#                     if lis.requisitos.requisito.id in [54, 62]:
#                         deta = DetalleEvidenciaRequisitosAspirante.objects.get(status=True, evidencia=lis)
#                         deta.estado_aprobacion = 3
#                         deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.save()
#
#         if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(status=True, id=crubro.id)
#                 rubro.status = False
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
#                     cursor.execute(sql)
#                     cursor.close()
#
#         postulante.cohortes_id = cohorte2023.id
#         postulante.save()
#         cont += 1
#
#         estado = 0
#         if postulante.estado_aprobador == 1:
#             estado = 'EN PROCESO'
#         elif postulante.estado_aprobador == 2:
#             estado = 'ADMITIDO'
#         elif postulante.estado_aprobador == 3:
#             estado = 'RECHAZADO'
#
#         print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado)
#
# except Exception as ex:
#     print('error: %s' % ex)
#
# print("Finaliza")



    # aspirantesmatri = []
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id__in=[91, 100])
    #
    # for aspi in aspirantes:
    #     if aspi.estado_aprobador == 3:
    #         aspirantesmatri.append(aspi.id)
    #     elif aspi.inscripcion:
    #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
    #             aspirantesmatri.append(aspi.id)
     #
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id__in=[91, 100]).exclude(id__in=aspirantesmatri)
    #
    # cont = 0
    # cohorte2023 = CohorteMaestria.objects.get(id=141, status=True)
    # for aspirante in aspirantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
    #     listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
    #                                                                        requisitos__requisito__claserequisito__clasificacion=3)
    #     for lis in listarequisitos:
    #         if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                              cohorte_id=cohorte2023.id, status=True):
    #             requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                       cohorte_id=cohorte2023.id,
    #                                                       status=True)[0]
    #             lis.requisitos = requi
    #             lis.save()
    #
    #     if aspirante.formapagopac:
    #         if aspirante.formapagopac.id == 2:
    #             for listf in listarequisitosfinan:
    #                 if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                      cohorte_id=cohorte2023.id, status=True,
    #                                                      requisito__claserequisito__clasificacion=3):
    #                     requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                                    cohorte_id=cohorte2023.id, status=True,
    #                                                                    requisito__claserequisito__clasificacion=3)[0]
    #                     listf.requisitos = requifinan
    #                     listf.save()
    #
    #     if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
    #         rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
    #         for rubro in rubros:
    #             if rubro.admisionposgradotipo == 2:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
    #                 rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #             elif rubro.admisionposgradotipo == 3:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     aspirante.cohortes_id = cohorte2023.id
    #     aspirante.save()
    #
    #     cont += 1
    #     print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)

    # print(u"********************Biotecnología***************************")
    #
    # aspirantesmatri = []
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=107)
    #
    # for aspi in aspirantes:
    #     if aspi.inscripcion:
    #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
    #             aspirantesmatri.append(aspi.id)
    #
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=107).exclude(id__in=aspirantesmatri)
    #
    # cont = 0
    # cohorte2023 = CohorteMaestria.objects.get(id=139, status=True)
    # for aspirante in aspirantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
    #     listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
    #                                                                        requisitos__requisito__claserequisito__clasificacion=3)
    #     for lis in listarequisitos:
    #         if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                              cohorte_id=cohorte2023.id, status=True):
    #             requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                       cohorte_id=cohorte2023.id,
    #                                                       status=True)[0]
    #             lis.requisitos = requi
    #             lis.save()
    #
    #     if aspirante.formapagopac:
    #         if aspirante.formapagopac.id == 2:
    #             for listf in listarequisitosfinan:
    #                 if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                      cohorte_id=cohorte2023.id, status=True,
    #                                                      requisito__claserequisito__clasificacion=3):
    #                     requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                                    cohorte_id=cohorte2023.id, status=True,
    #                                                                    requisito__claserequisito__clasificacion=3)[0]
    #                     listf.requisitos = requifinan
    #                     listf.save()
    #
    #     if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
    #         rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
    #         for rubro in rubros:
    #             if rubro.admisionposgradotipo == 2:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
    #                 rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #             elif rubro.admisionposgradotipo == 3:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     aspirante.cohortes_id = cohorte2023.id
    #     aspirante.save()
    #
    #     cont += 1
    #     print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)

    # print(u"********************Neuropsicología***************************")
    #
    # aspirantesmatri = []
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=86)
    #
    # for aspi in aspirantes:
    #     if aspi.inscripcion:
    #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
    #             aspirantesmatri.append(aspi.id)
    #
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=86).exclude(id__in=aspirantesmatri)
    #
    # cont = 0
    # cohorte2023 = CohorteMaestria.objects.get(id=140, status=True)
    # for aspirante in aspirantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
    #     listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
    #                                                                        requisitos__requisito__claserequisito__clasificacion=3)
    #     for lis in listarequisitos:
    #         if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                              cohorte_id=cohorte2023.id, status=True):
    #             requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                       cohorte_id=cohorte2023.id,
    #                                                       status=True)[0]
    #             lis.requisitos = requi
    #             lis.save()
    #
    #     if aspirante.formapagopac:
    #         if aspirante.formapagopac.id == 2:
    #             for listf in listarequisitosfinan:
    #                 if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                      cohorte_id=cohorte2023.id, status=True,
    #                                                      requisito__claserequisito__clasificacion=3):
    #                     requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                                    cohorte_id=cohorte2023.id, status=True,
    #                                                                    requisito__claserequisito__clasificacion=3)[0]
    #                     listf.requisitos = requifinan
    #                     listf.save()
    #
    #     if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
    #         rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
    #         for rubro in rubros:
    #             if rubro.admisionposgradotipo == 2:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
    #                 rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #             elif rubro.admisionposgradotipo == 3:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     aspirante.cohortes_id = cohorte2023.id
    #     aspirante.save()
    #
    #     cont += 1
    #     print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)



# try:
#     aspirantesmatri = []
#     aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=127)
#
#     for aspi in aspirantes:
#         if aspi.inscripcion:
#             if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
#                 aspirantesmatri.append(aspi.id)
#         elif Rubro.objects.filter(status=True, inscripcion=aspi).exists():
#             firstrubro = Rubro.objects.filter(status=True, inscripcion=aspi).order_by('id')[0]
#             if Pago.objects.filter(status=True, rubro=firstrubro).exists():
#                 aspirantesmatri.append(aspi.id)
#         elif Contrato.objects.filter(status=True, inscripcion=aspi).exists():
#             aspirantesmatri.append(aspi.id)
#
#     aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=127).exclude(id__in=aspirantesmatri)
#
#     cont = 0
#     cohorte2023 = CohorteMaestria.objects.get(id=137, status=True)
#     for aspirante in aspirantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
#         listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
#                                                                            requisitos__requisito__claserequisito__clasificacion=3)
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte_id=cohorte2023.id, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                           cohorte_id=cohorte2023.id,
#                                                           status=True)[0]
#                 lis.requisitos = requi
#                 lis.save()
#
#         if aspirante.formapagopac:
#             if aspirante.formapagopac.id == 2:
#                 for listf in listarequisitosfinan:
#                     if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                          cohorte_id=cohorte2023.id, status=True,
#                                                          requisito__claserequisito__clasificacion=3):
#                         requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                                        cohorte_id=cohorte2023.id, status=True,
#                                                                        requisito__claserequisito__clasificacion=3)[0]
#                         listf.requisitos = requifinan
#                         listf.save()
#
#         if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
#             rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
#             for rubro in rubros:
#                 if rubro.admisionposgradotipo == 2:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
#                     rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
#                         # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#                 elif rubro.admisionposgradotipo == 3:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         aspirante.cohortes_id = cohorte2023.id
#         aspirante.save()
#
#         cont += 1
#         print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     # leadsbasica = InscripcionCohorte.objects.filter(status=True, cohortes__id=122, inscripcion__isnull=False).values_list('inscripcion__id', flat=True)
#     # matriculados = Matricula.objects.filter(status=True, inscripcion__id__in=leadsbasica, nivel__periodo__id=162)
#
#     matriculados = Matricula.objects.filter(status=True, nivel__periodo__id__in=[182, 160, 167]).order_by('inscripcion__persona__apellido1',
#                                                                                             'inscripcion__persona__apellido2',
#                                                                                             'inscripcion__persona__nombres')
#
#     for matriculado in matriculados:
#         inscrito = InscripcionCohorte.objects.filter(status=True, inscripcion=matriculado.inscripcion).first()
#         if inscrito is not None:
#             rubrosinscrito = Rubro.objects.filter(status=True, inscripcion=inscrito,persona=inscrito.inscripcionaspirante.persona)
#             for rubro in rubrosinscrito:
#                 if not rubro.matricula:
#                     rubro.matricula = matriculado
#                     rubro.save()
#                     print(f"Se ha actualizado rubro de {inscrito.inscripcionaspirante.persona} {rubro.id} - {rubro.matricula}")
#                 print(f"Ya tiene matricula {inscrito.inscripcionaspirante.persona}  {rubro.id} - {rubro.matricula}")
#         else:
#             print(f"NO ID INSCRITO {matriculado.inscripcion.persona} - {matriculado.inscripcion.persona.cedula}")
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")




# try:
#     miarchivo = openpyxl.load_workbook("contabilidad_financiamiento.xlsx")
#     lista = miarchivo.get_sheet_by_name('resultados')
#     totallista = lista.rows
#     a=0
#     c=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             idinscripcioncohorte = str(filas[0].value)
#             # cedula = str((filas[6].value).replace(".", ""))
#             cedula = str((filas[6].value)).replace(".", "")
#             tipo = TipoFormaPagoPac.objects.get(id=2, status=True)
#             if idinscripcioncohorte != 'None':
#                 observacion = '30%+10 CUOTAS'
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     if AsesorComercial.objects.filter(status=True, persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#
#                         inscripcion.formapagopac = tipo
#                         inscripcion.estadoformapago = 1
#                         inscripcion.save()
#
#                         deta = DetalleAprobacionFormaPago(inscripcion_id=inscripcion.id,
#                                                           formapagopac=tipo,
#                                                           estadoformapago=1,
#                                                           observacion=observacion,
#                                                           persona=asesor.persona)
#                         deta.save()
#
#                         c += 1
#                         print(c,' ','Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor, ' - ', 'FormaPago:', inscripcion.formapagopac.descripcion)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")

# try:
#     # cohorte = CohorteMaestria.objects.get(id__in=[116, 118]).values_list('id')
#     id_inscripcion = Matricula.objects.filter(inscripcion__carrera__coordinacion__id=7).values_list('inscripcion__id', flat=True)
#
#     prospectos = InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__coordinacion__id=7,
#                            fecha_creacion__in=InscripcionCohorte.objects.values('inscripcionaspirante__id').annotate(fecha_creacion=Max('fecha_creacion')).values_list('fecha_creacion', flat=True).filter(status=True),
#                                                    estado_aprobador=2, fecha_creacion__lte = '2022-08-31').exclude(inscripcion__id__in=id_inscripcion)
#     tipo = TipoRespuestaProspecto.objects.get(id=5, status=True)
#     cont = 0
#     for prospecto in prospectos:
#         aspi = InscripcionCohorte.objects.get(pk=prospecto.id, status=True)
#
#         aspi.tiporespuesta = tipo
#         aspi.save()
#         cont += 1
#
#         print(cont, 'Lead:', ' ', aspi.inscripcionaspirante.persona, 'Respuesta:', ' ', aspi.tiporespuesta.descripcion)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     # cohorte = CohorteMaestria.objects.get(id__in=[116, 118]).values_list('id')
#     id_inscripcion = Matricula.objects.filter(inscripcion__carrera__coordinacion__id=7).values_list('inscripcion__id', flat=True)
#
#     prospectos = InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__coordinacion__id=7,
#                            fecha_creacion__in=InscripcionCohorte.objects.values('inscripcionaspirante__id').annotate(fecha_creacion=Max('fecha_creacion')).values_list('fecha_creacion', flat=True).filter(status=True),
#                                                    estado_aprobador__in=[1, 3], fecha_creacion__lte = '2022-08-31').exclude(inscripcion__id__in=id_inscripcion, cohortes__id__in=[120, 123, 122, 113, 121, 86, 124, 107, 125, 126, 127, 128])
#     tipo = TipoRespuestaProspecto.objects.get(id=4, status=True)
#     cont = 0
#     for prospecto in prospectos:
#         aspi = InscripcionCohorte.objects.get(pk=prospecto.id, status=True)
#
#         aspi.tiporespuesta = tipo
#         aspi.save()
#         cont += 1
#
#         print(cont, 'Lead:', ' ', aspi.inscripcionaspirante.persona, 'Respuesta:', ' ', aspi.tiporespuesta.descripcion)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")

# try:
#     aspirantesmatri = []
#     aspirantes = InscripcionCohorte.objects.filter(status=True, id__in=[28791, 21682, 27463, 19158, 27692, 28231, 25270, 28701, 26615, 25141])
#
#     # for aspi in aspirantes:
#     #     if aspi.inscripcion:
#     #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
#     #             aspirantesmatri.append(aspi.id)
#
#     # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=122).exclude(id__in=aspirantesmatri)
#
#     cont = 0
#     cohorte2023 = CohorteMaestria.objects.get(id=122, status=True)
#     for aspirante in aspirantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
#         listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
#                                                                            requisitos__requisito__claserequisito__clasificacion=3)
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte_id=cohorte2023.id, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                           cohorte_id=cohorte2023.id,
#                                                           status=True)[0]
#                 lis.requisitos = requi
#                 lis.save()
#
#         if aspirante.formapagopac:
#             if aspirante.formapagopac.id == 2:
#                 for listf in listarequisitosfinan:
#                     if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                          cohorte_id=cohorte2023.id, status=True,
#                                                          requisito__claserequisito__clasificacion=3):
#                         requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                                        cohorte_id=cohorte2023.id, status=True,
#                                                                        requisito__claserequisito__clasificacion=3)[0]
#                         listf.requisitos = requifinan
#                         listf.save()
#
#         if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
#             rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
#             for rubro in rubros:
#                 if rubro.admisionposgradotipo == 2:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
#                     rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#                 elif rubro.admisionposgradotipo == 3:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         aspirante.cohortes_id = cohorte2023.id
#         aspirante.save()
#
#         cont += 1
#         print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     miarchivo = openpyxl.load_workbook("masivo_finanaciamiento_original.xlsx")
#     lista = miarchivo.get_sheet_by_name('resultados')
#     totallista = lista.rows
#     a=0
#     c=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             idinscripcioncohorte = str(filas[1].value)
#             # cedula = str((filas[6].value).replace(".", ""))
#             cedula = str((filas[7].value)).replace(".", "")
#             tipo = TipoFormaPagoPac.objects.get(id=2, status=True)
#             if idinscripcioncohorte != 'None':
#                 if str(filas[9].value) == 'None':
#                     observacion = 'NINGUNA'
#                 else:
#                     observacion = str(filas[9].value)
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     if AsesorComercial.objects.filter(status=True, persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#
#                         inscripcion.formapagopac = tipo
#                         inscripcion.estadoformapago = 1
#                         inscripcion.save()
#
#                         deta = DetalleAprobacionFormaPago(inscripcion_id=inscripcion.id,
#                                                           formapagopac=tipo,
#                                                           estadoformapago=1,
#                                                           observacion=observacion,
#                                                           persona=asesor.persona)
#                         deta.save()
#
#                         c += 1
#                         print(c,' ','Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor, ' - ', 'FormaPago:', inscripcion.formapagopac.descripcion, ' - ', observacion)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     miarchivo = openpyxl.load_workbook("contabilidad_financiamiento.xlsx")
#     lista = miarchivo.get_sheet_by_name('ventas')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             cedula = str(filas[9].value)
#             if cedula == 'None':
#                 cedula = ''
#             else:
#                 cedula = str((filas[9].value).replace(".", ""))
#             idinscripcioncohorte = str(filas[1].value)
#             if idinscripcioncohorte != 'None':
#                 observacion = 'MASIVO VENTAS COHORTE II BASICA'
#                 # urlzoom=u"https://unemi-edu-ec.zoom.us/j/%s"%idzoom
#                 # print(u"%s"%correo)
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     asesoranti = inscripcion.asesor
#                     if AsesorComercial.objects.filter(status=True,persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#                         # if not inscripcion.asesor:
#                         inscripcion.asesor = asesor
#                         inscripcion.estado_asesor = 2
#                         inscripcion.save()
#
#                         if not asesoranti:
#                             histo = HistorialAsesor(inscripcion_id=inscripcion.id, fecha_inicio=inscripcion.fecha_modificacion,
#                                             fecha_fin=None, asesor=inscripcion.asesor, observacion=observacion)
#                             histo.save()
#                         else:
#                             if not asesoranti == asesor:
#                                 histoanti = HistorialAsesor.objects.get(inscripcion_id=inscripcion.id, fecha_fin=None)
#                                 histoanti.fecha_fin = inscripcion.fecha_modificacion
#                                 histoanti.save()
#                                 histo = HistorialAsesor(inscripcion_id=inscripcion.id, fecha_inicio=inscripcion.fecha_modificacion,
#                                                         fecha_fin=None, asesor=inscripcion.asesor, observacion=observacion)
#                                 histo.save()
#                             else:
#                                 print('El asesor ya tiene este lead asignado')
#
#                         print('Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")