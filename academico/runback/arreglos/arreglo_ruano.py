import os
import sys
import io
import xlsxwriter
import xlwt
import openpyxl
import xlwt
from xlwt import *
from django.http import HttpResponse
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))


YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from balcon.models import *
from inno.funciones import *
from Moodle_Funciones import crearhtmlphpmoodle


################################################# VERIFICADAS ################################################
###MIGRACION TRABAJO SOCIAL
# @transaction.atomic()
def homologacion_ts():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_ts_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ts.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=130
        mallaantigua_id=206
        mallanueva_id=485

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp= haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10724, 10730, 10777, 10787]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True



                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                            if equivalencia.asignaturamallasalto_id in [10777, 10787]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10777:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=8)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()


                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10787:
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if equivalencia.asignaturamallasalto_id in [10724, 10730]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10724 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                            horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas = horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                            )
                                            vinculacion.save()

                                        if equivalencia.asignaturamallasalto_id == 10730 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                            horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas = horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                            )
                                            vinculacion.save()


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)


                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_ts_rezagados():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_ts_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_ts.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=130
        mallaantigua_id=206
        mallanueva_id=485
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion_id=inscripcion).first()
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s - %s" % (matricula, cont, inscripcion))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp= haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10724, 10730, 10777, 10787]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10777, 10787]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10777:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=8)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=ioctavonuevo.horas_practicas,
                                                                                                             nivelmalla=ioctavonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=ioctavonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10787:
                                                itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=9)
                                                inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=9)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarionoveno).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarionoveno).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarionoveno).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=inovenonuevo.horas_practicas,
                                                                                                             nivelmalla=inovenonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=inovenonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarionoveno).update(
                                                    itinerario=inovenonuevo)


                                if equivalencia.asignaturamallasalto_id in [10724, 10730]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10724 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas = horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10730 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas = horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                )
                                                vinculacion.save()


                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(1)
                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

@transaction.atomic()
def llenar_tabla_equivalencias_ts():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_trabajosocial.xlsx")
        lista = miarchivo.get_sheet_by_name('malla_nuevas')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)

###MIGRACION IDIOMAS
def llenar_tabla_equivalencias_idiomas():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_pedagogia.xlsx")
        lista = miarchivo.get_sheet_by_name('datos')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)


def homologacion_idiomas():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("idiomas.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_idiomas_rezagados():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_idiomas.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion_id=inscripcion).first()
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s - %s" % (matricula, cont, inscripcion))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                                if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                    if not practicaspp or not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 11018:

                                                itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                                iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioprimero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioprimero).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioprimero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iprimeronuevo.horas_practicas,
                                                                                                             nivelmalla=iprimeronuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iprimeronuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioprimero).update(
                                                    itinerario=iprimeronuevo)

                                            if equivalencia.asignaturamallasalto_id == 11021:
                                                itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=2)
                                                isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=2)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosegundo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosegundo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosegundo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=isegundonuevo.horas_practicas,
                                                                                                             nivelmalla=isegundonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=isegundonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosegundo).update(
                                                    itinerario=isegundonuevo)

                                            if equivalencia.asignaturamallasalto_id == 11020:
                                                itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=3)
                                                iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=3)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariotercero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariotercero).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariotercero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iterceronuevo.horas_practicas,
                                                                                                             nivelmalla=iterceronuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iterceronuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariotercero).update(
                                                    itinerario=iterceronuevo)

                                            if equivalencia.asignaturamallasalto_id == 10990:
                                                itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=4)
                                                icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=4)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariocuarto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariocuarto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariocuarto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=icuartonuevo.horas_practicas,
                                                                                                             nivelmalla=icuartonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=icuartonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariocuarto).update(
                                                    itinerario=icuartonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10991:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iquintonuevo.horas_practicas,
                                                                                                             nivelmalla=iquintonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iquintonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 11005:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=isextonuevo.horas_practicas,
                                                                                                             nivelmalla=isextonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=isextonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10993:
                                                ####################################### VINCULACION ###################################################
                                                if not horasvinculacion:
                                                    if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                        horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                        vinculacion = ParticipantesMatrices(status=True,
                                                                                            matrizevidencia_id=2,
                                                                                            proyecto_id=601,
                                                                                            inscripcion=inscripcion,
                                                                                            horas=horasfalta,
                                                                                            registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                            registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                            estado=1
                                                                                            )
                                                        vinculacion.save()
                                                ######################################################################################################
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iseptimonuevo.horas_practicas,
                                                                                                             nivelmalla=iseptimonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iseptimonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10997:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)
                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=ioctavonuevo.horas_practicas,
                                                                                                             nivelmalla=ioctavonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=ioctavonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                            if equivalencia.asignaturamallasalto_id == 11007:
                                                ####################################### VINCULACION ###################################################
                                                if not horasvinculacion:
                                                    if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                        horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                        vinculacion = ParticipantesMatrices(status=True,
                                                                                            matrizevidencia_id=2,
                                                                                            proyecto_id=601,
                                                                                            inscripcion=inscripcion,
                                                                                            horas=horasfalta,
                                                                                            registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                            registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                            estado=1
                                                                                            )
                                                        vinculacion.save()
                                                ######################################################################################################
                                                itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=9)
                                                inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=9)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarionoveno).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarionoveno).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarionoveno).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=inovenonuevo.horas_practicas,
                                                                                                             nivelmalla=inovenonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=inovenonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarionoveno).update(
                                                    itinerario=inovenonuevo)


                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(1)

                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


###MIGRACION COMUNICACION
def llenar_tabla_equivalencias_com():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_comunicacion.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA_NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)


def homologacion_comunicacion():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_comunicacion.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]

        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("comlinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=131
        mallaantigua_id=205
        mallanueva_id=488

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp= haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10850,10853,10854,10859,10865]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True



                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                            if equivalencia.asignaturamallasalto_id in [10850,10853,10854]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10850:
                                            itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[0]
                                            itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[0]

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosanteriores).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosanteriores).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosanteriores).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=itinerariosnuevos.horas_practicas,
                                                                                                         nivelmalla=itinerariosnuevos.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=itinerariosnuevos,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosanteriores).update(
                                                itinerario=itinerariosnuevos)

                                        if equivalencia.asignaturamallasalto_id == 10854:
                                            itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[1]
                                            itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[1]

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosanteriores).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosanteriores).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosanteriores).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=itinerariosnuevos.horas_practicas,
                                                                                                         nivelmalla=itinerariosnuevos.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=itinerariosnuevos,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosanteriores).update(
                                                itinerario=itinerariosnuevos)

                                        if equivalencia.asignaturamallasalto_id == 10853:
                                            itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[2]
                                            itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[2]

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosanteriores).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosanteriores).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosanteriores).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=itinerariosnuevos.horas_practicas,
                                                                                                         nivelmalla=itinerariosnuevos.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=itinerariosnuevos,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosanteriores).update(
                                                itinerario=itinerariosnuevos)


                            if equivalencia.asignaturamallasalto_id in [10859,10865]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10859 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                            horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()

                                        if equivalencia.asignaturamallasalto_id == 10865 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                            horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                # if cont_asig_vinculacion_aprobadas != 0:
                #     if cont_asig_vinculacion_aprobadas == 1:
                #         horasfalta = 80
                #     elif cont_asig_vinculacion_aprobadas == 2:
                #         horasfalta = 160
                #     vinculacion = ParticipantesMatrices(status=True,
                #                                         matrizevidencia_id=2,
                #                                         proyecto_id=601,
                #                                         inscripcion=inscripcion,
                #                                         horas=horasfalta,
                #                                         registrohorasdesde=datetime.now().date(),
                #                                         registrohorashasta=datetime.now().date(),
                #                                         estado=1
                #                                         )
                #     vinculacion.save()

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_comunicacion_rezagados():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_comunicacion.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]

        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_comunicacion.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=131
        mallaantigua_id=205
        mallanueva_id=488
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion_id=inscripcion).first()
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp= haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10850,10853,10854,10859,10865]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10850,10853,10854]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10850:
                                                itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[0]
                                                itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[0]

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosanteriores).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosanteriores).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosanteriores).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=itinerariosnuevos.horas_practicas,
                                                                                                             nivelmalla=itinerariosnuevos.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=itinerariosnuevos,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosanteriores).update(
                                                    itinerario=itinerariosnuevos)

                                            if equivalencia.asignaturamallasalto_id == 10854:
                                                itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[1]
                                                itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[1]

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosanteriores).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosanteriores).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosanteriores).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=itinerariosnuevos.horas_practicas,
                                                                                                             nivelmalla=itinerariosnuevos.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=itinerariosnuevos,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosanteriores).update(
                                                    itinerario=itinerariosnuevos)

                                            if equivalencia.asignaturamallasalto_id == 10853:
                                                itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[2]
                                                itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[2]

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosanteriores).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosanteriores).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosanteriores).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=itinerariosnuevos.horas_practicas,
                                                                                                             nivelmalla=itinerariosnuevos.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=itinerariosnuevos,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosanteriores).update(
                                                    itinerario=itinerariosnuevos)


                                if equivalencia.asignaturamallasalto_id in [10859,10865]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10859 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10865 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    # if cont_asig_vinculacion_aprobadas != 0:
                    #     if cont_asig_vinculacion_aprobadas == 1:
                    #         horasfalta = 80
                    #     elif cont_asig_vinculacion_aprobadas == 2:
                    #         horasfalta = 160
                    #     vinculacion = ParticipantesMatrices(status=True,
                    #                                         matrizevidencia_id=2,
                    #                                         proyecto_id=601,
                    #                                         inscripcion=inscripcion,
                    #                                         horas=horasfalta,
                    #                                         registrohorasdesde=datetime.now().date(),
                    #                                         registrohorashasta=datetime.now().date(),
                    #                                         estado=1
                    #                                         )
                    #     vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


###MIGRACION ECONOMIA
def llenar_tabla_equivalencias_econo():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_economia.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)


def homologacion_economia():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_econo.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("econo.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=128
        mallaantigua_id=201
        mallanueva_id=489

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10733, 10742, 10770, 10774]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True

                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [10770, 10774]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10770:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10774:
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                            if equivalencia.asignaturamallasalto_id in [10733, 10742]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10733 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                            horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()

                                        if equivalencia.asignaturamallasalto_id == 10742 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                            horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_economia_rezagados():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_econo.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_economia.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("datos1")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=128
        mallaantigua_id=201
        mallanueva_id=489
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion_id=inscripcion).first()
                # if not matricula:
                #     sin_matricula.append(inscripcion)
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10733, 10742, 10770, 10774]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True

                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                                if equivalencia.asignaturamallasalto_id in [10770, 10774]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10770:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=isextonuevo.horas_practicas,
                                                                                                             nivelmalla=isextonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=isextonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10774:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iseptimonuevo.horas_practicas,
                                                                                                             nivelmalla=iseptimonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iseptimonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                if equivalencia.asignaturamallasalto_id in [10733, 10742]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10733 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10742 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()


                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

###MIGRACION INICIAL
def homologacion_inicial():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_inicial.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inicial.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=127
        mallaantigua_id=200
        mallanueva_id=491

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10831,10856,10880,10894,10899,10907,10910,10926,10928]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [10831,10856,10880,10894,10899,10907,10910,10928]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10831:
                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud=3,
                                                                                                               itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada :
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10856:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()


                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10880:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()


                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10894:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()


                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10899:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()


                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10907:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()


                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10910:
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10928:
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if equivalencia.asignaturamallasalto_id in [10926]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10926 and inscripcion.numero_horas_proyectos_vinculacion() < 320:
                                            horasfalta = 320 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_inicial_rezagados():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_inicial.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_inicial.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=127
        mallaantigua_id=200
        mallanueva_id=491
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion_id=inscripcion).first()
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10831,10856,10880,10894,10899,10907,10910,10926,10928]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                                if equivalencia.asignaturamallasalto_id in [10831,10856,10880,10894,10899,10907,10910,10928]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10831:
                                                itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                                iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioprimero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                                   inscripcion=inscripcion,
                                                                                                                   estadosolicitud=3,
                                                                                                                   itinerariomalla=itinerarioprimero).exists()

                                                if not practica or practicarechazada :
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioprimero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iprimeronuevo.horas_practicas,
                                                                                                             nivelmalla=iprimeronuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iprimeronuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioprimero).update(
                                                    itinerario=iprimeronuevo)

                                            if equivalencia.asignaturamallasalto_id == 10856:
                                                itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=2)
                                                isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=2)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosegundo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosegundo).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosegundo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=isegundonuevo.horas_practicas,
                                                                                                             nivelmalla=isegundonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=isegundonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosegundo).update(
                                                    itinerario=isegundonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10880:
                                                itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=3)
                                                iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=3)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariotercero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariotercero).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariotercero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iterceronuevo.horas_practicas,
                                                                                                             nivelmalla=iterceronuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iterceronuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariotercero).update(
                                                    itinerario=iterceronuevo)

                                            if equivalencia.asignaturamallasalto_id == 10894:
                                                itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=4)
                                                icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=4)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariocuarto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariocuarto).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariocuarto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=icuartonuevo.horas_practicas,
                                                                                                             nivelmalla=icuartonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=icuartonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariocuarto).update(
                                                    itinerario=icuartonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10899:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iquintonuevo.horas_practicas,
                                                                                                             nivelmalla=iquintonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iquintonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10907:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=isextonuevo.horas_practicas,
                                                                                                             nivelmalla=isextonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=isextonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10910:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=7)
                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iseptimonuevo.horas_practicas,
                                                                                                             nivelmalla=iseptimonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iseptimonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10928:
                                                itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=9)
                                                inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=9)
                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarionoveno).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarionoveno).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarionoveno).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=inovenonuevo.horas_practicas,
                                                                                                             nivelmalla=inovenonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=inovenonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarionoveno).update(
                                                    itinerario=inovenonuevo)


                                if equivalencia.asignaturamallasalto_id in [10926]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10926 and inscripcion.numero_horas_proyectos_vinculacion() < 320:
                                                horasfalta = 320 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def llenar_tabla_equivalencias_inicial():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_einicial.xlsx")
        lista = miarchivo.get_sheet_by_name('datos')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)


###MIGRACION BASICA
def llenar_tabla_equivalencias_basica():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_eb.xlsx")
        lista = miarchivo.get_sheet_by_name('datos')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)

###MIGRACION BASICA
def homologacion_basica():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_basica.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("basica.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=135
        mallaantigua_id=208
        mallanueva_id=490

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10822,10834,10846,10858,10870,10888,10913,10930,10944]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [10822,10834,10846,10858,10870,10888,10913,10930,10944]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10822:
                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)


                                        if equivalencia.asignaturamallasalto_id == 10834:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)


                                        if equivalencia.asignaturamallasalto_id == 10846:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)


                                        if equivalencia.asignaturamallasalto_id == 10858:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)


                                        if equivalencia.asignaturamallasalto_id == 10870:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)


                                        if equivalencia.asignaturamallasalto_id == 10888:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)


                                        if equivalencia.asignaturamallasalto_id == 10913:
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)


                                        if equivalencia.asignaturamallasalto_id == 10930:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)


                                        if equivalencia.asignaturamallasalto_id == 10944:
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if equivalencia.asignaturamallasalto_id in [10913,10930]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10913 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                            horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()

                                        if equivalencia.asignaturamallasalto_id == 10930 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                            horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                # if cont_asig_vinculacion_aprobadas != 0:
                #     if cont_asig_vinculacion_aprobadas == 1:
                #         horasfalta = 80
                #     elif cont_asig_vinculacion_aprobadas == 2:
                #         horasfalta = 160
                #     horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                #     vinculacion = ParticipantesMatrices(status=True,
                #                                         matrizevidencia_id=2,
                #                                         proyecto_id=601,
                #                                         inscripcion=inscripcion,
                #                                         horas=horasfalta,
                #                                         registrohorasdesde=fechainicioitinerario,
                #                                         registrohorashasta=fechafinitinerario,
                #                                         estado=1
                #                                         )
                #     vinculacion.save()

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_basica_rezagados():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_basica.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_basica.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("datos")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=135
        mallaantigua_id=208
        mallanueva_id=490
        sin_matricula = []
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion_id=inscripcion).first()
                # if not matricula:
                #     sin_matricula.append(inscripcion)

                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10822,10834,10846,10858,10870,10888,10913,10930,10944]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                                if equivalencia.asignaturamallasalto_id in [10822,10834,10846,10858,10870,10888,10913,10930,10944]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10822:
                                                itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                                iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioprimero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioprimero).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioprimero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iprimeronuevo.horas_practicas,
                                                                                                             nivelmalla=iprimeronuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iprimeronuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioprimero).update(
                                                    itinerario=iprimeronuevo)


                                            if equivalencia.asignaturamallasalto_id == 10834:
                                                itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=2)
                                                isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=2)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosegundo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosegundo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosegundo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=isegundonuevo.horas_practicas,
                                                                                                             nivelmalla=isegundonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=isegundonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosegundo).update(
                                                    itinerario=isegundonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10846:
                                                itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=3)
                                                iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=3)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariotercero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariotercero).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariotercero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iterceronuevo.horas_practicas,
                                                                                                             nivelmalla=iterceronuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iterceronuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariotercero).update(
                                                    itinerario=iterceronuevo)


                                            if equivalencia.asignaturamallasalto_id == 10858:
                                                itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=4)
                                                icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=4)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariocuarto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariocuarto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariocuarto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=icuartonuevo.horas_practicas,
                                                                                                             nivelmalla=icuartonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=icuartonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariocuarto).update(
                                                    itinerario=icuartonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10870:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iquintonuevo.horas_practicas,
                                                                                                             nivelmalla=iquintonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iquintonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10888:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=isextonuevo.horas_practicas,
                                                                                                             nivelmalla=isextonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=isextonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10913:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=7)
                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=iseptimonuevo.horas_practicas,
                                                                                                             nivelmalla=iseptimonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=iseptimonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10930:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=ioctavonuevo.horas_practicas,
                                                                                                             nivelmalla=ioctavonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=ioctavonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10944:
                                                itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=9)
                                                inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=9)
                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarionoveno).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarionoveno).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarionoveno).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=inovenonuevo.horas_practicas,
                                                                                                             nivelmalla=inovenonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=inovenonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarionoveno).update(
                                                    itinerario=inovenonuevo)


                                if equivalencia.asignaturamallasalto_id in [10913,10930]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10913 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10930 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    # if cont_asig_vinculacion_aprobadas != 0:
                    #     if cont_asig_vinculacion_aprobadas == 1:
                    #         horasfalta = 80
                    #     elif cont_asig_vinculacion_aprobadas == 2:
                    #         horasfalta = 160
                    #     horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                    #     vinculacion = ParticipantesMatrices(status=True,
                    #                                         matrizevidencia_id=2,
                    #                                         proyecto_id=601,
                    #                                         inscripcion=inscripcion,
                    #                                         horas=horasfalta,
                    #                                         registrohorasdesde=fechainicioitinerario,
                    #                                         registrohorashasta=fechafinitinerario,
                    #                                         estado=1
                    #                                         )
                    #     vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1




###TURISMO
def homologacion_turismo_rezagados():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_turismo_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_turismo.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("datos")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        sin_matricula = []
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion_id=inscripcion).first()
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')

                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def llenar_tabla_equivalencias_turismo():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_turismo.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA_NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)


###PSICOLOGÍA
def homologacion_psicologia_rezagados():
    try:
        libre_origen = '/homologacion_psico_2_8.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_psicologia.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("datos")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion_id=inscripcion).first()
                # if not matricula:
                #     sin_matricula.append(inscripcion)
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')

                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None

                    temporal = []
                    for e in equivalencias:
                        temporal.append(
                            [inscripcion.recordacademico_set.filter(status=True,
                                                                    asignaturamalla=e.asignaturamalla).first(),
                             e])

                    for t in temporal:
                        recordantiguo, equivalencia = t

                        old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                        print(f"Nueva - {equivalencia.asignaturamallasalto}")

                        if recordantiguo:
                            print(f"Anterior - {equivalencia.asignaturamalla}")
                            print(f"Record antiguo: {recordantiguo}")

                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            recordnuevo, homologada = None, False
                            if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                            if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                     status=True).first():
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()
                            else:
                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                             recordacademiconuevo=recordnuevo)
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()

                                if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10646:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)

                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()
                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10649:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=ioctavonuevo.horas_practicas,
                                                            nivelmalla=ioctavonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=ioctavonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10639:
                                                cont_asig_vinculacion_aprobadas += 1

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                cont_asig_vinculacion_aprobadas += 1
                                            fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                            fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                                print(u"Record actualizado %s" % recordnuevo)
                        else:
                            hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    if cont_asig_vinculacion_aprobadas != 0:
                        if cont_asig_vinculacion_aprobadas == 1:
                            horasfalta = 144
                        if cont_asig_vinculacion_aprobadas == 2:
                            horasfalta = 160
                        horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                        vinculacion = ParticipantesMatrices(status=True,
                                                            matrizevidencia_id=2,
                                                            proyecto_id=601,
                                                            inscripcion=inscripcion,
                                                            horas=horasfalta,
                                                            registrohorasdesde=fechainicioitinerario,
                                                            registrohorashasta=fechafinitinerario,
                                                            estado=1
                                                            )
                        vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(1)
                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def llenar_tabla_equivalencias_psicología():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_psicolo.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA_NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)
##########################################################################################################################


#################################################### POR VERIFICAR ######################################################

###MIGRACION TICS
def llenar_tabla_equivalencias_tics():
    try:
        miarchivo = openpyxl.load_workbook("equivalencia_malla_tics.xlsx")
        lista = miarchivo.get_sheet_by_name('datos')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[1].value}".isdigit():
                if filas[1].value is None:
                    break
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva = int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                   asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                           asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s" % tablaeq)
                else:
                    tablaeq = \
                    TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[
                        0]
                    tablaeq.asignaturamallasalto_id = idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s" % a)

    except Exception as ex:
            print('error: %s' % ex)

###MIGRACION TICS
def homologacion_tics():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tics.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=133
        mallaantigua_id=202
        mallanueva_id=478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10587,10618,10612]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [10618,10612]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10618:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallaantigua_id,
                                                                                            nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10612:
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallaantigua_id,
                                                                                            nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if equivalencia.asignaturamallasalto_id in [10587]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10587 and inscripcion.numero_horas_proyectos_vinculacion() < 96:
                                            horasfalta = 96 - inscripcion.numero_horas_proyectos_vinculacion()
                                            vinculacion = ParticipantesMatrices(status=True,
                                                                                matrizevidencia_id=2,
                                                                                proyecto_id=601,
                                                                                inscripcion=inscripcion,
                                                                                horas=horasfalta,
                                                                                registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                estado=1
                                                                                )
                                            vinculacion.save()


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def reporte_estudiantes_horas_PPP_culminadas_aprobadas():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/reporte_estudiantes_horas_PPP_culminadas_aprobadas.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                    (u"ITINERARO REPETIDO", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("idiomas.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        carrera_id = 129
        mallaantigua_id = 198
        mallanueva_id = 492
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                PPP = inscripcion.numero_horas_practicas_pre_profesionales()

                for i in range(1, 10):

                    practicaanterior = False
                    itinerarioviejo = ItinerariosMalla.objects.get(status=True,
                                                                    malla_id=mallaantigua_id,
                                                                    nivel_id=i)
                    itinerarionuevo = ItinerariosMalla.objects.get(status=True,
                                                                malla_id=mallanueva_id,
                                                                nivel_id=i)


                    practicanueva = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                           inscripcion=inscripcion,
                                                                                           estadosolicitud__in=[
                                                                                               1, 2, 4],
                                                                                           itinerariomalla=itinerarionuevo).exists()
                    practicaactividad = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                        inscripcion=inscripcion,
                                                                        estadosolicitud__in=[1, 2, 4],
                                                                        actividad__itinerariomalla=itinerarioviejo).exists()
                    if not practicaactividad:
                        practicaanterior = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud__in=[
                                                                                                   1, 2, 4],
                                                                                               itinerariomalla=itinerarioviejo).exists()

                    if practicaanterior:
                        if practicaanterior and practicanueva:
                            hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                            hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                            hojadestino.write(fila, 2, str(itinerarioviejo.nombre) + ' ' + str(itinerarioviejo.nivel) + ' - ' + str(itinerarionuevo.nombre) + ' ' + str(itinerarionuevo.nivel), fuentenormal)
                            fila += 1

                    if practicaactividad:
                        if practicaactividad and practicanueva:
                            hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                            hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                            hojadestino.write(fila, 2, str(itinerarioviejo.nombre) + ' ' + str(itinerarioviejo.nivel) + ' - ' + str(itinerarionuevo.nombre) + ' ' + str(itinerarionuevo.nivel), fuentenormal)
                            fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def reporte_estudiantes_sin_horas_vinculacion():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/reporte_estudiantes_sin_horas_vinculacion.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"HORAS_VINCULACION", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("ts.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        carrera_id = 130
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                horasvinculacion = inscripcion.numero_horas_proyectos_vinculacion()
                if horasvinculacion < 160:
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, horasvinculacion, fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def reporte_estudiantes_horas_PPP_Vinc():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/reporte_estudiantes_horas_PPP_Vinc_INICIAL2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PPP - 680", 6000),
                   (u"HORAS_VINCULACION - 160", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("basica.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        carrera_id = 135
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                horasvinculacion = inscripcion.numero_horas_proyectos_vinculacion()
                PPP = inscripcion.numero_horas_practicas_pre_profesionales()

                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, PPP, fuentenormal)
                hojadestino.write(fila, 3, horasvinculacion, fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


# llenar_tabla_equivalencias_basica()
# homologacion_basica_rezagados()
# reporte_estudiantes_horas_PPP_Vinc()
# reporte_estudiantes_horas_PPP_culminadas_aprobadas()