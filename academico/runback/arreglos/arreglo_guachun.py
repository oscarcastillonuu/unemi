import os
import statistics
import sys

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from webpush import send_user_notification
from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
import calendar
import datetime
from django.db import transaction
from sga.models import *
from sagest.models import *
from django.db.models import Sum, F, FloatField, IntegerField
from django.db.models.functions import Coalesce, ExtractYear
from settings import MEDIA_ROOT, BASE_DIR
from gdocumental.models import *
from bd.models import *
from balcon.models import EncuestaProceso, RespuestaEncuestaSatisfaccion
from empleo.models import ResponsableConvenio
# Trato de documentos xls
# Pandas libreria reciente más legible
import pandas as pd

# Openpyxl libreria para tratar cantidades grandes de registros e integración sencilla de graficos
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors

# xlwt libreria antigua utilizada para recorrer pocos registros en xls
import xlwt
from xlwt import easyxf


def actualizar_activos_fijos():
    with transaction.atomic():
        try:
            archivo_ = 'activos_fijos_actualizar'
            url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
            wb = openpyxl.load_workbook(filename=url_archivo)
            # ws = wb.get_sheet_by_name(archivo[1]) # Permite seleccionar una hoja en especifico del excel
            ws = wb.active  # Selecciona la hoja activa u hoja principal
            total_filas = ws.max_row - 2
            for row in ws.iter_rows(min_row=3):
                c_gobierno = row[0].value
                c_interno = int(row[1].value) if row[1].value and int(row[1].value) != 0 else ''
                filtro = Q(status=True)
                if c_gobierno:
                    filtro = filtro & Q(codigogobierno=c_gobierno)
                if c_interno:
                    filtro = filtro & Q(codigointerno=c_interno)
                if c_gobierno or c_interno:
                    if ActivoFijo.objects.filter(filtro).exists():
                        activo_f = ActivoFijo.objects.get(filtro)
                        activo_f.vidautil = row[3].value
                        activo_f.costo = row[5].value
                        activo_f.valorresidual = row[6].value
                        activo_f.valorlibros = row[7].value
                        activo_f.valordepreciacionacumulada = row[8].value
                        activo_f.save()
                        print(f'Código: {row[0].value}, '
                              f'Vida útil: {row[3].value}, '
                              f'Valor contable: {row[5].value}, '
                              f'Valor residual: {row[6].value}, '
                              f'Valor en libros: {row[7].value}, '
                              f'Valor depreciación acumulada: {row[8].value}')
                    else:
                        raise NameError(f'Error con c_gobierno: {c_gobierno}, c_interno:{c_interno}')
                else:
                    total_filas -= 1
            print(f'Se actulizo: {total_filas}')
        except Exception as ex:
            transaction.set_rollback(True)
            print(str(ex))
# actualizar_activos_fijos()

def actualizar_estado_edcon():
    inscritos = CapInscritoIpec.objects.filter(status=True)
    cont = 0
    for i in inscritos:
        nota_final = i.instructor_notasfinales()
        if nota_final[0][1] == 'APROBADO':
            cont += 1
            i.estado = 2
            i.save()
            print(f'{i.id} Estado cambiado: aprobado')
    print(cont)
# actualizar_estado_edcon()

def actualizar_calculo_valores_producto():
    with transaction.atomic():
        try:
            hoy = datetime.now()
            anio_anterior = hoy.date().year - 1
            inicio, fin = datetime(anio_anterior, 1, 1), datetime(anio_anterior, 12, 31)
            tiempo_reposicion = 15
            total_actualizado = 0
            productos_ids = KardexInventario.objects.filter(fecha__range=(inicio, fin), tipomovimiento=2, status=True).values_list('producto_id').distinct()
            productos = Producto.objects.filter(status=True).exclude(id__in=productos_ids)
            for p in productos:
                valores_salida = []
                total_salida = 0
                # CÁLCULO DE VALORES SALIDA
                ultimo_kardex = KardexInventario.objects.annotate(anio=ExtractYear('fecha')).filter(anio__lt=hoy.date().year, tipomovimiento=2, producto_id=p.id, status=True).order_by('fecha').last()
                if not ultimo_kardex:
                    ultimo_kardex = KardexInventario.objects.annotate(anio=ExtractYear('fecha')).filter(tipomovimiento=2, producto_id=p.id, status=True).order_by('fecha').last()
                if ultimo_kardex:
                    anio = ultimo_kardex.anio
                    ini, fi = datetime(anio, 1, 1), datetime(anio, 12, 31)
                    kardexs = KardexInventario.objects.filter(fecha__range=(ini, fi), tipomovimiento=2, producto_id=p.id, status=True)
                    for mes in range(1, 13):
                        ultimo_dia = calendar.monthrange(anio, mes)[1]
                        i_mes = datetime(anio, mes, 1)
                        f_mes = datetime(anio, mes, ultimo_dia)
                        kardexs_mes = kardexs.filter(fecha__range=(i_mes, f_mes)).aggregate(valor_salida=Sum('valor'), total_cantidad=Sum('cantidad'))
                        valor_salida = round(kardexs_mes['valor_salida'], 2) if kardexs_mes['valor_salida'] else 0
                        total_cantidad = round(kardexs_mes['total_cantidad'], 2) if kardexs_mes['total_cantidad'] else 0
                        total_salida += valor_salida
                        diccionario = {'mes': mes, 'valor_salida': valor_salida, 'total_cantidad': total_cantidad}
                        valores_salida.append(diccionario)
                    consumo_minimo = round(min(valores_salida, key=lambda x: x["valor_salida"])["valor_salida"], 2)
                    consumo_maximo = round(max(valores_salida, key=lambda x: x["valor_salida"])["valor_salida"], 2)
                    consumo_medio = round(statistics.mean([float(d["valor_salida"]) for d in valores_salida]), 2)
                    consumo_minimo = consumo_minimo if not consumo_minimo == 0 else 1
                    cantidad_minima = min(valores_salida, key=lambda x: x["total_cantidad"])["total_cantidad"]
                    kardex_maxima = max(valores_salida, key=lambda x: x["total_cantidad"])["total_cantidad"]
                    kardex_minima = cantidad_minima if not cantidad_minima == 0 else 1

                    p.consumo_minimo_diario = consumo_minimo
                    p.consumo_medio_diario = consumo_medio
                    p.consumo_maximo_diario = consumo_maximo
                    p.kardex_maximo = kardex_maxima
                    p.kardex_minimo = kardex_minima
                    p.tiempo_reposicion_inventario = tiempo_reposicion
                    p.save()
                    p.minimo = p.calcular_existencia_minima()
                    p.maximo = p.calcular_existencia_maxima()
                    p.save()
                    total_actualizado += 1
                    totales = {'id_producto': p.id, 'total_salida': total_salida,
                               'consumo_minimo': consumo_minimo,
                               'consumo_maximo': consumo_maximo,
                               'consumo_medio': consumo_medio,
                               'cantidad_minima': kardex_minima,
                               'cantidad_maxima': kardex_maxima,
                               }
                    print(totales)
            print(f'Actualizado:{total_actualizado}')
            print(f'Sin actualizar:{len(productos) - total_actualizado}')
        except Exception as ex:
            transaction.set_rollback(True)
            print(str(ex))
# actualizar_calculo_valores_producto()

def secuencia_activos():
    if not SecuenciaActivos.objects.exists():
        secuencia = SecuenciaActivos()
        secuencia.save()
        return secuencia
    else:
        return SecuenciaActivos.objects.all()[0]

def crear_constataciones_mal_estado():
    with transaction.atomic():
        try:
            activosfijo = ActivoFijo.objects.filter(status=True, procesobaja=True)
            creados, actualizados = 0, 0
            for activo in activosfijo:
                detalle_c = DetalleConstatacionFisica.objects.filter(status=True, activo=activo, codigoconstatacion__periodo_id=1).first()
                if not detalle_c:
                    constatacion = ConstatacionFisica.objects.filter(status=True, usuariobienes=activo.responsable, periodo_id=1).first()
                    if not constatacion:
                        secuencia = secuencia_activos()
                        secuencia.numeroconstatacion += 1
                        secuencia.save()
                        constatacion = ConstatacionFisica(usuariobienes=activo.responsable,
                                                          numero=secuencia.numeroconstatacion,
                                                          normativaconstatacion=secuencia.normativaconstatacion,
                                                          fechainicio=datetime.now(),
                                                          periodo_id=1,
                                                          ubicacionbienes=activo.ubicacion)
                        constatacion.save()
                    detalle_c = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                          activo=activo,
                                                          responsable_id=1204,
                                                          ubicacionbienes=activo.ubicacion,
                                                          estadooriginal=activo.estado,
                                                          estadoactual_id=3,
                                                          enuso=False,
                                                          encontrado=True)
                    detalle_c.save()
                    creados += 1
                    print(f'Constatación creada: {creados}')
                else:
                    detalle_c.ubicacionbienes = activo.ubicacion
                    detalle_c.estadooriginal = activo.estado
                    detalle_c.estadoactual_id = 3
                    detalle_c.enuso = False
                    detalle_c.encontrado = True
                    detalle_c.save()
                    actualizados += 1
                    print(f'Constatación actualizada: {actualizados}')
            print(f'Se constatado :{len(activosfijo)} activos, Creados: {creados}, Actualizados:{actualizados}')
        except Exception as ex:
            transaction.set_rollback(True)
            print(str(ex))
# crear_constataciones_mal_estado()

def actualizar_numero_hijos():
    try:
        personas = DistributivoPersona.objects.filter(status=True)
        for d in personas:
            total_hijos = len(d.persona.cargas())
            per_extension = d.persona.personaextension_set.filter(status=True).first()
            per_extension.hijos = total_hijos
            per_extension.save(update_fields=["hijos"])
            print(f'Cantidad de hijos actualizado: {total_hijos}')
        print(f'Cantidad personas actualizadas: {len(personas)}')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))
# actualizar_numero_hijos()

def migrar_estado():
    try:
        print(f'Inicio proceso')
        cronogramas = CronogramaPersonaConstatacionAT.objects.filter(status=True, periodo_id=2, estado=4)
        cont=0
        for cronograma in cronogramas:
            activos_c = cronograma.detalleconstatacionfisicaactivotecnologico_set.filter(status=True, cronograma__status=True, constatado=True)
            for ac in activos_c:
                activo = ac.activo
                if not activo.estado == ac.estadoactual:
                    cont += 1
                    activo.estado = ac.estadoactual
                    activo.save(update_fields=["estado"])
                    print(f'{cont}. Estado actualizado codigogobierno={activo.activotecnologico.codigogobierno} activo_id={activo.id} - {activo.estado} de {cronograma.persona}')
        print(f'Se actualizo satisfactoriamente {cont} de {len(cronogramas)} personas')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))
# migrar_estado()

def actualizar_encuesta_balcon():
    try:
        print('Inicio proceso de actualización de content type y object_id')
        e_procesos = EncuestaProceso.objects.filter(proceso__isnull=False, status=True)
        cont = 0
        for ep in e_procesos:
            content_type = ContentType.objects.get_for_model(ep.proceso)
            ep.content_type = content_type
            ep.categoria_id = 2
            ep.object_id = ep.proceso.id
            ep.save()
            cont += 1
            respuestas = RespuestaEncuestaSatisfaccion.objects.filter(pregunta__encuesta=ep, status=True, solicitud__isnull=False)
            for respuesta in respuestas:
                content_type = ContentType.objects.get_for_model(respuesta.solicitud)
                respuesta.content_type = content_type
                respuesta.object_id = respuesta.solicitud.id
                respuesta.save()
            print(f'{cont}. Modelo: {content_type}, id: {ep.id} - añadido y {len(respuestas)} respuestas actualizadas')
        print(f'{len(e_procesos)} actualizados.')
    except Exception as ex:
     transaction.set_rollback(True)
     print(str(ex))
# actualizar_encuesta_balcon()

def extraer_valores_excel():
    try:
        archivo_ = 'BATERIAS_BAJA_2019'
        url_archivo = "D:\\git\\academico\\media\\BATERIAS_BAJA_2019.xlsx"
        nombres_hojas = pd.ExcelFile(url_archivo).sheet_names
        for name in nombres_hojas:
            url_guardar = f"D:\\git\\academico\\media\\{name}.xlsx"
            df = pd.read_excel(url_archivo, sheet_name=name)
            # if not 'CÓDIGO DEL BIEN' in df.columns:
            #     raise NameError('Formato de archivo erróneo, columna código del bien faltante.')
            df['Valor contable'] = ""
            df['Fecha ingreso'] = ""
            df['Fecha baja'] = ""
            df['Códigos gob iguales'] = ""
            df['Códigos int iguales'] = ""
            cont=0
            for index, row in df.iterrows():
                codigo = str(row['CÓDIGO']).strip().split('.')[0]
                codigo_interno=''
                activo = ActivoFijo.objects.filter(Q(codigogobierno__iexact=codigo) |
                                                   Q(codigointerno__iexact=codigo)).first()
                if 'CÓDIGO ANTERIOR' in df.columns:
                    codigo_interno = str(row['CÓDIGO ANTERIOR']).strip().split('.')[0]

                if codigo_interno and not activo:
                    activo = ActivoFijo.objects.filter(Q(codigointerno__iexact=codigo_interno) |
                                                       Q(codigogobierno__iexact=codigo_interno)).first()

                if not activo:
                    cont += 1
                    print(f'Activo no encontrado {codigo}')
                else:
                    baja = DetalleBajaActivo.objects.filter(status=True, activo=activo, seleccionado=True).first()
                    df.at[index, 'Valor contable'] = f'$ {activo.costo}'
                    df.at[index, 'Fecha ingreso'] = activo.fechaingreso
                    df.at[index, 'Fecha baja'] = baja.codigobaja.fecha if baja and baja.codigobaja else 'S/F'
                    df.at[index, 'Códigos gob iguales'] = "SI" if activo.codigogobierno == codigo else 'NO'
                    df.at[index, 'Códigos int iguales'] = "SI" if codigo_interno and activo.codigointerno == codigo_interno or activo.codigointerno == codigo else 'NO'
                    print(f'Información extraida correctamente {codigo}')
            print(f'Activos no encontrados: {cont}, encontrados:')
            df.to_excel(url_guardar, index=False)
    except Exception as ex:
        print(f'{ex}')
# extraer_valores_excel()

def extraer_valores_contables_excel():
    try:
        archivo_ = 'adjunto'
        url_archivo = "D:\\git\\academico\\media\\adjunto.xlsx"
        url_guardar = f"D:\\git\\academico\\media\\acta_entrega.xlsx"
        df = pd.read_excel(url_archivo, sheet_name='ACTA DE ENTREGA')
        # if not 'CÓDIGO DEL BIEN' in df.columns:
        #     raise NameError('Formato de archivo erróneo, columna código del bien faltante.')
        df['Valor contable'] = ""
        df['Fecha ingreso'] = ""
        cont=0
        for index, row in df.iterrows():
            codigo = str(row['Cod. gobierno']).strip().split('.')[0]
            codigo_interno = str(row['Cod. interno']).strip().split('.')[0]
            activo = None
            if codigo:
                activo = ActivoFijo.objects.filter(Q(codigogobierno__iexact=codigo) |
                                                   Q(codigointerno__iexact=codigo)).first()

            if codigo_interno and not activo:
                activo = ActivoFijo.objects.filter(Q(codigointerno__iexact=codigo_interno) |
                                                   Q(codigogobierno__iexact=codigo_interno)).first()

            if not activo:
                cont += 1
                print(f'Activo no encontrado {codigo}')
            else:
                baja = DetalleBajaActivo.objects.filter(status=True, activo=activo, seleccionado=True).first()
                df.at[index, 'Valor contable'] = f'$ {activo.costo}'
                df.at[index, 'Fecha ingreso'] = activo.fechaingreso
                print(f'Información extraida correctamente {codigo}')
        print(f'Activos no encontrados: {cont}, encontrados:')
        df.to_excel(url_guardar, index=False)
    except Exception as ex:
        print(f'{ex}')
# extraer_valores_contables_excel()

def migrar_responsables_internos_convenios():
    try:
        cont=0
        print('Inicio proceso de migración')
        convenios = ConvenioEmpresa.objects.filter(status=True, responsableinterno__isnull=False)
        for convenio in convenios:
            cargo = convenio.cargo_denominaciones.all()
            cargo = cargo.first() if len(cargo) == 1 else None
            if not cargo:
                cargo = convenio.responsableinterno.mi_cargo_administrativo()
                if not cargo:
                    cargo = convenio.responsableinterno.distributivopersonahistorial_set.filter(status=True, regimenlaboral_id=1).order_by('-fecha_creacion').first()
                    cargo = cargo.denominacionpuesto if cargo else convenio.responsableinterno.mi_cargo()
            if not ResponsableConvenio.objects.filter(convenio=convenio, persona=convenio.responsableinterno):
                responsable = ResponsableConvenio(convenio=convenio, persona=convenio.responsableinterno, cargo=cargo)
                responsable.save()
                cont += 1
                print(f'Se migro el responsable con exito: {responsable}')
        convenios = ConvenioEmpresa.objects.filter(status=True, responsableinterno__isnull=True)
        print(f'{cont} convenios de responsable interno Migrados.')
        cont = 0
        for convenio in convenios:
            distributivos = DistributivoPersona.objects.filter(denominacionpuesto__in=convenio.cargo_denominaciones.all().values_list('id', flat=True), status=True)
            for distributivo in distributivos:
                if not ResponsableConvenio.objects.filter(convenio=convenio, persona=distributivo.persona):
                    responsable = ResponsableConvenio(convenio=convenio, persona=distributivo.persona, cargo=distributivo.denominacionpuesto)
                    responsable.save()
                    cont += 1
                    print(f'Se migro el responsable con exito: {responsable}')
        print(f'{cont} convenios por cargo Migrados.')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# migrar_responsables_internos_convenios()

def crear_sesiones_resolucion():
    try:
        cont=0
        print('Inicio proceso de migración')
        resoluciones = Resoluciones.objects.filter(status=True)
        fechas = resoluciones.values_list('fecha', 'tipo_id').order_by('fecha', 'tipo_id').distinct()
        for r in fechas:
            cont += 1
            fecha = r[0].strftime("%Y-%m-%d")
            nombre = f'Sesión: {fecha}'
            sesion = SesionResolucion(tipo_id=int(r[1]), fecha=fecha, nombre=nombre)
            sesion.save()
            resoluciones = Resoluciones.objects.filter(status=True, fecha=r[0], tipo=sesion.tipo)
            for resolucion in resoluciones:
                resolucion.sesion=sesion
                resolucion.save()
            print(f'Sesión creada correctamente {cont}')

        print(f'{len(fechas)} Sesiones creadas')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

#crear_sesiones_resolucion()

def mover_constataciones_duplicaddas():
    try:
        cont=0
        print('Inicio proceso de constatacion')
        responsables = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).values_list('usuariobienes_id', flat=True).order_by('usuariobienes_id').distinct()
        for responsable in responsables:
            constataciones = ConstatacionFisica.objects.filter(usuariobienes=responsable, periodo_id=1, estado=1).order_by('numero')
            if len(constataciones) > 1:
                first_constatacion = constataciones.first()
                constataciones = constataciones.exclude(id=first_constatacion.id)
                for constatacion in constataciones:
                    constatacion.detalle_constatacion().update(codigoconstatacion=first_constatacion)
                    # constatacion.delete()
                    print(f'Cabecera de constatación eliminara y sus detalles actualizados a primera cabcera salvada {first_constatacion}')

        constataciones = ConstatacionFisica.objects.filter(status=True, periodo__isnull=True, estado=2).order_by('numero')
        numero = constataciones.last().numero
        constataciones = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).order_by('numero')
        for constatacion in constataciones:
            numero += 1
            if not constatacion.numero == numero:
                constatacion.numero = numero
                constatacion.save()
                print(f'Constatación actualizada {numero} - {constatacion}')
        print(f'Finalizado con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# mover_constataciones_duplicaddas()

def delete_cabeceraconstatacion():
    try:
        print('Inicio proceso de constatacion')
        constataciones = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).order_by('numero')
        cont = 0
        for constatacion in constataciones:
            if not constatacion.detalle_constatacion().exists():
                cont += 1
                constatacion.delete()
                print(f'{cont}.Cabecera eliminada')
        constataciones = ConstatacionFisica.objects.filter(status=True, periodo__isnull=True, estado=2).order_by('numero')
        numero = constataciones.last().numero
        constataciones = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).order_by('numero')
        for constatacion in constataciones:
            numero += 1
            if not constatacion.numero == numero:
                constatacion.numero = numero
                constatacion.save()
                print(f'Constatación actualizada {numero} - {constatacion}')
        print(f'Finalizado con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# delete_cabeceraconstatacion()

def ordenar_sessiones_resoluciones():
    try:
        print('Inicio proceso de ordenamiento')
        tipos = TipoResolucion.objects.filter(status=True)
        for t in tipos:
            sesiones = t.sesiones().order_by('fecha_creacion')
            for s in sesiones:
                if s.orden == 0:
                    s.orden = s.orden_next()
                    s.save()
                    print(f'Sesión {s}--{s.orden}')
                for idx, r in enumerate(s.resoluciones()):
                    orden = r.numeroresolucion[-3:]
                    i = -1  # Índice para recorrer hacia la izquierda
                    # Encontrar la posición del primer carácter no numérico
                    while i >= -len(orden) and orden[i].isdigit():
                        i -= 1
                    # Extraer los dígitos numéricos y convertir a entero
                    num_part = orden[i + 1:]
                    if num_part.isdigit():
                        r.orden = int(num_part)
                    else:
                        r.orden = idx + 1
                    r.save()
                    print(f'Resolución {r.numeroresolucion}--{r.orden}')
            resoluciones = Resoluciones.objects.filter(status=True, sesion__isnull=True, tipo=t).order_by('fecha_creacion')
            for idx, r in enumerate(resoluciones):
                orden = r.numeroresolucion[-3:]
                i = -1  # Índice para recorrer hacia la izquierda
                # Encontrar la posición del primer carácter no numérico
                while i >= -len(orden) and orden[i].isdigit():
                    i -= 1
                # Extraer los dígitos numéricos y convertir a entero
                num_part = orden[i + 1:]
                if num_part.isdigit():
                    r.orden = int(num_part)
                else:
                    r.orden = idx
                r.save()
                print(f'Resolución {r.numeroresolucion}--{r.orden}')
        print(f'Proceso finalizado')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# ordenar_sessiones_resoluciones()

# Lógica de modelo para Preguntas de Paz Y Salvo
# Modelo DetalleDireccionFormatoPS, atributo logicamodelo
def verificar_pregunta(pazsalvo):
    from sagest.models import ActivoFijo
    funcionario = pazsalvo.persona
    activos = ActivoFijo.objects.filter(Q(responsable=funcionario) | Q(custodio=funcionario),status=True).exists()
    cargos = funcionario.mis_cargos().exclude(denominacionpuesto=pazsalvo.cargo).exists()
    return not activos or cargos


def crear_respuestas_jefe_ps():
    from sagest.models import PazSalvo, DetallePazSalvo
    try:
        print('Inicio de proceso')
        pazsalvos = PazSalvo.objects.filter(status=True, estado_requisito=1)
        cont=0
        for idx, ps in enumerate(pazsalvos):
            jefe = ps.jefeinmediato
            cargo = jefe.mi_cargo_administrativo() if jefe.mi_cargo_administrativo() else jefe.mi_cargo()
            cumplimiento = ps.cumplimiento(cargo.id, True)
            if not cumplimiento['respondio']:
                preguntas = ps.preguntas(cargo.id, True, True)
                preguntas_all = ps.preguntas_all()
                respuestas_all = ps.respuestas_all()
                total = len(preguntas_all) - len(respuestas_all)
                if total > len(preguntas):

                    for idxre, pregunta in enumerate(preguntas):
                        respuesta = ps.respuestas_jefe().filter(pregunta=pregunta).first()
                        if not respuesta:
                            respuesta = DetallePazSalvo(persona=jefe,
                                                        pazsalvo=ps,
                                                        pregunta=pregunta,
                                                        marcado=True)
                            respuesta.save()
                            print(f'{idx}.{idxre}. Se crea respuestas de jefe inmediato Paz y Salvo (pazsalvo_id : {ps.id} | respuesta_id: {respuesta.id})')
                            cont += 1
        print(f'Finalizo proceso de creación con éxito, se creo {cont}')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# crear_respuestas_jefe_ps()

def crear_respuestas_financiero_ps():
    from sagest.models import PazSalvo, DetallePazSalvo
    try:
        print('Inicio de proceso')
        pazsalvos = PazSalvo.objects.filter(status=True, estado=2)
        cont=0
        for idx, ps in enumerate(pazsalvos):
            jefe = Persona.objects.get(id=8294)
            cargo = jefe.mi_cargo_administrativo() if jefe.mi_cargo_administrativo() else jefe.mi_cargo()
            cumplimiento = ps.cumplimiento(cargo.id)
            if not cumplimiento['respondio']:
                preguntas = ps.preguntas(cargo.id)
                for idxre, pregunta in enumerate(preguntas):
                    respuesta = ps.respuestas(cargo).filter(pregunta=pregunta, persona=jefe).first()
                    if not respuesta:
                        respuesta = DetallePazSalvo(persona=jefe,
                                                    pazsalvo=ps,
                                                    pregunta=pregunta,
                                                    respondio=True,
                                                    marcado=True)
                        respuesta.save()
                        cont += 1
        print(f'Finalizo proceso de creación con éxito, se creo {cont}')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# crear_respuestas_financiero_ps()

def actualizar_estado_actas_postulate():
    from postulate.models import ActaPartida
    try:
        print('Inicio de proceso')
        actas = ActaPartida.objects.filter(status=True).exclude(estado=3)
        for idx, acta in enumerate(actas):
            ids_responsables = acta.responsables().values_list('id', flat=True).order_by('id').distinct()
            firmados = acta.historial_firmados().filter(personatribunal_id__in=ids_responsables).values_list('personatribunal_id', flat=True).order_by('personatribunal_id').distinct('personatribunal_id')
            acta.estado = 3 if len(firmados) >= len(ids_responsables) else 2
            acta.save()
        print(f'Finalizo proceso de creación con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

def revertir_estado_ps():
    from sagest.models import PazSalvo
    try:
        print('Inicio de proceso')
        pazsalvos = PazSalvo.objects.filter(status=True, estado=2)
        for idx, ps in enumerate(pazsalvos):
            respuestas_all = len(ps.detallepazsalvo_set.filter(status=True, respondio=True).order_by('pregunta_id').distinct('pregunta_id'))
            if not len(ps.preguntas_all()) <= respuestas_all:
                ps.estado = 1
                ps.save()
                print(f'{ps.id}. Actualizado')
        print(f'Finalizo proceso de creación con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

revertir_estado_ps()